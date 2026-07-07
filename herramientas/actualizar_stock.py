# -*- coding: utf-8 -*-
"""
Genera data/stock.json cruzando las 2 tablas de stock de Curifor con los códigos
de repuestos que usan las pautas de mantención.

Fuentes (viven en SharePoint, se van actualizando):
  - Stock bodegas.xlsx           -> giro CURIFOR (autos livianos: repuestos de las pautas)
  - Stock bodegas Frontera.xlsx  -> giro FRONTERA (camiones)

Uso:
  python herramientas/actualizar_stock.py             # usa el snapshot en herramientas/stock_fuente/
  python herramientas/actualizar_stock.py --descargar # baja primero las tablas frescas de SharePoint

La descarga reutiliza el módulo subir_sharepoint.py del proyecto Data BI
(perfil Playwright ya logueado + REST). Si la sesión expiró, correr allá:
  python subir_sharepoint.py login

Salida: data/stock.json  (solo los códigos que aparecen en alguna pauta) y un
resumen impreso + herramientas/stock_reporte.md.
"""
import glob
import json
import os
import re
import sys
from collections import defaultdict

AQUI = os.path.dirname(os.path.abspath(__file__))
FUENTE = os.path.join(AQUI, "stock_fuente")
DATA = os.path.normpath(os.path.join(AQUI, "..", "data"))
AUTOM = r"C:\Users\icalderon\OneDrive - Curifor S.A\Documentos\Desarrollos\Automatizaciones\3. Actualizacion\automatizacion"

ARCHIVO_CURIFOR = "Stock bodegas.xlsx"
ARCHIVO_FRONTERA = "Stock bodegas Frontera.xlsx"


def norm(c):
    return re.sub(r"[^A-Z0-9]", "", str(c).upper()) if c is not None else ""


def descargar_de_sharepoint():
    """Baja las 2 tablas desde SharePoint al folder stock_fuente/. Reusa subir_sharepoint."""
    import urllib.request
    from urllib.parse import quote

    sys.path.insert(0, AUTOM)
    cwd = os.getcwd()
    os.chdir(AUTOM)
    try:
        from subir_sharepoint import _cookies, SITE, FOLDER, _SSL_CTX
        ck = _cookies()
        os.makedirs(FUENTE, exist_ok=True)
        for nombre in (ARCHIVO_CURIFOR, ARCHIVO_FRONTERA):
            url = f"{SITE}/_api/web/GetFileByServerRelativeUrl('{quote(FOLDER + '/' + nombre)}')/$value"
            req = urllib.request.Request(url, headers={"Cookie": ck, "Accept": "application/octet-stream"})
            with urllib.request.urlopen(req, timeout=300, context=_SSL_CTX) as r:
                data = r.read()
            with open(os.path.join(FUENTE, nombre), "wb") as f:
                f.write(data)
            print(f"  descargado {nombre}: {len(data)/1024/1024:.2f} MB")
        return True
    except Exception as e:
        print(f"  AVISO: no se pudo descargar de SharePoint ({e}).")
        print("  Se usará el último snapshot en stock_fuente/. Para refrescar la sesión: "
              "python subir_sharepoint.py login (en el proyecto Data BI).")
        return False
    finally:
        os.chdir(cwd)


def leer_stock(ruta, con_rubro):
    """Lee una tabla de stock. Devuelve (idx, crudo):
      idx  = {codigoNorm: {stock, desc, precio, bodegas:set}}  (por código limpio)
      crudo = [ (codigoNorm, descNorm, stock, desc, precio, bodega) ]  (para cruce secundario)
    """
    import openpyxl
    idx = {}
    crudo = []
    if not os.path.exists(ruta):
        return idx, crudo
    wb = openpyxl.load_workbook(ruta, read_only=True, data_only=True)
    ws = wb.worksheets[0]
    filas = ws.iter_rows(min_row=1, values_only=True)
    encab = next(filas)
    col = {str(v).strip().lower(): i for i, v in enumerate(encab) if v is not None}

    def gi(row, *nombres, default=None):
        for n in nombres:
            if n in col and col[n] < len(row):
                return row[col[n]]
        return default

    for row in filas:
        prod = gi(row, "producto")
        if prod is None:
            continue
        prod = str(prod).strip()
        # código limpio: quitar rubro (primer token) en el stock Curifor
        if con_rubro and " " in prod:
            codigo = prod.split(" ", 1)[1].strip()
        else:
            codigo = prod
        nc = norm(codigo)
        if not nc:
            continue
        stock = gi(row, "stock", default=0) or 0
        if not isinstance(stock, (int, float)):
            stock = 0
        desc = gi(row, "descripción", "descripcion")
        precio = gi(row, "precio venta")
        bodega = gi(row, "bodega")
        precio = int(round(precio)) if isinstance(precio, (int, float)) and precio else None
        descs = str(desc).strip() if desc else None
        bod = str(bodega).strip() if bodega and str(bodega).strip() else None

        def acumular(clave):
            e = idx.setdefault(clave, {"stock": 0, "desc": None, "precio": None, "porBodega": {}})
            e["stock"] += stock
            if descs and not e["desc"]:
                e["desc"] = descs
            if precio and not e["precio"]:
                e["precio"] = precio
            if bod and stock:
                e["porBodega"][bod] = e["porBodega"].get(bod, 0) + stock

        acumular(nc)
        # el mismo stock también queda indexado por su código de REEMPLAZO (supersesión):
        # si la pauta usa el SKU antiguo y el stock lo tiene bajo el nuevo (o viceversa),
        # así igual cruza. Solo el stock Curifor tiene esta columna.
        reempl = gi(row, "reemplazo")
        nr = norm(reempl)
        if nr and nr != nc:
            acumular(nr)
        # tokens de la descripción (separados por no-alfanuméricos) para cruce por código
        tokens = frozenset(t for t in re.split(r"[^A-Z0-9]+", descs.upper()) if t) if descs else frozenset()
        crudo.append((nc, tokens, stock, descs, precio, bod))
    wb.close()
    return idx, crudo


def _acumular_matches(filas, es_match):
    """Agrega stock/bodega de las filas de crudo que cumplen es_match(fila).
    Devuelve (acc, alt) donde alt = código del SKU con más stock. (None, None) si nada."""
    acc = {"stock": 0, "desc": None, "precio": None, "porBodega": {}}
    por_codigo = {}
    encontrado = False
    for codigoNorm, tokens, stock, desc, precio, bod in filas:
        if not es_match(codigoNorm, tokens, desc):
            continue
        encontrado = True
        acc["stock"] += stock or 0
        por_codigo[codigoNorm] = por_codigo.get(codigoNorm, 0) + (stock or 0)
        if desc and not acc["desc"]:
            acc["desc"] = desc
        if precio and not acc["precio"]:
            acc["precio"] = precio
        if bod and stock:
            acc["porBodega"][bod] = acc["porBodega"].get(bod, 0) + stock
    if not encontrado:
        return None, None
    alt = max(por_codigo, key=por_codigo.get) if por_codigo else None
    return acc, alt


def buscar_secundario(nc, crudo):
    """Cruce difuso: el código como prefijo con sufijo de letras (104406 -> 104406-AG)
    o como token COMPLETO de la descripción. Devuelve (acc, alt) o (None, None)."""
    if len(nc) < 6:
        return None, None
    def es_match(codigoNorm, tokens, desc):
        if codigoNorm.startswith(nc) and (len(codigoNorm) == len(nc) or codigoNorm[len(nc)].isalpha()):
            return True
        return nc in tokens
    return _acumular_matches(crudo, es_match)


def buscar_por_tokens(tokens, crudo):
    """Cruce por NOMBRE: la descripción del producto en stock contiene TODAS las palabras.
    Usado para lubricantes cuyo SKU en bodega difiere del código de tambor de la pauta
    (mapeo curado en equivalencias.json). Devuelve (acc, alt) o (None, None)."""
    toks = [t.upper() for t in tokens]
    def es_match(codigoNorm, tks, desc):
        return bool(desc) and all(t in desc.upper() for t in toks)
    return _acumular_matches(crudo, es_match)


ARCHIVO_COMPLETO = "StockCurifor_completo.xlsx"


def leer_catalogo_completo():
    """Lee StockCurifor_completo.xlsx (catálogo enriquecido Ford, snapshot estable).
    Devuelve (equiv, aplic):
      equiv = {codNorm: set(codigos equivalentes norm)}  (Reemplazo + Equivalente Scrap + Supersesión)
      aplic = {codNorm: 'MODELOS...'}                     (Aplicabilidad por modelo)
    Si el archivo no está, devuelve mapas vacíos (la plataforma funciona igual)."""
    import openpyxl
    ruta = os.path.join(FUENTE, ARCHIVO_COMPLETO)
    equiv, aplic = {}, {}
    if not os.path.exists(ruta):
        print(f"  (aviso: {ARCHIVO_COMPLETO} no está; sin equivalencias/aplicabilidad Ford)")
        return equiv, aplic
    wb = openpyxl.load_workbook(ruta, read_only=True, data_only=True)
    ws = wb.worksheets[0]
    filas = ws.iter_rows(min_row=1, values_only=True)
    encab = next(filas)
    col = {str(v).strip().lower(): i for i, v in enumerate(encab) if v is not None}

    def gi(row, nombre):
        i = col.get(nombre)
        return row[i] if i is not None and i < len(row) else None

    def limpio_normal(prod):
        p = str(prod).strip()
        return norm(p.split(" ", 1)[1] if " " in p else p)

    def enlazar(a, b):
        a, b = norm(a), norm(b)
        if a and b and a != b:
            equiv.setdefault(a, set()).add(b)
            equiv.setdefault(b, set()).add(a)

    for row in filas:
        prod = gi(row, "producto")
        if prod is None:
            continue
        limpio = gi(row, "código limpio") or gi(row, "codigo limpio")
        base = norm(limpio) if limpio else limpio_normal(prod)
        for campo in ("reemplazo", "código equivalente (scrap)", "supersesión (scrap)", "supersesion (scrap)"):
            v = gi(row, campo)
            if v and str(v).strip() not in ("", "-", "0", "#N/A"):
                enlazar(base, v)
        ap = gi(row, "aplicabilidad (modelos)")
        if ap and str(ap).strip() not in ("", "-", "0", "#N/A") and base not in aplic:
            aplic[base] = re.sub(r"\s*\|\s*", " · ", str(ap).strip())[:180]
    wb.close()
    print(f"Catálogo completo: {len(equiv)} códigos con equivalencia, {len(aplic)} con aplicabilidad")
    return equiv, aplic


def cargar_mapeo_manual():
    """equivalencias.json: {codNorm: {'tokens': [...] | 'codigo': 'XXX', 'nota': ...}}."""
    ruta = os.path.join(AQUI, "equivalencias.json")
    if not os.path.exists(ruta):
        return {}
    data = json.load(open(ruta, encoding="utf-8")).get("mapeo", {})
    return {norm(k): v for k, v in data.items()}


def codigos_de_pautas():
    """Set de códigos normalizados usados en las pautas + su forma original."""
    usados = {}
    for f in glob.glob(os.path.join(DATA, "pautas", "*.json")):
        d = json.load(open(f, encoding="utf-8"))
        for pl in d["planes"]:
            for itv in pl["intervalos"]:
                for it in (itv.get("items") or []):
                    cod = it.get("codigo")
                    if cod:
                        nc = norm(cod)
                        # placeholders que no son códigos reales
                        if nc and not any(x in str(cod).upper() for x in
                                          ("COMPRA EN PLAZA", "PENDIENTE", "INGRESAR", "MAT-", "N/A")):
                            usados[nc] = str(cod).strip()
    return usados


def main(descargar=False):
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")
    if descargar:
        print("Descargando tablas de stock desde SharePoint...")
        descargar_de_sharepoint()

    cur, cur_crudo = leer_stock(os.path.join(FUENTE, ARCHIVO_CURIFOR), con_rubro=True)
    fro, fro_crudo = leer_stock(os.path.join(FUENTE, ARCHIVO_FRONTERA), con_rubro=False)
    print(f"Stock Curifor:  {len(cur)} códigos")
    print(f"Stock Frontera: {len(fro)} códigos")
    equiv_map, aplic_map = leer_catalogo_completo()
    mapeo = cargar_mapeo_manual()

    usados = codigos_de_pautas()
    print(f"Códigos de repuestos en pautas (reales): {len(usados)}")

    def resolver_curifor(nc):
        """Resuelve el stock Curifor de un código con precedencia:
        directo -> mapeo manual por nombre -> difuso -> equivalencia (catálogo).
        Devuelve (entry, alt, via) o (None, None, None)."""
        e = cur.get(nc)
        if e:
            return e, None, "directo"
        if nc in mapeo:
            m = mapeo[nc]
            if m.get("codigo"):
                e = cur.get(norm(m["codigo"]))
                if e:
                    return e, m["codigo"], "producto"
            if m.get("tokens"):
                e, alt = buscar_por_tokens(m["tokens"], cur_crudo)
                if e:
                    return e, alt, "producto"
        e, alt = buscar_secundario(nc, cur_crudo)
        if e:
            return e, (alt if alt != nc else None), "difuso"
        for eq in equiv_map.get(nc, ()):          # equivalencia/supersesión (catálogo Ford)
            if eq in cur:
                return cur[eq], eq, "equivalente"
        return None, None, None

    items = {}
    n_cur = n_fro = n_aprox = 0
    via_cnt = {}
    for nc, original in usados.items():
        ec, alt, via = resolver_curifor(nc)
        ef, _ = (fro.get(nc), None)
        if not ef:
            ef, _ = buscar_secundario(nc, fro_crudo)
        if not ec and not ef:
            continue
        aprox = bool(via and via != "directo")
        sc = int(ec["stock"]) if ec else None
        sf = int(ef["stock"]) if ef else None
        if ec:
            n_cur += 1
            via_cnt[via] = via_cnt.get(via, 0) + 1
        if ef:
            n_fro += 1
        if aprox:
            n_aprox += 1
        desc = (ec or ef or {}).get("desc")
        precio = (ec or ef or {}).get("precio")
        por_bodega = {}
        for src in (ec, ef):
            for b, q in (src or {}).get("porBodega", {}).items():
                por_bodega[b] = por_bodega.get(b, 0) + q
        bodegas = [{"n": b, "q": int(q)} for b, q in
                   sorted(por_bodega.items(), key=lambda kv: -kv[1]) if q > 0][:5]
        item = {
            "c": sc, "f": sf, "desc": desc, "precio": precio,
            "bodegas": bodegas,
            "aprox": aprox,
        }
        if alt:
            item["alt"] = alt          # SKU alternativo bajo el que está el stock
            item["via"] = via          # 'producto' | 'difuso' | 'equivalente'
        ap = aplic_map.get(nc) or (aplic_map.get(norm(alt)) if alt else None)
        if ap:
            item["aplica"] = ap        # modelos a los que aplica (referencia Ford)
        items[nc] = item

    # fecha del snapshot (mtime del archivo Curifor)
    try:
        import datetime
        ts = os.path.getmtime(os.path.join(FUENTE, ARCHIVO_CURIFOR))
        fecha = datetime.datetime.fromtimestamp(ts).strftime("%d-%m-%Y %H:%M")
    except Exception:
        fecha = "desconocida"

    salida = {
        "actualizado": fecha,
        "fuentes": {"curifor": ARCHIVO_CURIFOR, "frontera": ARCHIVO_FRONTERA},
        "items": items,
    }
    os.makedirs(DATA, exist_ok=True)
    with open(os.path.join(DATA, "stock.json"), "w", encoding="utf-8") as f:
        json.dump(salida, f, ensure_ascii=False, separators=(",", ":"))

    con_stock = sum(1 for v in items.values() if (v["c"] or 0) > 0 or (v["f"] or 0) > 0)
    rep = [
        "# Reporte de stock", "",
        f"- Snapshot: **{fecha}**",
        f"- Códigos de repuestos en pautas (reales): **{len(usados)}**",
        f"- Con registro en stock: **{len(items)}** (Curifor {n_cur}, Frontera {n_fro})",
        f"- Cruce: directo **{via_cnt.get('directo', 0)}**, por nombre/producto **{via_cnt.get('producto', 0)}**, "
        f"difuso **{via_cnt.get('difuso', 0)}**, equivalente/supersesión **{via_cnt.get('equivalente', 0)}**",
        f"- Con stock disponible (>0): **{con_stock}**",
        f"- Sin catalogar en stock: **{len(usados) - len(items)}**",
        "",
        "La plataforma marca cada repuesto con su disponibilidad y bodega. Cuando el SKU de la "
        "pauta difiere del de bodega (lubricantes, presentaciones, supersesión), se muestra el "
        "código alternativo bajo el que está el stock. El mapeo manual de lubricantes vive en "
        "`herramientas/equivalencias.json` (editable por Servicio).",
    ]
    with open(os.path.join(AQUI, "stock_reporte.md"), "w", encoding="utf-8") as f:
        f.write("\n".join(rep))

    print(f"OK: stock.json con {len(items)} códigos ({con_stock} con stock >0). "
          f"Sin catalogar: {len(usados) - len(items)}.")
    print(f"    cruce -> directo {via_cnt.get('directo',0)}, nombre {via_cnt.get('producto',0)}, "
          f"difuso {via_cnt.get('difuso',0)}, equivalente {via_cnt.get('equivalente',0)}")


if __name__ == "__main__":
    main(descargar="--descargar" in sys.argv)
