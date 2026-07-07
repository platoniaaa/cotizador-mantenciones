# -*- coding: utf-8 -*-
"""
Generador de datos del Cotizador de Mantenciones Curifor.

Lee las 8 pautas de mantención en Excel (carpeta padre del proyecto) y produce:
  - data/indice.json                     catálogo marca -> modelo -> versión
  - data/pautas/<id>.json                detalle por versión
  - herramientas/validacion.md           reporte de parseo y cuadratura de totales

Re-ejecutar cada vez que cambien las pautas:  python herramientas/generar_datos.py
"""
import json
import os
import re
import sys
import unicodedata
from difflib import SequenceMatcher

BASE = os.path.normpath(os.path.join(os.path.dirname(os.path.abspath(__file__)), "..", ".."))
OUT = os.path.normpath(os.path.join(os.path.dirname(os.path.abspath(__file__)), "..", "data"))

# ----------------------------------------------------------------------------- utilidades

def norm(s):
    """minúsculas, sin tildes, espacios colapsados"""
    if s is None:
        return ""
    s = str(s)
    s = unicodedata.normalize("NFD", s)
    s = "".join(c for c in s if unicodedata.category(c) != "Mn")
    return re.sub(r"\s+", " ", s).strip().lower()


def slug(s):
    s = norm(s)
    s = re.sub(r"[^a-z0-9]+", "-", s).strip("-")
    return s or "x"


def to_num(v):
    """Convierte celda a float; None si no es numérica (maneja #REF!, texto, etc.)."""
    if v is None:
        return None
    if isinstance(v, bool):
        return None
    if isinstance(v, (int, float)):
        return float(v)
    t = str(v).strip()
    if not t or t.startswith("#"):
        return None
    t = t.replace("$", "").replace(".", "").replace(",", ".").replace(" ", "")
    try:
        return float(t)
    except ValueError:
        return None


def rnd(v, nd=0):
    if v is None:
        return None
    return round(v, nd) if nd else int(round(v))


def similar(a, b):
    return SequenceMatcher(None, norm(a), norm(b)).ratio()


def texto(v):
    return str(v).strip() if v is not None and str(v).strip() else None


LOG = []          # líneas del reporte
CHEQUEOS = []     # (marca, version, intervalo, total_hoja, total_calculado, ok)


def log(nivel, msg):
    LOG.append(f"- **{nivel}** {msg}")


# ----------------------------------------------------------------------------- configuración

ARCHIVOS = {
    "baic":     ("Pauta de mantencion Baic Enero 2026.xlsx", "BAIC"),
    "jac":      ("Pauta de mantencion JAC Enero 2026.xlsx", "JAC"),
    "jim":      ("Pauta de mantencion JIM Enero 2026.xlsx", "JIM"),
    "mahindra": ("Pauta de mantencion Mahindra Enero 2026.xlsx", "Mahindra"),
    "shineray": ("Pauta de mantencion Shineray Enero 2026.xlsx", "Shineray"),
    "swm":      ("Pauta de mantencion Swm Enero 2026.xlsx", "SWM"),
}
OMODA_FILE = "Pauta Mantencion OMODA JAECOO Julio 2025 (1).xlsb"
FORD_FILE = "Pauta Servicio Ford - 17-06-2026.xlsm"

# hojas que NO son de costos por modelo
EXCLUIR = re.compile(
    r"(inicio|tabla|resumen|maestro|lubricantes|valores|pauta|hoja\d*|lista|mpp|fabrica|old)", re.I
)

# prefijos de modelo conocidos (para separar modelo de versión)
MODELOS_MULTIPALABRA = {
    "mahindra": ["PIK UP", "XUV 300", "XUV300", "XUV500", "KUV 100", "KUV100", "SCORPIO", "3XO"],
    "baic": ["X55 PLUS", "BJ40P", "X55", "X35", "X25", "X7", "PLUS", "UP"],
    "shineray": ["T30-T32-X30", "T30-T32", "T50-T52", "X30L", "X30"],
}

SEGMENTOS = ["suv alto", "suv medio", "suv bajo", "camioneta", "comercial", "hatchback", "sedan", "city car"]

NOTAS_DEFECTO = [
    "El valor de las mantenciones es un valor referencial sugerido; confírmalo con tu concesionario Curifor.",
    "En el costo de las mantenciones no se incluyen reparaciones ni piezas de desgaste (pastillas de freno, neumáticos, plumillas, etc.).",
    "Se recomienda su realización según kilometraje o tiempo, lo que ocurra primero.",
]


# ----------------------------------------------------------------------------- familia estándar

def celdas(ws, max_row=120, max_col=30):
    """Matriz de valores (lista de listas), 0-indexada."""
    filas = []
    for row in ws.iter_rows(min_row=1, max_row=max_row, max_col=max_col, values_only=True):
        filas.append(list(row))
    return filas


def _leer_eje(fila, en_miles):
    """Intenta leer un eje de km desde una fila. Devuelve (cols, etiquetas)."""
    cols, etiquetas = {}, {}
    ultimo = 0
    for c in range(2, len(fila)):
        v = fila[c]
        n = to_num(v)
        t = norm(v)
        if n is not None and n > 0:
            if en_miles or n <= 300:
                if n != int(n) or n > 500:
                    continue  # en miles: 5, 10, 20 ... 300
                km = n * 1000
            else:
                if n > 300000 or n != int(n) or int(n) % 500 != 0:
                    continue  # km crudos: múltiplos de 500
                km = n
            if km < ultimo:  # eje secundario o columna "Total"
                break
            ultimo = km
            cols[c] = int(km)
            etiquetas[c] = f"{int(km):,}".replace(",", ".") + " km"
        elif t and ("dia" in t or "gratis" in t):
            cols[c] = None
            etiquetas[c] = "30 días" if "dia" in t else "Entrega"
        elif t and ("total" in t or "mes" in t):
            break
    return cols, etiquetas


def encontrar_eje_km(filas):
    """Busca la fila del eje de kilometraje. Devuelve (idx_fila, {col: km|None}, {col: etiqueta})."""
    for i, fila in enumerate(filas[:15]):
        etiqueta_fila = " ".join(norm(v) for v in fila[:12] if v is not None)
        es_eje = "kilometraje" in etiqueta_fila or "kilometros" in etiqueta_fila
        if not es_eje:
            continue
        en_miles = "1.000" in etiqueta_fila or "x 1000" in etiqueta_fila
        cols, etiquetas = _leer_eje(fila, en_miles)
        if len([k for k in cols.values() if k]) >= 4:
            return i, cols, etiquetas
        # etiqueta en una fila y valores en la siguiente (p. ej. JAC)
        if i + 1 < len(filas):
            cols, etiquetas = _leer_eje(filas[i + 1], en_miles)
            if len([k for k in cols.values() if k]) >= 4:
                return i + 1, cols, etiquetas
    # sin etiqueta: fila con secuencia creciente de kilometrajes
    for i, fila in enumerate(filas[:12]):
        cols, etiquetas = _leer_eje(fila, False)
        kms = [k for k in cols.values() if k]
        if len(kms) >= 5 and kms == sorted(kms) and min(kms) >= 1000:
            return i, cols, etiquetas
    return None, {}, {}


def partes_item(fila, primera_col_km):
    """Extrae (nombre, codigo, cantidad, precio_unitario) de las celdas previas al eje km."""
    previas = [(c, fila[c]) for c in range(0, primera_col_km) if texto(fila[c]) is not None or to_num(fila[c]) is not None]
    if not previas:
        return None, None, None, None
    nombre = None
    codigo = None
    numeros = []
    for c, v in previas:
        n = to_num(v)
        t = texto(v)
        if nombre is None and t and n is None:
            nombre = t
        elif n is not None:
            numeros.append(n)
        elif t and codigo is None:
            codigo = t
    if nombre is None:
        return None, None, None, None
    # código puramente numérico (típico de lubricantes: 104406, 55004597301):
    # el patrón es nombre | código | cantidad | precio, así que si no hay código
    # de texto y hay 3+ números con el primero entero y "grande", ese es el código.
    if codigo is None and len(numeros) >= 3 and numeros[0] == int(numeros[0]) and numeros[0] >= 10000:
        codigo = str(int(numeros[0]))
        numeros = numeros[1:]
    precio = numeros[-1] if numeros else None
    cantidad = numeros[-2] if len(numeros) >= 2 else 1
    if cantidad is not None and cantidad > 400:  # dos precios seguidos, no hay cantidad real
        cantidad = 1
    return nombre, codigo, cantidad, precio


def parsear_hoja_estandar(ws, marca_id, tarifa_defecto=40000):
    filas = celdas(ws)
    idx_eje, km_cols, km_etiquetas = encontrar_eje_km(filas)
    if idx_eje is None:
        return None
    primera_col = min(km_cols)
    orden_cols = sorted(km_cols)

    meta = {"segmento": None, "categoria": None, "vigencia": None, "tarifa": None, "nombre": None}
    for fila in filas[:3]:
        for v in fila[:16]:
            t = texto(v)
            if not t:
                continue
            nt = norm(t)
            if nt in ("activo",) or re.match(r"^hasta \d{4}$", nt):
                meta["vigencia"] = t
            elif nt in SEGMENTOS:
                meta["segmento"] = t
            elif nt in ("alta", "media", "baja"):
                meta["categoria"] = t
            n = to_num(v)
            if n and 20000 <= n <= 120000 and meta["tarifa"] is None:
                meta["tarifa"] = rnd(n)
    nombre_hoja = texto(filas[0][0]) or texto(filas[1][0]) or ws.title.strip()
    if norm(nombre_hoja).startswith("costo") or norm(nombre_hoja).startswith("costos"):
        nombre_hoja = ws.title.strip()
    meta["nombre"] = nombre_hoja

    mano_obra = {}
    items = []          # (tipo, nombre, codigo, cant, unit, {col: subtotal})
    total_iva = {}
    total_neto = {}
    descuentos = {}
    adicionales = []
    notas = []
    seccion = "repuesto"
    en_notas = False

    i = idx_eje + 1
    fin_tabla = False
    while i < len(filas):
        fila = filas[i]
        primeras = " ".join(norm(fila[c]) for c in range(0, min(5, len(fila))) if fila[c] is not None)
        primera = norm(fila[0]) or norm(fila[1] if len(fila) > 1 else "")

        if en_notas:
            t = texto(fila[0]) or texto(fila[1] if len(fila) > 1 else None)
            if t and re.match(r"^\d", t.strip()):
                notas.append(re.sub(r"^\d+\.\-?\s*", "", t.strip()))
            elif t and "recomienda" in norm(t):
                notas.append(t.strip())

        if "alineacion" in primeras and "balanceo" in primeras:
            # eje propio: fila anterior con kilometros
            eje2 = filas[i - 1] if i > 0 else []
            en_miles2 = "1.000" in norm(eje2[0] if eje2 else "") or True
            por_km = {}
            precio_ali = None
            for c in range(2, min(len(fila), 25)):
                n = to_num(fila[c])
                k = to_num(eje2[c]) if c < len(eje2) else None
                if n and n > 1000 and k:
                    km2 = int(k * 1000 if k <= 300 else k)
                    por_km[str(km2)] = rnd(n)
                    precio_ali = rnd(n)
            if precio_ali:
                adicionales.append({"nombre": "Alineación y balanceo", "precio": precio_ali, "porKm": por_km})
            i += 1
            continue

        if not fin_tabla:
            if primera.startswith("total sin iva") or primera.startswith("total neto"):
                for c in orden_cols:
                    total_neto[c] = to_num(fila[c]) if c < len(fila) else None
                fin_tabla = True
                i += 1
                continue
            if re.match(r"^total(\s|$)|^total iva", primera):
                for c in orden_cols:
                    total_iva[c] = to_num(fila[c]) if c < len(fila) else None
                i += 1
                continue
            if "descuento" in primeras:
                for c in orden_cols:
                    descuentos[c] = to_num(fila[c]) if c < len(fila) else None
                i += 1
                continue
            if "mano de obra" in primera or primera == "mano obra":
                nums_previos = [to_num(fila[c]) for c in range(0, primera_col)]
                nums_previos = [n for n in nums_previos if n and 20000 <= n <= 120000]
                if nums_previos:
                    meta["tarifa"] = rnd(nums_previos[-1])
                for c in orden_cols:
                    mano_obra[c] = to_num(fila[c]) if c < len(fila) else None
                i += 1
                continue
            if re.match(r"^lubricantes", primera):
                seccion = "lubricante"
                i += 1
                continue
            if re.match(r"^(repuestos|revision)", primera):
                seccion = "repuesto"
                i += 1
                continue
            if primera.startswith("nota"):
                en_notas = True
                fin_tabla = True
                i += 1
                continue
            if "valores incluyen iva" in primeras and "descuento" not in primeras:
                i += 1
                continue
            if primera.startswith("horas") or primera == "":
                i += 1
                continue
            if "tempario" in primera or "valor mano de obra" in primera:
                # tarifa u horas: si es "valor mano de obra" con valores por km, capturar
                if "valor mano de obra" in primera:
                    nums_previos = [to_num(fila[c]) for c in range(0, primera_col)]
                    nums_previos = [n for n in nums_previos if n and 20000 <= n <= 120000]
                    if nums_previos:
                        meta["tarifa"] = rnd(nums_previos[-1])
                    vals = {c: to_num(fila[c]) for c in orden_cols if c < len(fila)}
                    if any(v for v in vals.values()):
                        mano_obra = vals
                i += 1
                continue

            nombre, codigo, cant, unit = partes_item(fila, primera_col)
            nn = norm(nombre) if nombre else ""
            ignorar = ("codigo", "cantidad", "kilometraje", "horas", "tempario",
                       "revision", "ajuste", "meses")
            if nombre and not any(nn.startswith(x) for x in ignorar):
                aplicacion = {}
                for c in orden_cols:
                    n = to_num(fila[c]) if c < len(fila) else None
                    if n and n > 100:  # descarta horas/factores; precios reales son miles
                        aplicacion[c] = n
                tipo = "material" if "material" in nn else seccion
                if aplicacion:
                    items.append((tipo, nombre, codigo, cant, unit, aplicacion))
        else:
            if primera.startswith("nota"):
                en_notas = True
        i += 1

    if not mano_obra and not items:
        return None

    tarifa = meta["tarifa"] or tarifa_defecto
    intervalos = []
    n_rev = 0
    for c in orden_cols:
        km = km_cols[c]
        mo = mano_obra.get(c)
        its = []
        suma = mo or 0
        for tipo, nombre, codigo, cant, unit, aplicacion in items:
            if c in aplicacion:
                sub = aplicacion[c]
                suma += sub
                its.append({
                    "tipo": tipo, "nombre": nombre, "codigo": codigo,
                    "cantidad": rnd(cant, 2) if cant else None,
                    "precioUnitario": rnd(unit) if unit else None,
                    "subtotal": rnd(sub),
                })
        tot = total_iva.get(c)
        gratis = False
        if tot is None:
            if km is None or (mo in (None, 0) and not its):
                gratis = km is None
                tot = 0 if gratis else (rnd(suma) if suma else None)
            else:
                tot = rnd(suma)
        if tot == 0 and not its:
            gratis = True
        if tot is None and mo is None and not its:
            continue  # columna rota (#REF!)
        n_rev += 1
        neto = total_neto.get(c)
        dcto = descuentos.get(c)
        meses = None
        if km and km % 10000 == 0:
            meses = int(km / 10000 * 12)
        intervalos.append({
            "n": n_rev,
            "km": km,
            "etiqueta": km_etiquetas[c],
            "meses": meses,
            "horas": rnd((mo or 0) / tarifa, 1) if mo else None,
            "manoObra": rnd(mo) if mo else 0,
            "items": its,
            "totalConIva": rnd(tot) if tot is not None else None,
            "totalNeto": rnd(neto) if neto else None,
            "conDescuento": rnd(dcto) if dcto else None,
            "gratis": gratis,
            "operaciones": None,
            "totalCalculado": rnd(suma),
        })

    return {
        "nombreHoja": ws.title.strip(),
        "nombre": meta["nombre"],
        "segmento": meta["segmento"],
        "categoria": meta["categoria"],
        "vigencia": meta["vigencia"] or "Activo",
        "tarifaMO": tarifa,
        "intervalos": intervalos,
        "adicionales": adicionales,
        "notas": notas or list(NOTAS_DEFECTO),
    }


def parsear_hoja_pauta(ws):
    """Hoja de actividades I/R. Devuelve {km: [{nombre, accion}]} o None."""
    filas = celdas(ws, max_row=100, max_col=30)
    idx_eje, km_cols = None, {}
    for i, fila in enumerate(filas[:12]):
        tf = " ".join(norm(v) for v in fila[:4] if v is not None)
        if "kilometros" in tf or "kilometraje" in tf:
            cols = {}
            ultimo = 0
            for c in range(1, len(fila)):
                n = to_num(fila[c])
                if n is not None and n > 0:
                    km = n * 1000 if n <= 300 else n
                    if km < ultimo:
                        break
                    ultimo = km
                    cols[c] = int(km)
            if len(cols) >= 4:
                idx_eje, km_cols = i, cols
                break
    if idx_eje is None:
        return None

    meses_cols = {}
    for j in (idx_eje - 1, idx_eje + 1):
        if 0 <= j < len(filas):
            tf = " ".join(norm(v) for v in filas[j][:4] if v is not None)
            if "mes" in tf:
                for c, km in km_cols.items():
                    m = to_num(filas[j][c]) if c < len(filas[j]) else None
                    if m:
                        meses_cols[km] = int(m)

    actividades = {km: [] for km in km_cols.values()}
    marcas_validas = {"R", "I", "A", "L", "T", "C", "AJ"}
    for fila in filas[idx_eje + 1:]:
        nombre = None
        for c in range(0, min(km_cols), 1):
            t = texto(fila[c])
            if t and to_num(fila[c]) is None and len(t) > 3:
                nombre = t
                break
        if not nombre:
            continue
        non = norm(nombre)
        if any(non.startswith(x) for x in ("tempario", "total", "mano de obra", "n°", "meses", "revision/")):
            continue
        marcas = {}
        for c, km in km_cols.items():
            v = texto(fila[c]) if c < len(fila) else None
            if v and v.strip().upper() in marcas_validas:
                marcas[km] = v.strip().upper()
        if marcas:
            for km, m in marcas.items():
                actividades[km].append({"nombre": nombre.strip().rstrip(","), "accion": m})
    if not any(actividades.values()):
        return None
    return {"actividades": actividades, "meses": meses_cols}


def procesar_estandar(marca_id, archivo, marca_nombre):
    import openpyxl
    ruta = os.path.join(BASE, archivo)
    wb = openpyxl.load_workbook(ruta, read_only=True, data_only=True)
    versiones = []
    pautas = {}
    for ws in wb.worksheets:
        nombre = ws.title.strip()
        es_pauta = norm(nombre).startswith("pauta")
        if es_pauta and ws.sheet_state == "visible":
            try:
                p = parsear_hoja_pauta(ws)
                if p:
                    pautas[nombre] = p
            except Exception as e:
                log("WARN", f"{marca_nombre}: hoja pauta '{nombre}' no parseada: {e}")
            continue
        if EXCLUIR.search(nombre) or ws.sheet_state != "visible":
            continue
        try:
            v = parsear_hoja_estandar(ws, marca_id)
            if v:
                versiones.append(v)
            else:
                log("WARN", f"{marca_nombre}: hoja '{nombre}' sin estructura reconocible, omitida")
        except Exception as e:
            log("ERROR", f"{marca_nombre}: hoja '{nombre}' falló: {e}")
    wb.close()

    # emparejar pautas de actividades con versiones
    for v in versiones:
        objetivo = norm(v["nombreHoja"]) + " " + norm(v["nombre"] or "")
        mejor, score = None, 0.45
        for pn, p in pautas.items():
            s = similar(re.sub(r"pauta", "", norm(pn)), v["nombreHoja"])
            s2 = similar(re.sub(r"pauta", "", norm(pn)), norm(v["nombre"] or ""))
            s = max(s, s2)
            if s > score:
                mejor, score = pn, s
        if mejor:
            p = pautas[mejor]
            for itv in v["intervalos"]:
                if itv["km"] and itv["km"] in p["actividades"]:
                    itv["operaciones"] = p["actividades"][itv["km"]]
                    if p["meses"].get(itv["km"]):
                        itv["meses"] = p["meses"][itv["km"]]
            v["hojaPauta"] = mejor
        else:
            v["hojaPauta"] = None
    n_con_pauta = sum(1 for v in versiones if v.get("hojaPauta"))
    log("INFO", f"{marca_nombre}: {len(versiones)} versiones, {len(pautas)} hojas de actividades, {n_con_pauta} emparejadas")
    return versiones


def modelo_de(nombre, marca_id, marca_nombre):
    """Separa 'modelo' del nombre de versión."""
    n = nombre.strip()
    # quitar prefijo de marca
    n = re.sub(r"^" + re.escape(marca_nombre) + r"\s+", "", n, flags=re.I)
    mayus = n.upper()
    for pref in MODELOS_MULTIPALABRA.get(marca_id, []):
        if mayus.startswith(pref):
            return n[: len(pref)].strip()
    return n.split()[0] if n.split() else n


# ----------------------------------------------------------------------------- omoda / jaecoo

# tipo de repuesto Omoda por columna en la hoja REPUESTOS 2025
OMODA_COL_REP = {
    2: "filtro_aire", 3: "filtro_aceite", 4: "filtro_combustible", 5: "filtro_ac",
    6: "bujia", 7: "correa", 8: "pastilla_del", 9: "pastilla_tra", 10: "filtro_transmision",
}


def clasificar_repuesto_omoda(nombre):
    n = norm(nombre)
    if "filtro" in n and ("transmision" in n or "cvt" in n or "caja" in n or "dct" in n):
        return "filtro_transmision"
    if "filtro" in n and "aceite" in n:
        return "filtro_aceite"
    if "filtro" in n and "aire" in n:
        return "filtro_aire"
    if "filtro" in n and ("a/c" in n or "a c" in n or "polen" in n or "cabina" in n or n.endswith("ac")):
        return "filtro_ac"
    if "filtro" in n and "combustible" in n:
        return "filtro_combustible"
    if "bujia" in n:
        return "bujia"
    if "correa" in n:
        return "correa"
    return None


def leer_repuestos_omoda(wb):
    """{modeloNorm: {tipo_repuesto: codigo}} desde la hoja REPUESTOS 2025."""
    if "REPUESTOS 2025" not in wb.sheets:
        return {}
    with wb.get_sheet("REPUESTOS 2025") as ws:
        filas = []
        for row in ws.rows():
            fila = [None] * 15
            for c in row[:15]:
                if c.c is not None and c.c < 15:
                    fila[c.c] = c.v
            filas.append(fila)
    mapa = {}
    modelo_actual = None
    for f in filas:
        etq0 = texto(f[0])
        etq1 = norm(f[1]) if len(f) > 1 else ""
        if etq0 and not etq1:
            modelo_actual = etq0
        if etq1 == "codigo" and modelo_actual:
            d = {}
            for col, tipo in OMODA_COL_REP.items():
                cod = texto(f[col]) if col < len(f) else None
                if cod and cod.upper() not in ("N/A", "-"):
                    d[tipo] = cod
            if d:
                mapa[norm(modelo_actual)] = d
    return mapa


def procesar_omoda():
    from pyxlsb import open_workbook
    ruta = os.path.join(BASE, OMODA_FILE)
    versiones = []
    packs_por_version = {}

    with open_workbook(ruta) as wb:
        repuestos_map = leer_repuestos_omoda(wb)
        # MPP: packs por versión
        if "MPP" in wb.sheets:
            with wb.get_sheet("MPP") as ws:
                filas = [[c.v for c in row] for row in ws.rows()]
            encabezado = None
            for i, f in enumerate(filas):
                if f and norm(f[0]) == "marca":
                    encabezado = i
                    break
            if encabezado is not None:
                for f in filas[encabezado + 1:]:
                    if not f or not texto(f[0] if len(f) > 0 else None):
                        continue
                    marca = texto(f[0])
                    modelo = texto(f[1]) if len(f) > 1 else None
                    version = texto(f[2]) if len(f) > 2 else None
                    if not (marca and version):
                        continue
                    packs = []
                    for idx, nombre_pack in ((9, "Pack 1 (1ª y 2ª mantención)"),
                                             (10, "Pack 2 (1ª a 3ª mantención)"),
                                             (11, "Pack 3 (1ª a 4ª mantención)")):
                        p = to_num(f[idx]) if len(f) > idx else None
                        if p and p > 0:
                            packs.append({"nombre": nombre_pack, "precio": rnd(p)})
                    limite = texto(f[12]) if len(f) > 12 else None
                    packs_por_version[f"{marca} {modelo} {version}"] = {"packs": packs, "limite": limite}

        for nombre in wb.sheets:
            if not (nombre.upper().startswith("OMODA") or nombre.upper().startswith("JAECOO")):
                continue
            with wb.get_sheet(nombre) as ws:
                filas = []
                for row in ws.rows():
                    fila = [None] * 22
                    for c in row[:22]:
                        if c.c is not None and c.c < 22:
                            fila[c.c] = c.v
                    filas.append(fila)
                    if len(filas) > 70:
                        break
            v = parsear_modelo_omoda(nombre, filas, packs_por_version, repuestos_map)
            if v:
                versiones.append(v)
            else:
                log("WARN", f"Omoda/Jaecoo: hoja '{nombre}' sin estructura reconocible")
    log("INFO", f"Omoda/Jaecoo: {len(versiones)} versiones")
    return versiones


def fila_omoda(filas, patron, max_row=30):
    for i, f in enumerate(filas[:max_row]):
        for c in range(0, 4):
            if c < len(f) and re.search(patron, norm(f[c]) if f[c] is not None else ""):
                return i
    return None


def parsear_modelo_omoda(nombre_hoja, filas, packs_map, repuestos_map=None):
    repuestos_map = repuestos_map or {}
    i_km = fila_omoda(filas, r"^kilometros")
    i_mo = fila_omoda(filas, r"^mano de obra$|^mano de obra")
    if i_km is None or i_mo is None:
        return None
    i_hrs = fila_omoda(filas, r"tiempo")
    i_mes = fila_omoda(filas, r"^mes$")
    i_neto = fila_omoda(filas, r"^neto")
    i_iva = fila_omoda(filas, r"^total con iva")

    km_cols = {}
    for c in range(2, 22):
        n = to_num(filas[i_km][c]) if c < len(filas[i_km]) else None
        if n and n > 0:
            km_cols[c] = int(n * 1000 if n <= 300 else n)
    if len(km_cols) < 4:
        return None
    orden = sorted(km_cols)

    nombre = None
    for f in filas[:4]:
        for c in range(0, 4):
            t = texto(f[c]) if c < len(f) else None
            if t and len(t) > 4 and to_num(f[c]) is None:
                nombre = t
                break
        if nombre:
            break
    nombre = nombre or nombre_hoja

    # items: entre REPUESTOS y NETO, saltando TOTAL REPUESTOS / LUBRICANTES headers
    items = []
    seccion = "repuesto"
    tope = i_neto if i_neto is not None else (i_iva or len(filas))
    for i in range(i_mo + 1, tope):
        f = filas[i]
        etiqueta = None
        for c in range(0, min(orden)):
            t = texto(f[c]) if c < len(f) else None
            if t:
                etiqueta = t
                break
        if not etiqueta:
            continue
        ne = norm(etiqueta)
        if ne.startswith("repuestos"):
            seccion = "repuesto"
            continue
        if ne.startswith("lubricantes"):
            seccion = "lubricante"
            continue
        if ne.startswith("total"):
            continue
        aplicacion = {}
        for c in orden:
            n = to_num(f[c]) if c < len(f) else None
            if n and n > 0:
                aplicacion[c] = n
        if aplicacion:
            tipo = "material" if "insumo" in ne else seccion
            items.append((tipo, etiqueta, aplicacion))

    # códigos de repuestos por tipo (hoja REPUESTOS 2025), match difuso del modelo
    codigos_rep = {}
    if repuestos_map:
        mejor, sc = None, 0.55
        for k, d in repuestos_map.items():
            s = max(similar(k, nombre), similar(k, nombre_hoja))
            if s > sc:
                mejor, sc = k, s
        if mejor:
            codigos_rep = repuestos_map[mejor]

    # operaciones (TEMPARIO & OPERACIONES): filas con minutos por km
    i_ops = fila_omoda(filas, r"tempario", max_row=60)
    operaciones = {km: [] for km in km_cols.values()}
    if i_ops is not None:
        # eje de esta sección
        eje2 = None
        for i in range(i_ops, min(i_ops + 4, len(filas))):
            for c in range(2, 22):
                if to_num(filas[i][c]) and to_num(filas[i][c]) <= 300:
                    eje2 = i
                    break
            if eje2:
                break
        if eje2 is not None:
            cols2 = {}
            for c in range(2, 22):
                n = to_num(filas[eje2][c])
                if n and n <= 300:
                    cols2[c] = int(n * 1000)
            for f in filas[eje2 + 1:]:
                nombre_op = None
                for c in range(0, min(cols2) if cols2 else 3):
                    t = texto(f[c]) if c < len(f) else None
                    if t and to_num(f[c]) is None:
                        nombre_op = t
                        break
                if not nombre_op:
                    continue
                non = norm(nombre_op)
                if any(non.startswith(x) for x in ("tempario", "operacion", "total", "mano de obra", "tiempo")):
                    continue  # filas de tiempo/labor, no operaciones
                accion = "R" if non.startswith("cambio") else "I"
                for c, km in cols2.items():
                    if c < len(f) and to_num(f[c]):
                        if km in operaciones:
                            operaciones[km].append({"nombre": nombre_op, "accion": accion})

    es_ev = "EV" in nombre_hoja.upper()
    tarifa = 60000 if es_ev else 45000
    intervalos = []
    for idx, c in enumerate(orden, 1):
        km = km_cols[c]
        mo = to_num(filas[i_mo][c]) if c < len(filas[i_mo]) else None
        hrs = to_num(filas[i_hrs][c]) if i_hrs is not None and c < len(filas[i_hrs]) else None
        meses = to_num(filas[i_mes][c]) if i_mes is not None and c < len(filas[i_mes]) else None
        neto = to_num(filas[i_neto][c]) if i_neto is not None and c < len(filas[i_neto]) else None
        con_iva = to_num(filas[i_iva][c]) if i_iva is not None and c < len(filas[i_iva]) else None
        its = []
        suma = mo or 0
        for tipo, etq, aplicacion in items:
            if c in aplicacion:
                suma += aplicacion[c]
                clase = clasificar_repuesto_omoda(etq)
                codigo = codigos_rep.get(clase) if clase else None
                its.append({"tipo": tipo, "nombre": etq, "codigo": codigo, "cantidad": None,
                            "precioUnitario": None, "subtotal": rnd(aplicacion[c])})
        if con_iva is None and neto is not None:
            con_iva = neto * 1.19
        if con_iva is None:
            continue
        intervalos.append({
            "n": idx, "km": km,
            "etiqueta": f"{km:,}".replace(",", ".") + " km",
            "meses": int(meses) if meses else (int(km / 10000 * 12) if km % 10000 == 0 else None),
            "horas": rnd(hrs, 2) if hrs else None,
            "manoObra": rnd(mo) if mo else 0,
            "items": its,
            "totalConIva": rnd(con_iva),
            "totalNeto": rnd(neto) if neto else None,
            "conDescuento": None,
            "gratis": False,
            "operaciones": operaciones.get(km) or None,
            "totalCalculado": rnd(suma * 1.19),
        })

    # packs por matching de nombre
    packs, limite = [], None
    mejor, score = None, 0.55
    for k, v in packs_map.items():
        s = similar(k, nombre)
        s2 = similar(k, nombre_hoja)
        s = max(s, s2)
        if s > score:
            mejor, score = k, s
    if mejor:
        packs = packs_map[mejor]["packs"]
        limite = packs_map[mejor]["limite"]

    marca_nombre = "OMODA" if nombre_hoja.upper().startswith("OMODA") else "JAECOO"
    adicionales = []
    for f in filas:
        for c in range(0, 4):
            t = texto(f[c]) if c < len(f) else None
            if t and "alineacion" in norm(t):
                precios = [to_num(x) for x in f if to_num(x) and to_num(x) > 5000]
                if precios:
                    adicionales.append({"nombre": t if len(t) < 60 else "Alineación y rotación",
                                        "precio": rnd(precios[0]), "porKm": {}})
                break

    return {
        "nombreHoja": nombre_hoja,
        "nombre": nombre,
        "marcaNombre": marca_nombre,
        "segmento": "SUV",
        "categoria": None,
        "vigencia": "Activo",
        "tarifaMO": tarifa,
        "intervalos": intervalos,
        "adicionales": adicionales,
        "packs": packs,
        "limitePrimeraMant": limite,
        "notas": [
            "Valores con IVA incluido, referenciales y sugeridos.",
            "Las mantenciones no incluyen piezas de desgaste (pastillas de freno, neumáticos, plumillas).",
        ] + ([f"Primera mantención: {limite}."] if limite else []),
    }


# ----------------------------------------------------------------------------- ford

def procesar_ford():
    import openpyxl
    ruta = os.path.join(BASE, FORD_FILE)
    wb = openpyxl.load_workbook(ruta, read_only=True, data_only=True)

    # --- Resumen: precios por (año, modelo, versión, motor)
    ws = wb["Resumen"]
    filas = celdas(ws, max_row=260, max_col=18)
    registros = []
    rev_info = []  # [(años, km)] por columna de revisión
    cols_rev = list(range(6, 16))
    anio_m = anio_c = modelo = version = None
    for fila in filas:
        c1 = texto(fila[1])
        if c1 and norm(c1) == "año modelo" or (c1 and "ano modelo" in norm(c1)):
            # fila de encabezado: extraer etiquetas de revisiones
            rev_info = []
            for c in cols_rev:
                t = texto(fila[c])
                m = re.search(r"(\d+)\s*a[ñn]os?\s*[oó]\s*(?:hasta\s*)?([\d\.]+)\s*km", norm(t or ""))
                if m:
                    rev_info.append((int(m.group(1)), int(m.group(2).replace(".", ""))))
                else:
                    rev_info.append(None)
            anio_m = anio_c = modelo = version = None
            continue
        precios = [to_num(fila[c]) for c in cols_rev]
        n_precios = len([p for p in precios if p and p > 10000])
        if n_precios < 3:
            continue
        anio_m = to_num(fila[1]) and int(to_num(fila[1])) or anio_m
        anio_c = to_num(fila[2]) and int(to_num(fila[2])) or anio_c
        modelo = texto(fila[3]) or modelo
        version = texto(fila[4]) or version
        motor = texto(fila[5])
        if not (anio_m and modelo and version):
            continue
        registros.append({
            "anio": anio_m, "anioComercial": anio_c, "modelo": modelo,
            "version": version, "motor": motor, "precios": precios, "revInfo": list(rev_info),
        })
    # deduplicar (año, modelo, versión, motor): conservar el de año comercial más nuevo
    unicos = {}
    for r in registros:
        k = (r["anio"], norm(r["modelo"]), norm(r["version"]), norm(r["motor"] or ""))
        if k not in unicos or (r["anioComercial"] or 0) > (unicos[k]["anioComercial"] or 0):
            unicos[k] = r
    registros = list(unicos.values())

    # --- Planes: operaciones por bloque (año, modelo, versión)
    bloques = []
    for nombre in wb.sheetnames:
        if not nombre.lower().startswith("plan mantenimiento"):
            continue
        wsp = wb[nombre]
        filasp = celdas(wsp, max_row=900, max_col=17)
        tarifa = None
        for f in filasp[:5]:
            for c in range(1, 6):
                n = to_num(f[c]) if c < len(f) else None
                if n and 30000 <= n <= 90000:
                    tarifa = rnd(n)
        i = 0
        while i < len(filasp):
            f = filasp[i]
            anio = to_num(f[1])
            if anio and 2015 <= anio <= 2030 and texto(f[2]):
                b_anio, b_modelo, b_version = int(anio), texto(f[2]), texto(f[3]) or ""
                # etiquetas de revisiones en fila siguiente
                ops = []
                tiempos = None
                j = i + 1
                n_revs = 0
                if j < len(filasp):
                    n_revs = len([1 for c in range(4, 15) if texto(filasp[j][c]) and "km" in norm(filasp[j][c])])
                j += 1
                while j < len(filasp):
                    fj = filasp[j]
                    nombre_op = texto(fj[1])
                    if nombre_op and norm(nombre_op) == "tiempo":
                        tiempos = [to_num(fj[c]) for c in range(4, 4 + max(n_revs, 10))]
                        break
                    aa = to_num(fj[1])
                    if aa and 2015 <= aa <= 2030 and texto(fj[2]):
                        break  # siguiente bloque sin fila Tiempo
                    if nombre_op and texto(fj[2]):
                        marcas_op = []
                        for c in range(4, 4 + max(n_revs, 10)):
                            v = fj[c] if c < len(fj) else None
                            marcas_op.append(bool(texto(v)) and to_num(v) is None)
                        precio_unit = None
                        for c in range(13, 16):
                            n = to_num(fj[c]) if c < len(fj) else None
                            if n and n > 100:
                                precio_unit = rnd(n)
                                break
                        ops.append({"nombre": nombre_op.strip(), "codigo": texto(fj[2]),
                                    "cantidad": to_num(fj[3]), "precioUnitario": precio_unit,
                                    "marcas": marcas_op})
                    j += 1
                if ops:
                    bloques.append({"anio": b_anio, "modelo": b_modelo, "version": b_version,
                                    "ops": ops, "tiempos": tiempos, "tarifa": tarifa})
                i = j
            i += 1
    wb.close()

    # --- combinar: registro de precios + bloque de operaciones
    versiones = {}
    emparejados = 0
    for r in registros:
        vid = f"{slug(r['modelo'])}--{slug(r['version'])}-{slug(r['motor'] or '')}"
        mejor, score = None, 0.5
        for b in bloques:
            if b["anio"] != r["anio"] or norm(b["modelo"]) != norm(r["modelo"]):
                continue
            s = similar(b["version"], r["version"] + " " + (r["motor"] or ""))
            s = max(s, similar(b["version"], r["version"]))
            if s > score:
                mejor, score = b, s
        if mejor:
            emparejados += 1
        intervalos = []
        for idx, precio in enumerate(r["precios"]):
            if not precio or precio < 10000:
                continue
            info = r["revInfo"][idx] if idx < len(r["revInfo"]) else None
            anios_i, km = info if info else (idx + 1, (idx + 1) * 10000)
            ops, its, horas, mo = None, [], None, 0
            if mejor:
                its = []
                ops = []
                for op in mejor["ops"]:
                    if idx < len(op["marcas"]) and op["marcas"][idx]:
                        sub = (op["precioUnitario"] or 0) * (op["cantidad"] or 1)
                        its.append({"tipo": "repuesto", "nombre": op["nombre"], "codigo": op["codigo"],
                                    "cantidad": op["cantidad"], "precioUnitario": op["precioUnitario"],
                                    "subtotal": rnd(sub) if sub else None})
                        ops.append({"nombre": op["nombre"], "accion": "R"})
                if mejor["tiempos"] and idx < len(mejor["tiempos"]) and mejor["tiempos"][idx]:
                    horas = rnd(mejor["tiempos"][idx], 1)
                    if mejor["tarifa"]:
                        mo = rnd(horas * mejor["tarifa"])
            intervalos.append({
                "n": idx + 1, "km": km,
                "etiqueta": f"{km:,}".replace(",", ".") + " km",
                "meses": anios_i * 12,
                "horas": horas, "manoObra": mo or 0,
                "items": its,
                "totalConIva": rnd(precio),
                "totalNeto": None, "conDescuento": None, "gratis": False,
                "operaciones": ops or None,
                "totalCalculado": None,
            })
        if vid not in versiones:
            versiones[vid] = {
                "modelo": r["modelo"],
                "version": r["version"],
                "motor": r["motor"],
                "planesPorAnio": {},
            }
        versiones[vid]["planesPorAnio"][str(r["anio"])] = intervalos
    log("INFO", f"Ford: {len(registros)} filas de precios, {len(bloques)} bloques de operaciones, {emparejados} emparejados")
    return versiones


# ----------------------------------------------------------------------------- salida

def construir():
    os.makedirs(os.path.join(OUT, "pautas"), exist_ok=True)
    indice = {"actualizado": "Enero 2026 (Ford: junio 2026, Omoda/Jaecoo: julio 2025)", "marcas": []}
    total_versiones = 0

    def guardar_pauta(pid, data):
        with open(os.path.join(OUT, "pautas", pid + ".json"), "w", encoding="utf-8") as f:
            json.dump(data, f, ensure_ascii=False, separators=(",", ":"))

    # ---- familia estándar
    for marca_id, (archivo, marca_nombre) in ARCHIVOS.items():
        try:
            versiones = procesar_estandar(marca_id, archivo, marca_nombre)
        except Exception as e:
            log("ERROR", f"{marca_nombre}: archivo no procesado: {e}")
            continue
        modelos = {}
        for v in versiones:
            base_nombre = v["nombre"] if v["nombre"] and len(v["nombre"]) < 60 else v["nombreHoja"]
            modelo = modelo_de(base_nombre, marca_id, marca_nombre)
            vid = f"{marca_id}__{slug(v['nombreHoja'])}"
            pauta = {
                "marca": marca_id, "marcaNombre": marca_nombre,
                "modelo": modelo, "version": base_nombre, "motor": None,
                "segmento": v["segmento"], "categoria": v["categoria"], "vigencia": v["vigencia"],
                "tarifaMO": v["tarifaMO"], "anios": None,
                "planes": [{"anio": None, "intervalos": v["intervalos"]}],
                "adicionales": v["adicionales"], "packs": [],
                "notas": v["notas"],
                "fuente": f"{archivo} / hoja '{v['nombreHoja']}'",
            }
            guardar_pauta(vid, pauta)
            for itv in v["intervalos"]:
                if itv["totalConIva"] and itv["totalCalculado"]:
                    ok = abs(itv["totalConIva"] - itv["totalCalculado"]) <= max(3000, itv["totalConIva"] * 0.03)
                    CHEQUEOS.append((marca_nombre, v["nombreHoja"], itv["etiqueta"],
                                     itv["totalConIva"], itv["totalCalculado"], ok))
            modelos.setdefault(modelo, []).append({
                "id": vid, "nombre": base_nombre, "vigencia": v["vigencia"],
                "segmento": v["segmento"], "anios": None,
            })
            total_versiones += 1
        indice["marcas"].append({
            "id": marca_id, "nombre": marca_nombre,
            "modelos": [{"nombre": m, "versiones": vs} for m, vs in sorted(modelos.items())],
        })

    # ---- omoda / jaecoo
    try:
        versiones_oj = procesar_omoda()
    except Exception as e:
        log("ERROR", f"Omoda/Jaecoo: archivo no procesado: {e}")
        versiones_oj = []
    for marca_nombre in ("OMODA", "JAECOO"):
        marca_id = marca_nombre.lower()
        modelos = {}
        for v in versiones_oj:
            if v["marcaNombre"] != marca_nombre:
                continue
            nombre = re.sub(r"^(OMODA|JAECOO)\s+", "", v["nombre"], flags=re.I)
            modelo = nombre.split()[0] if nombre.split() else nombre
            if re.match(r"^\d", modelo):  # 'JAECOO 5' -> J5
                modelo = "J" + modelo
            vid = f"{marca_id}__{slug(v['nombreHoja'])}"
            pauta = {
                "marca": marca_id, "marcaNombre": marca_nombre,
                "modelo": modelo, "version": nombre, "motor": None,
                "segmento": v["segmento"], "categoria": None, "vigencia": "Activo",
                "tarifaMO": v["tarifaMO"], "anios": None,
                "planes": [{"anio": None, "intervalos": v["intervalos"]}],
                "adicionales": v["adicionales"], "packs": v["packs"],
                "notas": v["notas"],
                "fuente": f"{OMODA_FILE} / hoja '{v['nombreHoja']}'",
            }
            guardar_pauta(vid, pauta)
            for itv in v["intervalos"]:
                if itv["totalConIva"] and itv["totalCalculado"]:
                    ok = abs(itv["totalConIva"] - itv["totalCalculado"]) <= max(3000, itv["totalConIva"] * 0.03)
                    CHEQUEOS.append((marca_nombre, v["nombreHoja"], itv["etiqueta"],
                                     itv["totalConIva"], itv["totalCalculado"], ok))
            modelos.setdefault(modelo, []).append({
                "id": vid, "nombre": nombre, "vigencia": "Activo", "segmento": v["segmento"], "anios": None,
            })
            total_versiones += 1
        if modelos:
            indice["marcas"].append({
                "id": marca_id, "nombre": marca_nombre.capitalize() if marca_nombre != "OMODA" else "Omoda",
                "modelos": [{"nombre": m, "versiones": vs} for m, vs in sorted(modelos.items())],
            })

    # ---- ford
    try:
        versiones_ford = procesar_ford()
    except Exception as e:
        log("ERROR", f"Ford: archivo no procesado: {e}")
        versiones_ford = {}
    modelos = {}
    for vid_base, v in versiones_ford.items():
        vid = f"ford__{vid_base}"
        anios = sorted(v["planesPorAnio"].keys(), reverse=True)
        nombre_version = v["version"] + (f" · {v['motor']}" if v["motor"] else "")
        pauta = {
            "marca": "ford", "marcaNombre": "Ford",
            "modelo": v["modelo"], "version": nombre_version, "motor": v["motor"],
            "segmento": None, "categoria": None, "vigencia": None,
            "tarifaMO": 53000, "anios": anios,
            "planes": [{"anio": a, "intervalos": v["planesPorAnio"][a]} for a in anios],
            "adicionales": [], "packs": [],
            "notas": [
                "Precios sugeridos de mantención con IVA incluido, publicados por Ford Chile para la red de concesionarios.",
                "Realiza las mantenciones cada 12 meses o el kilometraje indicado, lo que ocurra primero.",
            ],
            "fuente": f"{FORD_FILE} / hoja 'Resumen'",
        }
        guardar_pauta(vid, pauta)
        modelos.setdefault(v["modelo"], []).append({
            "id": vid, "nombre": nombre_version, "vigencia": None, "segmento": None, "anios": anios,
        })
        total_versiones += 1
    if modelos:
        indice["marcas"].append({
            "id": "ford", "nombre": "Ford",
            "modelos": [{"nombre": m, "versiones": sorted(vs, key=lambda x: x["nombre"])}
                        for m, vs in sorted(modelos.items())],
        })

    with open(os.path.join(OUT, "indice.json"), "w", encoding="utf-8") as f:
        json.dump(indice, f, ensure_ascii=False, indent=1)

    # ---- reporte
    ok = sum(1 for c in CHEQUEOS if c[5])
    reporte = ["# Validación de datos generados", "",
               f"- Versiones generadas: **{total_versiones}**",
               f"- Chequeos de cuadratura (total hoja vs suma de componentes): **{ok}/{len(CHEQUEOS)}** dentro de tolerancia (3% o $3.000)",
               "", "## Log de parseo", ""] + LOG + ["", "## Diferencias fuera de tolerancia", ""]
    for m, v, e, th, tc, esta_ok in CHEQUEOS:
        if not esta_ok:
            reporte.append(f"- {m} / {v} / {e}: hoja **${th:,.0f}** vs calculado **${tc:,.0f}**".replace(",", "."))
    reporte.append("")
    reporte.append("_Nota: la plataforma siempre muestra el total oficial de la hoja Excel; el total calculado se usa solo como control._")
    with open(os.path.join(os.path.dirname(os.path.abspath(__file__)), "validacion.md"), "w", encoding="utf-8") as f:
        f.write("\n".join(reporte))

    print(f"OK: {total_versiones} versiones | chequeos {ok}/{len(CHEQUEOS)} | marcas: {len(indice['marcas'])}")
    for m in indice["marcas"]:
        nv = sum(len(mo["versiones"]) for mo in m["modelos"])
        print(f"  - {m['nombre']}: {len(m['modelos'])} modelos, {nv} versiones")


if __name__ == "__main__":
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")
    construir()
