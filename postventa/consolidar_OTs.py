# -*- coding: utf-8 -*-
# =============================================================
#
#   CONSOLIDADOR DIARIO DE ÓRDENES DE TRABAJO
#   Curifor S.A
#
#   ¿Qué hace este script?
#   ---------------------
#   Cada día, en lugar de hacer los BUSCARVs a mano, este script:
#
#   1. Lee la sábana que descargaste del PBI (datos frescos de las OTs)
#   2. Lee el archivo de cada sucursal (con sus notas y avances)
#   3. Cruza todo por FOLIO OT
#   4. Actualiza el archivo maestro con los datos combinados
#   5. Guarda un respaldo automático por si necesitas volver atrás
#
#   ¿Cómo se usa?
#   -------------
#   1. Descarga la sábana del PBI y déjala en la carpeta PBI (ver config)
#   2. Haz doble clic en "Ejecutar_Consolidacion.bat"
#   3. Listo — el archivo maestro queda actualizado
#
# =============================================================


import os
import sys
import glob
import shutil
import json
import base64
import warnings
import unicodedata
import re
import pandas as pd
import requests
from datetime import datetime
from openpyxl import load_workbook

# Suprimir advertencias de SSL en Windows (certificados desactualizados)
warnings.filterwarnings("ignore", message="Unverified HTTPS request")


# =============================================================
#   CONFIGURACIÓN
#   -------------
#   Ajusta estas rutas antes de usar el script por primera vez.
#   Solo necesitas cambiar esto una vez.
# =============================================================

# Carpeta donde dejas la sábana del PBI cada día
CARPETA_PBI = r"C:\Users\cjerez\OneDrive - Curifor S.A\Documentos\Claude\Projects\Gestión OTs Pendientes\PBI"

# Carpeta donde están los archivos de las 9 sucursales
# (el script los encuentra aunque la fecha del nombre cambie cada día)
CARPETA_SUCURSALES = r"C:\Users\cjerez\OneDrive - Curifor S.A\Documentos\Claude\Projects\Gestión OTs Pendientes\Sucursales"

# Archivo maestro que se actualiza (el que tiene el panel y las macros)
ARCHIVO_MAESTRO = r"C:\Users\cjerez\OneDrive - Curifor S.A\Documentos\Claude\Projects\Gestión OTs Pendientes\Seguimiento_BASE.xlsm"

# Carpeta de respaldos (se crea sola si no existe)
CARPETA_RESPALDOS = r"C:\Users\cjerez\OneDrive - Curifor S.A\Documentos\Claude\Projects\Gestión OTs Pendientes\Respaldos"

# Ruta del JSON que se sube al dashboard web
RUTA_JSON = r"C:\Users\cjerez\OneDrive - Curifor S.A\Documentos\Claude\Projects\Gestión OTs Pendientes\datos_dashboard.json"

# Carpeta donde dejas el archivo de Seguimiento de Compras cada mañana
# (el script lo encuentra solo, igual que el PBI; toma el .xlsx más reciente)
CARPETA_COMPRAS = r"C:\Users\cjerez\OneDrive - Curifor S.A\Documentos\Claude\Projects\Gestión OTs Pendientes\Seguimiento de Compras"

# Nombre fijo del archivo de stock de repuestos en bodega.
# Déjalo cada mañana en la misma carpeta que el Seguimiento de Compras.
# Headers en fila 1; columnas requeridas: "Producto" y "Bodega".
NOMBRE_STOCK = "Stock Repestos Costo.xlsx"

# Nombre del archivo CSV con todos los Vale de Consumo históricos (tabla maestra).
# Separador ";", columna clave: "Id Prod.". Si el código de producto en el
# Seguimiento de Compras coincide con uno de aquí, significa que ese repuesto
# ya fue consumido en alguna OT (aunque no sea visible en la app) → se excluye
# de los paneles "Repuestos Pendientes" y "Patentes a Contactar".
NOMBRE_TABLA_VC = "tabla vc.csv"

# Carpeta donde se deja el archivo de "Revisión de Campañas" (Agenda Ford) —
# el nombre trae la fecha del día, ej. "2026-07-29_Consolidado_Curifor.xlsx".
# 29/07/2026: Cristóbal movió el archivo desde la carpeta de otro proyecto
# ("Agenda Ford") a esta misma carpeta del proyecto, y de paso dejó de usar
# el sufijo "_2Tandas" en el nombre — confirmó explícitamente que la
# estructura de columnas es la MISMA que el formato "_2Tandas" anterior, solo
# cambió dónde se deja el archivo y el nombre. El patrón de búsqueda
# (`leer_campanas_curifor`) acepta ambos nombres (con y sin "_2Tandas") por
# si algún día vuelve a usarse ese sufijo, y siempre toma el archivo más
# reciente entre los que matcheen cualquiera de los 2 patrones.
CARPETA_AGENDA_FORD = r"C:\Users\cjerez\OneDrive - Curifor S.A\Documentos\Claude\Projects\Gestión OTs Pendientes"

# =============================================================
#   CONFIGURACIÓN DASHBOARD WEB (Streamlit)
#   ------------------------------------------
#   Rellena estos datos una sola vez después de crear el repo
#   en GitHub. Si no quieres el dashboard web, deja en blanco.
# =============================================================
GITHUB_USUARIO   = "Cjerez-curi"          # ej: "ragnarcj12"
GITHUB_REPO      = "curifor-ots"          # ej: "curifor-ots"
# Token leído desde archivo local (nunca subir el token directamente al código)
_token_file = os.path.join(os.path.dirname(os.path.abspath(__file__)), "github_token.txt")
GITHUB_TOKEN = open(_token_file).read().strip() if os.path.exists(_token_file) else ""
GITHUB_HISTORIAL = "historial_cierres.json"   # Archivo de registro de OTs cerradas
GITHUB_RANKING   = "ranking_cierres.json"     # Ranking de OTs cerradas con >90 días
GITHUB_AGENDA    = "agenda_hoy.json"          # Citas del día (Agenda Curifor)
GITHUB_CTRL_TALLER = "control_taller.json"    # Ordenes/bloques del Planificador de Taller (JPCB)
GITHUB_STOCK_REPUESTOS = "stock_repuestos.json"  # Catalogo completo de Stock Repestos Costo.xlsx
GITHUB_PRODUCCION_TECNICOS = "produccion_tecnicos.json"  # Horas facturadas por tecnico (BDFlexline)
GITHUB_TECNICOS_SUCURSAL_MANUAL = "tecnicos_sucursal_manual.json"  # Mapeo manual tecnico->sucursal (respaldo, Admin -> Tecnicos)
GITHUB_CAMPANAS = "campanas_curifor.json"  # Revision de Campanas (Agenda Ford)
GITHUB_CUENTA_FICHA = "cuenta_ficha.json"  # Modulo Cuenta Ficha (saldos de cliente + historial de OT)
GITHUB_CUENTA_FICHA_REV = "cuenta_ficha_revisados.json"  # Marcas "Revisado" por cliente (las escribe la App)

# --- Modulo Cuenta Ficha -------------------------------------------------
# Ventana de historial de OT por cliente (decision de Cristobal 31/07/2026:
# "ultimos 24 meses"). Los clientes flota tienen miles de OT historicas
# (un RUT real llego a 14.093), asi que traer todo infla el archivo sin
# aportar: para revisar un saldo interesa lo reciente.
CUENTA_FICHA_MESES_HISTORIAL = 24

# Documentos posteriores que se muestran por cada OT en la ficha del cliente.
# (nombre visible, columna del folio/numero en el PBI, columna de fecha)
DOCS_CUENTA_FICHA = [
    ("Liquidación ST",      "FOLIO LIQUIDACIÓN ST",       "FECHA LIQUIDACIÓN ST"),
    ("Factura Cliente",     "FOLIO FACTURA CLIENTE",      "FECHA FACTURA CLIENTE"),
    ("NC Factura Cliente",  "FOLIO NC FACTURA CLIENTE",   "FECHA NC FACTURA CLIENTE"),
    ("Factura Compañía",    "FOLIO FACTURA COMPAÑÍA",     "FECHA FACTURA COMPAÑÍA"),
    ("NC Factura Compañía", "FOLIO NC FACTURA COMPAÑÍA",  "FECHA NC FACTURA COMPAÑÍA"),
    ("Cargo Interno",       "FOLIO CARGO INTERNO",        "FECHA CARGO INTERNO"),
    ("Cargo Garantía",      "FOLIO CARGO GARANTÍA",       "FECHA CARGO GARANTÍA"),
    ("Factura Garantía",    "NÚMERO FACTURA GARANTÍA",    "FECHA FACTURA GARANTÍA"),
    ("Vale de Consumo",     "NRO VALE DE CONSUMO",        "FECHA VALE DE CONSUMO"),
    ("Nota de Crédito",     "NÚMERO NOTA DE CRÉDITO",     "FECHA NOTA DE CRÉDITO"),
    ("Refacturación",       "NÚMERO REFACTURACIÓN",       "FECHA REFACTURACIÓN"),
    ("Cierre Gerencia",     "NÚMERO CIERRE GERENCIA",     "FECHA CIERRE GERENCIA"),
    ("Multifactura",        "FOLIO SOLCITUD MULTIFACTURA", "FECHA SOLCITUD MULTIFACTURA"),
]

# Columnas del PBI que necesita Cuenta Ficha (se conserva una copia acotada de
# TODOS los estados dentro de leer_pbi, sin volver a leer el archivo de 68 MB).
COLS_CUENTA_FICHA_PBI = [
    "FOLIO OT", "SUCURSAL", "PATENTE", "FECHA OT", "ESTADO", "ESTADO DETALLADO",
    "NETO", "TIPO VENTA", "MARCA", "MODELO", "ASESOR", "DIAS APERTURA",
    "GLOSA TRABAJO", "TIPO CIERRE", "IMPORTADOR",
] + [c for _n, c, _f in DOCS_CUENTA_FICHA] + [f for _n, _c, f in DOCS_CUENTA_FICHA]

# Patentes de prueba a excluir del backfill de ingresos historicos (prefijo SP + 4 digitos)
RE_PATENTE_EXCLUIDA = re.compile(r"^SP\d{4}$", re.IGNORECASE)

# =============================================================
#   PRODUCCION DE TECNICOS — BDFlexline (SQL Server, red interna)
#   ------------------------------------------------------------
#   Alimenta la pestaña "Producción Técnicos" del Planificador de Taller
#   (horas facturadas por tecnico y mes). Se consulta SOLO desde este
#   script local (nunca desde Streamlit Cloud, que no tiene acceso a la
#   red interna 10.50.15.2). Las credenciales viven en un archivo local
#   "sql_credenciales.txt" (2 lineas: usuario / clave) que NUNCA se sube
#   a GitHub (no esta en ARCHIVOS de Subir_App_GitHub.py). 20/07/2026.
# =============================================================
SQL_SERVER   = "10.50.15.2"
SQL_DATABASE = "BDFlexline"
_sql_cred_file = os.path.join(os.path.dirname(os.path.abspath(__file__)), "sql_credenciales.txt")

# =============================================================
#   AGENDA CURIFOR — Credenciales e IDs de sucursal
#   Las credenciales se guardan aquí porque este script corre
#   localmente (no en el servidor de Streamlit).
# =============================================================
AGENDA_USUARIO = "ADM_CUR"
AGENDA_CLAVE   = "ADM_CUR"

# IDs numéricos de cada sucursal en la Agenda Curifor
AGENDA_SUCURSALES = {
    "LINDEROS":       495,
    "CURICO":         496,
    "PLACILLA":       493,
    "CHILLAN":        497,
    "CHILLAN VIEJO":  524,
    "RANCAGUA":       525,
    "TALCA":          536,
    "LO BLANCO":      578,
    "TALLER MOVIL":   589,
    "TALCA BMW":      588,
    "TALCA CAMIONES": 590,
    "MACUL":          594,
}

# Ruta local del JSON de ranking (se genera junto al datos_dashboard.json)
RUTA_RANKING = r"C:\Users\cjerez\OneDrive - Curifor S.A\Documentos\Claude\Projects\Gestión OTs Pendientes\ranking_cierres.json"

# Listado de repuestos con códigos y descripciones (para enriquecer los vales de consumo)
RUTA_LISTADO_REPUESTOS = r"C:\Users\cjerez\OneDrive - Curifor S.A\Documentos\Claude\Projects\Gestión OTs Pendientes\Listado Repuestos.xlsx"

# Tabla Patente → RUT Cliente (puede haber varios RUTs por patente)
RUTA_PATENTE_CLIENTE = r"C:\Users\cjerez\OneDrive - Curifor S.A\Documentos\Claude\Projects\Gestión OTs Pendientes\PBI\Patente-Cliente.xlsx"

# Resumen de anticipos / Cuenta Ficha del taller
RUTA_ANTICIPO_TALLER = r"C:\Users\cjerez\OneDrive - Curifor S.A\Documentos\Claude\Projects\Gestión OTs Pendientes\PBI\RESUMEN_ANTICIPO_TALLER_2.xlsx"

# Columnas de fecha de los documentos de cierre en el PBI (sin acentos — se buscan con matching flexible)
# El script usará las que efectivamente existan en el PBI, ignorando las que no encuentre.
FECHAS_CIERRE_BUSCAR = [
    "FECHA LIQUIDACION ST",
    "FECHA LIQ ST",
    "FECHA FACTURA CLIENTE",
    "FECHA FACTURA COMPANIA",
    "FECHA CARGO INTERNO",
    "FECHA CARGO GARANTIA",
    "FECHA FACTURA GARANTIA",
    "FECHA CIERRE",
    "FECHA TERMINACION",
    "FECHA TERMINO",
    "FECHA RESOLUCION",
]

# Período desde el que se analiza el ranking (ISO format)
RANKING_DESDE = "2026-01-01"


# =============================================================
#   SUCURSALES
#   ----------
#   Los nombres deben coincidir con el inicio del nombre del archivo.
#   Ejemplo: "Linderos" encuentra "Linderos (11-05).xlsx"
# =============================================================

SUCURSALES = [
    "CHILLAN",
    "Chillán Viejo",
    "CURICO",
    "LO BLANCO",
    "LINDEROS",
    "RANCAGUA",
    "TALCA",
    "PLACILLA",
    "Taller Móvil",
]


# =============================================================
#   COLUMNAS QUE GESTIONAN LAS SUCURSALES
#   --------------------------------------
#   Estas columnas NO vienen del PBI — las llenan las sucursales
#   manualmente y el script las preserva en cada actualización.
# =============================================================

COLUMNAS_GESTION = [
    "CATEGORIA",
    "OBSERVACION OT",
    "NOTAS",
    "AVANCE - GESTIÓN",
    "ULTIMA_EDICION",
]

# Documentos asociados a cada OT en el PBI
# (nombre visible, columna de folio/número en PBI, clave interna)
DOCS_CONFIG = [
    ("Liquidación ST",   "FOLIO LIQUIDACIÓN ST",    "LIQ_ST"),
    ("Factura Cliente",  "FOLIO FACTURA CLIENTE",   "FACT_CLIENTE"),
    ("Factura Compañía", "FOLIO FACTURA COMPAÑÍA",  "FACT_COMPANIA"),
    ("Cargo Interno",    "FOLIO CARGO INTERNO",     "CARGO_INT"),
    ("Cargo Garantía",   "FOLIO CARGO GARANTÍA",    "CARGO_GTIA"),
    ("Factura Garantía", "NÚMERO FACTURA GARANTÍA", "FACT_GTIA"),
    ("Vale de Consumo",  "NRO VALE DE CONSUMO",     "VALE_CONSUMO"),
]

# Columna que identifica de forma única a cada OT
CLAVE = "FOLIO OT"

# Nombre de la hoja de datos en el archivo maestro
HOJA = "BASE"

# Hoja a ignorar en los archivos de sucursal (es un resumen, no los datos)
HOJA_RESUMEN_SUCURSAL = "Datos"

# Nombre de la hoja y fila de encabezado en la sábana del PBI
# (nuevo formato: encabezados en fila 9, índice 0-based = 8)
HOJA_PBI    = "Seguimiento Servicio Técnico"
HEADER_PBI  = 8      # índice 0-based -> fila 9 en Excel

# Estados del PBI que se consideran pendientes de gestión
ESTADOS_PENDIENTES = {"PENDIENTE", "ANULADO"}


# =============================================================
#   FUNCIONES AUXILIARES
# =============================================================

def log(mensaje):
    hora = datetime.now().strftime("%H:%M:%S")
    print(f"  [{hora}] {mensaje}")


def _quitar_acentos(s):
    """
    Elimina tildes/acentos de un string Unicode para comparación flexible de columnas.
    Ej: 'CATEGORÍA' → 'CATEGORIA', 'GESTIÓN' → 'GESTION', 'OBSERVACIÓN' → 'OBSERVACION'
    """
    return "".join(
        c for c in unicodedata.normalize("NFD", str(s))
        if unicodedata.category(c) != "Mn"
    )


# Mapa de nombre-sin-acento → nombre-oficial de cada columna de gestión.
# Se usa en leer_sucursal() para unificar variaciones con/sin tilde.
_GESTION_SIN_ACENTO = {_quitar_acentos(c): c for c in [
    "CATEGORIA", "OBSERVACION OT", "NOTAS", "AVANCE - GESTIÓN", "ULTIMA_EDICION"
]}


def calcular_rango(dias_str):
    """Calcula el rango de días a partir del valor numérico de DIAS APERTURA."""
    try:
        d = int(float(str(dias_str).strip()))
        if d <= 30:
            return "0-30"
        elif d <= 60:
            return "31-60"
        elif d <= 90:
            return "61-90"
        else:
            return "91 o más"
    except Exception:
        return ""


def normalizar_folio(serie):
    """
    Normaliza FOLIO OT: elimina ceros a la izquierda y espacios.
    PBI viene como '0001183961' -> maestro lo tiene como '1183961'.
    """
    return serie.astype(str).str.strip().str.lstrip("0")


def leer_maestro(ruta):
    """Lee la hoja BASE del archivo maestro y normaliza el FOLIO OT."""
    df = pd.read_excel(ruta, sheet_name=HOJA, engine="calamine", dtype=str)
    df[CLAVE] = normalizar_folio(df[CLAVE])
    return df


def _resolver_hoja_pbi(ruta_lectura):
    """
    Devuelve el nombre de hoja a usar para leer el PBI.
    Normalmente es HOJA_PBI tal cual, pero si calamine no la encuentra
    (puede pasar por un archivo que aun no termino de sincronizar desde
    OneDrive, o por una variacion minima del nombre), se busca la hoja
    real comparando sin tildes/mayusculas antes de rendirse.
    """
    try:
        hojas_reales = pd.ExcelFile(ruta_lectura, engine="calamine").sheet_names
    except Exception:
        return HOJA_PBI  # no se pudo listar hojas; se intenta con el nombre esperado igual

    if HOJA_PBI in hojas_reales:
        return HOJA_PBI

    objetivo = _quitar_acentos(HOJA_PBI).strip().lower()
    for h in hojas_reales:
        if _quitar_acentos(h).strip().lower() == objetivo:
            log(f"(!)  Hoja PBI encontrada con nombre ligeramente distinto: '{h}' (se esperaba '{HOJA_PBI}')")
            return h

    # Ultimo respaldo: si solo hay una hoja en el archivo, es casi seguro que es la correcta
    if len(hojas_reales) == 1:
        log(f"(!)  No se encontro la hoja '{HOJA_PBI}' pero el archivo solo tiene una hoja "
            f"('{hojas_reales[0]}') — se usa esa.")
        return hojas_reales[0]

    log(f"(!)  No se encontro la hoja '{HOJA_PBI}'. Hojas disponibles: {hojas_reales}")
    return HOJA_PBI  # se deja que falle mas abajo con un error claro


def encontrar_archivo_pbi():
    """
    Ubica la sábana del PBI ("Seguimiento Servicio Técnico DD-MM.xlsx") dentro
    de CARPETA_PBI, descartando los otros .xlsx que viven ahí mismo
    (Patente-Cliente, informes de anticipos / ficha de cuenta, etc.).

    Preferimos los archivos cuyo nombre contenga "Seguimiento" y, entre esos,
    el modificado más recientemente. Si ninguno calza, se cae al criterio
    antiguo (el .xlsx más reciente que no esté excluido) con un aviso.

    Devuelve la ruta o None si no hay ningún candidato. Se extrajo de main()
    para poder reutilizarla desde actualizar_cuenta_ficha.py — 31/07/2026.
    """
    _excluir_pbi = {"patente-cliente.xlsx"}
    _excluir_pbi_keywords = {
        "anticipo", "resumen_anticipo", "patente",
        "informe ficha cuenta", "ficha cuenta", "informe_ficha",
    }
    archivos_pbi_todos = [
        a for a in glob.glob(os.path.join(CARPETA_PBI, "*.xlsx"))
        if os.path.basename(a).lower() not in _excluir_pbi
        and not any(kw in os.path.basename(a).lower() for kw in _excluir_pbi_keywords)
    ]
    if not archivos_pbi_todos:
        return None

    archivos_pbi = [
        a for a in archivos_pbi_todos
        if "seguimiento" in _quitar_acentos(os.path.basename(a)).lower()
    ]
    if not archivos_pbi:
        log("(!)  Ningún archivo se llama 'Seguimiento...' — se usa el criterio "
            "anterior (más reciente entre todos los .xlsx no excluidos). "
            "Revisa que sea realmente la sábana del PBI.")
        archivos_pbi = archivos_pbi_todos

    return max(archivos_pbi, key=os.path.getmtime)


def leer_pbi(ruta, patentes_cf=None):
    """
    Lee la sábana del PBI (nuevo formato: hoja 'Seguimiento Servicio Técnico', header fila 9).
    - Filtra solo OTs con ESTADO PENDIENTE o ANULADO (case-insensitive)
    - Tiene múltiples filas por FOLIO OT (una por repuesto del Vale de Consumo)
    - Se deduplica a una fila por FOLIO OT para los datos principales
    - Calcula RANGO desde DIAS APERTURA (el nuevo PBI no trae columna RANGO)
    - Columnas normalizadas a mayúsculas
    """
    # Copiar a temporal primero: evita fallos si OneDrive todavia esta
    # sincronizando el archivo o si quedo parcialmente bloqueado.
    _tmp_pbi = None
    ruta_lectura = ruta
    try:
        import tempfile as _tf
        _fd, _tmp_pbi = _tf.mkstemp(suffix=os.path.splitext(ruta)[1] or ".xlsx")
        os.close(_fd)
        shutil.copy2(ruta, _tmp_pbi)
        ruta_lectura = _tmp_pbi
    except Exception:
        ruta_lectura = ruta
        if _tmp_pbi and os.path.exists(_tmp_pbi):
            try:
                os.remove(_tmp_pbi)
            except Exception:
                pass
            _tmp_pbi = None

    try:
        hoja_real = _resolver_hoja_pbi(ruta_lectura)
        df = pd.read_excel(ruta_lectura, sheet_name=hoja_real, header=HEADER_PBI,
                           engine="calamine", dtype=str)
    finally:
        if _tmp_pbi and os.path.exists(_tmp_pbi):
            try:
                os.remove(_tmp_pbi)
            except Exception:
                pass

    # Normalizar nombres de columnas a mayúsculas y eliminar columnas sin nombre
    df.columns = [c.upper().strip() if isinstance(c, str) else str(c) for c in df.columns]
    df = df.loc[:, ~df.columns.str.startswith("UNNAMED")]

    if CLAVE not in df.columns:
        raise ValueError(f"No se encontró columna '{CLAVE}' en el PBI. "
                         f"Columnas disponibles: {list(df.columns[:10])}")

    df[CLAVE] = normalizar_folio(df[CLAVE])

    # Eliminar filas con FOLIO OT vacío o inválido
    df = df[df[CLAVE].notna() & (df[CLAVE] != "") & (df[CLAVE] != "NAN")]

    # Mapa FOLIO OT -> PATENTE de TODAS las OTs (todos los estados, ANTES del filtro).
    # Sirve de "puente": una solicitud de repuesto puede referir una OT cerrada;
    # con su patente cruzamos al vehículo que SÍ tiene una OT pendiente en la app.
    mapa_ot_patente = {}
    if "PATENTE" in df.columns:
        for _folio, _pat in zip(df[CLAVE].astype(str), df["PATENTE"]):
            _f = str(_folio).strip()
            if _f and _f not in mapa_ot_patente:
                _p = str(_pat).strip().upper() if _pat is not None else ""
                mapa_ot_patente[_f] = "" if _p in ("NAN", "NONE") else _p
        log(f"Mapa OT->patente (todos los estados): {len(mapa_ot_patente)} OTs")

    # Indice de OT para el modulo Cuenta Ficha: se arma AQUI, antes del filtro de
    # estado de abajo, porque este es el unico punto donde estan en memoria TODAS
    # las OT (Finalizado/Cerrado/Pendiente/Anulado). Se recorre en streaming y
    # solo se guardan las OT de las patentes que interesan (las de los clientes
    # con ficha de cuenta) — NO se copia el DataFrame: son 260.000 filas x 70
    # columnas y duplicarlo dispara la memoria del PC.
    ots_cuenta_ficha = _indexar_ots_cuenta_ficha(df, patentes_cf)

    # Filtrar solo estados PENDIENTE y ANULADO (la nueva sábana incluye OTs cerradas)
    if "ESTADO" in df.columns:
        mask_estado = df["ESTADO"].astype(str).str.strip().str.upper().isin(ESTADOS_PENDIENTES)
        df = df[mask_estado].copy()
        log(f"Filtro ESTADO: {len(df)} filas corresponden a {ESTADOS_PENDIENTES}")
    else:
        log("(!)  Columna ESTADO no encontrada — se usan todas las filas")

    # Guardar copia completa antes de deduplicar (para extraer documentos y repuestos)
    df_completo = df.copy()

    # Calcular RANGO desde DIAS APERTURA (el nuevo PBI no trae columna RANGO)
    if "DIAS APERTURA" in df.columns and "RANGO" not in df.columns:
        df["RANGO"] = df["DIAS APERTURA"].apply(calcular_rango)
        df_completo["RANGO"] = df_completo["DIAS APERTURA"].apply(calcular_rango)
        log("Columna RANGO calculada desde DIAS APERTURA")

    # Deduplicar: el PBI trae una fila por repuesto dentro de cada OT.
    # Nos quedamos con la primera fila de cada FOLIO OT, que tiene los
    # datos principales de la OT (sucursal, asesor, estado, fechas, etc.)
    df = df.drop_duplicates(subset=[CLAVE], keep="first").reset_index(drop=True)

    return df, df_completo, mapa_ot_patente, ots_cuenta_ficha


def leer_sucursal(ruta):
    """
    Lee el archivo de una sucursal.
    - Tiene dos hojas: 'Datos' (resumen, se ignora) y 'DD-MM' (datos reales)
    - Columnas normalizadas a mayúsculas Y con acentos unificados.
      Ej: 'CATEGORÍA' y 'CATEGORIA' se tratan como la misma columna.
    - FOLIO OT normalizado (sin ceros adelante)
    """
    xf = pd.ExcelFile(ruta, engine="calamine")
    hojas_datos = [h for h in xf.sheet_names if h != HOJA_RESUMEN_SUCURSAL]
    if not hojas_datos:
        raise ValueError(f"No se encontró hoja de datos en {os.path.basename(ruta)}")

    df = xf.parse(hojas_datos[0], dtype=str)
    df.columns = [c.upper().strip() if isinstance(c, str) else c for c in df.columns]

    # ── CORRECCIÓN BUG ACENTOS ──────────────────────────────────────────────────
    # Los archivos de algunas sucursales usan "CATEGORÍA", "OBSERVACIÓN OT", etc.
    # con tilde, mientras que COLUMNAS_GESTION usa versiones sin tilde (o viceversa).
    # Mapeamos las columnas del archivo sucursal al nombre oficial de COLUMNAS_GESTION
    # comparando sus versiones sin acento.
    rename_acentos = {}
    for col_suc in df.columns:
        col_norm = _quitar_acentos(col_suc)
        if col_norm in _GESTION_SIN_ACENTO:
            nombre_oficial = _GESTION_SIN_ACENTO[col_norm]
            if col_suc != nombre_oficial:
                rename_acentos[col_suc] = nombre_oficial
    if rename_acentos:
        df = df.rename(columns=rename_acentos)
    # ────────────────────────────────────────────────────────────────────────────

    if CLAVE not in df.columns:
        # Intentar también con normalización de acentos en CLAVE
        clave_norm = _quitar_acentos(CLAVE)
        clave_alt  = next((c for c in df.columns if _quitar_acentos(c) == clave_norm), None)
        if clave_alt:
            df = df.rename(columns={clave_alt: CLAVE})
        else:
            raise ValueError(f"Columna '{CLAVE}' no encontrada en {os.path.basename(ruta)}")

    df[CLAVE] = normalizar_folio(df[CLAVE])
    return df


def encontrar_archivo_sucursal(nombre_sucursal):
    """
    Busca el archivo de una sucursal en CARPETA_SUCURSALES.
    El nombre del archivo puede tener fecha entre paréntesis, ej: 'Linderos (11-05).xlsx'
    Si hay más de un archivo para la misma sucursal, toma el más reciente.

    Nota: busca con espacio después del nombre (ej: 'Chillan (')
    para evitar que 'Chillan' encuentre 'Chillan Viejo'.
    """
    # El patrón incluye el paréntesis: "Chillan (*.xlsx"
    # Así "Chillan" NO coincide con "Chillan Viejo (11-05).xlsx"
    # porque ese archivo empieza con "Chillan V", no "Chillan ("
    patron_xlsx = os.path.join(CARPETA_SUCURSALES, f"{nombre_sucursal} (*.xlsx")
    patron_xlsm = os.path.join(CARPETA_SUCURSALES, f"{nombre_sucursal} (*.xlsm")
    # También buscar sin fecha por si alguien guarda el archivo sin ella
    patron_xlsx_exacto = os.path.join(CARPETA_SUCURSALES, f"{nombre_sucursal}.xlsx")
    patron_xlsm_exacto = os.path.join(CARPETA_SUCURSALES, f"{nombre_sucursal}.xlsm")
    archivos = (glob.glob(patron_xlsx) + glob.glob(patron_xlsm) +
                glob.glob(patron_xlsx_exacto) + glob.glob(patron_xlsm_exacto))

    if not archivos:
        return None  # No se encontró archivo para esta sucursal

    # Si hay varios archivos (de días distintos), toma el más reciente
    return max(archivos, key=os.path.getmtime)


def respaldar_maestro():
    """Guarda una copia del archivo maestro antes de modificarlo."""
    if not os.path.exists(ARCHIVO_MAESTRO):
        return
    os.makedirs(CARPETA_RESPALDOS, exist_ok=True)
    timestamp = datetime.now().strftime("%Y%m%d_%H%M")
    nombre = os.path.splitext(os.path.basename(ARCHIVO_MAESTRO))[0]
    destino = os.path.join(CARPETA_RESPALDOS, f"{nombre}_{timestamp}.xlsm")
    shutil.copy2(ARCHIVO_MAESTRO, destino)
    log(f"Respaldo guardado -> {os.path.basename(destino)}")


def aplicar_columnas_gestion(df_destino, df_fuente, sufijo, solo_rellenar_vacios=False):
    """
    Copia las COLUMNAS_GESTION desde df_fuente hacia df_destino
    usando CLAVE (FOLIO OT) como punto de cruce.

    Por defecto (solo_rellenar_vacios=False): sobreescribe el destino cada vez
    que la fuente tenga un valor no vacío — usar solo para fuentes que
    representan la edición más reciente y confiable (maestro Excel como línea
    base, GitHub/app web como la fuente de verdad de las ediciones).

    Con solo_rellenar_vacios=True: SOLO aplica el valor de la fuente si el
    destino está actualmente vacío — usar para fuentes que pueden estar
    desactualizadas (Excel de sucursal) y que por lo tanto NUNCA deben pisar
    un valor que ya vino de una fuente más confiable evaluada antes.

    Bug real corregido 10/07/2026: antes esta función siempre sobreescribía
    sin mirar el destino, sin importar el parámetro — como los Excel de
    sucursal se aplican al final "con prioridad máxima", cualquier valor
    viejo que hubiera quedado ahí (aunque llevara semanas sin tocarse) pisaba
    para siempre las ediciones hechas desde la app web en cada corrida del
    consolidador. Ver CLAUDE.md sesión 10/07/2026 para el diagnóstico completo
    (caso real: LINDEROS (11-05).xlsx, sin tocar desde el 25/05, seguía
    revirtiendo Categoría/Notas de OTs editadas después en la app).
    """
    df_fuente_cols = [CLAVE] + [c for c in COLUMNAS_GESTION if c in df_fuente.columns]
    df_fuente_slim = df_fuente[df_fuente_cols].rename(
        columns={c: f"{sufijo}_{c}" for c in COLUMNAS_GESTION}
    )

    df_resultado = df_destino.merge(df_fuente_slim, on=CLAVE, how="left")

    for col in COLUMNAS_GESTION:
        col_nueva = f"{sufijo}_{col}"
        if col_nueva not in df_resultado.columns:
            continue
        tiene_valor = df_resultado[col_nueva].notna() & (df_resultado[col_nueva].astype(str).str.strip() != "")
        if solo_rellenar_vacios:
            destino_vacio = df_resultado[col].isna() | (df_resultado[col].astype(str).str.strip() == "")
            aplica = tiene_valor & destino_vacio
        else:
            aplica = tiene_valor
        df_resultado.loc[aplica, col] = df_resultado.loc[aplica, col_nueva]
        df_resultado.drop(columns=[col_nueva], inplace=True)

    return df_resultado


def actualizar_dashboard(wb, df, df_anterior=None):
    """
    Recalcula y escribe todos los valores del DASHBOARD a partir del DataFrame.
    El Dashboard tiene valores hardcodeados (no formulas), por eso hay que
    recalcularlos cada vez que se actualiza la BASE.
    Maneja correctamente las celdas combinadas (merged cells).
    """
    ws = wb["DASHBOARD"]
    hoy = datetime.now().strftime("%d/%m/%Y")

    # Construir set de celdas "esclavas" de combinaciones (no se puede escribir en ellas)
    celdas_combinadas = set()
    for rango_merge in ws.merged_cells.ranges:
        celdas = list(rango_merge.cells)
        for row_c, col_c in celdas[1:]:   # La primera es la celda maestra, el resto son esclavas
            celdas_combinadas.add((row_c, col_c))

    def set_cell(row, col, value):
        """Escribe en una celda solo si no es esclava de una combinación."""
        if (row, col) not in celdas_combinadas:
            ws.cell(row=row, column=col).value = value

    def clear_range(min_row, max_row, cols):
        """Limpia un rango de celdas respetando las celdas combinadas."""
        for r in range(min_row, max_row + 1):
            for c in cols:
                set_cell(r, c, None)

    # Normalizar columnas clave
    df = df.copy()
    for col in ["RANGO", "SUCURSAL", "TIPO VENTA", "MARCA", "ASESOR"]:
        if col in df.columns:
            df[col] = df[col].fillna("").astype(str).str.strip()

    # --- Titulo con fecha actualizada ---
    set_cell(1, 3, f"\U0001f4ca  DASHBOARD — SEGUIMIENTO ÓRDENES DE TRABAJO  |  {hoy}")

    # Valores exactos del campo RANGO en BASE
    RANGO_0_30   = "0-30"
    RANGO_31_60  = "31-60"
    RANGO_61_90  = "61-90"
    RANGO_91_MAS = "91 o más"

    # --- KPIs principales ---
    total     = len(df)
    criticas  = int((df["RANGO"] == RANGO_91_MAS).sum()) if "RANGO" in df.columns else 0
    urgentes  = int((df["RANGO"] == RANGO_61_90).sum())  if "RANGO" in df.columns else 0
    atencion  = int((df["RANGO"] == RANGO_31_60).sum())  if "RANGO" in df.columns else 0
    recientes = int((df["RANGO"] == RANGO_0_30).sum())   if "RANGO" in df.columns else 0

    set_cell(4, 3,  total)
    set_cell(4, 6,  criticas)
    set_cell(4, 8,  urgentes)
    set_cell(4, 10, atencion)
    set_cell(4, 12, recientes)

    # --- OT POR SUCURSAL Y RANGO (cols C-H, desde fila 10) ---
    if "SUCURSAL" in df.columns and "RANGO" in df.columns:
        rangos = [RANGO_0_30, RANGO_31_60, RANGO_61_90, RANGO_91_MAS]
        pivot = df.groupby(["SUCURSAL", "RANGO"]).size().unstack(fill_value=0)
        for r in rangos:
            if r not in pivot.columns:
                pivot[r] = 0
        pivot["Total"] = pivot[rangos].sum(axis=1)
        pivot = pivot.sort_values("Total", ascending=False).reset_index()

        clear_range(10, 44, [3, 4, 5, 6, 7, 8])

        col_map = {RANGO_0_30: 4, RANGO_31_60: 5, RANGO_61_90: 6, RANGO_91_MAS: 7, "Total": 8}
        for idx, row in pivot.iterrows():
            fila = 10 + idx
            if fila > 44:
                break
            set_cell(fila, 3, row["SUCURSAL"])
            for rango, col in col_map.items():
                set_cell(fila, col, int(row.get(rango, 0)))

    # --- OT POR TIPO DE VENTA (cols J-L, desde fila 10) ---
    if "TIPO VENTA" in df.columns:
        pivot_tv = df[df["TIPO VENTA"] != ""].groupby("TIPO VENTA").size().reset_index(name="Cantidad")
        pivot_tv = pivot_tv.sort_values("Cantidad", ascending=False).reset_index(drop=True)

        clear_range(10, 44, [10, 11, 12])

        for idx, row in pivot_tv.iterrows():
            fila = 10 + idx
            if fila > 44:
                break
            pct = f"{row['Cantidad'] / total * 100:.1f}%" if total > 0 else "0%"
            set_cell(fila, 10, row["TIPO VENTA"])
            set_cell(fila, 11, int(row["Cantidad"]))
            set_cell(fila, 12, pct)

    # --- OT POR MARCA (cols N-O, desde fila 10) ---
    if "MARCA" in df.columns:
        pivot_marca = df[df["MARCA"] != ""].groupby("MARCA").size().reset_index(name="Cantidad")
        pivot_marca = pivot_marca.sort_values("Cantidad", ascending=False).reset_index(drop=True)

        clear_range(10, 44, [14, 15])

        for idx, row in pivot_marca.iterrows():
            fila = 10 + idx
            if fila > 44:
                break
            set_cell(fila, 14, row["MARCA"])
            set_cell(fila, 15, int(row["Cantidad"]))

    # --- OT POR ASESOR TOP 15 (cols C-D, desde fila 28) ---
    if "ASESOR" in df.columns:
        pivot_asesor = df[df["ASESOR"] != ""].groupby("ASESOR").size().reset_index(name="Cantidad")
        pivot_asesor = pivot_asesor.sort_values("Cantidad", ascending=False).head(15).reset_index(drop=True)

        clear_range(28, 49, [3, 4])

        for idx, row in pivot_asesor.iterrows():
            fila = 28 + idx
            set_cell(fila, 3, row["ASESOR"])
            set_cell(fila, 4, int(row["Cantidad"]))


    log(f"Dashboard actualizado: {total} OTs | {hoy}")


def escribir_en_maestro(df_datos, df_anterior=None):
    """
    Actualiza la hoja BASE del archivo maestro con los datos nuevos
    y recalcula el DASHBOARD. Las macros y demás hojas quedan intactos.
    """
    wb = load_workbook(ARCHIVO_MAESTRO, keep_vba=True)
    ws = wb[HOJA]

    # Leer los encabezados del Excel para respetar el orden original de columnas
    encabezados = [ws.cell(1, col).value for col in range(1, ws.max_column + 1)]
    mapa_columnas = {nombre: i + 1 for i, nombre in enumerate(encabezados) if nombre}

    # Borrar datos anteriores (la fila 1 con encabezados se conserva)
    for fila in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for celda in fila:
            celda.value = None

    # Escribir los datos nuevos fila por fila
    for i, (_, fila) in enumerate(df_datos.iterrows(), start=2):
        for nombre_col, num_col in mapa_columnas.items():
            if nombre_col not in df_datos.columns:
                continue
            valor = fila[nombre_col]
            # Convertir valores nulos y fechas al formato que acepta Excel
            if not isinstance(valor, str) and pd.isna(valor):
                valor = None
            elif isinstance(valor, pd.Timestamp):
                valor = valor.to_pydatetime()
            ws.cell(row=i, column=num_col, value=valor)

    # Recalcular el Dashboard con los datos nuevos
    actualizar_dashboard(wb, df_datos, df_anterior)

    wb.save(ARCHIVO_MAESTRO)




# =============================================================
#   DOCUMENTOS ASOCIADOS A CADA OT
# =============================================================

def agregar_info_documentos(df_ots, df_pbi_completo):
    """
    Para cada OT agrega conteos y números de cada tipo de documento
    (liquidaciones, facturas, vales de consumo, etc.) a partir de
    TODAS las filas que ese FOLIO OT tiene en el PBI.
    """
    resultado = df_ots.copy()

    for _, folio_col, key in DOCS_CONFIG:
        folio_col_norm = folio_col.upper().strip()
        if folio_col_norm not in df_pbi_completo.columns:
            resultado[f"N_{key}"]      = 0
            resultado[f"FOLIOS_{key}"] = ""
            continue

        docs_por_ot = {}
        for folio_ot, grupo in df_pbi_completo.groupby(CLAVE):
            valores = (
                grupo[folio_col_norm]
                .dropna()
                .astype(str)
                .str.strip()
            )
            validos = [v for v in valores if v and v.upper() not in ("", "NAN", "NONE", "NAT")]
            # Deduplicar manteniendo orden
            vistos = []
            for v in validos:
                if v not in vistos:
                    vistos.append(v)
            if vistos:
                docs_por_ot[folio_ot] = vistos

        resultado[f"N_{key}"]      = resultado[CLAVE].map(lambda x: len(docs_por_ot.get(x, [])))
        resultado[f"FOLIOS_{key}"] = resultado[CLAVE].map(lambda x: ", ".join(docs_por_ot.get(x, [])))

    # ── Fechas de Factura Cliente y Factura Compañía (para Facturas X) ──────
    COLS_FECHAS_DOC = [
        ("FECHA FACTURA CLIENTE",  "FECHA_FACT_CLIENTE"),
        ("FECHA FACTURA COMPA",    "FECHA_FACT_COMPANIA"),   # matching parcial
    ]
    for fecha_col_patron, clave_salida in COLS_FECHAS_DOC:
        # Buscar columna con matching flexible (sin tilde, parcial)
        col_real = next(
            (c for c in df_pbi_completo.columns
             if fecha_col_patron.upper() in c.upper().replace("Ñ","N").replace("Á","A")
                                                      .replace("É","E").replace("Ó","O")),
            None
        )
        if col_real is None:
            resultado[clave_salida] = ""
            continue

        fechas_por_ot = {}
        for folio_ot, grupo in df_pbi_completo.groupby(CLAVE):
            fechas = []
            for v in grupo[col_real].dropna():
                try:
                    if hasattr(v, "strftime"):
                        fechas.append(v.strftime("%d/%m/%Y"))
                    else:
                        s = str(v).strip()
                        if s and s.lower() not in ("nan","none","nat",""):
                            fechas.append(s)
                except Exception:
                    pass
            vistos = list(dict.fromkeys(fechas))   # deduplicar manteniendo orden
            if vistos:
                fechas_por_ot[folio_ot] = vistos

        resultado[clave_salida] = resultado[CLAVE].map(
            lambda x: ", ".join(fechas_por_ot.get(x, []))
        )

    log(f"Info de documentos agregada ({len(DOCS_CONFIG)} tipos)")
    return resultado


# =============================================================
#   REPUESTOS DEL VALE DE CONSUMO
# =============================================================

def cargar_listado_repuestos():
    """
    Carga el 'Listado Repuestos.xlsx' y devuelve un dict {codigo: descripcion}.
    El código se normaliza con strip() para evitar problemas de espacios.
    Si el archivo no existe o falla, devuelve un dict vacío (sin interrumpir el proceso).
    """
    if not os.path.exists(RUTA_LISTADO_REPUESTOS):
        log("(i)  Listado de repuestos no encontrado — los códigos se mostrarán sin descripción")
        return {}
    try:
        df_rep = pd.read_excel(RUTA_LISTADO_REPUESTOS, usecols=["Producto", "Descripción"],
                               dtype=str)
        lookup = {}
        for _, row in df_rep.iterrows():
            cod  = str(row.get("Producto",    "")).strip()
            desc = str(row.get("Descripción", "")).strip()
            if cod and cod.lower() not in ("nan", "none") and desc and desc.lower() not in ("nan", "none"):
                lookup[cod] = desc
        log(f"Listado de repuestos cargado: {len(lookup):,} códigos")
        return lookup
    except Exception as e:
        log(f"(!)  Error al cargar listado de repuestos: {e}")
        return {}


def inferir_trabajo(repuestos):
    """
    A partir de la lista de repuestos de una OT (con campo 'descripcion'),
    infiere el tipo de trabajo mecánico que se está realizando.

    Estrategia: cada descripción se normaliza (sin acentos, mayúsculas) y se
    puntúa contra grupos de palabras clave. El grupo con más hits gana.
    Si hay dos grupos con puntaje igual o muy cercano, se combinan.
    Retorna un string listo para poner en NOTAS, o "" si no se puede inferir.
    """
    if not repuestos:
        return ""

    # ── Reglas: (etiqueta, [palabras_clave_normalizadas]) ───────────────────
    # Orden importa para desempate: los más específicos van primero.
    REGLAS = [
        ("Reparación de motor", [
            "PISTON", "BIELA", "METAL BIELA", "METAL BANCADA",
            "CIGUENAL", "CULATA", "JUNTA CULATA", "JUNTA MOTOR",
            "KIT MOTOR", "ARBOL DE LEVAS", "LEVADOR", "ANILLO PISTON",
            "VALVULA ADMISION", "VALVULA ESCAPE", "GUIA VALVULA",
            "SELLO VALVULA", "RETEN CIGUENAL", "TAPON CARTER",
            "EMPAQUETADURA CULATA", "BUJE BIELA",
        ]),
        ("Reparación de embrague", [
            "EMBRAGUE", "DISCO EMBRAGUE", "KIT EMBRAG", "KIT EMBRAGUE",
            "PLATO PRESION", "CILINDRO EMBRAGUE", "HORQUILLA EMBRAGUE",
            "RODAMIENTO EMBRAGUE", "SERVO EMBRAGUE", "MAZA EMBRAGUE",
        ]),
        ("Reparación de caja de cambios", [
            "CAJA CAMBIO", "SINCRONIZADOR", "TELECOMANDO",
            "PIÑON RIEL", "PIÑON CAJA", "EJE CAJA", "HORQUILLA CAMBIO",
            "BUJE CAJA", "PALANCA CAMBIO", "KIT SINCRONIZADOR",
        ]),
        ("Reparación de diferencial / cardan", [
            "DIFERENCIAL", "CORONA DIFERENCIAL", "PLANETARIO",
            "SEMIEJES", "CRUCETA", "CARDAN", "RODTO CARDAN",
            "JUNTA HOMOCINETICA", "BUJE CARDAN",
        ]),
        ("Reparación de frenos", [
            "PASTILLA", "DISCO FRENO", "ZAPATA", "CALIPER",
            "CILINDRO FRENO", "BOMBA FRENO", "SERVO FRENO",
            "LIQUIDO FRENO", "MORDAZA FRENO", "KIT FRENO",
        ]),
        ("Reparación de suspensión y dirección", [
            "AMORTIGUADOR", "AMORTIG TRAS", "AMORTIG DEL",
            "RESORTE SUSPENSION", "BARRA ESTABILIZ", "ROTULA",
            "BUJE SUSPENSION", "BRAZO SUSPENSION", "TENSOR B/ESTAB",
            "BOMBA DIRECCION", "CREMALLERA", "CAJA DIRECCION",
            "BARRA DIRECCION", "ROTULA DIRECCION",
        ]),
        ("Reparación de sistema de refrigeración", [
            "RADIADOR", "TERMOSTATO", "BOMBA AGUA",
            "MANGUERA RADIADOR", "TAPA RADIADOR", "DEPOSITO AGUA",
            "DEPOSITO EXPANSION", "VENTILADOR MOTOR", "INTERCOOLER",
            "CORREA DISTRIBUCION",
        ]),
        ("Reparación de sistema eléctrico", [
            "ALTERNADOR", "BATERIA", "MOTOR ARRANQUE", "ARRANQUE",
            "RELAY", "RELE", "FUSIBLE", "BOBINA ENCENDIDO",
            "SENSOR CMP", "SENSOR CKP", "SENSOR MAF", "SENSOR MAP",
            "SENSOR TEMPERATURA", "ECU", "MODULO", "FAROL", "FARO",
        ]),
        ("Reparación de sistema de combustible / turbo", [
            "INYECTOR", "BOMBA INYECCION", "BOMBA COMBUSTIBLE",
            "TURBO", "TURBOCOMPRESOR", "FILTRO COMBUSTIBLE",
            "FILTRO PETROLEO", "RAIL COMBUSTIBLE", "TOBERA", "REGULADOR PRESION",
        ]),
        ("Reparación de frenos de aire / neumáticos", [
            "FILTRO SECADOR", "VALVULA FRENO", "VALVULA RELAY",
            "VALVULA HAND", "COMPRESOR AIRE", "CAMARA DE FRENO",
            "REGULADOR FRENO", "PURGA AIRE", "DEPOSITO AIRE",
        ]),
        ("Mantención preventiva", [
            "FILTRO ACEITE", "FILTRO AIRE MOTOR", "FILTRO AIRE",
            "ACEITE MOTOR", "LUBRICANTE", "GRASA", "KIT MANTENCION",
            "FILTRO HIDRAULICO", "FILTRO TRANSMISION",
        ]),
        ("Reparación de carrocería / accesorios", [
            "PARACHOQUE", "PLUMILLA", "LIMPIA PARABRISAS",
            "EMBLEMA", "MANILLA", "VIDRIO", "ESPEJO",
            "TAPA DEPOSITO", "SOPORTE",
        ]),
    ]

    def _norm(s):
        """Elimina acentos y pasa a mayúsculas para comparación."""
        sin_tilde = "".join(
            c for c in unicodedata.normalize("NFD", str(s))
            if unicodedata.category(c) != "Mn"
        )
        return sin_tilde.upper()

    # Texto completo de todas las descripciones de la OT (normalizado)
    texto_total = " ".join(
        _norm(r.get("descripcion", "")) for r in repuestos if r.get("descripcion")
    )
    if not texto_total.strip():
        return ""

    # Puntuar cada categoría
    puntajes = {}
    for etiqueta, keywords in REGLAS:
        pts = sum(1 for kw in keywords if _norm(kw) in texto_total)
        if pts > 0:
            puntajes[etiqueta] = pts

    if not puntajes:
        return ""

    # Ordenar de mayor a menor puntaje
    ranking = sorted(puntajes.items(), key=lambda x: x[1], reverse=True)
    top_pts = ranking[0][1]

    # Incluir todas las categorías que igualen el puntaje máximo (máx. 2)
    ganadoras = [et for et, pts in ranking if pts == top_pts][:2]

    return " + ".join(ganadoras)


def agregar_repuestos_detalle(df_pbi_completo, lookup_repuestos=None):
    """
    Extrae los repuestos del Vale de Consumo por OT a partir del PBI completo
    (todas las filas, antes de deduplicar). Devuelve un dict:
        { folio_ot: [ {"vale":..., "producto":..., "descripcion":...,
                        "cantidad":..., "costo_unitario":..., "costo_total":...}, ... ] }
    Deduplica por (VALE, PRODUCTO) para evitar filas repetidas que trae el PBI.
    Solo incluye filas que tengan PRODUCTO con valor real.
    Si se pasa lookup_repuestos ({codigo: descripcion}), agrega el campo "descripcion".
    """
    if lookup_repuestos is None:
        lookup_repuestos = {}

    resultado = {}
    cols_rep = ["NRO VALE DE CONSUMO", "PRODUCTO", "CANTIDAD", "COSTO UNITARIO", "COSTO TOTAL"]
    cols_disp = [c for c in cols_rep if c in df_pbi_completo.columns]

    for folio, grupo in df_pbi_completo.groupby(CLAVE):
        # Solo filas con PRODUCTO real (no vacío / nan)
        mask_prod = (
            grupo["PRODUCTO"].notna() &
            (grupo["PRODUCTO"].astype(str).str.strip() != "") &
            (~grupo["PRODUCTO"].astype(str).str.strip().str.upper().isin(["NAN", "NONE", "NAT", "0"]))
        ) if "PRODUCTO" in grupo.columns else pd.Series(False, index=grupo.index)

        grupo_prod = grupo[mask_prod][cols_disp].drop_duplicates()

        if grupo_prod.empty:
            resultado[folio] = []
            continue

        repuestos = []
        for _, row in grupo_prod.iterrows():
            cod = _limpiar_str(row.get("PRODUCTO", ""))
            r = {
                "vale":           _limpiar_str(row.get("NRO VALE DE CONSUMO", "")),
                "producto":       cod,
                "descripcion":    lookup_repuestos.get(cod, ""),   # nombre del repuesto
                "cantidad":       _limpiar_str(row.get("CANTIDAD", "")),
                "costo_unitario": _limpiar_str(row.get("COSTO UNITARIO", "")),
                "costo_total":    _limpiar_str(row.get("COSTO TOTAL", "")),
            }
            repuestos.append(r)

        resultado[folio] = repuestos

    total_con_reps = sum(1 for v in resultado.values() if v)
    log(f"Repuestos extraídos: {total_con_reps} OT(s) con repuestos en Vale de Consumo")
    return resultado


def _limpiar_str(v):
    """Convierte un valor a string limpio; retorna '' si es nan/None/NAT."""
    s = str(v).strip() if v is not None else ""
    return "" if s.lower() in ("nan", "none", "nat", "<na>") else s


def _clave_repuesto(r):
    """Clave de identidad de un repuesto: (vale, producto)."""
    return (r.get("vale", ""), r.get("producto", ""))


def _merge_historico_repuestos(hist_prev, actuales, hoy):
    """
    Fusiona el historial anterior de repuestos con la lista actual del PBI.

    Lógica:
    - Repuesto en hist_prev Y en actuales  → se mantiene, ultima_vista = hoy, eliminado=False
    - Repuesto en hist_prev PERO NO en actuales → eliminado=True, fecha_eliminacion=hoy (si no lo era ya)
    - Repuesto NUEVO (en actuales pero no en hist_prev) → se agrega con primera_vista=hoy
    """
    actuales_keys = {_clave_repuesto(r) for r in actuales}
    nuevos_historico = []
    vistos_en_hist = set()

    for h in hist_prev:
        k = _clave_repuesto(h)
        h_nuevo = dict(h)
        if k in actuales_keys:
            h_nuevo["ultima_vista"] = hoy
            h_nuevo["eliminado"] = False
            h_nuevo.pop("fecha_eliminacion", None)
        else:
            if not h_nuevo.get("eliminado"):
                h_nuevo["eliminado"] = True
                h_nuevo["fecha_eliminacion"] = hoy
        nuevos_historico.append(h_nuevo)
        vistos_en_hist.add(k)

    # Agregar repuestos nuevos que no estaban en el historial anterior
    for r in actuales:
        k = _clave_repuesto(r)
        if k not in vistos_en_hist:
            nuevos_historico.append({
                **r,
                "primera_vista":  hoy,
                "ultima_vista":   hoy,
                "eliminado":      False,
            })

    return nuevos_historico


# =============================================================
#   EXPORTACIÓN PARA DASHBOARD WEB
# =============================================================

# =============================================================
#   SEGUIMIENTO DE COMPRAS  (repuestos: en espera / en bodega)
#   ------------------------------------------------------------
#   Archivo con 2 hojas (Importación + Nacional). El N° de OT
#   viene "sucio" dentro de la columna 'Comentario Lineal', así
#   que lo detectamos por regex (números de 6 o 7 dígitos) y lo
#   cruzamos contra las OTs pendientes. Solo se aceptan números
#   que coincidan con una OT real -> se filtran falsos positivos.
# =============================================================

# Candidatas de columna "fecha en bodega" (Nacional: P/E; Importación: Recepción).
# Se prueba en orden hasta encontrar la que exista en la hoja.
FECHAS_BODEGA_COMPRAS = [
    "Fecha Documento P/E",
    "Fecha Documento Recepción",
    "Fecha Documento Recepcion",
]
ESTADO_EN_BODEGA = "CERRADO POR SISTEMA"   # repuesto llegó a bodega
# Estados que significan "aún no llega": "Pendiente" y "Abierto con referencia"
RE_OT_COMPRAS = re.compile(r"(?<!\d)\d{6,7}(?!\d)")


def _detectar_hojas_compras(sheetnames):
    """
    Reconoce las hojas del archivo de compras por palabra clave en su nombre,
    sin importar cómo el usuario las llame. Devuelve [(nombre_real, origen)]
    con origen ∈ {"Importación", "Nacional"}.

    Nota: 'internacional' contiene 'nacional', por eso se evalúa import/internac
    PRIMERO. Así el usuario puede pegar la sábana en hojas llamadas, por ejemplo,
    "Seguimiento Nacional" / "Nacional" e "Importacion" / "Internacional".
    """
    pares = []
    for s in sheetnames:
        sl = str(s).strip().lower()
        if "import" in sl or "internac" in sl:
            pares.append((s, "Importación"))
        elif "nacional" in sl:
            pares.append((s, "Nacional"))
    return pares


def _fmt_fecha_compras(v):
    """Normaliza la fecha de bodega a texto DD/MM/AAAA."""
    if v is None:
        return ""
    if isinstance(v, datetime):
        return v.strftime("%d/%m/%Y")
    s = str(v).strip()
    return "" if s.lower() in ("nan", "none", "nat", "") else s


def _col_idx_compras(header, nombre):
    """Índice de la primera columna cuyo encabezado empieza con `nombre`."""
    objetivo = nombre.strip().lower()
    for i, h in enumerate(header):
        if str(h).strip().lower().startswith(objetivo[:18]):
            return i
    return None


def _encontrar_header_compras(ws, max_scan=30):
    """Localiza la fila de encabezado (la que contiene 'Comentario Lineal')."""
    for i, row in enumerate(ws.iter_rows(min_row=1, max_row=max_scan, values_only=True)):
        vals = [str(c).strip() if c is not None else "" for c in row]
        if any(v.startswith("Comentario Lineal") for v in vals):
            return i, [str(c).strip() if c is not None else "" for c in row]
    return None, None


def leer_seguimiento_compras(folios_pendientes, mapa_ot_patente=None, mapa_patente_pendiente=None, mapa_stock=None):
    """
    Lee el archivo de Seguimiento de Compras (ambas hojas), detecta el N° de OT
    de 6-7 dígitos dentro de 'Comentario Lineal' y lo enlaza con una OT PENDIENTE
    de la app.

    El N° de OT referido en la solicitud de repuesto suele estar CERRADO. Por eso:
      1. Validamos el N° contra TODAS las OTs del PBI (mapa_ot_patente, todos los
         estados) -> filtra falsos positivos sin gastar tokens.
      2. Si el N° es una OT pendiente -> se asocia directo.
      3. Si el N° es una OT cerrada/otra -> usamos su PATENTE como puente y lo
         asociamos a la(s) OT(s) pendiente(s) del MISMO vehículo (misma patente).

    Parámetros:
      folios_pendientes      : set de FOLIO OT pendientes (los que están en la app)
      mapa_ot_patente        : { folio(cualquier estado): patente }  (puente)
      mapa_patente_pendiente : { patente: [folios_pendientes] }       (puente)

    Devuelve dict { folio_pendiente: [ {producto, descripcion, cantidad, total,
                    estado, fecha_bodega, origen, en_bodega, ot_origen,
                    via_patente, patente} ] }
    """
    if mapa_ot_patente is None:
        mapa_ot_patente = {}
    if mapa_patente_pendiente is None:
        mapa_patente_pendiente = {}

    archivos = (glob.glob(os.path.join(CARPETA_COMPRAS, "*.xlsx")) +
                glob.glob(os.path.join(CARPETA_COMPRAS, "*.xlsm")))
    archivos = [a for a in archivos if not os.path.basename(a).startswith("~$")]
    if not archivos:
        log(f"(i)  No se encontró archivo de compras en: {CARPETA_COMPRAS}")
        return {}

    # Preferir el archivo fijo "Seguimiento Compras.xlsx" (el usuario pega los datos
    # ahí cada día, en sus 2 hojas). Si no existe, tomar el más reciente de la carpeta.
    _fijo = [a for a in archivos
             if os.path.splitext(os.path.basename(a))[0].strip().lower() == "seguimiento compras"]
    ruta = _fijo[0] if _fijo else max(archivos, key=os.path.getmtime)
    log(f"Compras: leyendo {os.path.basename(ruta)}")

    folios_pendientes = set(str(f).strip().lstrip("0") for f in folios_pendientes)
    # Universo de OTs reales para validar la detección (pendientes + todas las del PBI)
    folios_reales = set(mapa_ot_patente.keys()) | folios_pendientes

    # El archivo suele quedar ABIERTO en Excel (bloqueo de Windows) -> "Permission
    # denied". Para evitarlo, copiamos el archivo a un temporal y leemos esa copia.
    _tmp_compras = None
    ruta_lectura = ruta
    try:
        import tempfile
        _fd, _tmp_compras = tempfile.mkstemp(suffix=os.path.splitext(ruta)[1] or ".xlsx")
        os.close(_fd)
        shutil.copy2(ruta, _tmp_compras)
        ruta_lectura = _tmp_compras
    except Exception:
        # Si no se pudo copiar (p.ej. bloqueo total), intentamos abrir el original
        ruta_lectura = ruta
        if _tmp_compras and os.path.exists(_tmp_compras):
            try:
                os.remove(_tmp_compras)
            except Exception:
                pass
            _tmp_compras = None

    try:
        wb = load_workbook(ruta_lectura, read_only=True, data_only=True)
    except Exception as e:
        log(f"(!)  No se pudo abrir el archivo de compras: {e}")
        log("     -> Cierra 'Seguimiento Compras.xlsx' en Excel y vuelve a correr el BAT.")
        if _tmp_compras and os.path.exists(_tmp_compras):
            try:
                os.remove(_tmp_compras)
            except Exception:
                pass
        return {}

    resultado = {}
    vistos = {}   # folio_pendiente -> set de claves (para deduplicar líneas idénticas)
    n_directas = 0
    n_puente   = 0

    hojas_detectadas = _detectar_hojas_compras(wb.sheetnames)
    if not hojas_detectadas:
        log("(!)  Compras: no se reconocieron hojas 'Nacional' ni 'Importación/Internacional'. "
            f"Hojas presentes: {wb.sheetnames}")
    for hoja_real, origen in hojas_detectadas:
        ws = wb[hoja_real]
        h_idx, header = _encontrar_header_compras(ws)
        if header is None:
            log(f"(!)  Compras: sin encabezado (no se halló 'Comentario Lineal') en '{hoja_real}'")
            continue

        ci_com  = _col_idx_compras(header, "Comentario Lineal")
        ci_est  = _col_idx_compras(header, "Estado Documento Base")
        ci_prod = _col_idx_compras(header, "Producto")
        ci_desc = _col_idx_compras(header, "Descripción Producto")
        ci_cant = _col_idx_compras(header, "Cantidad")
        ci_tot  = _col_idx_compras(header, "Total")
        # Fecha en bodega: probar candidatas (P/E o Recepción), usar la primera que exista
        ci_fec = None
        for _cf in FECHAS_BODEGA_COMPRAS:
            ci_fec = _col_idx_compras(header, _cf)
            if ci_fec is not None:
                break

        if ci_com is None:
            log(f"(!)  Compras: 'Comentario Lineal' no encontrada en '{hoja_real}'")
            continue

        n_cruz = 0
        for j, row in enumerate(ws.iter_rows(min_row=1, values_only=True)):
            if j <= h_idx:
                continue
            if row is None or all(c is None for c in row):
                continue
            com = row[ci_com] if ci_com < len(row) else None
            if not com:
                continue
            nums = [n.lstrip("0") for n in RE_OT_COMPRAS.findall(str(com))]
            # Solo números que correspondan a una OT real del PBI (filtra falsos positivos)
            ots_ref = [n for n in dict.fromkeys(nums) if n in folios_reales]
            if not ots_ref:
                continue

            estado_raw = (str(row[ci_est]).strip()
                          if (ci_est is not None and ci_est < len(row) and row[ci_est] is not None)
                          else "")
            en_bodega = estado_raw.upper() == ESTADO_EN_BODEGA

            def _v(ci):
                return (str(row[ci]).strip()
                        if (ci is not None and ci < len(row) and row[ci] is not None) else "")

            _cod_prod = _v(ci_prod)
            base = {
                "producto":     _cod_prod,
                "descripcion":  _v(ci_desc),
                "cantidad":     _v(ci_cant),
                "total":        _v(ci_tot),
                "estado":       estado_raw,
                "fecha_bodega": _fmt_fecha_compras(row[ci_fec]) if (ci_fec is not None and ci_fec < len(row)) else "",
                "origen":       origen,
                "en_bodega":    en_bodega,
                "bodega":       (mapa_stock.get(_norm_cod_producto(_cod_prod), {}).get("bodega", "")
                                 if mapa_stock else ""),
                "stock":        (mapa_stock.get(_norm_cod_producto(_cod_prod), {}).get("stock", None)
                                 if mapa_stock else None),
                "costo":        (mapa_stock.get(_norm_cod_producto(_cod_prod), {}).get("costo", None)
                                 if mapa_stock else None),
            }

            for ot_ref in ots_ref:
                patente = mapa_ot_patente.get(ot_ref, "")
                # Determinar a qué OT(s) PENDIENTE(s) asociar
                if ot_ref in folios_pendientes:
                    destinos = [(ot_ref, False)]           # asociación directa
                else:
                    pend = mapa_patente_pendiente.get(patente, []) if patente else []
                    destinos = [(f, True) for f in pend]   # puente por patente

                for folio_pend, via_pat in destinos:
                    rec = dict(base)
                    rec["ot_origen"]   = ot_ref
                    rec["via_patente"] = via_pat
                    rec["patente"]     = patente
                    clave = (rec["producto"], rec["descripcion"], rec["cantidad"],
                             rec["total"], rec["estado"], rec["fecha_bodega"],
                             rec["origen"], ot_ref)
                    vs = vistos.setdefault(folio_pend, set())
                    if clave in vs:
                        continue
                    vs.add(clave)
                    resultado.setdefault(folio_pend, []).append(rec)
                    n_cruz += 1
                    if via_pat:
                        n_puente += 1
                    else:
                        n_directas += 1

        log(f"  {hoja_real}: {n_cruz} línea(s) de repuesto enlazada(s)")

    try:
        wb.close()
    except Exception:
        pass

    # Borrar la copia temporal del archivo de compras
    if _tmp_compras and os.path.exists(_tmp_compras):
        try:
            os.remove(_tmp_compras)
        except Exception:
            pass

    log(f"Compras: {len(resultado)} OT(s) pendientes con repuestos "
        f"({n_directas} directas · {n_puente} por patente)")
    return resultado


def leer_stock_repuestos():
    """
    Lee 'Stock repuestos.xlsx' desde CARPETA_COMPRAS (headers en fila 1, hoja única).
    Construye un mapa { codigo_producto_norm: bodega } que se usa para enriquecer
    los repuestos del Seguimiento de Compras con su ubicación física en bodega.

    Si un mismo código aparece en varias bodegas, se concatenan con " / ".
    Si el archivo no existe o falla, devuelve {} sin interrumpir el proceso.
    """
    ruta = os.path.join(CARPETA_COMPRAS, NOMBRE_STOCK)
    if not os.path.exists(ruta):
        log(f"(i)  '{NOMBRE_STOCK}' no encontrado en Seguimiento de Compras — "
            "se omite la ubicación en bodega")
        return {}

    # Copiar a temporal para evitar bloqueo de Windows / OneDrive
    _tmp = None
    ruta_lectura = ruta
    try:
        import tempfile as _tf
        _fd, _tmp = _tf.mkstemp(suffix=".xlsx")
        os.close(_fd)
        shutil.copy2(ruta, _tmp)
        ruta_lectura = _tmp
    except Exception:
        ruta_lectura = ruta
        if _tmp and os.path.exists(_tmp):
            try:
                os.remove(_tmp)
            except Exception:
                pass
            _tmp = None

    try:
        df_stock = pd.read_excel(ruta_lectura, header=0, engine="calamine", dtype=str)
        df_stock.columns = [str(c).strip() for c in df_stock.columns]

        # Búsqueda flexible de columnas requeridas
        col_prod = next(
            (c for c in df_stock.columns if c.strip().lower() == "producto"), None
        )
        col_bod = next(
            (c for c in df_stock.columns if c.strip().lower() == "bodega"), None
        )
        # Columna de cantidad: "Stock" (nombre confirmado); fallbacks por si cambia el nombre
        col_stock = next(
            (c for c in df_stock.columns
             if c.strip().lower() in ("stock", "cantidad", "saldo", "existencia")),
            None
        )
        # Columna de costo unitario real del repuesto (viene del nuevo Stock Repestos Costo.xlsx)
        col_costo = next(
            (c for c in df_stock.columns if c.strip().lower() == "costo"),
            None
        )
        # Columnas de clasificacion — usadas para buscar variantes/repuestos compatibles
        # (Pre-picking, sesion 13/07/2026): mismo producto en otro formato/envase.
        col_desc = next((c for c in df_stock.columns if c.strip().lower() in ("descripcion", "descripción")), None)
        col_fam = next((c for c in df_stock.columns if c.strip().lower() == "familia"), None)
        col_subfam = next((c for c in df_stock.columns if c.strip().lower() in ("subfamilia", "sub familia")), None)

        if not col_prod or not col_bod:
            log(f"(!)  '{NOMBRE_STOCK}': no se encontraron columnas 'Producto' y/o 'Bodega'. "
                f"Columnas disponibles: {list(df_stock.columns[:10])}")
            return {}
        if not col_stock:
            log(f"(!)  '{NOMBRE_STOCK}': no se encontró columna de cantidad "
                f"(Stock/Cantidad/Saldo). Columnas: {list(df_stock.columns[:10])}")
        if col_costo:
            log(f"(i)  '{NOMBRE_STOCK}': columna Costo encontrada → se usará como costo unitario")
        else:
            log(f"(i)  '{NOMBRE_STOCK}': columna Costo no encontrada — se mostrará sin costo unitario")

        # mapa: { cod_normalizado: {"bodega": "...", "stock": N, "costo": C} }
        # Si el mismo código aparece en múltiples bodegas: bodega concatenada, stock sumado,
        # costo = promedio ponderado (o primera ocurrencia si no hay stock).
        # detalle: { cod_normalizado: {"descripcion","familia","subfamilia",
        #            "por_bodega": {bodega: {"stock":N,"costo":C}}} } — mismo recorrido,
        # pero SIN aplanar por bodega, para poder filtrar por sucursal despues y para
        # poder agrupar por Familia/SubFamilia al buscar repuestos compatibles/variantes
        # (Pre-picking, sesion 13/07/2026).
        mapa = {}
        detalle = {}
        for _, row in df_stock.iterrows():
            cod = _norm_cod_producto(str(row.get(col_prod, "") or ""))
            bod = str(row.get(col_bod, "") or "").strip()
            if not cod or not bod or bod.lower() in ("nan", "none", ""):
                continue
            try:
                qty = float(str(row.get(col_stock, "0") or "0").replace(",", ".")) if col_stock else 0.0
            except (ValueError, TypeError):
                qty = 0.0
            try:
                costo_val = float(str(row.get(col_costo, "0") or "0").replace(",", ".")) if col_costo else 0.0
            except (ValueError, TypeError):
                costo_val = 0.0
            if cod not in mapa:
                mapa[cod] = {"bodega": bod, "stock": qty, "costo": costo_val}
            else:
                if bod not in mapa[cod]["bodega"].split(" / "):
                    mapa[cod]["bodega"] = mapa[cod]["bodega"] + " / " + bod
                mapa[cod]["stock"] += qty
                # Costo: si ya teníamos uno, mantener el primero (precio más específico)
                if mapa[cod]["costo"] == 0.0 and costo_val > 0.0:
                    mapa[cod]["costo"] = costo_val

            d = detalle.get(cod)
            if d is None:
                d = {
                    "descripcion": str(row.get(col_desc, "") or "").strip() if col_desc else "",
                    "familia": str(row.get(col_fam, "") or "").strip() if col_fam else "",
                    "subfamilia": str(row.get(col_subfam, "") or "").strip() if col_subfam else "",
                    "por_bodega": {},
                }
                detalle[cod] = d
            pb = d["por_bodega"]
            if bod not in pb:
                pb[bod] = {"stock": qty, "costo": costo_val}
            else:
                pb[bod]["stock"] += qty
                if pb[bod]["costo"] == 0.0 and costo_val > 0.0:
                    pb[bod]["costo"] = costo_val

        _set_stock_detalle_cache(detalle)
        log(f"Stock repuestos: {len(mapa):,} código(s) mapeados")
        return mapa

    except Exception as e:
        log(f"(!)  Error al leer '{NOMBRE_STOCK}': {e}")
        return {}
    finally:
        if _tmp and os.path.exists(_tmp):
            try:
                os.remove(_tmp)
            except Exception:
                pass


def exportar_stock_repuestos_completo():
    """
    Lee 'Stock Repestos Costo.xlsx' COMPLETO (catálogo de repuestos, ~33.000
    productos) para exportarlo a stock_repuestos.json en GitHub.

    A diferencia de leer_stock_repuestos() (que solo arma un mapa liviano
    bodega/stock/costo por código, usado para enriquecer repuestos_compras —
    y por lo tanto solo cubre repuestos que alguna vez se pidieron para una
    OT), este catálogo es independiente de las OTs: incluye TODOS los
    productos del archivo, tengan o no movimiento asociado a una OT pendiente.
    Es lo que alimenta la búsqueda general de "stock de repuestos" del
    Asistente App. Se sube siempre completo (no incremental). 08/07/2026.

    Si el archivo no existe o falla la lectura, devuelve [] sin interrumpir
    el resto de la consolidación.
    """
    ruta = os.path.join(CARPETA_COMPRAS, NOMBRE_STOCK)
    if not os.path.exists(ruta):
        log(f"(i)  '{NOMBRE_STOCK}' no encontrado — se omite el export del catálogo de stock")
        return []

    _tmp = None
    ruta_lectura = ruta
    try:
        import tempfile as _tf
        _fd, _tmp = _tf.mkstemp(suffix=".xlsx")
        os.close(_fd)
        shutil.copy2(ruta, _tmp)
        ruta_lectura = _tmp
    except Exception:
        ruta_lectura = ruta
        if _tmp and os.path.exists(_tmp):
            try:
                os.remove(_tmp)
            except Exception:
                pass
            _tmp = None

    try:
        df = pd.read_excel(ruta_lectura, header=0, engine="calamine", dtype=str)
        df.columns = [str(c).strip() for c in df.columns]

        _cols_map = {
            "Producto": "producto", "Descripción": "descripcion", "Stock": "stock",
            "Stock Proyectado": "stock_proyectado", "Precio Venta": "precio_venta",
            "Bodega": "bodega", "Costo": "costo", "Familia": "familia",
            "SubFamilia": "subfamilia", "Procedencia": "procedencia",
            "Categoria": "categoria", "Clasificacion Stock": "clasificacion_stock",
        }
        _cols_presentes = {c: v for c, v in _cols_map.items() if c in df.columns}
        if "Producto" not in _cols_presentes:
            log(f"(!)  '{NOMBRE_STOCK}': no se encontró columna 'Producto' — se omite el export")
            return []

        df = df[list(_cols_presentes.keys())].rename(columns=_cols_presentes)
        df = df[df["producto"].fillna("").astype(str).str.strip() != ""]

        for _c in ["stock", "stock_proyectado", "precio_venta", "costo"]:
            if _c in df.columns:
                df[_c] = pd.to_numeric(
                    df[_c].astype(str).str.replace(",", "."), errors="coerce"
                ).fillna(0)
        for _c in ["producto", "descripcion", "bodega", "familia", "subfamilia",
                   "procedencia", "categoria", "clasificacion_stock"]:
            if _c in df.columns:
                df[_c] = df[_c].fillna("").astype(str).str.strip()

        registros = df.to_dict(orient="records")
        log(f"Catálogo de Stock de Repuestos: {len(registros):,} producto(s) leídos "
            f"({os.path.basename(ruta)})")
        return registros

    except Exception as e:
        log(f"(!)  Error al leer catálogo completo de '{NOMBRE_STOCK}': {e}")
        return []
    finally:
        if _tmp and os.path.exists(_tmp):
            try:
                os.remove(_tmp)
            except Exception:
                pass


def leer_tabla_vc():
    """
    Lee 'tabla vc.csv' desde CARPETA_COMPRAS y devuelve un set con los códigos
    de producto (columna 'Id Prod.') normalizados.

    Un repuesto del Seguimiento de Compras que aparece en este set ya fue
    consumido en alguna OT (posiblemente cerrada y fuera del radar de la app).
    Esos repuestos se excluyen de "en bodega pendiente de instalar" en los
    paneles Repuestos Pendientes y Patentes a Contactar.

    Si el archivo no existe o falla, devuelve set() sin interrumpir el proceso.
    """
    ruta = os.path.join(CARPETA_COMPRAS, NOMBRE_TABLA_VC)
    if not os.path.exists(ruta):
        log(f"(i)  '{NOMBRE_TABLA_VC}' no encontrado en Seguimiento de Compras — "
            "no se aplicará filtro de VC global")
        return set()
    try:
        df_vc = pd.read_csv(
            ruta, sep=";", encoding="utf-8-sig", dtype=str,
            low_memory=False, usecols=["Id Prod."]
        )
        codigos = {
            _norm_cod_producto(v)
            for v in df_vc["Id Prod."].dropna()
            if str(v).strip()
        }
        log(f"Tabla VC global cargada: {len(codigos):,} códigos de productos consumidos")
        return codigos
    except Exception as e:
        log(f"(!)  Error al leer '{NOMBRE_TABLA_VC}': {e}")
        return set()


def leer_patente_cliente():
    """
    Lee 'Patente-Cliente.xlsx' desde CARPETA_PBI.
    Devuelve un dict { PATENTE_UPPER: "RUT1 / RUT2 / ..." }.
    Si una patente tiene más de un RUT asociado, se concatenan con " / ".
    Si el archivo no existe o falla, devuelve {} sin interrumpir el proceso.
    """
    if not os.path.exists(RUTA_PATENTE_CLIENTE):
        log(f"(i)  'Patente-Cliente.xlsx' no encontrado en PBI — sin datos de RUT cliente")
        return {}
    try:
        df_pc = pd.read_excel(RUTA_PATENTE_CLIENTE, dtype=str)
        df_pc.columns = [str(c).strip() for c in df_pc.columns]
        col_pat = next(
            (c for c in df_pc.columns if c.strip().lower() in ("patente", "pat", "placa")),
            None
        )
        col_rut = next(
            (c for c in df_pc.columns if c.strip().lower() in ("rut", "rut cliente", "run", "rut_cliente")),
            None
        )
        if not col_pat or not col_rut:
            log(f"(!)  'Patente-Cliente.xlsx': no se encontraron columnas 'Patente' y/o 'RUT'. "
                f"Columnas disponibles: {list(df_pc.columns)}")
            return {}
        mapa = {}   # { patente: [rut1, rut2, ...] }
        for _, row in df_pc.iterrows():
            pat = str(row.get(col_pat, "") or "").strip().upper()
            rut = str(row.get(col_rut, "") or "").strip()
            if not pat or pat.lower() in ("nan", "none", ""):
                continue
            if not rut or rut.lower() in ("nan", "none", ""):
                continue
            if pat not in mapa:
                mapa[pat] = [rut]
            elif rut not in mapa[pat]:
                mapa[pat].append(rut)
        resultado = {pat: " / ".join(ruts) for pat, ruts in mapa.items()}
        multi = sum(1 for ruts in mapa.values() if len(ruts) > 1)
        log(f"Patente-Cliente: {len(resultado):,} patentes ({multi:,} con múltiples RUTs)")
        return resultado
    except Exception as e:
        log(f"(!)  Error al leer 'Patente-Cliente.xlsx': {e}")
        return {}


_RE_RUT_ANTICIPO = re.compile(r'^\d{6,8}-[\dkK]$')


def _aq(v):
    """
    Extrae el valor string de una celda, quitando el apóstrofe de prefijo
    que Excel usa para forzar formato texto (ej: la celda muestra 10031659-5
    pero al leer con openpyxl/calamine aparece como '10031659-5').
    También normaliza None / nan / vacío a "".
    """
    if v is None:
        return ""
    s = str(v).strip()
    if s.lower() in ("none", "nan", ""):
        return ""
    if s.startswith("'"):
        s = s[1:].strip()
    return s


def _detectar_indices_anticipo(rows):
    """
    Auto-detecta las posiciones de columna del archivo de anticipos.

    Hay dos formatos posibles:
      A) "Pegar Datos" (RESUMEN_ANTICIPO_TALLER_2.xlsx):
           SALDO=G(6) · LOCAL=H(7) · DOC=I(8) · FECHA=J(9) · NRO=L(11)
      B) XLS bruto del ERP (ej: INFORME CUENTA FICHA...xls):
           SALDO=H(7) · LOCAL=I(8) · DOC=J(9) · FECHA=K(10) · NRO=N(13)

    Detecta el formato buscando la primera fila de detalle (col A vacía, col D con glosa)
    y comprobando si col G (índice 6) tiene valor numérico o es vacía.
    """
    for row in rows[:100]:
        if not row or len(row) < 8:
            continue
        col_a = _aq(row[0])
        if col_a:                                   # las filas de detalle tienen col A vacía
            continue
        glosa = _aq(row[3]) if len(row) > 3 else ""
        if not glosa:
            continue
        # Primera fila de detalle real — detectar dónde está el saldo
        col_g = row[6] if len(row) > 6 else None
        col_g_val = _aq(col_g)
        if col_g_val and col_g_val not in ("nan", "0", ""):
            # col G tiene valor → formato A (Pegar Datos)
            return dict(saldo=6, local=7, doc=8, fecha=10, nro=11)
        else:
            # col G vacía → formato B (XLS bruto ERP)
            return dict(saldo=7, local=8, doc=9, fecha=10, nro=13)
    # Por defecto: formato B (XLS bruto)
    return dict(saldo=7, local=8, doc=9, fecha=10, nro=13)


def _parsear_sabana_anticipo_raw(rows):
    """
    Parsea el formato jerárquico que entrega el ERP o la hoja "Pegar Datos".

    Detecta automáticamente dos variantes de columnas:
      - XLS bruto del ERP:  SALDO en H(7), FECHA en K(10), NRO en N(13)
      - Hoja "Pegar Datos": SALDO en G(6), FECHA en J(9),  NRO en L(11)

    Estructura de filas (común a ambos formatos):
      - Fila RUT header:     col A = RUT (ej "12345678-9"), col B = NOMBRE
      - Fila de detalle:     col A = vacía, col D = GLOSA, col H/G = SALDO
      - Fila TOTAL ANALISIS: col A = "TOTAL ANALISIS 1", col H/G = saldo total del RUT
      - Resto (totales, encabezados de cuenta): se ignoran

    El saldo es NEGATIVO cuando el cliente tiene dinero depositado.
    Se almacena como positivo (monto disponible).

    Devuelve:
        { RUT_UPPER: {
            "nombre":      str,
            "total":       int,
            "tiene_saldo": bool,
            "movimientos": [{"documento", "nro", "saldo", "fecha", "local", "glosa"}]
          } }
    """
    idx  = _detectar_indices_anticipo(rows)
    i_sal  = idx["saldo"]
    i_loc  = idx["local"]
    i_doc  = idx["doc"]
    i_fec  = idx["fecha"]
    i_nro  = idx["nro"]

    resultado   = {}
    current_rut = None

    def _get_num(row, i):
        try:
            v = row[i] if i < len(row) else None
            return int(float(_aq(v))) if v is not None and _aq(v) not in ("", "nan") else 0
        except (ValueError, TypeError):
            return 0

    def _get_fecha(row, i):
        if i >= len(row) or row[i] is None:
            return ""
        try:
            f = row[i]
            if hasattr(f, "strftime"):
                return f.strftime("%d/%m/%Y")
            s = _aq(f)
            # Formato "YYYY-MM-DD HH:MM:SS" (común en .xls leídos con calamine)
            if s and len(s) >= 10 and s[4] == "-":
                from datetime import datetime
                return datetime.strptime(s[:10], "%Y-%m-%d").strftime("%d/%m/%Y")
            return s
        except Exception:
            return _aq(row[i])

    for row in rows:
        if not row or all(v is None for v in row):
            continue

        col_a = _aq(row[0])

        # ── Encabezado de RUT ─────────────────────────────────────────────────
        if _RE_RUT_ANTICIPO.match(col_a):
            current_rut = col_a.upper()
            if current_rut not in resultado:
                resultado[current_rut] = {
                    "nombre":      _aq(row[1]) if len(row) > 1 else "",
                    "total":       0,
                    "tiene_saldo": False,
                    "movimientos": [],
                }
            continue

        if current_rut is None:
            continue

        col_a_up = col_a.upper()

        # ── Total acumulado por RUT (fila "TOTAL ANALISIS 1") ─────────────────
        if col_a_up.startswith("TOTAL ANALISIS"):
            total_raw = _get_num(row, i_sal)
            resultado[current_rut]["total"]       = abs(total_raw)
            resultado[current_rut]["tiene_saldo"] = total_raw < 0
            continue

        # ── Saltear totales de cuenta y encabezados de cuenta ─────────────────
        if col_a_up.startswith("TOTAL") or "2107004" in col_a:
            continue

        # ── Fila de detalle (col A vacía, col D = GLOSA) ──────────────────────
        if not col_a and len(row) > 3:
            glosa = _aq(row[3])
            if not glosa:
                continue

            saldo = _get_num(row, i_sal)

            resultado[current_rut]["movimientos"].append({
                "documento": _aq(row[i_doc]) if i_doc < len(row) else "",
                "nro":       _aq(row[i_nro]) if i_nro < len(row) else "",
                "saldo":     abs(saldo),
                "fecha":     _get_fecha(row, i_fec),
                "local":     _aq(row[i_loc]) if i_loc < len(row) else "",
                "glosa":     glosa,
            })

    return resultado


def _aplicar_sabana_virgen_a_resumen(filas_virgen, ruta_resumen=RUTA_ANTICIPO_TALLER):
    """
    Pega las filas de la sábana virgen (.xls bruto exportado del ERP) en la hoja
    "Pegar Datos" de RESUMEN_ANTICIPO_TALLER_2.xlsx — el mismo paso que antes se
    hacía a mano (copiar/pegar el reporte crudo a partir de A3).

    Se escribe SOLO en columnas A:M (los datos crudos). Las columnas N:AA
    (fórmulas que arman el detalle limpio, ya precargadas hasta la fila 10000)
    y las filas 1-2 (comentario del código de cuenta + instrucción de pegado)
    quedan intactas, igual que las hojas "Datos" y "Resumen por Rut".

    Limpia primero cualquier dato viejo en A3:M10000 antes de escribir el nuevo,
    para que no queden filas residuales de una sábana anterior más larga.

    Devuelve True si se aplicó correctamente, False si no (archivo/hoja faltante).
    """
    if not os.path.exists(ruta_resumen):
        log(f"(i)  {os.path.basename(ruta_resumen)} no existe — no se aplica sábana virgen")
        return False

    FILA_INICIO = 3
    FILA_MAX    = 10000
    CAP_FILAS   = FILA_MAX - FILA_INICIO + 1   # 9998

    n_filas = len(filas_virgen)
    n_cols  = max((len(f) for f in filas_virgen), default=0)

    if n_filas > CAP_FILAS:
        log(f"(!)  Sábana virgen tiene {n_filas:,} filas — excede el espacio de la plantilla "
            f"({CAP_FILAS:,}). Se truncará a las primeras {CAP_FILAS:,}.")
        filas_virgen = filas_virgen[:CAP_FILAS]
        n_filas = CAP_FILAS

    def _limpio(v):
        if v is None:
            return None
        if isinstance(v, float) and pd.isna(v):
            return None
        return v

    try:
        wb = load_workbook(ruta_resumen, data_only=False)
        if "Pegar Datos" not in wb.sheetnames:
            log(f"(!)  Hoja 'Pegar Datos' no encontrada en {os.path.basename(ruta_resumen)} — "
                f"no se aplica sábana virgen")
            wb.close()
            return False
        ws = wb["Pegar Datos"]

        # Limpiar datos viejos en A3:M10000 (sin tocar filas 1-2 ni fórmulas N:AA)
        for r in range(FILA_INICIO, FILA_MAX + 1):
            for c in range(1, 14):   # A..M
                ws.cell(row=r, column=c).value = None

        # Escribir la sábana nueva a partir de A3
        for i, fila in enumerate(filas_virgen):
            r = FILA_INICIO + i
            for c_idx in range(min(len(fila), 13)):   # nunca más allá de M
                ws.cell(row=r, column=c_idx + 1).value = _limpio(fila[c_idx])

        wb.save(ruta_resumen)
        wb.close()
        log(f"Sábana virgen aplicada a 'Pegar Datos': {n_filas:,} filas × {n_cols} columnas "
            f"→ {os.path.basename(ruta_resumen)}")
        return True
    except Exception as e:
        log(f"(!)  No se pudo aplicar la sábana virgen a {os.path.basename(ruta_resumen)}: {e}")
        return False


def leer_anticipo_taller():
    """
    Lee el archivo de anticipos taller y devuelve un mapa RUT → datos.

    PRIORIDAD de lectura:
      1. Cualquier .xls en CARPETA_PBI  → formato jerárquico bruto del ERP
         ("sábana virgen"). El usuario solo deja el archivo que exporta del
         sistema; no hace nada más. Este paso, ADEMÁS, aplica esa sábana
         virgen a RESUMEN_ANTICIPO_TALLER_2.xlsx (hoja "Pegar Datos") para
         que el archivo plantilla quede siempre actualizado — es el mismo
         paso de "pegar a mano" que antes se hacía manualmente, ahora
         automatizado dentro de la consolidación.
      2. RESUMEN_ANTICIPO_TALLER_2.xlsx hoja "Pegar Datos"
         → misma sábana bruta, pero pegada dentro del Excel plantilla.
      3. RESUMEN_ANTICIPO_TALLER_2.xlsx hojas "Datos" + "Resumen por Rut"
         → formato procesado (respaldo / compatibilidad con versiones anteriores).

    El saldo en origen es NEGATIVO cuando el cliente tiene dinero depositado.
    Se almacena como positivo (monto disponible).

    Devuelve:
        { RUT_UPPER: {
            "nombre":      str,
            "total":       int,
            "tiene_saldo": bool,
            "movimientos": [{"documento", "nro", "saldo", "fecha", "local", "glosa"}]
          } }
    """

    # ── 1. Buscar .xls bruto del ERP en CARPETA_PBI ──────────────────────────
    xls_raw = [
        f for f in glob.glob(os.path.join(CARPETA_PBI, "*.xls"))
        if not os.path.basename(f).startswith("~$")
    ]
    if xls_raw:
        ruta_raw = max(xls_raw, key=os.path.getmtime)
        try:
            df_raw = pd.read_excel(ruta_raw, sheet_name=0, header=None,
                                   engine="calamine", dtype=object)
            resultado = _parsear_sabana_anticipo_raw(df_raw.values.tolist())
            if resultado:
                con_saldo = sum(1 for v in resultado.values() if v["tiene_saldo"])
                log(f"Anticipos (XLS bruto): {len(resultado):,} RUTs | "
                    f"{con_saldo:,} con saldo  ←  {os.path.basename(ruta_raw)}")
                # Aplicar la sábana virgen a RESUMEN_ANTICIPO_TALLER_2.xlsx (hoja
                # "Pegar Datos"), igual que el pegado manual de antes, para que
                # el archivo plantilla quede siempre actualizado.
                _aplicar_sabana_virgen_a_resumen(df_raw.values.tolist())
                return resultado
        except Exception as e:
            log(f"(!)  No se pudo leer XLS bruto de anticipos ({os.path.basename(ruta_raw)}): {e}")

    # ── 2 & 3. Usar RESUMEN_ANTICIPO_TALLER_2.xlsx ───────────────────────────
    ruta = RUTA_ANTICIPO_TALLER
    if not os.path.exists(ruta):
        candidatos = (glob.glob(os.path.join(CARPETA_PBI, "*ANTICIPO*TALLER*.*"))
                      + glob.glob(os.path.join(CARPETA_PBI, "*RESUMEN*ANTICIPO*.*")))
        candidatos = [c for c in candidatos if not os.path.basename(c).startswith("~$")]
        if candidatos:
            ruta = max(candidatos, key=os.path.getmtime)
        else:
            log("(i)  Archivo de anticipos no encontrado en PBI — sin datos de Cuenta Ficha")
            return {}
    try:
        wb = load_workbook(ruta, read_only=True, data_only=True)

        # ── 2. Hoja "Pegar Datos" (sábana bruta pegada en la plantilla) ───────
        if "Pegar Datos" in wb.sheetnames:
            rows_pd = list(wb["Pegar Datos"].iter_rows(values_only=True))
            resultado = _parsear_sabana_anticipo_raw(rows_pd)
            if resultado:
                con_saldo = sum(1 for v in resultado.values() if v["tiene_saldo"])
                log(f"Anticipos (Pegar Datos): {len(resultado):,} RUTs | {con_saldo:,} con saldo")
                wb.close()
                return resultado

        # ── 3. Hojas procesadas "Resumen por Rut" + "Datos" (respaldo) ────────
        resumen = {}
        if "Resumen por Rut" in wb.sheetnames:
            rows_r = list(wb["Resumen por Rut"].iter_rows(values_only=True))
            hdridx = next((i for i, r in enumerate(rows_r)
                           if r and str(r[0]).strip().upper() == "RUT"), None)
            if hdridx is not None:
                for row in rows_r[hdridx + 1:]:
                    if not row or not row[0]:
                        continue
                    rut = str(row[0]).strip()
                    if not rut or rut.lower() in ("nan", "none", ""):
                        continue
                    nombre = str(row[1]).strip() if len(row) > 1 and row[1] else ""
                    try:
                        total = int(float(str(row[2] or 0))) if len(row) > 2 else 0
                    except (ValueError, TypeError):
                        total = 0
                    resumen[rut.upper()] = (nombre, total)

        movimientos = {}
        if "Datos" in wb.sheetnames:
            rows_d   = list(wb["Datos"].iter_rows(values_only=True))
            hdridx_d = next((i for i, r in enumerate(rows_d)
                             if r and str(r[0]).strip().upper() == "ID"), None)
            if hdridx_d is not None:
                hdr_d = [str(c).strip().upper() if c else "" for c in rows_d[hdridx_d]]
                def _ci(name):
                    return hdr_d.index(name) if name in hdr_d else None
                idx_rut = _ci("RUT");      idx_doc = _ci("DOCUMENTO")
                idx_nro = _ci("NRO");      idx_sal = _ci("SALDO")
                idx_fec = _ci("FECHA DE EMISION")
                idx_loc = _ci("LOCAL");    idx_glo = _ci("GLOSA")

                for row in rows_d[hdridx_d + 1:]:
                    if not row:
                        continue
                    rut_raw = str(row[idx_rut]).strip() if idx_rut is not None and idx_rut < len(row) else ""
                    if not rut_raw or rut_raw.lower() in ("nan", "none", ""):
                        continue
                    rut_key = rut_raw.upper()

                    def _get(idx):
                        return str(row[idx]).strip() if idx is not None and idx < len(row) and row[idx] else ""

                    try:
                        saldo = int(float(str(row[idx_sal] or 0))) if idx_sal is not None and idx_sal < len(row) else 0
                    except (ValueError, TypeError):
                        saldo = 0

                    fecha = ""
                    if idx_fec is not None and idx_fec < len(row) and row[idx_fec]:
                        try:
                            f = row[idx_fec]
                            fecha = f.strftime("%d/%m/%Y") if hasattr(f, "strftime") else str(f).strip()
                        except Exception:
                            fecha = str(row[idx_fec]).strip()

                    movimientos.setdefault(rut_key, []).append({
                        "documento": _get(idx_doc),
                        "nro":       _get(idx_nro),
                        "saldo":     abs(saldo),
                        "fecha":     fecha,
                        "local":     _get(idx_loc),
                        "glosa":     _get(idx_glo),
                    })

        wb.close()

        resultado = {}
        for rut_key, (nombre, total_raw) in resumen.items():
            resultado[rut_key] = {
                "nombre":      nombre,
                "total":       abs(total_raw),
                "tiene_saldo": total_raw < 0,
                "movimientos": movimientos.get(rut_key, []),
            }

        con_saldo = sum(1 for v in resultado.values() if v["tiene_saldo"])
        log(f"Anticipos (procesado): {len(resultado):,} RUTs | {con_saldo:,} con saldo")
        return resultado

    except Exception as e:
        log(f"(!)  Error al leer archivo de anticipos: {e}")
        return {}


# =============================================================
#   MODULO CUENTA FICHA
#   -------------------
#   Cruza 3 fuentes que hoy viven separadas:
#     1. INFORME FICHA CUENTA (.xls del ERP, carpeta PBI) -> saldo por RUT
#        (ya lo lee leer_anticipo_taller(), se reutiliza ese mapa tal cual).
#     2. Patente-Cliente.xlsx -> el puente: patente <-> RUT del cliente.
#     3. Seguimiento Servicio Tecnico -> historial de OT del cliente, con
#        TODOS los estados (Finalizado / Cerrado / Pendiente / Anulado) y
#        todos sus documentos posteriores.
#   El resultado se sube comprimido (gzip+base64) porque en crudo son ~9 MB.
# =============================================================

_ESTADOS_CF_GRUPO = {
    "PENDIENTE":  "Pendiente",
    "ANULADO":    "Anulado",
    "CERRADO":    "Cerrado",
    "FINALIZADO": "Finalizado",
}


def _cf_txt(v):
    """Normaliza cualquier celda del PBI a texto limpio ('' si viene vacia/nan)."""
    if v is None:
        return ""
    s = str(v).strip()
    return "" if s.lower() in ("nan", "none", "nat", "") else s


def _cf_folio(v):
    """Quita los ceros a la izquierda de un folio/numero de documento del ERP."""
    s = _cf_txt(v)
    if not s:
        return ""
    s2 = s.lstrip("0")
    return s2 if s2 else s


def _cf_fecha_ord(fecha_ddmmaaaa):
    """'31/07/2026' -> '2026-07-31' (ordenable). '' si no se puede parsear."""
    s = _cf_txt(fecha_ddmmaaaa)
    if len(s) >= 10 and s[2] == "/" and s[5] == "/":
        return f"{s[6:10]}-{s[3:5]}-{s[0:2]}"
    if len(s) >= 10 and s[4] == "-":          # ya viene ISO
        return s[:10]
    return ""


def mapear_patentes_cuenta_ficha(mapa_anticipo, mapa_patente_rut):
    """
    Invierte Patente-Cliente para los RUT que aparecen en el Informe Ficha
    Cuenta. Devuelve (rut_patentes, pat_ruts):
        rut_patentes = { RUT: {PATENTE, ...} }
        pat_ruts     = { PATENTE: [RUT, ...] }

    Se calcula ANTES de leer el PBI, porque leer_pbi() necesita saber qué
    patentes indexar para el historial (asi no hay que copiar el DataFrame).
    """
    if not mapa_anticipo:
        return {}, {}

    ruts_informe = set(mapa_anticipo.keys())
    rut_patentes = {}
    for _pat, _ruts_txt in (mapa_patente_rut or {}).items():
        pat = _cf_txt(_pat).upper().replace(" ", "")
        if not pat:
            continue
        for r in str(_ruts_txt).split("/"):
            r = r.strip().upper()
            if r and r in ruts_informe:
                rut_patentes.setdefault(r, set()).add(pat)

    pat_ruts = {}
    for r, pats in rut_patentes.items():
        for p in pats:
            pat_ruts.setdefault(p, []).append(r)

    log(f"Cuenta Ficha: {len(ruts_informe):,} RUT del informe · "
        f"{len(rut_patentes):,} con patente(s) asociada(s) · "
        f"{len(pat_ruts):,} patentes a rastrear en el PBI")
    return rut_patentes, pat_ruts


def _indexar_ots_cuenta_ficha(df, patentes_cf, meses=CUENTA_FICHA_MESES_HISTORIAL):
    """
    Recorre el DataFrame COMPLETO del PBI (todos los estados, antes del filtro
    de PENDIENTE/ANULADO) y arma el historial de OT de las patentes indicadas.

    Se hace en streaming (itertuples) devolviendo solo un dict con las OT que
    interesan: copiar el DataFrame entero para procesarlo despues costaba ~1 GB
    extra de memoria (260.000 filas x 70 columnas de texto).

    Devuelve { FOLIO_OT: {...datos de la OT..., "docs": {tipo: [{n, f}]}} }
    """
    if not patentes_cf or df is None or df.empty or "PATENTE" not in df.columns:
        return {}

    # Corte de la ventana de historial (por FECHA OT)
    _hoy = datetime.now()
    _anio_c = _hoy.year - (meses // 12)
    _mes_c  = _hoy.month - (meses % 12)
    if _mes_c <= 0:
        _mes_c += 12
        _anio_c -= 1
    corte = f"{_anio_c:04d}-{_mes_c:02d}-01"

    cols_pedidas = [c for c in COLS_CUENTA_FICHA_PBI if c in df.columns]
    if "PATENTE" not in cols_pedidas or CLAVE not in cols_pedidas:
        return {}
    pos = {c: k for k, c in enumerate(cols_pedidas)}

    def gv(row, col):
        j = pos.get(col)
        return _cf_txt(row[j]) if j is not None else ""

    docs_cfg = [(n, c, f) for n, c, f in DOCS_CUENTA_FICHA if c in df.columns]

    ots     = {}
    fuera   = set()
    n_filas = 0

    # Se itera columna por columna con zip() en vez de df[cols].itertuples():
    # seleccionar un sub-DataFrame materializa una copia de 260.000 filas x 43
    # columnas de texto (~1 GB) y eso reventaba la memoria. Cada .to_numpy()
    # de una sola columna son ~2 MB.
    _arrays = [df[c].to_numpy() for c in cols_pedidas]

    for row in zip(*_arrays):
        pat = gv(row, "PATENTE").upper().replace(" ", "")
        if pat not in patentes_cf:
            continue
        folio = _cf_folio(gv(row, CLAVE))
        if not folio or folio in fuera:
            continue
        n_filas += 1

        o = ots.get(folio)
        if o is None:
            fec     = gv(row, "FECHA OT")
            fec_ord = _cf_fecha_ord(fec)
            est     = gv(row, "ESTADO")
            # Las OT PENDIENTES entran SIEMPRE, aunque sean mas antiguas que la
            # ventana: son justamente las que hay que destacar antes de tocar el
            # saldo del cliente (verificado con datos reales: 2 OT pendientes de
            # 2024 se perdian con el corte de 24 meses).
            if fec_ord and fec_ord < corte and est.strip().upper() != "PENDIENTE":
                fuera.add(folio)
                continue
            o = ots[folio] = {
                "ot":      folio,
                "imp":     gv(row, "IMPORTADOR"),
                "suc":     gv(row, "SUCURSAL"),
                "pat":     pat,
                "fec":     fec,
                "fec_ord": fec_ord,
                "est":     _ESTADOS_CF_GRUPO.get(est.upper(), est or "Sin estado"),
                "estd":    gv(row, "ESTADO DETALLADO"),
                "neto":    gv(row, "NETO"),
                "tv":      gv(row, "TIPO VENTA"),
                "mar":     gv(row, "MARCA"),
                "mod":     gv(row, "MODELO"),
                "ase":     gv(row, "ASESOR"),
                "dias":    gv(row, "DIAS APERTURA"),
                "cierre":  gv(row, "TIPO CIERRE"),
                "glosa":   gv(row, "GLOSA TRABAJO")[:300],
                "docs":    {},
            }

        # Documentos posteriores: el PBI repite la OT una vez por repuesto del
        # Vale de Consumo, asi que se acumulan sin duplicar.
        for nombre, col_folio, col_fecha in docs_cfg:
            num = _cf_folio(gv(row, col_folio))
            if not num:
                continue
            lst = o["docs"].setdefault(nombre, [])
            if not any(d["n"] == num for d in lst):
                lst.append({"n": num, "f": gv(row, col_fecha)})

    log(f"Cuenta Ficha: {len(ots):,} OT de clientes con ficha dentro de los "
        f"ultimos {meses} meses ({len(fuera):,} OT mas antiguas omitidas · "
        f"{n_filas:,} filas del PBI revisadas)")
    return ots


def generar_cuenta_ficha(mapa_anticipo, rut_patentes, ots,
                         meses=CUENTA_FICHA_MESES_HISTORIAL):
    """
    Arma el payload del modulo Cuenta Ficha.

    Parametros
    ----------
    mapa_anticipo : dict  { RUT: {nombre, total, tiene_saldo, movimientos:[...]} }
                    tal cual lo devuelve leer_anticipo_taller().
    rut_patentes  : dict  { RUT: {PATENTE, ...} } de mapear_patentes_cuenta_ficha().
    ots           : dict  { FOLIO: {...} } de _indexar_ots_cuenta_ficha()
                    (leer_pbi lo devuelve como 4to valor).
    meses         : ventana de historial usada (informativa, va en el resumen).

    Devuelve { "clientes": [...], "resumen": {...} } o {} si no hay datos.
    """
    if not mapa_anticipo:
        log("(i)  Cuenta Ficha: sin datos de anticipos — se omite")
        return {}

    ots          = ots or {}
    rut_patentes = rut_patentes or {}

    pat_ruts = {}
    for r, pats in rut_patentes.items():
        for p in pats:
            pat_ruts.setdefault(p, []).append(r)

    # ── Armar la ficha de cada cliente ────────────────────────────────────────
    rut_ots = {}
    for folio, o in ots.items():
        for r in pat_ruts.get(o["pat"], []):
            rut_ots.setdefault(r, []).append(o)

    clientes = []
    for rut, datos in mapa_anticipo.items():
        movs = datos.get("movimientos", []) or []

        # Saldo por sucursal (un mismo cliente puede tener plata en varias)
        por_suc = {}
        for m in movs:
            loc = _cf_txt(m.get("local")) or "Sin sucursal"
            d = por_suc.setdefault(loc, {"suc": loc, "monto": 0, "n": 0})
            d["monto"] += int(m.get("saldo") or 0)
            d["n"] += 1
        sucursales = sorted(por_suc.values(), key=lambda x: -x["monto"])

        lista_ots = sorted(rut_ots.get(rut, []),
                           key=lambda o: (o.get("fec_ord") or "", o["ot"]), reverse=True)
        n_pend = sum(1 for o in lista_ots if o["est"] == "Pendiente")

        clientes.append({
            "rut":         rut,
            "nombre":      _cf_txt(datos.get("nombre")) or "(sin nombre)",
            "saldo":       int(datos.get("total") or 0),
            "tiene_saldo": bool(datos.get("tiene_saldo")),
            "n_mov":       len(movs),
            "sucursales":  sucursales,
            "suc_principal": sucursales[0]["suc"] if sucursales else "",
            "movimientos": movs,
            "patentes":    sorted(rut_patentes.get(rut, [])),
            "ots":         lista_ots,
            "n_ot":        len(lista_ots),
            "n_ot_pend":   n_pend,
        })

    clientes.sort(key=lambda c: (-c["saldo"] if c["tiene_saldo"] else 0, c["nombre"]))

    con_saldo = [c for c in clientes if c["tiene_saldo"] and c["saldo"] > 0]
    total_disp = sum(c["saldo"] for c in con_saldo)
    resumen = {
        "total_clientes":   len(clientes),
        "clientes_saldo":   len(con_saldo),
        "monto_total":      total_disp,
        "saldo_promedio":   int(total_disp / len(con_saldo)) if con_saldo else 0,
        "saldo_mayor":      max((c["saldo"] for c in con_saldo), default=0),
        "total_movimientos": sum(c["n_mov"] for c in clientes),
        "clientes_con_ot":  sum(1 for c in clientes if c["n_ot"]),
        "total_ot":         len(ots),
        "total_ot_pend":    sum(1 for o in ots.values() if o["est"] == "Pendiente"),
        "meses_historial":  meses,
    }

    log(f"Cuenta Ficha: {resumen['total_clientes']:,} cliente(s) · "
        f"{resumen['clientes_saldo']:,} con saldo (${resumen['monto_total']:,}) · "
        f"{resumen['clientes_con_ot']:,} con historial de OT · "
        f"{resumen['total_ot_pend']:,} OT pendiente(s)")

    return {"clientes": clientes, "resumen": resumen}


def subir_cuenta_ficha(payload):
    """
    Sube cuenta_ficha.json a GitHub con el payload comprimido (gzip + base64).

    En crudo el JSON pesa ~9 MB (20.000+ OT con sus documentos); comprimido
    baja a ~1,5 MB — mismo patron ya usado por cotizador_data.json. El archivo
    se sube por la Git Data API (no la Contents API), que es la que aguanta
    archivos grandes sin el 403 "Timed out validating rule".
    """
    if not payload or not payload.get("clientes"):
        return False

    import gzip as _gzip

    crudo = json.dumps(payload, ensure_ascii=False, separators=(",", ":")).encode("utf-8")
    gz_b64 = base64.b64encode(_gzip.compress(crudo, 6)).decode("ascii")

    archivo = {
        "fecha_actualizacion": datetime.now().strftime("%d/%m/%Y %H:%M"),
        "resumen":             payload.get("resumen", {}),
        "gz":                  gz_b64,
    }

    ok = _subir_json_github_gitdata(
        GITHUB_CUENTA_FICHA, archivo,
        f"Cuenta Ficha actualizada {datetime.now().strftime('%d/%m/%Y %H:%M')}",
        timeout=120,
    )
    if ok:
        log(f"Cuenta Ficha subida a GitHub -> {GITHUB_CUENTA_FICHA} "
            f"({len(crudo)/1_000_000:.1f} MB en crudo -> {len(gz_b64)/1_000_000:.1f} MB comprimida)")
    else:
        log("(!) Error al subir Cuenta Ficha a GitHub")
    return ok


COLS_EXPORTAR = [
    "FOLIO OT", "SUCURSAL", "RANGO", "DIAS APERTURA", "FECHA OT", "AÑO",
    # TIPO CLIENTE: columna BM del "Seguimiento Servicio Tecnico" del PBI.
    # Valores reales: *PARTICULAR, *CURIFOR, *GARANTIA y las companias de
    # seguro (CIA-SEG BCI, CIA-SEG HDI, etc.). Se exporta desde el 06/08/2026
    # para mostrarla en el listado de OTs. leer_pbi() ya conserva todas las
    # columnas del PBI, asi que solo hacia falta incluirla aca.
    "TIPO VENTA", "TIPO CLIENTE", "MARCA", "MODELO", "PATENTE", "ASESOR",
    "ESTADO", "IMPORTADOR", "NETO", "GLOSA TRABAJO",
    "CATEGORIA", "OBSERVACION OT", "NOTAS", "AVANCE - GESTIÓN", "ULTIMA_EDICION",
    "_FECHA_APERTURA",   # fecha real de apertura (para recalcular DIAS correctamente)
    "_MARCA_COLOR_",     # marca visual asignada desde el dashboard web
    "ETAPA_JPCB",        # etapa de taller asignada desde el kanban JPCB del dashboard
    # Documentos asociados
    "N_LIQ_ST",       "FOLIOS_LIQ_ST",
    "N_FACT_CLIENTE",  "FOLIOS_FACT_CLIENTE",
    "N_FACT_COMPANIA", "FOLIOS_FACT_COMPANIA",
    "N_CARGO_INT",     "FOLIOS_CARGO_INT",
    "N_CARGO_GTIA",    "FOLIOS_CARGO_GTIA",
    "N_FACT_GTIA",     "FOLIOS_FACT_GTIA",
    "N_VALE_CONSUMO",  "FOLIOS_VALE_CONSUMO",
    # Fechas de facturas (para Facturas X)
    "FECHA_FACT_CLIENTE", "FECHA_FACT_COMPANIA",
]


def _norm_cod_producto(p):
    """Normaliza un código de producto para comparar Compras vs Vale de Consumo."""
    return re.sub(r"\s+", " ", str(p).strip().upper())


def exportar_json(df, repuestos_actuales=None, repuestos_compras=None, codigos_vc_global=None, mapa_patente_rut=None, mapa_anticipo=None):
    """
    Genera datos_dashboard.json con los datos para el dashboard web.

    repuestos_actuales: dict { folio: [lista de repuestos del Vale de Consumo] }
        Si se pasa, cada OT en el JSON tendrá:
          - "repuestos_actual"   : lista de repuestos en el PBI de hoy
          - "repuestos_historico": lista acumulada; los eliminados quedan marcados

    repuestos_compras: dict { folio: [lista de repuestos del Seguimiento de Compras] }
        Si se pasa, cada OT en el JSON tendrá:
          - "repuestos_compras"  : repuestos cruzados desde el archivo de compras
                                   (en espera de bodega / en bodega)
    """
    if repuestos_actuales is None:
        repuestos_actuales = {}
    if repuestos_compras is None:
        repuestos_compras = {}
    if codigos_vc_global is None:
        codigos_vc_global = set()
    if mapa_patente_rut is None:
        mapa_patente_rut = {}

    # Leer JSON anterior para preservar colores, campos de gestión y repuestos históricos
    ruta_prev = RUTA_JSON
    _colores_prev    = {}
    _etapas_prev     = {}   # { folio: etapa_jpcb }
    _historico_prev  = {}
    _gestion_prev    = {}   # { folio: { "CATEGORIA": "...", "NOTAS": "...", ... } }

    # Campos de gestión que se preservan como fallback desde el JSON anterior.
    # Si el Excel maestro / archivos de sucursal ya tienen el valor, ese tiene prioridad.
    # Este fallback solo aplica cuando el campo en el DataFrame está vacío.
    _CAMPOS_GESTION_FALLBACK = [
        "CATEGORIA", "OBSERVACION OT", "NOTAS", "AVANCE - GESTIÓN", "ULTIMA_EDICION"
    ]

    if os.path.exists(ruta_prev):
        try:
            with open(ruta_prev, "r", encoding="utf-8") as _f:
                _prev = json.load(_f)
            for _ot in _prev.get("ots", []):
                _folio = str(_ot.get("FOLIO OT", "")).strip()
                if not _folio:
                    continue
                _color = str(_ot.get("_MARCA_COLOR_", "")).strip()
                if _color:
                    _colores_prev[_folio] = _color
                _etapa = str(_ot.get("ETAPA_JPCB", "")).strip()
                if _etapa:
                    _etapas_prev[_folio] = _etapa
                _hist = _ot.get("repuestos_historico", [])
                if _hist:
                    _historico_prev[_folio] = _hist
                # Preservar campos de gestión (CATEGORIA, NOTAS, etc.) como fallback
                _vals_gestion = {}
                for _campo in _CAMPOS_GESTION_FALLBACK:
                    _v = str(_ot.get(_campo, "")).strip()
                    if _v:
                        _vals_gestion[_campo] = _v
                if _vals_gestion:
                    _gestion_prev[_folio] = _vals_gestion
            if _colores_prev:
                log(f"Marcas de color preservadas: {len(_colores_prev)} OT(s)")
            if _historico_prev:
                log(f"Histórico de repuestos cargado: {len(_historico_prev)} OT(s)")
            if _gestion_prev:
                log(f"Campos de gestión del JSON anterior cargados: {len(_gestion_prev)} OT(s)")
        except Exception as e:
            log(f"(i)  No se pudo leer JSON anterior para histórico: {e}")

    # Aplicar colores y etapas JPCB anteriores al DataFrame
    df = df.copy()
    if "_MARCA_COLOR_" not in df.columns:
        df["_MARCA_COLOR_"] = ""
    if _colores_prev:
        df["_MARCA_COLOR_"] = df[CLAVE].astype(str).map(
            lambda f: _colores_prev.get(f, "")
        )
    if "ETAPA_JPCB" not in df.columns:
        df["ETAPA_JPCB"] = ""
    if _etapas_prev:
        df["ETAPA_JPCB"] = df[CLAVE].astype(str).map(
            lambda f: _etapas_prev.get(f, "")
        )

    cols = [c for c in COLS_EXPORTAR if c in df.columns]
    df_exp = df[cols].copy()

    # Convertir todo a string para serialización JSON segura
    for col in df_exp.columns:
        df_exp[col] = df_exp[col].fillna("").astype(str).str.strip()
        df_exp[col] = df_exp[col].replace("nan", "").replace("None", "")

    registros = df_exp.to_dict(orient="records")

    # Fusionar repuestos actuales con histórico, y aplicar fallback de campos de gestión
    hoy = datetime.now().strftime("%d/%m/%Y")
    _recuperados_gestion = 0
    for rec in registros:
        folio    = str(rec.get("FOLIO OT", "")).strip()

        # ── RUT Cliente (cruce por patente) ───────────────────────────────────
        _patente_ot = str(rec.get("PATENTE", "") or "").strip().upper()
        rec["rut_cliente"] = mapa_patente_rut.get(_patente_ot, "") if _patente_ot else ""

        # ── Anticipo / Cuenta Ficha (cruce por RUT) ───────────────────────────
        # El campo rut_cliente puede tener varios RUTs separados por " / ".
        # Se busca cada uno en mapa_anticipo y se combinan si hay más de uno.
        rec["anticipo"] = {}
        if mapa_anticipo:
            _ruts_ot = [r.strip() for r in str(rec.get("rut_cliente", "") or "").split("/") if r.strip()]
            _ant_merged = None
            _ruts_con_saldo = []
            for _rut in _ruts_ot:
                _akey = _rut.upper()
                if _akey in mapa_anticipo:
                    _a = mapa_anticipo[_akey]
                    if _a["tiene_saldo"]:
                        _ruts_con_saldo.append(_rut)
                    if _ant_merged is None:
                        _ant_merged = {
                            "nombre":      _a["nombre"],
                            "total":       _a["total"],
                            "tiene_saldo": _a["tiene_saldo"],
                            "movimientos": list(_a["movimientos"]),
                        }
                    else:
                        _ant_merged["total"]       += _a["total"]
                        _ant_merged["tiene_saldo"]  = _ant_merged["tiene_saldo"] or _a["tiene_saldo"]
                        _ant_merged["movimientos"] += _a["movimientos"]
            if _ant_merged:
                _ant_merged["rut_saldo"] = _ruts_con_saldo   # RUT(s) específicos con saldo
                rec["anticipo"] = _ant_merged

        # ── Repuestos ──────────────────────────────────────────────────────────
        actuales = repuestos_actuales.get(folio, [])
        hist_ant = _historico_prev.get(folio, [])
        rec["repuestos_actual"]    = actuales
        rec["repuestos_historico"] = _merge_historico_repuestos(hist_ant, actuales, hoy)

        # ── Repuestos del Seguimiento de Compras (en espera / en bodega) ──────
        rec["repuestos_compras"]   = repuestos_compras.get(folio, [])

        # ── Filtrar los EN BODEGA que YA están en el Vale de Consumo ──────────
        # Si un repuesto en bodega ya figura en el Vale (mismo código de producto),
        # significa que ya fue consumido/asignado a la OT -> no es "pendiente".
        # Se conservan los en espera y los en bodega que NO están en el Vale.
        _compras_ot = rec["repuestos_compras"]
        if _compras_ot and actuales:
            _vale_cods = {_norm_cod_producto(v.get("producto", ""))
                          for v in actuales if v.get("producto")}
            if _vale_cods:
                rec["repuestos_compras"] = [
                    _rc for _rc in _compras_ot
                    if not (_rc.get("en_bodega")
                            and _norm_cod_producto(_rc.get("producto", "")) in _vale_cods)
                ]
                _compras_ot = rec["repuestos_compras"]

        # ── Filtrar los EN BODEGA sin stock disponible en el archivo de Stock ──
        # Si el repuesto llegó a bodega pero el stock es 0 (o no figura en el
        # archivo de stock), no tiene sentido mostrarlo como pendiente de instalar.
        # Se conservan los en espera (aún no llegan) y los en bodega con stock > 0.
        if _compras_ot:
            rec["repuestos_compras"] = [
                _rc for _rc in _compras_ot
                if not (
                    _rc.get("en_bodega")
                    and _rc.get("stock") is not None   # stock fue cruzado con el archivo
                    and (_rc.get("stock") or 0) <= 0   # stock = 0 → excluir
                )
            ]
            _compras_ot = rec["repuestos_compras"]

        # ── Filtrar EN BODEGA que aparecen en la tabla VC global ──────────────
        # Si el código del producto figura en "tabla vc.csv" significa que ya fue
        # consumido en alguna OT (aunque esa OT esté cerrada o fuera del radar de
        # la app) → no es un repuesto pendiente de instalar.
        # Se conservan: (a) los en espera, (b) los en bodega cuyo código NO aparece
        # en la tabla VC global.
        if codigos_vc_global and _compras_ot:
            rec["repuestos_compras"] = [
                _rc for _rc in _compras_ot
                if not (
                    _rc.get("en_bodega")
                    and _norm_cod_producto(_rc.get("producto", "")) in codigos_vc_global
                )
            ]

        # ── Fallback campos de gestión ─────────────────────────────────────────
        # Si el valor en el registro está vacío Y el JSON anterior tenía algo,
        # recuperamos ese valor. Prioridad: sucursales/maestro > JSON anterior.
        _prev_vals = _gestion_prev.get(folio, {})
        if _prev_vals:
            _recupero = False
            for _campo in _CAMPOS_GESTION_FALLBACK:
                _val_actual = str(rec.get(_campo, "")).strip()
                if not _val_actual and _campo in _prev_vals:
                    rec[_campo] = _prev_vals[_campo]
                    _recupero = True
            if _recupero:
                _recuperados_gestion += 1

        # ── Auto-completar OBSERVACION OT con repuesto de mayor costo ──────────
        # Solo si la OT tiene repuestos con descripción Y la observación sigue vacía.
        _obs = str(rec.get("OBSERVACION OT", "")).strip()
        if not _obs and actuales:
            _mejor = None
            _max_costo = -1
            for _r in actuales:
                _desc = str(_r.get("descripcion", "")).strip()
                if not _desc:
                    continue
                try:
                    _costo = float(str(_r.get("costo_total", "0")).replace(",", ".") or "0")
                except (ValueError, TypeError):
                    _costo = 0
                if _costo > _max_costo:
                    _max_costo = _costo
                    _mejor = _desc
            if _mejor:
                rec["OBSERVACION OT"] = _mejor

        # ── Auto-completar NOTAS con tipo de trabajo inferido ─────────────────
        # Solo si NOTAS está vacía Y hay repuestos con descripción.
        _notas = str(rec.get("NOTAS", "")).strip()
        if not _notas and actuales:
            _trabajo = inferir_trabajo(actuales)
            if _trabajo:
                rec["NOTAS"] = _trabajo

        # ── Auto-completar CATEGORIA con "DYP" ────────────────────────────────
        # Aplica cuando CATEGORIA está vacía y se detecta trabajo de carrocería,
        # ya sea por el TIPO VENTA o por los repuestos del Vale de Consumo.
        _cat = str(rec.get("CATEGORIA", "")).strip()
        if not _cat or _cat.lower() in ("sin categoría", "sin categoria"):
            _es_dyp = False

            # Señal 1 — TIPO VENTA contiene indicador de colisión / DYP
            _tipo_venta = _quitar_acentos(str(rec.get("TIPO VENTA", ""))).upper()
            _palabras_tipo_dyp = {
                "COLISION", "COLISIÓN", "DYP", "D.Y.P", "DAÑOS",
                "DANOS", "PINTURA", "CHAPA", "CARROCERIA", "CARROCERÍA",
                "ACCIDENTE", "SINIESTRO",
            }
            if any(p in _tipo_venta for p in _palabras_tipo_dyp):
                _es_dyp = True

            # Señal 2 — Repuestos del Vale de Consumo con descripción de carrocería
            if not _es_dyp and actuales:
                _kw_carroceria = {
                    "PARACHOQUE", "PARAGOLPE", "BUMPER",
                    "PUERTA", "PORTEZUELA",
                    "CAPO", "CAPOT",
                    "GUARDABARRO", "GUARDABARROS", "ALETA",
                    "TECHO", "TOLDO",
                    "VIDRIO", "PARABRISAS", "LUNETA", "CRISTAL",
                    "ESPEJO RETROVISOR", "RETROVISOR",
                    "BISAGRA PUERTA", "BISAGRA CAPO",
                    "MOLDURA", "FACIA",
                    "PANEL LATERAL", "PANEL FRONTAL", "PANEL TRASERO",
                    "PINTURA", "LACA", "MASILLA", "IMPRIMANTE",
                    "BASE PINTURA", "BARNIZ",
                    "TAPIZADO", "TAPIZ", "ALFOMBRA",
                    "SOPORTE PARAGOLPE", "SOPORTE PARACHOQUE",
                    "REJILLA FRONTAL", "REJILLA RADIADOR",
                    "PLUMILLA LIMPIAPARABRISAS", "PLUMILLA",
                    "EMBLEMA",
                }
                _texto_reps = " ".join(
                    _quitar_acentos(str(r.get("descripcion", "")))
                    for r in actuales
                ).upper()
                if any(kw in _texto_reps for kw in _kw_carroceria):
                    _es_dyp = True

            if _es_dyp:
                rec["CATEGORIA"] = "DYP"

    if _recuperados_gestion:
        log(f"Campos de gestión recuperados del JSON anterior: {_recuperados_gestion} OT(s)")

    datos = {
        "fecha_actualizacion": datetime.now().strftime("%d/%m/%Y %H:%M"),
        "total_ots":           len(registros),
        "ots":                 registros,
    }

    with open(RUTA_JSON, "w", encoding="utf-8") as f:
        json.dump(datos, f, ensure_ascii=False, indent=2)

    return RUTA_JSON


def _leer_json_github_blob(nombre_archivo):
    """
    Lee un JSON directo del blob del último commit en 'main' vía Git Data API.

    A diferencia de pedir raw.githubusercontent.com (o el download_url de la
    Contents API para archivos >1MB, que internamente es la misma URL), esto
    NUNCA pasa por el CDN público (Fastly) — ese CDN puede seguir sirviendo una
    copia cacheada de datos_dashboard.json varios minutos (a veces más) después
    de un commit nuevo. Es exactamente la causa raíz confirmada de que la
    consolidación revertía ediciones de CATEGORIA/OBSERVACION/NOTAS/AVANCE
    hechas desde la app web: obtener_ots_github_actuales() pedía la CDN, a
    veces recibía una copia vieja (sin la edición reciente), la mezclaba con
    el Excel maestro/sucursales, y subía esa versión vieja pisando la edición
    para siempre. Mismo patrón que _leer_json_github_blob() en app.py
    (sesión 07/07/2026, ya validado en producción para lectura de
    datos_dashboard.json) — replicado aquí para el consolidador.
    Devuelve el dict, o None si algo falla.
    """
    try:
        base_url = f"https://api.github.com/repos/{GITHUB_USUARIO}/{GITHUB_REPO}"
        hdrs = {"Authorization": f"token {GITHUB_TOKEN}", "Accept": "application/vnd.github.v3+json"}
        r = requests.get(f"{base_url}/git/ref/heads/main", headers=hdrs, timeout=15, verify=False)
        r.raise_for_status()
        commit_sha = r.json()["object"]["sha"]
        r = requests.get(f"{base_url}/git/trees/{commit_sha}", headers=hdrs,
                         params={"recursive": "1"}, timeout=15, verify=False)
        r.raise_for_status()
        blob_sha = None
        for item in r.json().get("tree", []):
            if item.get("path") == nombre_archivo:
                blob_sha = item.get("sha")
                break
        if not blob_sha:
            return None
        r = requests.get(f"{base_url}/git/blobs/{blob_sha}", headers=hdrs, timeout=30, verify=False)
        r.raise_for_status()
        raw = r.json().get("content", "").replace("\n", "")
        return json.loads(base64.b64decode(raw).decode("utf-8"))
    except Exception as _e:
        log(f"(i)  No se pudo leer via Git Data API: {_e}")
        return None


def obtener_ots_github_actuales():
    """
    Descarga las OTs del datos_dashboard.json actual en GitHub
    ANTES de que se suba la nueva versión.
    Devuelve un DataFrame con los folios actuales, o vacío si no hay datos.

    Usa _leer_json_github_blob() (Git Data API, sin CDN) en vez de
    raw.githubusercontent.com — ver docstring de esa función para el porqué:
    es la fuente autoritativa de las ediciones hechas desde la app web
    (CATEGORIA, OBSERVACION OT, NOTAS, AVANCE - GESTIÓN) y NO puede leerse
    de una copia potencialmente cacheada/vieja, o esas ediciones se pierden
    en cada consolidación.
    """
    if not all([GITHUB_USUARIO, GITHUB_REPO, GITHUB_TOKEN]):
        return pd.DataFrame()
    try:
        data = _leer_json_github_blob("datos_dashboard.json")
        if data:
            ots = data.get("ots", [])
            if ots:
                df = pd.DataFrame(ots)
                df[CLAVE] = normalizar_folio(df[CLAVE])
                return df
        log("(!)  Git Data API no devolvió datos_dashboard.json — "
            "reintentando via raw.githubusercontent.com como respaldo "
            "(puede traer una copia cacheada/desactualizada)")
        url = (f"https://raw.githubusercontent.com/"
               f"{GITHUB_USUARIO}/{GITHUB_REPO}/main/datos_dashboard.json")
        resp = requests.get(url, timeout=15, verify=False,
                            params={"_": datetime.now().timestamp()})
        if resp.status_code == 200:
            data = resp.json()
            ots  = data.get("ots", [])
            if ots:
                df = pd.DataFrame(ots)
                df[CLAVE] = normalizar_folio(df[CLAVE])
                return df
    except Exception as e:
        log(f"(i)  No se pudo obtener datos anteriores de GitHub: {e}")
    return pd.DataFrame()


def registrar_cierres_github(df_nuevo, df_anterior_github):
    """
    Compara el set actual de OTs con el anterior en GitHub.
    Registra en historial_cierres.json las OTs que ya no aparecen (cerradas)
    y las que son nuevas en esta actualización.
    """
    if not all([GITHUB_USUARIO, GITHUB_REPO, GITHUB_TOKEN]):
        return
    if df_anterior_github is None or df_anterior_github.empty:
        log("(i)  Sin datos anteriores en GitHub — omitiendo registro de cierres")
        return

    folios_nuevos    = set(df_nuevo[CLAVE].astype(str).str.strip())
    folios_anteriores = set(df_anterior_github[CLAVE].astype(str).str.strip())

    folios_cerrados = folios_anteriores - folios_nuevos
    folios_entraron = folios_nuevos - folios_anteriores

    # Detalle de las OTs cerradas (tomado del DataFrame anterior)
    cols_registrar = [c for c in [CLAVE, "SUCURSAL", "RANGO", "ASESOR",
                                   "DIAS APERTURA", "TIPO VENTA", "CATEGORIA"]
                      if c in df_anterior_github.columns]
    ots_cerradas = []
    for folio in folios_cerrados:
        filt = df_anterior_github[df_anterior_github[CLAVE].astype(str) == folio]
        if not filt.empty:
            row = filt.iloc[0]
            ots_cerradas.append({c: str(row.get(c, "")).strip() for c in cols_registrar})

    # Cargar historial existente en GitHub
    url_hist = (f"https://api.github.com/repos/{GITHUB_USUARIO}/"
                f"{GITHUB_REPO}/contents/{GITHUB_HISTORIAL}")
    headers  = {
        "Authorization": f"token {GITHUB_TOKEN}",
        "Accept": "application/vnd.github.v3+json",
    }
    sha      = None
    historial = {"registros": []}

    try:
        resp = requests.get(url_hist, headers=headers, timeout=15, verify=False)
        if resp.status_code == 200:
            info     = resp.json()
            sha      = info["sha"]
            historial = json.loads(base64.b64decode(info["content"]).decode("utf-8"))
    except Exception:
        pass

    # Agregar registro de esta actualización
    historial["registros"].append({
        "fecha":          datetime.now().strftime("%d/%m/%Y %H:%M"),
        "total_cerradas": len(folios_cerrados),
        "total_nuevas":   len(folios_entraron),
        "total_activas":  len(folios_nuevos),
        "ots_cerradas":   ots_cerradas,
    })
    # Mantener solo los últimos 90 registros para no inflar el archivo
    historial["registros"] = historial["registros"][-90:]

    nuevo_b64 = base64.b64encode(
        json.dumps(historial, ensure_ascii=False, indent=2).encode("utf-8")
    ).decode()
    payload = {
        "message": f"Historial cierres {datetime.now().strftime('%d/%m/%Y %H:%M')}",
        "content": nuevo_b64,
    }
    if sha:
        payload["sha"] = sha

    try:
        resp_put = requests.put(url_hist, headers=headers,
                                json=payload, timeout=30, verify=False)
        if resp_put.status_code in [200, 201]:
            log(f"Historial de cierres actualizado: "
                f"{len(folios_cerrados)} cerradas / {len(folios_entraron)} nuevas")
        else:
            log(f"(!)  Error al actualizar historial de cierres: {resp_put.status_code}")
    except Exception as e:
        log(f"(!)  No se pudo guardar historial de cierres: {e}")


def subir_a_github(ruta_json):
    """Sube datos_dashboard.json al repositorio GitHub via API REST."""
    if not all([GITHUB_USUARIO, GITHUB_REPO, GITHUB_TOKEN]):
        log("(!)  GitHub no configurado — dashboard web no actualizado")
        log("     Rellena GITHUB_USUARIO, GITHUB_REPO y GITHUB_TOKEN en el script.")
        return

    try:
        archivo = "datos_dashboard.json"
        url = f"https://api.github.com/repos/{GITHUB_USUARIO}/{GITHUB_REPO}/contents/{archivo}"

        with open(ruta_json, "rb") as f:
            contenido_b64 = base64.b64encode(f.read()).decode()

        headers = {
            "Authorization": f"token {GITHUB_TOKEN}",
            "Accept": "application/vnd.github.v3+json",
        }

        # Obtener SHA actual del archivo (necesario para actualizar)
        sha = None
        try:
            resp_get = requests.get(url, headers=headers, timeout=15, verify=False)
            if resp_get.status_code == 200:
                sha = resp_get.json()["sha"]
        except Exception:
            pass  # El archivo no existe aún (primera vez)

        payload = {
            "message": f"OTs actualizadas {datetime.now().strftime('%d/%m/%Y %H:%M')}",
            "content": contenido_b64,
        }
        if sha:
            payload["sha"] = sha

        resp_put = requests.put(url, headers=headers, json=payload, timeout=30, verify=False)

        if resp_put.status_code in [200, 201]:
            log("Dashboard web actualizado correctamente en GitHub")
        else:
            log(f"(!)  GitHub respondio con estado {resp_put.status_code}: {resp_put.text[:200]}")

    except Exception as e:
        log(f"(!)  No se pudo subir a GitHub: {e}")


# =============================================================
#   RANKING DE OTs CERRADAS CON MÁS DE 90 DÍAS
# =============================================================

def generar_ranking_cierres(ruta_pbi):
    """
    Lee el PBI completo (TODAS las OTs, incluyendo cerradas) y genera un ranking
    de asesores y sucursales que han cerrado más OTs con más de 90 días de apertura,
    desde RANKING_DESDE hasta hoy.

    La fecha de cierre se determina como la más temprana entre las columnas de fecha
    de documento disponibles en el PBI (FECHA LIQUIDACIÓN ST, FECHA FACTURA CLIENTE, etc.).
    Los días de apertura se calculan como (fecha_cierre - fecha_apertura).

    Retorna un dict listo para serializar a JSON, o None si no hay datos suficientes.
    """
    log("Leyendo PBI completo para ranking de cierres...")

    # ── 1. Leer PBI sin filtro de estado ──────────────────────────────────────
    try:
        df_full = pd.read_excel(ruta_pbi, sheet_name=HOJA_PBI, header=HEADER_PBI,
                                engine="calamine", dtype=str)
    except Exception as e:
        log(f"(!)  No se pudo leer PBI para ranking: {e}")
        return None

    # Normalizar columnas: mayúsculas, sin acentos, sin columnas sin nombre
    df_full.columns = [
        _quitar_acentos(c.upper().strip()) if isinstance(c, str) else str(c)
        for c in df_full.columns
    ]
    df_full = df_full.loc[:, ~df_full.columns.str.startswith("UNNAMED")]

    # Normalizar FOLIO OT y eliminar filas inválidas
    clave_norm = _quitar_acentos(CLAVE)
    if clave_norm not in df_full.columns:
        log(f"(!)  Columna '{CLAVE}' no encontrada en PBI para ranking")
        return None
    df_full[clave_norm] = normalizar_folio(df_full[clave_norm])
    df_full = df_full[
        df_full[clave_norm].notna() &
        (df_full[clave_norm] != "") &
        (df_full[clave_norm] != "NAN")
    ].copy()

    # ── 2. Filtrar solo OTs CERRADAS (excluir PENDIENTE y ANULADO) ─────────────
    estado_col = _quitar_acentos("ESTADO")
    if estado_col in df_full.columns:
        mask_cerradas = ~df_full[estado_col].astype(str).str.strip().str.upper().isin(
            {_quitar_acentos(e) for e in ESTADOS_PENDIENTES}
        )
        df_cerradas = df_full[mask_cerradas].copy()
    else:
        log("(i)  Columna ESTADO no encontrada — se usan todas las filas para ranking")
        df_cerradas = df_full.copy()

    # Deduplicar a 1 fila por OT
    df_cerradas = df_cerradas.drop_duplicates(subset=[clave_norm], keep="first").reset_index(drop=True)
    log(f"OTs cerradas en PBI: {len(df_cerradas)}")

    # ── 3. Fecha de apertura ───────────────────────────────────────────────────
    apertura_candidatas = ["FECHA OT", "FECHA APERTURA", "FECHA"]
    fecha_apertura_col  = next(
        (c for c in [_quitar_acentos(x) for x in apertura_candidatas]
         if c in df_cerradas.columns),
        None
    )
    if not fecha_apertura_col:
        log("(!)  No se encontró columna de fecha de apertura — ranking no generado")
        return None

    df_cerradas["_FECH_APER"] = pd.to_datetime(
        df_cerradas[fecha_apertura_col], dayfirst=True, errors="coerce"
    )

    # ── 4. Fecha de cierre: mínimo entre todos los documentos disponibles ──────
    cols_fecha_cierre_pbi = [
        c for c in df_cerradas.columns
        if any(_quitar_acentos(patron) in c for patron in FECHAS_CIERRE_BUSCAR)
        and c != fecha_apertura_col
    ]

    # Fallback: cualquier columna que contenga "FECHA" y no sea la de apertura
    if not cols_fecha_cierre_pbi:
        cols_fecha_cierre_pbi = [
            c for c in df_cerradas.columns
            if "FECHA" in c and c != fecha_apertura_col
        ]

    if not cols_fecha_cierre_pbi:
        log("(!)  No se encontraron columnas de fecha de cierre — ranking no generado")
        return None

    log(f"Columnas de fecha de cierre detectadas: {cols_fecha_cierre_pbi}")

    # Convertir cada columna de fecha de cierre a datetime y tomar el mínimo
    _dt_cols = []
    for col in cols_fecha_cierre_pbi:
        dcol = f"_dt_{col}"
        df_cerradas[dcol] = pd.to_datetime(df_cerradas[col], dayfirst=True, errors="coerce")
        _dt_cols.append(dcol)

    df_cerradas["_FECH_CIERR"] = df_cerradas[_dt_cols].min(axis=1)
    df_cerradas.drop(columns=_dt_cols, inplace=True)

    # ── 5. Filtrar por período y calcular días ─────────────────────────────────
    fecha_inicio = pd.Timestamp(RANKING_DESDE)
    df_periodo   = df_cerradas[
        df_cerradas["_FECH_CIERR"].notna() &
        (df_cerradas["_FECH_CIERR"] >= fecha_inicio)
    ].copy()

    if df_periodo.empty:
        log(f"(i)  No hay OTs cerradas desde {RANKING_DESDE} — ranking vacío")
        return {
            "fecha_generacion": datetime.now().strftime("%d/%m/%Y %H:%M"),
            "periodo_desde":    RANKING_DESDE,
            "total_ots_90mas":  0,
            "por_asesor": [], "por_sucursal": [],
            "por_asesor_sucursal": [], "top_ots": [], "por_mes": [],
        }

    df_periodo["_DIAS_CIERR"] = (
        df_periodo["_FECH_CIERR"] - df_periodo["_FECH_APER"]
    ).dt.days

    # Solo OTs con más de 90 días al momento de cerrarse
    df_90 = df_periodo[df_periodo["_DIAS_CIERR"] > 90].copy()
    log(f"OTs cerradas con >90 días ({RANKING_DESDE} → hoy): {len(df_90)}")

    if df_90.empty:
        return {
            "fecha_generacion": datetime.now().strftime("%d/%m/%Y %H:%M"),
            "periodo_desde":    RANKING_DESDE,
            "total_ots_90mas":  0,
            "por_asesor": [], "por_sucursal": [],
            "por_asesor_sucursal": [], "top_ots": [], "por_mes": [],
        }

    # ── 6. Mes de cierre ──────────────────────────────────────────────────────
    df_90["_MES"] = df_90["_FECH_CIERR"].dt.strftime("%Y-%m")

    def _grp(df, cols_agr):
        """Agrupa y retorna lista de dicts."""
        g = df.groupby(cols_agr).agg(
            total        =(clave_norm, "count"),
            dias_promedio=("_DIAS_CIERR", "mean"),
            dias_max     =("_DIAS_CIERR", "max"),
        ).reset_index().sort_values("total", ascending=False)
        g["dias_promedio"] = g["dias_promedio"].round(1)
        g["dias_max"]      = g["dias_max"].astype(int)
        g["total"]         = g["total"].astype(int)
        return g.to_dict(orient="records")

    # Columnas clave normalizadas
    asesor_col   = _quitar_acentos("ASESOR")
    sucursal_col = _quitar_acentos("SUCURSAL")
    tipo_col     = _quitar_acentos("TIPO VENTA")
    marca_col    = _quitar_acentos("MARCA")

    por_asesor            = _grp(df_90, [asesor_col])   if asesor_col   in df_90.columns else []
    por_sucursal          = _grp(df_90, [sucursal_col]) if sucursal_col in df_90.columns else []
    por_asesor_sucursal   = (_grp(df_90, [asesor_col, sucursal_col])
                             if asesor_col in df_90.columns and sucursal_col in df_90.columns else [])

    # Evolución mensual
    por_mes = []
    if "_MES" in df_90.columns:
        gm = df_90.groupby("_MES").agg(total=(clave_norm, "count")).reset_index()
        gm = gm.sort_values("_MES")
        gm.columns = ["mes", "total"]
        por_mes = gm.to_dict(orient="records")

    # Top 50 OTs más antiguas al cerrar
    cols_top = [c for c in [
        clave_norm, sucursal_col, asesor_col, "_DIAS_CIERR",
        "_FECH_APER", "_FECH_CIERR", tipo_col, marca_col
    ] if c in df_90.columns]
    top_df = df_90.nlargest(50, "_DIAS_CIERR")[cols_top].copy()
    for dc in ["_FECH_APER", "_FECH_CIERR"]:
        if dc in top_df.columns:
            top_df[dc] = top_df[dc].dt.strftime("%d/%m/%Y")
    top_df = top_df.rename(columns={
        clave_norm:   "folio_ot",
        sucursal_col: "sucursal",
        asesor_col:   "asesor",
        "_DIAS_CIERR":"dias_al_cierre",
        "_FECH_APER": "fecha_apertura",
        "_FECH_CIERR":"fecha_cierre",
        tipo_col:     "tipo_venta",
        marca_col:    "marca",
    })
    top_ots = top_df.fillna("").astype(str).to_dict(orient="records")

    resultado = {
        "fecha_generacion":    datetime.now().strftime("%d/%m/%Y %H:%M"),
        "periodo_desde":       RANKING_DESDE,
        "total_ots_90mas":     int(len(df_90)),
        "por_asesor":          por_asesor,
        "por_sucursal":        por_sucursal,
        "por_asesor_sucursal": por_asesor_sucursal,
        "top_ots":             top_ots,
        "por_mes":             por_mes,
    }
    return resultado



def subir_ranking_github(datos):
    """Sube ranking_cierres.json al repositorio GitHub."""
    if not all([GITHUB_USUARIO, GITHUB_REPO, GITHUB_TOKEN]):
        log("(i)  GitHub no configurado — ranking no subido")
        return

    # Guardar copia local primero
    try:
        with open(RUTA_RANKING, "w", encoding="utf-8") as f:
            json.dump(datos, f, ensure_ascii=False, indent=2)
        log(f"Ranking guardado localmente -> {os.path.basename(RUTA_RANKING)}")
    except Exception as e:
        log(f"(!)  No se pudo guardar ranking local: {e}")

    try:
        url = (f"https://api.github.com/repos/{GITHUB_USUARIO}/"
               f"{GITHUB_REPO}/contents/{GITHUB_RANKING}")
        headers = {
            "Authorization": f"token {GITHUB_TOKEN}",
            "Accept": "application/vnd.github.v3+json",
        }

        with open(RUTA_RANKING, "rb") as f:
            contenido_b64 = base64.b64encode(f.read()).decode()

        sha = None
        try:
            resp_get = requests.get(url, headers=headers, timeout=15, verify=False)
            if resp_get.status_code == 200:
                sha = resp_get.json()["sha"]
        except Exception:
            pass

        payload = {
            "message": f"Ranking cierres {datetime.now().strftime('%d/%m/%Y %H:%M')}",
            "content": contenido_b64,
        }
        if sha:
            payload["sha"] = sha

        resp_put = requests.put(url, headers=headers, json=payload, timeout=30, verify=False)

        if resp_put.status_code in [200, 201]:
            log("Ranking de cierres actualizado en GitHub")
        else:
            log(f"(!)  GitHub respondio con estado {resp_put.status_code}: {resp_put.text[:200]}")

    except Exception as e:
        log(f"(!)  No se pudo subir ranking a GitHub: {e}")


# =============================================================
#   AGENDA CURIFOR — Descarga de citas (hoy + 4 dias) + detalle por cita
# =============================================================
#
#   La agenda Curifor (agenda.curifor.cl) tiene reservas por sucursal.
#   Credenciales genericas: ADM_CUR / ADM_CUR
#
#   Para cada sucursal necesitas configurar su id_cons en SUCURSALES_AGENDA.
#   Si no conoces el id_cons de alguna sucursal, dejala sin mapear (None).
#
#   Como obtener el id_cons:
#   1. Inicia sesion en agenda.curifor.cl
#   2. Abre DevTools (F12) -> Network
#   3. Ve a la vista de reservas de cada sucursal
#   4. Busca la llamada a getReporteReservas.jsp y anota el valor de id_cons
# =============================================================

# Mapeo id_cons de la agenda Curifor -> nombre de sucursal que usa el PBI (asi el
# Planificador cruza directo con el selector de sucursal del dashboard). Es una LISTA
# de pares (nombre, id_cons), no un dict, porque dos id_cons distintos pueden
# corresponder a la MISMA sucursal fisica del PBI — en ese caso se repite el mismo
# nombre y sus citas se fusionan automaticamente en agenda_hoy.json.
#
# Caso real (13/07/2026, confirmado por Cristobal): en la Agenda Curifor, "Talca" (Ford)
# y "Talca BMW" son la misma sucursal fisica de Talca (autos Ford y BMW comparten
# taller) — se fusionan bajo el nombre "TALCA". "Talca Camiones" es una sucursal
# fisica distinta, que en el PBI aparece como "TALCA (2)" (no como "TALCA CAMIONES") —
# se renombro para que coincida exacto y deje de aparecer vacia en el Planificador.
SUCURSALES_AGENDA = [
    ("CHILLAN",       "497"),
    ("CHILLAN VIEJO", "524"),  # En la Agenda figura como "Chillán Viejo", pero el PBI (y por
                               # lo tanto el selector del Planificador) la llama "CHILLAN VIEJO".
                               # Con el nombre acentuado se creaba una SEGUNDA sucursal vacia en
                               # control_taller.json, conviviendo con la real: el equipo podia
                               # terminar cargando trabajo en la que no era. Se usa el nombre del
                               # PBI, mismo criterio que CD REPUESTOS y TALCA (2). (05/08/2026)
    ("CURICO",        "496"),
    ("LO BLANCO",     "578"),
    ("LINDEROS",      "495"),
    ("RANCAGUA",      "525"),
    ("TALCA",         "536"),  # Talca (Ford)
    ("TALCA",         "588"),  # Talca BMW — misma sucursal fisica que la anterior, se fusiona
    ("PLACILLA",      "493"),
    ("CD REPUESTOS",  "589"),  # "Taller Movil" en la Agenda es la misma operacion que
                                 # "CD REPUESTOS" en el PBI — se usa el nombre del PBI para
                                 # que coincida directo con el selector de sucursal del Planificador.
    # Comercios adicionales detectados en el metodo de extraccion del equipo (07/2026).
    # OJO: si el PBI usa otro nombre para estas sucursales, renombrar para que coincida
    # (igual que se hizo con CD REPUESTOS/TALCA (2)), o no cruzaran en el Planificador.
    ("MACUL",         "594"),  # "Curifor Macul (auto-park)" en la Agenda
    ("TALCA (2)",     "590"),  # antes "TALCA CAMIONES" — el PBI identifica esta sucursal como "TALCA (2)"
]

AGENDA_LOGIN_URL   = "https://agenda.curifor.cl/config.jsp"
AGENDA_REPORT_URL  = "https://agenda.curifor.cl/ajax/getReporteReservas.jsp"
AGENDA_DETALLE_URL = "https://agenda.curifor.cl/motorAgenda2.jsp"
AGENDA_USER        = "ADM_CUR"
AGENDA_PASS        = "ADM_CUR"

# Dias de agenda a descargar: hoy + 4 hacia adelante (ventana del metodo del equipo,
# antes era hoy + 2). Mas horizonte para pre-picking de repuestos.
AGENDA_DIAS_ADELANTE = 4

# Cache local del detalle de citas (motorAgenda2.jsp): 1 llamada extra por cita, asi que
# se guarda por OC y solo se vuelve a pedir si la entrada no tiene asesor todavia.
RUTA_CACHE_DETALLE = os.path.join(
    os.path.dirname(os.path.abspath(__file__)), "agenda_detalle_cache.json"
)


def _login_agenda():
    """Inicia sesion en agenda.curifor.cl. Retorna la sesion requests o None."""
    import requests as _req
    ses = _req.Session()
    try:
        resp = ses.post(
            AGENDA_LOGIN_URL,
            data={"usuario": AGENDA_USER, "clave": AGENDA_PASS, "acc": "LOGIN"},
            timeout=15,
            verify=False,
        )
        if resp.ok:
            return ses
    except Exception as _e:
        log(f"(!) Error al iniciar sesion en agenda: {_e}")
    return None


def _es_login_agenda(html_text):
    """True si la respuesta es el formulario de login (sesion caducada)."""
    t = html_text or ""
    return 'name="clave"' in t or "name='clave'" in t


def _relogin_agenda(ses):
    """Re-autentica la MISMA sesion (refresca cookies in-place)."""
    try:
        ses.post(
            AGENDA_LOGIN_URL,
            data={"usuario": AGENDA_USER, "clave": AGENDA_PASS, "acc": "LOGIN"},
            timeout=20, verify=False,
        )
    except Exception:
        pass


def _http_get_agenda(ses, url, params=None, timeout=40, intentos=4):
    """
    GET a la agenda con reintentos + backoff + re-login automatico si la sesion
    caduco a mitad de corrida (patron tomado del metodo de extraccion del equipo,
    generar_csv_agenda.py). Retorna el HTML o None si fallo tras los reintentos.
    """
    import time as _time
    ultimo = None
    for _i in range(intentos):
        try:
            r = ses.get(url, params=params, timeout=timeout, verify=False)
            if r.ok and _es_login_agenda(r.text):
                _relogin_agenda(ses)
                r = ses.get(url, params=params, timeout=timeout, verify=False)
            if r.ok and not _es_login_agenda(r.text):
                return r.text
            ultimo = f"HTTP {r.status_code}" if not r.ok else "sesion invalida"
        except Exception as _e:
            ultimo = _e
        _time.sleep(2 * (_i + 1))
    log(f"  (!) GET agenda fallo tras {intentos} intentos: {ultimo}")
    return None


def _inp_agenda(h, fid):
    """Extrae el value de un input por id (formato id='...' value='...' de motorAgenda2.jsp)."""
    m = re.search(r"id='%s'[^>]*?value\s*=\s*'([^']*)'" % re.escape(fid), h or "", re.S)
    return re.sub(r"\s+", " ", m.group(1)).strip() if m else ""


def _sel_agenda(h, sid):
    """Extrae el texto de la opcion seleccionada de un select por id."""
    m = re.search(r"id='%s'.*?</select>" % re.escape(sid), h or "", re.S)
    if not m:
        return ""
    o = re.search(r"<option[^>]*\bselected\b[^>]*>(.*?)</option>", m.group(0), re.S)
    return re.sub(r"\s+", " ", re.sub(r"<[^>]+>", "", o.group(1))).strip() if o else ""


def _detalle_cita_web(ses, oc):
    """
    Descarga el detalle de una cita (motorAgenda2.jsp?id_agenda=OC&EDIT=1) y extrae
    asesor/tecnico, RUT, telefonos, mail, VIN, tipo de cliente y patente.
    Retorna dict o None si fallo la descarga.
    """
    h = _http_get_agenda(ses, AGENDA_DETALLE_URL, params={"id_agenda": oc, "EDIT": "1"})
    if h is None:
        return None
    rut = _inp_agenda(h, "rut")
    dv  = _inp_agenda(h, "dv")
    tp  = _inp_agenda(h, "hTipPersona")
    return {
        "asesor":       _sel_agenda(h, "id_tecnico"),
        "rut":          f"{rut}-{dv}" if rut else "",
        "celular":      _inp_agenda(h, "fono_celular"),
        "fono":         _inp_agenda(h, "fono_fijo"),
        "mail":         _inp_agenda(h, "email"),
        "vin":          _inp_agenda(h, "vin_agenda"),
        "patente":      _inp_agenda(h, "patente"),
        "tipo_cliente": "Juridica" if tp == "J" else ("Natural" if tp == "N" else ""),
        "guardado":     datetime.now().strftime("%Y-%m-%d"),
    }


def _cargar_cache_detalle():
    """Carga el cache local de detalles por OC. Poda entradas de mas de 90 dias."""
    try:
        if os.path.exists(RUTA_CACHE_DETALLE):
            with open(RUTA_CACHE_DETALLE, encoding="utf-8") as f:
                cache = json.load(f)
            from datetime import timedelta as _td
            limite = (datetime.now() - _td(days=90)).strftime("%Y-%m-%d")
            return {k: v for k, v in cache.items()
                    if isinstance(v, dict) and v.get("guardado", "9999") >= limite}
    except Exception as _e:
        log(f"(!) Cache de detalle ilegible ({_e}) — se parte de cero")
    return {}


def _guardar_cache_detalle(cache):
    try:
        with open(RUTA_CACHE_DETALLE, "w", encoding="utf-8") as f:
            json.dump(cache, f, ensure_ascii=False)
    except Exception as _e:
        log(f"(!) No se pudo guardar el cache de detalle: {_e}")


def _enriquecer_citas_con_detalle(ses, resultado, hilos=3):
    """
    Enriquece las citas de agenda_hoy.json con el detalle de motorAgenda2.jsp:
    asesor (vacio hasta ahora — lo usa Control de Taller), RUT, telefonos, mail,
    VIN y tipo de cliente. Usa cache por OC: solo va a la web por citas nuevas o
    cuyo detalle cacheado aun no tiene asesor asignado. Concurrencia 3 con sesion
    compartida (validado por el equipo en generar_csv_agenda.py).
    """
    cache = _cargar_cache_detalle()
    todas = [c
             for suc in (resultado.get("sucursales") or {}).values()
             for citas in suc.values()
             for c in citas]

    pendientes, vistos = [], set()
    for c in todas:
        oc = (c.get("oc") or "").strip()
        if not oc or oc in vistos:
            continue
        ent = cache.get(oc)
        if ent and (ent.get("asesor") or "").strip():
            continue  # detalle completo en cache
        vistos.add(oc)
        pendientes.append(oc)

    if pendientes:
        log(f"Detalle de citas: {len(pendientes)} por descargar ({len(cache)} en cache)...")
        from concurrent.futures import ThreadPoolExecutor
        with ThreadPoolExecutor(max_workers=hilos) as ex:
            for oc, det in zip(pendientes, ex.map(lambda o: _detalle_cita_web(ses, o), pendientes)):
                if det is not None:
                    cache[oc] = det
        _guardar_cache_detalle(cache)

    n_enr = 0
    for c in todas:
        det = cache.get((c.get("oc") or "").strip())
        if not det:
            continue
        for campo in ("asesor", "rut", "celular", "fono", "mail", "vin", "tipo_cliente"):
            if det.get(campo) and not (c.get(campo) or "").strip():
                c[campo] = det[campo]
        if det.get("patente") and not (c.get("patente") or "").strip():
            c["patente"] = det["patente"].upper()
        n_enr += 1
    if n_enr:
        log(f"Citas enriquecidas con detalle: {n_enr}")
    return resultado


# =============================================================
#   TEMPARIO DE MANO DE OBRA (mantenciones) — Cotizador Curifor
#   ------------------------------------------------------------
#   tempario.json vive junto a este script (se sube a GitHub con
#   Subir_App_GitHub.bat) y trae, por marca/modelo, las horas de
#   mano de obra oficiales por kilometraje de mantencion (8 marcas:
#   BAIC, Ford, JAC, Jaecoo, JIM, Mahindra, Omoda, Shineray, SWM).
#   Se usa para estimar cuanto va a demorar el servicio de cada
#   cita de Agenda que trae "modelo" + "mantencion" (km), y esa
#   estimacion viaja dentro de agenda_hoy.json como "horas_tempario"
#   para que el Planificador de Taller la use como duracion por
#   defecto del bloque en Tecnico x Hora (editable si hay atrasos).
# =============================================================
RUTA_TEMPARIO = os.path.join(os.path.dirname(os.path.abspath(__file__)), "tempario.json")
RUTA_PAUTA_REPUESTOS = os.path.join(os.path.dirname(os.path.abspath(__file__)), "pauta_repuestos.json")

_TEMPARIO_CACHE = None
_PAUTA_REP_CACHE = None

# Mapa de stock real { cod_normalizado: {"bodega":..,"stock":..,"costo":..} } —
# se setea una sola vez desde main() (ya se calcula ahi para otros cruces) y lo
# usa _buscar_repuestos_pauta() para enriquecer el listado de repuestos sugeridos
# del Pre-picking con Stock/Ubicacion/Costo reales, sin tener que releer el Excel.
_STOCK_IDX_CACHE = None

# Indice secundario por "sufijo" (codigo real sin el prefijo numerico de familia
# que trae la columna Producto del catalogo de Stock, ej. "25 JL3Z6731A" -> clave
# "JL3Z6731A") — se arma una sola vez, la primera vez que hace falta. La pauta del
# Cotizador trae el codigo "pelado" (sin ese prefijo), asi que el cruce exacto
# contra _STOCK_IDX_CACHE solo pegaba ~7% de las veces; con el sufijo sube a ~73%
# (verificado 13/07/2026 contra el catalogo real, 209/285 codigos de la pauta).
_STOCK_SUFFIX_CACHE = None
_RE_PREFIJO_FAMILIA = re.compile(r"^\d{1,3}\s+(.+)$")


def _set_stock_idx_cache(mapa_stock):
    global _STOCK_IDX_CACHE, _STOCK_SUFFIX_CACHE
    _STOCK_IDX_CACHE = mapa_stock or {}
    _STOCK_SUFFIX_CACHE = None  # se reconstruye perezosamente en _stock_lookup()


# Catalogo detallado por bodega (no aplanado) + clasificacion (Familia/SubFamilia) —
# lo arma leer_stock_repuestos() como efecto secundario. Se usa para: (a) filtrar el
# Stock/Ubicacion del Pre-picking a SOLO la sucursal de la cita (en vez de sumar todas
# las bodegas del pais), y (b) buscar repuestos "compatibles" — mismo producto en otro
# formato/envase (misma Familia+SubFamilia, descripcion parecida). Sesion 13/07/2026.
_STOCK_DETALLE_CACHE = None
_STOCK_DETALLE_SUFFIX_CACHE = None  # { sufijo: cod_norm_completo } — para el mismo
# problema de prefijo numerico de familia que _STOCK_SUFFIX_CACHE, pero apuntando al
# codigo completo (no al bodega/stock ya aplanado) para poder buscar en _STOCK_DETALLE_CACHE.


def _set_stock_detalle_cache(detalle):
    global _STOCK_DETALLE_CACHE, _STOCK_DETALLE_SUFFIX_CACHE
    _STOCK_DETALLE_CACHE = detalle or {}
    _STOCK_DETALLE_SUFFIX_CACHE = None


def _detalle_lookup_cod_completo(cod_norm):
    """Devuelve el codigo COMPLETO (con prefijo de familia) del catalogo que matchea
    cod_norm, exacto o por sufijo — para poder indexar _STOCK_DETALLE_CACHE."""
    global _STOCK_DETALLE_SUFFIX_CACHE
    if not cod_norm or not _STOCK_DETALLE_CACHE:
        return None
    if cod_norm in _STOCK_DETALLE_CACHE:
        return cod_norm
    if _STOCK_DETALLE_SUFFIX_CACHE is None:
        idx = {}
        for cod_completo in _STOCK_DETALLE_CACHE:
            m = _RE_PREFIJO_FAMILIA.match(cod_completo)
            sufijo = m.group(1) if m else cod_completo
            idx.setdefault(sufijo, cod_completo)
        _STOCK_DETALLE_SUFFIX_CACHE = idx
    return _STOCK_DETALLE_SUFFIX_CACHE.get(cod_norm)


def _bodega_pertenece_sucursal(bodega, sucursal):
    """
    True si una bodega del catalogo de Stock corresponde a la sucursal indicada.
    Match exacto, o el nombre de la sucursal seguido solo de espacios/digitos (para
    tolerar bodegas secundarias tipo "CHILLAN2"/"TALCA (2)") — pero NUNCA si sigue
    otra palabra (asi "CHILLAN VIEJO" no matchea contra la sucursal "CHILLAN").
    Bodegas centrales/internas (BODEGA ML, CD REPUESTOS, PE X REGULARIZAR, etc.) no
    matchean ninguna sucursal — quedan como "otro lugar" en el fallback.
    """
    b = _norm_temp(bodega)
    s = _norm_temp(sucursal)
    if not b or not s:
        return False
    if b == s:
        return True
    if b.startswith(s):
        resto = b[len(s):]
        if re.fullmatch(r"\s*\d*\s*", resto):
            return True
    return False


def _stock_para_sucursal(cod_norm, sucursal):
    """
    Devuelve (stock_en_sucursal:float|None, otras_bodegas:[(bodega,stock),...]) para
    un codigo (exacto o por sufijo de familia) del catalogo detallado de Stock.
    stock_en_sucursal es None si el codigo no esta en el catalogo o si ninguna bodega
    de esa sucursal lo tiene (no es lo mismo que 0: None = "no hay dato de esa
    sucursal en el catalogo", listado en otras_bodegas si existe en otro lado).
    """
    if not _STOCK_DETALLE_CACHE or not cod_norm:
        return None, []
    cod_completo = _detalle_lookup_cod_completo(cod_norm)
    det = _STOCK_DETALLE_CACHE.get(cod_completo) if cod_completo else None
    if not det:
        return None, []
    stock_suc = 0.0
    encontrado = False
    otras = []
    for bod, info in det["por_bodega"].items():
        if _bodega_pertenece_sucursal(bod, sucursal):
            stock_suc += info["stock"]
            encontrado = True
        elif info["stock"] > 0:
            otras.append((bod, info["stock"]))
    return (stock_suc if encontrado else None), otras


def _codigos_relacionados(cod_base, cod_cand, min_prefijo=6, min_ratio=0.6):
    """
    Determina si dos codigos de repuesto son variantes del mismo producto —
    UNICO criterio confiable (a pedido explicito de Cristobal, 14/07/2026): los
    codigos deben "conversar entre si", compartiendo un prefijo sustancial y
    variando solo en las letras/numeros finales (ej. 'XO5W30Q1SP' vs
    'XO5W30Q1SP-PLZ', o 'XO5W30Q1SP' vs 'XO5W30Q1FS'). Categoria (Familia/
    SubFamilia) y descripcion NO se usan — se comprobo que producen falsos
    positivos graves (ej. sugerir un aceite de TRANSMISION como alternativa de
    uno de MOTOR solo porque ambos son "lubricantes", o un filtro de aceite de
    otro vehiculo solo porque dice "filtro de aceite").
    Se quita prefijo de familia numerico + espacios/guiones antes de comparar.
    """
    def _limpiar(c):
        c = (c or "").upper().strip()
        m = _RE_PREFIJO_FAMILIA.match(c)
        if m:
            c = m.group(1)
        return re.sub(r"[\s\-]", "", c)
    a = _limpiar(cod_base)
    b = _limpiar(cod_cand)
    if not a or not b or a == b:
        return False
    n = 0
    for ca, cb in zip(a, b):
        if ca != cb:
            break
        n += 1
    if n < min_prefijo:
        return False
    # el prefijo compartido debe ser una porcion grande de AMBOS codigos, no
    # solo de uno largo con un codigo corto que por casualidad calza al inicio
    if n / len(a) < min_ratio or n / len(b) < min_ratio:
        return False
    return True


def _buscar_repuestos_compatibles(cod_norm_base, nombre_item, sucursal, max_alt=3):
    """
    Busca variantes/derivaciones del MISMO repuesto (mismo producto en otro
    formato/envase, ej. distinto litraje) que tambien sirven. Criterio UNICO:
    el codigo del candidato debe estar relacionado con el codigo base segun
    `_codigos_relacionados()` (prefijo compartido sustancial). No se usa
    Familia/SubFamilia ni similitud de descripcion como criterio de inclusion
    — solo sirven, si acaso, para ordenar entre varios candidatos ya validos
    por codigo. Devuelve lista de hasta `max_alt` dicts:
    {"codigo","descripcion","stock_sucursal","stock_otro"}. Si no hay catalogo
    detallado cargado, o no se encuentra ningun codigo relacionado, devuelve
    lista vacia (nunca inventa una alternativa).
    """
    if not _STOCK_DETALLE_CACHE or not cod_norm_base:
        return []
    cod_completo = _detalle_lookup_cod_completo(cod_norm_base)
    candidatos = []
    for cod_c, info in _STOCK_DETALLE_CACHE.items():
        if cod_c == cod_completo:
            continue
        if not _codigos_relacionados(cod_norm_base, cod_c):
            continue
        desc = info.get("descripcion") or ""
        stock_suc = sum(v["stock"] for bod, v in info["por_bodega"].items() if _bodega_pertenece_sucursal(bod, sucursal))
        stock_otro = sum(v["stock"] for bod, v in info["por_bodega"].items() if not _bodega_pertenece_sucursal(bod, sucursal))
        # ratio solo para ordenar (mayor similitud de codigo primero), no para filtrar
        ratio = len(os.path.commonprefix([cod_norm_base.upper(), cod_c.upper()]))
        candidatos.append((ratio, cod_c, desc, stock_suc, stock_otro))
    candidatos.sort(key=lambda t: (-t[0], -t[3]))
    return [
        {"codigo": _RE_PREFIJO_FAMILIA.sub(r"\1", cc) if _RE_PREFIJO_FAMILIA.match(cc) else cc,
         "descripcion": desc, "stock_sucursal": (stock_suc if stock_suc > 0 else None),
         "stock_otro": (stock_otro if stock_suc <= 0 and stock_otro > 0 else None)}
        for _, cc, desc, stock_suc, stock_otro in candidatos[:max_alt]
    ]


def _stock_lookup(cod_norm):
    """
    Busca un codigo de repuesto en el catalogo de Stock: primero exacto, y si no
    hay match, por sufijo (ignorando el prefijo numerico de familia que usa la
    columna Producto, ej. "25 JL3Z6731A"). Si varias filas del catalogo comparten
    el mismo sufijo, se agregan (stock sumado, bodegas concatenadas, costo del
    primero que tenga valor) — mismo criterio que ya usa leer_stock_repuestos().
    """
    global _STOCK_SUFFIX_CACHE
    if not cod_norm or not _STOCK_IDX_CACHE:
        return None
    directo = _STOCK_IDX_CACHE.get(cod_norm)
    if directo:
        return directo
    if _STOCK_SUFFIX_CACHE is None:
        idx = {}
        for cod_completo, info in _STOCK_IDX_CACHE.items():
            m = _RE_PREFIJO_FAMILIA.match(cod_completo)
            sufijo = m.group(1) if m else cod_completo
            if sufijo not in idx:
                idx[sufijo] = {"bodega": info["bodega"], "stock": info["stock"], "costo": info["costo"]}
            else:
                if info["bodega"] not in idx[sufijo]["bodega"].split(" / "):
                    idx[sufijo]["bodega"] = idx[sufijo]["bodega"] + " / " + info["bodega"]
                idx[sufijo]["stock"] += info["stock"]
                if idx[sufijo]["costo"] == 0.0 and info["costo"] > 0.0:
                    idx[sufijo]["costo"] = info["costo"]
        _STOCK_SUFFIX_CACHE = idx
    return _STOCK_SUFFIX_CACHE.get(cod_norm)


def _norm_temp(s):
    """Normaliza texto para cruzar marca/modelo contra el tempario: mayusculas, sin tildes."""
    if not s:
        return ""
    s = str(s).upper().strip()
    s = unicodedata.normalize("NFKD", s)
    s = "".join(c for c in s if not unicodedata.combining(c))
    s = re.sub(r"[^A-Z0-9 ]", " ", s)
    s = re.sub(r"\s+", " ", s).strip()
    return s


def _cargar_tempario_idx():
    """Carga tempario.json (una sola vez por corrida) y arma el indice marca -> [(modelo, km_horas)]."""
    global _TEMPARIO_CACHE
    if _TEMPARIO_CACHE is not None:
        return _TEMPARIO_CACHE
    idx = {}
    try:
        if os.path.exists(RUTA_TEMPARIO):
            with open(RUTA_TEMPARIO, "r", encoding="utf-8") as f:
                datos = json.load(f)
            for m in datos.get("modelos", []):
                marca = _norm_temp(m.get("marca"))
                modelo = _norm_temp(m.get("modelo"))
                km_horas = {}
                for k, v in (m.get("km_horas") or {}).items():
                    try:
                        km_horas[int(k)] = float(v)
                    except Exception:
                        continue
                if marca and modelo and km_horas:
                    idx.setdefault(marca, []).append((modelo, km_horas))
            # Modelos mas largos/especificos primero (evita que "RANGER" se cuele
            # antes que una variante mas especifica del mismo prefijo, si la hubiera)
            for marca in idx:
                idx[marca].sort(key=lambda t: -len(t[0]))
        else:
            log("(!) tempario.json no encontrado junto al script — sin horas estimadas de mantencion")
    except Exception as _e:
        log(f"(!) No se pudo cargar tempario.json: {_e}")
        idx = {}
    _TEMPARIO_CACHE = idx
    return idx


def _parse_km_mantencion(mantencion_str):
    """'20.000 KMS' / '20.000 km' -> 20000. None si no trae un numero de km reconocible."""
    if not mantencion_str:
        return None
    digitos = re.sub(r"[^0-9]", "", str(mantencion_str))
    if not digitos:
        return None
    try:
        km = int(digitos)
    except Exception:
        return None
    return km if km > 0 else None


def _es_servicio_mantencion(servicio_texto):
    """
    El tempario de mano de obra SOLO aplica a mantenciones por kilometraje — no a
    Diagnostico Tecnico, Recall, reparaciones puntuales, etc. (esos servicios no
    tienen una pauta de horas fija, dependen de lo que se encuentre). Se usa como
    guardia antes de cruzar contra el tempario, para no asignarle una duracion
    estimada de mantencion a un servicio que no lo es.
    """
    return "mant" in (servicio_texto or "").lower()


def _buscar_horas_tempario(modelo_texto, mantencion_texto):
    """
    Cruza el 'modelo' (ej. 'FORD RANGER') y 'mantencion' (ej. '20.000 KMS') de una
    cita de Agenda contra el tempario de mano de obra. Devuelve horas (float) o None
    si la marca/modelo no esta cubierto por el tempario (solo las 8 marcas del
    cotizador) o si no hay un kilometraje de mantencion reconocible. El llamador debe
    verificar antes con `_es_servicio_mantencion()` que el servicio sea mantencion.
    """
    idx = _cargar_tempario_idx()
    if not idx:
        return None
    km = _parse_km_mantencion(mantencion_texto)
    if km is None:
        return None
    modelo_norm = _norm_temp(modelo_texto)
    if not modelo_norm:
        return None

    for marca, modelos in idx.items():
        prefijo = marca + " "
        if modelo_norm == marca:
            resto = ""
        elif modelo_norm.startswith(prefijo):
            resto = modelo_norm[len(prefijo):]
        elif marca in modelo_norm.split():
            resto = modelo_norm.replace(marca, "").strip()
        else:
            continue
        # Comparacion sin espacios de respaldo: la Agenda a veces escribe el modelo
        # pegado (ej. "F150") mientras el tempario (derivado del nombre oficial con
        # guion "F-150") queda normalizado con espacio ("F 150") — sin esto, ese caso
        # no matcheaba y se perdia el dato aunque la marca/modelo si estan cubiertos.
        resto_sin_esp = resto.replace(" ", "")
        for modelo_key, km_horas in modelos:
            modelo_key_sin_esp = modelo_key.replace(" ", "")
            calza = (
                resto == modelo_key
                or resto.startswith(modelo_key + " ")
                or (modelo_key and modelo_key in resto)
                or (modelo_key_sin_esp and modelo_key_sin_esp == resto_sin_esp)
                or (modelo_key_sin_esp and resto_sin_esp.startswith(modelo_key_sin_esp))
            )
            if calza:
                if km in km_horas:
                    return km_horas[km]
                # Sin match exacto del km: usar el mas cercano, solo si esta a <=5.000 km
                cercano = min(km_horas.keys(), key=lambda k: abs(k - km))
                if abs(cercano - km) <= 5000:
                    return km_horas[cercano]
                return None
    return None


def _cargar_pauta_repuestos_idx():
    """
    Carga pauta_repuestos.json (una sola vez por corrida) y arma el indice
    marca -> [(modelo, km_data)], igual patron que _cargar_tempario_idx().
    km_data: { km(int): {"horas":.., "mano_obra":.., "items":[...]} }
    Cada item: {"nombre":.., "codigo":.., "cantidad":.., "precio_unitario":..}

    pauta_repuestos.json se genero una sola vez (manualmente, fuera de este script)
    a partir de "Cotizador mantenciones.zip" (plataforma/data/pautas/*.json) — por
    cada marca/modelo/km se tomo la version con mas horas, para mantener
    consistencia con el mismo criterio de tempario.json.
    """
    global _PAUTA_REP_CACHE
    if _PAUTA_REP_CACHE is not None:
        return _PAUTA_REP_CACHE
    idx = {}
    try:
        if os.path.exists(RUTA_PAUTA_REPUESTOS):
            with open(RUTA_PAUTA_REPUESTOS, "r", encoding="utf-8") as f:
                datos = json.load(f)
            for m in datos.get("modelos", []):
                marca = _norm_temp(m.get("marca"))
                modelo = _norm_temp(m.get("modelo"))
                km_data = {}
                for k, v in (m.get("km_data") or {}).items():
                    try:
                        km_data[int(k)] = v
                    except Exception:
                        continue
                if marca and modelo and km_data:
                    idx.setdefault(marca, []).append((modelo, km_data))
            for marca in idx:
                idx[marca].sort(key=lambda t: -len(t[0]))
        else:
            log("(!) pauta_repuestos.json no encontrado junto al script — "
                "sin repuestos sugeridos para el Pre-picking")
    except Exception as _e:
        log(f"(!) No se pudo cargar pauta_repuestos.json: {_e}")
        idx = {}
    _PAUTA_REP_CACHE = idx
    return idx


def _buscar_repuestos_pauta(modelo_texto, mantencion_texto, sucursal=None):
    """
    Cruza marca/modelo + km de mantencion contra pauta_repuestos.json (mismo
    algoritmo de match que _buscar_horas_tempario). Devuelve un dict
    {"horas":.., "mano_obra":.., "items":[{"nombre","codigo","cantidad",
    "precio_unitario","stock","ubicacion","stock_otro_lugar","alternativas"}]}
    o None si no hay cobertura.

    Stock/Ubicacion quedan acotados a la SUCURSAL de la cita (13/07/2026, a pedido
    de Cristobal — antes se sumaba el stock de todas las bodegas del pais, lo que no
    servia para saber si HAY que pedirlo). Si la sucursal no tiene stock del codigo
    exacto, `stock_otro_lugar` indica en que otra bodega si hay. `alternativas` trae
    hasta 3 codigos compatibles — variantes del MISMO codigo (prefijo compartido
    sustancial, ej. 'XO5W30Q1SP' vs 'XO5W30Q1SP-PLZ'/'XO5W30Q1FS'), ver
    `_codigos_relacionados()`. Ya NO se usa Familia/SubFamilia+descripcion (14/07/2026
    — producia falsos positivos graves, ej. sugerir aceite de transmision para uno de
    motor, o filtros sin ninguna relacion real).
    """
    idx = _cargar_pauta_repuestos_idx()
    if not idx:
        return None
    km = _parse_km_mantencion(mantencion_texto)
    if km is None:
        return None
    modelo_norm = _norm_temp(modelo_texto)
    if not modelo_norm:
        return None

    for marca, modelos in idx.items():
        prefijo = marca + " "
        if modelo_norm == marca:
            resto = ""
        elif modelo_norm.startswith(prefijo):
            resto = modelo_norm[len(prefijo):]
        elif marca in modelo_norm.split():
            resto = modelo_norm.replace(marca, "").strip()
        else:
            continue
        resto_sin_esp = resto.replace(" ", "")
        for modelo_key, km_data in modelos:
            modelo_key_sin_esp = modelo_key.replace(" ", "")
            calza = (
                resto == modelo_key
                or resto.startswith(modelo_key + " ")
                or (modelo_key and modelo_key in resto)
                or (modelo_key_sin_esp and modelo_key_sin_esp == resto_sin_esp)
                or (modelo_key_sin_esp and resto_sin_esp.startswith(modelo_key_sin_esp))
            )
            if not calza:
                continue
            datos_km = None
            if km in km_data:
                datos_km = km_data[km]
            else:
                cercano = min(km_data.keys(), key=lambda k: abs(k - km))
                if abs(cercano - km) <= 5000:
                    datos_km = km_data[cercano]
            if datos_km is None:
                return None
            items_out = []
            for it in (datos_km.get("items") or []):
                cod = str(it.get("codigo") or "").strip()
                nombre_it = it.get("nombre") or ""
                cod_norm = _norm_cod_producto(cod) if cod else ""
                stock_info = _stock_lookup(cod_norm) if cod_norm else None  # costo (nacional, sin variar)

                stock_suc, otras_bodegas = (None, [])
                if cod_norm and sucursal:
                    stock_suc, otras_bodegas = _stock_para_sucursal(cod_norm, sucursal)
                stock_otro_lugar = None
                if (not stock_suc) and otras_bodegas:
                    # Solo se menciona donde SI hay stock si en esta sucursal NO hay.
                    otras_bodegas.sort(key=lambda t: -t[1])
                    stock_otro_lugar = ", ".join(f"{b} ({int(q)})" for b, q in otras_bodegas[:3])

                alternativas = (
                    _buscar_repuestos_compatibles(cod_norm, nombre_it, sucursal)
                    if cod_norm and sucursal else []
                )

                items_out.append({
                    "nombre": nombre_it,
                    "codigo": cod,
                    "cantidad": it.get("cantidad") or 0,
                    "precio_unitario": it.get("precio_unitario") or 0,
                    "stock": stock_suc,
                    "ubicacion": (sucursal if stock_suc is not None else None),
                    "stock_otro_lugar": stock_otro_lugar,
                    "costo_real": (stock_info.get("costo") if stock_info else None),
                    "alternativas": alternativas,
                })
            return {
                "horas": datos_km.get("horas") or 0,
                "mano_obra": datos_km.get("mano_obra") or 0,
                "items": items_out,
            }
    return None


# Muestras de iconos de la columna Accion que el parser no reconocio (diagnostico)
_ACCION_NO_RECONOCIDA = set()


def _parsear_citas_html(html_text, fecha_str, nombre_sucursal):
    """
    Parsea el HTML del reporte de reservas de la agenda.
    Detecta el ESTADO de la cita desde el icono de la columna Accion (3 estados):
      - auto (fa-car)      -> estado = "pendiente"   (no ha ingresado)
      - ticket (fa-ticket) -> estado = "ingresado"   (en servicio, en taller)
      - persona (fa-user)  -> estado = "finalizado"  (servicio terminado/retirado)
    `ingresado` (bool) se mantiene por compatibilidad: True solo si estado=="ingresado".
    Retorna lista de dicts con los datos de cada cita.
    """
    import re as _re
    citas = []
    try:
        from html.parser import HTMLParser

        class AgendaParser(HTMLParser):
            def __init__(self):
                super().__init__()
                self.in_table = False
                self.in_tbody = False
                self.in_row = False
                self.in_td = False
                self.cols = []
                self.current_cell = ""
                self.current_cell_html = ""
                self.rows = []
                self.depth = 0
                self.table_depth = 0
                self.found_first_table = False

            def handle_starttag(self, tag, attrs):
                attrs_d = dict(attrs)
                if tag == "table" and not self.found_first_table:
                    # Buscar la tabla principal de reservas
                    cls = attrs_d.get("class", "")
                    if "table" in cls or "reserva" in cls or "reporte" in cls:
                        self.in_table = True
                        self.found_first_table = True
                        self.table_depth = self.depth
                if self.in_table:
                    if tag == "tbody":
                        self.in_tbody = True
                    elif tag == "tr" and self.in_tbody:
                        self.in_row = True
                        self.cols = []
                    elif tag == "td" and self.in_row:
                        self.in_td = True
                        self.current_cell = ""
                        self.current_cell_html = f"<{tag}"
                        for k, v in attrs:
                            self.current_cell_html += f' {k}="{v}"'
                        self.current_cell_html += ">"
                    elif self.in_td:
                        self.current_cell_html += f"<{tag}"
                        for k, v in attrs:
                            self.current_cell_html += f' {k}="{v}"'
                        self.current_cell_html += ">"
                self.depth += 1

            def handle_endtag(self, tag):
                self.depth -= 1
                if self.in_table:
                    if tag == "td" and self.in_td:
                        self.cols.append((self.current_cell.strip(), self.current_cell_html))
                        self.in_td = False
                    elif tag == "tr" and self.in_row and self.cols:
                        self.rows.append(self.cols[:])
                        self.in_row = False
                    elif tag == "tbody":
                        self.in_tbody = False
                    elif tag == "table" and self.depth == self.table_depth:
                        self.in_table = False

            def handle_data(self, data):
                if self.in_td:
                    self.current_cell += data

        parser = AgendaParser()
        parser.feed(html_text)
        rows = parser.rows

        # Si el parser no encontro tabla, intentar con regex simple
        if not rows:
            # Fallback: extraer filas de cualquier tabla
            rows_raw = _re.findall(r"<tr[^>]*>(.*?)</tr>", html_text, _re.DOTALL | _re.IGNORECASE)
            for row_html in rows_raw:
                cells_raw = _re.findall(r"<td[^>]*>(.*?)</td>", row_html, _re.DOTALL | _re.IGNORECASE)
                if len(cells_raw) >= 8:
                    clean = [_re.sub(r"<[^>]+>", " ", c).strip() for c in cells_raw]
                    rows.append([(c, cells_raw[i]) for i, c in enumerate(clean)])

        for row in rows:
            if len(row) < 8:
                continue
            texts = [c[0] for c in row]
            htmls = [c[1] for c in row]
            # Columnas esperadas: OC, Fecha, Horario, Nombre, Modelo, Año, KM, Patente, Servicio, Mant, Accion
            # El orden puede variar; usamos posicion relativa
            if len(texts) >= 9:
                oc       = texts[0].strip()
                fecha    = texts[1].strip()
                horario  = texts[2].strip()
                nombre   = texts[3].strip()
                modelo   = texts[4].strip()
                anio     = texts[5].strip()
                km       = texts[6].strip()
                patente  = texts[7].strip()
                servicio = texts[8].strip() if len(texts) > 8 else ""
                mantencion = texts[9].strip() if len(texts) > 9 else ""
                accion_html = htmls[-1] if htmls else ""

                # Detectar el estado de la cita desde el icono de la columna Accion.
                # La agenda tiene 3 estados (confirmado por Cristobal 07/07/2026):
                #   auto    -> "pendiente"  (no ha ingresado a servicio)
                #   ticket  -> "ingresado"  (entro a servicio, sigue en taller)
                #   persona -> "finalizado" (servicio terminado / vehiculo retirado)
                # IMPORTANTE: persona se evalua PRIMERO — su icono/tooltip puede contener
                # palabras que tambien matchean el ticket (ej. fa-user-check, "ingresado y
                # finalizado"), y antes eso lo clasificaba mal como "ingresado" (causa de la
                # inundacion de Vehiculos en Taller tras el backfill).
                accion_lower = accion_html.lower()
                estado = "pendiente"
                if any(kw in accion_lower for kw in ["fa-user", "fa-person", "fa-male", "fa-street-view",
                                                     "fa-walking", "person", "peaton", "final",
                                                     "entreg", "retir", "termin"]):
                    estado = "finalizado"
                elif any(kw in accion_lower for kw in ["fa-ticket", "ticket", "fa-check", "check", "ingres"]):
                    estado = "ingresado"
                elif any(kw in accion_lower for kw in ["fa-car", "fa-automobile", "auto", "vehicle"]):
                    estado = "pendiente"
                else:
                    # Diagnostico: icono no reconocido — guardar muestra para el log
                    # (maximo 5 por corrida) y asi poder ajustar las palabras clave.
                    if accion_html.strip() and len(_ACCION_NO_RECONOCIDA) < 5:
                        _ACCION_NO_RECONOCIDA.add(accion_html.strip()[:150])
                # Compatibilidad: `ingresado` sigue siendo bool (True solo si esta EN taller).
                # Un finalizado NO cuenta como ingresado — asi ni autoImportarCitas ni el
                # backfill vuelven a crear ordenes de vehiculos que ya se fueron.
                ingresado = (estado == "ingresado")

                # Filtrar filas de encabezado
                if oc.upper() in ("OC", "N° OC", "NUMERO", "FOLIO") or not oc:
                    continue
                # Limpiar OC (solo digitos)
                oc_clean = _re.sub(r"[^0-9]", "", oc)

                # Horas de mano de obra estimadas (tempario) segun marca/modelo + km de
                # mantencion — solo para servicios de Mantencion (no Diagnostico/Recall/
                # reparaciones puntuales); None si la marca no esta cubierta o no hay km
                # reconocible.
                horas_tempario = (
                    _buscar_horas_tempario(modelo, mantencion)
                    if _es_servicio_mantencion(servicio) else None
                )
                # Pauta de repuestos sugeridos (Pre-picking): mismo criterio que arriba
                # -- solo mantenciones por kilometraje -- pero ademas del listado de
                # repuestos (codigo/cantidad/precio de la pauta oficial), enriquecido
                # con Stock/Ubicacion reales cuando el codigo existe en el catalogo.
                pauta_rep = (
                    _buscar_repuestos_pauta(modelo, mantencion, nombre_sucursal)
                    if _es_servicio_mantencion(servicio) else None
                )

                citas.append({
                    "oc":        oc_clean,
                    "fecha":     fecha,
                    "horario":   horario,
                    "nombre":    nombre,
                    "modelo":    modelo,
                    "anio":      anio,
                    "km":        km,
                    "patente":   patente.upper(),
                    "servicio":  servicio,
                    "mantencion": mantencion,
                    "horas_tempario": horas_tempario,
                    "repuestos_sugeridos": (pauta_rep.get("items") if pauta_rep else []),
                    "mano_obra_monto": (pauta_rep.get("mano_obra") if pauta_rep else 0),
                    "ingresado": ingresado,
                    "estado":    estado,       # pendiente / ingresado / finalizado
                    "sucursal":  nombre_sucursal,
                })
    except Exception as _e:
        log(f"(!) Error parseando citas de {nombre_sucursal}: {_e}")
    return citas


def leer_agenda_curifor():
    """
    Descarga las citas de agenda.curifor.cl para las sucursales configuradas,
    para hoy + AGENDA_DIAS_ADELANTE dias (ventana de pre-picking ampliada), y
    enriquece cada cita con el detalle de motorAgenda2.jsp (asesor, RUT,
    telefonos, mail, VIN) usando cache local por OC.
    Retorna dict con estructura:
    {
        "fecha_actualizacion": "DD/MM/YYYY HH:MM",
        "sucursales": {
            "LINDEROS": {
                "30/06/2026": [ {...cita...}, ... ],
                "01/07/2026": [ {...cita...}, ... ],
                "02/07/2026": [ {...cita...}, ... ]
            },
            ...
        }
    }
    """
    import requests as _req
    sucursales_con_id = [(n, v) for n, v in SUCURSALES_AGENDA if v is not None]
    if not sucursales_con_id:
        log("(!) SUCURSALES_AGENDA sin id_cons configurados — omitiendo agenda")
        log("    Configura los id_cons en la constante SUCURSALES_AGENDA del script")
        return None

    ses = _login_agenda()
    if ses is None:
        log("(!) No se pudo autenticar en agenda.curifor.cl")
        return None

    # Fechas: hoy + AGENDA_DIAS_ADELANTE dias hacia adelante
    from datetime import timedelta as _td
    hoy = datetime.now()
    fechas = [hoy + _td(days=i) for i in range(AGENDA_DIAS_ADELANTE + 1)]

    resultado = {
        "fecha_actualizacion": hoy.strftime("%d/%m/%Y %H:%M"),
        "sucursales": {},
    }

    # nombre_suc puede repetirse (ej. "TALCA" con 2 id_cons distintos, Ford y BMW) — en
    # ese caso las citas de ambos id_cons se fusionan bajo la misma clave de salida.
    for nombre_suc, id_cons in sucursales_con_id:
        log(f"Agenda {nombre_suc} (id_cons={id_cons})...")
        if nombre_suc not in resultado["sucursales"]:
            resultado["sucursales"][nombre_suc] = {}

        for fecha_d in fechas:
            fecha_str = fecha_d.strftime("%d/%m/%Y")   # DD/MM/YYYY para el JSON
            resultado["sucursales"][nombre_suc].setdefault(fecha_str, [])
            html_dia = _http_get_agenda(
                ses, AGENDA_REPORT_URL,
                params={"fecha": fecha_str, "id_cons": id_cons, "id_tipo_serv": "-1"},
                timeout=40,
            )
            if html_dia is None:
                log(f"  (!) {nombre_suc} {fecha_str}: sin respuesta tras reintentos")
                continue

            citas = _parsear_citas_html(html_dia, fecha_str, nombre_suc)
            resultado["sucursales"][nombre_suc][fecha_str].extend(citas)
            log(f"  {fecha_str}: {len(citas)} cita(s)")

    # Enriquecer con detalle (asesor, RUT, telefonos, mail, VIN) — cache por OC
    try:
        resultado = _enriquecer_citas_con_detalle(ses, resultado)
    except Exception as _e:
        log(f"(!) Enriquecimiento de detalle fallo (se sube la agenda sin detalle): {_e}")

    if _ACCION_NO_RECONOCIDA:
        log(f"(!) {len(_ACCION_NO_RECONOCIDA)} icono(s) de la columna Accion no reconocidos — "
            f"muestras (para ajustar palabras clave del estado):")
        for _m in sorted(_ACCION_NO_RECONOCIDA):
            log(f"    {_m}")

    return resultado


def _detect_tipo_backfill(servicio, mantencion):
    """Replica detectTipo() del JS: clasifica Mantencion/Reparacion/Diagnostico/Otro."""
    s = (servicio or mantencion or "").lower()
    if "mant" in s or mantencion:
        return "mant"
    if "diag" in s:
        return "diag"
    if "rep" in s:
        return "rep"
    return "ot"


def _fecha_ddmmyyyy_a_iso(fecha_str):
    """'DD/MM/YYYY' o 'DD/MM/YY' -> 'YYYY-MM-DD'. Si no calza, devuelve la fecha de hoy en ISO."""
    try:
        p = fecha_str.strip().split("/")
        if len(p) == 3:
            anio = p[2].strip()
            if len(anio) == 2:
                anio = "20" + anio  # la agenda a veces trae el ano en 2 digitos (DD/MM/YY)
            return f"{anio}-{p[1].zfill(2)}-{p[0].zfill(2)}"
    except Exception:
        pass
    return datetime.now().strftime("%Y-%m-%d")


def _sanear_patente(patente):
    """Deja solo letras/digitos, mayusculas, maximo 8 caracteres (evita basura tipo comillas/HTML)."""
    limpia = re.sub(r"[^A-Za-z0-9]", "", (patente or "")).upper()
    return limpia[:8]


def _norm_suc_campana(s):
    """Normaliza un nombre de sucursal para comparar sin tildes/mayusculas/espacios extra."""
    return _quitar_acentos(str(s or "")).upper().strip()


def _sucursal_desde_campana(valor_crudo, sucursales_reales):
    """
    El archivo de Campañas (Agenda Ford) trae la sucursal con un formato propio
    (ej. '21 - Curifor - Linderos', '131 - Curifor Chillan - Comerciales') que
    NO coincide textualmente con el nombre real de SUCURSAL que usa el resto
    de la app (viene del PBI, ej. 'LINDEROS', 'CHILLAN'). Se identifica "por
    lógica" (pedido explícito de Cristóbal, 28/07/2026): se quita el prefijo
    numérico + 'Curifor' y el sufijo '- Comerciales', y el texto que queda se
    compara (sin tildes/mayúsculas) contra las sucursales reales conocidas del
    PBI, aceptando coincidencia exacta o por contención en cualquier dirección
    (mismo criterio que normSuc() en el Planificador de Taller para el caso
    "Taller Móvil"/"Chillán Viejo", sesión 01/07/2026).

    Devuelve el nombre REAL de la sucursal (tal como aparece en el PBI) si hay
    match, o el texto ya limpio (sin normalizar) si no se encontró ninguno —
    para que quede visible en el módulo en vez de perderse en silencio.
    """
    s = re.sub(r"^\s*\d+\s*-\s*Curifor\s*-?\s*", "", str(valor_crudo or ""), flags=re.I).strip()
    s = re.sub(r"\s*-\s*Comerciales\s*$", "", s, flags=re.I).strip()
    s_norm = _norm_suc_campana(s)
    if not s_norm:
        return s or str(valor_crudo or "").strip()

    mejor = None
    for suc in sucursales_reales:
        suc_norm = _norm_suc_campana(suc)
        if not suc_norm:
            continue
        if suc_norm == s_norm:
            return suc
        if mejor is None and (suc_norm in s_norm or s_norm in suc_norm):
            mejor = suc
    return mejor or s


# Códigos FSA (Field Service Action) de campañas de RECALL que Ford considera
# OBLIGATORIAS — lista entregada por Cristóbal el 28/07/2026. Si el texto de
# la columna W (Campañas/Boletín) de una cita trae AL MENOS uno de estos
# códigos, la cita se marca con la alerta "🚨 Recall Obligatorio".
CODIGOS_RECALL_OBLIGATORIO = {
    "24B24", "24B50", "24B79", "24B83", "24P37", "24S72",
    "25B08", "25B09", "25B52", "25C05", "25C31", "25C43", "25C69",
    "25P04", "25P15", "25S15", "25S21", "25S30", "25S72",
    "25SA3", "25SB7", "25SC3",
}


def _detectar_recall_obligatorio(campanas_txt):
    """
    Compara los códigos de la columna W (Campañas/Boletín, ej. '23S55,25S22')
    contra CODIGOS_RECALL_OBLIGATORIO. La columna trae los códigos separados
    por coma (confirmado con datos reales) -> se tokeniza por coma/punto y
    coma/espacio y se compara cada token (mayúscula, sin espacios) contra el
    set. Devuelve la lista de códigos que matchearon (orden estable), o []
    si ninguno.
    """
    if not campanas_txt:
        return []
    tokens = re.split(r"[,;/\s]+", str(campanas_txt).strip())
    encontrados = []
    for t in tokens:
        t_norm = t.strip().upper()
        if t_norm in CODIGOS_RECALL_OBLIGATORIO and t_norm not in encontrados:
            encontrados.append(t_norm)
    return encontrados


def leer_campanas_curifor(sucursales_reales):
    """
    Lee el archivo de "Revisión de Campañas" (Agenda Ford), que se deja en
    CARPETA_AGENDA_FORD con nombre 'AAAA-MM-DD_Consolidado_Curifor.xlsx' (o,
    en el formato anterior, 'AAAA-MM-DD_Consolidado_Curifor_2Tandas.xlsx' —
    ambos con la misma estructura de columnas, confirmado 29/07/2026) — la
    fecha del nombre cambia cada día, el resto del nombre/formato es siempre
    igual. Siempre se toma el archivo MÁS RECIENTE entre los que matcheen
    cualquiera de los 2 patrones.

    Columnas usadas (hoja 'Consolidado', encabezados en fila 1), identificadas
    por letra tal como las indicó Cristóbal:
      A=Sucursal, D=Fecha de Programación, E=Hora, F=ID Cita, K=Asesor,
      L=Modelo, M=Placa/Patente, N=Chasis, P=Nombre Propietario, V=Servicio,
      W=Campañas/Boletín, X=Orden de servicio, AA=Revisado, AB=Status,
      AK=Fecha de cierre del cronograma real (se muestra en la app como
      "Fecha Cierre" — solo viene con dato una vez que la cita se cerró).

    Solo viajan al módulo las filas que tienen datos en la columna W
    (Campañas/Boletín) — es la que de verdad importa (pedido explícito).

    Estado calculado por fila (no viene resuelto en el archivo — se deriva):
      - Fecha de Programación anterior a hoy               -> "no_realizada"
        (🔴 "Campaña No Realizada"), sin importar la columna AA.
      - Fecha de Programación de hoy en adelante, y AA dice
        "Revisado"                                          -> "revisada"
        (🟢 "Cita Revisada").
      - Fecha de Programación de hoy en adelante, y AA dice
        "No revisado" (o cualquier otra cosa)                -> "no_revisada"
        (🟡 "No revisada").
    (Además: fecha de hoy con Status "En Curso"/"Agendado" -> "cita_hoy", 🔵.)

    Cada registro trae también "recall_obligatorio" (bool) y "recall_codigos"
    (string) — si la columna W trae al menos uno de los códigos FSA de
    CODIGOS_RECALL_OBLIGATORIO, la cita queda marcada con la alerta
    "🚨 Recall Obligatorio" en la app (pedido de Cristóbal, 28/07/2026).

    Devuelve (lista_de_registros, nombre_archivo_usado) — lista vacía y string
    vacío si no se encontró ningún archivo. Nunca lanza excepción hacia afuera
    en el caso de fallo de lectura (se llama siempre dentro de un try/except
    en main(), igual que el resto de fuentes opcionales).
    """
    patron_nuevo = os.path.join(CARPETA_AGENDA_FORD, "*_Consolidado_Curifor.xlsx")
    patron_viejo = os.path.join(CARPETA_AGENDA_FORD, "*_Consolidado_Curifor_2Tandas.xlsx")
    archivos = [
        a for a in (glob.glob(patron_nuevo) + glob.glob(patron_viejo))
        if not os.path.basename(a).startswith("~$")
    ]
    if not archivos:
        log(f"(i) No se encontró ningún archivo de Campañas en "
            f"'{CARPETA_AGENDA_FORD}' con el patrón "
            f"'*_Consolidado_Curifor.xlsx' (ni '*_Consolidado_Curifor_2Tandas.xlsx') "
            f"— se omite Revisión de Campañas")
        return [], ""

    ruta = max(archivos, key=os.path.getmtime)
    nombre_archivo = os.path.basename(ruta)

    # Copiar a temporal (evita bloqueo si OneDrive/Excel lo tiene abierto)
    ruta_lectura = ruta
    _tmp = None
    try:
        import tempfile as _tf
        _fd, _tmp = _tf.mkstemp(suffix=".xlsx")
        os.close(_fd)
        shutil.copy2(ruta, _tmp)
        ruta_lectura = _tmp
    except Exception:
        ruta_lectura = ruta
        _tmp = None

    try:
        df_c = pd.read_excel(ruta_lectura, sheet_name="Consolidado", header=0,
                              engine="calamine", dtype=str)
    except Exception as _e:
        log(f"(!) Error leyendo '{nombre_archivo}' (Campañas): {_e}")
        return [], nombre_archivo
    finally:
        if _tmp and os.path.exists(_tmp):
            try:
                os.remove(_tmp)
            except Exception:
                pass

    cols = list(df_c.columns)

    def _col(letra):
        idx = 0
        for ch in letra:
            idx = idx * 26 + (ord(ch.upper()) - 64)
        idx -= 1
        return cols[idx] if 0 <= idx < len(cols) else None

    c_suc         = _col("A")
    c_fecha       = _col("D")
    c_hora        = _col("E")
    c_idcita      = _col("F")
    c_asesor      = _col("K")
    c_modelo      = _col("L")
    c_patente     = _col("M")
    c_chasis      = _col("N")
    c_propietario = _col("P")
    c_servicio    = _col("V")
    c_campanas    = _col("W")
    c_orden       = _col("X")
    c_revisado    = _col("AA")
    c_status      = _col("AB")
    c_fecha_cierre = _col("AK")

    log(f"'{nombre_archivo}' — columnas detectadas: A(Sucursal)={c_suc!r} "
        f"D(FechaProgramacion)={c_fecha!r} L(Modelo)={c_modelo!r} M(Patente)={c_patente!r} "
        f"N(Chasis)={c_chasis!r} P(Propietario)={c_propietario!r} X(OrdenServicio)={c_orden!r} "
        f"W(Campañas)={c_campanas!r} AA(Revisado)={c_revisado!r} "
        f"AK(FechaCierreCronograma)={c_fecha_cierre!r}")

    if not c_campanas or not c_suc:
        log(f"(!) '{nombre_archivo}': no se pudieron ubicar las columnas esperadas "
            f"(A=Sucursal / W=Campañas/Boletín). Revisar si cambió el formato del archivo.")
        return [], nombre_archivo

    def _v(fila, col):
        """Valor de una celda como string limpio — nunca deja pasar el texto
        literal 'nan' que produce str(NaN) cuando la celda viene vacía
        (dtype=str deja las celdas vacías como float NaN, no como '')."""
        if not col:
            return ""
        val = fila.get(col)
        if val is None or (isinstance(val, float) and pd.isna(val)):
            return ""
        s = str(val).strip()
        return "" if s.lower() == "nan" else s

    hoy = datetime.now().date()
    excluidos_no_asistio = 0
    excluidos_duplicados = 0
    ids_cita_vistos = set()
    registros = []
    for _, fila in df_c.iterrows():
        campanas = _v(fila, c_campanas)
        if not campanas:
            continue

        status_txt = _v(fila, c_status)
        if "no asist" in _quitar_acentos(status_txt).lower():
            # Citas marcadas como "No asistió" no valen -> se excluyen del listado
            excluidos_no_asistio += 1
            continue

        # El archivo "2Tandas" trae la MISMA cita 2 veces (una vez por cada
        # "Tanda": Historico y Proximos) -> deduplicar por ID Cita (columna F)
        # para que cada cita real aparezca solo una vez en el listado.
        id_cita_val = _v(fila, c_idcita)
        clave_dedup = id_cita_val or f"{_v(fila, c_patente)}|{_v(fila, c_orden)}|{_v(fila, c_fecha)}"
        if clave_dedup in ids_cita_vistos:
            excluidos_duplicados += 1
            continue
        ids_cita_vistos.add(clave_dedup)

        sucursal_cruda = _v(fila, c_suc)
        sucursal = _sucursal_desde_campana(sucursal_cruda, sucursales_reales)

        fecha_txt = _v(fila, c_fecha)
        fecha_prog = None
        try:
            d, m, a = fecha_txt.split("/")
            fecha_prog = datetime(int(a), int(m), int(d)).date()
        except Exception:
            fecha_prog = None

        revisado = _v(fila, c_revisado)
        _status_norm = _quitar_acentos(status_txt).lower().strip()

        if fecha_prog is not None and fecha_prog < hoy:
            estado, estado_color, estado_texto = "no_realizada", "rojo", "Campaña No Realizada"
        elif fecha_prog is not None and fecha_prog == hoy and _status_norm in ("en curso", "agendado"):
            # Citas de HOY (fecha de la consulta) que todavia estan En Curso o
            # Agendado -> se destacan aparte, ni "no realizada" (no paso la fecha
            # todavia) ni confundidas con el resto de "hoy en adelante".
            estado, estado_color, estado_texto = "cita_hoy", "azul", "Cita de Hoy"
        elif revisado.lower() == "revisado":
            estado, estado_color, estado_texto = "revisada", "verde", "Cita Revisada"
        else:
            estado, estado_color, estado_texto = "no_revisada", "amarillo", "No revisada"

        _codigos_recall = _detectar_recall_obligatorio(campanas)

        registros.append({
            "sucursal":            sucursal,
            "sucursal_original":   sucursal_cruda,
            "fecha_programacion":  fecha_txt,
            "hora":                _v(fila, c_hora),
            "id_cita":             _v(fila, c_idcita),
            "asesor":              _v(fila, c_asesor),
            "modelo":              _v(fila, c_modelo),
            "patente":             _v(fila, c_patente).upper(),
            "chasis":              _v(fila, c_chasis),
            "propietario":         _v(fila, c_propietario),
            "servicio":            _v(fila, c_servicio),
            "campanas":            campanas,
            "orden_servicio":      _v(fila, c_orden),
            "revisado":            revisado,
            "status":              _v(fila, c_status),
            "fecha_cierre":        _v(fila, c_fecha_cierre),
            "estado":              estado,
            "estado_color":        estado_color,
            "estado_texto":        estado_texto,
            "recall_obligatorio":  bool(_codigos_recall),
            "recall_codigos":      ", ".join(_codigos_recall),
        })

    if excluidos_no_asistio:
        log(f"'{nombre_archivo}': {excluidos_no_asistio} caso(s) con Status "
            f"'No asistió' excluidos del listado de Campañas.")
    if excluidos_duplicados:
        log(f"'{nombre_archivo}': {excluidos_duplicados} fila(s) duplicada(s) "
            f"(misma cita en 2 'Tandas') excluidas del listado de Campañas.")
    _n_recall = sum(1 for r in registros if r["recall_obligatorio"])
    if _n_recall:
        log(f"'{nombre_archivo}': {_n_recall} caso(s) con código FSA de Recall "
            f"Obligatorio detectado (🚨).")

    return registros, nombre_archivo


def _leer_json_github_simple(nombre_archivo):
    """Lee un JSON del repo GitHub. Retorna (sha, dict). dict={} si no existe/vacio."""
    url = f"https://api.github.com/repos/{GITHUB_USUARIO}/{GITHUB_REPO}/contents/{nombre_archivo}"
    headers = {"Authorization": f"token {GITHUB_TOKEN}", "Accept": "application/vnd.github.v3+json"}
    try:
        r = requests.get(url, headers=headers, timeout=15, verify=False)
        if r.status_code != 200:
            return None, {}
        info = r.json()
        sha = info.get("sha")
        contenido = info.get("content", "")
        if contenido:
            datos = json.loads(base64.b64decode(contenido).decode("utf-8"))
        else:
            # Archivo grande (>1MB): usar download_url
            dl = info.get("download_url")
            datos = requests.get(dl, timeout=20, verify=False).json() if dl else {}
        return sha, datos
    except Exception as _e:
        log(f"(!) Error leyendo {nombre_archivo} de GitHub: {_e}")
        return None, {}


def _subir_json_github_simple(nombre_archivo, datos, sha, mensaje, timeout=30):
    """Sube un JSON al repo GitHub via API REST (create/update)."""
    url = f"https://api.github.com/repos/{GITHUB_USUARIO}/{GITHUB_REPO}/contents/{nombre_archivo}"
    headers = {"Authorization": f"token {GITHUB_TOKEN}", "Accept": "application/vnd.github.v3+json"}
    content_b64 = base64.b64encode(json.dumps(datos, ensure_ascii=False, indent=2).encode("utf-8")).decode()
    payload = {"message": mensaje, "content": content_b64}
    if sha:
        payload["sha"] = sha
    try:
        r = requests.put(url, headers=headers, json=payload, timeout=timeout, verify=False)
        if r.status_code in (200, 201):
            return True
        # Diagnostico: antes esto se perdia en silencio (solo se logueaba en
        # excepcion) — ahora se ve el status y el motivo real que da GitHub.
        log(f"(!) GitHub respondio {r.status_code} al subir {nombre_archivo}: {r.text[:300]}")
        return False
    except Exception as _e:
        log(f"(!) Error subiendo {nombre_archivo} a GitHub: {_e}")
        return False


def _subir_json_github_gitdata(nombre_archivo, datos, mensaje, timeout=60):
    """
    Sube un JSON a GitHub via Git Data API (blob -> tree -> commit -> ref),
    en vez de la Contents API que usa _subir_json_github_simple().

    Se agregó porque la Contents API (PUT /contents/{path}) devolvió, al subir
    stock_repuestos.json (~13 MB en base64), un error 403 "Timed out validating
    rule, please try again" — un timeout del lado de GitHub al validar el
    archivo (p.ej. el escaneo de secretos) para archivos grandes vía esa API
    puntual. La Git Data API arma el commit directo a partir de blobs/trees sin
    pasar por esa validación por archivo, y es la misma vía que ya usa
    guardar_en_github() en app.py para escribir datos_dashboard.json sin
    problemas. No requiere SHA del archivo anterior (arma el tree nuevo a
    partir del último commit de 'main').

    29/07/2026 — FIX de carrera ("Update is not a fast forward", HTTP 422):
    la version anterior leia el head de 'main' UNA sola vez y despues hacia
    tree -> commit -> patch. Si entre esa lectura y el PATCH entraba cualquier
    otro commit en main (cosa habitual: la app web escribe control_taller.json
    /prepicking cada vez que alguien edita algo, y la propia consolidacion sube
    varios archivos seguidos antes de este), el commit nuevo quedaba colgando
    de un padre que ya no era el head y GitHub rechazaba el PATCH con 422 —
    perdiendo TODO el trabajo del paso (caso real del 29/07: Produccion
    Tecnicos se calculo bien, con la homologacion aplicada, pero nunca llego a
    GitHub, asi que la app siguio mostrando los datos viejos del 22/07).
    Ahora el ciclo head->tree->commit->patch se reintenta hasta 4 veces,
    releyendo el head fresco en cada intento y reconstruyendo el commit sobre
    esa base nueva. El blob se crea una sola vez (es content-addressed, sigue
    siendo valido entre intentos). Nunca se hace force push: si el reintento
    tampoco alcanza, se reporta el error en vez de pisar commits ajenos.
    """
    hdrs = {"Authorization": f"token {GITHUB_TOKEN}", "Accept": "application/vnd.github.v3+json"}
    base_url = f"https://api.github.com/repos/{GITHUB_USUARIO}/{GITHUB_REPO}"
    try:
        contenido_b64 = base64.b64encode(
            json.dumps(datos, ensure_ascii=False, indent=2).encode("utf-8")
        ).decode()

        r = requests.post(f"{base_url}/git/blobs", headers=hdrs,
                           json={"content": contenido_b64, "encoding": "base64"},
                           timeout=timeout, verify=False)
        if r.status_code not in (200, 201):
            log(f"(!) GitHub respondio {r.status_code} creando blob de {nombre_archivo}: {r.text[:300]}")
            return False
        blob_sha = r.json()["sha"]

        _INTENTOS = 4
        for _intento in range(1, _INTENTOS + 1):
            # Head FRESCO en cada intento (esto es lo que evita el 422).
            r = requests.get(f"{base_url}/git/ref/heads/main", headers=hdrs, timeout=15, verify=False)
            r.raise_for_status()
            commit_sha = r.json()["object"]["sha"]

            r = requests.get(f"{base_url}/git/commits/{commit_sha}", headers=hdrs, timeout=15, verify=False)
            r.raise_for_status()
            tree_sha = r.json()["tree"]["sha"]

            r = requests.post(f"{base_url}/git/trees", headers=hdrs,
                               json={"base_tree": tree_sha,
                                     "tree": [{"path": nombre_archivo, "mode": "100644",
                                               "type": "blob", "sha": blob_sha}]},
                               timeout=30, verify=False)
            if r.status_code not in (200, 201):
                log(f"(!) GitHub respondio {r.status_code} creando tree para {nombre_archivo}: {r.text[:300]}")
                return False
            new_tree_sha = r.json()["sha"]

            r = requests.post(f"{base_url}/git/commits", headers=hdrs,
                               json={"message": mensaje, "tree": new_tree_sha, "parents": [commit_sha]},
                               timeout=30, verify=False)
            if r.status_code not in (200, 201):
                log(f"(!) GitHub respondio {r.status_code} creando commit para {nombre_archivo}: {r.text[:300]}")
                return False
            new_commit_sha = r.json()["sha"]

            r = requests.patch(f"{base_url}/git/refs/heads/main", headers=hdrs,
                                json={"sha": new_commit_sha}, timeout=15, verify=False)
            if r.status_code in (200, 201):
                if _intento > 1:
                    log(f"    {nombre_archivo}: subido en el intento {_intento} "
                        f"(la rama main habia avanzado, se reintento sobre el head nuevo).")
                return True

            # 422 = alguien mas movio main entremedio -> reintentar con head fresco.
            _es_carrera = (r.status_code == 422 and "fast forward" in r.text.lower())
            if _es_carrera and _intento < _INTENTOS:
                log(f"    {nombre_archivo}: la rama main avanzo mientras se subia "
                    f"(intento {_intento}/{_INTENTOS}) — reintentando sobre el head nuevo...")
                continue

            log(f"(!) GitHub respondio {r.status_code} actualizando la rama main "
                f"tras subir {nombre_archivo}: {r.text[:300]}")
            return False

        return False
    except Exception as _e:
        log(f"(!) Error subiendo {nombre_archivo} a GitHub (Git Data API): {_e}")
        return False


def _leer_credenciales_sql():
    """Lee usuario/clave de sql_credenciales.txt (2 lineas: usuario, clave).
    Devuelve (None, None) si el archivo no existe o esta mal formado."""
    if not os.path.exists(_sql_cred_file):
        return None, None
    try:
        with open(_sql_cred_file, "r", encoding="utf-8") as f:
            lineas = [l.strip() for l in f.readlines() if l.strip()]
        if len(lineas) >= 2:
            return lineas[0], lineas[1]
    except Exception:
        pass
    return None, None


# Consulta SQL — union de las 2 tablas de produccion de mano de obra por
# tecnico/mecanico, mismos tipos de documento y logica que la consulta M/Power
# Query que ya se usa para el PBI de Produccion Mensual de Mecanicos (pasada
# tal cual a T-SQL, sin modificar la logica de negocio).
_SQL_PRODUCCION_TECNICOS = """
Select mecanico as 'Mecanico', [Rut] as 'Rut Mecanico', [Área] as 'Area Mecanico',
    [Cod. Sucursal Mecanico] as 'Cod Sucursal Mecanico', [Nombre Sucursal Mecanico],
    tipodocto as 'Tipo Docto', Numero, [local] as 'Sucursal', fecha as 'Fecha',
    [Area Venta], Marca, [Nº OT] as 'Nro OT', producto, [Precio Lista],
    [Total Horas], [Comi_Vta], 'PMM' as [_Origen]
From Tmp_ProduccionMensualMecanicos
Where Empresa = 'E01'
And UPPER(LTRIM(RTRIM(tipodocto))) IN ('NOTA CREDITO ST','REFACTURACION C/RS','SOLICITUD N/C ST','FACTURA S/T','CARGO INTERNO S/T','FACTURA GARANTIA S/T','FACTURA ST','CARGO INTERNO','FACTURA SEGURO S/T')
UNION ALL
Select Mecanico, NULL, NULL, NULL, NULL,
    Documento, Folio, NULL, NULL,
    TipoCargo, Marca, NroOT,
    Producto, Precio, Cantidadd, Total, 'HPT' as [_Origen]
From flexline.Tmp_HorasPorTecnico
Where UPPER(LTRIM(RTRIM(Documento))) IN ('NOTA CREDITO ST','REFACTURACION C/RS','SOLICITUD N/C ST','FACTURA S/T','CARGO INTERNO S/T','FACTURA GARANTIA S/T','FACTURA ST','CARGO INTERNO','FACTURA SEGURO S/T')
"""


def _conectar_bdflexline():
    """Abre conexion pyodbc a BDFlexline (10.50.15.2). Devuelve None (sin lanzar
    excepcion) si falta pyodbc, faltan credenciales, no hay driver ODBC de SQL
    Server instalado, o el servidor no responde (ej. fuera de la red interna)."""
    try:
        import pyodbc
    except ImportError:
        log("(!) pyodbc no esta instalado — 'pip install pyodbc' (Ejecutar_Consolidacion.bat "
            "lo instala solo). Produccion Tecnicos no se actualizara esta corrida.")
        return None

    usuario, clave = _leer_credenciales_sql()
    if not usuario or not clave:
        log(f"(!) No se encontraron credenciales SQL en {os.path.basename(_sql_cred_file)} "
            f"(2 lineas: usuario / clave) — Produccion Tecnicos no se actualizara.")
        return None

    try:
        drivers_disponibles = pyodbc.drivers()
    except Exception:
        drivers_disponibles = []
    _preferidos = [
        "ODBC Driver 18 for SQL Server",
        "ODBC Driver 17 for SQL Server",
        "ODBC Driver 13 for SQL Server",
        "ODBC Driver 11 for SQL Server",
        "SQL Server Native Client 11.0",
        "SQL Server",
    ]
    driver = next((d for d in _preferidos if d in drivers_disponibles), None)
    if not driver:
        log(f"(!) No se encontro un driver ODBC de SQL Server instalado en esta PC. "
            f"Disponibles: {drivers_disponibles or 'ninguno'}. Instala 'ODBC Driver 17 for "
            f"SQL Server' de Microsoft (el mismo que usa Power BI Desktop para conectarse).")
        return None

    conn_str = (
        f"DRIVER={{{driver}}};SERVER={SQL_SERVER};DATABASE={SQL_DATABASE};"
        f"UID={usuario};PWD={clave};TrustServerCertificate=yes;Connection Timeout=15;"
    )
    try:
        return pyodbc.connect(conn_str, timeout=15)
    except Exception as _e:
        log(f"(!) No se pudo conectar a BDFlexline ({SQL_SERVER}): {_e}")
        return None


# Cuantos meses hacia atras se conservan en produccion_tecnicos.json. La consulta
# real trae el HISTORICO COMPLETO de BDFlexline (90.000+ filas la primera vez) —
# subir eso fila por fila hizo que la API de blobs de GitHub rechazara la subida
# ("your input was too large to process", HTTP 422). Se agrega por
# mecanico+sucursal+mes (sumando horas, contando OT distintos) y se recorta a
# los ultimos N meses para el resumen; el detalle por producto (mas pesado, un
# registro por cada combinacion mecanico+mes+producto) se recorta ademas a
# PRODUCCION_DETALLE_MESES para no volver a pasarse de tamaño.
PRODUCCION_MESES_HISTORIAL = 12
PRODUCCION_DETALLE_MESES = 3

# 20/07/2026 (a pedido de Cristobal, tras ver la primera corrida real): los
# datos de BDFlexline mezclan filas de Mano de Obra (horas reales del tecnico)
# con filas de repuestos/insumos — sin filtrar, "Total Horas" sumaba cantidades
# de repuestos y daba cifras disparatadas (millones de "horas"). Confirmado con
# Cristobal: solo cuentan como horas los productos cuyo codigo empieza con
# "MO_". Al principio se acepto CUALQUIER codigo con ese prefijo (mas robusto
# a futuro, en teoria) en vez de una lista exacta.
#
# 23/07/2026 — FIX: esa regla de prefijo resulto demasiado permisiva. Cristobal
# reporto un caso real imposible (HECTOR ANDRADE con 30.081,7 "horas
# facturadas" / 349 OT / 15003% de productividad) — se aislo el problema al
# bucket "Sin fecha" de ese tecnico: 30.047,2 horas en solo 338 filas (~89 h
# promedio por fila, un valor sin sentido para mano de obra real). Esto viene
# de filas de origen "HPT" (Tmp_HorasPorTecnico, sin Sucursal/Fecha) cuyo
# codigo de producto tiene el prefijo "MO_" pero NO es uno de los codigos
# reales de mano de obra — Cristobal envio la lista completa real (captura de
# "Etiquetas de fila" de un pivote/filtro), confirmando que el universo real
# de codigos de mano de obra es un conjunto CERRADO y conocido, no "cualquier
# cosa con MO_". Se cambia de un prefijo abierto a esta lista exacta
# (whitelist) — cualquier producto con prefijo "MO_" que NO este en esta
# lista ya no cuenta como hora real, sin importar cuanto "Total Horas" traiga.
#
# 30/07/2026 — Cristobal pidio SACAR la whitelist y volver al prefijo abierto:
# ahora cuenta como Mano de Obra cualquier codigo que empiece con "MO" (tanto
# "MO_XXX" como "MO XXX", "MO-XXX" o el codigo "MO" solo). Se exige que despues
# del "MO" venga un separador (_ - espacio) o el fin del codigo, para no barrer
# por accidente codigos de repuesto que solo empiecen con esas 2 letras
# (ej. "MOTOR...", "MOLDURA..."). La lista de abajo SE CONSERVA pero ya NO
# filtra nada: solo se usa como diagnostico en el log, para dejar visible que
# codigos nuevos entraron al calculo que antes no estaban considerados.
_CODIGOS_MANO_DE_OBRA = {
    "MO_RECTIFICADO DIS", "MO_ALL-BALANCED_EX", "MO_IVECO", "MO_BRILLANCE",
    "MO_CHEVROLET", "MO_CITROEN", "MO_FORD", "MO_GEELY", "MO_GTA_BMW",
    "MO_GTA_FORTALEZA", "MO_GTA_HYUNDAI", "MO_GTA_JAC", "MO_GTA_MAHINDRA",
    "MO_GTIA_FORD", "MO_HYUNDAI", "MO_INTERNA FORD", "MO_INTERNA FORTALEZA",
    "MO_INTERNA HYUNDAI", "MO_INTERNA IVECO", "MO_INTERNA JAC", "MO_JAC",
    "MO_JAECOO", "MO_JIM", "MO_JMC-LIVIANO", "MO_MAHINDRA", "MO_MEC_BMW",
    "MO_MECTECNORED", "MO_OMODA", "MO_PRE ENTREGA", "MO_PRE_ENTR_VEH FORD",
    "MO_REPORT", "MO_SERV_BSIBMW",
}


# 30/07/2026 (parte 3) — Codigos que, pese a tener el prefijo "MO", NO son horas
# de tecnico y quedan EXCLUIDOS a pedido expreso de Cristobal: son servicios de
# terceros / insumos donde la columna "Total Horas" no representa horas reales
# (ej. el caso detectado de CORDOVA VIVANCO con 40.000,0 "horas" en
# 'MO SERVICIO 3º', que en realidad es un monto en pesos).
# Se comparan normalizados (sin tildes, "_"/"-" tratados como espacio), asi que
# 'MO_SERVICIO 3º', 'MO SERVICIO 3RO' y variantes de escritura caen igual.
_CODIGOS_MO_EXCLUIDOS = (
    "MO SERVICIO 3",      # cubre el codigo real 'MO SERVICIO 3º'
    "MO SERVICIO 3RO",
    "MO SERVICIO 3ERO",
    "MO NITRO",
    "MO SANITIZACION",
)


def _norm_cod_mo(producto):
    """Normaliza un codigo de producto para comparar contra _CODIGOS_MO_EXCLUIDOS."""
    _p = _quitar_acentos(str(producto or "")).upper().replace("_", " ").replace("-", " ")
    return " ".join(_p.split())


def _es_cod_mo_excluido(producto):
    """
    True si el codigo esta en la lista negra. Se exige que, tras el texto de la
    lista, NO venga otra letra ASCII — asi 'MO NITRO' excluye 'MO NITRO' y
    'MO NITRO EXPRESS', pero NO excluye un hipotetico 'MO NITROGENO ...'.
    (El ordinal 'º' de 'MO SERVICIO 3º' no es letra ASCII, por eso calza.)
    """
    _n = _norm_cod_mo(producto)
    for _x in _CODIGOS_MO_EXCLUIDOS:
        if _n.startswith(_x):
            _sig = _n[len(_x):]
            if _sig == "" or not (_sig[0].isascii() and _sig[0].isalpha()):
                return True
    return False


def _es_mano_de_obra(producto):
    # 30/07/2026 — prefijo abierto "MO" (ver comentario de arriba). Se acepta
    # "MO" solo, o "MO" seguido de _ / - / espacio; NO se acepta "MO" pegado a
    # otra letra o digito (MOTOR, MOLDURA, MO123...), para no colar repuestos.
    _p = str(producto or "").strip().upper()
    if not _p.startswith("MO"):
        return False
    _resto = _p[2:]
    if not (_resto == "" or _resto[0] in ("_", "-", " ")):
        return False
    # ...salvo los codigos de la lista negra de arriba.
    return not _es_cod_mo_excluido(producto)


# 29/07/2026 — Roster real de tecnicos/mecanicos, extraido de "Nomina Area PV
# (Clasificada).xlsx" (cargos MECANICO, MECANICO 1, MECANICO - ALINEADOR,
# MECANICO TALLER MOVIL, AYUDANTE MECANICO — 49 personas). A pedido explicito
# de Cristobal: "consideremos a todos los tecnicos que estan en este archivo,
# no tomemos en cuenta los RUT, guiemonos solamente por el nombre" — el RUT no
# sirve como cruce porque, segun Cristobal, "la lista de tecnicos es imposible
# de obtener completa por RUT" (BDFlexline no siempre trae un RUT limpio/
# completo por tecnico). El nombre tampoco alcanza con una comparacion exacta:
# el MISMO tecnico puede aparecer en BDFlexline bajo varias escrituras
# distintas (ejemplo real dado por Cristobal: "Luis Riquelme" / "Luis  Riquelme"
# (doble espacio) / "Luis Riquelme Q" — 3 variantes para la misma persona) y
# ademas en orden distinto al de la nomina (BDFlexline suele escribir
# "Nombre Apellido", la nomina trae "Apellido Apellido Nombre Nombre").
#
# Por eso el cruce se hace por TOKENS (palabras), no por el string completo:
# se toman las palabras "significativas" (3+ letras, para ignorar iniciales
# sueltas tipo la "Q" del ejemplo) del nombre que trae BDFlexline, y se
# consideran un match si TODAS esas palabras aparecen (exactas, sin importar
# el orden) entre las palabras de un tecnico de la nomina — exigiendo al menos
# 2 palabras en comun (nombre + apellido) para no adivinar con solo un nombre
# de pila. Si el mismo conjunto de palabras calza con mas de un tecnico de la
# nomina a la vez, se considera AMBIGUO y no se asigna nada (se loguea para
# revision manual) — mejor no adivinar que asignar a la persona equivocada.
#
# "TALCA (2)" es una sucursal real y DISTINTA de "TALCA" (confirmado por
# Cristobal: "Talca 2 es Talca 2") — no se fusionan.
#
# Si en el futuro la nomina cambia (ingresos/egresos/traslados), hay que
# regenerar esta lista desde el Excel actualizado (misma logica: filtrar por
# Nombre Cargo en los 5 valores de arriba, tomar Nombre + Lugar de Trabajo,
# normalizar sin tildes/mayuscula/espacios).
NOMINA_TECNICOS = [
    ('ACEVEDO ASTROZA JOSE TOMAS', 'LINDEROS'),
    ('AGUILAR CUEVAS MARCELO ALEJANDRO', 'LINDEROS'),
    ('ALARCON NORAMBUENA JESUS ANTONIO', 'CURICO'),
    ('ALVARADO JARA JUAN JOSE', 'CURICO'),
    ('ANDRADE CABELLO HECTOR MATIAS', 'LINDEROS'),
    ('AVALOS GARRIDO JORGE ALONZO', 'CURICO'),
    ('BAHAMONDEZ LEFIMAN HECTOR RAUL', 'CD REPUESTOS'),
    ('BASTIAS CERNA RAMON ISMAEL', 'CHILLAN'),
    ('BETANCOURT RUBILAR TOMAS AGUSTIN', 'RANCAGUA'),
    ('BIZARRO GOMEZ JUAN PABLO', 'CHILLAN'),
    ('BRAVO ABURTO DAVID IGNACIO', 'TALCA (2)'),
    ('CARDENAS VILLEGAS ALEXIS WILLIAM', 'TALCA'),
    ('CARRENO PENAILILLO ERROL FRANCHESCO', 'PLACILLA'),
    ('CORDOVA VIVANCO FELIPE SEGUNDO', 'LINDEROS'),
    ('FUENTES FLORES ABRAHAM EDUARDO', 'CHILLAN'),
    ('GARCIA OVALLE CRISTIAN CLAUDIO', 'LINDEROS'),
    ('GOMEZ AVILA BENJAMIN IGNACIO', 'TALCA'),
    ('HERNANDEZ RODRIGUEZ NAAMAN', 'CHILLAN'),
    ('HUAIQUI RAMIREZ JUAN SEBASTIAN', 'TALCA'),
    ('JARA MEDINA ARNALDO ANDRES', 'CURICO'),
    ('JARA SEPULVEDA CLAUDIO ALEJANDRO', 'CD REPUESTOS'),
    ('JIMENEZ ARROYO HECTOR RENE', 'CHILLAN'),
    ('LEIVA RAMIREZ ENRIQUE ALFONSO', 'CURICO'),
    ('LOAIZA CASTILLO DANIEL ANDRES', 'RANCAGUA'),
    ('LOBOS LOYOLA JOSE PATRICIO', 'LINDEROS'),
    ('LOPEZ AGUIRRE HECTOR IGNACIO', 'CHILLAN VIEJO'),
    ('MAIDANA MAIDANA SIMON AGUSTIN', 'PLACILLA'),
    ('MARTINEZ AGUILAR ESTEBAN VALENTIN', 'LINDEROS'),
    ('MEDINA GUZMAN NICOLAS ANDRES', 'RANCAGUA'),
    ('MONDACA MOYA ALEJANDRO ANTONIO', 'TALCA'),
    ('MONTECINOS ORMAZABAL CRISTIAN JAVIER', 'RANCAGUA'),
    ('MORENO VERA JORGE ESTEBAN', 'LINDEROS'),
    ('NEIRA BUSTOS JONATHAN IGNACIO', 'CURICO'),
    ('NUNEZ LUCERO HANS YERCO', 'RANCAGUA'),
    ('PACHECO IBARRAT EUGENIO ALEXIS', 'LINDEROS'),
    ('PALMA BUSTAMANTE RODRIGO ALEXANDER', 'CHILLAN VIEJO'),
    ('PEREZ CANTILLANA FRANCISCO JAVIER', 'RANCAGUA'),
    ('PINA QUEZADA DIEGO ANDRES', 'CURICO'),
    ('RAMIREZ ARAYA RODOLFO LEONARDO', 'CURICO'),
    ('RIQUELME BECERRA LUIS HERNAN', 'CHILLAN'),
    ('RIVEROS ALIAGA NICOLAS ANDRES', 'LINDEROS'),
    ('ROBLES MACHUCA FABIAN ALFREDO', 'TALCA'),
    ('RODRIGUEZ MUNOZ NICOLAS MARCELO', 'TALCA'),
    ('SAN MARTIN CARRASCO VICTOR HERMOGENES', 'CHILLAN'),
    ('URIBE ARRIAGADA ARIEL ESTEBAN', 'CD REPUESTOS'),
    ('VALDEBENITO VALDEBENITO EDER PAOLO', 'TALCA (2)'),
    ('VALLEJOS MUENA NICOLAS GABRIEL', 'CHILLAN'),
    ('VILLALOBOS MORENO NICOLAS ANDRES', 'RANCAGUA'),
    ('VIO BELMAR EDGAR ANTONIO', 'TALCA'),
]


# ALIAS MANUAL (escape hatch): nombre crudo tal como viene de BDFlexline ->
# nombre canonico exacto de NOMINA_TECNICOS. Solo hace falta para los casos que
# el matcher automatico por tokens NO puede resolver solo, es decir cuando el
# nombre de BDFlexline no comparte 2 palabras con el de la nomina — por ejemplo
# un apodo, un apellido mal escrito, o un nombre cargado con el apellido de otra
# persona. Se consulta ANTES del matcher automatico.
# Se deja vacio a proposito: al 29/07/2026 los 9 nombres de BDFlexline que no
# homologan solo por tokens NO corresponden a ningun tecnico de la nomina
# actual de Post Venta — son ex-trabajadores o gente que ya no figura ahi:
#   - Angelo Cifuentes, Benjamin Urrutia, Nicolas Parra: ex-trabajadores
#     (confirmado por Cristobal en la sesion del 23/07/2026).
#   - ENRIQUE JARA: ex-trabajador de la sucursal RANCAGUA (confirmado por
#     Cristobal el 29/07/2026). **OJO — NO es "LEIVA RAMIREZ ENRIQUE ALFONSO"**,
#     que es un trabajador ACTIVO y distinto, de la sucursal CURICO. Comparten
#     solo el nombre de pila "ENRIQUE" (y "JARA" es apellido de otras 2 personas
#     de la nomina), asi que a simple vista parecen el mismo — no lo son. NUNCA
#     agregar un alias que los una: fusionaria las horas de dos personas
#     distintas de dos sucursales distintas. El matcher automatico ya los deja
#     correctamente separados (exige 2 palabras en comun, y aca solo hay 1).
#   - Alvaro Carrasco, Claudio Flores, Guillermo Dominguez, Hector Ortega E,
#     Jesus Muñoz: ex-trabajadores (confirmado por Cristobal el 29/07/2026).
# Es decir: los 9 estan correctamente fuera del calculo — ninguno es un tecnico
# activo que se este perdiendo por un problema de nombre.
# Este caso deja la leccion de fondo: compartir UNA sola palabra (nombre de pila
# o apellido) NO significa que sean la misma persona — antes de agregar
# cualquier alias aca hay que confirmarlo con Cristobal, nunca deducirlo por
# parecido de nombre.
# Formato, cuando haga falta agregar uno confirmado:
#     "NOMBRE CRUDO EN BDFLEXLINE": "NOMBRE EXACTO DE NOMINA_TECNICOS",
ALIAS_TECNICOS_BDFLEX = {
}


def _norm_nombre_tecnico(s):
    """Normaliza un nombre de tecnico/mecanico: sin tildes, mayuscula, espacios
    (simples o dobles) colapsados a uno solo."""
    s = _quitar_acentos(s).upper().strip()
    return re.sub(r"\s+", " ", s)


def _tokens_significativos(nombre_norm):
    """Palabras de 3+ letras de un nombre ya normalizado — se ignoran
    iniciales sueltas (ej. la 'Q' de 'Luis Riquelme Q') porque no alcanzan
    para confirmar ni descartar un match por si solas."""
    return frozenset(t for t in nombre_norm.split(" ") if len(t) >= 3)


# Tokens precalculados de cada tecnico de la nomina (una sola vez al cargar
# el modulo) — evita recomputar el split en cada fila de BDFlexline.
_NOMINA_TECNICOS_TOKENS = [
    (nombre, sucursal, _tokens_significativos(_norm_nombre_tecnico(nombre)))
    for nombre, sucursal in NOMINA_TECNICOS
]

# Cache de resultados por nombre crudo de BDFlexline (muchas filas repiten el
# mismo Mecanico) — evita repetir el escaneo de 49 tecnicos por cada fila.
_CACHE_MATCH_NOMINA = {}


def _match_nomina_tecnico(nombre_bdflex):
    """Busca a que tecnico de la NOMINA corresponde un nombre tal como viene
    de BDFlexline (formato y orden pueden ser distintos a los de la nomina).
    Devuelve (nombre_nomina, sucursal) si hay un match unico y sin ambiguedad,
    o (None, None) si no calza con nadie o calza con mas de un tecnico a la
    vez (en ese caso se loguea para revision manual, nunca se adivina)."""
    _clave = str(nombre_bdflex or "").strip()
    if _clave in _CACHE_MATCH_NOMINA:
        return _CACHE_MATCH_NOMINA[_clave]

    _norm = _norm_nombre_tecnico(_clave)

    # 1) Alias manual (tiene prioridad sobre el matcher automatico).
    _alias = ALIAS_TECNICOS_BDFLEX.get(_norm)
    if _alias:
        _suc_alias = next((s for n, s in NOMINA_TECNICOS if n == _alias), None)
        if _suc_alias is not None:
            _CACHE_MATCH_NOMINA[_clave] = (_alias, _suc_alias)
            return _CACHE_MATCH_NOMINA[_clave]
        log(f"(!) Produccion Tecnicos: el alias '{_norm}' -> '{_alias}' apunta a un "
            f"nombre que NO existe en NOMINA_TECNICOS — revisar ALIAS_TECNICOS_BDFLEX.")

    # 2) Matcher automatico por tokens.
    _tokens_cand = _tokens_significativos(_norm)
    resultado = (None, None)
    if len(_tokens_cand) >= 2:
        _candidatos = [
            (nombre, sucursal) for nombre, sucursal, tokens_nom in _NOMINA_TECNICOS_TOKENS
            if _tokens_cand <= tokens_nom
        ]
        if len(_candidatos) == 1:
            resultado = _candidatos[0]
        elif len(_candidatos) > 1:
            log(f"(!) Produccion Tecnicos: '{_clave}' calza por nombre con "
                f"{len(_candidatos)} tecnicos distintos de la nomina "
                f"({', '.join(n for n, _ in _candidatos)}) — ambiguo, no se "
                f"asigna nada (revisar a mano si hace falta).")

    _CACHE_MATCH_NOMINA[_clave] = resultado
    return resultado


def leer_produccion_tecnicos():
    """
    Consulta la produccion de mano de obra de cada tecnico/mecanico contra
    BDFlexline (union de Tmp_ProduccionMensualMecanicos + Tmp_HorasPorTecnico),
    para alimentar la pestaña "Producción Técnicos" del Planificador (horas
    facturadas por tecnico y mes, detalle por producto, comparacion contra la
    jornada laboral). Devuelve un dict {"resumen": [...], "detalle_producto":
    [...], "detalle_ot": [...]} listo para exportar a JSON, o {} si algo fallo
    — nunca lanza excepcion hacia main() para no interrumpir el resto de la
    consolidacion.
    20/07/2026 (reescrita el mismo dia tras la primera corrida real: se agrego
    el filtro de Mano de Obra "MO_", el descarte de tecnicos sin RUT valido
    —placeholders como "SIN COMISION-LIN"/"ST LINDEROS" que traen Rut vacio—,
    normalizacion de nombre a mayusculas (el mismo tecnico aparecia repetido
    con distinta may/min, ej. "JUAN AYALA" y "juan ayala"), y el detalle por
    producto para el drill-down de cada tecnico.
    21/07/2026: agregado "detalle_ot" (linea por linea, sin agrupar, con
    Nº OT/Producto/Precio Lista/Total Horas/Comi_Vta — a pedido de Cristobal
    para replicar la sabana de ejemplo que muestra por tecnico).
    21/07/2026 (parte 4): fix del filtro de RUT valido — excluia por error el
    100% de las filas de flexline.Tmp_HorasPorTecnico (esa tabla no tiene RUT,
    el SQL pone NULL a proposito ahi, no significa que el tecnico sea falso).
    El filtro de RUT ahora solo aplica a las filas de Tmp_ProduccionMensualMecanicos.
    """
    conn = _conectar_bdflexline()
    if conn is None:
        return {}
    try:
        df = pd.read_sql(_SQL_PRODUCCION_TECNICOS, conn)
    except Exception as _e:
        log(f"(!) Error al ejecutar la consulta de Produccion Tecnicos: {_e}")
        return {}
    finally:
        try:
            conn.close()
        except Exception:
            pass

    if df.empty:
        log("Produccion Tecnicos: la consulta no devolvio filas.")
        return {}

    df.columns = [str(c).strip() for c in df.columns]
    filas_leidas = len(df)

    df["Mecanico"] = df.get("Mecanico", "").fillna("").astype(str).str.strip().str.upper()
    df["Rut Mecanico"] = df.get("Rut Mecanico", "").fillna("").astype(str).str.strip()
    df["producto"] = df.get("producto", "").fillna("").astype(str).str.strip()
    df["_Origen"] = df.get("_Origen", "").fillna("").astype(str).str.strip().str.upper()

    # Solo tecnicos con RUT valido — descarta placeholders sin RUT real, tipo
    # "SIN COMISION-LIN"/"ST LINDEROS", que en Tmp_ProduccionMensualMecanicos (origen
    # "PMM") traen Rut vacio o literalmente "SIN RUT".
    #
    # IMPORTANTE (bug real detectado 21/07/2026, reportado por Cristobal — "faltan
    # muchos tecnicos por sucursal"): el SQL es un UNION de 2 tablas, y la segunda
    # (flexline.Tmp_HorasPorTecnico, origen "HPT") NUNCA trae RUT — el SELECT pone
    # NULL en esa columna a proposito, porque esa tabla no tiene ese dato, no porque
    # el tecnico sea un placeholder. Aplicar el filtro de RUT a esa mitad borraba el
    # 100% de sus filas sin importar si el tecnico era real. Ahora el filtro de RUT
    # solo se aplica a las filas de origen "PMM" (que si tienen RUT real disponible);
    # las de origen "HPT" se aceptan salvo que el propio Mecanico sea uno de los
    # nombres placeholder conocidos (no hay forma de validar por RUT en esa tabla).
    # 29/07/2026 — si el Mecanico calza por NOMBRE (por tokens, tolerando
    # variantes de escritura/orden — ver _match_nomina_tecnico) con algun
    # tecnico de NOMINA_TECNICOS, la fila se acepta SIEMPRE, sin importar el
    # RUT — a pedido explicito de Cristobal ("no tomemos en cuenta los RUT,
    # guiemonos solamente por el nombre").
    _en_nomina = df["Mecanico"].apply(
        lambda m: _match_nomina_tecnico(m)[0] is not None
    )

    _es_pmm = df["_Origen"] == "PMM"
    _rut_ok = (
        df["Rut Mecanico"].str.len() > 0
    ) & (~df["Rut Mecanico"].str.upper().isin(["SIN RUT", "NAN", "NONE"])) & (
        df["Rut Mecanico"].str.contains(r"\d", regex=True, na=False)
    )
    _placeholder_nombre = df["Mecanico"].str.contains(
        r"SIN COMISION|SIN RUT", regex=True, na=False
    )
    _valido = _en_nomina | (_es_pmm & _rut_ok) | ((~_es_pmm) & (~_placeholder_nombre))
    df = df[
        (df["Mecanico"].str.len() > 0)
        & (~df["Mecanico"].str.upper().isin(["NAN", "NONE"]))
        & _valido
    ]
    if df.empty:
        log("Produccion Tecnicos: sin filas con Mecanico/RUT validos.")
        return {}

    # 29/07/2026 — HOMOLOGACION por nombre: si el Mecanico calza (por tokens)
    # con un tecnico de la nomina, se reemplaza por el nombre CANONICO de la
    # nomina antes de agrupar. Sin esto, el ejemplo real de Cristobal ("Luis
    # Riquelme" / "Luis  Riquelme" / "Luis Riquelme Q" — 3 escrituras para la
    # misma persona) quedaria como 3 "tecnicos" distintos en el resumen, cada
    # uno con solo una fraccion de las horas reales — con la homologacion, las
    # 3 variantes se fusionan bajo un unico nombre y sus horas se suman juntas.
    _nombre_homologado = df["Mecanico"].apply(lambda m: _match_nomina_tecnico(m)[0])
    _n_homologados = int((_nombre_homologado.notna() & (_nombre_homologado != df["Mecanico"])).sum())
    df["Mecanico"] = _nombre_homologado.where(_nombre_homologado.notna(), df["Mecanico"])
    if _n_homologados:
        log(f"Produccion Tecnicos: {_n_homologados:,} fila(s) homologadas a un nombre "
            f"canonico de la nomina (fusiona variantes de escritura de la misma persona).")

    n_mecanicos_crudo = df["Mecanico"].nunique()
    _n_pmm = int((df["_Origen"] == "PMM").sum())
    _n_hpt = int((df["_Origen"] == "HPT").sum())
    _n_por_nomina = int(_en_nomina.reindex(df.index, fill_value=False).sum())
    log(f"Produccion Tecnicos: tras filtro de validez -> {_n_pmm:,} fila(s) de "
        f"Tmp_ProduccionMensualMecanicos + {_n_hpt:,} fila(s) de Tmp_HorasPorTecnico "
        f"({_n_por_nomina:,} de ellas aceptadas por calzar con la nomina de tecnicos "
        f"por nombre, sin importar su RUT).")

    # Solo Mano de Obra (producto empieza con "MO") cuenta como horas — el
    # resto (repuestos/insumos) queda excluido del calculo de horas/OT.
    df_mo = df[df["producto"].apply(_es_mano_de_obra)].copy()
    if df_mo.empty:
        log("Produccion Tecnicos: sin filas de Mano de Obra (producto 'MO...') tras filtrar.")
        return {}

    # Diagnostico 30/07/2026: al pasar de la whitelist exacta al prefijo abierto,
    # dejar visible en el log que codigos entraron al calculo que ANTES no contaban.
    _cods_mo = sorted(
        str(_c).strip().upper()
        for _c in df_mo["producto"].dropna().unique()
    )
    _cods_nuevos = [_c for _c in _cods_mo if _c not in _CODIGOS_MANO_DE_OBRA]
    if _cods_nuevos:
        log(f"(i) Produccion Tecnicos: {len(_cods_nuevos)} codigo(s) de Mano de Obra entraron al "
            f"calculo que NO estaban en la whitelist anterior de 32 codigos (el prefijo abierto "
            f"'MO' se activo el 30/07/2026): {', '.join(_cods_nuevos[:25])}"
            + (f" (+{len(_cods_nuevos)-25} mas)" if len(_cods_nuevos) > 25 else ""))

    # Mes (YYYY-MM) desde la Fecha del documento. Las filas de origen "HPT"
    # (Tmp_HorasPorTecnico) nunca traen Fecha (columna NULL a proposito, ver
    # comentario de mas arriba) — quedan bajo el bucket especial "Sin fecha"
    # del selector de Mes.
    #
    # REVERTIDO 23/07/2026 (ronda 2): en la primera ronda de hoy se habia
    # cambiado esto a "asumir mes en curso" (fillna con el mes actual) a
    # pedido de Cristobal, para que un tecnico recien mapeado a su sucursal no
    # pareciera "sin datos" en el mes actual. Pero el reporte real (Hector
    # Andrade con 30.081,7 "horas facturadas") mostro que el bucket "Sin
    # fecha" de HPT trae datos sucios en cantidad (una sola fila puede traer
    # decenas de horas sin sentido) — al fusionarlo con el mes en curso, esa
    # basura contaminaba SIEMPRE la vista principal (el mes que la gente mira
    # a diario), y de paso hacia aparecer tecnicos que ya no trabajan en la
    # empresa (su ultimo registro sin fecha quedaba "hoy" para siempre) y OTs
    # que no corresponden al periodo actual. Se revierte a mantener "Sin
    # fecha" como bucket SEPARADO (hay que elegirlo a mano en el selector de
    # Mes) — mas seguro: la vista de "mes en curso" ya no se ensucia sola.
    _fechas = pd.to_datetime(df_mo.get("Fecha"), errors="coerce")
    df_mo["_mes"] = _fechas.dt.strftime("%Y-%m")
    df_mo["_mes"] = df_mo["_mes"].fillna("Sin fecha")

    # Sucursal del mecanico: prioridad al nombre real de sucursal del mecanico
    # (tabla Tmp_ProduccionMensualMecanicos, origen "PMM"); si viene vacio, usa la
    # columna Sucursal (local del documento) como respaldo.
    _suc_mec = df_mo.get("Nombre Sucursal Mecanico", "").fillna("").astype(str).str.strip()
    _suc_doc = df_mo.get("Sucursal", "").fillna("").astype(str).str.strip()
    df_mo["_suc"] = _suc_mec.where(_suc_mec.str.len() > 0, _suc_doc)

    # 29/07/2026 — respaldo por NOMINA (ver NOMINA_TECNICOS/_match_nomina_tecnico
    # arriba): si BDFlexline no trajo sucursal para esta fila (tipico de origen
    # "HPT", que nunca la registra), pero el Mecanico calza por nombre (por
    # tokens, tolerando variantes de escritura) con algun tecnico de la nomina,
    # se usa su "Lugar de Trabajo" real — se evalua ANTES del cruce por nombre
    # dentro del propio dataset y antes del mapeo manual de Admin, porque es la
    # fuente mas confiable y curada a mano por Cristobal.
    #
    # 30/07/2026 — AMPLIADO a pedido de Cristobal ("hay tecnicos que no me aparecen
    # por sucursal"): la sucursal de la nomina ahora MANDA SIEMPRE, no solo cuando
    # BDFlexline vino vacia. Motivo (confirmado con datos reales de la corrida del
    # 30/07): BDFlexline devuelve para varios tecnicos una sucursal que NO existe en
    # el Planificador — codigos crudos tipo "SUC070"/"SUC130" o nombres de otra
    # operacion como "LA FLORIDA" — y como el Planificador filtra por nombre exacto
    # de sucursal, esos tecnicos quedaban invisibles en TODAS las sucursales aunque
    # sus horas se calcularan bien. Casos reales detectados:
    #   MARCELO AGUILAR (nomina LINDEROS)            -> BDFlexline "SUC070"
    #   NICOLAS MEDINA (nomina RANCAGUA)             -> BDFlexline "SUC130"
    #   CRISTIAN MONTECINOS (nomina RANCAGUA)        -> BDFlexline "SUC070" Y "SUC130"
    #                                                   (aparecia partido en 2 filas)
    #   ESTEBAN MARTINEZ (nomina LINDEROS)           -> BDFlexline "LA FLORIDA"
    #   JUAN PABLO BIZARRO (nomina CHILLAN)          -> BDFlexline "TALCA"
    # La nomina (Lugar de Trabajo, curada a mano por Cristobal) es la fuente de
    # verdad de a que sucursal pertenece cada tecnico. Los 3 respaldos de mas abajo
    # (cruce por nombre dentro del dataset y mapeo manual de Admin) siguen vigentes
    # para los mecanicos que NO estan en la nomina.
    _suc_nomina = df_mo["Mecanico"].apply(lambda m: _match_nomina_tecnico(m)[1] or "")
    _tiene_nomina = _suc_nomina.str.len() > 0
    _reasignados = _tiene_nomina & (df_mo["_suc"].str.len() > 0) & (df_mo["_suc"] != _suc_nomina)
    if _reasignados.any():
        _det = (
            df_mo.loc[_reasignados, ["Mecanico", "_suc"]]
            .assign(_nueva=_suc_nomina[_reasignados])
            .drop_duplicates()
        )
        log(f"(i) Produccion Tecnicos: {int(_reasignados.sum()):,} fila(s) cambiaron de sucursal "
            f"a la de la nomina (la sucursal que traia BDFlexline no coincidia). "
            f"{len(_det)} combinacion(es) distintas, muestra (hasta 15):")
        for _, _r in _det.head(15).iterrows():
            log(f"    {_r['Mecanico']}: BDFlexline '{_r['_suc']}' -> nomina '{_r['_nueva']}'")
    _n_resueltos_nomina = int((_tiene_nomina & (df_mo["_suc"].str.len() == 0)).sum())
    df_mo.loc[_tiene_nomina, "_suc"] = _suc_nomina[_tiene_nomina]
    if _n_resueltos_nomina:
        log(f"Produccion Tecnicos: {_n_resueltos_nomina} fila(s) sin sucursal resueltas por "
            f"nombre contra la nomina de tecnicos (Lugar de Trabajo real).")

    # FIX 21/07/2026 (parte 5, mismo reporte de Cristobal de tecnicos faltantes por
    # sucursal): las filas de origen "HPT" (Tmp_HorasPorTecnico) no traen NI Nombre
    # Sucursal Mecanico NI Sucursal(local) — el SQL las deja NULL en ambas columnas (esa
    # tabla no registra la sucursal de la transaccion), asi que el "respaldo" de arriba
    # nunca alcanzaba a esas filas — quedaban con _suc="" y jamas calzaban con ninguna
    # sucursal real seleccionada en el Planificador (normSuc('')!==normSuc(cualquier
    # sucursal)), por mas que ya pasaran el filtro de RUT (fix de la parte 4). Como
    # ultimo respaldo: si el MISMO tecnico (por nombre) aparece en alguna fila de origen
    # "PMM" con una sucursal real, se usa esa. Si el tecnico SOLO aparece en HPT (nunca
    # en PMM), sigue sin poder ubicarse por sucursal — limitacion real del origen de
    # datos (falta esa columna en Tmp_HorasPorTecnico), no de este filtro.
    _mapa_mec_suc = (
        df_mo[df_mo["_suc"].str.len() > 0]
        .groupby("Mecanico")["_suc"]
        .agg(lambda s: s.value_counts().idxmax())
        .to_dict()
    )
    _sin_suc = df_mo["_suc"].str.len() == 0
    if _sin_suc.any():
        df_mo.loc[_sin_suc, "_suc"] = df_mo.loc[_sin_suc, "Mecanico"].map(_mapa_mec_suc).fillna("")

    # ULTIMO respaldo (23/07/2026, a pedido de Cristobal): mapeo manual
    # tecnico->sucursal, editable desde Admin -> Tecnicos, para los tecnicos que
    # ni siquiera el cruce por nombre de arriba pudo ubicar (existen SOLO en
    # Tmp_HorasPorTecnico, nunca en Tmp_ProduccionMensualMecanicos — limitacion
    # real del origen de datos). Confirmado con Cristobal via AskUserQuestion:
    # se prefiere este mapeo manual en vez de (o ademas de) seguir investigando
    # si Tmp_HorasPorTecnico tiene una columna de sucursal oculta.
    _sin_suc2 = df_mo["_suc"].str.len() == 0
    _n_resueltos_manual = 0
    if _sin_suc2.any():
        _, _mapa_manual = _leer_json_github_simple(GITHUB_TECNICOS_SUCURSAL_MANUAL)
        if isinstance(_mapa_manual, dict) and _mapa_manual:
            _mapa_manual_norm = {str(k).strip().upper(): str(v).strip() for k, v in _mapa_manual.items()}
            _asignado_manual = df_mo.loc[_sin_suc2, "Mecanico"].map(_mapa_manual_norm).fillna("")
            _n_resueltos_manual = int((_asignado_manual.str.len() > 0).sum())
            df_mo.loc[_sin_suc2, "_suc"] = _asignado_manual
            if _n_resueltos_manual:
                log(f"Produccion Tecnicos: {_n_resueltos_manual} fila(s) resueltas por el "
                    f"mapeo manual tecnico->sucursal ({os.path.basename(GITHUB_TECNICOS_SUCURSAL_MANUAL)}).")

    _mecanicos_sin_suc = sorted(df_mo.loc[df_mo["_suc"].str.len() == 0, "Mecanico"].unique().tolist())
    if _mecanicos_sin_suc:
        log(f"(!) Produccion Tecnicos: {len(_mecanicos_sin_suc)} tecnico(s) sin sucursal "
            f"identificable (no aparecen en ninguna sucursal del Planificador — solo tienen "
            f"filas de Tmp_HorasPorTecnico, que no registra sucursal, nunca aparecen en "
            f"Tmp_ProduccionMensualMecanicos para poder cruzar, y tampoco tienen un mapeo manual "
            f"guardado en Admin -> Tecnicos): "
            f"{', '.join(_mecanicos_sin_suc[:15])}"
            + (f" (+{len(_mecanicos_sin_suc)-15} mas)" if len(_mecanicos_sin_suc) > 15 else "")
            + " — se pueden asignar a mano desde 🛡️ Admin -> 🔧 Técnicos -> Sucursal manual "
              "para Produccion Tecnicos.")

    df_mo["Total Horas"] = pd.to_numeric(df_mo.get("Total Horas"), errors="coerce").fillna(0.0)
    df_mo["Nro OT"] = df_mo.get("Nro OT", "").fillna("").astype(str).str.strip()
    df_mo["Precio Lista"] = pd.to_numeric(df_mo.get("Precio Lista"), errors="coerce").fillna(0.0)
    df_mo["Comi_Vta"] = pd.to_numeric(df_mo.get("Comi_Vta"), errors="coerce").fillna(0.0)

    # FIX 23/07/2026 (ronda 2): se EXCLUIAN del calculo las filas con mas de 24 horas
    # en una sola linea, por considerarse dato sucio.
    #
    # 30/07/2026 — Cristobal pidio SACAR ese tope: ahora se cuenta la cantidad de horas
    # tal cual venga de BDFlexline, sin importar cuanto sea. El aviso del log SE
    # MANTIENE como diagnostico (para poder detectar a simple vista si alguna fila
    # disparatada esta inflando el total de un tecnico), pero ya NO filtra nada.
    # 30/07/2026 (parte 3) — CORRECCION DE VALORES INVERTIDOS, a pedido de Cristobal:
    # "malamente la gente introduce en la cantidad el precio, y el precio en la
    # cantidad. La solucion que propongo es que en esos casos particulares aplique la
    # teoria inversa". Es decir: hay filas donde "Total Horas" trae el MONTO y
    # "Precio Lista" trae las HORAS — se detectan y se dan vuelta.
    #
    # Criterio (deliberadamente conservador, tienen que cumplirse las 2 condiciones):
    #   a) "Total Horas" > 100  -> imposible como horas de una sola linea (las lineas
    #      legitimas mas largas vistas en datos reales llegan a 40 h).
    #   b) 0 < "Precio Lista" <= 100 -> ese valor SI es plausible como horas, y a la
    #      vez es absurdo como precio (ningun repuesto/servicio vale $100 o menos).
    # Si solo se cumple (a) — ej. 40.000 h con un precio real de $250.000 — NO se
    # toca nada: no hay de donde recuperar las horas reales, y se listan aparte en el
    # log para revisarlas en Flexline.
    _UMBRAL_INVERSION = 100.0
    _pl = df_mo["Precio Lista"]
    _invertidas = (df_mo["Total Horas"] > _UMBRAL_INVERSION) & (_pl > 0) & (_pl <= _UMBRAL_INVERSION)
    if _invertidas.any():
        _muestra_inv = (
            df_mo.loc[_invertidas, ["Mecanico", "Nro OT", "producto", "Total Horas", "Precio Lista", "_Origen", "_mes"]]
            .sort_values("Total Horas", ascending=False)
        )
        log(f"(i) Produccion Tecnicos: {int(_invertidas.sum())} fila(s) con cantidad y precio "
            f"INVERTIDOS (el monto quedo cargado en 'Total Horas' y las horas en 'Precio Lista') "
            f"— se dan vuelta automaticamente. Muestra (hasta 15):")
        for _, _r in _muestra_inv.head(15).iterrows():
            log(f"    {_r.get('Mecanico','')} | OT {_r.get('Nro OT','')} | producto "
                f"'{_r.get('producto','')}' | {_r.get('Total Horas',0):,.1f} h -> "
                f"{_r.get('Precio Lista',0):,.1f} h (precio ${_r.get('Total Horas',0):,.0f}) | "
                f"origen {_r.get('_Origen','')} | mes {_r.get('_mes','')}")
        _horas_ok = df_mo.loc[_invertidas, "Precio Lista"].copy()
        _precio_ok = df_mo.loc[_invertidas, "Total Horas"].copy()
        df_mo.loc[_invertidas, "Total Horas"] = _horas_ok
        df_mo.loc[_invertidas, "Precio Lista"] = _precio_ok

    _horas_sospechosas = df_mo[df_mo["Total Horas"] > 24].sort_values("Total Horas", ascending=False)
    if not _horas_sospechosas.empty:
        log(f"(i) Produccion Tecnicos: {len(_horas_sospechosas)} fila(s) con 'Total Horas' > 24 "
            f"en una sola linea — SI se estan contando (el tope de 24 h se quito el 30/07/2026 "
            f"a pedido de Cristobal). Se listan las 25 mas grandes, de mayor a menor:")
        for _, _r in _horas_sospechosas.head(25).iterrows():
            log(f"    {_r.get('Mecanico','')} | OT {_r.get('Nro OT','')} | producto "
                f"'{_r.get('producto','')}' | {_r.get('Total Horas',0):.1f} h | origen "
                f"{_r.get('_Origen','')} | mes {_r.get('_mes','')}")

    def _meses_desde_hoy(n):
        _hoy = datetime.now()
        out = set()
        for _i in range(n):
            _y, _m = _hoy.year, _hoy.month - _i
            while _m <= 0:
                _m += 12
                _y -= 1
            out.add(f"{_y:04d}-{_m:02d}")
        return out

    # --- 1) Resumen: mecanico x sucursal x mes (tabla principal + selector) --
    _meses_resumen = _meses_desde_hoy(PRODUCCION_MESES_HISTORIAL)
    df_resumen = df_mo[df_mo["_mes"].isin(_meses_resumen) | (df_mo["_mes"] == "Sin fecha")]
    resumen = []
    if not df_resumen.empty:
        _ag = (
            df_resumen.groupby(["Mecanico", "_suc", "_mes"], dropna=False)
            .agg(total_horas=("Total Horas", "sum"),
                 n_ot=("Nro OT", lambda s: s[s.str.len() > 0].nunique()))
            .reset_index()
        )
        for _, row in _ag.iterrows():
            resumen.append({
                "mecanico":          row["Mecanico"],
                "sucursal_mecanico": row["_suc"],
                "mes":               row["_mes"],
                "total_horas":       round(float(row["total_horas"]), 2),
                "n_ot":              int(row["n_ot"]),
            })

    # --- 2) Detalle por producto: mecanico x mes x producto (drill-down) -----
    # Mas granular -> se limita a los ultimos PRODUCCION_DETALLE_MESES para no
    # inflar el archivo (el resumen de arriba ya cubre el historial completo).
    _meses_detalle = _meses_desde_hoy(PRODUCCION_DETALLE_MESES)
    df_detalle = df_mo[df_mo["_mes"].isin(_meses_detalle)]
    detalle_producto = []
    if not df_detalle.empty:
        _agd = (
            df_detalle.groupby(["Mecanico", "_mes", "producto"], dropna=False)
            .agg(horas=("Total Horas", "sum"),
                 cantidad=("Nro OT", lambda s: s[s.str.len() > 0].nunique()))
            .reset_index()
        )
        for _, row in _agd.iterrows():
            detalle_producto.append({
                "mecanico":  row["Mecanico"],
                "mes":       row["_mes"],
                "producto":  row["producto"],
                "horas":     round(float(row["horas"]), 2),
                "cantidad":  int(row["cantidad"]),
            })

    # --- 3) Detalle por OT: mecanico x mes x Nro OT x producto (drill-down) --
    # A pedido de Cristobal (21/07/2026, con captura de ejemplo de Nº OT / Producto /
    # Precio Lista / Total Horas / Comi_Vta): el detalle por producto de arriba ya suma
    # las horas de un mismo producto dentro del mes, perdiendo el detalle linea por linea
    # que se necesita ver por tecnico (incluye lineas repetidas del mismo OT+producto con
    # horas/comision distintas — por eso NO se agrupan, se listan tal cual vienen de
    # BDFlexline). Mismo recorte de meses que detalle_producto (PRODUCCION_DETALLE_MESES)
    # para no inflar el archivo — el resumen de arriba ya cubre el historial completo.
    detalle_ot = []
    if not df_detalle.empty:
        _dfo = df_detalle[[
            "Mecanico", "_mes", "Fecha", "Nro OT", "producto",
            "Precio Lista", "Total Horas", "Comi_Vta",
        ]].copy()
        _dfo["_fecha_dt"] = pd.to_datetime(_dfo["Fecha"], errors="coerce")
        _dfo = _dfo.sort_values(["Mecanico", "_fecha_dt"], ascending=[True, False])
        for _, row in _dfo.iterrows():
            _fecha_dt = row["_fecha_dt"]
            detalle_ot.append({
                "mecanico":     row["Mecanico"],
                "mes":          row["_mes"],
                "fecha":        _fecha_dt.strftime("%d/%m/%Y") if pd.notna(_fecha_dt) else "",
                "nro_ot":       row["Nro OT"],
                "producto":     row["producto"],
                "precio_lista": round(float(row["Precio Lista"]), 0),
                "horas":        round(float(row["Total Horas"]), 2),
                "comi_vta":     round(float(row["Comi_Vta"]), 0),
            })

    n_mecanicos_mo = df_mo["Mecanico"].nunique()
    log(f"Produccion Tecnicos: {filas_leidas:,} fila(s) leidas de BDFlexline "
        f"({n_mecanicos_crudo} mecanico(s) con RUT valido) -> {len(df_mo):,} fila(s) de "
        f"Mano de Obra ({n_mecanicos_mo} mecanico(s)) -> {len(resumen):,} registro(s) de "
        f"resumen (ultimos {PRODUCCION_MESES_HISTORIAL} meses) + {len(detalle_producto):,} "
        f"registro(s) de detalle por producto + {len(detalle_ot):,} registro(s) de detalle "
        f"por OT (ambos ultimos {PRODUCCION_DETALLE_MESES} meses).")
    return {"resumen": resumen, "detalle_producto": detalle_producto, "detalle_ot": detalle_ot}


def _descargar_citas_dia(ses, nombre_suc, id_cons, fecha_str):
    """Descarga y parsea un solo dia/sucursal (con reintentos/backoff/re-login).
    Pensado para correr en un hilo."""
    html_dia = _http_get_agenda(
        ses, AGENDA_REPORT_URL,
        params={"fecha": fecha_str, "id_cons": id_cons, "id_tipo_serv": "-1"},
        timeout=40,
    )
    if html_dia is None:
        return nombre_suc, fecha_str, []
    return nombre_suc, fecha_str, _parsear_citas_html(html_dia, fecha_str, nombre_suc)


def backfill_ingresos_historicos(dias=60, hilos=3):
    """
    Backfill UNICA VEZ: recorre agenda.curifor.cl dia por dia (hoy-1 hasta hoy-dias) para
    cada sucursal configurada, y agrega a control_taller.json (ordenes) las citas marcadas
    con el icono Ticket (ingresado=True) que aun no existan (por patente), para que la vista
    "Vehiculos en Taller" tenga cobertura real de los ultimos ~60 dias desde el arranque.

    Las descargas se hacen en paralelo (varios hilos a la vez, no las ~540 una por una) para
    que la primera corrida no demore tanto. El merge de resultados en control_taller.json se
    hace despues, en un solo hilo, para no tener condiciones de carrera.

    IMPORTANTE (detectado 01/07/2026): usar UNA sola sesion de requests compartida entre
    muchos hilos en paralelo hace que agenda.curifor.cl encole/serialice las peticiones de
    esa misma sesion, y con 20 hilos a la vez las respuestas tardan mas de 20s y todas dan
    timeout (confirmado: las ~540 consultas de una corrida real dieron 100% timeout, incluso
    fechas de 1-2 dias atras). Por eso ahora se abre UNA sesion (login) independiente por
    cada hilo, y cada tarea se asigna siempre al mismo hilo/sesion (round-robin), evitando
    que dos peticiones compartan sesion al mismo tiempo.

    ACTUALIZACION 07/07/2026: hilos bajado a 3 (parametro validado por el equipo en
    generar_csv_agenda.py, que si logro leer fechas historicas con esa concurrencia) y
    cada request ahora usa _http_get_agenda (reintentos con backoff + re-login).

    Es idempotente: si control_taller.json ya tiene la marca '_backfill_ingresos_2meses',
    no vuelve a correr (para no repetir cientos de consultas cada vez que se corre el BAT).
    Para forzar un nuevo backfill, borra esa clave del JSON en GitHub.
    """
    if not all([GITHUB_USUARIO, GITHUB_REPO, GITHUB_TOKEN]):
        log("(!) GitHub no configurado — omitiendo backfill de ingresos historicos")
        return

    sucursales_con_id = [(n, v) for n, v in SUCURSALES_AGENDA if v is not None]
    if not sucursales_con_id:
        log("(!) SUCURSALES_AGENDA sin id_cons configurados — omitiendo backfill")
        return
    # Nombres de sucursal unicos (sin duplicar cuando 2 id_cons comparten el mismo nombre
    # de salida, ej. "TALCA" con Ford + BMW) — para todo lo que agrupa por sucursal.
    nombres_unicos = []
    for _n, _ in sucursales_con_id:
        if _n not in nombres_unicos:
            nombres_unicos.append(_n)

    sha, ctrl_data = _leer_json_github_simple(GITHUB_CTRL_TALLER)
    if ctrl_data.get("_backfill_ingresos_2meses"):
        log(f"Backfill de ingresos historicos ya se hizo el {ctrl_data['_backfill_ingresos_2meses']} — se omite.")
        return

    # Una sesion (login) independiente por hilo — evita que agenda.curifor.cl encole/serialice
    # peticiones concurrentes de una misma sesion compartida (ver nota en el docstring).
    sesiones = []
    for _i in range(hilos):
        _s = _login_agenda()
        if _s is not None:
            sesiones.append(_s)
    if not sesiones:
        log("(!) No se pudo autenticar en agenda.curifor.cl — se omite backfill")
        return
    if len(sesiones) < hilos:
        log(f"(!) Solo se pudieron abrir {len(sesiones)}/{hilos} sesiones — se continua con esas.")

    from datetime import timedelta as _td
    from concurrent.futures import ThreadPoolExecutor, as_completed
    hoy = datetime.now()
    fechas = [(hoy - _td(days=i)).strftime("%d/%m/%Y") for i in range(1, dias + 1)]

    for nombre_suc in nombres_unicos:
        if nombre_suc not in ctrl_data:
            ctrl_data[nombre_suc] = {"tecnicos": [], "ordenes": [], "bloques": {}}
        ctrl_data[nombre_suc].setdefault("ordenes", [])
        ctrl_data[nombre_suc].setdefault("tecnicos", [])
        ctrl_data[nombre_suc].setdefault("bloques", {})

    patentes_existentes = {
        suc: {(o.get("patente") or "").strip().upper() for o in ctrl_data[suc]["ordenes"]}
        for suc in nombres_unicos
    }
    agregadas_por_suc = {suc: 0 for suc in nombres_unicos}
    nuevas_ordenes_por_suc = {suc: [] for suc in nombres_unicos}  # copia para re-merge con datos frescos
    total_agregadas = 0

    tareas = [
        (nombre_suc, id_cons, fecha_str)
        for nombre_suc, id_cons in sucursales_con_id
        for fecha_str in fechas
    ]
    n_ses = len(sesiones)
    log(f"Backfill: descargando {len(tareas)} reportes ({len(nombres_unicos)} sucursales x {dias} dias) "
        f"con {n_ses} sesion(es) independiente(s) en paralelo...")

    completadas = 0
    fechas_devueltas_distintas = set()  # diagnostico: cuantas fechas reales distintas devolvio el sitio
    with ThreadPoolExecutor(max_workers=n_ses) as executor:
        futuros = [
            executor.submit(_descargar_citas_dia, sesiones[i % n_ses], nombre_suc, id_cons, fecha_str)
            for i, (nombre_suc, id_cons, fecha_str) in enumerate(tareas)
        ]
        for fut in as_completed(futuros):
            nombre_suc, fecha_solicitada, citas = fut.result()
            completadas += 1

            for c in citas:
                if not c.get("ingresado"):
                    continue
                patente = _sanear_patente(c.get("patente"))
                if not patente or RE_PATENTE_EXCLUIDA.match(patente):
                    continue
                # Fecha real de la cita segun la agenda (no la que pedimos) — asi detectamos
                # si el sitio esta devolviendo el dia solicitado o siempre el mismo (ej. hoy).
                fecha_real = (c.get("fecha") or "").strip() or fecha_solicitada
                fechas_devueltas_distintas.add(fecha_real)
                if patente in patentes_existentes[nombre_suc]:
                    continue
                patentes_existentes[nombre_suc].add(patente)
                orden_id = f"bf{nombre_suc[:3]}{patente}{int(datetime.now().timestamp()*1000)}"
                _nueva_orden = {
                    "id": orden_id, "patente": patente,
                    "cliente": c.get("nombre", ""), "modelo": c.get("modelo", ""),
                    "ot": c.get("oc", ""), "km": c.get("km", ""), "asesor": c.get("asesor", ""),
                    "ingreso": _fecha_ddmmyyyy_a_iso(fecha_real), "salida": "",
                    "tecnico": None, "esp": False, "tra": False, "lav": False, "patio": False,
                    "comentarios": c.get("servicio") or c.get("mantencion") or "",
                    "tipo": _detect_tipo_backfill(c.get("servicio"), c.get("mantencion")),
                    "etapa": "recepcion", "stop": None,
                    "comentario2": "", "numero_caso": "", "n_pedido": "", "eta": "", "auto_reemplazo": "",
                    "servicio": c.get("servicio") or c.get("mantencion") or "",
                    "hora_rec": c.get("horario", ""), "mantencion": c.get("mantencion", ""),
                    "hora_compromiso": "", "duracion_min": 60,
                    "horas_tempario": c.get("horas_tempario"),
                    "cerrada": False, "fecha_cierre": "",
                }
                ctrl_data[nombre_suc]["ordenes"].append(_nueva_orden)
                nuevas_ordenes_por_suc[nombre_suc].append(_nueva_orden)
                agregadas_por_suc[nombre_suc] += 1
                total_agregadas += 1

            if completadas % 20 == 0 or completadas == len(tareas):
                log(f"  ... {completadas}/{len(tareas)} reportes descargados, "
                    f"{total_agregadas} ingreso(s) nuevo(s) hasta ahora")

    if len(fechas_devueltas_distintas) <= 1 and len(tareas) > len(sucursales_con_id):
        log(f"(!) AVISO: la agenda devolvio solo {len(fechas_devueltas_distintas)} fecha(s) distinta(s) "
            f"({', '.join(sorted(fechas_devueltas_distintas)) or 'ninguna'}) para {len(tareas)} consultas de "
            f"{dias} dias distintos — probable que el reporte de agenda.curifor.cl no soporte fechas "
            f"pasadas y siempre devuelva el dia actual, no un backfill historico real.")
    else:
        log(f"Diagnostico: se detectaron {len(fechas_devueltas_distintas)} fecha(s) distinta(s) en los "
            f"resultados — la cobertura historica parece real.")

    for nombre_suc in nombres_unicos:
        log(f"  {nombre_suc}: {agregadas_por_suc[nombre_suc]} ingreso(s) historico(s) agregado(s)")

    # El backfill puede tardar varios minutos (540 consultas). Para no perder el trabajo si
    # alguien edito control_taller.json desde la app mientras corria (SHA quedaria vieja y
    # GitHub rechazaria el PUT), se relee el archivo FRESCO justo antes de subir y se
    # mezclan ahi las ordenes nuevas encontradas (dedup por patente), en vez de subir la
    # copia vieja que se tenia al empezar.
    sha_fresco, ctrl_fresco = _leer_json_github_simple(GITHUB_CTRL_TALLER)
    if sha_fresco is None:
        log("(!) No se pudo releer control_taller.json fresco — se sube con los datos originales.")
        sha_fresco, ctrl_fresco = sha, ctrl_data
    else:
        for nombre_suc in nombres_unicos:
            if nombre_suc not in ctrl_fresco:
                ctrl_fresco[nombre_suc] = {"tecnicos": [], "ordenes": [], "bloques": {}}
            ctrl_fresco[nombre_suc].setdefault("ordenes", [])
            ctrl_fresco[nombre_suc].setdefault("tecnicos", [])
            ctrl_fresco[nombre_suc].setdefault("bloques", {})
            _patentes_frescas = {(o.get("patente") or "").strip().upper() for o in ctrl_fresco[nombre_suc]["ordenes"]}
            for _orden in nuevas_ordenes_por_suc[nombre_suc]:
                if _orden["patente"] not in _patentes_frescas:
                    ctrl_fresco[nombre_suc]["ordenes"].append(_orden)
                    _patentes_frescas.add(_orden["patente"])

    ctrl_fresco["_backfill_ingresos_2meses"] = datetime.now().strftime("%d/%m/%Y %H:%M")
    ok = _subir_json_github_simple(
        GITHUB_CTRL_TALLER, ctrl_fresco, sha_fresco,
        f"Backfill ingresos historicos ({dias}d) — {datetime.now().strftime('%d/%m/%Y %H:%M')}",
    )
    if ok:
        log(f"Backfill completo: {total_agregadas} vehiculo(s) historico(s) agregados a control_taller.json")
    else:
        log("(!) Error al subir control_taller.json con el backfill (revisar conexion/GitHub).")


def _subir_agenda_github(agenda_data):
    """Sube agenda_hoy.json al repositorio GitHub via API REST."""
    if not all([GITHUB_USUARIO, GITHUB_REPO, GITHUB_TOKEN]):
        return
    try:
        url     = f"https://api.github.com/repos/{GITHUB_USUARIO}/{GITHUB_REPO}/contents/{GITHUB_AGENDA}"
        headers = {"Authorization": f"token {GITHUB_TOKEN}", "Accept": "application/vnd.github.v3+json"}
        sha = None
        try:
            rg = requests.get(url, headers=headers, timeout=15, verify=False)
            if rg.status_code == 200:
                sha = rg.json()["sha"]
        except Exception:
            pass
        content_b64 = base64.b64encode(
            json.dumps(agenda_data, ensure_ascii=False, indent=2).encode("utf-8")
        ).decode()
        payload = {
            "message": f"Agenda Curifor {datetime.now().strftime('%d/%m/%Y %H:%M')}",
            "content": content_b64,
        }
        if sha:
            payload["sha"] = sha
        rp = requests.put(url, headers=headers, json=payload, timeout=30, verify=False)
        if rp.status_code in [200, 201]:
            log(f"agenda_hoy.json subido a GitHub ({len(json.dumps(agenda_data))} bytes)")
        else:
            log(f"(!) Error al subir agenda: {rp.status_code}")
    except Exception as _e:
        log(f"(!) No se pudo subir agenda: {_e}")


# =============================================================
#   PROCESO PRINCIPAL
# =============================================================

def main():

    print()
    print("=" * 55)
    print("  CONSOLIDADOR OTs — CURIFOR S.A")
    print(f"  {datetime.now().strftime('%A %d/%m/%Y  %H:%M')}")
    print("=" * 55)
    print()

    # -- PASO 1: Leer la sábana del PBI ----------------------
    print("PASO 1 — Leyendo sábana del PBI")
    ruta_pbi = encontrar_archivo_pbi()
    if not ruta_pbi:
        print(f"\n  ERROR No se encontró ningún archivo .xlsx en:\n     {CARPETA_PBI}")
        print("  -> Descarga la sábana del PBI y déjala en esa carpeta.\n")
        input("  Presiona ENTER para cerrar...")
        sys.exit(1)
    # Cuenta Ficha: los saldos de cliente y el puente patente<->RUT se leen
    # ANTES del PBI porque leer_pbi() necesita saber que patentes indexar para
    # el historial de OT (asi se arma de una pasada, sin copiar el DataFrame).
    # Son 2 archivos chicos (~3 MB) y se reutilizan despues en el PASO 6.
    mapa_patente_rut   = leer_patente_cliente()
    mapa_anticipo      = leer_anticipo_taller()
    _cf_rut_patentes, _cf_pat_ruts = mapear_patentes_cuenta_ficha(mapa_anticipo, mapa_patente_rut)

    df_pbi, df_pbi_completo, mapa_ot_patente, _cf_ots = leer_pbi(
        ruta_pbi, patentes_cf=set(_cf_pat_ruts.keys())
    )
    log(f"{os.path.basename(ruta_pbi)}  ->  {len(df_pbi)} OTs (deduplicadas) | {len(df_pbi_completo)} filas totales")
    print()

    # -- PASO 2: Leer el maestro anterior --------------------
    print("PASO 2 — Leyendo maestro anterior (para preservar columnas de gestión)")
    if os.path.exists(ARCHIVO_MAESTRO):
        df_maestro_anterior = leer_maestro(ARCHIVO_MAESTRO)
        log(f"Maestro anterior  ->  {len(df_maestro_anterior)} OTs")
    else:
        df_maestro_anterior = pd.DataFrame(columns=[CLAVE])
        log("(!)  No se encontró maestro anterior — se parte desde cero")
    print()

    # -- PASO 3: (DESACTIVADO 13/07/2026) ---------------------
    # A pedido de Cristobal: los Excel de sucursal (carpeta Sucursales/) DEJARON de
    # alimentar Categoria/Observacion OT/Notas/Avance-Gestion. Motivo: llevaban semanas
    # o meses sin tocarse y, pese al fix de prioridad del 10/07/2026
    # (solo_rellenar_vacios=True), seguian generando confusion/riesgo de pisar ediciones
    # reales hechas en la app. A partir de ahora la UNICA fuente de esas 4 columnas es
    # la app web (GitHub) — el cruce de la consolidacion es solo App <-> PBI (Seguimiento
    # de la carpeta PBI). encontrar_archivo_sucursal()/leer_sucursal() y la constante
    # SUCURSALES se dejan definidas en el script por si se necesitan reactivar alguna
    # vez, pero ya no se llaman desde aqui.
    encontradas = 0
    df_sucursales_consolidado = pd.DataFrame(columns=[CLAVE] + COLUMNAS_GESTION)
    print("PASO 3 — Excel de sucursales: OMITIDO (ya no alimentan Categoria/Observacion/Notas/Avance — solo manda la App)")
    print()

    # -- PASO 4: Cruzar todo por FOLIO OT --------------------
    print("PASO 4 — Cruzando datos")

    # Descargar datos actuales de GitHub ANTES del cruce.
    # Estos contienen las ediciones hechas desde la app (CATEGORIA, NOTAS, etc.)
    # que no están en el Excel maestro ni en los archivos de sucursal.
    df_anterior_github = obtener_ots_github_actuales()
    if not df_anterior_github.empty:
        log(f"Datos de GitHub descargados: {len(df_anterior_github)} OTs "
            f"(ediciones de la app incluidas)")
    else:
        log("(i)  Sin datos previos en GitHub — se usará solo el maestro")

    df_final = df_pbi.copy()

    cols_solo_maestro = [c for c in df_maestro_anterior.columns
                         if c not in df_final.columns and c != CLAVE]
    if cols_solo_maestro and not df_maestro_anterior.empty:
        df_extras = df_maestro_anterior[[CLAVE] + cols_solo_maestro].copy()
        df_final = df_final.merge(df_extras, on=CLAVE, how="left")
        log(f"Columnas del maestro recuperadas: {len(cols_solo_maestro)}")

    for col in COLUMNAS_GESTION:
        if col not in df_final.columns:
            df_final[col] = pd.NA
        else:
            df_final[col] = pd.NA

    # Orden de prioridad — ACTUALIZADO 13/07/2026: los Excel de sucursal ya NO
    # participan de este cruce (ver nota del PASO 3). La app web (GitHub) es la
    # UNICA fuente de Categoria/Observacion OT/Notas/Avance-Gestion de aqui en
    # adelante — el cruce de la consolidacion es solo App <-> PBI (Seguimiento).
    #   1. Maestro Excel anterior  — línea base, puede estar desactualizado
    #      (sobreescribe blancos únicamente, es el punto de partida)
    #   2. GitHub (JSON publicado) — ediciones hechas desde la app web.
    #      SIEMPRE gana si tiene valor — esto es lo que garantiza que tus
    #      ediciones en la app nunca se pierdan al correr el consolidador.
    df_final = aplicar_columnas_gestion(df_final, df_maestro_anterior, "ant")
    df_final = aplicar_columnas_gestion(df_final, df_anterior_github,  "git")

    nuevas  = set(df_final[CLAVE]) - set(df_maestro_anterior[CLAVE]) if not df_maestro_anterior.empty else set()
    salidas = set(df_maestro_anterior[CLAVE]) - set(df_final[CLAVE])  if not df_maestro_anterior.empty else set()
    log(f"Total OTs en resultado:              {len(df_final)}")
    if nuevas:
        log(f"OTs nuevas (entraron hoy del PBI):   {len(nuevas)}")
    if salidas:
        log(f"OTs cerradas (salieron del PBI):     {len(salidas)}")

    df_final = agregar_info_documentos(df_final, df_pbi_completo)
    print()

    # -- PASO 5: Guardar el archivo maestro ------------------
    print("PASO 5 — Guardando")
    respaldar_maestro()
    escribir_en_maestro(df_final, df_maestro_anterior)
    log(f"Maestro actualizado -> {os.path.basename(ARCHIVO_MAESTRO)}")
    print()

    # -- PASO 6: Exportar datos para dashboard web -----------
    print("PASO 6 — Exportando datos para dashboard web")

    # Extraer detalle de repuestos del Vale de Consumo por OT
    lookup_repuestos   = cargar_listado_repuestos()
    repuestos_actuales = agregar_repuestos_detalle(df_pbi_completo, lookup_repuestos)

    # Cruzar repuestos del Seguimiento de Compras (en espera / en bodega) por OT.
    # Mapa patente -> OT(s) pendiente(s) para el puente desde OTs cerradas.
    folios_pend = set(df_final[CLAVE].astype(str))
    mapa_patente_pendiente = {}
    if "PATENTE" in df_final.columns:
        for _folio, _pat in zip(df_final[CLAVE].astype(str), df_final["PATENTE"]):
            _p = str(_pat).strip().upper()
            if _p and _p not in ("NAN", "NONE"):
                mapa_patente_pendiente.setdefault(_p, []).append(str(_folio).strip())
    mapa_stock         = leer_stock_repuestos()
    _set_stock_idx_cache(mapa_stock)  # disponible para _buscar_repuestos_pauta() en PASO 8 (agenda)
    codigos_vc_global  = leer_tabla_vc()
    # mapa_patente_rut y mapa_anticipo ya se leyeron en el PASO 1 (los necesita
    # el indice de Cuenta Ficha para saber que patentes rastrear en el PBI).
    repuestos_compras  = leer_seguimiento_compras(
        folios_pend, mapa_ot_patente, mapa_patente_pendiente, mapa_stock
    )

    ruta_json = exportar_json(df_final, repuestos_actuales, repuestos_compras, codigos_vc_global, mapa_patente_rut, mapa_anticipo)
    log(f"JSON exportado -> {os.path.basename(ruta_json)}")
    subir_a_github(ruta_json)
    print()

    # -- PASO 7: Generar ranking de cierres >90 días ---------
    print("PASO 7 — Generando ranking de cierres >90 días")
    datos_ranking = generar_ranking_cierres(ruta_pbi)
    if datos_ranking:
        subir_ranking_github(datos_ranking)
        log(f"ado correctamente en GitHub")
    else:
        log("(!) No se pudo generar el ranking de cierres")
    print()

    # -- PASO 8: Descargar agenda Curifor (proximos 3 dias) ------
    print("PASO 8 — Descargando agenda Curifor (proximos 3 dias)")
    try:
        agenda_data = leer_agenda_curifor()
        if agenda_data:
            _subir_agenda_github(agenda_data)
            total_citas = sum(
                len(citas)
                for suc in agenda_data.get("sucursales", {}).values()
                for citas in suc.values()
            )
            log(f"Agenda subida: {total_citas} cita(s)")
        else:
            log("(!) No se pudieron obtener datos de la agenda Curifor")
    except Exception as _e_ag:
        log(f"(!) Error al obtener agenda: {_e_ag}")
    print()

    # -- PASO 9: Backfill unico de ingresos historicos (~60 dias) ----
    # Se salta solo despues de la primera vez (queda marcado en control_taller.json).
    # Puede tardar varios minutos la primera vez porque consulta la agenda dia por dia.
    print("PASO 9 — Backfill historico de ingresos (Vehiculos en Taller)")
    try:
        backfill_ingresos_historicos(dias=60)
    except Exception as _e_bf:
        log(f"(!) Error en backfill de ingresos historicos: {_e_bf}")
    print()

    # -- PASO 10: Exportar catalogo completo de Stock de Repuestos --
    # Independiente de las OTs (a diferencia de repuestos_compras) — alimenta
    # la busqueda general de "stock de repuestos" del Asistente App. 08/07/2026.
    # Se sube via Git Data API (_subir_json_github_gitdata), no Contents API:
    # con este archivo (~13 MB en base64) la Contents API devolvia 403 "Timed
    # out validating rule, please try again" (timeout de GitHub validando el
    # archivo, no un error nuestro) — confirmado en ultimo_resultado.txt real.
    print("PASO 10 — Exportando catalogo completo de Stock de Repuestos")
    try:
        _stock_completo = exportar_stock_repuestos_completo()
        if _stock_completo:
            _stock_json = {
                "fecha_actualizacion": datetime.now().strftime("%d/%m/%Y %H:%M"),
                "total_productos": len(_stock_completo),
                "productos": _stock_completo,
            }
            if _subir_json_github_gitdata(
                GITHUB_STOCK_REPUESTOS, _stock_json,
                f"Stock de Repuestos actualizado {datetime.now().strftime('%d/%m/%Y %H:%M')}",
                timeout=120,
            ):
                log(f"Catálogo de Stock subido a GitHub -> {GITHUB_STOCK_REPUESTOS} "
                    f"({len(_stock_completo):,} productos)")
            else:
                log("(!) Error al subir el catálogo de Stock de Repuestos a GitHub")
    except Exception as _e_stk:
        log(f"(!) Error inesperado exportando catálogo de Stock: {_e_stk}")
    print()

    # -- PASO 11: Produccion de Tecnicos (horas facturadas, BDFlexline) --
    # Se conecta a la base SQL de la red interna (10.50.15.2) desde esta PC —
    # Streamlit Cloud nunca llega a ese servidor. Si falla (sin credenciales,
    # sin driver ODBC, sin VPN/red), no interrumpe el resto de la consolidacion.
    print("PASO 11 — Produccion de Tecnicos (horas facturadas, BDFlexline)")
    try:
        _prod_data = leer_produccion_tecnicos()
        _resumen_prod = _prod_data.get("resumen", []) if _prod_data else []
        _detalle_prod = _prod_data.get("detalle_producto", []) if _prod_data else []
        _detalle_ot_prod = _prod_data.get("detalle_ot", []) if _prod_data else []
        if _resumen_prod:
            _prod_json = {
                "fecha_actualizacion": datetime.now().strftime("%d/%m/%Y %H:%M"),
                "resumen": _resumen_prod,
                "detalle_producto": _detalle_prod,
                "detalle_ot": _detalle_ot_prod,
            }
            if _subir_json_github_gitdata(
                GITHUB_PRODUCCION_TECNICOS, _prod_json,
                f"Produccion Tecnicos actualizada {datetime.now().strftime('%d/%m/%Y %H:%M')}",
                timeout=90,
            ):
                log(f"Produccion Tecnicos subida a GitHub -> {GITHUB_PRODUCCION_TECNICOS} "
                    f"({len(_resumen_prod):,} registro(s) de resumen, {len(_detalle_prod):,} de "
                    f"detalle por producto, {len(_detalle_ot_prod):,} de detalle por OT)")
            else:
                log("(!) Error al subir Produccion Tecnicos a GitHub")
        else:
            log("(!) Sin datos de Produccion Tecnicos esta corrida (revisar credenciales/conexion/filtro Mano de Obra arriba).")
    except Exception as _e_prod:
        log(f"(!) Error inesperado en Produccion Tecnicos: {_e_prod}")
    print()

    # -- PASO 12: Revision de Campanas (Agenda Ford) --
    # Archivo local (CARPETA_AGENDA_FORD, ahora dentro de esta misma carpeta
    # del proyecto — antes vivia en la carpeta de otro proyecto, movido
    # 29/07/2026) — no depende del
    # PBI ni de GitHub para leerlo, solo se usan las sucursales reales del
    # PBI (df_final) para reconocer "por logica" a que sucursal corresponde
    # cada fila. Si el archivo no esta (aun no se genero, o el proceso que lo
    # arma no corrio hoy) no interrumpe el resto de la consolidacion. 28/07/2026.
    print("PASO 12 — Revision de Campanas (Agenda Ford)")
    try:
        _sucursales_reales_camp = sorted(
            {s for s in df_final["SUCURSAL"].dropna().astype(str).str.strip().unique() if s}
        )
        _campanas_regs, _campanas_archivo = leer_campanas_curifor(_sucursales_reales_camp)
        if _campanas_regs:
            _campanas_json = {
                "fecha_actualizacion": datetime.now().strftime("%d/%m/%Y %H:%M"),
                "archivo_origen": _campanas_archivo,
                "campanas": _campanas_regs,
            }
            _sha_camp, _ = _leer_json_github_simple(GITHUB_CAMPANAS)
            if _subir_json_github_simple(
                GITHUB_CAMPANAS, _campanas_json, _sha_camp,
                f"Revision de Campanas actualizada {datetime.now().strftime('%d/%m/%Y %H:%M')}",
            ):
                _n_rojo = sum(1 for r in _campanas_regs if r["estado_color"] == "rojo")
                _n_amar = sum(1 for r in _campanas_regs if r["estado_color"] == "amarillo")
                _n_verde = sum(1 for r in _campanas_regs if r["estado_color"] == "verde")
                log(f"Campañas subidas a GitHub -> {GITHUB_CAMPANAS} "
                    f"({len(_campanas_regs)} caso(s) desde '{_campanas_archivo}': "
                    f"🔴 {_n_rojo} no realizada(s) · 🟡 {_n_amar} no revisada(s) · "
                    f"🟢 {_n_verde} revisada(s))")
            else:
                log("(!) Error al subir Revision de Campanas a GitHub")
        elif _campanas_archivo:
            log(f"(i) '{_campanas_archivo}' leido pero sin ninguna fila con datos "
                f"en la columna W (Campañas/Boletín) — nada que subir.")
    except Exception as _e_camp:
        log(f"(!) Error inesperado en Revision de Campanas: {_e_camp}")
    print()

    # -- PASO 13: Cuenta Ficha (saldos de cliente + historial de OT) --------
    print("PASO 13 — Cuenta Ficha (saldos de cliente + historial de OT)")
    try:
        _cf_payload = generar_cuenta_ficha(mapa_anticipo, _cf_rut_patentes, _cf_ots)
        if _cf_payload:
            subir_cuenta_ficha(_cf_payload)
        else:
            log("(i) Cuenta Ficha: sin datos para subir (revisa que el .xls del "
                "Informe Ficha Cuenta este en la carpeta PBI).")
    except Exception as _e_cf:
        log(f"(!) Error inesperado en Cuenta Ficha: {_e_cf}")
    print()

    # -- RESULTADO FINAL ----------------------------------------
    print("=" * 55)
    print("  COMPLETADO  PROCESO COMPLETADO")
    print(f"  OTs actualizadas : {len(df_final)}")
    print(f"  Fuente Categoria/Observacion/Notas/Avance: solo App (GitHub) — Excel de sucursal desactivados")
    print(f"  Fecha y hora     : {datetime.now().strftime('%d/%m/%Y %H:%M')}")
    print("=" * 55)
    print()

    input("  Presiona ENTER para cerrar...")


if __name__ == "__main__":
    try:
        main()
    except Exception as e:
        print(f"\n  ERROR Error inesperado: {e}")
        import traceback
        traceback.print_exc()
        print()
        input("  Presiona ENTER para cerrar...")
        sys.exit(1)
