# -*- coding: utf-8 -*-
# =============================================================
#   DASHBOARD OTs PENDIENTES - CURIFOR S.A
#   Aplicación web Streamlit
# =============================================================

import streamlit as st
import streamlit.components.v1 as components
import pandas as pd
import requests
import json
import base64
import gzip
import hashlib
import secrets
import uuid
import io
from datetime import datetime, timedelta as _timedelta
from zoneinfo import ZoneInfo

# Capa de datos: Supabase si hay credencial, GitHub si no. Ver datos_supabase.py.
# Si el módulo no estuviera (despliegue viejo), la app sigue con GitHub.
try:
    import datos_supabase as _datos
except Exception:                                    # pragma: no cover
    class _SinSupabase:
        @staticmethod
        def disponible():
            return False
    _datos = _SinSupabase()

# PDF / chart (solo se importan cuando se generan)
def _importar_pdf_libs():
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib import colors
    from reportlab.lib.units import cm, mm
    from reportlab.pdfgen import canvas as rl_canvas
    from reportlab.platypus import Table, TableStyle
    from reportlab.lib.utils import ImageReader
    import matplotlib
    matplotlib.use("Agg")
    import matplotlib.pyplot as plt
    import matplotlib.patches as mpatches
    return (A4, landscape, colors, cm, mm, rl_canvas, Table, TableStyle, plt, mpatches, ImageReader)

# Hora local Chile
_TZ_CHILE = ZoneInfo("America/Santiago")

def ahora_chile() -> str:
    return datetime.now(_TZ_CHILE).strftime("%d/%m/%Y %H:%M")

# ============================================================
#   CONFIGURACIÓN DE PÁGINA
# ============================================================
st.set_page_config(
    page_title="Control y Gestión Post Venta — Curifor S.A",
    page_icon="🔧",
    layout="wide",
    initial_sidebar_state="expanded",
)

# Token desde Streamlit Secrets
GITHUB_TOKEN = st.secrets.get("GITHUB_TOKEN", "")

# Logo
LOGO_B64 = "/9j/4AAQSkZJRgABAQAAAQABAAD/4gHYSUNDX1BST0ZJTEUAAQEAAAHIAAAAAAQwAABtbnRyUkdCIFhZWiAH4AABAAEAAAAAAABhY3NwAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAQAA9tYAAQAAAADTLQAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAlkZXNjAAAA8AAAACRyWFlaAAABFAAAABRnWFlaAAABKAAAABRiWFlaAAABPAAAABR3dHB0AAABUAAAABRyVFJDAAABZAAAAChnVFJDAAABZAAAAChiVFJDAAABZAAAAChjcHJ0AAABjAAAADxtbHVjAAAAAAAAAAEAAAAMZW5VUwAAAAgAAAAcAHMAUgBHAEJYWVogAAAAAAAAb6IAADj1AAADkFhZWiAAAAAAAABimQAAt4UAABjaWFlaIAAAAAAAACSgAAAPhAAAts9YWVogAAAAAAAA9tYAAQAAAADTLXBhcmEAAAAAAAQAAAACZmYAAPKnAAANWQAAE9AAAApbAAAAAAAAAABtbHVjAAAAAAAAAAEAAAAMZW5VUwAAACAAAAAcAEcAbwBvAGcAbABlACAASQBuAGMALgAgADIAMAAxADb/2wBDAAMCAgICAgMCAgIDAwMDBAYEBAQEBAgGBgUGCQgKCgkICQkKDA8MCgsOCwkJDRENDg8QEBEQCgwSExIQEw8QEBD/2wBDAQMDAwQDBAgEBAgQCwkLEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBD/wAARCABQAXYDASIAAhEBAxEB/8QAHgAAAgICAwEBAAAAAAAAAAAAAAkHCAYKAwQFAgH/xABZEAABAwMCAgUFCAkODQUAAAABAgMEAAUGBxEIEgkTITFBFFFhldMZIjI4V3GBkTNCUlNicnaxsxUXGDQ2N0NzdHV3tLXBFiMkRYKSlJahoqOy4SU1VJPS/8QAGAEBAQEBAQAAAAAAAAAAAAAAAAECAwT/xAAcEQEBAQEBAQEBAQAAAAAAAAAAARECIRIxQVH/2gAMAwEAAhEDEQA/AF/0UUVh6BRRRQ0UUUVYzaKKKKuoKKKKaCiiipoKKKKugoooqmiiiihooooolooooogJAG5PYKZJwodGBjGVYNbdQ9f5t0L17jolw7BBfMYMMLHMhUhwDnK1JIPIkp5d9iSdwFvxnUMSWX3Gg6hpxK1NnuWAQSn6e6tizTjKrDnGA49l2LzGpVqu1tYkxXWzukoUgdnoIO4I8CCPCiVRLiM6KvBW8OnZLw+zLrBvdtYU+LLOlGUxPSkbltta/ftuEA7bqKSdhsN96V6tC21qbcQpC0kpUlQ2II7wR4Gtka83i2Y/aJt9vU1qHb7dHclSpDqglDTSElSlKJ7gACa1085vEHIc4yPILW11UK6XebNjII25WnX1rQNvDZKh2VqEeJRRRVzVFFFFawFFFFAUUUUBRRRQFFFFAUUUVMBRRRUBRRRRYKKKKYoooooCsz0x0a1R1mu5smmOEXO/yEEdcqM1/iWAfF11WyGx+MRv4b1MfBJwgXHigzR+ZfHZMDBsfcQbtLa965JcPamKyo9yiO1SvtUkeJTTncFwDDNM8aiYhgeOQrJaISeVqLEbCU77dqlHvUo7dqlEk+JrjJal6wrjBeiN1qvbLUnOs8xrGUr+Gwwlye+j5+XkQfoXUoQuhzxlDYFy12urq/EsWNpsfUp1VMNul1tljt0m8Xm4RoMGG2XpEmS6ltppA71KUrYADzmqsZl0nfCjiVyctkTIb3kimlFC3rLbC6zuPM46psLHpTuD4GteRnbUK3DocrAtB/UrXm4tL8PKbC24P+V5NRPnvRL6848w7LwnK8aytCNyljnXCkKA8wcBRv6Oer3aM8ePDhrjfWMVxnK5NsvktXJFt96imI5JV9y2rdTaleZIVzHwFWFpkptjXQz7TbPdLL+5i+ouJXPH7mgb9RNYKOdP3SFfBcT+EkkemvBhxjMmR4aVBJfdQ0FEd3MoDf8A41sF64aF6e8QGDy8H1As7cll1KjEmJSPKID23vXmV96VA7bjuI7DuKRVqFplftG9Zbhplkuyp9gvDcZTqRsl9vnSpt5P4K0FKh5t9vCs3xrm6uinod82UkKGttk7Rv8A+0O+0r99x2zXx1usnqh32lNCa+xI/FFfdaxn6rXCy2wOYplV5xZ6SmQ5Z58iAp5KeUOFpwoKgD3A8u+1eTWX6x/vuZv+UVx/rC6xCstvptPWOIbB251BO/znamB23og80uVui3FGtVlQmUwh4JNpdJSFJB2+yeml/R/2wz/GI/OK2O8X/czaP5BH/Rpqxnq2Fne47Zt8t1k9UO+0qmOBaRTM71wtuiUe9MxJVyvjlkTPWyVNoUhS09YUA7kHk7t/GthOkccPPx98W/L2T+leq/iSrBe47Zt8t1k9UO+0o9x2zb5brJ6od9pTRa61yuEW026VdZqyiPDYXIeUASQhCSpR2Hf2A1U0sP3HbNvlusnqh32lfh6HbNtuzW2yb/zQ77SrNp6TjhBUkKGc3TYjcf8Aocr/APFZdpzx2cLuqWRRMSxjUxpF3nuBqJFuEKRE69w9yELdQEFR8E8258BQ2qUTuh71SQgm26vYq8rwD8OQ0PrSFVAmuHAlxDaDWqRkuR43HvGPxe1+6WV4yGmE/duoIDjafwinYeJFPPrilRY02M9CmR2348hCmnWnEhSHEKGykqB7CCCQRQ1rjYnYXMrymzYuzJTHXeLhHgIeUnmDZdcSgKIHeBzb7U2bQfg84q+HWI5ZcB4ksefsbrhdVabnYnZEZKz3qbHWBTZPjykA+I3qimrulVt0X46E4BY2w3aouX2uXb2x3NR5DjL6Gx6Ehzl/0aeHV/FtU41z4WOLfX+yrxbLuJDGrdj7pBftlosL0dqRsQQHVdaVuJBG/KTy7+FKpsGmEu/a1QtGG7syzJm5QMZE9TRLaVmV5P1vJvvtv77bffbsrYdpEenXx6bF/Suj+1jSJFl/cds2+W6yeqHfaUe47Zt8t1k9UO+0potFPqmlde47Zt8t1k9UO+0o9x2zb5brJ6od9pV3JHGhwqxJDsWTrrirbzK1NuIVL7UqB2IPZ4EVx/s2OE/5ecT/ANr/APFXabVJ/cds2+W6yeqHfaVXzi14Nb3wnx8YkXjOYOQjJnJbbYjQ1sdT1AbJJ5lHffrR9VNb/ZscJ/y84n/tf/iqJdKPrfpJrBA06b0xz+0ZIq1P3NU0QHufqA4mOEFXZ2b8itvmNWW6aoMSANydhU5aOcFfEXrhHZumI4E/Dsz+xRdburyOMtPnb5xzOD0oSR6auVwAcA+PtY/bNc9bbGi4T7ihEyw2OW3zMRmD2tyX0HsW4obKSk9iUkEgk9jFUIQ2hLbaAlCQEpSkbAAdwApev8NLDxXod8okMpczfWy2QXO9TVrtLkkfNzuON/8AaazZnod9PA2BI1oyVa/Eot0dI+ok/nq6equs2mWiWPHKNT8vhWOColDXXEqdkL235GmkgrcV6Eg7eO1VauHS1cOEaaqPBx3NpjCVbdem3soCvSEqd32+faptpqPLz0OdlW2o49rtPacA96mbY0OAn0lDydvqNQZqX0WvElhMd2fixsmaxWgVcltkFmTt6GngkKPoSo0xfRLjd4eNeri3YMPy9cG+vfYrTeGDEku/xe5KHD+ChRPoqeafVhrW7v1gvuLXeTYMms0203OGsokQ5rCmXmleZSFAEV0KfVxMcKmmvExij1rye3tQr/HaULVfmGwJMRz7UE/wjRPwkHs2J22OxpHuqGm+U6Q55etOc0heTXeySCw8B2ocT3odQfFC0kKSfMa1LqyrbaQ9F3lmremWOalwtW7TbmMigonIiu2xxamQrf3pUFgE9nftWYe47Zt8t1k9UO+0q7nBf8VbTH8n2P76mms21NK69x2zb5brJ6od9pR7jtm3y3WT1Q77Srmas8bnDzolm0rT7ULKJ0K9w2mX3WWrW++kIdQFoPOhJB3BHjWKw+kt4P5chDC9RZsYLO3WPWSYED5yGztT02qrv9DxqClJ8l1ox5avAOWx9I/4KNYFnnRU8SWKQHLhjk7GctDSSsx4EpbEggeCUvJSlR9AVTcsVyvG84x+DleIXuJd7Pcmg9EmxHA406g+II9O4I7wQQa9aptNrW5vdkvONXeZYMhtUq23O3uqYlRJTSmnWXB3pUlXaDRTIulv0atjjOI6xWKAyzd5cldjuTiU8vlKA2pxlS9u9SQhxO/mIHgKK3LrUq3XBvpVC0f4ccLxhiMluZKtzd1uSgnZTkuSkOrKvSOYI+ZAqaq6trRHbtkRETbqEsNhrl7uTlG230bV2q5sFY9K7r/fLlnMLh9sdxdj2W0RWbjem2l7CXKdHM02vbvShvlVyns3WD4DZfVWH6QVEpHGFqN5Zzc5lQynf735DH5P+Xaq8Vyt9blyOSPJkw5DUyFIcjyI7iXWXW1FKm3EndKgR3EEAg+imH2Hpgb7bbHbrfdtGGrjOixWmZMw3so8pdSkBTnL1J5eYgnbc99LsopOsL6ZF7sbN+QVr1+fYVUfiJ16Y4ktb4eqCMQTjj0huDCeipl+UBaml7Bzm5U9pBA228BUK13bJ2Xq3E//AC2f+8U+tJ42Q2vsSPxRX3Xw19iR+KK+66sNdvWP99zN/wAorj/WF1iFZfrH++7m/wCUVx/rC6xCuTeuSP8Athn+MR+cVseYv+5m0fyGP+jTWuHH/bDP8Yj84rY8xj9zVp/kLH6NNa5qdPTpHHDz8ffFvy9k/pXqePSOOHn4++Lfl7J/SPVqpDx68TOI78vCsgixWVuvPWuW222gbqWotKAAHiSa9uiqjX0Z4ZOIsMoB0OzcEJH+ZX/N+LWf6L8F3EtmeoVhYTpjf8fhsXGO/JutzjmI1EbQ4lSnAV7KUoAEgJBJO1PMoouvxIISATvsNt/PXTvV6tGOWiZfr9co9vt1vZVIlSpDgQ0y2kbqUpR7AAK7Mhtx5hbTUhbC1JIS6gJKkHzgKBH1g1SrjG4NdeNZselzMa4gb1fWoqTIaxS5sMxoshSe0BK46UIK/BPWII38R30RQ7U/VyBrlxusak2dKha5+W2qPbiobFcVhxlltZHgVBvn2/Cp5Va8GlUOXb9aMPt8+M5HlRcngMvsuJ5VtuJlICkqHgQQQa2H6ApEenXx6bF/Suj+1jT3KRFp38eix/0ro/tY1Z4s8PdoooqI1+8o0H1xeya8PM6M50425cJCkLTjsshSS4ogg9X2ivM/WD12+RXPP93Jns62F6K19DXo/WD12+RXPP8AdyZ7Osi4f9GbpmXEnhmkua4/Ptq5N2YVdIE6Kth5MZA65aVtrAUApCfEDcKp+tLbhvRWOmGkGURsuOpCN/uzZU8v/Gr9aGQtNNsNIZZbShttIShKRsEgdgAFfj77UVhyS+sIaaQVrUe5KQNyfqrkrHdRRIOn2TiJv15s00Nbd/P1C9tvp2rARVxTa75BxC6x3zNbrOdXbGZLkOxxCvduLBQohsJHduoDnUfEq+aojr5a+xI/FFfVdRyxZUqDKZnQZLsaTHcS6y80soW2tJ3SpKh2gg9oIp6PA9rlcdfeHuyZZkEjrr9bXHLNd3PF2Szy7OH0rbU2s+lRpFNNd6IAzP1m82DnN5L/AITJ6rfu6zyVrn2+jkqdewX2pY/S/wCmcOJdMH1dgxgh+eh+xXBaU7c5bHWsFR8TsXR8wpnFUZ6XdTA4e8YQsp605ewUefbyOVzf3Vnn9E+8F/xVtMfyfY/vqaahbgv+Ktpj+T7H99TTUv6FG9IlojrFm3FJfshw7S/J71bHrdbkNzINtdeZWpMdIUApIIJB7DVcIfCxxJXCSiJE0MzRTrh2SFWl1A39KlAAfSa2AaKv0uq/cC2jmaaG8OtlwjP0oZvS5Uq4PRUuhzyRLy+ZLJUOwkDtO3ZuTVgaKibW/SDUbVCySrVhmv2SYKX0FO1vhRVoO47ufkS8Ae73rgqfqKSdLDr7YrjJxzRPErmxLudlmLut6U2oLRFcLZbaYUR/CbLWojwBTv30VTHiM0D1B4d9R5OF6hLRMkSUeXRLo0pS2riypRHWpUr33NzAhQV2g9++4JK6TyNbDneEHVeDrJw8YblseSl2Yzb2rZckhW6m5kdIbcCvMTyhfzLFTLSN+DPi+v3C3mLyZsd+6YXfFoF4travftqHYmSxv2dYkdhHYFpGxIIBDmNMNXNOtZcZYy3TbK4N6t7wHMWHB1rCtvgOtn3zax9yoA/RXDnqVLFAOlP4YsmuF+jcRWF2h64QvIm4GRtR0FbkctbhqUUjtKOU8ij9ryp37DS2gQRuCCK2UloQ4gocSFJUNiCNwR5qhXLuCzhZzi5O3fINFMeVMfPM47EbXE5leJKWVISSfE7bmp1xt2EpENstlyvdxjWezQJE6fNdSxGix2y4684o7JSlI7SSacdoB0e+hti0fxmBq9pbabzmJh9fd5LzjvMH3FFZa96sD3gUEdg+1qcdNOGvQfR+WbjpvpbYrLOI5fLG2C7ISPEJdcKlpB8QCN6kunPGfpagT9gfwi/IdYv/ALH/AGlL46S3T3RHRzN8NwrSDCbdj9wbgv3S6mGtaisLWlLCV86jsR1bhHd8KmO8RnFZpTw2Y29ccvvLMq9uNKNusMV1Kpkte3vfe/wbe/e4rYDt23PZSQNW9Usp1o1EvWpeZSEuXO9SC6pCCerjtjsbZbB7kISAkfNue0mp3ZPCH/6aZhB1B08xrOLa6hyPfbVGnpKDuAXG0qKfnBJB9INZLSvejj43Mcwa0M6BavXhu221DylY7eJK+VhjrFbqivKPYhPMSpCz2DmIO3ZTP2XmZLKJEd1DrTiQpC0KCkqSe4gjsIrfN2IRxx4aI5Do5xDZNJmW51Nhyqe9ebNMCD1LqHlc62gru521qUkjv2APcRVdq2NsywXDNRLI7jed4vbL9bHTuqLcIyHm+bwUAoe9UPBQ2I8DUIPdHpwfvTfLTo5CT27lpE6UGz/o9ZWbyulC8OGi2Sa9au2DA8fgvOsOS2n7pKQglEKEhYLrqz3D3oISD3qIA762BGm22W0MtICUISEpSO4AdwrF9PdLNOdKLQbFpvhdpx2EohTjcGMlsukdxWr4Sz6VEmsmffZjMrkSXkNNNJK1uLUEpSkd5JPYBWpMLdebluSW/DsWu+WXV1LcOzQX5z6lKAAQ0grPafmpGvCJcV3jjG06u7hJVOyvyk79+6+sV2/XVqOkU45ccyywzNAtG7y3cokpwIyO9RV8zDiEnfyRhY7FgqA51jsIHKCQTVTOC742Gln5RM/9i6zbtIfXXhZ4441g2ROtLUhaLTLUlSTsUkMq2IPga92vA1A/cHkn80TP0K62jXqZ1I1GLLZOoWUfBH+eZPm/Hr1se1v1lxG5s3rHNVMrhzI6gtC03Z9Q3HbsUqUUqHnBBBrB2PsDf4g/NX3WF1sQaM5pI1G0kw7PJoQJV+skOfJDY2SHnGklzYeA5+baszqB+BS6C78JOmkrrAtSLR5OrY9xbdWjb6kip4raEpcROMtYn0h9wtkdhLTb+bWu4pSkbD/KVR3zt9Lhp1tKm4+cSfsXHtgORloiPkq7HIQvbsU6zKDKx9CUtfXTWagKRFp38eix/wBK6P7WNPdpEWnfx6LH/Suj+1jVD3aKKKBEGS8X3FDFyO7Ro+u+YttMzpDbaE3FQCUhxQAHoAFeb+zF4p/l8zP1kqoyyz91V6/nGT+lVXlVZRMX7MXin+XzM/WSq6emWu+U2HiIxfXPOb/Ovdxt92jSLlMluFx56OkBpYJ7zs1uAPQBUU0VRsl264Qrtb4t1tspuTEmsokR3m1BSHG1pCkqSR3gggg+muV5lqQyth5AW24koWk9xSRsRStOAnpAbVp1aIWimuFwWzYY5DVjvy91iCgnsjyPHqgfgr+17j2bENDtV2tV9tzF3slyi3CDKQHGJMV5LrTqD3KStJIUPSDWQifi74b8o4c9WLrZpttf/wAGrnLdlWC5BB6l+OtRUGubuDje/KpPf2AjcEGoPrY8yjEcWzezvY9mOOW292yQP8ZEuEVD7SvMeVYI3Hge8VBc/o9uEGfL8rXo5BZJO6kMTJLaCfxQ5sPo2rc6QkWxWK9ZReImPY3aZVzuk91LEWHFaLjrzhOwSlI7TT0uC/QaXw76C2bCLz1f6uy3HLreerVzJTLe23bB8eRCUI3Hfyb1memnD/oto7zr0001sdhfcHKuTHjBUhSfuS8vdzb0c21Z/wB3aaluq/aVt0u2rEG8ZdiOjtslJcXj7Lt3uaUK35HnwEsoV6Q2lSvmcHnqz/Fdx96ZaBWqbjuJ3GHlOeKQpti3RXQ4xBcI+HKcT2J5e/qwec93vQd6TZl2WZDneTXTMssujtxvF4krlzZTp9844o7k+gDsAA7AAAOwVeZ/U09Tgv8AiraY/k+x/fU01C3Bf8VbTH8n2P76mmsVScOkozPMbNxZZBAs2X3y3xUW22qSxFuLzLYJjJJISlQA3Poqr7epepLS0uN6iZQlSTuCLzJ3B/16sR0nXxvMi/my1/1ZNVTrrPwNw6LDXDPNUsCy/Fc+ySbfZGKzYqocua6XXxHkIc2bUs9qglTKiN9yArbuAq8dK66Ha7oZzbUmwlYCpdqgTAnzhp51BP8A1h9dNFrn1+hePTDYkq44Rp7k8OMkyot1lQFO7dvVuMhYTv5t2iaKs/xb6ITtedO7ZittCA/CvTVw5lEDZCWH2yO30uJ+qitc2SBDde5h+c5np7eEZBguVXSwXJvuk2+UtlZHmVyn3w9B3FeHRXiytatrhnSg8VeKMIi3O82DJ2mwEg3a1jrNh51sKbJPpVuakOP0wGr6GgmTpRiDrg71IkSUA/RzH89UHorX10ni9V06XjXaSgptGnmEQVH7Z1Ep/b/qoqJs86Q/ixz1h2G7qOLFFd7CzY4bcQgeYObF3b/Tqt1FN6v9PHZuVzuV5nvXS8XGVPmyVc70mU8p11xXnUtRJJ+c11qKKzi6O/sNSvpXxU8QOi7TcPT3U66woDXwbfIUmVET8zLwUlP0AVFFFWTE1dqw9LVxG21tDV6xnCbwEjYuKhPsOK9JKHeX6kishPTBatcuw0jxLfbv8rk/m3qglFa+qi7F/wClo4kLm2tqy47hVmChsHEQX33EekFx3l+tNV71T4pdf9Z2nImoep12nwHfhW9hSYsQ/OyyEpP0g1FdFPqgrItOs8v2l+dWTUPF/Jv1Wx+WmbD8pb6xrrACBzJ3G47T2b1jtFWUXD91V4qfvmIep1e0rq3TpQ+KG8WyZaJjmJGPOjuRneW0KB5FpKTses7DsTVRqKu0fiUhCQkdyRsK/aKKCxOi3Hjr5oNgcbTnBpNhXZ4b7z7An28vuILiuZSQoLHvdySBt2b1nXuqvFT98xD1Or2lU8oq6Jw1i4wtW9cr/ieT5yzj5uOGTPLLY9Dt5ZPNzoXyue/PMnmbSduzxqUz0qvFTv8AZMP9Tq9rVPKKui4fuqvFT98w/wBTq9rVY7TqJkVl1MjaswjF/V6JfRkLfO1uz5WH+v7Ub/A5/td+7s3rGKKouH7qrxU/fMQ9Tq9pR7qrxU/fMQ9Tq9pVPKKaOefMeuM6TcJHL1sp5b7nKNhzKUVHYebc1wUUVrQUUUVYgrPtMde9ZNGn+t0z1FvNjaKudUVl/nirP4TK92z/AKtYDRWmVzsZ6V7iaszLbF7gYfkAT8J2TbXGHVfSy4lA/wBSstT0wWrYQArSTEirbtIlyQPq3qglFPF1d6+dLbxET0KbsuJYRagobBfkkiQtPpBU8E/Wk1BepPGbxM6rsOwss1YuqILw5VwrZywGVJ8xSwElQ+cmoVorWQ0HtJJ7ydyfOaKKKItDp30jPEZphg9l0+xdeMC1WGImHE8otZcd6tPdzK6wbnt79qyH3VTiq++Yh6nV7Sqe0VMi6zjWfWPMdeM+l6kZ4YJvE1llh3yJjqWuVpAQnZO52OwG/bWD0UVTUjaFa96hcOuYvZxpxJhN3CTCXb3kTGOuZcZUpKiCncdoUhJHb3ip991V4qfvmIep1e0qnlFM1dXD91V4qfvmIep1e0oqnlFTIa//2Q=="
LOGO_DATA_URI = f"data:image/jpeg;base64,{LOGO_B64}"

# Imagenes reales del documento oficial Ford "Hoja Multipuntos" (VCU) — el mismo
# archivo escaneado exacto que exige la marca, no una recreacion en HTML. Se hospedan
# como archivos aparte en GitHub (no embebidas en base64 aqui — pesan ~300-420KB c/u)
# y se cargan por URL directa en el navegador dentro del modulo del Planificador.
VCU_FORD_P1_URL = f"https://raw.githubusercontent.com/Cjerez-curi/curifor-ots/main/vcu_ford_p1.jpg"
VCU_FORD_P2_URL = f"https://raw.githubusercontent.com/Cjerez-curi/curifor-ots/main/vcu_ford_p2.jpg"

# ============================================================
#   CONFIGURACIÓN GITHUB  (antes de check_password)
# ============================================================
GITHUB_USUARIO        = "Cjerez-curi"
GITHUB_REPO           = "curifor-ots"

# Supabase para el tablero embebido. La clave `anon` es pública por diseño: sin
# un vale válido no abre absolutamente nada (todas las tablas tienen RLS y las
# dos funciones del tablero exigen el vale). Ver herramientas/setup_supabase_tablero.sql.
SUPABASE_URL      = st.secrets.get("SUPABASE_URL", "https://ordgsglujssgzmnlmcus.supabase.co")
SUPABASE_ANON_KEY = st.secrets.get(
    "SUPABASE_ANON_KEY",
    "eyJhbGciOiJIUzI1NiIsInR5cCI6IkpXVCJ9."
    "eyJpc3MiOiJzdXBhYmFzZSIsInJlZiI6Im9yZGdzZ2x1anNzZ3ptbmxtY3VzIiwicm9sZSI6ImFub24iLCJpYXQiOjE3ODUzMzM4NTgsImV4cCI6MjEwMDkwOTg1OH0."
    "n15xGwipVso0hRC9_LuWfFEe34eP9O1J1NC4LlenwUM")
GITHUB_ARCHIVO        = "datos_dashboard.json"
GITHUB_COMENTARIOS    = "comentarios_log.json"
GITHUB_USUARIOS       = "usuarios_curifor.json"
GITHUB_NOTIFICACIONES = "notificaciones.json"
GITHUB_AUDIT          = "audit_log.json"
GITHUB_ONLINE         = "online_users.json"
GITHUB_AGENDA         = "agenda_hoy.json"
GITHUB_CTRL_TALLER    = "control_taller.json"
GITHUB_PRODUCCION_TECNICOS = "produccion_tecnicos.json"
# Mapeo manual tecnico->sucursal (23/07/2026, a pedido de Cristobal): respaldo
# para cuando BDFlexline no trae sucursal para un tecnico (ver
# _cargar_tecnicos_sucursal_manual / seccion Admin -> Tecnicos).
GITHUB_TECNICOS_SUCURSAL_MANUAL = "tecnicos_sucursal_manual.json"
GITHUB_CAMPANAS       = "campanas_curifor.json"  # Revision de Campanas (Agenda Ford), 28/07/2026
GITHUB_CUENTA_FICHA     = "cuenta_ficha.json"            # Cuenta Ficha: saldos + historial de OT, 31/07/2026
GITHUB_INFORMES_GESTION = "informes_gestion.json"        # Informes de Gestion por marca (AG / Ford), 04/08/2026

# Meses usados por el modulo "Informes de Gestion": los reportes AG traen los
# 12 meses como columnas (Ene..Dic) y el IMOP de Ford una hoja por mes.
MESES_CORTOS_IG = ["Ene", "Feb", "Mar", "Abr", "May", "Jun",
                   "Jul", "Ago", "Sept", "Oct", "Nov", "Dic"]
MESES_LARGOS_IG = ["Enero", "Febrero", "Marzo", "Abril", "Mayo", "Junio",
                   "Julio", "Agosto", "Septiembre", "Octubre", "Noviembre", "Diciembre"]


def _norm_txt_ig(s):
    """Normaliza texto (mayusculas, sin tildes) para comparar nombres de
    sucursal entre el informe IMOP y los codigos del ERP."""
    import unicodedata as _ud
    s = str(s or "").upper().strip()
    s = "".join(c for c in _ud.normalize("NFD", s)
                if _ud.category(c) != "Mn")
    return " ".join(s.split())


GITHUB_CUENTA_FICHA_REV = "cuenta_ficha_revisados.json"  # Marcas "Revisado" por cliente (compartidas)
GITHUB_LOANERS          = "loaners.json"                 # Flota de vehiculos de cortesia (modulo Loaners)
ADMIN_EMAIL           = "cjerez@curifor.com"
DOMINIO_PERMITIDO     = "@curifor.com"

# ============================================================
#   NOMINA — mapa correo -> sucursal (columna J / columna E de
#   "Nomina Area PV (Actualizada).xlsx", subida por Cristobal el
#   20/07/2026). Se usa SOLO para asignar automaticamente la
#   sucursal de origen de un usuario nuevo (o uno que aun no la
#   tenga guardada) la primera vez que inicia sesion o se
#   registra — a partir de ahi queda restringido a esa sola
#   sucursal en toda la app (Control y Gestion Post Venta,
#   Planificador de Taller, Asistente App) hasta que el admin le
#   habilite otras desde 🛡️ Admin -> 👥 Usuarios -> 🏢 Acceso por
#   Sucursal. Si un correo no aparece aca, el usuario NO queda
#   restringido (ve todas las sucursales) hasta que el admin se lo
#   asigne a mano. Para actualizar esta lista en el futuro: volver
#   a exportar la nomina y regenerar este dict (correo en minuscula
#   -> sucursal tal como aparece en SUCURSAL del PBI, mayuscula).
# ============================================================
NOMINA_SUCURSAL_POR_EMAIL = {
    'aacevedo@curifor.com': 'CURICO',
    'acardenas@curifor.com': 'TALCA',
    'adelafuente@curifor.com': 'BRASIL 18',
    'afuentes@curifor.com': 'CHILLAN',
    'agutierrez@curifor.com': 'CHILLAN',
    'ahernandez@curifor.com': 'LINDEROS',
    'ajara@curifor.com': 'CURICO',
    'alejandrom@curifor.com': 'TALCA',
    'aretamal@curifor.com': 'CD REPUESTOS',
    'ariquelme@curifor.com': 'RANCAGUA',
    'asalinas@curifor.com': 'CD REPUESTOS',
    'asandoval@curifor.com': 'CHILLAN',
    'bgomez@curifor.com': 'TALCA',
    'bodegabrasil@curifor.com': 'BRASIL 18',
    'bodegachillan2@curifor.com': 'CHILLAN VIEJO',
    'bodegalinderos@curifor.com': 'LINDEROS',
    'bodegaloblanco@curifor.com': 'LO BLANCO',
    'bodegatalca@curifor.com': 'TALCA',
    'bpoblete@curifor.com': 'TALCA',
    'cabarzua@curifor.com': 'CD REPUESTOS',
    'cajabrasil@curifor.com': 'BRASIL 18',
    'cajaddj602@curifor.com': 'DIEZ DE JULIO 602',
    'caliaga@curifor.com': 'CURICO',
    'ccarrasco@curifor.com': 'CHILLAN VIEJO',
    'ccatalan@curifor.com': 'LINDEROS',
    'centrodistribucion@curifor.com': 'CD REPUESTOS',
    'cflores@curifor.com': 'LINDEROS',
    'cgarcia@curifor.com': 'LINDEROS',
    'cjerez@curifor.com': 'LINDEROS',
    'cmartin@curifor.com': 'CHILLAN',
    'cmontecinos@curifor.com': 'RANCAGUA',
    'cpasten@curifor.com': 'TALCA',
    'crios@curifor.com': 'CD REPUESTOS',
    'cromero@curifor.com': 'LINDEROS',
    'csalazar@curifor.com': 'TALCA',
    'ctroncoso@curifor.com': 'LO BLANCO',
    'culloa@curifor.com': 'CHILLAN',
    'cvalenzuela@curifor.com': 'CHILLAN',
    'dbravo@curifor.com': 'TALCA (2)',
    'dfigueroa@curifor.com': 'RANCAGUA',
    'dloiza@curifor.com': 'RANCAGUA',
    'dmartinez@curifor.com': 'RANCAGUA',
    'dmondaca@curifor.com': 'TALCA',
    'dpina@curifor.com': 'CURICO',
    'ecarvajal@curifor.com': 'PLACILLA',
    'ecifuentes@curifor.com': 'CURICO',
    'edgarv@curifor.com': 'TALCA',
    'eleiva@curifor.com': 'CURICO',
    'eortiz@curifor.com': 'LINDEROS',
    'epacheco@curifor.com': 'LINDEROS',
    'equezada@curifor.com': 'CURICO',
    'erodriguezs@curifor.com': 'CD REPUESTOS',
    'evaldevenito@curifor.com': 'TALCA (2)',
    'evalenzuela@curifor.com': 'CD REPUESTOS',
    'fagurto@curifor.com': 'LINDEROS',
    'fcordova@curifor.com': 'LINDEROS',
    'fespinoza@curifor.com': 'CURICO',
    'ffuentes@curifor.com': 'TALCA',
    'fmanriquez@curifor.com': 'CD REPUESTOS',
    'fmora@curifor.com': 'CD REPUESTOS',
    'fordonez@curifor.com': 'LINDEROS',
    'fperez@curifor.com': 'RANCAGUA',
    'frobles@curifor.com': 'TALCA',
    'gmaldonado@curifor.com': 'RANCAGUA',
    'griquelme@curifor.com': 'CD REPUESTOS',
    'hcastro@curifor.com': 'TALCA',
    'hcoronado@curifor.com': 'CD REPUESTOS',
    'hgarcia@curifor.com': 'CD REPUESTOS',
    'hjimenez@curifor.com': 'CHILLAN',
    'hlopez@curifor.com': 'CHILLAN VIEJO',
    'hmillar@curifor.com': 'BRASIL 18',
    'icabrera@curifor.com': 'CHILLAN',
    'icalderon@curifor.com': 'CD REPUESTOS',
    'icastro@curifor.com': 'RANCAGUA',
    'importaciones@curifor.com': 'CD REPUESTOS',
    'jacevedo@curifor.com': 'LINDEROS',
    'jalarcon@curifor.com': 'CURICO',
    'jalvarado@curifor.com': 'CURICO',
    'jarenas@curifor.com': 'CHILLAN VIEJO',
    'jarmijo@curifor.com': 'RANCAGUA',
    'javalos@curifor.com': 'CURICO',
    'jchavez@curifor.com': 'BRASIL 18',
    'jfuenzalida@curifor.com': 'CURICO',
    'jgomez@curifor.com': 'LINDEROS',
    'jgonzalez@curifor.com': 'CHILLAN',
    'jherrera@curifor.com': 'PLACILLA',
    'jhuaiqui@curifor.com': 'TALCA',
    'jmunoz@curifor.com': 'CHILLAN',
    'jneira@curifor.com': 'CURICO',
    'jorge.moreno@curifor.com': 'LINDEROS',
    'jortega@curifor.com': 'TALCA',
    'jose.lobos@curifor.com': 'LINDEROS',
    'jpbizarro@curifor.com': 'CHILLAN',
    'jploreto@curifor.com': 'LINDEROS',
    'jrojas@curifor.com': 'TALCA',
    'jsolis@curifor.com': 'CURICO',
    'juanr@curifor.com': 'DIEZ DE JULIO 602',
    'kbarrera@curifor.com': 'TALCA (2)',
    'kquintana@curifor.com': 'TALCA',
    'lfigueroa@curifor.com': 'DIEZ DE JULIO 602',
    'lfuentealba@curifor.com': 'TALCA (2)',
    'lmarin@curifor.com': 'AUTOPARK',
    'lmartinez@curifor.com': 'LINDEROS',
    'lriquelme@curifor.com': 'CHILLAN',
    'maguilar@curifor.com': 'LINDEROS',
    'matiasr@curifor.com': 'PLACILLA',
    'mcanales@curifor.com': 'CD REPUESTOS',
    'mcares@curifor.com': 'CHILLAN',
    'mcarrasquero@curifor.com': 'AUTOPARK',
    'mfernandez@curifor.com': 'CD REPUESTOS',
    'mfigueroa@curifor.com': 'LINDEROS',
    'mguerrero@curifor.com': 'CD REPUESTOS',
    'miguelm@curifor.com': 'TALCA (2)',
    'mlopez@curifor.com': 'LINDEROS',
    'mmartinez@curifor.com': 'LO BLANCO',
    'mmatta@curifor.com': 'PLACILLA',
    'mmunoz@curifor.com': 'TALCA',
    'mnunez@curifor.com': 'BRASIL 18',
    'mparada@curifor.com': 'CD REPUESTOS',
    'mramos@curifor.com': 'CD REPUESTOS',
    'mriquelme@curifor.com': 'BRASIL 18',
    'msanchez@curifor.com': 'CHILLAN',
    'nhernandez@curifor.com': 'CHILLAN',
    'nmaltes@curifor.com': 'CD REPUESTOS',
    'nmartinez@curifor.com': 'CHILLAN VIEJO',
    'nmedina@curifor.com': 'RANCAGUA',
    'nmorales@curifor.com': 'CHILLAN VIEJO',
    'nriveros@curifor.com': 'LINDEROS',
    'nrodriguez@curifor.com': 'TALCA',
    'nvallejos@curifor.com': 'CHILLAN',
    'nvillalobos@curifor.com': 'RANCAGUA',
    'oescobar@curifor.com': 'LO BLANCO',
    'pabaca@curifor.com': 'CURICO',
    'pbriones@curifor.com': 'CD REPUESTOS',
    'rbastias@curifor.com': 'CHILLAN',
    'repuestosdyp.lf@curifor.com': 'LO BLANCO',
    'repuestosfordplacilla@curifor.com': 'PLACILLA',
    'repuestoshyundailf@curifor.com': 'DIEZ DE JULIO 602',
    'repuestoslinderos@curifor.com': 'LINDEROS',
    'repuestosrancagua3@curifor.com': 'RANCAGUA',
    'rmorales@curifor.com': 'DIEZ DE JULIO 602',
    'rmunoz@curifor.com': 'LINDEROS',
    'rpalma@curifor.com': 'CHILLAN VIEJO',
    'rramirez@curifor.com': 'CURICO',
    'rsoto@curifor.com': 'CURICO',
    'rvaldivia@curifor.com': 'CURICO',
    'slandaeta@curifor.com': 'TALCA (2)',
    'ssarmiento@curifor.com': 'LO BLANCO',
    'tbetancourt@curifor.com': 'RANCAGUA',
    'tgomez@curifor.com': 'CURICO',
    'vcaceres@curifor.com': 'TALCA',
    'vsanmartin@curifor.com': 'CHILLAN',
}

# ============================================================
#   ASESORES — permiso limitado "Confirmar Citas" en el Planificador
#   -------------------------------------------------------------
#   Listado de asesores (subido por Cristobal el 23/07/2026,
#   "Asesores.xlsx", 29 correos) que deben poder usar SOLO los
#   botones "Asiste"/"No Asiste"/"Reagenda" de la columna "Citas
#   <fecha>" del JPCB (confirmar si el vehiculo agendado llego o
#   no) — sin el resto de permisos de edicion del Planificador
#   (no pueden arrastrar tarjetas, asignar tecnicos/horarios ni
#   editar Control de Taller). Se usa para PRE-REGISTRAR a estos
#   usuarios en usuarios_curifor.json con el flag
#   `puede_confirmar_citas: True` de una sola vez desde el boton
#   "🚀 Aplicar a la lista de asesores" en 🛡️ Admin -> 👥 Usuarios
#   -> 🔐 Permisos de modulos (no hace falta que cada uno haya
#   iniciado sesion antes) — si el correo ya existe, solo se le
#   agrega el flag (no se pisa nada mas); si no existe, se crea sin
#   password_hash, asi la primera vez que esa persona entre con su
#   correo, la app le va a pedir crear su contraseña normalmente.
#   Para actualizar esta lista en el futuro: volver a exportar el
#   Excel de asesores y regenerar este dict (correo en minuscula ->
#   nombre tal como viene en la columna "Nombre").
#   NOTA: 'galiaga@curifor.onmicrosoft.com' NO termina en
#   "@curifor.com" (DOMINIO_PERMITIDO) — con ese dominio esa
#   persona no va a poder iniciar sesion en la app. Confirmar con
#   Cristobal si es un error de tipeo (deberia ser @curifor.com) o
#   si de verdad usa esa cuenta de Microsoft.
# ============================================================
ASESORES_CONFIRMAR_CITAS = {
    'lmarin@curifor.com': 'MARIN GARCIA LUIS IGNACIO',
    'agutierrez@curifor.com': 'GUTIERREZ FUENTES ALEX MATIAS',
    'asandoval@curifor.com': 'SANDOVAL BAEZA ANA MARIA',
    'jgonzalez@curifor.com': 'GONZALEZ ARIAS JUAN IGNACIO',
    'ccarrasco@curifor.com': 'CARRASCO CANTO CHRISTOPHER NAIN ALEXANDER',
    'fespinoza@curifor.com': 'ESPINOZA SAAVEDRA FELIPE IGNACIO',
    'aacevedo@curifor.com': 'ACEVEDO MUÑOZ ANYELA CHARLOTTE',
    'ecifuentes@curifor.com': 'CIFUENTES BARRERA ERICK GONZALO',
    'caliaga@curifor.com': 'ALIAGA CABRERA CRISTIAN ORLANDO',
    'galiaga@curifor.onmicrosoft.com': 'ALIAGA CORREA GONZALO ANDRES',
    'eortiz@curifor.com': 'ORTIZ MORENO EDUARDO ANDRES',
    'mfigueroa@curifor.com': 'FIGUEROA JARA MATIAS IGNACIO',
    'rmunoz@curifor.com': 'MUÑOZ OVALLE RODRIGO ANTONIO',
    'oescobar@curifor.com': 'ESCOBAR GARRIDO OSCAR DANIEL',
    'csaavedra@curifor.com': 'SAAVEDRA POBLETE CRISTOBAL HERNAN',
    'jherrera@curifor.com': 'HERRERA PEREZ JOCELYN ELIZABETH',
    'mmatta@curifor.com': 'MATTA ZAMORA MARIO ANDRES',
    'dmartinez@curifor.com': 'MARTINEZ ROMERO DANIEL IGNACIO',
    'icastro@curifor.com': 'CASTRO GOMEZ IGNACIO ALEJANDRO',
    'nmartinez@curifor.com': 'MARTINEZ VALDERRAMA NURY GABRIELA',
    'ariquelme@curifor.com': 'RIQUELME PAVEZ ALEXANDER IGNACIO',
    'jarmijo@curifor.com': 'ARMIJO VON JENTSCHYK JUAN ANTONIO',
    'jortega@curifor.com': 'ORTEGA GONZALEZ JUAN LUIS SEBASTIAN',
    'bpoblete@curifor.com': 'POBLETE SANCHEZ BORIS IGNACIO FELIPE',
    'vcaceres@curifor.com': 'CACERES GUTIERREZ VICENTE PATRICIO',
    'ffuentes@curifor.com': 'FUENTES MUÑOZ FELIPE IGNACIO',
    'lfuentealba@curifor.com': 'FUENTEALBA GUERRERO LUIS PATRICIO',
    'slandaeta@curifor.com': 'LANDAETA TREJO STEEPHANY ALEXANDRA',
    'miguelm@curifor.com': 'MARTINEZ RODRIGUEZ MIGUEL ANGEL',
}

URL_DATOS = (
    f"https://raw.githubusercontent.com/"
    f"{GITHUB_USUARIO}/{GITHUB_REPO}/main/{GITHUB_ARCHIVO}"
)
URL_COMENTARIOS = (
    f"https://raw.githubusercontent.com/"
    f"{GITHUB_USUARIO}/{GITHUB_REPO}/main/{GITHUB_COMENTARIOS}"
)
URL_RANKING = (
    f"https://raw.githubusercontent.com/"
    f"{GITHUB_USUARIO}/{GITHUB_REPO}/main/ranking_cierres.json"
)
URL_AGENDA = (
    f"https://raw.githubusercontent.com/"
    f"{GITHUB_USUARIO}/{GITHUB_REPO}/main/{GITHUB_AGENDA}"
)
_GITHUB_API_BASE = (
    f"https://api.github.com/repos/{GITHUB_USUARIO}/{GITHUB_REPO}/contents/"
)

DOCS_CONFIG = [
    ("Liquidación ST",   "LIQ_ST"),
    ("Factura Cliente",  "FACT_CLIENTE"),
    ("Factura Compañía", "FACT_COMPANIA"),
    ("Cargo Interno",    "CARGO_INT"),
    ("Cargo Garantía",   "CARGO_GTIA"),
    ("Factura Garantía", "FACT_GTIA"),
    ("Vale de Consumo",  "VALE_CONSUMO"),
]

# ============================================================
#   PLANIFICADOR DE TALLER — HTML Component
# ============================================================
def _emitir_vale_tablero(usuario, sucursal, horas=12):
    """Un permiso mínimo para que el tablero embebido hable con Supabase.

    Antes se le pasaba al navegador el token de GitHub, con permiso de escritura
    sobre TODO el repositorio: cualquiera que abriera las herramientas del
    navegador podía reescribir `usuarios_curifor.json` y darse permisos de
    administrador. El vale solo abre los dos documentos del tablero, dura una
    jornada y queda registrado a nombre de quien lo pidió.

    Devuelve "" si no hay Supabase, y ahí el tablero sigue con GitHub como antes.
    """
    if not _datos.disponible():
        return ""
    try:
        vale = secrets.token_urlsafe(32)
        conn = _datos._conn()
        if conn is None:
            return ""
        cur = conn.cursor()
        cur.execute(
            """insert into public.taller_vales (vale, usuario, sucursal, expira)
               values (%s, %s, %s, now() + make_interval(hours => %s))""",
            (vale, usuario or "?", sucursal or "?", int(horas)))
        # Los vencidos no sirven para nada y no hay proceso que los limpie;
        # se barren acá, que es donde se sabe que la tabla está en uso.
        cur.execute("delete from public.taller_vales where expira < now() - interval '1 day'")
        conn.commit()
        return vale
    except Exception:
        try:
            conn.rollback()
        except Exception:
            pass
        return ""


def _generar_html_planificador(sucursal, usuario, puede_editar, token, github_user, github_repo,
                               agenda_data=None, ctrl_data=None, ctrl_sha="",
                               puede_prepicking=False, prepicking_data=None,
                               prepicking_sha="", logo_data_uri="", produccion_data=None,
                               cotizador_gz="", stock_completo_gz="",
                               puede_confirmar_citas=False,
                               puede_disponibilidad=False):
    """Planificador de Taller — 5 dias desde Agenda Curifor.
    Los datos se inyectan desde Python — sin fetch en el browser."""
    import json as _json
    api_base   = f"https://api.github.com/repos/{github_user}/{github_repo}/contents/"
    puede_str  = "true" if puede_editar else "false"
    # Permiso limitado (23/07/2026, asesores) — independiente de PUEDE_EDITAR:
    # solo habilita los botones Asiste/No Asiste/Reagenda de la columna "Citas".
    puede_confirmar_str = "true" if (puede_editar or puede_confirmar_citas) else "false"
    # Permiso limitado (29/07/2026, Torre de Control) — marcar a un tecnico como
    # no disponible (vacaciones/licencia/permiso/capacitacion) en Produccion
    # Tecnicos, para que esos dias no cuenten como horas disponibles.
    puede_disp_str = "true" if puede_disponibilidad else "false"
    puede_pp_str = "true" if puede_prepicking else "false"
    # Con Supabase, el navegador NO recibe el token de GitHub: recibe un vale
    # que solo abre los dos documentos del tablero (ver _emitir_vale_tablero).
    sb_vale = _emitir_vale_tablero(usuario, sucursal)
    sb_url  = SUPABASE_URL if sb_vale else ""
    sb_key  = SUPABASE_ANON_KEY if sb_vale else ""
    token_safe = "" if sb_vale else (token or "")
    # Serializar datos como JSON para inyectarlos en el JS
    # ensure_ascii=False para que tildes/ñ viajen legibles; se escapa "</" a "<\/"
    # para que un comentario/campo que contenga literalmente "</script" no cierre
    # la etiqueta <script> antes de tiempo y trunque todo el JS que viene después
    # (incluyendo el loadData() final) — eso dejaría el spinner de carga pegado
    # para siempre sin ningún error visible.
    _agenda_js  = _json.dumps(agenda_data  or {}, ensure_ascii=False).replace("</", "<\\/")
    _ctrl_js    = _json.dumps(ctrl_data    or {}, ensure_ascii=False).replace("</", "<\\/")
    _pp_js      = _json.dumps(prepicking_data or {}, ensure_ascii=False).replace("</", "<\\/")
    _prod_js    = _json.dumps(produccion_data or {}, ensure_ascii=False).replace("</", "<\\/")
    _sha_safe   = (ctrl_sha or "").replace('"', '')
    _pp_sha_safe = (prepicking_sha or "").replace('"', '')
    _logo_safe  = (logo_data_uri or "").replace('"', '')
    # OJO: no se puede usar backslash dentro de una expresion {} de un f-string en
    # Python <3.12 (SyntaxError) — por eso el HTML del boton se arma ANTES, como
    # variable simple, y solo se interpola su nombre (sin backslash) mas abajo.
    _pp_tab_btn = ('<button class="sntab" onclick="switchView(\'pp\',this)">🧩 Pre-picking</button>'
                   if puede_prepicking else '')
    _cotiz_gz_safe = (cotizador_gz or "")
    _stock_gz_safe = (stock_completo_gz or "")

    return f"""<!DOCTYPE html>
<html lang="es">
<head>
<meta charset="UTF-8">
<style>
:root{{--azul:#2b5ea7;--azulOsc:#0d2f5a;--azulGrad:linear-gradient(135deg,var(--azulOsc),var(--azul));--borde:#dbe2ea;--gris:#f4f6f9;--sombra:0 2px 10px rgba(13,47,90,.10);--sombraH:0 6px 18px rgba(13,47,90,.18);}}
*{{box-sizing:border-box;font-family:'Segoe UI',Arial,sans-serif;font-size:13px;margin:0;padding:0;}}
body{{background:#fff;padding:8px;overflow-x:hidden;}}
.subnav{{display:flex;align-items:center;gap:6px;margin-bottom:10px;padding:7px 9px;background:var(--gris);border-radius:10px;border:1px solid var(--borde);flex-wrap:wrap;box-shadow:var(--sombra);}}
.sntab{{border:1px solid var(--borde);border-radius:7px;padding:7px 16px;cursor:pointer;font-weight:600;background:#fff;color:#456;transition:background .15s,box-shadow .15s;}}
.sntab:hover{{background:#eef3fb;}}
.sntab.active{{background:var(--azulGrad);color:#fff;border-color:var(--azul);box-shadow:var(--sombra);}}
.save-st{{margin-left:auto;font-size:12px;color:#667;}}
.area-bar{{display:flex;align-items:center;gap:8px;margin-bottom:10px;padding:8px 11px;background:#fff;border:1px solid var(--borde);border-radius:10px;flex-wrap:wrap;box-shadow:var(--sombra);}}
.area-lbl{{font-weight:700;color:var(--azulOsc);font-size:12px;}}
.area-tab{{border:1px solid var(--borde);border-radius:20px;padding:5px 16px;cursor:pointer;font-weight:700;background:var(--gris);color:#456;font-size:12px;transition:background .15s,box-shadow .15s;}}
.area-tab:hover{{background:#e9eef5;}}
.area-tab.active{{background:var(--azulGrad);color:#fff;border-color:var(--azul);box-shadow:var(--sombra);}}
.area-hint{{font-size:10.5px;color:#889;margin-left:4px;}}
/* JPCB */
.kanban{{display:flex;gap:8px;overflow-x:auto;padding-bottom:8px;align-items:flex-start;}}
.col{{flex:0 0 160px;background:var(--gris);border:1px solid var(--borde);border-radius:10px;display:flex;flex-direction:column;box-shadow:var(--sombra);}}
/* Fila de titulos "responsable" sobre el JPCB (Asesor/Torre Control/Tecnico/Asesor) —
   mismo ancho fijo de columna (160px) + gap (8px) que .kanban/.col, para que cada
   bloque quede alineado exactamente sobre las columnas que agrupa. */
.kanban-groups{{display:flex;gap:8px;overflow-x:auto;margin-bottom:6px;}}
.kg-block{{flex:0 0 auto;background:#333f4d;color:#fff;text-align:center;font-size:10.5px;font-weight:700;text-transform:uppercase;letter-spacing:.4px;padding:5px 4px;border-radius:5px;}}
.col-head{{padding:7px 8px;color:#fff;font-weight:700;border-radius:5px 5px 0 0;text-align:center;font-size:11px;line-height:1.3;}}
.cnt{{font-weight:400;opacity:.9;}}
.drop{{flex:1;padding:5px;display:flex;flex-direction:column;gap:5px;min-height:80px;}}
.drop.over{{background:#ddeeff;outline:2px dashed var(--azul);}}
.card{{border:1px solid #c5cdd5;border-left:5px solid #999;border-radius:7px;padding:6px 8px;background:#fff;cursor:grab;user-select:none;box-shadow:0 1px 4px rgba(0,0,0,.06);transition:box-shadow .15s;}}
.card:hover{{box-shadow:0 4px 12px rgba(0,0,0,.12);}}
.card.dragging{{opacity:.35;}}
.card b{{font-size:13px;color:var(--azulOsc);}}
.cinfo{{color:#445;margin:2px 0;font-size:12px;}}
.cmeta{{color:#778;font-size:11px;}}
.cbadge{{display:inline-block;background:#e23b3b;color:#fff;border-radius:3px;font-size:9px;padding:1px 4px;margin-bottom:2px;}}
.cot{{display:inline-block;background:var(--azulOsc);color:#fff;border-radius:3px;font-size:10.5px;font-weight:700;padding:1px 6px;margin:2px 0;letter-spacing:.2px;}}
.ccoment{{color:#333;font-size:11px;background:#fff8dd;border-left:3px solid #d6ad00;border-radius:2px;padding:2px 5px;margin-top:3px;word-break:break-word;}}
.centrega{{color:#fff;font-size:11px;font-weight:700;background:#c0392b;border-radius:3px;padding:2px 6px;margin-top:3px;display:inline-block;}}
.cacts{{margin-top:3px;display:flex;gap:3px;flex-wrap:wrap;}}
.cacts button{{border:none;border-radius:3px;padding:2px 6px;cursor:pointer;font-size:10px;background:#f0f0f0;color:#333;}}
.cacts button:hover{{background:#ddd;}}
/* Columna "Citas <fecha>" del JPCB (23/07/2026) — tarjetas de citas aun sin confirmar,
   con 3 botones grandes (Asiste/No Asiste/Reagenda) en vez de las acciones normales. */
.card.cita-pend{{cursor:default;border-left-color:#0b7d43;}}
.cacts-cita{{margin-top:5px;display:flex;flex-direction:column;gap:3px;}}
.cacts-cita button{{border:none;border-radius:4px;padding:5px 4px;cursor:pointer;font-size:11px;font-weight:700;color:#fff;}}
.btn-asiste{{background:#1e8449;}}
.btn-asiste:hover{{background:#166638;}}
.btn-noasiste{{background:#b33;}}
.btn-noasiste:hover{{background:#8f2828;}}
.btn-reagenda{{background:#c87900;}}
.btn-reagenda:hover{{background:#a56300;}}
.cita-reagenda{{font-size:9px;font-weight:700;color:#fff;background:#c87900;border-radius:2px;padding:1px 4px;display:inline-block;margin-bottom:2px;}}
.stopzone{{margin-top:12px;border:1px solid rgba(226,59,59,.35);border-radius:10px;overflow:hidden;box-shadow:var(--sombra);}}
.sz-head{{background:linear-gradient(135deg,#8a1f1f,#e23b3b);color:#fff;padding:8px 12px;font-weight:700;}}
/* Date tabs */
.dtabs{{display:flex;gap:6px;margin-bottom:8px;flex-wrap:wrap;}}
.dtab{{border:1px solid var(--borde);border-radius:7px;padding:6px 14px;cursor:pointer;font-weight:600;background:#fff;color:#456;font-size:12px;transition:background .15s,box-shadow .15s;}}
.dtab:hover{{background:#eef3fb;}}
.dtab.active{{background:var(--azulGrad);color:#fff;border-color:var(--azul);box-shadow:var(--sombra);}}
.dtab.prepick{{border-color:#c87900;}}
.dtab.prepick.active{{background:linear-gradient(135deg,#8a5600,#c87900);color:#fff;border-color:#c87900;box-shadow:0 2px 10px rgba(200,121,0,.25);}}
/* Plan split layout */
.plan-board{{display:flex;gap:0;height:calc(100vh - 210px);min-height:480px;}}
.prog-panel{{flex:0 0 215px;border-right:2px solid #bbb;overflow-y:auto;background:#fff;display:flex;flex-direction:column;}}
.prog-head-bar{{background:var(--azulGrad);color:#fff;padding:8px 10px;font-weight:700;font-size:13px;position:sticky;top:0;z-index:5;flex-shrink:0;}}
.prog-group{{margin-bottom:4px;}}
.prog-asesor{{background:#e6edf5;padding:4px 8px;font-weight:700;font-size:11px;color:#0b2e63;border-bottom:1px solid #ccc;}}
/* Cita cards */
.cita-card{{background:#d6dce5;border:1px solid #b9c2cd;border-radius:6px;margin:4px 6px;padding:6px 8px;cursor:grab;user-select:none;position:relative;box-shadow:0 1px 3px rgba(0,0,0,.05);transition:box-shadow .15s;}}
.cita-card:hover{{background:#c5cfd9;box-shadow:0 3px 10px rgba(0,0,0,.1);}}
.cita-card.dragging{{opacity:.35;}}
.cita-card.prepicking{{background:#fff3cd;border-color:#c87900;border-left:4px solid #c87900;}}
.cita-card.asignado{{opacity:.65;cursor:not-allowed;background:#d4edda;border-color:#5aa84a;}}
.cita-top{{display:flex;justify-content:space-between;align-items:center;margin-bottom:2px;}}
.cita-hora{{font-weight:700;font-size:12px;color:var(--azulOsc);}}
.cita-status{{font-size:13px;}}
.cita-plate{{font-weight:700;color:#1a1a1a;font-size:12px;}}
.cita-info{{color:#667;font-size:10px;}}
.cita-svc{{color:#334;font-size:11px;margin-top:1px;}}
.cita-cliente{{color:#667;font-size:10px;white-space:nowrap;overflow:hidden;text-overflow:ellipsis;}}
.cita-asig{{color:#147a3d;font-size:10px;font-weight:700;}}
.cita-card.no-asiste{{background:#f2f2f2 !important;opacity:.6;}}
.cita-card.no-asiste .cita-plate{{text-decoration:line-through;}}
.cita-noasiste{{font-size:9px;font-weight:700;color:#fff;background:#b33;border-radius:2px;padding:1px 4px;display:inline-block;margin-bottom:2px;}}
.cita-noasiste-btn{{margin-top:4px;border:none;border-radius:3px;padding:2px 6px;cursor:pointer;font-size:10px;background:#f0f0f0;color:#333;}}
.cita-noasiste-btn:hover{{background:#ddd;}}
/* Plan right panel */
.plan-panel{{flex:1;display:flex;flex-direction:column;overflow:hidden;}}
.plan-head-bar{{background:var(--azulGrad);color:#fff;padding:8px 10px;font-weight:700;font-size:13px;flex-shrink:0;}}
.plan-scroll{{flex:1;overflow:auto;}}
table.pgrid{{border-collapse:collapse;}}
th.corner{{position:sticky;left:0;top:0;z-index:6;background:#f4f6f8;border:1px solid #e3e8ec;min-width:120px;width:120px;font-size:11px;color:#667;padding:4px 8px;}}
th.tec{{position:sticky;left:0;z-index:3;background:#fff;border:1px solid #e3e8ec;padding:5px 8px;font-weight:700;text-align:left;min-width:120px;width:120px;font-size:12px;}}
th.time{{position:sticky;top:0;z-index:4;background:#d9e1e7;border:1px solid #e3e8ec;min-width:56px;width:56px;height:26px;text-align:center;font-size:10px;font-weight:600;white-space:nowrap;}}
td.slot{{border:1px solid #eaecee;min-width:56px;width:56px;height:72px;padding:0;position:relative;vertical-align:top;}}
td.slot.over{{background:#ddeeff;}}
td.slot.prepick-zone{{background:#fffdf0;}}
.gblock{{box-sizing:border-box;position:absolute;top:2px;bottom:2px;left:1px;background:#d6dce5;border-radius:6px;padding:3px 5px;font-size:11px;line-height:1.2;overflow:hidden;cursor:pointer;z-index:2;border:1px solid #b9c2cd;border-left:4px solid var(--azul);box-shadow:0 1px 3px rgba(0,0,0,.08);transition:box-shadow .12s;}}
.gblock:hover{{background:#c5cfd9;box-shadow:0 3px 9px rgba(0,0,0,.16);}}
.gblock b{{color:var(--azulOsc);display:block;font-size:11px;white-space:nowrap;overflow:hidden;text-overflow:ellipsis;}}
.gtime{{font-size:9px;color:#667;white-space:nowrap;}}
.gtemp{{font-size:8px;color:#8a5a00;font-weight:600;white-space:normal;line-height:1.15;}}
.gentrega{{font-size:8px;color:#fff;font-weight:700;background:#c0392b;border-radius:2px;padding:0 3px;display:inline-block;margin-top:1px;white-space:nowrap;max-width:100%;overflow:hidden;text-overflow:ellipsis;}}
.rm-btn{{position:absolute;top:2px;right:2px;background:#e23b3b;color:#fff;border:none;border-radius:50%;width:14px;height:14px;font-size:9px;cursor:pointer;display:flex;align-items:center;justify-content:center;padding:0;line-height:1;}}
/* Leyenda */
.legend{{display:flex;gap:14px;flex-wrap:wrap;padding:6px 0 2px;font-size:11px;align-items:center;}}
.legend .it{{display:flex;align-items:center;gap:5px;}}
.legend .sw{{width:14px;height:14px;border-radius:3px;border:1px solid #999;flex-shrink:0;}}
/* Modal */
.overlay{{position:fixed;inset:0;background:rgba(8,18,32,.5);display:none;align-items:center;justify-content:center;z-index:50;}}
.overlay.open{{display:flex;}}
.modal{{background:#fff;border-radius:14px;width:460px;max-height:88vh;overflow:auto;box-shadow:0 18px 50px rgba(0,0,0,.32);}}
.modal h3{{background:var(--azulGrad);color:#fff;padding:14px 18px;border-bottom:none;border-radius:14px 14px 0 0;font-size:14px;margin:0;}}
.mbody{{padding:16px 18px;}}
.det-g{{display:grid;grid-template-columns:1fr 1fr;gap:6px 16px;}}
.mfoot{{display:flex;justify-content:flex-end;gap:8px;padding:12px 18px;border-top:1px solid #eee;}}
.mbtn{{border:none;border-radius:7px;padding:8px 18px;font-weight:600;cursor:pointer;background:var(--azulGrad);color:#fff;font-size:13px;box-shadow:0 2px 8px rgba(13,47,90,.25);transition:box-shadow .15s,transform .1s;}}
.mbtn:hover{{box-shadow:0 4px 14px rgba(13,47,90,.35);transform:translateY(-1px);}}
.mbtn.danger{{background:linear-gradient(135deg,#7a1620,#c0392b);box-shadow:0 2px 8px rgba(192,57,43,.28);}}
hr{{border:none;border-top:1px solid #eee;margin:10px 0;}}
select,input[type=time],input[type=number]{{border:1px solid #ccc;border-radius:4px;padding:4px 6px;font-size:12px;}}
#toast{{position:fixed;bottom:18px;right:18px;background:#16324f;color:#fff;padding:9px 18px;border-radius:24px;font-size:12px;opacity:0;transition:opacity .3s;z-index:100;pointer-events:none;box-shadow:0 6px 18px rgba(0,0,0,.28);}}
#toast.show{{opacity:1;}}
/* VCU — Hoja Multipuntos Ford */
.vcu-modal{{width:1300px;max-width:97vw;}}
.vcu-badge{{display:inline-block;margin-top:5px;padding:2px 7px;border-radius:3px;font-size:9.5px;font-weight:700;cursor:pointer;}}
.vcu-badge.ok{{background:#147a3d;color:#fff;}}
.vcu-badge.pend{{background:#c0392b;color:#fff;}}
.vcu-chip{{display:inline-block;margin-top:2px;padding:1px 5px;border-radius:3px;font-size:9px;font-weight:700;cursor:pointer;}}
.vcu-chip.ok{{background:#147a3d;color:#fff;}}
.vcu-chip.pend{{background:#c0392b;color:#fff;}}
/* Botones rapidos de Recepcion / Lavado (23/07/2026) — badges persistentes que quedan
   marcadas aunque la orden se mueva despues a otra Etapa. */
.qbadge{{display:inline-block;margin-top:5px;margin-right:4px;padding:2px 7px;border-radius:3px;font-size:9.5px;font-weight:700;color:#fff;}}
.qbadge.recepcion{{background:#c87900;}}
.qbadge.lavado{{background:#0077b6;}}
.vcu-sec{{margin-bottom:10px;border:1px solid #e3e8ec;border-radius:6px;overflow:hidden;}}
.vcu-sec-title{{background:var(--azulGrad);color:#fff;font-weight:700;font-size:11.5px;padding:5px 10px;}}
.vcu-sec-body{{display:grid;grid-template-columns:1fr 1fr;gap:6px 14px;padding:8px 10px;}}
.vcu-field{{display:flex;flex-direction:column;gap:2px;font-size:11px;}}
.vcu-field label{{color:#445;font-weight:600;}}
.vcu-req{{color:#c0392b;}}
.vcu-field select,.vcu-field input,.vcu-field textarea{{width:100%;}}
.vcu-sem-na{{background:#fff;}}
.vcu-sem-verde{{background:#d4f5dd;}}
.vcu-sem-amarillo{{background:#fff3cd;}}
.vcu-sem-rojo{{background:#f8d7da;}}
/* VCU — overlay de campos ENCIMA de la imagen real del PDF Ford (15/07/2026).
   A pedido explicito de Cristobal: no es una recreacion, es el documento escaneado
   real con inputs posicionados de forma absoluta sobre las casillas/lineas reales. */
.vcuf-overlay-wrap{{position:relative;width:100%;margin:0 0 12px;line-height:0;border:1px solid #cdd7e2;border-radius:4px;overflow:hidden;}}
.vcuf-overlay-wrap img{{display:block;width:100%;height:auto;}}
.vcuf-of-box{{position:absolute;margin:0;padding:0;cursor:pointer;z-index:2;
  appearance:none;-webkit-appearance:none;-moz-appearance:none;
  background:rgba(255,255,255,.001);border:1px solid transparent;border-radius:1px;
  box-shadow:0 0 0 1px rgba(20,60,120,.22);
  box-sizing:border-box;outline:none;}}
.vcuf-of-box:hover{{box-shadow:0 0 0 1.5px rgba(20,60,120,.7);}}
.vcuf-of-box:checked{{background:rgba(17,17,17,.62);box-shadow:0 0 0 1.5px #111;}}
.vcuf-of-box:disabled{{cursor:not-allowed;box-shadow:0 0 0 1px rgba(20,60,120,.1);}}
.vcuf-of-box:disabled:checked{{background:rgba(17,17,17,.45);}}
.vcuf-of-text{{position:absolute;border:none;border-bottom:1px solid rgba(26,58,92,.55);border-radius:0;background:rgba(255,255,255,.72);font-size:.85vw;line-height:1.15;padding:0 2px;box-sizing:border-box;z-index:2;font-family:inherit;}}
.vcuf-of-text.vcuf-of-empty{{background:rgba(255,244,214,.85);}}
.vcuf-of-text:focus{{background:#fff;outline:1.5px solid #1a3a5c;}}
.vcuf-of-textarea{{position:absolute;border:1px solid rgba(26,58,92,.35);border-radius:3px;background:rgba(255,255,255,.82);font-size:.85vw;padding:3px 5px;box-sizing:border-box;resize:none;font-family:inherit;z-index:2;}}
.vcuf-of-textarea:focus{{background:#fff;outline:1.5px solid #1a3a5c;}}
.vcuf-of-range{{position:absolute;height:1.7%;margin:0;padding:0;z-index:3;cursor:pointer;accent-color:#111;background:transparent;}}
.vcuf-bat-marker{{position:absolute;width:.35%;background:#111;z-index:2;pointer-events:none;}}
.vcuf-bat-marker-lbl{{position:absolute;font-size:.8vw;font-weight:700;color:#111;background:rgba(255,255,255,.8);padding:0 3px;border-radius:2px;transform:translate(-50%,-100%);white-space:nowrap;z-index:3;pointer-events:none;}}
@media (max-width:700px){{.vcuf-of-text,.vcuf-of-textarea,.vcuf-bat-marker-lbl{{font-size:9px;}}}}
#loading{{padding:40px;text-align:center;color:#667;font-size:15px;}}
/* Control de Taller */
.ct-toolbar{{display:flex;align-items:center;gap:12px;margin-bottom:8px;flex-wrap:wrap;}}
.ct-head-bar{{background:var(--azulGrad);color:#fff;padding:8px 10px;font-weight:700;font-size:13px;border-radius:8px;box-shadow:var(--sombra);}}
.ct-scroll{{overflow:auto;max-height:calc(100vh - 230px);border:1px solid var(--borde);border-radius:10px;box-shadow:var(--sombra);}}
table.ctgrid{{border-collapse:collapse;width:max-content;min-width:100%;font-size:11px;}}
table.ctgrid th{{position:sticky;top:0;z-index:4;background:var(--azulGrad);border:1px solid #0a2650;padding:6px 7px;font-size:10px;font-weight:700;color:#fff;white-space:nowrap;}}
table.ctgrid td{{border:1px solid #e3e8ec;padding:4px 5px;white-space:nowrap;vertical-align:middle;}}
table.ctgrid tbody tr{{transition:filter .1s;}}
table.ctgrid tbody tr:hover{{filter:brightness(.95);}}
table.ctgrid tr.ct-salida{{opacity:.55;}}
table.ctgrid tr.ct-marca-sep td{{background:var(--azulGrad);color:#fff;font-weight:700;font-size:11.5px;padding:6px 8px;text-align:left;letter-spacing:.3px;white-space:nowrap;}}
table.ctgrid input[type=text],table.ctgrid input[type=number],table.ctgrid input[type=date]{{border:1px solid #ccc;border-radius:4px;padding:4px 5px;font-size:11px;width:88px;background:#fff;}}
table.ctgrid input.ct-wide{{width:150px;}}
table.ctgrid select{{border:1px solid #ccc;border-radius:4px;padding:4px 5px;font-size:11px;background:#fff;}}
table.ctgrid select.et-select{{font-size:10.5px;padding:4px 3px;}}
table.ctgrid td.ct-pat{{font-weight:700;color:var(--azulOsc);position:sticky;left:0;z-index:2;font-size:12.5px;}}
table.ctgrid td.ct-dias{{text-align:center;color:#445;font-weight:700;}}
/* Ingreso Taller / Salida Taller — destacadas: son las que alimentan el Planificador
   de Tecnicos y el Tiempo Estimado, a diferencia de Horario Ingreso/Entrega (informativos) */
table.ctgrid th.ct-th-taller{{background:#1e8449!important;}}
table.ctgrid td.ct-td-taller{{background:#eafaf1;}}
table.ctgrid input.ct-input-taller{{border:2px solid #1e8449;background:#fff;color:#145a32;font-weight:700;box-shadow:0 0 0 1px rgba(30,132,73,.25);}}
.prod-select,.prod-input{{border:1px solid var(--borde);border-radius:4px;padding:5px 8px;font-size:12px;background:#fff;}}
.prod-input{{width:180px;}}
.prod-total{{margin-top:8px;font-weight:700;color:var(--azulOsc);font-size:13px;}}
.prod-stats{{display:flex;gap:10px;flex-wrap:wrap;margin:10px 0;}}
.prod-kpi{{background:#eef3fb;border:1px solid var(--borde);border-radius:10px;padding:8px 14px;text-align:center;min-width:120px;box-shadow:0 1px 6px rgba(13,47,90,.08);}}
.prod-kpi b{{display:block;font-size:16px;color:var(--azulOsc);}}
.prod-kpi span{{font-size:10.5px;color:#667;}}
table.ctgrid tr.prod-row{{cursor:pointer;}}
table.ctgrid tr.prod-row:hover{{background:#eef3fb;}}
table.ctgrid tr.prod-row.sel{{background:#d6e6fb;}}
.prod-detalle{{margin-top:12px;border:1px solid var(--borde);border-radius:12px;padding:12px 16px;background:#fbfdff;box-shadow:var(--sombra);}}
.prod-det-head{{font-weight:700;color:var(--azulOsc);font-size:14px;margin-bottom:8px;display:flex;justify-content:space-between;align-items:center;}}
.prod-det-close{{border:none;background:#eee;border-radius:4px;padding:3px 10px;cursor:pointer;font-size:11px;}}
.prod-kpis-det{{display:flex;gap:10px;flex-wrap:wrap;margin-bottom:10px;}}
.prod-det-sub{{font-weight:700;color:var(--azulOsc);font-size:12px;margin:10px 0 4px;}}
/* Productividad: fraccion vendidas/disponibles bajo el % y disponibilidad del tecnico */
.prod-frac{{font-size:10px;color:#778;line-height:1.2;margin-top:1px;}}
.nodisp-badge{{display:inline-block;background:#fdecea;color:#a3302a;border:1px solid #f0b3ad;
  border-radius:10px;padding:1px 7px;font-size:10px;font-weight:700;margin-left:4px;white-space:nowrap;}}
.nodisp-btn{{background:#eef3fb;border:1px solid var(--borde);border-radius:4px;padding:3px 8px;
  font-size:11px;cursor:pointer;color:var(--azulOsc);white-space:nowrap;}}
.nodisp-btn:hover{{background:#dde7f6;}}
.nodisp-btn.add{{background:linear-gradient(135deg,#0d5c2c,#1b7f3a);border-color:#1b7f3a;color:#fff;font-weight:700;padding:6px 12px;box-shadow:0 2px 8px rgba(27,127,58,.25);}}
.nodisp-btn.add:hover{{background:#166a31;}}
.nodisp-form{{display:flex;gap:10px;align-items:flex-end;flex-wrap:wrap;}}
.nodisp-form label{{display:flex;flex-direction:column;gap:3px;font-size:11px;font-weight:700;color:#0b2e63;}}
.nodisp-form .prod-input{{width:150px;}}
.prod-det-empty{{color:#889;font-size:12px;padding:8px 0;}}
table.ctgrid.prod-mini{{width:100%;min-width:0;}}
.prod-ot-flex{{display:flex;gap:14px;align-items:flex-start;flex-wrap:wrap;}}
.prod-ot-tabla{{flex:1 1 320px;min-width:0;}}
.prod-ot-chart{{flex:0 1 280px;min-width:220px;background:#fbfdff;border:1px solid var(--borde);border-radius:6px;padding:8px;}}
.ct-del{{border:none;border-radius:3px;background:#e23b3b;color:#fff;cursor:pointer;padding:3px 7px;font-size:11px;}}
/* Pre-picking */
.pp-toolbar{{margin-bottom:8px;display:flex;align-items:center;justify-content:space-between;gap:12px;flex-wrap:wrap;}}
.pp-btn-xls{{background:linear-gradient(135deg,#0d5c2c,#147a3d);color:#fff;border:none;border-radius:8px;padding:8px 14px;font-weight:600;cursor:pointer;font-size:12px;white-space:nowrap;box-shadow:0 2px 8px rgba(20,122,61,.25);}}
.pp-btn-xls:hover{{background:#106631;}}
.pp-cards{{display:flex;flex-direction:column;gap:8px;}}
.pp-card{{border:1px solid #c5cdd5;border-left:6px solid #889;border-radius:12px;background:#fff;overflow:hidden;box-shadow:0 1px 6px rgba(0,0,0,.06);}}
.pp-card.pp-pendiente{{border-left-color:#c87900;}}
.pp-card.pp-realizado{{border-left-color:#147a3d;background:#f4faf6;}}
.pp-head{{display:flex;align-items:center;gap:14px;padding:10px 14px;cursor:pointer;flex-wrap:wrap;}}
.pp-head:hover{{background:#f8fafc;}}
.pp-hora{{font-weight:700;color:#0b2e63;font-size:13px;min-width:50px;}}
.pp-plate{{font-weight:800;font-size:16px;color:#1a1a1a;background:#eef2f6;border:1px solid #cfd6dd;border-radius:4px;padding:2px 8px;}}
.pp-mmv{{color:#334;font-size:12px;}}
.pp-svc{{color:#556;font-size:12px;}}
.pp-cliente{{color:#778;font-size:11px;margin-left:auto;}}
.pp-status-badge{{font-size:10.5px;font-weight:700;border-radius:10px;padding:2px 10px;}}
.pp-status-badge.pendiente{{background:#fff3cd;color:#8a5a00;}}
.pp-status-badge.realizado{{background:#d4edda;color:#147a3d;}}
.pp-chevron{{color:#889;font-size:12px;}}
.pp-body{{display:none;border-top:1px solid #e3e8ec;padding:14px;background:#fafbfc;}}
.pp-body.open{{display:block;}}
.pp-det-grid{{display:grid;grid-template-columns:1fr 1fr 1fr;gap:6px 18px;margin-bottom:14px;}}
.pp-modelo-sel{{background:#f3f6fa;border:1px solid #dde5ec;border-radius:10px;padding:9px 12px;margin-bottom:12px;}}
.pp-modelo-sel-tit{{font-size:10.5px;color:#556;font-weight:700;margin-bottom:6px;text-transform:uppercase;letter-spacing:.3px;}}
.pp-modelo-auto{{color:#147a3d;font-weight:600;text-transform:none;letter-spacing:0;}}
.pp-modelo-manual{{color:#c87900;font-weight:600;text-transform:none;letter-spacing:0;}}
.pp-modelo-sel-row{{display:flex;gap:8px;flex-wrap:wrap;align-items:center;}}
.pp-modelo-sel-row select{{flex:1;min-width:150px;padding:5px 7px;border:1px solid #ccd4dc;border-radius:5px;font-size:11.5px;color:#223;background:#fff;}}
.pp-btn-reset-modelo{{background:none;border:1px solid #ccd4dc;border-radius:5px;padding:5px 10px;font-size:11px;color:#556;cursor:pointer;}}
.pp-det-row .pp-lbl{{color:#889;font-size:10.5px;display:block;}}
.pp-det-row .pp-val{{color:#223;font-size:12.5px;font-weight:600;}}
table.pptable{{border-collapse:collapse;width:100%;font-size:11.5px;margin-bottom:12px;}}
table.pptable th{{background:var(--azulGrad);color:#fff;padding:6px 8px;text-align:left;font-size:10.5px;}}
table.pptable td{{border:1px solid #e3e8ec;padding:5px 8px;}}
table.pptable tr.pp-mo-row td{{background:#f8d7da;font-weight:700;color:#7a1f1f;}}
table.pptable tr.pp-tot-row td{{background:#0b2e63;color:#fff;font-weight:700;font-size:12.5px;}}
table.pptable tr.pp-desc-row td{{background:#fff3cd;color:#8a5a00;font-weight:700;}}
table.pptable tr.pp-tot-desc td{{background:#147a3d;color:#fff;}}
.pp-alt{{margin-top:5px;padding-top:5px;border-top:1px dashed #d8dee3;}}
.pp-alt-lbl{{font-size:9.5px;color:#889;text-transform:uppercase;letter-spacing:.3px;margin-bottom:3px;}}
.pp-alt-item{{display:flex;align-items:baseline;gap:6px;font-size:10.5px;color:#556;line-height:1.5;}}
.pp-alt-item b{{color:#0b2e63;font-size:11px;flex-shrink:0;}}
.pp-alt-item .pp-alt-desc{{color:#778;}}
.pp-alt-item .pp-alt-stock{{margin-left:auto;flex-shrink:0;font-weight:600;white-space:nowrap;}}
.pp-alt-item .pp-alt-stock.si{{color:#147a3d;}}
.pp-alt-item .pp-alt-stock.otro{{color:#c87900;}}
.pp-alt-item .pp-alt-stock.no{{color:#a33;}}
.pp-actions{{display:flex;gap:8px;flex-wrap:wrap;}}
.pp-actions button{{border:none;border-radius:8px;padding:8px 14px;font-weight:600;cursor:pointer;font-size:12px;transition:box-shadow .15s,transform .1s;}}
.pp-actions button:hover{{transform:translateY(-1px);}}
.pp-btn-pdf{{background:var(--azulGrad);color:#fff;box-shadow:0 2px 8px rgba(13,47,90,.25);}}
.pp-btn-real{{background:linear-gradient(135deg,#0d5c2c,#147a3d);color:#fff;box-shadow:0 2px 8px rgba(20,122,61,.25);}}
.pp-btn-pend{{background:linear-gradient(135deg,#8a5600,#c87900);color:#fff;box-shadow:0 2px 8px rgba(200,121,0,.25);}}
.pp-btn-desc{{background:#eef1f4;color:#334;border:1px solid #ccd4dc !important;}}
.pp-btn-desc.activo{{background:#c0392b;color:#fff;border:1px solid #c0392b !important;}}
/* Modal detalle cita */
#cita-modal-overlay{{display:none;position:fixed;inset:0;background:rgba(0,0,0,.5);z-index:9000;align-items:center;justify-content:center;}}
#cita-modal-overlay.open{{display:flex;}}
#cita-modal{{background:#fff;border-radius:14px;padding:24px 28px;min-width:300px;max-width:420px;width:90%;box-shadow:0 18px 50px rgba(0,0,0,.3);position:relative;}}
#cita-modal h3{{margin:0 0 16px;color:var(--azulOsc);font-size:16px;border-bottom:2px solid #e0e9f5;padding-bottom:8px;}}
.cm-row{{display:flex;gap:8px;margin-bottom:8px;font-size:13px;}}
.cm-lbl{{color:#888;min-width:90px;flex-shrink:0;}}
.cm-val{{color:#222;font-weight:600;word-break:break-all;}}
.cm-close{{position:absolute;top:12px;right:14px;background:none;border:none;font-size:20px;cursor:pointer;color:#888;line-height:1;}}
</style>
</head>
<body>
<div id="cita-modal-overlay" onclick="if(event.target===this)cerrarDetalleCita()">
  <div id="cita-modal">
    <button class="cm-close" onclick="cerrarDetalleCita()">✕</button>
    <h3 id="cm-title">Detalle de cita</h3>
    <div id="cm-body"></div>
  </div>
</div>
<div id="loading">⏳ Cargando agenda y tablero de taller...</div>
<div id="main" style="display:none">

  <div class="subnav">
    <button class="sntab active" onclick="switchView('jpcb',this)">📋 JPCB — Avance del trabajo</button>
    <button class="sntab" onclick="switchView('plan',this)">📅 Planificador (Tecnico x Hora)</button>
    <button class="sntab" onclick="switchView('ct',this)">🗂️ Control de Taller</button>
    <button class="sntab" onclick="switchView('vt',this)">🅿️ Vehiculos en Taller</button>
    <button class="sntab" onclick="switchView('hist',this)">📜 Historial de Taller</button>
    <button class="sntab" onclick="switchView('prod',this)">📊 Producción Técnicos</button>
    {_pp_tab_btn}
    <span class="save-st" id="saveStatus"></span>
  </div>

  <div class="area-bar">
    <span class="area-lbl">🗂️ Área de trabajo:</span>
    <button class="area-tab active" data-area="st" onclick="setArea('st',this)">🔧 Servicio Técnico</button>
    <button class="area-tab" data-area="dyp" onclick="setArea('dyp',this)">🎨 Desabolladura y Pintura</button>
    <span class="area-hint">Se detecta sola desde el Servicio/Comentario de la Agenda — afecta JPCB, Planificador, Control de Taller, Vehículos en Taller e Historial.</span>
  </div>

  <!-- JPCB VIEW -->
  <div id="v-jpcb">
    <div class="kanban-groups" id="jpcbGroups"></div>
    <div class="kanban" id="jpcbBoard"></div>
    <div class="stopzone">
      <div class="sz-head">⛔ Detencion del trabajo — arrastra aqui si hay una irregularidad</div>
      <div class="kanban" id="stopBoard" style="padding:8px"></div>
    </div>
    <div class="stopzone" style="border-color:#1b7f3a;margin-top:10px">
      <div class="sz-head" style="background:#1b7f3a">✅ Finalizados de esta semana — se limpian solos del JPCB la semana siguiente (quedan siempre en el Historial de Taller)</div>
      <div id="finBoard" style="padding:8px;display:flex;flex-wrap:wrap;gap:8px"></div>
    </div>
  </div>

  <!-- PLANIFICADOR VIEW -->
  <div id="v-plan" style="display:none">
    <div class="dtabs" id="dateTabs"></div>
    <div class="plan-board">
      <div class="prog-panel">
        <div class="prog-head-bar">📋 Programacion — {sucursal}</div>
        <div id="progList" style="flex:1;overflow-y:auto;"></div>
      </div>
      <div class="plan-panel">
        <div class="plan-head-bar">🔧 Planificador — {sucursal}</div>
        <div class="plan-scroll">
          <table class="pgrid" id="planGrid"></table>
        </div>
      </div>
    </div>
    <div class="legend" id="planLegend"></div>
  </div>

  <!-- CONTROL DE TALLER VIEW -->
  <div id="v-ct" style="display:none">
    <div class="ct-toolbar">
      <div class="ct-head-bar">🗂️ Control de Taller — {sucursal}</div>
      <button class="mbtn" onclick="agregarPatenteManual()">➕ Agregar patente manual</button>
      <button class="mbtn" onclick="agregarAsesorManual()" title="Agrega un asesor a la lista aunque todavia no tenga citas en la Agenda">➕ Agregar asesor</button>
      <span style="font-size:11px;color:#778;">Las patentes con cita marcada 🎟️ Ingresado en la Agenda de hoy se agregan solas. La lista de Asesor se arma sola desde la Agenda — usa "➕ Agregar asesor" si falta alguno.</span>
    </div>
    <div class="legend" id="ctLegend"></div>
    <div class="ct-scroll">
      <table class="ctgrid" id="ctGrid"></table>
    </div>
  </div>

  <!-- VEHICULOS EN TALLER VIEW (detenidos, aun no retirados) -->
  <div id="v-vt" style="display:none">
    <div class="ct-toolbar">
      <div class="ct-head-bar">🅿️ Vehiculos en Taller — {sucursal}</div>
      <span style="font-size:11px;color:#778;">
        🎟️ En taller (ingresado, sin marcar salida) · 🧍 Retirado · Ventana: ingresos de los ultimos 60 dias.
        Se excluyen patentes de prueba tipo SP0000.
      </span>
    </div>
    <div class="legend" id="vtLegend"></div>
    <div class="ct-scroll">
      <table class="ctgrid" id="vtGrid"></table>
    </div>
  </div>

  <!-- HISTORIAL DE TALLER VIEW (citas cerradas) -->
  <div id="v-hist" style="display:none">
    <div class="ct-toolbar">
      <div class="ct-head-bar">📜 Historial de Taller — {sucursal}</div>
      <span style="font-size:11px;color:#778;">
        Ordenes con "🔒 Cerrar" presionado — ya no ocupan espacio en JPCB, Control de Taller,
        Vehiculos en Taller ni Tecnico x Hora. Se pueden reabrir si fue un error.
      </span>
    </div>
    <div class="ct-scroll">
      <table class="ctgrid" id="histGrid"></table>
    </div>
  </div>

  <!-- PRODUCCION TECNICOS VIEW (horas facturadas por tecnico, BDFlexline) -->
  <div id="v-prod" style="display:none">
    <div class="ct-toolbar">
      <div class="ct-head-bar">📊 Producción Técnicos — {sucursal}</div>
      <label style="font-size:12px;font-weight:700;color:#0b2e63;">Mes:
        <select id="prodMes" class="prod-select"></select>
      </label>
      <input id="prodBuscarTec" type="text" placeholder="Buscar técnico..." class="prod-input" oninput="renderProdTabla()">
      <span style="font-size:11px;color:#778;">
        Solo cuenta como hora la Mano de Obra (código "MO_..."). Haz clic en un técnico para ver su detalle.
        <b>% Productividad = horas vendidas ÷ horas disponibles acumuladas</b> (horas del día × días hábiles
        transcurridos, el día de hoy incluido). Jornada Lun-Mar 08:30-18:15 (9,0 h) · Mié-Vie 08:30-17:15
        (8,0 h), ya descontados 45 min de colación, sobre días hábiles del calendario chileno (aproximado,
        no incluye feriados extraordinarios como elecciones). Los días marcados como
        <b>no disponible</b> (vacaciones, licencia, permiso, capacitación) no suman horas disponibles.
      </span>
    </div>
    <div id="prodStats" class="prod-stats"></div>
    <div id="prodDiagNombres"></div>
    <div class="ct-scroll">
      <table class="ctgrid" id="prodGrid">
        <thead><tr><th>Técnico</th><th style="text-align:right">Horas Vendidas</th><th style="text-align:right"># OT</th><th style="text-align:right">% Productividad<div class="prod-frac" style="font-weight:400">vendidas / disponibles</div></th><th style="text-align:center">Disponibilidad</th></tr></thead>
        <tbody id="prodTbody"></tbody>
      </table>
    </div>
    <div class="prod-total">Total horas del mes: <span id="prodTotalHoras">0.0</span> h</div>
    <div id="prodDetalle" class="prod-detalle" style="display:none"></div>
  </div>

  <!-- Modal de Disponibilidad de un tecnico (solo con permiso PUEDE_DISPONIBILIDAD) -->
  <div class="overlay" id="nodisp-modal-overlay" onclick="if(event.target===this)cerrarNoDisponible()">
    <div class="modal" style="max-width:560px">
      <div id="nodispBody"></div>
    </div>
  </div>

  <!-- PRE-PICKING VIEW (una tarjeta por cita, detalle + repuestos + PDF + estado) -->
  <div id="v-pp" style="display:none">
    <div class="dtabs" id="ppDateTabs"></div>
    <div class="pp-toolbar">
      <span style="font-size:11px;color:#778;">
        🧩 Preparación anticipada de mantenciones — clic en una tarjeta para ver el detalle completo,
        el listado de repuestos sugeridos (marca/modelo/kilometraje según Agenda) y exportar el presupuesto.
      </span>
      <button class="pp-btn-xls" onclick="exportarListadoPrepickingExcel()">📊 Exportar listado a Excel</button>
    </div>
    <div id="ppCards" class="pp-cards"></div>
  </div>

</div>

<div class="overlay" id="modal-overlay" onclick="if(event.target===this)closeModal()">
  <div class="modal">
    <h3 id="modal-title">Detalle</h3>
    <div class="mbody" id="modal-body"></div>
    <div class="mfoot">
      <button class="mbtn danger" id="modal-del-btn" style="display:none" onclick="deleteBloque(modalBloqueId)">🗑 Quitar</button>
      <button class="mbtn" onclick="closeModal()">Cerrar</button>
    </div>
  </div>
</div>
<div class="overlay" id="vcu-modal-overlay" onclick="if(event.target===this)cerrarVCU()">
  <div class="modal vcu-modal">
    <h3 id="vcu-modal-title">VCU</h3>
    <div id="vcu-modal-estado" style="padding:0 16px 8px;font-size:11.5px;font-weight:700"></div>
    <div class="mbody" id="vcu-modal-body"></div>
    <div class="mfoot" style="flex-wrap:wrap">
      <button class="mbtn" style="background:#667" onclick="vcuGuardar(false)">💾 Guardar borrador</button>
      <button class="mbtn" style="background:#147a3d" onclick="vcuGuardar(true)">✅ Marcar completo y guardar</button>
      <button class="mbtn" style="background:#0b2e63" onclick="vcuDescargarPDF()">📄 Descargar PDF</button>
      <button class="mbtn" onclick="cerrarVCU()">Cerrar</button>
    </div>
  </div>
</div>
<div id="toast"></div>

<script>
const GITHUB_TOKEN  = "{token_safe}";
const API_BASE      = "{api_base}";
// Acceso a Supabase para el tablero. SB_VALE es un permiso mínimo: dura una
// jornada, es de este usuario, y solo sirve para los dos documentos del
// tablero. Reemplaza al token de GitHub, que daba escritura sobre TODO el
// repositorio y viajaba dentro de esta misma página. SB_KEY es la clave
// pública de Supabase: por sí sola no abre nada (todo tiene RLS).
const SB_URL        = "{sb_url}";
const SB_KEY        = "{sb_key}";
const SB_VALE       = "{sb_vale}";
const USA_SUPABASE  = !!(SB_URL && SB_KEY && SB_VALE);

async function sbTablero(fn, args){{
  const r = await fetch(SB_URL + '/rest/v1/rpc/' + fn, {{
    method: 'POST',
    headers: {{'apikey': SB_KEY, 'Authorization': 'Bearer ' + SB_KEY,
               'Content-Type': 'application/json'}},
    body: JSON.stringify(args)}});
  if(!r.ok) throw new Error('HTTP ' + r.status);
  return await r.json();
}}
const SUCURSAL      = "{sucursal}";
const USUARIO       = "{usuario}";
const PUEDE_EDITAR  = {puede_str};
const PUEDE_CONFIRMAR_CITAS = {puede_confirmar_str};
const PUEDE_DISPONIBILIDAD = {puede_disp_str};
const PUEDE_PREPICK = {puede_pp_str};
const LOGO_URI      = "{_logo_safe}";
const _AGENDA_INIT  = {_agenda_js};
const _CTRL_INIT    = {_ctrl_js};
const _CTRL_SHA     = "{_sha_safe}";
const _PP_INIT      = {_pp_js};
const _PP_SHA       = "{_pp_sha_safe}";
const _PROD_INIT    = {_prod_js};
const _COTIZ_GZ     = "{_cotiz_gz_safe}";
const _STOCK_GZ     = "{_stock_gz_safe}";

// Bundle del Cotizador de Mantenciones (indice/stock/pautas), descomprimido en el
// navegador desde _COTIZ_GZ (gzip+base64, mismo mecanismo que el modulo standalone
// "Cotizador de Mantenciones"). Se usa en Pre-picking para sacar repuestos+stock
// directamente de las pautas del cotizador en vez del pipeline viejo del
// consolidador (pauta_repuestos.json + Stock Repestos Costo.xlsx). 22/07/2026.
let COTIZ_PP = null;
(async function(){{
  if(!_COTIZ_GZ) return;
  try{{
    if(typeof DecompressionStream==='undefined') return;
    const bin=Uint8Array.from(atob(_COTIZ_GZ), c=>c.charCodeAt(0));
    const ds=new DecompressionStream('gzip');
    const stream=new Blob([bin]).stream().pipeThrough(ds);
    const buf=await new Response(stream).arrayBuffer();
    const pkg=JSON.parse(new TextDecoder().decode(buf));
    COTIZ_PP={{indice:pkg.indice, stock:pkg.stock, pautas:pkg.pautas}};
    if(typeof renderPrepickingView==='function' && document.getElementById('ppCards')) renderPrepickingView();
  }}catch(e){{ console.warn('No se pudo cargar el bundle del Cotizador para Pre-picking', e); }}
}})();

// Catalogo COMPLETO de Stock de Repuestos (stock_repuestos.json, ~33.000
// filas producto+bodega), descomprimido en el navegador desde _STOCK_GZ y
// agrupado por codigo normalizado -> {{descripcion, bodegas:[{{n,q}}]}}. Se usa
// SOLO para ampliar las alternativas "tambien sirve" de Pre-picking (busqueda
// de codigo relacionado contra todo el catalogo real, no solo los ~400
// codigos acotados del bundle del Cotizador) y como respaldo de stock si un
// codigo de la pauta no esta en ese bundle chico. 22/07/2026.
let STOCK_FULL = null;
(async function(){{
  if(!_STOCK_GZ) return;
  try{{
    if(typeof DecompressionStream==='undefined') return;
    const bin=Uint8Array.from(atob(_STOCK_GZ), c=>c.charCodeAt(0));
    const ds=new DecompressionStream('gzip');
    const stream=new Blob([bin]).stream().pipeThrough(ds);
    const buf=await new Response(stream).arrayBuffer();
    const productos=JSON.parse(new TextDecoder().decode(buf));
    const idx={{}};
    for(const p of productos){{
      // OJO: el codigo crudo de Stock Repestos Costo.xlsx trae un prefijo
      // numerico de familia pegado con espacio (ej. "13 XO5W30Q1SP") — hay
      // que quitarlo (_cotizCodBase) ANTES de normalizar, si no el prefijo
      // queda pegado al codigo real y nunca calza con el codigo limpio que
      // trae la pauta del Cotizador (ej. "XO5W30Q1SP"). 22/07/2026.
      const cod=_cotizCodBase(p.p);
      if(!cod) continue;
      if(!idx[cod]) idx[cod]={{descripcion:p.d||'', bodegas:[]}};
      idx[cod].bodegas.push({{n:p.b||'', q:Number(p.s)||0}});
    }}
    STOCK_FULL=idx;
    if(typeof renderPrepickingView==='function' && document.getElementById('ppCards')) renderPrepickingView();
  }}catch(e){{ console.warn('No se pudo cargar el catalogo completo de Stock para Pre-picking', e); }}
}})();

const START=8*60+30,STEP=30,COLW=56;
const ETAPAS=[
  {{id:"recepcion",      t:"1 · Recepcion",       color:"#1b6ec2", bg:"#eaf3fc"}},
  {{id:"ingreso_taller", t:"2 · Ingreso Taller",  color:"#c87900", bg:"#fff3e0"}},
  {{id:"en_proceso",     t:"3 · En Proceso",      color:"#147a3d", bg:"#eafaf0"}},
  {{id:"en_prueba",      t:"4 · En Prueba",       color:"#6c4fc4", bg:"#f2eefc"}},
  {{id:"lavado",         t:"5 · Lavado",          color:"#0077b6", bg:"#e8f4fb"}},
  {{id:"entrega",        t:"6 · Entrega",         color:"#0d5c2e", bg:"#e3f6ea"}},
];
const etapaInfo=id=>ETAPAS.find(e=>e.id===id)||{{color:"#999",bg:"#fff"}};
const STOPS=[
  {{id:"decision", t:"Esperando diagnostico"}},
  {{id:"aprob",    t:"Esperando aprobacion"}},
  {{id:"repuestos",t:"Esperando repuestos"}},
  {{id:"terceros", t:"Terceros (sublet)"}},
];
// Estado de Campaña (garantia/recall) — 14/07/2026, a pedido de Cristobal. Los 3
// estados de "salida" (Quiebre Stock/Cliente desiste/Falla servidor) sacan la orden
// del tablero JPCB (sigue visible en Control de Taller/Vehiculos en Taller); "Realizada"
// no la oculta, solo la marca.
const ESTADOS_CAMPANA=[
  {{id:"realizada",       t:"Realizada"}},
  {{id:"quiebre_stock",   t:"Quiebre Stock"}},
  {{id:"cliente_desiste", t:"Cliente desiste"}},
  {{id:"falla_servidor",  t:"Falla servidor"}},
];
const _ESTADOS_CAMPANA_OCULTAN_JPCB=["quiebre_stock","cliente_desiste","falla_servidor"];
// Responsable de cada tramo del JPCB (14/07/2026, a pedido de Cristobal) — solo es una
// fila de titulos informativos sobre el tablero, no cambia ninguna logica de datos.
// "_no_asiste_" es un id ficticio para la columna No asiste (no es una Etapa real).
// "_citas_" (23/07/2026) es otro id ficticio para la columna nueva "Citas <fecha>" —
// las citas de la Agenda aterrizan ahi primero (sin confirmar) antes de pasar a
// Recepcion; confirmar asistencia es tarea del Asesor, igual que No Asiste.
const GRUPOS_RESPONSABLE=[
  {{label:"Asesor",        ids:["_citas_","_no_asiste_"]}},
  {{label:"Torre Control",  ids:["recepcion","ingreso_taller"]}},
  {{label:"Tecnico",        ids:["en_proceso","en_prueba","lavado"]}},
  {{label:"Asesor",        ids:["entrega"]}},
];
const IDX_EN_PROCESO=ETAPAS.findIndex(e=>e.id==='en_proceso');

// =============================================================
// VCU — Hoja Multipuntos Ford: imagenes reales del documento oficial (15/07/2026).
// A pedido explicito de Cristobal, el formulario NO es una recreacion en HTML: se
// dibujan los campos ENCIMA de la imagen escaneada real de las 2 paginas del PDF que
// exige la marca. Las imagenes se hospedan en GitHub (vcu_ford_p1.jpg/p2.jpg).
const VCU_IMG_P1='{VCU_FORD_P1_URL}';
const VCU_IMG_P2='{VCU_FORD_P2_URL}';

// =============================================================
// VCU — Hoja Multipuntos Ford (14/07/2026, a pedido de Cristobal)
// Solo aplica a vehiculos Ford. Obligatorio completar antes de avanzar
// el JPCB mas alla de la etapa "3 - En Proceso". Datos guardados en
// ctrlData[SUCURSAL].vcu[ordenId] = {{datos, completo, tecnico, fecha}}.
// =============================================================
const VCU_SCHEMA=[
  {{section:"Datos del Vehiculo", fields:[
    {{id:"fecha",   label:"Fecha",        type:"date", req:true}},
    {{id:"or",      label:"N° OR",        type:"text", req:true}},
    {{id:"linea",   label:"Linea",        type:"text", req:false}},
    {{id:"modelo",  label:"Modelo",       type:"text", req:true}},
    {{id:"vin",     label:"N° de Serie / VIN", type:"text", req:true}},
  ]}},
  {{section:"Niveles de Fluidos (Asesor)", fields:[
    {{id:"fl_fugas",        label:"Fugas visibles",                       type:"sino", req:true}},
    {{id:"fl_aceite_motor", label:"Aceite Motor",                          type:"sino", req:true}},
    {{id:"fl_fluido_freno", label:"Fluido de Freno",                       type:"sino", req:true}},
    {{id:"fl_embrague",     label:"Revision Embrague",                     type:"sino", req:true}},
    {{id:"fl_dir_hid",      label:"Direccion Hidraulica",                  type:"sino", req:true}},
    {{id:"fl_limpiaparab",  label:"Nivel Deposito Limpiaparabrisas",       type:"sino", req:true}},
    {{id:"fl_lineas_comb",  label:"Revision lineas de Combustible",        type:"sino", req:true}},
    {{id:"fl_transmision",  label:"Transmision",                           type:"sino", req:true}},
    {{id:"fl_refrigerante", label:"Deposito recuperacion Refrigerante",    type:"sino", req:true}},
    {{id:"fl_diferencial",  label:"Revision Fugas Diferencial",            type:"sino", req:true}},
  ]}},
  {{section:"Plumillas / Luces / Cristales (Asesor)", fields:[
    {{id:"plumillas",       label:"Plumillas",                              type:"sino", req:true}},
    {{id:"luces",           label:"Luces",                                  type:"sino", req:true}},
    {{id:"parabrisas",      label:"Parabrisas",                             type:"sino", req:true}},
    {{id:"cristales",       label:"Cristales",                              type:"sino", req:true}},
  ]}},
  {{section:"Bateria (Asesor)", fields:[
    {{id:"bat_estado",      label:"Estado de la Bateria",                   type:"semaforo", req:true}},
    {{id:"bat_nivel_carga", label:"Nivel de carga de Bateria",              type:"trislider", req:false}},
    {{id:"bat_cca_real",    label:"CCA real",                               type:"text", req:true}},
    {{id:"bat_cca_fabrica", label:"CCA de fabrica",                         type:"text", req:true}},
    {{id:"bat_recuperacion",label:"Recuperacion",                           type:"sino", req:false}},
  ]}},
  {{section:"Codigos de Falla (Asesor)", fields:[
    {{id:"cod_verificacion",label:"Verificacion de Codigos",  type:"radio", req:true,
      opts:[["ok","Sin codigos"],["pendiente","Con codigos pendientes"]]}},
    {{id:"cod_relenti",     label:"Funcionamiento de motor en relenti", type:"radio", req:true,
      opts:[["normal","Normal"],["anormal","Anormal"]]}},
  ]}},
  {{section:"Correas / Mangueras (Tecnico)", fields:[
    {{id:"correa_accesorios", label:"Correa de accesorios",   type:"semaforo", req:true}},
    {{id:"mangueras_motor",   label:"Mangueras de motor",     type:"semaforo", req:true}},
    {{id:"mangueras_refrig",  label:"Mangueras de refrigeracion", type:"semaforo", req:true}},
  ]}},
  {{section:"Sistema de Frenos (Tecnico)", fields:[
    {{id:"frenos_sistema",    label:"Sistema de frenos completo", type:"semaforo", req:true}},
  ]}},
  {{section:"Direccion / Suspension (Tecnico)", fields:[
    {{id:"direccion",         label:"Sistema de direccion",   type:"semaforo", req:true}},
    {{id:"suspension",        label:"Sistema de suspension",  type:"semaforo", req:true}},
  ]}},
  {{section:"Sistema de Escape (Tecnico)", fields:[
    {{id:"escape",            label:"Sistema de escape",      type:"semaforo", req:true}},
  ]}},
  {{section:"Tren Motriz (Tecnico)", fields:[
    {{id:"tren_motriz_del",   label:"Tren motriz delantero",  type:"semaforo", req:true}},
    {{id:"tren_motriz_tra",   label:"Tren motriz trasero",    type:"semaforo", req:true}},
  ]}},
  {{section:"Aire Acondicionado (Tecnico)", fields:[
    {{id:"ac_funcionamiento", label:"Funcionamiento A/C",     type:"semaforo", req:true}},
    {{id:"ac_filtro_cabina",  label:"Filtro de cabina",       type:"semaforo", req:true}},
  ]}},
  {{section:"Filtros (Tecnico)", fields:[
    {{id:"filtro_aire",       label:"Filtro de aire",         type:"semaforo", req:true}},
    {{id:"filtro_combustible",label:"Filtro de combustible",  type:"semaforo", req:true}},
  ]}},
  {{section:"Parte Inferior del Vehiculo (Tecnico)", fields:[
    {{id:"parte_inferior_obs",label:"Observaciones parte inferior", type:"textarea", req:false}},
  ]}},
  {{section:"Neumatico Delantero Izquierdo", fields:[
    {{id:"ndi_labrado",   label:"Profundidad de labrado (mm)",  type:"semaforo_num", req:true}},
    {{id:"ndi_desgaste",  label:"Patron de desgaste / dano",    type:"text", req:true}},
    {{id:"ndi_presion",   label:"Presion de inflado (PSI)",     type:"text", req:true}},
    {{id:"ndi_pastillas", label:"Espesor de Pastillas (mm)",    type:"semaforo_num", req:true}},
    {{id:"ndi_disco",     label:"Espesor de Disco (mm)",        type:"semaforo_num", req:true}},
  ]}},
  {{section:"Neumatico Delantero Derecho", fields:[
    {{id:"ndd_labrado",   label:"Profundidad de labrado (mm)",  type:"semaforo_num", req:true}},
    {{id:"ndd_desgaste",  label:"Patron de desgaste / dano",    type:"text", req:true}},
    {{id:"ndd_presion",   label:"Presion de inflado (PSI)",     type:"text", req:true}},
    {{id:"ndd_pastillas", label:"Espesor de Pastillas (mm)",    type:"semaforo_num", req:true}},
    {{id:"ndd_disco",     label:"Espesor de Disco (mm)",        type:"semaforo_num", req:true}},
  ]}},
  {{section:"Neumatico Trasero Izquierdo", fields:[
    {{id:"nti_labrado",   label:"Profundidad de labrado (mm)",  type:"semaforo_num", req:true}},
    {{id:"nti_desgaste",  label:"Patron de desgaste / dano",    type:"text", req:true}},
    {{id:"nti_presion",   label:"Presion de inflado (PSI)",     type:"text", req:true}},
    {{id:"nti_pastillas", label:"Espesor de Pastillas (mm)",    type:"semaforo_num", req:true}},
    {{id:"nti_tambor",    label:"Diametro del tambor (mm)",     type:"semaforo_num", req:false}},
  ]}},
  {{section:"Neumatico Trasero Derecho", fields:[
    {{id:"ntd_labrado",   label:"Profundidad de labrado (mm)",  type:"semaforo_num", req:true}},
    {{id:"ntd_desgaste",  label:"Patron de desgaste / dano",    type:"text", req:true}},
    {{id:"ntd_presion",   label:"Presion de inflado (PSI)",     type:"text", req:true}},
    {{id:"ntd_pastillas", label:"Espesor de Pastillas (mm)",    type:"semaforo_num", req:true}},
    {{id:"ntd_tambor",    label:"Diametro del tambor (mm)",     type:"semaforo_num", req:false}},
  ]}},
  {{section:"Neumatico de Repuesto", fields:[
    {{id:"nrep_presion",  label:"Presion de inflado (PSI)",     type:"text", req:true}},
  ]}},
  {{section:"Mantencion / Comentarios", fields:[
    {{id:"reinicio_aceite", label:"Reinicio indicador de cambio de Aceite", type:"check", req:false}},
    {{id:"comentarios",     label:"Comentarios",                 type:"textarea", req:false}},
  ]}},
  {{section:"Diagnostico", fields:[
    {{id:"diag_sintoma",     label:"Sintoma",      type:"textarea", req:false}},
    {{id:"diag_componente",  label:"Componente",   type:"textarea", req:false}},
    {{id:"diag_causa_raiz",  label:"Causa Raiz",   type:"textarea", req:false}},
  ]}},
  {{section:"Firmas", fields:[
    {{id:"nombre_asesor",   label:"Nombre del Asesor",  type:"text", req:true}},
    {{id:"nombre_tecnico",  label:"Nombre del Tecnico", type:"text", req:true}},
  ]}},
];

function esFord(o){{return marcaDeOrden(o).toUpperCase()==='FORD';}}

function _vcuMap(){{
  if(!ctrlData[SUCURSAL])ctrlData[SUCURSAL]={{}};
  if(!ctrlData[SUCURSAL].vcu)ctrlData[SUCURSAL].vcu={{}};
  return ctrlData[SUCURSAL].vcu;
}}
function vcuEstado(o){{return _vcuMap()[o.id]||null;}}
function vcuDatos(o){{const e=vcuEstado(o);return(e&&e.datos)||{{}};}}
function vcuCompleto(o){{const e=vcuEstado(o);return!!(e&&e.completo);}}

function _vcuCamposRequeridos(){{
  const out=[];
  VCU_SCHEMA.forEach(sec=>sec.fields.forEach(f=>{{if(f.req)out.push(f);}}));
  return out;
}}
function vcuFaltantes(datos){{
  return _vcuCamposRequeridos().filter(f=>{{
    const v=datos[f.id];
    if(v===undefined||v===null||String(v).trim()==='')return true;
    if(f.type==='semaforo_num'){{
      const vn=datos[f.id+'_valor'];
      if(vn===undefined||vn===null||String(vn).trim()==='')return true;
    }}
    return false;
  }});
}}

// Antes bloqueaba el avance de etapa de una orden Ford mas alla de "En Proceso" si su
// VCU no estaba completo. A pedido de Cristobal (15/07/2026) se elimino esa restriccion:
// el VCU se sigue pidiendo/mostrando (badge, formulario, PDF), pero ya no impide mover
// la orden de etapa en el JPCB/Control de Taller/modal aunque falte completarlo.
function _avanceBloqueadoPorVCU(o,nuevaEtapaId){{
  return false;
}}

const TIPOS={{
  recall:{{color:"#ffcc99",border:"#d2691e",label:"Recall"}},
  mant:{{color:"#f8c6d6",border:"#e87aa0",label:"Mantencion"}},
  rep: {{color:"#bfe3b0",border:"#5aa84a",label:"Reparacion"}},
  diag:{{color:"#d2b3e8",border:"#9a5fc4",label:"Diagnostico"}},
  ot:  {{color:"#d9e1e7",border:"#7a8ba0",label:"Otro"}},
}};
const DIAS=['Dom','Lun','Mar','Mie','Jue','Vie','Sab'];
const MESES=['ene','feb','mar','abr','may','jun','jul','ago','sep','oct','nov','dic'];

const hhmm=m=>String(Math.floor(m/60)).padStart(2,'0')+':'+String(m%60).padStart(2,'0');
const parseHH=t=>{{const[a,b]=(t||'00:00').split(':').map(Number);return a*60+b;}};
const byId=id=>ordenes.find(o=>o.id===String(id));
const tipoInfo=o=>TIPOS[o.tipo||'ot']||TIPOS.ot;

function formatDate(d){{return String(d.getDate()).padStart(2,'0')+'/'+String(d.getMonth()+1).padStart(2,'0')+'/'+d.getFullYear();}}
function parseDate(s){{const p=s.split('/');return new Date(+p[2],+p[1]-1,+p[0]);}}
function getEnd(dateStr){{const dow=parseDate(dateStr).getDay();return(dow===1||dow===2)?18*60:17*60;}}
function toast(msg,dur=2500){{const t=document.getElementById('toast');t.textContent=msg;t.classList.add('show');setTimeout(()=>t.classList.remove('show'),dur);}}
function setSaveStatus(msg){{const el=document.getElementById('saveStatus');el.textContent=msg;if(msg&&msg.startsWith('✅'))setTimeout(()=>el.textContent='',4000);}}

let agendaData=null,ctrlData=null,ctrlSha=null;
let prodData=null,_prodTecSel=null;
let ppData=null,ppSha=null,ppSelectedDate='',ppOpenKey=null;
// Descuento del 10% aplicado (o no) al presupuesto de cada cita del Pre-picking —
// estado solo de la sesion actual (como ppOpenKey), no se guarda en GitHub: es una
// herramienta rapida para simular el presupuesto con descuento antes de exportarlo.
let ppDescuentos={{}};
let ordenes=[],tecnicos=[],asesoresSucursal=[];
let currentView='jpcb',selectedDate='',modalBloqueId=null;
let currentArea='st';
// Ultima version "sincronizada" de ordenes/bloques (id/fecha -> JSON), usada para
// mezclar los cambios propios con los de otros usuarios en vez de pisar todo el
// registro de la sucursal — ver _refrescarCtrlSha() (22/07/2026).
let _ordenesBaseline=new Map(), _bloquesBaseline=new Map();

/* ─── Area de trabajo: Servicio Tecnico vs Desabolladura y Pintura ───
   Se detecta automaticamente segun palabras clave en el texto que viaja desde
   la Agenda Curifor al campo Comentarios/Servicio de cada orden o cita — no hay
   que elegirlo a mano. Si no calza con ninguna palabra de DyP, se asume Servicio
   Tecnico (comportamiento por defecto, igual que antes de esta funcionalidad). */
function detectArea(text){{
  const s=String(text||'').toUpperCase();
  if(/DESABOLL|PINTURA|\bCHAPA\b|\bDYP\b|D\s*&\s*P|LATONER/.test(s))return'dyp';
  return'st';
}}
function ordenArea(o){{return detectArea(o.comentarios||o.servicio||o.mantencion||'');}}
function citaArea(c){{return detectArea(c.servicio||c.mantencion||'');}}
function setArea(area,btn){{
  currentArea=area;
  document.querySelectorAll('.area-tab').forEach(b=>b.classList.remove('active'));
  if(btn)btn.classList.add('active');
  renderJPCB();
  renderControlTaller();
  renderVehiculosTaller();
  renderHistorialTaller();
  if(currentView==='plan')renderPlanView();
}}
const today_d=new Date();
const planDates=[0,1,2,3,4].map(i=>{{const d=new Date(today_d);d.setDate(d.getDate()+i);return d;}});

function switchView(v,btn){{
  document.querySelectorAll('.sntab').forEach(b=>b.classList.remove('active'));
  if(btn)btn.classList.add('active');
  document.getElementById('v-jpcb').style.display=v==='jpcb'?'block':'none';
  document.getElementById('v-plan').style.display=v==='plan'?'block':'none';
  document.getElementById('v-ct').style.display=v==='ct'?'block':'none';
  document.getElementById('v-vt').style.display=v==='vt'?'block':'none';
  document.getElementById('v-hist').style.display=v==='hist'?'block':'none';
  document.getElementById('v-prod').style.display=v==='prod'?'block':'none';
  const _vpp=document.getElementById('v-pp');
  if(_vpp)_vpp.style.display=v==='pp'?'block':'none';
  currentView=v;
  if(v==='plan')renderPlanView();
  if(v==='ct')renderControlTaller();
  if(v==='vt')renderVehiculosTaller();
  if(v==='hist')renderHistorialTaller();
  if(v==='prod')renderProduccion();
  if(v==='pp')renderPrepickingView();
}}

/* Puntaje de "cuanto avance/datos tiene" una orden — usado por
   _dedupOrdenesPorPatenteOT para decidir cual de 2 duplicados exactos (misma
   patente+OT) conservar cuando aparecen ambos. Prioriza la que ya tiene trabajo
   real encima (tecnico asignado, etapa avanzada, cerrada, confirmada, comentarios,
   etc.) para no perder nada si alguien ya empezo a operar sobre una de las 2. */
function _ordenScore(o){{
  let s=0;
  if(o.tecnico!==null&&o.tecnico!==undefined&&o.tecnico!=='')s+=2;
  if(o.etapa&&ETAPAS.length&&o.etapa!==ETAPAS[0].id)s+=3;
  if(o.cerrada)s+=3;
  if(o.estadoCita==='asiste')s+=2;
  if(o.estadoCita==='no_asiste'||o.estadoCita==='reagenda')s+=1;
  if(o.stop)s+=1;
  ['comentario2','numero_caso','n_pedido','eta','auto_reemplazo',
   'ingreso_taller','salida_taller','tecnico_x_hora'].forEach(f=>{{
    if(String(o[f]||'').trim())s+=1;
  }});
  return s;
}}
/* Deduplica ordenes que comparten la MISMA patente y el MISMO Folio OT — nunca toca
   patentes con OTs distintas (una patente puede tener varias citas activas a la vez
   por diseno, ver autoImportarCitas). Un duplicado real de patente+OT solo puede pasar
   por una carrera: 2 sesiones/pestañas abriendo el Planificador casi al mismo tiempo,
   cada una creando su propia orden nueva para la misma cita porque todavia no veian la
   que la otra acababa de crear (autoImportarCitas compara solo contra lo que esa
   sesion tiene cargado localmente) — el merge por id de mas abajo no lo detectaba,
   asi que ambas quedaban guardadas para siempre. 29/07/2026, a pedido de Cristobal
   ("estaria bueno... siempre y cuando el numero de OT este duplicado"). Ordenes sin
   OT (altas manuales sin folio) nunca se tocan aqui — no hay forma confiable de saber
   si son la misma cita o 2 vehiculos distintos agregados a mano. */
function _dedupOrdenesPorPatenteOT(lista){{
  const elegidas=new Map(); // patente|OT -> orden elegida
  const resultado=[];
  lista.forEach(o=>{{
    const pat=normPat(o.patente);
    const ot=String(o.ot||'').trim();
    if(!pat||!ot){{resultado.push(o);return;}}
    const key=pat+'|'+ot;
    if(!elegidas.has(key)){{
      elegidas.set(key,o);
      resultado.push(o);
    }} else {{
      const previa=elegidas.get(key);
      if(_ordenScore(o)>_ordenScore(previa)){{
        const idx=resultado.indexOf(previa);
        if(idx>=0)resultado[idx]=o;
        elegidas.set(key,o);
      }}
      // la orden descartada (duplicado real de patente+OT) se pierde a proposito
    }}
  }});
  return resultado;
}}

/* Mezcla ordenes/bloques/no_show/vcu/asesores_extra de la sucursal actual entre lo que
   trae GitHub (fresco — puede incluir cambios de OTROS usuarios, ej. Torre Control
   asignando tecnico/horario a una orden) y lo que hay en memoria local (mis propios
   cambios sin guardar todavia). Antes de esto, saveCtrl() simplemente pisaba TODA la
   sucursal con la copia local — si dos personas editaban la misma sucursal a la vez,
   el ultimo en guardar borraba silenciosamente los cambios del otro sin ningun aviso.
   Ahora se compara cada registro contra su "ultima version sincronizada"
   (_ordenesBaseline/_bloquesBaseline, tomada la ultima vez que se cargo o se guardo):
   solo lo que YO modifique desde entonces se superpone sobre lo fresco — todo lo demas
   (incluyendo ordenes nuevas o editadas por otro usuario) se toma tal cual viene de
   GitHub. 22/07/2026, a pedido de Cristobal ("un torre control asigno un trabajo... yo
   no lo veo").

   04/08/2026 — FIX REAL del bug "cambios que se revierten solos" (reportado con video:
   un usuario marca 'Asiste'/mueve una tarjeta de etapa y el cambio desaparece poco
   despues, y "sigue ocurriendo" pese a 5 rondas previas de fixes de concurrencia).
   Las rondas anteriores (22/07 a 03/08) resolvian las carreras DENTRO de una misma
   sesion (encolar saveCtrl, refrescar el SHA antes de guardar) pero el merge seguia
   comparando el OBJETO COMPLETO de cada orden/bloque contra el baseline: si YO cambiaba
   CUALQUIER campo, mi copia local ENTERA (incluyendo campos que NO toque) reemplazaba
   lo que hubiera en el servidor — borrando en silencio cualquier campo que OTRO usuario
   hubiera cambiado en esa MISMA orden mientras mi pestaña seguia abierta. Ejemplo real:
   Torre Control asigna tecnico a las 09:00; un asesor tiene el JPCB abierto desde las
   08:00 y a las 09:05 marca "Asiste" en esa misma orden — su guardado trae su copia
   vieja de la orden (sin el tecnico que Torre Control acaba de asignar) y la pisa
   ENTERA, o viceversa (el guardado de Torre Control pisa el "Asiste" del asesor si sale
   despues). Con varias pestañas abiertas todo el dia (lo normal en un taller), esta
   colision es cuestion de tiempo — de ahi que "siguiera ocurriendo" pese a los fixes
   anteriores, que nunca atacaron esto. Ahora la mezcla es CAMPO A CAMPO (ver
   _mergeCampoACampo) para ordenes, y BLOQUE A BLOQUE por id (ver _mergeArrayPorId) para
   el grid Tecnico x Hora: un campo que YO no toque desde mi ultima sincronizacion
   siempre respeta lo que traiga el servidor (incluyendo cambios de otros), y solo los
   campos que SI cambie se superponen — sin arrastrar de paso el resto de la orden. */
function _mergeCampoACampo(fresco,local,baseline){{
  const result={{...(fresco||{{}})}};
  const claves=new Set([...Object.keys(local||{{}}),...Object.keys(baseline||{{}})]);
  claves.forEach(k=>{{
    const lv=local?local[k]:undefined, bv=baseline?baseline[k]:undefined;
    if(JSON.stringify(lv)!==JSON.stringify(bv))result[k]=lv; // yo cambie este campo -> mi valor gana
  }});
  return result;
}}
/* Mismo criterio que _mergeCampoACampo pero para un ARRAY de objetos con "id" (los
   bloques del grid Tecnico x Hora — varios por fecha). Antes se decidia por FECHA
   completa (si yo tocaba cualquier bloque de un dia, mi lista entera de ese dia pisaba
   la del servidor, perdiendo un bloque que otro usuario hubiera agregado/movido ese
   mismo dia) — ahora se decide bloque por bloque, por su id. */
function _mergeArrayPorId(freshArr,localArr,baselineArr){{
  freshArr=Array.isArray(freshArr)?freshArr:[];
  localArr=Array.isArray(localArr)?localArr:[];
  baselineArr=Array.isArray(baselineArr)?baselineArr:[];
  const baseMap=new Map(baselineArr.map(x=>[String(x.id),JSON.stringify(x)]));
  const freshMap=new Map(freshArr.map(x=>[String(x.id),x]));
  const localMap=new Map(localArr.map(x=>[String(x.id),x]));
  const result=new Map(freshMap);
  localMap.forEach((val,id)=>{{
    const baseStr=baseMap.get(id);
    if(baseStr===undefined||JSON.stringify(val)!==baseStr)result.set(id,val); // lo agregue/edite yo
  }});
  baseMap.forEach((baseStr,id)=>{{
    if(!localMap.has(id)){{ // ya no esta en mi copia local -> lo elimine yo
      const freshVal=freshMap.get(id);
      // solo se respeta mi eliminacion si nadie mas lo cambio desde entonces; si el
      // servidor SI lo cambio, se respeta ese cambio ajeno (ya quedo en result via freshMap)
      if(freshVal!==undefined&&JSON.stringify(freshVal)===baseStr)result.delete(id);
    }}
  }});
  return [...result.values()];
}}
function _mergeOrdenesYBloques(fresco){{
  const frescoSuc=(fresco&&fresco[SUCURSAL])||{{}};
  if(!ctrlData[SUCURSAL])ctrlData[SUCURSAL]={{}};

  // --- ordenes: mezcla CAMPO A CAMPO por id (ver _mergeCampoACampo) ---
  const freshOrdenes=Array.isArray(frescoSuc.ordenes)?frescoSuc.ordenes:[];
  const mergedOrd=new Map(freshOrdenes.map(o=>[String(o.id),o]));
  ordenes.forEach(o=>{{
    const id=String(o.id);
    const baseStr=_ordenesBaseline.get(id);
    if(baseStr===undefined){{mergedOrd.set(id,o);return;}} // orden creada localmente, aun no sincronizada
    let baseObj={{}};
    try{{baseObj=JSON.parse(baseStr);}}catch(e){{}}
    const freshObj=mergedOrd.get(id);
    if(freshObj===undefined){{
      // el servidor ya no tiene esta orden (otro usuario la elimino). Si yo tampoco
      // cambie nada respecto al baseline, se respeta el borrado; si SI cambie algo, se
      // recrea con mi version para no perder mi trabajo.
      if(JSON.stringify(o)!==baseStr)mergedOrd.set(id,o);
      return;
    }}
    mergedOrd.set(id,_mergeCampoACampo(freshObj,o,baseObj));
  }});
  const idsActuales=new Set(ordenes.map(o=>String(o.id)));
  for(const id of _ordenesBaseline.keys()){{
    if(!idsActuales.has(id))mergedOrd.delete(id); // eliminada localmente (eliminarOrdenCT)
  }}
  // Limpia duplicados reales de patente+OT (misma cita creada 2 veces por una carrera
  // de autoImportarCitas entre sesiones concurrentes — ver _dedupOrdenesPorPatenteOT).
  // Corre en CADA merge, o sea en cada guardado (_saveCtrlInterno siempre arranca con
  // _refrescarCtrlSha) — asi cualquier duplicado que ya haya quedado guardado en
  // GitHub tambien se limpia solo apenas alguien vuelva a guardar algo.
  ordenes=_dedupOrdenesPorPatenteOT([...mergedOrd.values()]);

  // --- bloques (grid Tecnico x Hora): mezcla bloque a bloque, por id (ver _mergeArrayPorId) ---
  const freshBloques=frescoSuc.bloques||{{}};
  const localBloques=ctrlData[SUCURSAL].bloques||{{}};
  const fechasBloques=new Set([...Object.keys(freshBloques),...Object.keys(localBloques),..._bloquesBaseline.keys()]);
  const mergedBloq={{}};
  fechasBloques.forEach(fecha=>{{
    const baseStr=_bloquesBaseline.get(fecha);
    let baseArr=[];
    if(baseStr){{try{{baseArr=JSON.parse(baseStr);}}catch(e){{}}}}
    mergedBloq[fecha]=_mergeArrayPorId(freshBloques[fecha],localBloques[fecha],baseArr);
  }});
  ctrlData[SUCURSAL].bloques=mergedBloq;

  // --- no_show / vcu / eliminadas (mapas por clave) y asesores_extra (lista):
  // union simple, los cambios locales tienen prioridad sobre la MISMA clave, pero
  // no se pierden las claves que solo existan del lado fresco. ---
  ['no_show','vcu','eliminadas','no_disponible'].forEach(key=>{{
    const freshMap=frescoSuc[key]||{{}};
    const localMap=ctrlData[SUCURSAL][key]||{{}};
    ctrlData[SUCURSAL][key]={{...freshMap,...localMap}};
  }});
  if(frescoSuc.asesores_extra||ctrlData[SUCURSAL].asesores_extra){{
    ctrlData[SUCURSAL].asesores_extra=[...new Set([...(frescoSuc.asesores_extra||[]),...(ctrlData[SUCURSAL].asesores_extra||[])])];
  }}

  _snapshotOrdenes();_snapshotBloques();
}}

async function _refrescarCtrlSha(){{
  // Relee control_taller.json fresco desde GitHub (SHA + datos actuales) y mezcla los
  // cambios (ver _mergeOrdenesYBloques) — asi no se pisan cambios guardados por otra
  // sesion (ej. Torre Control) mientras esta pestaña estaba abierta.
  // Camino Supabase: una sola llamada, sin limite de tamano y sin CDN de por
  // medio, asi que no hace falta el respaldo por Git Data API de mas abajo.
  if(USA_SUPABASE){{
    try{{
      const j = await sbTablero('tablero_leer', {{p_vale: SB_VALE, p_nombre: 'control_taller.json'}});
      if(!j || j.ok !== true){{ _ctrlLecturaOk=false; return false; }}
      if(j.existe === false){{ _ctrlLecturaOk=true; return true; }} // aun no existe: es valido crearlo
      const fresco = j.data;
      if(!fresco || typeof fresco !== 'object'){{ _ctrlLecturaOk=false; return false; }}
      ctrlSha = j.sello || ctrlSha;
      _ctrlLecturaOk=true;
      _ctrlSucursalesServidor=Object.keys(fresco).filter(k=>fresco[k]&&typeof fresco[k]==='object'&&!Array.isArray(fresco[k]));
      _ctrlOrdenesServidor=(((fresco[SUCURSAL]||{{}}).ordenes)||[]).length;
      _ctrlTecnicosServidor=((fresco[SUCURSAL]||{{}}).tecnicos)||[];
      if(!ctrlData)ctrlData={{}};
      for(const k of Object.keys(fresco)){{
        if(k!==SUCURSAL)ctrlData[k]=fresco[k];
      }}
      _mergeOrdenesYBloques(fresco);
      try{{ if(typeof render==='function')render(); }}catch(e){{}}
      return true;
    }}catch(e){{ _ctrlLecturaOk=false; return false; }}
  }}

  try{{
    const r=await fetch(API_BASE+'control_taller.json',{{
      headers:{{'Authorization':`token ${{GITHUB_TOKEN}}`,'Accept':'application/vnd.github.v3+json'}}}});
    if(r.status===404){{_ctrlLecturaOk=true;return true;}} // el archivo aun no existe: es valido crearlo
    if(!r.ok)return false;
    const j=await r.json();
    ctrlSha=j.sha||ctrlSha;
    let fresco=null;
    if(j.content){{
      try{{fresco=JSON.parse(decodeURIComponent(escape(atob(j.content.replace(/\\n/g,'')))));}}catch(e){{fresco=null;}}
    }}else{{
      // INCIDENTE 05/08/2026: la Contents API devuelve content vacio si el archivo
      // pesa mas de 1 MB. Antes esto dejaba fresco={{}} y el guardado subia SOLO la
      // sucursal actual, borrando las otras 10 y los 46 tecnicos. Ahora se lee el
      // blob por Git Data API, que no tiene limite de tamano.
      try{{
        const rr=await fetch(API_BASE.replace('/contents/','/git/')+'ref/heads/main',{{
          headers:{{'Authorization':`token ${{GITHUB_TOKEN}}`,'Accept':'application/vnd.github.v3+json'}}}});
        const commitSha=(await rr.json()).object.sha;
        const rt=await fetch(API_BASE.replace('/contents/','/git/')+'trees/'+commitSha+'?recursive=1',{{
          headers:{{'Authorization':`token ${{GITHUB_TOKEN}}`,'Accept':'application/vnd.github.v3+json'}}}});
        const item=((await rt.json()).tree||[]).find(x=>x.path==='control_taller.json');
        if(item){{
          const rb=await fetch(API_BASE.replace('/contents/','/git/')+'blobs/'+item.sha,{{
            headers:{{'Authorization':`token ${{GITHUB_TOKEN}}`,'Accept':'application/vnd.github.v3+json'}}}});
          const jb=await rb.json();
          fresco=JSON.parse(decodeURIComponent(escape(atob((jb.content||'').replace(/\\n/g,'')))));
          ctrlSha=item.sha||ctrlSha;
        }}
      }}catch(e){{fresco=null;}}
    }}
    // Si el archivo existe pero NO se pudo leer, se aborta: guardar a ciegas es
    // exactamente lo que borro los datos el 05/08/2026.
    if(fresco===null){{_ctrlLecturaOk=false;return false;}}
    _ctrlLecturaOk=true;
    _ctrlSucursalesServidor=Object.keys(fresco).filter(k=>fresco[k]&&typeof fresco[k]==='object'&&!Array.isArray(fresco[k]));
    _ctrlOrdenesServidor=(((fresco[SUCURSAL]||{{}}).ordenes)||[]).length;
    _ctrlTecnicosServidor=((fresco[SUCURSAL]||{{}}).tecnicos)||[];
    if(!ctrlData)ctrlData={{}};
    for(const k of Object.keys(fresco)){{
      if(k!==SUCURSAL)ctrlData[k]=fresco[k]; // otras sucursales: se toman tal cual (no las edita esta sesion)
    }}
    _mergeOrdenesYBloques(fresco);
    // El merge puede traer cambios de otro usuario (ej. una orden nueva/editada por
    // Torre Control) — se refleja de inmediato en la pantalla, no solo cuando se
    // recarga la pagina.
    renderJPCB();renderControlTaller();renderVehiculosTaller();
    if(typeof renderHistorialTaller==='function')renderHistorialTaller();
    if(currentView==='plan')renderPlanView();
    return true;
  }}catch(e){{return false;}}
}}

/* FIX CRITICO DE CONCURRENCIA (22/07/2026, ronda 2) — un cambio (mover etapa en el
   JPCB, eliminar una cita, marcar "no asiste", etc.) podia revertirse solo a los
   pocos segundos, y en cambios simultaneos "solo se aplicaba el de uno". Causa
   real: saveCtrl() se llamaba SIN esperar (fire-and-forget) desde el drag&drop y
   otros botones — si el usuario disparaba una segunda accion (otro drag, otro
   click) ANTES de que el primer guardado terminara su ciclo completo (fetch fresco
   -> merge -> re-snapshot del baseline -> PUT), el SEGUNDO guardado arrancaba su
   propio fetch mientras el PRIMERO ya habia avanzado el baseline (_ordenesBaseline)
   al mezclar. Como el baseline ya reflejaba el cambio local (puesto por el primer
   guardado), el merge del SEGUNDO guardado comparaba contra ESE baseline ya
   actualizado -> veia "sin diferencia respecto al baseline" -> tomaba tal cual la
   copia del servidor que su propio fetch (posiblemente aun sin la primera edicion
   ya subida) habia traido, revirtiendo el cambio en memoria Y subiendolo de vuelta
   a GitHub. Esto podia pasar con dos acciones seguidas de UNA misma persona, o con
   dos personas editando casi al mismo tiempo.
   Fix: encolar todas las llamadas a saveCtrl() en una cadena de promesas
   (_ctrlSaveChain) para que nunca haya dos ciclos fetch->merge->PUT corriendo en
   paralelo dentro de la misma pestaña — cada guardado espera a que el anterior
   termine por completo (incluyendo su propio reintento por SHA desactualizada)
   antes de empezar el suyo, asi el baseline siempre esta al dia cuando se compara. */
let _ctrlSaveChain=Promise.resolve();
/* Marca si la ultima relectura de control_taller.json fue exitosa. Si NO se pudo
   leer el archivo, no se guarda nada (ver el incidente del 05/08/2026 en
   _refrescarCtrlSha y en _cargar_ctrl_taller). */
let _ctrlLecturaOk=false;
/* Sucursales que el servidor tenia en la ultima lectura exitosa. Se usa como
   guardia antes del PUT: si a la copia local le falta alguna, no se sube. */
let _ctrlSucursalesServidor=[];
/* Cuantas ordenes tenia el servidor para ESTA sucursal en la ultima lectura.
   Guardia contra el caso del 05/08/2026 14:13: un guardado dejo LINDEROS con 1
   orden de 248, borrando 247 con sus tecnicos, etapas y bloques. */
let _ctrlOrdenesServidor=0;
/* Tecnicos que el servidor tenia para ESTA sucursal en la ultima lectura. Se usa
   como respaldo: los tecnicos se configuran en Admin, no en el Planificador, asi
   que una lista local vacia significa "no los cargue", nunca "borralos". */
let _ctrlTecnicosServidor=[];
function saveCtrl(){{
  _ctrlSaveChain=_ctrlSaveChain.then(()=>_saveCtrlInterno(false)).catch(e=>{{console.warn('saveCtrl encolado fallo',e);}});
  return _ctrlSaveChain;
}}
async function _saveCtrlInterno(_reintento){{
  if(!USA_SUPABASE && !GITHUB_TOKEN){{setSaveStatus('Sin token');return;}}
  setSaveStatus('💾 Guardando...');
  const _okLectura=await _refrescarCtrlSha();
  if(!_okLectura){{
    // No se pudo leer la version actual del servidor. Guardar igual significaria
    // subir SOLO lo que tiene esta pestaña en memoria y borrar el resto de las
    // sucursales/tecnicos — exactamente el incidente del 05/08/2026.
    setSaveStatus('⛔ No se guardo: no se pudo leer el archivo del servidor');
    alert('No se pudo leer la version actual del Planificador desde el servidor, '
          +'asi que NO se guardo el cambio (para no borrar los datos de las otras '
          +'sucursales).\\n\\nRevisa tu conexion y presiona "Actualizar datos". '
          +'Si el problema sigue, avisa al administrador antes de seguir editando.');
    return;
  }}
  const now=new Date();
  const nowStr=now.toLocaleDateString('es-CL')+' '+now.toLocaleTimeString('es-CL',{{hour:'2-digit',minute:'2-digit'}});
  if(!ctrlData)ctrlData={{}};
  ctrlData.fecha_actualizacion=nowStr;
  if(!ctrlData[SUCURSAL])ctrlData[SUCURSAL]={{}};
  // Los tecnicos NO se editan desde el Planificador (se configuran en Admin), asi que
  // una lista local vacia casi siempre significa "esta pestaña cargo sin datos", no
  // "borren los tecnicos". Subirla igual fue lo que dejo las sucursales sin tecnicos
  // para asignar el 05/08/2026. Solo se pisa la del servidor si la local tiene algo.
  const _tecServidor=(_ctrlTecnicosServidor&&_ctrlTecnicosServidor.length)
        ?_ctrlTecnicosServidor:((ctrlData[SUCURSAL].tecnicos)||[]);
  ctrlData[SUCURSAL].tecnicos=(Array.isArray(tecnicos)&&tecnicos.length)?tecnicos:_tecServidor;
  if(!Array.isArray(tecnicos)||!tecnicos.length)tecnicos=ctrlData[SUCURSAL].tecnicos;
  ctrlData[SUCURSAL].ordenes=ordenes;
  // Guardia final: nunca subir un archivo al que le falten sucursales que el
  // servidor SI tenia hace un instante (defensa en profundidad del 05/08/2026).
  const _faltan=(_ctrlSucursalesServidor||[]).filter(k=>!(k in ctrlData));
  if(_faltan.length){{
    setSaveStatus('⛔ No se guardo: faltaban sucursales');
    alert('No se guardo el cambio porque la copia local perdio estas sucursales: '
          +_faltan.join(', ')+'.\\n\\nRecarga la pagina y vuelve a intentarlo.');
    return;
  }}
  // Guardia sobre la PROPIA sucursal: si esta pestaña esta a punto de subir muchas
  // menos ordenes de las que el servidor acaba de mostrar, algo salio mal en la
  // mezcla (05/08/2026 14:13: un guardado dejo LINDEROS con 1 orden de 248). Un
  // borrado normal es de a una o dos ordenes, nunca decenas de golpe.
  const _perdidas=(_ctrlOrdenesServidor||0)-((ordenes||[]).length);
  if(_perdidas>5){{
    setSaveStatus('⛔ No se guardo: se perderian '+_perdidas+' ordenes');
    alert('NO se guardo el cambio.\\n\\nEsta pestaña tiene '+((ordenes||[]).length)
          +' ordenes pero el servidor tiene '+_ctrlOrdenesServidor+' en '+SUCURSAL
          +'. Guardar habría borrado '+_perdidas+' ordenes con sus tecnicos y etapas.'
          +'\\n\\nRecarga la pagina (Ctrl+F5) y vuelve a hacer el cambio.');
    return;
  }}
  // JSON compacto (sin indentacion): la version indentada pesaba 1.058.920 bytes y
  // cruzo el limite de 1 MB de la Contents API, que es lo que gatillo la perdida de
  // datos del 05/08/2026. Compacto, el mismo contenido pesa ~750 KB.
  // Camino Supabase. Todas las guardias de arriba siguen valiendo: esto solo
  // cambia por donde viaja el guardado. El `sello` hace lo que hacia el `sha`
  // de GitHub, incluido el reintento cuando otro guardo primero.
  if(USA_SUPABASE){{
    try{{
      const j = await sbTablero('tablero_guardar', {{
        p_vale: SB_VALE, p_nombre: 'control_taller.json',
        p_data: ctrlData, p_sello: ctrlSha || null}});
      if(j && j.ok === true){{
        ctrlSha=j.sello||ctrlSha;_snapshotOrdenes();_snapshotBloques();setSaveStatus('✅ Guardado');
      }}else if(j && j.motivo==='conflicto' && !_reintento){{
        ctrlSha=j.sello||ctrlSha;          // alguien guardo primero: releer y reintentar
        await _saveCtrlInterno(true);
      }}else{{
        setSaveStatus('Error: '+((j&&j.motivo)||'desconocido'));
      }}
    }}catch(e){{setSaveStatus('Error de red');}}
    return;
  }}

  const content=btoa(unescape(encodeURIComponent(JSON.stringify(ctrlData))));
  const payload={{message:`Taller ${{SUCURSAL}} - ${{USUARIO}} ${{nowStr}}`,content,...(ctrlSha?{{sha:ctrlSha}}:{{}})}};
  try{{
    const r=await fetch(API_BASE+'control_taller.json',{{method:'PUT',
      headers:{{'Authorization':`token ${{GITHUB_TOKEN}}`,'Accept':'application/vnd.github.v3+json','Content-Type':'application/json'}},
      body:JSON.stringify(payload)}});
    const j=await r.json();
    if(r.ok){{ctrlSha=j.content&&j.content.sha||ctrlSha;_snapshotOrdenes();_snapshotBloques();setSaveStatus('✅ Guardado');}}
    else if(!_reintento){{
      // SHA desactualizada (chocó con otro guardado simultáneo de otra sesión/usuario): reintenta una vez más.
      await _saveCtrlInterno(true);
    }}
    else setSaveStatus('Error: '+(j.message||r.status));
  }}catch(e){{setSaveStatus('Error de red');}}
}}

function mostrarDetalleCita(el){{
  // No mostrar si estaba arrastrando
  if(el.classList.contains('dragging'))return;
  let c;
  try{{c=JSON.parse(el.dataset.cita||'{{}}');}}catch(e){{return;}}
  const pat=(c.patente||'').replace(/\?/g,'').trim()||'—';
  const ingr=c.estado==='finalizado'?'🧍 Servicio finalizado (retirado)':(c.ingresado?'✅ Ingresado al taller':'🚗 Pendiente de ingreso');
  const filas=[
    ['OC / Folio', c.oc||'—'],
    ['Patente',    pat],
    ['Cliente',    c.nombre||c.cliente||'—'],
    ['Modelo',     c.modelo||'—'],
    ['Año',        c.anio||'—'],
    ['Kilómetros', c.km?Number(c.km).toLocaleString('es-CL')+' km':'—'],
    ['Horario',    c.horario||'—'],
    ['Fecha',      c.fecha||'—'],
    ['Servicio',   c.servicio||'—'],
    ['Mantención', c.mantencion||'—'],
    ['Asesor',     c.asesor||'—'],
    ['Sucursal',   c.sucursal||SUCURSAL],
    ['Estado',     ingr],
  ].filter(([,v])=>v&&v!=='—');
  const _noShow=esNoAsiste(c.oc,c.patente);
  document.getElementById('cm-title').textContent=`🚘 ${{pat}} — ${{c.modelo||''}}`;
  let _cmHtml=filas.map(([l,v])=>
    `<div class="cm-row"><span class="cm-lbl">${{l}}</span><span class="cm-val">${{v}}</span></div>`
  ).join('');
  if(_noShow)_cmHtml=`<div class="cita-noasiste" style="margin-bottom:8px">🚫 Cliente no asiste</div>`+_cmHtml;
  if(PUEDE_EDITAR){{
    const _ocEsc=String(c.oc||'').replace(/'/g,"\\'");
    const _patEsc=pat.replace(/'/g,"\\'");
    _cmHtml+=`<div style="margin-top:10px"><button class="cita-noasiste-btn" onclick="toggleNoAsiste('${{_ocEsc}}','${{_patEsc}}');cerrarDetalleCita();">${{_noShow?'↩️ Reactivar':'🚫 Marcar cliente no asiste'}}</button></div>`;
  }}
  document.getElementById('cm-body').innerHTML=_cmHtml;
  document.getElementById('cita-modal-overlay').classList.add('open');
}}
function cerrarDetalleCita(){{
  document.getElementById('cita-modal-overlay').classList.remove('open');
}}

/* ══════════════════════════════════════════════════════════════
   PRE-PICKING — tarjeta por cita, con detalle desplegable, tabla
   de repuestos sugeridos (pauta de mantencion x marca/modelo/km,
   ya calculada en el consolidador) y estado Realizado/Pendiente
   persistente (prepicking_estados.json en GitHub, por sucursal +
   fecha + OC — independiente de agenda_hoy.json). 13/07/2026.
   ══════════════════════════════════════════════════════════════ */
const fmtCLP=n=>'$'+Math.round(Number(n||0)).toLocaleString('es-CL');
function ppKey(fecha,oc){{return fecha+'__'+oc;}}
function getPpEstado(fecha,oc){{
  const suc=ppData&&ppData[SUCURSAL];
  const dia=suc&&suc[fecha];
  return(dia&&dia[oc])||'pendiente';
}}
async function setPpEstado(fecha,oc,estado){{
  if(!ppData)ppData={{}};
  if(!ppData[SUCURSAL])ppData[SUCURSAL]={{}};
  if(!ppData[SUCURSAL][fecha])ppData[SUCURSAL][fecha]={{}};
  ppData[SUCURSAL][fecha][oc]=estado;
  renderPrepickingView();
  await savePrepicking();
}}
// Override manual de Marca/Modelo/Version por cita (para cuando el texto de la
// Agenda no matchea bien, o el usuario quiere ver otra motorizacion/ano) — se
// guarda junto al estado Realizado/Pendiente en el mismo prepicking_estados.json,
// bajo una clave separada "__overrides" para no chocar con el formato existente
// (ppData[SUC][fecha][oc] sigue siendo el string de estado). 22/07/2026.
function getPpOverride(fecha,oc){{
  const suc=ppData&&ppData[SUCURSAL];
  const ov=suc&&suc.__overrides;
  return (ov&&ov[ppKey(fecha,oc)])||null;
}}
async function setPpOverride(fecha,oc,override){{
  if(!ppData)ppData={{}};
  if(!ppData[SUCURSAL])ppData[SUCURSAL]={{}};
  if(!ppData[SUCURSAL].__overrides)ppData[SUCURSAL].__overrides={{}};
  const k=ppKey(fecha,oc);
  if(override) ppData[SUCURSAL].__overrides[k]=override;
  else delete ppData[SUCURSAL].__overrides[k];
  renderPrepickingView();
  await savePrepicking();
}}
function _cotizModeloObjPorNombre(marcaNombre,modeloNombre){{
  if(!COTIZ_PP||!COTIZ_PP.indice) return null;
  const m=(COTIZ_PP.indice.marcas||[]).find(x=>x.nombre===marcaNombre);
  if(!m) return null;
  return (m.modelos||[]).find(md=>md.nombre===modeloNombre)||null;
}}
// Maneja el cambio de cualquiera de los 3 selects (Marca/Modelo/Version) del
// Pre-picking: parte del override actual (o de lo auto-detectado si aun no hay
// override) y solo pisa el campo que cambio, reseteando los campos "hijos"
// (cambiar Marca limpia Modelo+Version; cambiar Modelo limpia Version).
async function ppCambiarModeloSel(fecha,oc,campo,valor){{
  if(campo==='reset'){{ await setPpOverride(fecha,oc,null); return; }}
  const cita=(getCitas(fecha)||[]).find(c=>String(c.oc||c.patente)===String(oc));
  let marca='',modelo='',anio='',versionId='';
  const _ovPrev=getPpOverride(fecha,oc);
  if(_ovPrev){{ marca=_ovPrev.marca||'';modelo=_ovPrev.modelo||'';anio=_ovPrev.anio||'';versionId=_ovPrev.versionId||''; }}
  else if(cita){{
    const _autoModelo=_cotizBuscarModelo(String(cita.modelo||''));
    if(_autoModelo){{
      for(const m of (COTIZ_PP.indice.marcas||[])){{ if((m.modelos||[]).includes(_autoModelo)){{marca=m.nombre;break;}} }}
      modelo=_autoModelo.nombre;
      versionId=(_autoModelo.versiones&&_autoModelo.versiones[0])?_autoModelo.versiones[0].id:'';
    }}
  }}
  // El selector de Ano (agregado 23/07/2026, a pedido de Cristobal) es solo un
  // filtro visual para acortar la lista de Version cuando un modelo tiene muchas
  // motorizaciones/anos — cambiar Marca o Modelo limpia Ano+Version (empiezan de
  // nuevo); cambiar Ano limpia solo Version (para forzar a elegir de nuevo dentro
  // de las versiones que cubren ese ano).
  if(campo==='marca'){{ marca=valor;modelo='';anio='';versionId=''; }}
  else if(campo==='modelo'){{ modelo=valor;anio='';versionId=''; }}
  else if(campo==='anio'){{ anio=valor;versionId=''; }}
  else if(campo==='version'){{ versionId=valor; }}
  if(!marca&&!modelo&&!anio&&!versionId){{ await setPpOverride(fecha,oc,null); return; }}
  await setPpOverride(fecha,oc,{{marca,modelo,anio,versionId}});
}}
async function savePrepicking(_reintento){{
  if(!USA_SUPABASE && !GITHUB_TOKEN){{setSaveStatus('Sin token');return;}}
  setSaveStatus('💾 Guardando...');

  // Camino Supabase: se relee para no pisar las otras sucursales y se guarda
  // con el sello, igual que hacia con el sha.
  if(USA_SUPABASE){{
    try{{
      const j0 = await sbTablero('tablero_leer', {{p_vale: SB_VALE, p_nombre: 'prepicking_estados.json'}});
      if(j0 && j0.ok === true && j0.existe !== false && j0.data && typeof j0.data === 'object'){{
        ppSha = j0.sello || ppSha;
        for(const k of Object.keys(j0.data)){{ if(k!==SUCURSAL) ppData[k]=j0.data[k]; }}
      }}
    }}catch(e){{}}
    try{{
      const j = await sbTablero('tablero_guardar', {{
        p_vale: SB_VALE, p_nombre: 'prepicking_estados.json',
        p_data: ppData, p_sello: ppSha || null}});
      if(j && j.ok === true){{ ppSha=j.sello||ppSha; setSaveStatus('✅ Guardado'); }}
      else if(j && j.motivo==='conflicto' && !_reintento){{ ppSha=j.sello||ppSha; await savePrepicking(true); }}
      else setSaveStatus('Error: '+((j&&j.motivo)||'desconocido'));
    }}catch(e){{setSaveStatus('Error de red');}}
    return;
  }}

  try{{
    const r0=await fetch(API_BASE+'prepicking_estados.json',{{
      headers:{{'Authorization':`token ${{GITHUB_TOKEN}}`,'Accept':'application/vnd.github.v3+json'}}}});
    if(r0.ok){{
      const j0=await r0.json();
      ppSha=j0.sha||ppSha;
      let fresco={{}};
      if(j0.content){{try{{fresco=JSON.parse(decodeURIComponent(escape(atob(j0.content.replace(/\\n/g,'')))));}}catch(e){{fresco={{}};}}}}
      for(const k of Object.keys(fresco)){{if(k!==SUCURSAL)ppData[k]=fresco[k];}}
    }}
  }}catch(e){{}}
  const content=btoa(unescape(encodeURIComponent(JSON.stringify(ppData,null,2))));
  const now=new Date();
  const nowStr=now.toLocaleDateString('es-CL')+' '+now.toLocaleTimeString('es-CL',{{hour:'2-digit',minute:'2-digit'}});
  const payload={{message:`Pre-picking ${{SUCURSAL}} - ${{USUARIO}} ${{nowStr}}`,content,...(ppSha?{{sha:ppSha}}:{{}})}};
  try{{
    const r=await fetch(API_BASE+'prepicking_estados.json',{{method:'PUT',
      headers:{{'Authorization':`token ${{GITHUB_TOKEN}}`,'Accept':'application/vnd.github.v3+json','Content-Type':'application/json'}},
      body:JSON.stringify(payload)}});
    const j=await r.json();
    if(r.ok){{ppSha=j.content&&j.content.sha||ppSha;setSaveStatus('✅ Guardado');}}
    else if(!_reintento){{await savePrepicking(true);}}
    else setSaveStatus('Error: '+(j.message||r.status));
  }}catch(e){{setSaveStatus('Error de red');}}
}}

function ppSelectDate(btn){{
  document.querySelectorAll('#ppDateTabs .dtab').forEach(b=>b.classList.remove('active'));
  btn.classList.add('active');ppSelectedDate=btn.dataset.date;renderPrepickingView();
}}

function renderPrepickingView(){{
  document.getElementById('ppDateTabs').innerHTML=planDates.map((d,i)=>{{
    const dow=d.getDay(),fecha=formatDate(d);
    const lbl=i===0?'📅 Hoy':i===1?'📅 Manana':i===2?'📅 Pasado manana':'📅 +'+i+' dias';
    const dayStr=DIAS[dow]+' '+d.getDate()+' '+MESES[d.getMonth()];
    const cls='dtab'+(fecha===ppSelectedDate?' active':'');
    return`<button class="${{cls}}" data-date="${{fecha}}" onclick="ppSelectDate(this)">${{lbl}} — ${{dayStr}}</button>`;
  }}).join('');

  const citas=getCitas(ppSelectedDate).filter(c=>citaArea(c)===currentArea)
    .slice().sort((a,b)=>(a.horario||'').localeCompare(b.horario||''));
  const cont=document.getElementById('ppCards');
  if(!citas.length){{cont.innerHTML='<div style="padding:24px;color:#888;text-align:center">Sin citas agendadas para este dia.</div>';return;}}

  cont.innerHTML=citas.map(c=>{{
    const oc=String(c.oc||c.patente||'');
    const key=ppKey(ppSelectedDate,oc);
    const estado=getPpEstado(ppSelectedDate,oc);
    const pat=(c.patente||'').replace(/\?/g,'').trim()||'--';
    const marcaModelo=String(c.modelo||'').trim();
    const partes=marcaModelo.split(' ');
    const marca=partes[0]||'--';
    const modelo=partes.slice(1).join(' ')||'--';
    const open=(ppOpenKey===key);
    return`<div class="pp-card pp-${{estado}}" data-key="${{key}}">
      <div class="pp-head" onclick="togglePpCard('${{key}}')">
        <span class="pp-hora">${{c.horario||'--'}}</span>
        <span class="pp-plate">${{pat}}</span>
        <span class="pp-mmv">${{marcaModelo||'--'}}${{c.anio?' ('+c.anio+')':''}}</span>
        <span class="pp-svc">${{c.servicio||c.mantencion||'--'}}</span>
        <span class="pp-cliente">${{c.nombre||c.cliente||''}}</span>
        <span class="pp-status-badge ${{estado}}">${{estado==='realizado'?'✅ Realizado':'🕒 Pendiente'}}</span>
        <span class="pp-chevron">${{open?'▲':'▼'}}</span>
      </div>
      <div class="pp-body${{open?' open':''}}" id="ppbody-${{key}}">
        ${{open?ppDetalleHTML(c,oc,marca,modelo):''}}
      </div>
    </div>`;
  }}).join('');
}}

function togglePpCard(key){{
  ppOpenKey=(ppOpenKey===key)?null:key;
  renderPrepickingView();
}}

// Descuento rapido del 10% sobre el total del presupuesto de una cita del
// Pre-picking (14/07/2026, a pedido de Cristobal) — no se guarda en GitHub,
// es solo para simular/mostrar el presupuesto con descuento en pantalla y en
// el PDF exportado mientras se decide con el cliente.
function togglePpDescuento(key){{
  if(ppDescuentos[key])delete ppDescuentos[key];
  else ppDescuentos[key]=true;
  renderPrepickingView();
}}

/* ---- Mini-generador de .xlsx real (ZIP sin compresion + XML OOXML) ----
   El truco anterior de "tabla HTML con extension .xls" hacia que Excel mostrara
   la advertencia "el formato y la extension no coinciden" (reportado por Cristobal,
   14/07/2026) porque el contenido real era HTML, no un archivo Excel valido. Como
   el iframe del Planificador no tiene salida a CDNs externos (no se puede usar
   SheetJS), se arma un .xlsx real a mano: un ZIP (metodo "stored", sin compresion,
   valido segun el spec de ZIP) con las partes XML minimas que pide el formato
   OOXML de Excel — sin ninguna libreria externa, 100% API nativa del navegador
   (TextEncoder + Blob). Excel lo abre sin ninguna advertencia. */
function _crc32Tabla(){{
  const t=new Uint32Array(256);
  for(let n=0;n<256;n++){{
    let c=n;
    for(let k=0;k<8;k++)c=(c&1)?(0xEDB88320^(c>>>1)):(c>>>1);
    t[n]=c>>>0;
  }}
  return t;
}}
const _CRC32_TABLA=_crc32Tabla();
function _crc32(bytes){{
  let crc=0xFFFFFFFF;
  for(let i=0;i<bytes.length;i++)crc=_CRC32_TABLA[(crc^bytes[i])&0xFF]^(crc>>>8);
  return (crc^0xFFFFFFFF)>>>0;
}}
function _u16le(n){{return [n&0xFF,(n>>8)&0xFF];}}
function _u32le(n){{return [n&0xFF,(n>>8)&0xFF,(n>>16)&0xFF,(n>>24)&0xFF];}}
function _construirZip(archivos){{
  // archivos: [{{nombre, contenido}}] — todos guardados sin compresion (metodo 0)
  const enc=new TextEncoder();
  const locales=[],centrales=[];
  let offset=0;
  archivos.forEach(f=>{{
    const nombreB=enc.encode(f.nombre), datosB=enc.encode(f.contenido);
    const crc=_crc32(datosB), size=datosB.length;
    const local=new Uint8Array([
      ..._u32le(0x04034b50), ..._u16le(20), ..._u16le(0x0800), ..._u16le(0),
      ..._u16le(0), ..._u16le(0x21),
      ..._u32le(crc), ..._u32le(size), ..._u32le(size),
      ..._u16le(nombreB.length), ..._u16le(0),
      ...nombreB, ...datosB,
    ]);
    locales.push(local);
    centrales.push(new Uint8Array([
      ..._u32le(0x02014b50), ..._u16le(20), ..._u16le(20), ..._u16le(0x0800), ..._u16le(0),
      ..._u16le(0), ..._u16le(0x21),
      ..._u32le(crc), ..._u32le(size), ..._u32le(size),
      ..._u16le(nombreB.length), ..._u16le(0), ..._u16le(0),
      ..._u16le(0), ..._u16le(0), ..._u32le(0),
      ..._u32le(offset), ...nombreB,
    ]));
    offset+=local.length;
  }});
  const centralOffset=offset;
  const centralSize=centrales.reduce((s,c)=>s+c.length,0);
  const fin=new Uint8Array([
    ..._u32le(0x06054b50), ..._u16le(0), ..._u16le(0),
    ..._u16le(archivos.length), ..._u16le(archivos.length),
    ..._u32le(centralSize), ..._u32le(centralOffset), ..._u16le(0),
  ]);
  const total=offset+centralSize+fin.length;
  const out=new Uint8Array(total);
  let pos=0;
  locales.forEach(p=>{{out.set(p,pos);pos+=p.length;}});
  centrales.forEach(p=>{{out.set(p,pos);pos+=p.length;}});
  out.set(fin,pos);
  return out;
}}
function _xmlEsc(s){{return String(s==null?'':s).replace(/&/g,'&amp;').replace(/</g,'&lt;').replace(/>/g,'&gt;').replace(/"/g,'&quot;').replace(/'/g,'&apos;');}}
const _COL_LETRAS=['A','B','C','D','E','F','G','H','I','J'];

/* Exporta el listado completo de citas agendadas del dia/area seleccionado en el
   Pre-picking a un archivo Excel real (.xlsx) — Asesor, Modelo, Placa/Patente, Chasis,
   Kilometraje, Mantencion (segun kilometraje, igual criterio que Control de Taller),
   Nombre Propietario y Servicio (sin Motorizacion). Incluye titulo con Sucursal y
   fecha, encabezado con estilo, columnas anchas y fila superior congelada — mejora
   de diseno a pedido de Cristobal, 14/07/2026. */
function exportarListadoPrepickingExcel(){{
  const citas=getCitas(ppSelectedDate).filter(c=>citaArea(c)===currentArea)
    .slice().sort((a,b)=>(a.horario||'').localeCompare(b.horario||''));
  if(!citas.length){{toast('No hay citas agendadas para este dia');return;}}

  const cols=['Asesor','Modelo','Placa / Patente','Chasis','Kilometraje','Mantencion','Nombre Propietario','Servicio'];
  const nCols=cols.length;
  const ultCol=_COL_LETRAS[nCols-1];

  const dTab=planDates.find(d=>formatDate(d)===ppSelectedDate);
  const fechaLbl=dTab?(DIAS[dTab.getDay()]+' '+dTab.getDate()+' '+MESES[dTab.getMonth()]+' '+dTab.getFullYear()):ppSelectedDate;
  const tituloTxt=`PRE-PICKING — ${{SUCURSAL}} — ${{fechaLbl}}`;
  const ahoraStr=new Date().toLocaleDateString('es-CL')+' '+new Date().toLocaleTimeString('es-CL',{{hour:'2-digit',minute:'2-digit'}});
  const subTxt=`Generado ${{ahoraStr}} · ${{citas.length}} cita(s) agendada(s)`;

  const filasData=citas.map(c=>{{
    const pat=(c.patente||'').replace(/\?/g,'').trim()||'--';
    return [
      c.asesor||'', c.modelo||'', pat, c.vin||'',
      c.km?Number(c.km).toLocaleString('es-CL'):'',
      c.mantencion||'',
      c.nombre||c.cliente||'', c.servicio||c.mantencion||'',
    ];
  }});

  const HEADER_ROW=3, FIRST_DATA_ROW=4;
  const celda=(ref,val,s)=>`<c r="${{ref}}" t="inlineStr" s="${{s}}"><is><t xml:space="preserve">${{_xmlEsc(val)}}</t></is></c>`;

  const rowTitulo=`<row r="1" ht="22" customHeight="1">${{celda('A1',tituloTxt,1)}}</row>`;
  const rowSub=`<row r="2" ht="16" customHeight="1">${{celda('A2',subTxt,2)}}</row>`;
  const rowHead=`<row r="${{HEADER_ROW}}">${{cols.map((c,i)=>celda(_COL_LETRAS[i]+HEADER_ROW,c,3)).join('')}}</row>`;
  const rowsDatos=filasData.map((fila,idx)=>{{
    const r=FIRST_DATA_ROW+idx;
    const est=(idx%2===0)?4:5;
    const celdas=fila.map((val,cIdx)=>celda(_COL_LETRAS[cIdx]+r,val,est)).join('');
    return `<row r="${{r}}">${{celdas}}</row>`;
  }}).join('');

  const colsXml=`<cols>
<col min="1" max="1" width="22" customWidth="1"/>
<col min="2" max="2" width="24" customWidth="1"/>
<col min="3" max="3" width="16" customWidth="1"/>
<col min="4" max="4" width="20" customWidth="1"/>
<col min="5" max="5" width="14" customWidth="1"/>
<col min="6" max="6" width="22" customWidth="1"/>
<col min="7" max="7" width="32" customWidth="1"/>
<col min="8" max="8" width="30" customWidth="1"/>
</cols>`;

  const sheetXml=`<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<worksheet xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main">
<sheetViews><sheetView workbookViewId="0"><pane ySplit="${{HEADER_ROW}}" topLeftCell="A${{FIRST_DATA_ROW}}" activePane="bottomLeft" state="frozen"/></sheetView></sheetViews>
${{colsXml}}
<sheetData>${{rowTitulo}}${{rowSub}}${{rowHead}}${{rowsDatos}}</sheetData>
<mergeCells count="2"><mergeCell ref="A1:${{ultCol}}1"/><mergeCell ref="A2:${{ultCol}}2"/></mergeCells>
</worksheet>`;

  const stylesXml=`<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<styleSheet xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main">
<fonts count="4">
<font><sz val="10.5"/><name val="Calibri"/></font>
<font><b/><sz val="14"/><color rgb="FFFFFFFF"/><name val="Calibri"/></font>
<font><i/><sz val="9.5"/><color rgb="FF667788"/><name val="Calibri"/></font>
<font><b/><sz val="10"/><color rgb="FFFFFFFF"/><name val="Calibri"/></font>
</fonts>
<fills count="4">
<fill><patternFill patternType="none"/></fill>
<fill><patternFill patternType="gray125"/></fill>
<fill><patternFill patternType="solid"><fgColor rgb="FF0B2E63"/><bgColor indexed="64"/></patternFill></fill>
<fill><patternFill patternType="solid"><fgColor rgb="FFF3F6FA"/><bgColor indexed="64"/></patternFill></fill>
</fills>
<borders count="2">
<border><left/><right/><top/><bottom/><diagonal/></border>
<border><left style="thin"><color rgb="FFD8DEE3"/></left><right style="thin"><color rgb="FFD8DEE3"/></right><top style="thin"><color rgb="FFD8DEE3"/></top><bottom style="thin"><color rgb="FFD8DEE3"/></bottom></border>
</borders>
<cellStyleXfs count="1"><xf numFmtId="0" fontId="0" fillId="0" borderId="0"/></cellStyleXfs>
<cellXfs count="6">
<xf numFmtId="0" fontId="0" fillId="0" borderId="0" xfId="0"/>
<xf numFmtId="0" fontId="1" fillId="2" borderId="0" xfId="0" applyFont="1" applyFill="1" applyAlignment="1"><alignment horizontal="left" vertical="center"/></xf>
<xf numFmtId="0" fontId="2" fillId="0" borderId="0" xfId="0" applyFont="1"/>
<xf numFmtId="0" fontId="3" fillId="2" borderId="1" xfId="0" applyFont="1" applyFill="1" applyBorder="1" applyAlignment="1"><alignment vertical="center"/></xf>
<xf numFmtId="0" fontId="0" fillId="0" borderId="1" xfId="0" applyBorder="1"/>
<xf numFmtId="0" fontId="0" fillId="3" borderId="1" xfId="0" applyFill="1" applyBorder="1"/>
</cellXfs>
<cellStyles count="1"><cellStyle name="Normal" xfId="0" builtinId="0"/></cellStyles>
</styleSheet>`;

  const contentTypesXml=`<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<Types xmlns="http://schemas.openxmlformats.org/package/2006/content-types">
<Default Extension="rels" ContentType="application/vnd.openxmlformats-package.relationships+xml"/>
<Default Extension="xml" ContentType="application/xml"/>
<Override PartName="/xl/workbook.xml" ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet.main+xml"/>
<Override PartName="/xl/worksheets/sheet1.xml" ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.worksheet+xml"/>
<Override PartName="/xl/styles.xml" ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.styles+xml"/>
</Types>`;

  const rootRelsXml=`<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">
<Relationship Id="rId1" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/officeDocument" Target="xl/workbook.xml"/>
</Relationships>`;

  const workbookXml=`<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<workbook xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main" xmlns:r="http://schemas.openxmlformats.org/officeDocument/2006/relationships">
<sheets><sheet name="Pre-picking" sheetId="1" r:id="rId1"/></sheets></workbook>`;

  const workbookRelsXml=`<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">
<Relationship Id="rId1" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/worksheet" Target="worksheets/sheet1.xml"/>
<Relationship Id="rId2" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/styles" Target="styles.xml"/>
</Relationships>`;

  const zipBytes=_construirZip([
    {{nombre:'[Content_Types].xml', contenido:contentTypesXml}},
    {{nombre:'_rels/.rels', contenido:rootRelsXml}},
    {{nombre:'xl/workbook.xml', contenido:workbookXml}},
    {{nombre:'xl/_rels/workbook.xml.rels', contenido:workbookRelsXml}},
    {{nombre:'xl/styles.xml', contenido:stylesXml}},
    {{nombre:'xl/worksheets/sheet1.xml', contenido:sheetXml}},
  ]);

  const blob=new Blob([zipBytes],{{type:'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'}});
  const url=URL.createObjectURL(blob);
  const a=document.createElement('a');
  a.href=url;
  a.download=`Prepicking_${{SUCURSAL}}_${{ppSelectedDate}}.xlsx`;
  document.body.appendChild(a);
  a.click();
  document.body.removeChild(a);
  setTimeout(()=>URL.revokeObjectURL(url),1000);
  toast(`📊 Excel generado con ${{citas.length}} cita(s)`);
}}

/* ============================================================
   Repuestos/stock del Pre-picking desde el Cotizador de Mantenciones
   ------------------------------------------------------------
   Reemplaza el pipeline viejo (consolidar_OTs.py -> pauta_repuestos.json +
   Stock Repestos Costo.xlsx -> agenda_hoy.json -> cita.repuestos_sugeridos)
   por una lectura directa del bundle ya embebido del modulo "Cotizador de
   Mantenciones" (COTIZ_PP: indice/stock/pautas, ver bootstrap mas arriba),
   manteniendo la misma logica/alcance que tenia Pre-picking antes: repuestos
   de la pauta segun marca/modelo/km, con Stock/Ubicacion acotados a ESTA
   sucursal (si no hay stock aqui, se avisa en que otra bodega si hay) y
   codigos "tambien sirve" (alternativas). 22/07/2026, a pedido de Cristobal.
   Se mantiene un fallback a c.repuestos_sugeridos (pipeline viejo) por si el
   bundle del cotizador aun no cargo o la marca/modelo no esta cubierta ahi. */
function _cotizNormTxt(s){{
  return String(s||'').toUpperCase().normalize('NFD').replace(/[̀-ͯ]/g,'').replace(/[^A-Z0-9 ]/g,' ').replace(/\s+/g,' ').trim();
}}
function _cotizNormCod(c){{
  return c==null?'':String(c).toUpperCase().replace(/[^A-Z0-9]/g,'');
}}
function _esServicioMantencionPP(txt){{
  // Ademas de "Mantencion" (MANT), se acepta "Revision" (REVIS) — servicios tipo
  // "Revision 30 dias"/"Revision gratuita" son mantenciones programadas reales que
  // SI estan cubiertas por el Cotizador (misma pauta por km), solo que la Agenda
  // las etiqueta con otro texto. Bug real: Hyundai Grand i10 100km "Revision 30
  // dias" tenia pauta en el Cotizador pero Pre-picking la descartaba por el gate
  // de texto. 23/07/2026, a pedido de Cristobal.
  const t=_cotizNormTxt(txt);
  return t.indexOf('MANT')>=0 || t.indexOf('REVIS')>=0;
}}
function _cotizBuscarModelo(marcaModeloTexto){{
  if(!COTIZ_PP||!COTIZ_PP.indice) return null;
  const txt=_cotizNormTxt(marcaModeloTexto);
  if(!txt) return null;
  let mejorMarca=null, mejorScore=0;
  for(const m of (COTIZ_PP.indice.marcas||[])){{
    const nm=_cotizNormTxt(m.nombre);
    if(nm && txt.indexOf(nm)===0 && nm.length>mejorScore){{ mejorMarca=m; mejorScore=nm.length; }}
  }}
  if(!mejorMarca) return null;
  const resto=txt.slice(mejorScore).trim();
  if(!resto) return null;
  // Respaldo sin espacios: la Agenda escribe modelos con guion pegado
  // ("FORD F150"), pero el nombre oficial del Cotizador trae el guion
  // normalizado a espacio ("F-150" -> "F 150") — sin este respaldo la
  // comparacion exacta nunca calza y la pauta se pierde en silencio para
  // esos modelos (mismo bug ya corregido para el tempario en el
  // consolidador, sesion 09/07/2026 — replicado aca 22/07/2026).
  const restoSinEsp=resto.replace(/ /g,'');
  const restoTokens=resto.split(' ').filter(Boolean);
  // Match por PALABRAS (tokens), no por substring crudo — reescrito 23/07/2026 tras
  // confirmar con datos reales de la Agenda que el substring crudo dejaba fuera (o
  // matcheaba MAL) muchos vehiculos que el Cotizador si tiene cubiertos:
  //  - "FORD BRONCO" (Agenda) vs "Bronco Sport" (Cotizador): el texto de la Agenda
  //    es una ABREVIACION del nombre real — antes no matcheaba nada (null).
  //  - "HYUNDAI GRAND CRETA" (Agenda) vs "Creta Grand" (Cotizador): mismas palabras
  //    en OTRO ORDEN — antes matcheaba mal contra el modelo corto "Creta" (pauta
  //    equivocada, silenciosamente) en vez de "Creta Grand".
  // Se prueban 5 niveles de coincidencia (de mas a menos estricto) y se elige el de
  // mayor "tier"; dentro del mismo tier, el modelo con mas palabras (mas especifico).
  let mejorModelo=null, mejorPuntaje=-1;
  for(const mod of (mejorMarca.modelos||[])){{
    const nmod=_cotizNormTxt(mod.nombre);
    if(!nmod) continue;
    const nmodSinEsp=nmod.replace(/ /g,'');
    const nmodTokens=nmod.split(' ').filter(Boolean);
    let tier=0;
    if(restoSinEsp===nmodSinEsp) tier=4;                                                   // "F150" == "F 150"
    else if(nmodTokens.length===restoTokens.length && nmodTokens.every(t=>restoTokens.includes(t))) tier=3;  // mismas palabras, cualquier orden
    else if(nmodTokens.every(t=>restoTokens.includes(t))) tier=2;                           // el modelo completo esta dentro del texto (con palabras extra, ej. "RANGER RAPTOR" ⊇ "RANGER")
    else if(restoTokens.every(t=>nmodTokens.includes(t))) tier=1;                           // el texto es una abreviacion del modelo (ej. "BRONCO" ⊂ "BRONCO SPORT")
    else if(resto.indexOf(nmod)>=0 || restoSinEsp.indexOf(nmodSinEsp)>=0) tier=0.5;         // respaldo: substring crudo (compat con el comportamiento anterior)
    if(tier<=0) continue;
    const puntaje=tier*1000 + nmodTokens.length*10 + nmod.length*0.01;
    if(puntaje>mejorPuntaje){{ mejorModelo=mod; mejorPuntaje=puntaje; }}
  }}
  if(!mejorModelo||!mejorModelo.versiones||!mejorModelo.versiones.length) return null;
  return mejorModelo;
}}
function _cotizBuscarIntervalo(mejorModelo, kmTexto){{
  // Busca el km (exacto, o el mas cercano dentro de 5.000 km) recorriendo
  // TODAS las versiones/motorizaciones del modelo — no solo la primera. Un
  // mismo modelo (ej. Maverick) puede tener varias versiones con tablas de
  // mantencion DISTINTAS (unas parten en 16.000 km, otras en 10.000 km); si
  // solo se miraba la primera version, un kilometraje real que esa version
  // no cubre hacia que se perdiera la pauta (o peor, tomara repuestos de un
  // intervalo "cercano" de la version equivocada). Bug real encontrado en
  // produccion 22/07/2026 (Ford Maverick 10.000 km sin datos porque su
  // primera version solo tenia 16.000/32.000/...).
  if(!COTIZ_PP||!COTIZ_PP.pautas||!mejorModelo) return null;
  const kmNum=parseInt(String(kmTexto||'').replace(/[^0-9]/g,''),10);
  if(!kmNum) return null;
  let cercano=null, mejorDif=Infinity;
  for(const v of (mejorModelo.versiones||[])){{
    const pauta=COTIZ_PP.pautas[v.id];
    if(!pauta||!pauta.planes) continue;
    for(const plan of pauta.planes){{
      for(const iv of (plan.intervalos||[])){{
        if(Number(iv.km)===kmNum) return {{pauta,intervalo:iv}};
        const dif=Math.abs(Number(iv.km)-kmNum);
        if(dif<=5000 && dif<mejorDif){{ mejorDif=dif; cercano={{pauta,intervalo:iv}}; }}
      }}
    }}
  }}
  return cercano;
}}
function _cotizItemsParaVersion(versionId, kmTexto){{
  // Igual que _cotizBuscarIntervalo pero acotado a UNA sola version — se usa
  // cuando el usuario elige a mano la Version exacta en el selector manual de
  // Pre-picking (en vez de dejar que el sistema recorra todas las versiones
  // del modelo buscando la mas cercana). 22/07/2026.
  if(!COTIZ_PP||!COTIZ_PP.pautas||!versionId) return null;
  const pauta=COTIZ_PP.pautas[versionId];
  if(!pauta||!pauta.planes) return null;
  const kmNum=parseInt(String(kmTexto||'').replace(/[^0-9]/g,''),10);
  let cercano=null, mejorDif=Infinity;
  for(const plan of pauta.planes){{
    for(const iv of (plan.intervalos||[])){{
      if(kmNum && Number(iv.km)===kmNum) return {{pauta,intervalo:iv}};
      if(!kmNum) continue;
      const dif=Math.abs(Number(iv.km)-kmNum);
      if(dif<mejorDif){{ mejorDif=dif; cercano={{pauta,intervalo:iv}}; }}
    }}
  }}
  // Sin km reconocible: si la version solo tiene un intervalo, se usa ese.
  if(!cercano){{
    const todos=[]; for(const plan of pauta.planes){{ for(const iv of (plan.intervalos||[])) todos.push({{pauta,intervalo:iv}}); }}
    if(todos.length===1) cercano=todos[0];
  }}
  return cercano;
}}
function _cotizStockDe(codigo){{
  if(!COTIZ_PP||!COTIZ_PP.stock||!codigo) return null;
  return (COTIZ_PP.stock.items||{{}})[_cotizNormCod(codigo)]||null;
}}
// Respaldo: si el codigo no esta en el bundle chico del Cotizador (~400
// codigos), se busca en el catalogo completo (STOCK_FULL, ~33.000 codigos
// reales de Stock Repestos Costo.xlsx) — amplia la cobertura de stock a
// cualquier repuesto de la pauta, no solo a los que el Cotizador ya trae
// precalculados. 22/07/2026.
function _cotizCodBase(c){{
  // Quita el prefijo numerico de familia (ej. "13 XO5W30Q1SP" -> "XO5W30Q1SP")
  // ANTES de normalizar — mismo criterio que usa el consolidador Python al
  // limpiar codigos de Stock Repestos Costo.xlsx. Los codigos de las pautas
  // del Cotizador ya vienen sin ese prefijo, asi que aplicarselo tambien a
  // ellos es un no-op seguro (no hay digitos+espacio que quitar).
  return _cotizNormCod(String(c||'').replace(/^\s*\d+\s+/, ''));
}}
function _cotizStockDeCompleto(codigo){{
  if(!STOCK_FULL||!codigo) return null;
  const e=STOCK_FULL[_cotizCodBase(codigo)];
  return e?{{bodegas:e.bodegas}}:null;
}}
function _cotizLimpiarCod(c){{
  // Mismo criterio que el consolidador (Python, _codigos_relacionados): quita
  // un prefijo numerico de familia (ej. "13 BC4518D334DD" -> "BC4518D334DD")
  // y espacios/guiones, para comparar solo la parte real del codigo.
  return String(c||'').toUpperCase().replace(/^\d+\s+/, '').replace(/[\s\-]/g,'');
}}
function _cotizPrefijoComun(a,b){{
  let i=0; const n=Math.min(a.length,b.length);
  while(i<n && a[i]===b[i]) i++;
  return i;
}}
function _cotizCodigosRelacionados(codBase, codCand, minPrefijo, minRatio){{
  minPrefijo=minPrefijo||6; minRatio=minRatio||0.6;
  if(!codBase||!codCand||codBase===codCand) return false;
  const a=_cotizLimpiarCod(codBase), b=_cotizLimpiarCod(codCand);
  if(!a||!b) return false;
  const pref=_cotizPrefijoComun(a,b);
  return pref>=minPrefijo && pref>=a.length*minRatio && pref>=b.length*minRatio;
}}
// Busca codigos "relacionados" (mismo criterio que el pipeline viejo del
// consolidador, sesion 14/07/2026: prefijo comun >=6 caracteres Y >=60% del
// largo de AMBOS codigos) contra el catalogo COMPLETO de Stock — a diferencia
// del "alt" precalculado del bundle chico (a lo mas 1 equivalente), esto
// recorre los ~33.000 codigos reales y devuelve hasta `maxAlt` variantes,
// cada una con su propio stock ya calculado para esta sucursal. 22/07/2026.
function _cotizBuscarAlternativasCompletas(codigoOriginal, maxAlt){{
  maxAlt=maxAlt||3;
  if(!STOCK_FULL||!codigoOriginal) return [];
  const aLimpio=_cotizLimpiarCod(codigoOriginal);
  const candidatos=[];
  for(const cod in STOCK_FULL){{
    if(_cotizCodigosRelacionados(codigoOriginal, cod)) candidatos.push(cod);
  }}
  candidatos.sort((x,y)=>_cotizPrefijoComun(aLimpio,_cotizLimpiarCod(y))-_cotizPrefijoComun(aLimpio,_cotizLimpiarCod(x)));
  return candidatos.slice(0,maxAlt).map(cod=>{{
    const e=STOCK_FULL[cod];
    const r=_cotizStockEnSucursal({{bodegas:e.bodegas}});
    return {{
      codigo:cod,
      descripcion:e.descripcion||'',
      stock_sucursal:r.stockAqui,
      stock_otro:(!r.hayAqui && r.totalOtro>0)?r.totalOtro:null,
    }};
  }});
}}
function _cotizStockEnSucursal(s){{
  // Devuelve el desglose de un item de stock (s = COTIZ_PP.stock.items[cod])
  // para la sucursal actual: cuanto hay aqui, y cuanto suma en total en el
  // resto de bodegas (para el aviso "hay en otra sucursal"). Compartido entre
  // el item principal y sus alternativas/equivalentes.
  const bodegas=Array.isArray(s.bodegas)?s.bodegas:[];
  const normSuc=_cotizNormTxt(SUCURSAL);
  let stockAqui=0, huboBodegaAqui=false, totalOtro=0;
  const otras=[];
  for(const b of bodegas){{
    const nB=_cotizNormTxt(b.n);
    if(nB===normSuc || nB.indexOf(normSuc+' ')===0){{
      stockAqui+=Number(b.q)||0; huboBodegaAqui=true;
    }} else if(Number(b.q)>0){{
      otras.push(`${{b.n}} (${{b.q}})`);
      totalOtro+=Number(b.q)||0;
    }}
  }}
  return {{hayAqui:huboBodegaAqui&&stockAqui>0, stockAqui:huboBodegaAqui?stockAqui:null, otras, totalOtro}};
}}
function _cotizItemConStock(it){{
  const base={{nombre:it.nombre, codigo:it.codigo, cantidad:it.cantidad, precio_unitario:it.precioUnitario}};
  // Stock principal: primero el bundle chico del Cotizador (rapido, cubre los
  // ~400 codigos de las pautas); si no esta ahi, respaldo con el catalogo
  // completo (STOCK_FULL) para no quedar en "Sin dato" solo porque el bundle
  // acotado no incluye ese codigo puntual. 22/07/2026.
  const s=_cotizStockDe(it.codigo) || _cotizStockDeCompleto(it.codigo);
  // OJO: antes, si el codigo propio no tenia NINGUN registro de stock (s=null,
  // "Sin dato" en Stock/Ubicacion), la funcion cortaba aca con `alternativas:[]`
  // sin siquiera intentar buscar "tambien sirve" — asi, cualquier modelo cuyos
  // codigos no esten cargados directo en Stock (ej. Ford Territory) se quedaba
  // sin ninguna alternativa sugerida. Ahora se sigue el flujo igual y la
  // busqueda de alternativas (independiente de si el codigo propio tiene stock)
  // corre siempre. 23/07/2026, a pedido de Cristobal.
  const r=s?_cotizStockEnSucursal(s):{{hayAqui:false, stockAqui:null, otras:[], totalOtro:0}};
  // "Tambien sirve": se busca primero contra el catalogo COMPLETO (33.000
  // codigos reales, misma logica de codigo relacionado del pipeline viejo) —
  // mucho mas amplio que el "alt" precalculado del bundle chico del Cotizador
  // (a lo mas 1 equivalente). Si el catalogo completo aun no cargo o no
  // encuentra nada, cae al "alt" del bundle chico como respaldo. 22/07/2026.
  let alternativas=_cotizBuscarAlternativasCompletas(it.codigo, 3);
  if(!alternativas.length){{
    const sBundle=_cotizStockDe(it.codigo);
    if(sBundle && sBundle.alt){{
      const descripcion=(sBundle.via==='difuso'?'Equivalente (codigo aproximado)':'Mismo producto, otro formato/envase');
      const sAlt=_cotizStockDe(sBundle.alt);
      if(sAlt){{
        const rAlt=_cotizStockEnSucursal(sAlt);
        alternativas=[{{codigo:sBundle.alt, descripcion,
          stock_sucursal: rAlt.stockAqui,
          stock_otro: (!rAlt.hayAqui && rAlt.totalOtro>0)?rAlt.totalOtro:null}}];
      }} else {{
        alternativas=[{{codigo:sBundle.alt, descripcion, stock_sucursal:null, stock_otro:null}}];
      }}
    }}
  }}
  return {{
    ...base,
    stock: r.stockAqui,
    ubicacion: r.hayAqui?SUCURSAL:null,
    stock_otro_lugar: (!r.hayAqui && r.otras.length)?r.otras.join(', '):null,
    alternativas,
  }};
}}
function _cotizItemsParaCita(c, override){{
  if(!COTIZ_PP) return null;
  if(!_esServicioMantencionPP(c.servicio||c.mantencion||'')) return null;
  const km=c.mantencion||c.km||'';
  let res=null;
  if(override&&override.versionId){{
    // El usuario eligio una Version exacta a mano — se usa esa sola, sin
    // recorrer las demas versiones del modelo. 22/07/2026.
    res=_cotizItemsParaVersion(override.versionId, km);
  }} else if(override&&override.modelo){{
    // Eligio Marca+Modelo (y opcionalmente Ano) pero no una Version puntual: se
    // recorren todas las versiones de ESE modelo (mismo criterio automatico,
    // pero forzando el modelo elegido en vez del auto-detectado desde el texto
    // de la Agenda). Si ademas eligio Ano, se acota a solo las versiones que
    // cubren ese ano — 23/07/2026, a pedido de Cristobal.
    const modeloObj=_cotizModeloObjPorNombre(override.marca||'', override.modelo);
    if(modeloObj){{
      let modeloParaBuscar=modeloObj;
      if(override.anio){{
        const versionesAnio=(modeloObj.versiones||[]).filter(v=>
          Array.isArray(v.anios) && v.anios.map(String).includes(String(override.anio)));
        if(versionesAnio.length) modeloParaBuscar={{...modeloObj, versiones:versionesAnio}};
      }}
      res=_cotizBuscarIntervalo(modeloParaBuscar, km);
    }}
  }} else {{
    // Deteccion automatica (sin override): ademas de Marca/Modelo (desde c.modelo),
    // se usa el Ano del vehiculo que ya viaja en la cita (c.anio, campo real de la
    // Agenda) para acotar a las versiones que cubren ese ano — mismo criterio que
    // el override manual, pero automatico. Si el modelo no tiene el dato de Ano
    // cargado (la mayoria de las marcas del Cotizador aun no traen `anios` por
    // version) o el ano de la cita no calza con ninguna, se recorren TODAS las
    // versiones igual que antes (no se pierde pauta por esto). 23/07/2026, a
    // pedido de Cristobal.
    const mejorModelo=_cotizBuscarModelo(String(c.modelo||''));
    if(mejorModelo){{
      let modeloParaBuscar=mejorModelo;
      if(c.anio){{
        const versionesAnio=(mejorModelo.versiones||[]).filter(v=>
          Array.isArray(v.anios) && v.anios.map(String).includes(String(c.anio)));
        if(versionesAnio.length) modeloParaBuscar={{...mejorModelo, versiones:versionesAnio}};
      }}
      res=_cotizBuscarIntervalo(modeloParaBuscar, km);
    }}
  }}
  if(!res) return null;
  const items=(res.intervalo.items||[]).filter(it=>it.tipo==='repuesto').map(_cotizItemConStock);
  if(!items.length) return null;
  return {{items, horas:Number(res.intervalo.horas)||0, manoObra:Number(res.intervalo.manoObra)||0}};
}}
/* Punto unico de acceso: cotizador primero (respetando un override manual de
   Marca/Modelo/Version si el usuario eligio uno para esta cita puntual desde
   el selector de Pre-picking), con fallback al pipeline viejo
   (c.repuestos_sugeridos/horas_tempario/mano_obra_monto) si no hay match aun
   (bundle no cargado, marca no cubierta, o servicio no es mantencion). */
function _ppRepuestosDeCita(c, oc, fecha){{
  const override=(oc!=null)?getPpOverride(fecha||ppSelectedDate, oc):null;
  const cot=_cotizItemsParaCita(c, override);
  if(cot) return {{...cot, fuente:'cotizador'}};
  const viejos=Array.isArray(c.repuestos_sugeridos)?c.repuestos_sugeridos:[];
  return {{items:viejos, horas:Number(c.horas_tempario)||0, manoObra:Number(c.mano_obra_monto||0), fuente:'consolidador'}};
}}

// Selector manual de Marca/Modelo/Version/Ano para Pre-picking — agregado
// 22/07/2026 a pedido de Cristobal, ADEMAS de la deteccion automatica (no la
// reemplaza): por defecto todo sigue funcionando igual (auto-deteccion desde
// el texto de la Agenda), pero el usuario puede forzar a mano la version
// exacta cuando el texto no matchea bien o quiere ver otra motorizacion/ano.
function _ppSelectorModeloHTML(c,oc){{
  if(!COTIZ_PP||!COTIZ_PP.indice||!COTIZ_PP.indice.marcas||!COTIZ_PP.indice.marcas.length) return '';
  const override=getPpOverride(ppSelectedDate,oc);
  let marcaSel='', modeloSel='', anioSel='', versionSel='';
  if(override){{ marcaSel=override.marca||''; modeloSel=override.modelo||''; anioSel=override.anio||''; versionSel=override.versionId||''; }}
  else {{
    const auto=_cotizBuscarModelo(String(c.modelo||''));
    if(auto){{
      for(const m of COTIZ_PP.indice.marcas){{ if((m.modelos||[]).includes(auto)){{marcaSel=m.nombre;break;}} }}
      modeloSel=auto.nombre;
    }}
  }}
  const marcaObj=COTIZ_PP.indice.marcas.find(m=>m.nombre===marcaSel);
  const modeloObj=marcaObj?(marcaObj.modelos||[]).find(md=>md.nombre===modeloSel):null;
  const optMarca=`<option value="">-- Marca --</option>`+COTIZ_PP.indice.marcas.map(m=>
    `<option value="${{esc(m.nombre)}}"${{m.nombre===marcaSel?' selected':''}}>${{esc(m.nombre)}}</option>`).join('');
  const optModelo=marcaObj?(`<option value="">-- Modelo --</option>`+(marcaObj.modelos||[]).map(md=>
    `<option value="${{esc(md.nombre)}}"${{md.nombre===modeloSel?' selected':''}}>${{esc(md.nombre)}}</option>`).join('')):`<option value="">-- Modelo --</option>`;
  // Selector de Ano (nuevo, 23/07/2026, a pedido de Cristobal): lista los anos
  // cubiertos por CUALQUIER version del modelo elegido (union de v.anios de todas
  // las versiones), ordenados de mas reciente a mas antiguo. Es solo un filtro
  // visual para acortar la lista de Version en modelos con muchas motorizaciones —
  // al elegir un Ano, la lista de Version se acota a las versiones que lo cubren.
  let aniosDisponibles=[];
  if(modeloObj){{
    const set=new Set();
    for(const v of (modeloObj.versiones||[])){{ for(const a of (Array.isArray(v.anios)?v.anios:[])) set.add(String(a)); }}
    aniosDisponibles=Array.from(set).sort((a,b)=>Number(b)-Number(a));
  }}
  // El Ano de la cita (c.anio, dato real que ya trae la Agenda) se usa para
  // preseleccionar el Ano automaticamente cuando NO hay override manual — mismo
  // criterio que ya aplica _cotizItemsParaCita para acotar la busqueda de la
  // pauta. Si el modelo detectado no tiene ese ano cargado (marca aun sin
  // `anios` por version) o el ano de la cita no calza con ninguno, el select
  // queda en "todos" igual que antes (no se fuerza un valor invalido). 23/07/2026.
  if(!override && c.anio && aniosDisponibles.includes(String(c.anio))) anioSel=String(c.anio);
  const optAnio=modeloObj?(`<option value="">-- Ano (todos) --</option>`+aniosDisponibles.map(a=>
    `<option value="${{esc(a)}}"${{a===anioSel?' selected':''}}>${{esc(a)}}</option>`).join('')):`<option value="">-- Ano --</option>`;
  const versionesFiltradas=modeloObj?(modeloObj.versiones||[]).filter(v=>
    !anioSel || (Array.isArray(v.anios) && v.anios.map(String).includes(anioSel))
  ):[];
  const optVersion=modeloObj?(`<option value="">-- Version (todas) --</option>`+versionesFiltradas.map(v=>{{
      const anios=Array.isArray(v.anios)&&v.anios.length?` (${{v.anios.join('/')}})`:'';
      return `<option value="${{esc(v.id)}}"${{v.id===versionSel?' selected':''}}>${{esc(v.nombre)}}${{anios}}</option>`;
    }}).join('')):`<option value="">-- Version --</option>`;
  const estadoTxt=override?`<span class="pp-modelo-manual">✋ Manual</span>`:`<span class="pp-modelo-auto">🤖 Automatico (desde Agenda)</span>`;
  const btnReset=override?`<button class="pp-btn-reset-modelo" onclick="ppCambiarModeloSel('${{ppSelectedDate}}','${{oc}}','reset','')">↩️ Volver a automatico</button>`:'';
  return `<div class="pp-modelo-sel">
    <div class="pp-modelo-sel-tit">🚗 Marca / Modelo / Ano / Version &nbsp;—&nbsp; ${{estadoTxt}}</div>
    <div class="pp-modelo-sel-row">
      <select onchange="ppCambiarModeloSel('${{ppSelectedDate}}','${{oc}}','marca',this.value)">${{optMarca}}</select>
      <select onchange="ppCambiarModeloSel('${{ppSelectedDate}}','${{oc}}','modelo',this.value)">${{optModelo}}</select>
      <select onchange="ppCambiarModeloSel('${{ppSelectedDate}}','${{oc}}','anio',this.value)">${{optAnio}}</select>
      <select onchange="ppCambiarModeloSel('${{ppSelectedDate}}','${{oc}}','version',this.value)">${{optVersion}}</select>
      ${{btnReset}}
    </div>
  </div>`;
}}
function ppDetalleHTML(c,oc,marca,modelo){{
  const _rep=_ppRepuestosDeCita(c,oc,ppSelectedDate);
  const items=_rep.items;
  const filas=[
    ['OC / Folio', oc||'--'],
    ['Patente', (c.patente||'').replace(/\?/g,'').trim()||'--'],
    ['Nombre completo cliente', c.nombre||c.cliente||'--'],
    ['Rut cliente', c.rut||'--'],
    ['Marca', marca],
    ['Modelo', modelo],
    ['Ano', c.anio||'--'],
    ['VIN', c.vin||'--'],
    ['Kilometraje', c.km?Number(c.km).toLocaleString('es-CL')+' km':'--'],
    ['Servicio', c.servicio||'--'],
    ['Mantencion kilometraje', c.mantencion||'--'],
    ['Asesor', c.asesor||'--'],
    ['Sucursal', c.sucursal||SUCURSAL],
    ['Estado agenda', c.estado==='finalizado'?'🧍 Finalizado':(c.ingresado?'🎟️ Ingresado':'🚗 Pendiente de ingreso')],
  ];
  const detGrid=`<div class="pp-det-grid">${{filas.map(([l,v])=>
    `<div class="pp-det-row"><span class="pp-lbl">${{esc(l)}}</span><span class="pp-val">${{esc(v)}}</span></div>`
  ).join('')}}</div>`;
  const selectorModelo=_ppSelectorModeloHTML(c,oc);

  const manoObra=Number(_rep.manoObra||0);
  const _ppKey=ppKey(ppSelectedDate,oc);
  const _conDescuento=!!ppDescuentos[_ppKey];
  const _totalRep=(items||[]).reduce((s,it)=>s+(Number(it.precio_unitario||0)*Number(it.cantidad||0)),0);
  const _totalGral=_totalRep+manoObra;
  const _totalDesc=_totalGral*0.9;
  let tabla;
  if(items&&items.length){{
    // Stock/Ubicacion se calculan aqui mismo (_cotizItemConStock) acotados a ESTA
    // sucursal desde el bundle del Cotizador de Mantenciones (COTIZ_PP.stock) — si
    // esta sucursal no tiene stock, se avisa en que otra bodega si hay
    // (it.stock_otro_lugar). Cada repuesto puede traer un codigo "alternativas"
    // (mismo producto equivalente, precomputado en el bundle) que tambien sirve.
    // 22/07/2026: repuestos/stock ahora vienen del Cotizador embebido en vez del
    // pipeline viejo (pauta_repuestos.json + Stock Repestos Costo.xlsx).
    tabla=`<table class="pptable"><thead><tr>
        <th>Nombre / Descripcion repuesto</th><th>Codigo</th><th>Cantidad</th>
        <th>Stock (${{esc(SUCURSAL)}})</th><th>Ubicacion</th><th>Valor Neto</th></tr></thead><tbody>
      ${{items.map(it=>{{
        const alts=Array.isArray(it.alternativas)?it.alternativas:[];
        const altsHtml=alts.length?`<div class="pp-alt"><div class="pp-alt-lbl">🔁 Tambien sirve</div>${{alts.map(a=>{{
            let stCls='no', stTxt='Sin stock';
            if(a.stock_sucursal!=null){{ stCls='si'; stTxt=`Stock aqui: ${{a.stock_sucursal}}`; }}
            else if(a.stock_otro!=null){{ stCls='otro'; stTxt=`En otra sucursal: ${{a.stock_otro}}`; }}
            return `<div class="pp-alt-item"><b>${{esc(a.codigo)}}</b>
              <span class="pp-alt-desc">${{esc(a.descripcion||'')}}</span>
              <span class="pp-alt-stock ${{stCls}}">${{stTxt}}</span></div>`;
          }}).join('')}}</div>`:'';
        // OJO: usar it.stock (truthy, >0) y no it.stock!=null — 0 es un valor valido
        // (hay registro de esa sucursal pero con cero unidades) y en ese caso igual
        // hay que avisar donde SI hay stock, no mostrar la sucursal como si tuviera.
        const ubicCell=it.stock?esc(it.ubicacion||SUCURSAL):
          (it.stock_otro_lugar?`<span style="color:#c87900">Sin stock en ${{esc(SUCURSAL)}}<br>Hay en: ${{esc(it.stock_otro_lugar)}}</span>`:'Sin dato');
        return `<tr>
        <td>${{esc(it.nombre||'--')}}${{altsHtml}}</td>
        <td>${{esc(it.codigo||'--')}}</td>
        <td>${{esc(it.cantidad||'--')}}</td>
        <td>${{it.stock!=null?it.stock:'Sin dato'}}</td>
        <td>${{ubicCell}}</td>
        <td>${{fmtCLP(it.precio_unitario)}}</td>
      </tr>`;
      }}).join('')}}
      <tr class="pp-mo-row"><td colspan="2">Cantidad de mano de obra</td>
        <td>${{(Number(_rep.horas)||0).toFixed(1)}} h</td>
        <td colspan="2"></td><td>${{fmtCLP(manoObra)}}</td></tr>
      <tr class="pp-tot-row"><td colspan="5">TOTAL NETO</td><td>${{fmtCLP(_totalGral)}}</td></tr>
      ${{_conDescuento?`
      <tr class="pp-desc-row"><td colspan="5">Descuento (10%)</td><td>- ${{fmtCLP(_totalGral-_totalDesc)}}</td></tr>
      <tr class="pp-tot-row pp-tot-desc"><td colspan="5">TOTAL CON DESCUENTO</td><td>${{fmtCLP(_totalDesc)}}</td></tr>`:''}}
    </tbody></table>`;
  }} else {{
    tabla=`<div style="padding:10px;background:#fff3cd;border-radius:6px;font-size:12px;color:#8a5a00;margin-bottom:12px">
      ⚠️ Sin pauta de repuestos disponible — no es una mantencion por kilometraje reconocida,
      o la marca/modelo/kilometraje no esta cubierto por la pauta cargada.
    </div>`;
  }}

  // Se pasa la cita via atributo data-cita (mismo patron ya probado que usa
  // data-cita en las tarjetas de Programacion) y se lee con JSON.parse(el.dataset.cita)
  // en vez de inyectar el JSON como argumento del onclick — con doble comillas
  // escapadas a &quot; el navegador las decodifica ANTES de ejecutar el JS del
  // atributo, dejando codigo invalido en tiempo de ejecucion y el boton sin hacer
  // nada (bug real encontrado en produccion, corregido 13/07/2026).
  const _citaJson=JSON.stringify(c).replace(/'/g,"&#39;");
  return detGrid+selectorModelo+tabla+`<div class="pp-actions">
      <button class="pp-btn-pdf" data-cita='${{_citaJson}}' data-descuento="${{_conDescuento?'1':'0'}}" onclick="exportarPresupuestoPDF(this)">📄 Exportar Presupuesto (PDF)</button>
      <button class="pp-btn-desc${{_conDescuento?' activo':''}}" onclick="togglePpDescuento('${{_ppKey}}')">${{_conDescuento?'✖️ Quitar descuento 10%':'🏷️ Aplicar descuento 10%'}}</button>
      <button class="pp-btn-real" onclick="setPpEstado('${{ppSelectedDate}}','${{oc}}','realizado')">✅ Marcar Realizado</button>
      <button class="pp-btn-pend" onclick="setPpEstado('${{ppSelectedDate}}','${{oc}}','pendiente')">🕒 Marcar Pendiente</button>
    </div>`;
}}

/* Exporta el presupuesto como una pagina imprimible (con el logo Curifor) en una
   pestana nueva y dispara el dialogo de impresion del navegador — el usuario elige
   "Guardar como PDF" ahi. No depende de ninguna libreria externa (el iframe del
   Planificador no tiene salida a CDNs externos), asi que este es el camino robusto. */
function exportarPresupuestoPDF(el){{
  let c;
  try{{c=JSON.parse(el.dataset.cita||'{{}}');}}catch(e){{toast('No se pudo generar el presupuesto');return;}}
  const conDescuento=el.dataset.descuento==='1';
  const oc=String(c.oc||c.patente||'');
  const pat=(c.patente||'').replace(/\?/g,'').trim()||'--';
  const marcaModelo=String(c.modelo||'').trim();
  const partes=marcaModelo.split(' ');
  const marca=partes[0]||'--';
  const modelo=partes.slice(1).join(' ')||'--';
  const _rep=_ppRepuestosDeCita(c,oc,ppSelectedDate);
  const items=_rep.items;
  const manoObra=Number(_rep.manoObra||0);
  const totalRep=items.reduce((s,it)=>s+(Number(it.precio_unitario||0)*Number(it.cantidad||0)),0);
  const totalGral=totalRep+manoObra;
  const totalDesc=totalGral*0.9;
  const hoy=new Date().toLocaleDateString('es-CL');
  const filas=[
    ['OC / Folio', oc||'--'],['Patente', pat],
    ['Cliente', c.nombre||c.cliente||'--'],['Rut cliente', c.rut||'--'],
    ['Marca', marca],['Modelo', modelo],['VIN', c.vin||'--'],
    ['Kilometraje', c.km?Number(c.km).toLocaleString('es-CL')+' km':'--'],
    ['Servicio', c.servicio||'--'],['Mantencion', c.mantencion||'--'],
    ['Asesor', c.asesor||'--'],['Sucursal', c.sucursal||SUCURSAL],
  ];
  const filasHtml=filas.map(([l,v])=>`<tr><td class="lbl">${{esc(l)}}</td><td>${{esc(v)}}</td></tr>`).join('');
  const itemsHtml=items.map(it=>`<tr>
      <td>${{esc(it.nombre||'--')}}</td><td>${{esc(it.codigo||'--')}}</td>
      <td style="text-align:center">${{esc(it.cantidad||'--')}}</td>
      <td style="text-align:right">${{fmtCLP(it.precio_unitario)}}</td>
      <td style="text-align:right">${{fmtCLP(Number(it.precio_unitario||0)*Number(it.cantidad||0))}}</td>
    </tr>`).join('');
  const html=`<!DOCTYPE html><html lang="es"><head><meta charset="UTF-8">
    <title>Presupuesto ${{esc(pat)}}</title>
    <style>
      *{{box-sizing:border-box;font-family:Arial,Helvetica,sans-serif;}}
      body{{padding:28px;color:#222;}}
      .hd{{display:flex;align-items:center;gap:16px;border-bottom:3px solid #0b2e63;padding-bottom:12px;margin-bottom:18px;}}
      .hd img{{height:54px;}}
      .hd h1{{font-size:18px;color:#0b2e63;margin:0;}}
      .hd p{{margin:2px 0 0;color:#667;font-size:12px;}}
      table.det{{width:100%;border-collapse:collapse;margin-bottom:18px;}}
      table.det td{{padding:4px 8px;font-size:12px;border-bottom:1px solid #eee;}}
      table.det td.lbl{{color:#889;width:180px;font-weight:600;}}
      table.rep{{width:100%;border-collapse:collapse;margin-bottom:12px;}}
      table.rep th{{background:#0b2e63;color:#fff;padding:6px 8px;font-size:11px;text-align:left;}}
      table.rep td{{border:1px solid #ddd;padding:6px 8px;font-size:12px;}}
      .tot-row td{{font-weight:700;background:#f4f6f8;}}
      .foot{{margin-top:24px;font-size:10px;color:#999;text-align:center;}}
      @media print{{.no-print{{display:none;}}}}
    </style></head><body>
    <div class="hd">${{LOGO_URI?`<img src="${{LOGO_URI}}"/>`:''}}
      <div><h1>Presupuesto de Mantencion — Curifor S.A</h1>
      <p>Generado ${{hoy}} · ${{SUCURSAL}}</p></div></div>
    <table class="det">${{filasHtml}}</table>
    <table class="rep"><thead><tr><th>Nombre / Descripcion repuesto</th><th>Codigo</th>
      <th>Cantidad</th><th>Valor Unitario</th><th>Subtotal</th></tr></thead><tbody>
      ${{itemsHtml||'<tr><td colspan="5" style="text-align:center;color:#888">Sin repuestos sugeridos</td></tr>'}}
      <tr class="tot-row"><td colspan="2">Mano de obra</td>
        <td style="text-align:center">${{(Number(_rep.horas)||0).toFixed(1)}} h</td>
        <td></td><td style="text-align:right">${{fmtCLP(manoObra)}}</td></tr>
      <tr class="tot-row"><td colspan="4">TOTAL NETO</td><td style="text-align:right">${{fmtCLP(totalGral)}}</td></tr>
      ${{conDescuento?`
      <tr><td colspan="4" style="color:#c0392b">Descuento (10%)</td><td style="text-align:right;color:#c0392b">- ${{fmtCLP(totalGral-totalDesc)}}</td></tr>
      <tr class="tot-row" style="background:#e8f5e9"><td colspan="4">TOTAL CON DESCUENTO</td><td style="text-align:right">${{fmtCLP(totalDesc)}}</td></tr>`:''}}
    </tbody></table>
    <div class="foot">Curifor S.A — Presupuesto referencial generado desde el Pre-picking del Planificador de Taller.</div>
    <div class="no-print" style="text-align:center;margin-top:16px">
      <button onclick="window.print()" style="padding:8px 20px;font-size:13px;cursor:pointer">🖨️ Imprimir / Guardar como PDF</button>
    </div>
    </body></html>`;
  const w=window.open('','_blank');
  if(!w){{toast('El navegador bloqueo la ventana emergente — habilitala para exportar el presupuesto');return;}}
  w.document.write(html);
  w.document.close();
}}

function loadData(){{
  // Datos inyectados desde Python — sin fetch, sin async, carga inmediata
  agendaData = Object.keys(_AGENDA_INIT).length ? _AGENDA_INIT : null;
  ctrlData   = Object.keys(_CTRL_INIT).length   ? _CTRL_INIT   : null;
  ctrlSha    = _CTRL_SHA || null;
  ppData     = Object.keys(_PP_INIT).length     ? _PP_INIT     : {{}};
  ppSha      = _PP_SHA || null;
  prodData   = Object.keys(_PROD_INIT).length   ? _PROD_INIT   : null;
  asesoresSucursal = getAsesoresSucursal();
  buildOrdenes();
  const _importadas=autoImportarCitas();
  const _reparadas=_repararBloquesMultiDia();
  document.getElementById('loading').style.display='none';
  document.getElementById('main').style.display='block';
  renderJPCB();
  selectedDate=formatDate(planDates[0]);
  renderDateTabs();
  ppSelectedDate=formatDate(planDates[1]); // Pre-picking abre por defecto en "Manana"
  if(_importadas>0||_reparadas>0){{
    saveCtrl();
  }} else if(PUEDE_EDITAR){{
    // No hubo citas nuevas que autoimportar (no dispara ningun saveCtrl), pero igual
    // conviene refrescar una vez al abrir la pagina: trae cambios recientes de otra
    // sesion Y limpia duplicados de patente+OT que hayan quedado guardados de una
    // carrera anterior (ver _dedupOrdenesPorPatenteOT), sin esperar a la primera
    // edicion propia. Solo con PUEDE_EDITAR — un usuario de solo lectura no tiene
    // por que disparar este refresco.
    _refrescarCtrlSha();
  }}
}}

function buildOrdenes(){{
  const s=ctrlData&&ctrlData[SUCURSAL];
  ordenes=s&&s.ordenes?s.ordenes.map(o=>Object.assign({{}},o)):[];
  tecnicos=s&&s.tecnicos?[...s.tecnicos]:[];
  _snapshotOrdenes();_snapshotBloques();
}}
/* Guarda el estado "ya sincronizado" de ordenes/bloques — se usa como punto de
   comparacion para saber que cambio localmente desde la ultima vez que se
   leyo/escribio GitHub, y asi mezclar en vez de pisar los cambios de otros
   usuarios que hayan editado la MISMA sucursal mientras tanto. */
function _snapshotOrdenes(){{
  _ordenesBaseline=new Map(ordenes.map(o=>[String(o.id),JSON.stringify(o)]));
}}
function _snapshotBloques(){{
  const b=(ctrlData&&ctrlData[SUCURSAL]&&ctrlData[SUCURSAL].bloques)||{{}};
  _bloquesBaseline=new Map(Object.keys(b).map(k=>[k,JSON.stringify(b[k])]));
}}
/* ─── Auto-reparacion de bloques multi-dia al cargar la pagina ───
   30/07/2026: antes de este fix, un vehiculo con "Salida (fecha)" varios dias despues
   de "Ingreso (fecha)" solo quedaba con un bloque en el primer dia (bug de
   upsertBloqueDesdeOrden, ver comentario de esa funcion) — y ese bloque ademas dejaba de
   verse apenas esa fecha salia de la ventana visible del Planificador (hoy+4). Ordenes
   que ya quedaron guardadas asi (creadas ANTES del fix) no se corrigen solas con solo
   subir el codigo nuevo — el bloque ya esta persistido en control_taller.json y nada
   vuelve a llamar a upsertBloqueDesdeOrden a menos que alguien edite un campo de esa
   orden a mano. Esta funcion revisa, en cada carga de la pagina, las ordenes activas que
   son multi-dia (tecnico+horarios+fechas completos, con Salida (fecha) 1+ dia despues de
   Ingreso (fecha)) y, si falta el bloque de ALGUNO de los dias que deberia cubrir,
   recalcula toda la orden con upsertBloqueDesdeOrden — asi el caso ya reportado (y
   cualquier otro igual) se autocorrige la primera vez que alguien abre el Planificador
   despues de subir este fix, sin tener que volver a tocar cada orden a mano. No toca
   ordenes de un solo dia (turno normal o cruce de una sola noche) que ya esten bien. */
function _repararBloquesMultiDia(){{
  if(!ctrlData||!ctrlData[SUCURSAL])return 0;
  if(!ctrlData[SUCURSAL].bloques)ctrlData[SUCURSAL].bloques={{}};
  let reparadas=0;
  const conCoflicto=[];
  ordenes.forEach(o=>{{
    if(o.cerrada)return;
    const tecOk=o.tecnico!==null&&o.tecnico!==undefined&&o.tecnico!=='';
    if(!tecOk||!o.ingreso_taller||!o.salida_taller||!o.ingreso||!o.salida)return;
    const fIni=parseDateISO(o.ingreso), fSal=parseDateISO(o.salida);
    if(!fIni||!fSal)return;
    const nDias=Math.round((fSal-fIni)/86400000);
    if(nDias<1)return; // no es multi-dia — nada que reparar aca
    const dateStr=isoToDdmmyyyy(o.ingreso);
    const diasSpan=[];
    for(let i=0;i<=nDias;i++)diasSpan.push(addDiasFecha(dateStr,i));
    const prefijo='ct'+o.id;
    const cubreTodos=diasSpan.every(ds=>(ctrlData[SUCURSAL].bloques[ds]||[]).some(b=>b.id===prefijo||String(b.id).startsWith(prefijo+'_')));
    if(!cubreTodos){{
      // Con la regla imperativa (30/07/2026) upsertBloqueDesdeOrden puede rechazar la
      // reparacion si, al recalcular, esta orden ahora chocaria con otro vehiculo del
      // mismo tecnico (dato viejo inconsistente) — en ese caso no se fuerza nada, se
      // deja para revision manual y se avisa junto al resto al terminar.
      const res=upsertBloqueDesdeOrden(o);
      if(res&&res.ok===false){{conCoflicto.push(o.patente||o.ot||o.id);}}
      else{{reparadas++;}}
    }}
  }});
  if(conCoflicto.length){{
    toast('⚠️ '+conCoflicto.length+' vehiculo(s) con choque de horario detectado al reparar ('+conCoflicto.join(', ')+') — revisar tecnico/horario a mano', 5000);
  }}
  return reparadas;
}}
function detectTipo(c){{const s=(c.servicio||c.mantencion||'').toLowerCase();if(s.includes('recall'))return'recall';if(s.includes('mant')||c.mantencion)return'mant';if(s.includes('diag'))return'diag';if(s.includes('rep'))return'rep';return'ot';}}

/* ─── Control de Taller — helpers de fecha/patente ─── */
const normPat=p=>String(p||'').replace(/[^A-Za-z0-9]/g,'').toUpperCase().slice(0,8);
function isoToday(){{const d=new Date();return d.getFullYear()+'-'+String(d.getMonth()+1).padStart(2,'0')+'-'+String(d.getDate()).padStart(2,'0');}}
function ddmmyyyyToIso(s){{
  if(!s)return isoToday();
  const p=String(s).split('/');
  if(p.length!==3)return isoToday();
  let anio=p[2].trim();
  if(anio.length===2)anio='20'+anio; // la agenda a veces trae el año en 2 digitos (DD/MM/YY)
  return anio+'-'+p[1].padStart(2,'0')+'-'+p[0].padStart(2,'0');
}}
function esc(s){{return String(s==null?'':s).replace(/&/g,'&amp;').replace(/</g,'&lt;').replace(/>/g,'&gt;').replace(/"/g,'&quot;').replace(/'/g,'&#39;');}}
function parseDateISO(s){{if(!s)return null;const p=String(s).split('-');if(p.length!==3)return null;return new Date(+p[0],+p[1]-1,+p[2]);}}
function calcDiasEnTaller(o){{const ini=parseDateISO(o.ingreso);if(!ini)return'--';const fin=o.salida?parseDateISO(o.salida):new Date();if(!fin)return'--';const d=Math.floor((fin-ini)/86400000);return d>=0?d:0;}}
function calcDiasFaltantesEta(o){{if(!o.eta)return'--';const e=parseDateISO(o.eta);if(!e)return'--';const d=Math.ceil((e-new Date())/86400000);return d;}}
function isoToDdmmyyyy(iso){{const p=String(iso||'').split('-');if(p.length!==3)return formatDate(planDates[0]);return p[2]+'/'+p[1]+'/'+p[0];}}
function isWithinLastDays(iso,days){{const d=parseDateISO(iso);if(!d)return false;const diff=(new Date(new Date().toDateString())-d)/86400000;return diff<=days&&diff>=-1;}}
function esPatenteExcluida(p){{return/^SP\d{{4}}$/i.test(normPat(p));}}
function ctEstadoIcon(o){{return o.salida?'🧍':'🎟️';}}
function ctEstadoTitle(o){{return o.salida?'Retirado ('+o.salida+')':'En taller (ingresado, sin marcar salida)';}}
function timeToMin(t){{if(!t)return null;const p=String(t).split(':');if(p.length<2)return null;const h=+p[0],m=+p[1];if(isNaN(h)||isNaN(m))return null;return h*60+m;}}

/* ─── Turnos que cruzan a la jornada siguiente (ej. ingreso 16:00, salida 10:00 del
   dia siguiente) — 22/07/2026, a pedido de Cristobal. Un turno se considera "overnight"
   cuando Salida Taller es menor o igual a Ingreso Taller: en ese caso ya no cabe en un
   solo dia, y hay que partirlo en 2 bloques (uno en la jornada de ingreso, hasta el
   cierre del dia; otro en la jornada siguiente, desde la apertura hasta la Salida
   Taller real) para que el tecnico lo vea asignado en AMBOS dias en el grid. ─── */
function addDiasFecha(dateStr,n){{const d=parseDate(dateStr);d.setDate(d.getDate()+n);return formatDate(d);}}
function esCruceNoche(o){{
  const i=timeToMin(o.ingreso_taller),f=timeToMin(o.salida_taller);
  return i!==null&&f!==null&&f<=i;
}}
/* Fecha+hora real de Salida Taller, agregando el dia siguiente si el turno cruza la
   medianoche — para mostrar en el JPCB/modal/Planificador el dato completo (no solo la
   hora, que por si sola confundiria un turno nocturno con uno diurno). */
function fechaSalidaTallerTexto(o){{
  if(!o.salida_taller)return '';
  if(esCruceNoche(o))return addDiasFecha(isoToDdmmyyyy(o.ingreso),1)+' '+o.salida_taller;
  return o.salida_taller;
}}

/* ─── Duracion estimada (tempario de mantencion) ───
   Si la orden trae horas_tempario (cruce marca/modelo + km de mantencion contra el
   tempario de mano de obra, calculado en el consolidador desde la cita de Agenda),
   se usa esa duracion para proponer la Salida Taller apenas se asigna tecnico.
   Sigue siendo 100% editable a mano (salida_taller o duracion_min) por si hay
   atrasos — esto solo fija un valor inicial razonable, no un limite.
   NOTA (17/07/2026): Horario Ingreso (hora_rec) y Horario Entrega (hora_compromiso)
   son ahora puramente informativos — vienen de la Agenda / son el compromiso de
   entrega al cliente, y ya NO alimentan el Planificador de Tecnicos ni el Tiempo
   Estimado. Esos dos pasaron a depender de Ingreso Taller / Salida Taller. */
function calcularSalidaTaller(o){{
  const ini=timeToMin(o.ingreso_taller);
  if(ini===null)return '';
  const horas=(typeof o.horas_tempario==='number'&&o.horas_tempario>0)?o.horas_tempario:((o.duracion_min||60)/60);
  const dateStr=isoToDdmmyyyy(o.ingreso);
  const end=getEnd(dateStr);
  let fin=Math.min(ini+Math.round(horas*60),end);
  fin=Math.max(fin,Math.min(ini+STEP,end));
  fin=Math.ceil(fin/STEP)*STEP;
  return hhmm(fin);
}}
/* Duracion en minutos entre Ingreso Taller y Salida Taller, considerando turnos que
   cruzan a la jornada siguiente (Salida <= Ingreso => se cuenta lo que falta hasta
   medianoche + lo que ya paso desde la apertura del dia siguiente). */
function duracionTallerMin(o){{
  const iniM=timeToMin(o.ingreso_taller),finM=timeToMin(o.salida_taller);
  if(iniM===null||finM===null)return null;
  return finM<=iniM?(24*60-iniM)+finM:finM-iniM;
}}
function tiempoEstimadoTexto(o){{
  const dur=duracionTallerMin(o);
  const cruce=esCruceNoche(o)?' · 🌙 cruza a la jornada siguiente ('+fechaSalidaTallerTexto(o)+')':'';
  if(typeof o.horas_tempario==='number'&&o.horas_tempario>0){{
    const extendido=(dur!==null&&dur>Math.round(o.horas_tempario*60)+5);
    return o.horas_tempario.toFixed(1)+' h (tempario)'+(extendido?' · ⏳ extendido':'')+cruce;
  }}
  if(dur!==null&&dur>0)return(dur/60).toFixed(1)+' h (manual)'+cruce;
  return '--';
}}

/* ─── Regla imperativa (30/07/2026, a pedido de Cristobal): un tecnico NO puede tener
   2 trabajos asignados al mismo tiempo ───
   Antes, si el horario de un vehiculo se cruzaba con otro del mismo tecnico, el sistema
   corria uno en silencio al primer espacio libre (acomodarSinCruce, ahora eliminada).
   Cristobal pidio explicitamente que esto deje de "arreglarse solo": si asignar/editar
   un vehiculo dejaria a un tecnico con 2 trabajos superpuestos, la asignacion se debe
   RECHAZAR por completo (no se guarda nada, se avisa el choque) — hay que elegir otro
   tecnico o ajustar el horario a mano antes de continuar. Esto aplica tanto a asignar/
   editar tecnico y horarios desde Control de Taller o el modal de detalle, como a
   arrastrar una cita de Programacion directo a una celda del grid.

   _calcularSegmentosOrden(o): calcula que dia(s)/horario(s) OCUPARIA una orden con sus
   valores ACTUALES (tecnico, ingreso_taller, salida_taller, ingreso, salida) — sin tocar
   ctrlData ni bloques, solo como consulta. Mismo criterio de siempre: si Salida (fecha)
   es 1+ dia posterior a Ingreso (fecha), un segmento por cada dia del rango (jornada
   completa en los dias intermedios); si no, turno normal de un dia o cruce de una sola
   noche (Salida Taller <= Ingreso Taller). */
function _calcularSegmentosOrden(o){{
  const tecOk=o.tecnico!==null&&o.tecnico!==undefined&&o.tecnico!=='';
  const iniRaw=timeToMin(o.ingreso_taller),finRaw=timeToMin(o.salida_taller);
  if(!tecOk||iniRaw===null||finRaw===null||finRaw===iniRaw)return {{kind:'none',segmentos:[]}};
  const dateStr=isoToDdmmyyyy(o.ingreso);
  const fIni=parseDateISO(o.ingreso), fSal=parseDateISO(o.salida);
  let nDiasSalida=0;
  if(fIni&&fSal)nDiasSalida=Math.round((fSal-fIni)/86400000);

  if(nDiasSalida>=1){{
    const diasSpan=[];
    for(let i=0;i<=nDiasSalida;i++)diasSpan.push(addDiasFecha(dateStr,i));
    const segmentos=diasSpan.map((ds,idx)=>{{
      const end=getEnd(ds);
      const esPrimero=idx===0, esUltimo=idx===diasSpan.length-1;
      let iniMin=esPrimero?Math.max(START,Math.min(iniRaw,end-STEP)):START;
      iniMin=Math.floor(iniMin/STEP)*STEP;
      let finMin=esUltimo?Math.max(iniMin+STEP,Math.min(finRaw,end)):end;
      finMin=Math.ceil(finMin/STEP)*STEP;
      if(finMin<=iniMin)finMin=Math.min(iniMin+STEP,end);
      return {{fecha:ds,ini:iniMin,fin:finMin,esPrimero,esUltimo}};
    }});
    return {{kind:'multi',dateStr,segmentos}};
  }}

  const dateStr2=addDiasFecha(dateStr,1);
  if(finRaw<iniRaw){{
    const end1=getEnd(dateStr);
    let iniMin=Math.max(START,Math.min(iniRaw,end1-STEP));
    iniMin=Math.floor(iniMin/STEP)*STEP;
    const end2=getEnd(dateStr2);
    let finMin2=Math.max(START+STEP,Math.min(finRaw,end2));
    finMin2=Math.ceil(finMin2/STEP)*STEP;
    return {{kind:'overnight',dateStr,dateStr2,
      segmentos:[{{fecha:dateStr,ini:iniMin,fin:end1}},{{fecha:dateStr2,ini:START,fin:finMin2}}]}};
  }}

  const end=getEnd(dateStr);
  let iniMin=Math.max(START,Math.min(iniRaw,end-STEP));
  iniMin=Math.floor(iniMin/STEP)*STEP;
  let finMin=Math.max(iniMin+STEP,Math.min(finRaw,end));
  finMin=Math.ceil(finMin/STEP)*STEP;
  return {{kind:'same',dateStr,segmentos:[{{fecha:dateStr,ini:iniMin,fin:finMin}}]}};
}}

/* Busca si los segmentos de una orden chocan con: (1) otra orden ACTIVA con el mismo
   tecnico (comparando sus propios segmentos calculados en vivo, cubre tambien el caso
   multi-dia), o (2) un bloque suelto ya existente en el grid con ese tecnico ese dia
   (ej. una cita asignada por drag&drop desde Programacion, que no tiene una orden propia
   detras). Se excluyen los bloques/orden de la MISMA orden que se esta evaluando, para
   poder recalcular sin toparse consigo misma. Devuelve null si no hay choque. */
function _buscarChoqueTecnico(o,calc){{
  const segmentos=calc.segmentos;
  if(!segmentos.length)return null;
  const tec=+o.tecnico;
  const prefijo='ct'+o.id;

  for(const o2 of ordenes){{
    if(o2.id===o.id||o2.cerrada)continue;
    if(o2.tecnico===null||o2.tecnico===undefined||o2.tecnico===''||+o2.tecnico!==tec)continue;
    const calc2=_calcularSegmentosOrden(o2);
    for(const s of segmentos){{
      for(const s2 of calc2.segmentos){{
        if(s.fecha===s2.fecha&&s.ini<s2.fin&&s.fin>s2.ini){{
          return {{tipo:'orden',otra:o2,fecha:s.fecha,ini:s2.ini,fin:s2.fin}};
        }}
      }}
    }}
  }}

  for(const s of segmentos){{
    const bls=(ctrlData?.[SUCURSAL]?.bloques?.[s.fecha]||[]);
    for(const b of bls){{
      if(+b.tec!==tec)continue;
      if(b.id===prefijo||String(b.id).startsWith(prefijo+'_'))continue;
      const bIni=parseHH(b.ini),bFin=bIni+(b.dur||60);
      if(s.ini<bFin&&s.fin>bIni){{
        return {{tipo:'bloque',otra:b,fecha:s.fecha,ini:bIni,fin:bFin}};
      }}
    }}
  }}
  return null;
}}

function _mensajeChoqueTecnico(o,conflicto){{
  const tecNom=(o.tecnico!==null&&o.tecnico!==undefined&&tecnicos[+o.tecnico])?tecnicos[+o.tecnico]:'Este tecnico';
  const otraPat=conflicto.otra?.patente||conflicto.otra?.oc||conflicto.otra?.ot||'otro vehiculo';
  return `🚫 ${{tecNom}} ya tiene asignado ${{otraPat}} el ${{conflicto.fecha}} de ${{hhmm(conflicto.ini)}} a ${{hhmm(conflicto.fin)}} — un tecnico no puede tener 2 trabajos al mismo tiempo. Elige otro tecnico o ajusta el horario/fecha antes de continuar.`;
}}

/* ─── Sincroniza el bloque del Planificador (Tecnico x Hora) con el horario de la orden ───
   Antes de crear/actualizar cualquier bloque, valida la regla de arriba — si hay choque,
   NO TOCA NADA (ni limpia ni crea bloques) y devuelve {{ok:false,conflicto}} para que el
   llamador revierta el cambio que lo provoco y avise. Si no hay choque, limpia los
   bloques viejos de esta orden en TODAS las fechas (por si el rango de dias/tipo de turno
   cambio) y crea los nuevos segun _calcularSegmentosOrden, devolviendo {{ok:true}}. */
function upsertBloqueDesdeOrden(o){{
  if(!ctrlData)ctrlData={{}};
  if(!ctrlData[SUCURSAL])ctrlData[SUCURSAL]={{tecnicos,ordenes,bloques:{{}}}};
  if(!ctrlData[SUCURSAL].bloques)ctrlData[SUCURSAL].bloques={{}};

  const prefijo='ct'+o.id;
  const _limpiarTodo=()=>{{
    Object.keys(ctrlData[SUCURSAL].bloques).forEach(ds=>{{
      ctrlData[SUCURSAL].bloques[ds]=ctrlData[SUCURSAL].bloques[ds].filter(b=>!(b.id===prefijo||String(b.id).startsWith(prefijo+'_')));
    }});
  }};

  const calc=_calcularSegmentosOrden(o);
  if(!calc.segmentos.length){{ _limpiarTodo(); return {{ok:true}}; }}

  const conflicto=_buscarChoqueTecnico(o,calc);
  if(conflicto)return {{ok:false,conflicto}};

  _limpiarTodo();
  const baseBloque={{
    tec:+o.tecnico, oc:o.ot||o.patente, patente:o.patente, cliente:o.cliente||'',
    modelo:o.modelo||'', servicio:o.comentarios||o.servicio||'',
    horas_tempario:(typeof o.horas_tempario==='number'?o.horas_tempario:null),
  }};

  if(calc.kind==='multi'){{
    const ultimaFecha=calc.segmentos[calc.segmentos.length-1].fecha;
    calc.segmentos.forEach((s,idx)=>{{
      if(!ctrlData[SUCURSAL].bloques[s.fecha])ctrlData[SUCURSAL].bloques[s.fecha]=[];
      const bid=(s.esPrimero&&s.esUltimo)?prefijo:(prefijo+'_d'+idx);
      let contInfo='';
      if(!s.esPrimero&&!s.esUltimo)contInfo='🗓️ en taller todo el dia (ingreso '+calc.dateStr+' · salida '+ultimaFecha+')';
      else if(s.esPrimero&&!s.esUltimo)contInfo='🗓️ sigue en taller hasta el '+ultimaFecha;
      else if(s.esUltimo&&!s.esPrimero)contInfo='🗓️ en taller desde el '+calc.dateStr;
      ctrlData[SUCURSAL].bloques[s.fecha].push({{...baseBloque,
        id:bid, ini:hhmm(s.ini), dur:s.fin-s.ini, contInfo,
      }});
    }});
    return {{ok:true}};
  }}

  if(calc.kind==='overnight'){{
    if(!ctrlData[SUCURSAL].bloques[calc.dateStr])ctrlData[SUCURSAL].bloques[calc.dateStr]=[];
    if(!ctrlData[SUCURSAL].bloques[calc.dateStr2])ctrlData[SUCURSAL].bloques[calc.dateStr2]=[];
    const[s1,s2]=calc.segmentos;
    ctrlData[SUCURSAL].bloques[calc.dateStr].push({{...baseBloque,
      id:prefijo+'_a', ini:hhmm(s1.ini), dur:s1.fin-s1.ini,
      cont:'sigue', contInfo:'🌙 continua en la jornada siguiente ('+calc.dateStr2+' '+o.salida_taller+')',
    }});
    ctrlData[SUCURSAL].bloques[calc.dateStr2].push({{...baseBloque,
      id:prefijo+'_b', ini:hhmm(s2.ini), dur:s2.fin-s2.ini,
      cont:'viene', contInfo:'🌙 viene de la jornada anterior ('+calc.dateStr+' '+o.ingreso_taller+')',
    }});
    return {{ok:true}};
  }}

  // same
  if(!ctrlData[SUCURSAL].bloques[calc.dateStr])ctrlData[SUCURSAL].bloques[calc.dateStr]=[];
  const s=calc.segmentos[0];
  ctrlData[SUCURSAL].bloques[calc.dateStr].push({{...baseBloque,
    id:prefijo, ini:hhmm(s.ini), dur:s.fin-s.ini,
  }});
  return {{ok:true}};
}}

/* Aplica un cambio de tecnico/horario/fecha (los 5 campos que alimentan el Planificador
   de Tecnicos) validando PRIMERO la regla de arriba. Si hay choque, revierte TODOS los
   campos relacionados a su valor anterior (incluyendo el auto-llenado de Salida Taller
   que pudo haberse disparado) y devuelve {{ok:false,mensaje}} SIN tocar bloques ni guardar
   — el llamador debe mostrar el mensaje y volver a pintar la UI para que los campos
   reflejen el valor original. Si no hay choque, el bloque ya queda actualizado dentro de
   upsertBloqueDesdeOrden y devuelve {{ok:true}}. */
function _aplicarCambioAgenda(o,field,val){{
  const _snap={{tecnico:o.tecnico,ingreso_taller:o.ingreso_taller,salida_taller:o.salida_taller,ingreso:o.ingreso,salida:o.salida}};
  o[field]=val;
  const tecOk=o.tecnico!==null&&o.tecnico!==undefined&&o.tecnico!=='';
  if((field==='tecnico'||field==='ingreso_taller')&&tecOk&&o.ingreso_taller&&!o.salida_taller){{
    o.salida_taller=calcularSalidaTaller(o);
  }}
  const res=upsertBloqueDesdeOrden(o);
  if(res&&res.ok===false){{
    const mensaje=_mensajeChoqueTecnico(o,res.conflicto);
    Object.assign(o,_snap);
    return {{ok:false,mensaje}};
  }}
  return {{ok:true}};
}}

/* ─── Control de Taller — importa citas ingresadas de la agenda de hoy ─── */
function autoImportarCitas(){{
  const hoyStr=formatDate(planDates[0]);
  const citas=getCitas(hoyStr).filter(c=>c.ingresado);
  let agregadas=0, actualizadas=0;
  citas.forEach(c=>{{
    const pat=normPat(c.patente);
    if(!pat)return;
    // Patentes de prueba tipo SP0000 (usadas para probar la Agenda, sin vehiculo real
    // detras) nunca se auto-importan al tablero — ya se excluian de "Vehiculos en
    // Taller" (esPatenteExcluida) pero no de esta funcion, asi que si alguien las
    // eliminaba a mano del JPCB volvian a aparecer solas en la proxima carga de la
    // pagina (la cita seguia "ingresada" en la Agenda). 22/07/2026, a pedido de
    // Cristobal ("esas patentes las eliminé del tablero y volvieron a aparecer").
    if(esPatenteExcluida(pat))return;
    const oc=String(c.oc||'').trim();
    // Si el usuario ya elimino a mano esta combinacion exacta de patente+OT del
    // tablero (eliminarOrdenCT), no se vuelve a crear aunque la cita siga
    // "ingresada" en la Agenda — ver _ordenFueEliminada()/_marcarOrdenEliminada().
    if(_ordenFueEliminada(pat,oc))return;
    // Emparejamiento por patente + Folio OT (no solo patente): una misma patente puede
    // tener 2+ citas activas a la vez con OTs distintas (ej. una de Mantencion y otra de
    // Garantia/Recall/Diagnostico) — cada una debe quedar como su propia orden en Control
    // de Taller/JPCB, no fusionarse en una sola. Si la cita no trae OT (caso raro), se
    // mantiene el comportamiento anterior (match solo por patente) para no duplicar altas
    // manuales sin folio.
    const existente=ordenes.find(o=>normPat(o.patente)===pat &&
      (oc?String(o.ot||'').trim()===oc:!String(o.ot||'').trim()));
    if(existente){{
      // Ya existe (creada en una importacion anterior, o por alta manual) — se
      // SINCRONIZA con lo que hoy trae la Agenda para los campos que vienen de ahi
      // (modelo, mantencion, km, asesor, cliente, horario de ingreso): si en la Agenda
      // se corrigio un dato (ej. el modelo estaba mal escrito), el cambio ahora se
      // refleja aca tambien, no solo se rellena cuando esta vacio. Los campos que son
      // 100% de gestion manual del taller (tecnico, etapa, detencion, comentario
      // adicional, numero de caso, N de pedido, ETA, auto de reemplazo, checkboxes
      // ESP/TRA/LAV/PATIO) nunca se tocan aca — siguen siendo responsabilidad exclusiva
      // de quien los edita en Control de Taller.
      let cambio=false;
      if(c.mantencion&&existente.mantencion!==c.mantencion){{existente.mantencion=c.mantencion;cambio=true;}}
      if(c.modelo&&existente.modelo!==c.modelo){{existente.modelo=c.modelo;cambio=true;}}
      if(c.km&&existente.km!==c.km){{existente.km=c.km;cambio=true;}}
      if(c.asesor&&existente.asesor!==c.asesor){{existente.asesor=c.asesor;cambio=true;}}
      if(!existente.ot&&c.oc){{existente.ot=c.oc;cambio=true;}}
      const _nomC=c.nombre||c.cliente||'';
      if(_nomC&&existente.cliente!==_nomC){{existente.cliente=_nomC;cambio=true;}}
      if(c.horario&&existente.hora_rec!==c.horario){{existente.hora_rec=c.horario;cambio=true;}}
      if(c.vin&&existente.vin!==c.vin){{existente.vin=c.vin;cambio=true;}}
      if(typeof c.horas_tempario==='number'&&existente.horas_tempario!==c.horas_tempario){{existente.horas_tempario=c.horas_tempario;cambio=true;}}
      // Comentarios/Motivo (Recall/Mantencion/Diagnostico/etc.) tambien vienen del texto
      // de Servicio de la Agenda — se resincronizan solo cuando ESE texto cambia desde
      // la ultima vez que se importo (guardado en _servicio_src), para no pisar una
      // reclasificacion manual del Motivo que alguien haya hecho a mano en Control de
      // Taller sin que el texto de origen en la Agenda haya cambiado.
      const _svcNow=c.servicio||c.mantencion||'';
      if(_svcNow&&existente._servicio_src!==_svcNow){{
        existente.comentarios=_svcNow;
        existente.servicio=_svcNow;
        existente.tipo=detectTipo(c);
        existente._servicio_src=_svcNow;
        cambio=true;
      }}
      if(cambio)actualizadas++;
      return;
    }}
    const _svcIni=c.servicio||c.mantencion||'';
    const id='a'+Date.now()+Math.random().toString(36).slice(2,6);
    ordenes.push({{
      id, patente:pat, cliente:c.nombre||c.cliente||'', modelo:c.modelo||'',
      ot:c.oc||'', km:c.km||'', asesor:c.asesor||'', vin:c.vin||'',
      ingreso:ddmmyyyyToIso(c.fecha), salida:'',
      tecnico:null,
      comentarios:_svcIni, tipo:detectTipo(c), _servicio_src:_svcIni,
      etapa:ETAPAS[0].id, stop:null,
      comentario2:'', numero_caso:'', n_pedido:'', eta:'', auto_reemplazo:'',
      servicio:_svcIni, hora_rec:c.horario||'',
      mantencion:c.mantencion||'', hora_compromiso:'', duracion_min:60,
      ingreso_taller:'', salida_taller:'',
      horas_tempario:(typeof c.horas_tempario==='number'?c.horas_tempario:null),
      reloj_inicio_ts:null, reloj_inicio_txt:'', reloj_fin_ts:null, reloj_fin_txt:'',
      cerrada:false, fecha_cierre:'', estado_campana:'',
      // 23/07/2026, a pedido de Cristobal: toda cita nueva aterriza "sin confirmar" en
      // la columna "Citas <fecha>" del JPCB — recien pasa a Recepcion (y se vuelve
      // visible en Control de Taller/Vehiculos en Taller/grid Tecnico x Hora) cuando
      // alguien confirma "Asiste". Ver citaConfirmada()/marcarAsisteCita().
      estadoCita:'pendiente', fecha_reagenda:'',
    }});
    agregadas++;
  }});
  // Citas FINALIZADAS (icono persona 🧍 en la agenda = entregado al cliente): el estado
  // de la cita es el indicador final — se le marca Salida (dato informativo) y ademas
  // se cierra sola (pasa al Historial de Taller), asi no se acumulan en Control de
  // Taller/Vehiculos en Taller vehiculos que la Agenda ya dice que se fueron.
  let salidas=0;
  getCitas(hoyStr).filter(c=>c.estado==='finalizado').forEach(c=>{{
    const pat=normPat(c.patente);
    if(!pat)return;
    const oc=String(c.oc||'').trim();
    // Igual que en la importacion: si la patente tiene varias ordenes activas (varias
    // OTs), solo se cierra la que corresponde a ESTA cita (misma OT) — no todas las de
    // esa patente, para no cerrar por error la de Mantencion cuando solo la de Garantia
    // ya fue entregada (o viceversa).
    const o=ordenes.find(x=>normPat(x.patente)===pat&&!x.cerrada &&
      (oc?String(x.ot||'').trim()===oc:true));
    if(o){{
      if(!(o.salida||'').trim())o.salida=ddmmyyyyToIso(hoyStr);
      _cerrarOrdenInterno(o);
      salidas++;
    }}
  }});
  return agregadas+actualizadas+salidas;
}}

/* ─── Control de Taller — edicion y render ─── */
function editFieldCT(id,field,val){{
  const o=byId(id);if(!o)return;
  if(field==='etapa'&&_avanceBloqueadoPorVCU(o,val)){{
    alert(`🚫 No se puede avanzar de etapa — falta completar el VCU (Hoja Multipuntos Ford) de ${{o.patente}}.`);
    renderControlTaller();renderVehiculosTaller();renderJPCB();
    return;
  }}
  // Tecnico/Ingreso Taller/Salida Taller/Ingreso (fecha)/Salida (fecha) van por
  // _aplicarCambioAgenda(), que valida la regla imperativa (un tecnico no puede tener
  // 2 trabajos a la vez) ANTES de aplicar el cambio — si hay choque, revierte todo y
  // avisa, sin guardar. El resto de los campos sigue el flujo normal.
  if(['tecnico','ingreso_taller','salida_taller','ingreso','salida'].includes(field)){{
    const res=_aplicarCambioAgenda(o,field,val);
    if(!res.ok){{
      alert(res.mensaje);
      renderControlTaller();renderVehiculosTaller();renderJPCB();
      return;
    }}
    if(currentView==='plan')renderPlanView();
    renderControlTaller();renderVehiculosTaller();renderJPCB();
    saveCtrl();
    return;
  }}
  o[field]=val;
  if(field==='etapa')marcarCambioEtapa(o);
  renderControlTaller();
  renderVehiculosTaller();
  renderJPCB();
  saveCtrl();
}}
/* Quita del grid Tecnico x Hora TODOS los bloques de una orden (el bloque unico normal
   'ct'+id, el par overnight 'ct'+id+'_a'/'_b', o los N bloques de un rango multi-dia
   'ct'+id+'_d0'/'_d1'/...) — recorre TODAS las fechas con bloques guardados (no solo
   Ingreso/Ingreso+1) para no dejar huerfanos si la orden abarcaba varios dias. Ver
   upsertBloqueDesdeOrden(). */
function _quitarBloquesOrden(o){{
  if(!ctrlData?.[SUCURSAL]?.bloques)return;
  const prefijo='ct'+o.id;
  Object.keys(ctrlData[SUCURSAL].bloques).forEach(ds=>{{
    ctrlData[SUCURSAL].bloques[ds]=ctrlData[SUCURSAL].bloques[ds].filter(b=>!(b.id===prefijo||String(b.id).startsWith(prefijo+'_')));
  }});
}}
function eliminarOrdenCT(id){{
  const o=byId(id);if(!o)return;
  if(!confirm(`¿Quitar ${{o.patente}} de Control de Taller? Tambien desaparecera del tablero JPCB.`))return;
  _marcarOrdenEliminada(o.patente,o.ot);
  ordenes=ordenes.filter(x=>x.id!==id);
  _quitarBloquesOrden(o);
  renderControlTaller();renderVehiculosTaller();renderJPCB();
  if(currentView==='plan')renderPlanView();
  saveCtrl();
  toast('Patente eliminada de Control de Taller');
}}
/* ─── Cierre de cita — archiva la orden al Historial de Taller ───
   A diferencia de eliminarOrdenCT() (que borra el registro), esto solo la saca de
   los tableros activos (JPCB, Control de Taller, Vehiculos en Taller, Tecnico x
   Hora) y la deja disponible en el Historial. La columna Salida queda como un dato
   informativo aparte — ya no es lo que decide si la orden sigue apareciendo.
   `_cerrarOrdenInterno` hace el trabajo real (sin confirm/toast/re-render, para
   poder reutilizarlo tanto en el cierre manual como en el auto-cierre por Agenda). */
function _cerrarOrdenInterno(o){{
  if(o.cerrada)return;
  // Si el reloj de taller quedo corriendo (ej. se finaliza/cierra la cita directo desde
  // "En Proceso", sin pasar formalmente por Lavado/Entrega), se detiene aca — asi el
  // Historial de Taller nunca muestra un reloj "en curso" para una orden ya cerrada.
  if(o.reloj_inicio_ts&&!o.reloj_fin_ts){{
    o.reloj_fin_ts=Date.now();
    o.reloj_fin_txt=nowStrCorto();
  }}
  o.cerrada=true;
  o.fecha_cierre=isoToday();
  _quitarBloquesOrden(o);
}}
function cerrarCita(id){{
  const o=byId(id);if(!o)return;
  if(!confirm(`¿Cerrar la cita de ${{o.patente}}? Desaparecera de JPCB, Control de Taller, Vehiculos en Taller y Tecnico x Hora, y quedara en el Historial de Taller.`))return;
  _cerrarOrdenInterno(o);
  renderControlTaller();renderVehiculosTaller();renderJPCB();
  if(typeof renderHistorialTaller==='function')renderHistorialTaller();
  if(currentView==='plan')renderPlanView();
  saveCtrl();
  toast(`🔒 Cita de ${{o.patente}} cerrada — pasó al Historial de Taller`);
}}
function reabrirCita(id){{
  const o=byId(id);if(!o)return;
  if(!confirm(`¿Reabrir la cita de ${{o.patente}}? Volvera a aparecer en JPCB, Control de Taller y Vehiculos en Taller.`))return;
  o.cerrada=false;
  o.fecha_cierre='';
  o.finalizado_usuario='';
  o.finalizado_fecha='';
  renderControlTaller();renderVehiculosTaller();renderJPCB();
  if(typeof renderHistorialTaller==='function')renderHistorialTaller();
  if(currentView==='plan')renderPlanView();
  saveCtrl();
  toast(`↩️ Cita de ${{o.patente}} reabierta`);
}}
/* Boton "✅ Finalizado" en cada tarjeta del JPCB (24/07/2026, a pedido de Cristobal) —
   mismo permiso limitado que Asiste/No Asiste/Reagenda (PUEDE_CONFIRMAR_CITAS), sin
   requerir el permiso completo de edicion del Planificador. Reutiliza exactamente la
   misma logica de cierre que el boton "🔒 Cerrar" (_cerrarOrdenInterno): la orden pasa
   al Historial de Taller y desaparece de JPCB/Control de Taller/Vehiculos en
   Taller/Tecnico x Hora. Se deja registrado ademas quien y cuando finalizo, para
   mostrarlo en el acumulado de "Finalizados de esta semana" (ver renderFinalizadosSemana). */
function marcarFinalizadoCita(id){{
  const o=byId(id);if(!o||o.cerrada)return;
  if(!confirm(`¿Marcar la cita de ${{o.patente}} como Finalizada? Pasara al Historial de Taller.`))return;
  _cerrarOrdenInterno(o);
  o.finalizado_usuario=USUARIO;
  o.finalizado_fecha=nowStrCorto();
  renderControlTaller();renderVehiculosTaller();renderJPCB();
  if(typeof renderHistorialTaller==='function')renderHistorialTaller();
  if(currentView==='plan')renderPlanView();
  saveCtrl();
  toast(`✅ ${{o.patente}} finalizado — paso al Historial de Taller`);
}}
/* Semana (Lunes-Domingo) de una fecha ISO "AAAA-MM-DD" — usada para que el acumulado
   de "Finalizados de esta semana" se limpie solo al llegar la semana siguiente, sin
   necesidad de ningun proceso de limpieza aparte: la orden sigue con cerrada=true para
   siempre (nunca vuelve a aparecer en las columnas normales del JPCB ni en Control de
   Taller/Vehiculos en Taller) y sigue siempre visible en el Historial de Taller — solo
   deja de calzar con la semana actual en este panel puntual. */
function _lunesSemana(fechaIso){{
  if(!fechaIso)return'';
  const partes=fechaIso.split('-').map(Number);
  const y=partes[0],m=partes[1],d=partes[2];
  if(!y||!m||!d)return'';
  const dt=new Date(y,m-1,d);
  const dow=dt.getDay();
  const diff=(dow===0?-6:1-dow);
  dt.setDate(dt.getDate()+diff);
  return dt.getFullYear()+'-'+String(dt.getMonth()+1).padStart(2,'0')+'-'+String(dt.getDate()).padStart(2,'0');
}}
function esFinalizadoEstaSemana(o){{
  if(!o.cerrada||!o.fecha_cierre)return false;
  return _lunesSemana(o.fecha_cierre)===_lunesSemana(isoToday());
}}
/* Tarjeta simplificada y de solo informacion para el acumulado de "Finalizados de esta
   semana" — no es arrastrable ni tiene los botones normales del JPCB, solo un boton de
   "↩️ Reabrir" (mismo criterio de permiso que en el Historial de Taller: PUEDE_EDITAR). */
function cardFinalizadaHTML(o){{
  const ti=tipoInfo(o);
  return `<div class="card" data-id="${{o.id}}" style="width:190px;border-left-color:${{ti.border}};background:${{ti.color}}">
    <b>${{o.patente}}</b> <span class="cmeta">${{o.cliente||''}}</span>
    <div class="cinfo">${{o.modelo||''}}${{o.modelo&&(o.servicio||o.mantencion)?' · ':''}}${{o.servicio||o.mantencion||ti.label}}</div>
    <div class="cmeta">✅ ${{o.fecha_cierre||'--'}}${{o.finalizado_usuario?' · '+esc(o.finalizado_usuario):''}}</div>
    ${{PUEDE_EDITAR?`<div class="cacts"><button onclick="event.stopPropagation();reabrirCita('${{o.id}}')">↩️ Reabrir</button></div>`:''}}
  </div>`;
}}
function renderFinalizadosSemana(){{
  const el=document.getElementById('finBoard');
  if(!el)return;
  const cards=ordenes.filter(o=>esFinalizadoEstaSemana(o)&&ordenArea(o)===currentArea);
  const orden=[...cards].sort((a,b)=>(b.fecha_cierre||'').localeCompare(a.fecha_cierre||''));
  if(!orden.length){{
    el.innerHTML='<div style="color:#888;font-size:12px;padding:4px">Sin citas finalizadas esta semana.</div>';
    return;
  }}
  el.innerHTML=orden.map(cardFinalizadaHTML).join('');
}}
function agregarPatenteManual(){{
  const pat=normPat(prompt('Patente del vehiculo:')||'');
  if(!pat)return;
  if(ordenes.some(o=>normPat(o.patente)===pat)){{alert('Esa patente ya esta en Control de Taller.');return;}}
  const id='m'+Date.now()+Math.random().toString(36).slice(2,6);
  ordenes.push({{
    id, patente:pat, cliente:'', modelo:'', ot:'', km:'', asesor:'',
    ingreso:isoToday(), salida:'', tecnico:null,
    comentarios:'', tipo:'ot', etapa:ETAPAS[0].id, stop:null,
    comentario2:'', numero_caso:'', n_pedido:'', eta:'', auto_reemplazo:'', servicio:'', hora_rec:'',
    mantencion:'', hora_compromiso:'', duracion_min:60, ingreso_taller:'', salida_taller:'', horas_tempario:null,
    reloj_inicio_ts:null, reloj_inicio_txt:'', reloj_fin_ts:null, reloj_fin_txt:'',
    cerrada:false, fecha_cierre:'', estado_campana:'',
    // Alta manual = el vehiculo ya esta fisicamente en el taller, no pasa por la
    // columna "Citas" (esa es solo para lo que trae la Agenda sin confirmar).
    estadoCita:'asiste', fecha_reagenda:'',
  }});
  renderControlTaller();renderVehiculosTaller();renderJPCB();saveCtrl();
  toast(`✅ ${{pat}} agregado a Control de Taller`);
}}

const CT_COLS=28;
function ctTableHead(){{
  return `<thead><tr>
    <th>Estado</th><th>Patente</th><th>Modelo</th><th>Mantencion</th><th>OT</th><th>KM</th><th>Asesor</th>
    <th title="Horario que viene de la Agenda — informativo, no alimenta el Planificador de Tecnicos">Horario Ingreso</th>
    <th title="Compromiso de entrega al cliente — informativo, no alimenta el Planificador de Tecnicos">Horario Entrega</th>
    <th>Tiempo Estimado</th>
    <th class="ct-th-taller" title="Ingreso real al taller — alimenta el Planificador de Tecnicos y el Tiempo Estimado">⚙️ Ingreso Taller</th>
    <th class="ct-th-taller" title="Salida real/estimada del taller — alimenta el Planificador de Tecnicos y el Tiempo Estimado">⚙️ Salida Taller</th>
    <th>Ingreso</th><th>Salida</th><th>Dias en taller</th><th>Tecnico</th>
    <th>Comentarios</th><th>Motivo</th><th>Etapa (JPCB)</th><th>Detencion</th><th>Estado Campaña</th>
    <th>Comentario adicional</th><th>Numero de caso</th><th>N° de pedido</th>
    <th>ETA</th><th>Dias faltantes ETA</th><th>Auto de reemplazo</th><th>Acciones</th>
  </tr></thead>`;
}}
function ctRowHTML(o){{
  const dis=PUEDE_EDITAR?'':'disabled';
  const opcTec=`<option value="">--</option>`+tecnicos.map((t,i)=>`<option value="${{i}}" ${{o.tecnico===i?'selected':''}}>${{t}}</option>`).join('');
  const opcMotivo=Object.entries(TIPOS).map(([k,v])=>`<option value="${{k}}" ${{o.tipo===k?'selected':''}}>${{v.label}}</option>`).join('');
  const opcEtapa=ETAPAS.map(e=>`<option value="${{e.id}}" ${{o.etapa===e.id?'selected':''}}>${{e.t}}</option>`).join('');
  const opcStop=`<option value="">— Sin detencion —</option>`+STOPS.map(s=>`<option value="${{s.id}}" ${{o.stop===s.id?'selected':''}}>${{s.t}}</option>`).join('');
  const opcEstadoCamp=`<option value="">— Sin estado —</option>`+ESTADOS_CAMPANA.map(e=>`<option value="${{e.id}}" ${{o.estado_campana===e.id?'selected':''}}>${{e.t}}</option>`).join('');
  const _asesorActual=(o.asesor||'').trim();
  const _listaAsesores=_asesorActual&&!asesoresSucursal.includes(_asesorActual)?[...asesoresSucursal,_asesorActual]:asesoresSucursal;
  const opcAsesor=`<option value="">--</option>`+_listaAsesores.map(a=>`<option value="${{esc(a)}}" ${{_asesorActual===a?'selected':''}}>${{esc(a)}}</option>`).join('');
  const dias=calcDiasEnTaller(o);
  const diasEta=calcDiasFaltantesEta(o);
  const _et=etapaInfo(o.etapa);
  const _noShow=esNoAsiste(o.ot,o.patente);
  return `<tr style="background:${{_noShow?'#f2f2f2':_et.bg}}">
    <td style="text-align:center;font-size:15px" title="${{esc(ctEstadoTitle(o))}}">${{ctEstadoIcon(o)}}${{_noShow?'<br><span class="cita-noasiste" style="font-size:8px">🚫 No asiste</span>':''}}</td>
    <td class="ct-pat" style="background:${{_noShow?'#f2f2f2':_et.bg}};border-left:4px solid ${{_noShow?'#b33':_et.color}}${{_noShow?';text-decoration:line-through':''}}">${{esc(o.patente)}}${{esFord(o)?`<br><span class="vcu-chip ${{vcuCompleto(o)?'ok':'pend'}}" onclick="abrirVCU('${{o.id}}')" title="Hoja Multipuntos Ford (VCU)">📋 VCU</span>`:''}}</td>
    <td><input type="text" class="ct-wide" value="${{esc(o.modelo)}}" ${{dis}} onchange="editFieldCT('${{o.id}}','modelo',this.value)"></td>
    <td><input type="text" class="ct-wide" value="${{esc(o.mantencion)}}" ${{dis}} placeholder="ej: 10.000 KMS" onchange="editFieldCT('${{o.id}}','mantencion',this.value)"></td>
    <td><input type="text" value="${{esc(o.ot)}}" ${{dis}} onchange="editFieldCT('${{o.id}}','ot',this.value)"></td>
    <td><input type="text" value="${{esc(o.km)}}" ${{dis}} style="width:60px" onchange="editFieldCT('${{o.id}}','km',this.value)"></td>
    <td><select ${{dis}} onchange="editFieldCT('${{o.id}}','asesor',this.value)" title="Asesores de la Agenda para esta sucursal">${{opcAsesor}}</select></td>
    <td><input type="time" value="${{esc(o.hora_rec)}}" ${{dis}} onchange="editFieldCT('${{o.id}}','hora_rec',this.value)" title="Viene de la Agenda — informativo, no alimenta el Planificador de Tecnicos"></td>
    <td><input type="time" value="${{esc(o.hora_compromiso)}}" ${{dis}} onchange="editFieldCT('${{o.id}}','hora_compromiso',this.value)" title="Compromiso de entrega al cliente — informativo, no alimenta el Planificador de Tecnicos"></td>
    <td class="ct-dias" title="Calculado desde Ingreso Taller / Salida Taller (y el tempario de mantencion cuando aplica); se puede extender editando Salida Taller">${{tiempoEstimadoTexto(o)}}</td>
    <td class="ct-td-taller"><input type="time" class="ct-input-taller" value="${{esc(o.ingreso_taller)}}" ${{dis}} onchange="editFieldCT('${{o.id}}','ingreso_taller',this.value)" title="Alimenta el Planificador de Tecnicos y el Tiempo Estimado"></td>
    <td class="ct-td-taller"><input type="time" class="ct-input-taller" value="${{esc(o.salida_taller)}}" ${{dis}} onchange="editFieldCT('${{o.id}}','salida_taller',this.value)" title="Alimenta el Planificador de Tecnicos y el Tiempo Estimado"></td>
    <td><input type="date" value="${{esc(o.ingreso)}}" ${{dis}} onchange="editFieldCT('${{o.id}}','ingreso',this.value)"></td>
    <td><input type="date" value="${{esc(o.salida)}}" ${{dis}} onchange="editFieldCT('${{o.id}}','salida',this.value)"></td>
    <td class="ct-dias">${{dias}}</td>
    <td><select ${{dis}} onchange="editFieldCT('${{o.id}}','tecnico',this.value===''?null:+this.value)">${{opcTec}}</select></td>
    <td><input type="text" class="ct-wide" value="${{esc(o.comentarios)}}" ${{dis}} onchange="editFieldCT('${{o.id}}','comentarios',this.value)"></td>
    <td><select ${{dis}} onchange="editFieldCT('${{o.id}}','tipo',this.value)">${{opcMotivo}}</select></td>
    <td><select class="et-select" style="background:${{_et.color}};color:#fff;font-weight:700;border-color:${{_et.color}}" ${{dis}} onchange="editFieldCT('${{o.id}}','etapa',this.value)">${{opcEtapa}}</select></td>
    <td><select ${{dis}} onchange="editFieldCT('${{o.id}}','stop',this.value===''?null:this.value)">${{opcStop}}</select></td>
    <td><select ${{dis}} onchange="editFieldCT('${{o.id}}','estado_campana',this.value)" title="Al elegir Quiebre Stock/Cliente desiste/Falla servidor la orden se oculta del tablero JPCB">${{opcEstadoCamp}}</select></td>
    <td><input type="text" class="ct-wide" value="${{esc(o.comentario2)}}" ${{dis}} onchange="editFieldCT('${{o.id}}','comentario2',this.value)"></td>
    <td><input type="text" value="${{esc(o.numero_caso)}}" ${{dis}} onchange="editFieldCT('${{o.id}}','numero_caso',this.value)"></td>
    <td><input type="text" value="${{esc(o.n_pedido)}}" ${{dis}} onchange="editFieldCT('${{o.id}}','n_pedido',this.value)"></td>
    <td><input type="date" value="${{esc(o.eta)}}" ${{dis}} onchange="editFieldCT('${{o.id}}','eta',this.value)"></td>
    <td class="ct-dias">${{diasEta}}</td>
    <td><input type="text" value="${{esc(o.auto_reemplazo)}}" ${{dis}} onchange="editFieldCT('${{o.id}}','auto_reemplazo',this.value)"></td>
    <td style="white-space:nowrap">${{PUEDE_EDITAR?`<button class="mbtn" style="padding:2px 6px;font-size:11px" title="Cerrar cita — pasa al Historial de Taller" onclick="cerrarCita('${{o.id}}')">🔒 Cerrar</button> <button class="mbtn" style="padding:2px 6px;font-size:11px" title="${{_noShow?'Reactivar':'Marcar cliente no asiste'}}" onclick="toggleNoAsiste('${{(o.ot||'').replace(/'/g,"\\'")}}','${{(o.patente||'').replace(/'/g,"\\'")}}')">${{_noShow?'↩️':'🚫'}} No asiste</button> <button class="ct-del" onclick="eliminarOrdenCT('${{o.id}}')">🗑</button>`:''}}</td>
  </tr>`;
}}

function etapaLegendHTML(){{
  return ETAPAS.map(e=>`<div class="it"><span class="sw" style="background:${{e.bg}};border-color:${{e.color}}"></span>${{e.t}}</div>`).join('');
}}
/* Marca del vehiculo (primera palabra de Modelo, ej. "FORD F150" -> "FORD") — Control
   de Taller/Vehiculos en Taller no guardan la Marca como campo propio, solo Modelo. */
function marcaDeOrden(o){{return (o.modelo||'').trim().split(/\s+/)[0]||'Sin marca';}}

/* Agrupa las filas por Marca con una barra separadora horizontal entre grupos —
   14/07/2026, a pedido de Cristobal ("alguna barra horizontal que separe por Marcas").
   No reordena dentro de cada marca (respeta el orden ya definido por el caller). */
function ctFilasConMarca(lista){{
  let html='', marcaAnt=null;
  const conteos={{}};
  lista.forEach(o=>{{const m=marcaDeOrden(o);conteos[m]=(conteos[m]||0)+1;}});
  lista.forEach(o=>{{
    const marca=marcaDeOrden(o);
    if(marca!==marcaAnt){{
      html+=`<tr class="ct-marca-sep"><td colspan="${{CT_COLS}}">🚗 ${{esc(marca)}} (${{conteos[marca]}})</td></tr>`;
      marcaAnt=marca;
    }}
    html+=ctRowHTML(o);
  }});
  return html;
}}

function renderControlTaller(){{
  const g=document.getElementById('ctGrid');
  const leg=document.getElementById('ctLegend');
  if(leg)leg.innerHTML='<b>Color de fila = Etapa:</b>'+etapaLegendHTML();
  const head=ctTableHead();
  // 23/07/2026, a pedido de Cristobal: solo cuentan los vehiculos con asistencia
  // confirmada (citaConfirmada) — las citas todavia sin confirmar viven en la columna
  // "Citas <fecha>" del JPCB, y las marcadas No Asiste/Reagenda quedan fuera de aca
  // tambien (antes seguian viendose con un badge; ahora se sacan por completo).
  const activas=ordenes.filter(o=>!o.cerrada&&citaConfirmada(o)&&ordenArea(o)===currentArea);
  if(!activas.length){{
    g.innerHTML=head+`<tbody><tr><td colspan="${{CT_COLS}}" style="padding:16px;text-align:center;color:#888">Sin vehiculos en Control de Taller. Se agregan solos al confirmar "Asiste" en la columna "Citas" del JPCB, o usa "➕ Agregar patente manual".</td></tr></tbody>`;
    return;
  }}
  const orden=[...activas].sort((a,b)=>{{
    const ma=marcaDeOrden(a),mb=marcaDeOrden(b);
    if(ma!==mb)return ma.localeCompare(mb);
    return normPat(a.patente).localeCompare(normPat(b.patente));
  }});
  g.innerHTML=head+'<tbody>'+ctFilasConMarca(orden)+'</tbody>';
}}

/* ─── Vehiculos en Taller — misma tabla, filtrada a los que no estan cerrados.
   La columna Salida es solo informativa (fecha real de retiro) — ya no saca al
   vehiculo de este listado; para eso esta el boton "🔒 Cerrar" (Cierre de cita). */
function renderVehiculosTaller(){{
  const g=document.getElementById('vtGrid');
  if(!g)return;
  const leg=document.getElementById('vtLegend');
  if(leg)leg.innerHTML='<b>Color de fila = Etapa:</b>'+etapaLegendHTML();
  const head=ctTableHead();
  const filtradas=ordenes.filter(o=>!o.cerrada&&citaConfirmada(o)&&!esPatenteExcluida(o.patente)&&isWithinLastDays(o.ingreso,60)&&ordenArea(o)===currentArea);
  if(!filtradas.length){{
    g.innerHTML=head+`<tbody><tr><td colspan="${{CT_COLS}}" style="padding:16px;text-align:center;color:#888">Sin vehiculos detenidos en taller dentro de los ultimos 60 dias.</td></tr></tbody>`;
    return;
  }}
  const orden=[...filtradas].sort((a,b)=>{{
    const ma=marcaDeOrden(a),mb=marcaDeOrden(b);
    if(ma!==mb)return ma.localeCompare(mb);
    return normPat(a.patente).localeCompare(normPat(b.patente));
  }});
  g.innerHTML=head+'<tbody>'+ctFilasConMarca(orden)+'</tbody>';
}}

/* ─── Historial de Taller — ordenes cerradas con "🔒 Cerrar" (solo lectura + reabrir) ─── */
const HIST_COLS=14;
function histTableHead(){{
  return `<thead><tr>
    <th>Patente</th><th>Modelo</th><th>Mantencion</th><th>OT</th><th>Asesor</th><th>Tecnico</th>
    <th>Ingreso</th><th>Salida</th><th>Fecha Cierre</th><th title="Tiempo real desde que entro a En Proceso hasta que paso a Lavado/Entrega">⏱ Reloj Taller</th><th>Etapa (JPCB)</th>
    <th>Comentarios</th><th>Motivo</th><th>Acciones</th>
  </tr></thead>`;
}}
function histRowHTML(o){{
  const tecNom=o.tecnico!==null&&o.tecnico!==undefined&&tecnicos[o.tecnico]?tecnicos[o.tecnico]:'--';
  const etNom=ETAPAS.find(e=>e.id===o.etapa)?.t||o.etapa||'--';
  const motNom=tipoInfo(o).label;
  return `<tr class="ct-salida">
    <td class="ct-pat">${{esc(o.patente)}}</td>
    <td>${{esc(o.modelo)||'--'}}</td>
    <td>${{esc(o.mantencion)||'--'}}</td>
    <td>${{esc(o.ot)||'--'}}</td>
    <td>${{esc(o.asesor)||'--'}}</td>
    <td>${{esc(tecNom)}}</td>
    <td>${{esc(o.ingreso)||'--'}}</td>
    <td>${{esc(o.salida)||'--'}}</td>
    <td>${{esc(o.fecha_cierre)||'--'}}</td>
    <td>${{relojTallerTexto(o)}}</td>
    <td>${{esc(etNom)}}</td>
    <td>${{esc(o.comentarios)||'--'}}</td>
    <td>${{esc(motNom)}}</td>
    <td>${{PUEDE_EDITAR?`<button class="mbtn" style="padding:2px 6px;font-size:11px" onclick="reabrirCita('${{o.id}}')">↩️ Reabrir</button>`:''}}</td>
  </tr>`;
}}
function renderHistorialTaller(){{
  const g=document.getElementById('histGrid');
  if(!g)return;
  const head=histTableHead();
  const cerradas=ordenes.filter(o=>o.cerrada&&ordenArea(o)===currentArea);
  if(!cerradas.length){{
    g.innerHTML=head+`<tbody><tr><td colspan="${{HIST_COLS}}" style="padding:16px;text-align:center;color:#888">Sin ordenes cerradas todavia. Usa "🔒 Cerrar" en Control de Taller o Vehiculos en Taller.</td></tr></tbody>`;
    return;
  }}
  const orden=[...cerradas].sort((a,b)=>(b.fecha_cierre||'').localeCompare(a.fecha_cierre||''));
  g.innerHTML=head+'<tbody>'+orden.map(histRowHTML).join('')+'</tbody>';
}}

/* ─── Produccion Tecnicos (horas facturadas, BDFlexline via consolidador) ───
   Datos inyectados en _PROD_INIT/prodData (mismo mecanismo que agenda/ctrl):
   {{fecha_actualizacion, resumen:[{{mecanico, sucursal_mecanico, mes:'YYYY-MM',
   total_horas, n_ot}}], detalle_producto:[{{mecanico, mes, producto, horas,
   cantidad}}], detalle_ot:[{{mecanico, mes, fecha, nro_ot, producto,
   precio_lista, horas, comi_vta}}]}}. "resumen" y "detalle_producto" vienen
   agregados desde el consolidador (nunca fila por fila — el historico completo
   de BDFlexline no cabia en un solo blob de GitHub, ver PASO 11); "detalle_ot"
   (agregado 21/07/2026, a pedido de Cristobal) es la excepcion — viene SIN
   agregar, linea por linea tal cual BDFlexline, para poder ver el detalle real
   de Nº OT/Producto/Precio Lista/Total Horas/Comi_Vta por tecnico. Todo ya
   viene filtrado a solo Mano de Obra (producto "MO_...") y tecnicos con RUT
   valido. "resumen" cubre 12 meses (tabla principal + selector);
   "detalle_producto"/"detalle_ot" (drill-down por tecnico) solo los ultimos 3
   meses. Se filtra por la sucursal del Planificador (tolerante a
   tildes/mayusculas via normSuc). Puramente lectura — no guarda nada en
   control_taller.json. 20/07/2026 (detalle_ot: 21/07/2026). */
function _prodMesLabel(mesKey){{
  if(mesKey==='Sin fecha')return mesKey;
  const p=(mesKey||'').split('-');
  if(p.length<2)return mesKey||'';
  const mi=parseInt(p[1],10)-1;
  return (MESES[mi]||p[1])+' '+p[0];
}}
// FIX 23/07/2026 (ronda 2, a pedido de Cristobal — "estas mostrando personas que ya no
// trabajan desde hace años"): se intento filtrar por el roster de Admin -> Tecnicos
// (Tecnico x Hora), pero Cristobal reporto que con eso "aparecen muy pocos mecanicos
// de los que realmente hay" — ese roster esta pensado para armar el grid de horarios,
// no es una lista completa de todos los que facturan Mano de Obra, y el matching por
// nombre (exacto/substring) fallaba con variantes reales de escritura, excluyendo
// tecnicos activos de verdad. SE QUITO el filtro por roster (ronda 3): el problema real
// de "gente antigua"/horas disparatadas ya se resolvio con el revert de "Sin fecha ->
// mes actual" (los registros viejos/sin fecha ya no se cuelan en el mes en curso, solo
// aparecen si se elige "Sin fecha" a mano) + la exclusion de filas con Total Horas > 24.
// Si en el futuro hace falta sacar a un ex-empleado puntual que SI tenga horas reales
// recientes, hay que resolverlo desde el dato de origen (BDFlexline), no filtrando por
// este roster incompleto.
function _prodRegistrosSucursal(){{
  if(!prodData||!Array.isArray(prodData.resumen))return[];
  const target=normSuc(SUCURSAL);
  return prodData.resumen.filter(r=>normSuc(r.sucursal_mecanico||'')===target);
}}

/* ─── Feriados de Chile (aproximado) + calculo de jornada esperada del mes ───
   Fijos + Semana Santa (algoritmo de Gauss) + San Pedro y San Pablo (ultimo
   lunes de junio, Ley 19.973) + Encuentro de Dos Mundos (se traslada al lunes
   mas cercano si cae martes/miercoles, o al viernes si cae jueves — Ley
   19.668) + Dia Nacional de los Pueblos Indigenas (aprox. 21 de junio, el
   gobierno lo fija oficialmente cada año y puede variar 20/21/22). NO incluye
   feriados extraordinarios ad-hoc (ej. elecciones) — es una aproximacion para
   medir productividad, no un calendario legal oficial. */
function _pascuaGregoriana(year){{
  const a=year%19,b=Math.floor(year/100),c=year%100,d=Math.floor(b/4),e=b%4;
  const f=Math.floor((b+8)/25),g=Math.floor((b-f+1)/3);
  const h=(19*a+b-d-g+15)%30;
  const i=Math.floor(c/4),k=c%4;
  const l=(32+2*e+2*i-h-k)%7;
  const m=Math.floor((a+11*h+22*l)/451);
  const mes=Math.floor((h+l-7*m+114)/31);
  const dia=((h+l-7*m+114)%31)+1;
  return new Date(year,mes-1,dia);
}}
function _feriadosChile(year){{
  const set=new Set();
  const add=(y,m,d)=>set.add(`${{y}}-${{String(m).padStart(2,'0')}}-${{String(d).padStart(2,'0')}}`);
  add(year,1,1); add(year,5,1); add(year,5,21); add(year,7,16); add(year,8,15);
  add(year,9,18); add(year,9,19); add(year,10,31); add(year,11,1); add(year,12,8); add(year,12,25);
  const pascua=_pascuaGregoriana(year);
  const vSanto=new Date(pascua);vSanto.setDate(pascua.getDate()-2);
  const sSanto=new Date(pascua);sSanto.setDate(pascua.getDate()-1);
  add(vSanto.getFullYear(),vSanto.getMonth()+1,vSanto.getDate());
  add(sSanto.getFullYear(),sSanto.getMonth()+1,sSanto.getDate());
  {{
    let d=new Date(year,5,30);
    while(d.getDay()!==1)d.setDate(d.getDate()-1);
    add(d.getFullYear(),d.getMonth()+1,d.getDate());
  }}
  {{
    const base=new Date(year,9,12),dow=base.getDay();
    let d=new Date(base);
    if(dow===2)d=new Date(year,9,11);
    else if(dow===3)d=new Date(year,9,10);
    else if(dow===4)d=new Date(year,9,13);
    add(d.getFullYear(),d.getMonth()+1,d.getDate());
  }}
  add(year,6,21);
  return set;
}}
/* 29/07/2026 (pedido de Cristobal) — Jornada real y productividad acumulada.
   Antes: Lun/Mar 9.75 h y Mie-Vie 8.75 h (horario de puertas abiertas, SIN
   descontar colacion), y el % de productividad se calculaba contra el mes
   COMPLETO — asi, a mitad de mes, todos aparecian con un porcentaje bajisimo
   porque se comparaban las horas vendidas hasta hoy contra la capacidad de
   los 30 dias. Ahora:
     - Cada dia habil descuenta 45 min de colacion:
         Lun/Mar 08:30-18:15 = 9.75 - 0.75 = 9.00 h
         Mie-Vie 08:30-17:15 = 8.75 - 0.75 = 8.00 h
     - El denominador son las HORAS DISPONIBLES ACUMULADAS: horas por dia x
       dias habiles TRANSCURRIDOS (el dia de hoy cuenta completo, confirmado
       por Cristobal). Un mes ya cerrado acumula el mes entero; uno futuro, 0.
     - Si el tecnico esta marcado NO DISPONIBLE (vacaciones/licencia/permiso/
       capacitacion, ver _noDispPeriodos) esos dias habiles NO suman horas
       disponibles — no se le castiga la productividad por dias que no estuvo. */
const PROD_COLACION_H=0.75;
function _prodHorasDia(dow){{
  if(dow===0||dow===6)return 0;
  return ((dow===1||dow===2)?9.75:8.75)-PROD_COLACION_H;
}}
// Hasta que dia del mes se acumulan horas disponibles.
function _prodDiaTopeMes(mesKey){{
  const[y,m]=String(mesKey||'').split('-').map(Number);
  if(!y||!m)return 0;
  const hoy=new Date(),hy=hoy.getFullYear(),hm=hoy.getMonth()+1;
  if(y>hy||(y===hy&&m>hm))return 0;                    // mes futuro
  if(y===hy&&m===hm)return hoy.getDate();              // mes en curso: hasta hoy
  return new Date(y,m,0).getDate();                    // mes cerrado: completo
}}
/* Nucleo unico de calculo. `diaMax` null = mes completo. `tec` null = sin
   descontar dias de no disponibilidad. Devuelve tambien lo descontado, para
   poder mostrarlo en pantalla. */
function _prodJornada(mesKey,diaMax,tec){{
  const vacio={{horas:0,dias:0,horasNoDisp:0,diasNoDisp:0}};
  if(!mesKey||mesKey==='Sin fecha')return vacio;
  const[y,m]=mesKey.split('-').map(Number);
  if(!y||!m)return vacio;
  const feriados=_feriadosChile(y);
  const ultimoDia=new Date(y,m,0).getDate();
  const tope=Math.max(0,Math.min(diaMax==null?ultimoDia:diaMax,ultimoDia));
  const periodos=tec?_noDispPeriodos(tec):[];
  let horas=0,dias=0,horasNoDisp=0,diasNoDisp=0;
  for(let d=1;d<=tope;d++){{
    const dow=new Date(y,m-1,d).getDay();
    const hDia=_prodHorasDia(dow);
    if(!hDia)continue;
    const iso=`${{y}}-${{String(m).padStart(2,'0')}}-${{String(d).padStart(2,'0')}}`;
    if(feriados.has(iso))continue;
    if(periodos.length&&_prodDiaEnPeriodos(iso,periodos)){{horasNoDisp+=hDia;diasNoDisp++;continue;}}
    dias++;horas+=hDia;
  }}
  return{{horas,dias,horasNoDisp,diasNoDisp}};
}}
// Mes completo (capacidad teorica total) — se usa para proyectar el cierre.
function _prodJornadaMes(mesKey){{return _prodJornada(mesKey,null,null);}}
// Horas disponibles ACUMULADAS de un tecnico (denominador de la productividad).
function _prodJornadaDisponible(mesKey,tec){{
  return _prodJornada(mesKey,_prodDiaTopeMes(mesKey),tec||null);
}}

function renderProduccion(){{
  const sel=document.getElementById('prodMes');
  if(!sel)return;
  const regsSuc=_prodRegistrosSucursal();
  const _todos=[...new Set(regsSuc.map(r=>r.mes).filter(Boolean))];
  const meses=_todos.filter(m=>m!=='Sin fecha').sort().reverse();
  if(_todos.includes('Sin fecha'))meses.push('Sin fecha');
  const mesesKey=meses.join(',');
  if(sel.dataset.mesesKey!==mesesKey){{
    sel.innerHTML=meses.length
      ? meses.map(m=>`<option value="${{m}}">${{esc(_prodMesLabel(m))}}</option>`).join('')
      : '<option value="">Sin datos</option>';
    sel.dataset.mesesKey=mesesKey;
    sel.onchange=renderProdTabla;
  }}
  renderProdDiagNombres(regsSuc);
  renderProdTabla();
}}
// Diagnostico visual (pedido de Cristobal — "en el planificador tecnico x hora estan todos
// los tecnicos, pero en produccion no aparece ni la mitad"): compara la lista de tecnicos
// configurados para esta sucursal (grid Tecnico x Hora, Admin -> Tecnicos) contra los
// nombres reales que trajo BDFlexline para esta misma sucursal (cualquier mes, no solo el
// seleccionado) — usando comparacion normalizada (normSuc: sin tildes, mayusculas). Si un
// tecnico configurado no tiene NINGUNA coincidencia, lo mas probable es que el nombre este
// escrito distinto entre ambos sistemas (son 2 fuentes de datos independientes que nunca se
// cruzan por nombre en ningun otro punto del codigo) — no necesariamente que le falten datos.
// 29/07/2026 — matching por TOKENS (palabras), no por igualdad exacta.
// Desde que el consolidador homologa los nombres de BDFlexline al nombre
// canonico de la nomina ("LOBOS LOYOLA JOSE PATRICIO"), el roster de
// Admin -> Tecnicos sigue escrito en el formato corto/invertido de siempre
// ("JOSE LOBOS LOYOLA") — comparados con === nunca calzarian y el aviso de
// abajo se dispararia para casi todos los tecnicos, aunque sus datos esten
// perfectos. Se replica el criterio de _match_nomina_tecnico() de
// consolidar_OTs.py: palabras de 3+ letras (ignora iniciales sueltas), y hay
// coincidencia si el conjunto mas chico esta contenido en el mas grande
// compartiendo al menos 2 palabras.
function _prodTokensNombre(s){{
  return normSuc(s).split(/\\s+/).filter(t=>t.length>=3);
}}
function _prodMismoTecnico(a,b){{
  const ta=_prodTokensNombre(a), tb=_prodTokensNombre(b);
  if(ta.length<2||tb.length<2)return normSuc(a)===normSuc(b);
  const chico=ta.length<=tb.length?ta:tb, grande=new Set(ta.length<=tb.length?tb:ta);
  const comunes=chico.filter(t=>grande.has(t));
  return comunes.length>=2&&comunes.length===chico.length;
}}
function renderProdDiagNombres(regsSuc){{
  const el=document.getElementById('prodDiagNombres');
  if(!el)return;
  if(!tecnicos.length||!regsSuc.length){{el.innerHTML='';return;}}
  const mecanicosReales=[...new Set(regsSuc.map(r=>r.mecanico||'').filter(Boolean))];
  const sinCoincidencia=tecnicos.filter(t=>t&&!mecanicosReales.some(m=>_prodMismoTecnico(t,m)));
  if(!sinCoincidencia.length){{el.innerHTML='';return;}}
  el.innerHTML=`<div class="prod-det-empty" style="background:#fff8e6;border:1px solid #e8c46b;`
    +`border-radius:6px;padding:8px 10px;margin-bottom:8px;color:#7a5c00;">`
    +`⚠️ ${{sinCoincidencia.length}} de ${{tecnicos.length}} técnico(s) configurados en Técnico × Hora `
    +`para esta sucursal no tienen ninguna coincidencia por nombre en Producción Técnicos (BDFlexline): `
    +`<b>${{sinCoincidencia.map(esc).join(', ')}}</b>. Puede ser que el nombre esté escrito distinto entre `
    +`ambos sistemas, o que ese técnico no tenga Mano de Obra facturada en ningún mes disponible.</div>`;
}}
function renderProdStats(filas,mes){{
  const el=document.getElementById('prodStats');
  if(!el)return;
  if(!filas.length){{el.innerHTML='';return;}}
  const totalHoras=filas.reduce((a,f)=>a+f.horas,0);
  const nTec=filas.length;
  const promedio=nTec?totalHoras/nTec:0;
  // La capacidad de la sucursal es la SUMA de las horas disponibles de cada
  // tecnico (no jornada x N), porque cada uno puede tener dias no disponibles
  // distintos — asi el % de la sucursal cuadra con el de las filas de abajo.
  const capacidad=filas.reduce((a,f)=>a+(f.dispHoras||0),0);
  const diasNoDisp=filas.reduce((a,f)=>a+(f.diasNoDisp||0),0);
  const pct=capacidad?(totalHoras/capacidad*100):null;
  const jornadaRef=filas.length?_prodJornadaDisponible(mes,null):{{horas:0,dias:0}};
  el.innerHTML=`
    <div class="prod-kpi"><b>${{totalHoras.toFixed(1)}} h</b><span>Total horas vendidas</span></div>
    <div class="prod-kpi"><b>${{nTec}}</b><span>Técnicos activos</span></div>
    <div class="prod-kpi"><b>${{promedio.toFixed(1)}} h</b><span>Promedio por técnico</span></div>
    <div class="prod-kpi"><b>${{jornadaRef.horas.toFixed(1)}} h</b><span>Horas disponibles acumuladas/técnico (${{jornadaRef.dias}} días hábiles transcurridos)</span></div>
    <div class="prod-kpi"><b>${{pct==null?'--':pct.toFixed(0)+'%'}}</b><span>Productividad promedio sucursal (${{totalHoras.toFixed(1)}} / ${{capacidad.toFixed(1)}} h)</span></div>
    ${{diasNoDisp?`<div class="prod-kpi"><b>${{diasNoDisp}}</b><span>Días no disponibles descontados</span></div>`:''}}
  `;
}}
function renderProdTabla(){{
  const sel=document.getElementById('prodMes');
  const tbody=document.getElementById('prodTbody');
  const totEl=document.getElementById('prodTotalHoras');
  if(!sel||!tbody)return;
  const mes=sel.value;
  const filtro=(document.getElementById('prodBuscarTec').value||'').toLowerCase().trim();
  const regsSuc=_prodRegistrosSucursal();
  const regsMes=mes?regsSuc.filter(r=>r.mes===mes):regsSuc;
  const porTec={{}};
  regsMes.forEach(r=>{{
    const tec=(r.mecanico||'').trim();
    if(!tec)return;
    if(filtro&&!tec.toLowerCase().includes(filtro))return;
    if(!porTec[tec])porTec[tec]={{horas:0,nOt:0}};
    porTec[tec].horas+=Number(r.total_horas)||0;
    porTec[tec].nOt+=Number(r.n_ot)||0;
  }});
  const filas=Object.keys(porTec).map(tec=>{{
    const horas=porTec[tec].horas;
    // Denominador propio de cada tecnico: horas disponibles acumuladas, ya
    // descontados sus dias marcados como no disponible.
    const disp=_prodJornadaDisponible(mes,tec);
    const pct=disp.horas?(horas/disp.horas*100):null;
    return{{tec,horas,nOt:porTec[tec].nOt,pct,
            dispHoras:disp.horas,dispDias:disp.dias,
            diasNoDisp:disp.diasNoDisp,noDispHoy:_noDispActivoHoy(tec)}};
  }}).sort((a,b)=>b.horas-a.horas);
  renderProdStats(filas,mes);
  if(!filas.length){{
    tbody.innerHTML='<tr><td colspan="5" style="text-align:center;color:#889;padding:16px;">'
      +(prodData?'Sin datos de producción para esta sucursal/mes.'
        :'Aún no se ha cargado Producción Técnicos — corre el consolidador (PASO 11).')+'</td></tr>';
    if(totEl)totEl.textContent='0.0';
    return;
  }}
  tbody.innerHTML=filas.map(f=>{{
    const tecEsc=esc(f.tec).replace(/'/g,"&#39;");
    // Formato pedido por Cristobal: horas vendidas / horas disponibles acumuladas.
    const frac=`${{f.horas.toFixed(1)}} / ${{f.dispHoras.toFixed(1)}} h`;
    const badge=f.noDispHoy?`<span class="nodisp-badge" title="Hoy está marcado como no disponible">🚫 No disponible</span>`:'';
    const btn=PUEDE_DISPONIBILIDAD
      ? `<button class="nodisp-btn" onclick="event.stopPropagation();abrirNoDisponible('${{tecEsc}}')">🕐 Disponibilidad</button>`
      : (f.diasNoDisp?`<span class="prod-det-empty" style="padding:0">${{f.diasNoDisp}} día(s) descontado(s)</span>`:'--');
    return `<tr class="prod-row${{f.tec===_prodTecSel?' sel':''}}" onclick="prodSeleccionarTec('${{tecEsc}}')">`
      +`<td>${{esc(f.tec)}} ${{badge}}</td>`
      +`<td style="text-align:right">${{f.horas.toFixed(1)}}</td>`
      +`<td style="text-align:right">${{f.nOt}}</td>`
      +`<td style="text-align:right">${{f.pct==null?'--':f.pct.toFixed(0)+'%'}}`
      +`<div class="prod-frac">${{frac}}${{f.diasNoDisp?` · −${{f.diasNoDisp}}d`:''}}</div></td>`
      +`<td style="text-align:center">${{btn}}</td></tr>`;
  }}).join('');
  if(totEl)totEl.textContent=filas.reduce((a,f)=>a+f.horas,0).toFixed(1);
  if(_prodTecSel)renderProdDetalle(_prodTecSel,mes);
}}
function prodSeleccionarTec(tecEsc){{
  // El nombre viaja HTML-escapado desde el onclick (para que comillas/tildes no
  // rompan el atributo) — se revierte al valor real antes de comparar/guardar.
  const tmp=document.createElement('textarea');tmp.innerHTML=tecEsc;const tec=tmp.value;
  _prodTecSel=(_prodTecSel===tec)?null:tec; // clic de nuevo = cierra el panel
  const det=document.getElementById('prodDetalle');
  if(det)det.style.display=_prodTecSel?'block':'none';
  renderProdTabla();
}}
// Variante de _prodJornadaMes() acotada a los primeros `diaMax` dias del mes — se usa
// para saber cuanta jornada esperada ya transcurrio "hasta hoy" (o hasta el ultimo dia
// del mes si el mes seleccionado ya es un mes cerrado) y poder proyectar el cierre.
function _prodJornadaHastaFecha(mesKey,diaMax){{return _prodJornada(mesKey,diaMax,null);}}
// Calcula la serie de horas acumuladas dia a dia (a partir de detalle_ot, ya filtrado por
// tecnico+mes) y proyecta el cierre del mes segun la productividad real hasta hoy (horas
// reales / jornada esperada hasta hoy * jornada esperada total del mes) — mismo criterio
// de "% Productividad" que ya usa el resto del panel (renderProdDetalle/renderProdStats).
// Si el mes seleccionado ya termino (no es el mes en curso), "hasta hoy" = mes completo,
// por lo que la "proyeccion" coincide con el total real (no hay nada que proyectar).
function _prodDatosProyeccion(otsDetalle,mes){{
  const out={{puntos:[],proyeccion:null,productividad:null,jornadaTotal:{{horas:0,dias:0}},
             horasAcum:0,diaHoy:0,ultimoDia:0,esMesActual:false}};
  if(!mes||mes==='Sin fecha')return out;
  const[y,m]=mes.split('-').map(Number);
  if(!y||!m)return out;
  const ultimoDia=new Date(y,m,0).getDate();
  out.ultimoDia=ultimoDia;
  const hoyReal=new Date();
  out.esMesActual=(hoyReal.getFullYear()===y&&(hoyReal.getMonth()+1)===m);
  out.diaHoy=out.esMesActual?hoyReal.getDate():ultimoDia;

  const porDia={{}};
  (otsDetalle||[]).forEach(r=>{{
    const p=(r.fecha||'').split('/');
    if(p.length!==3)return;
    const d=parseInt(p[0],10);
    if(!d||d<1||d>ultimoDia)return;
    porDia[d]=(porDia[d]||0)+(Number(r.horas)||0);
  }});
  let acum=0;
  Object.keys(porDia).map(Number).sort((a,b)=>a-b).forEach(d=>{{
    acum+=porDia[d];
    out.puntos.push({{dia:d,acumulado:Math.round(acum*100)/100}});
  }});
  out.horasAcum=Math.round(acum*100)/100;

  out.jornadaTotal=_prodJornadaMes(mes);
  const jornadaHastaHoy=_prodJornadaHastaFecha(mes,out.diaHoy);
  if(jornadaHastaHoy.horas>0){{
    out.productividad=out.horasAcum/jornadaHastaHoy.horas;
    out.proyeccion=Math.round(out.productividad*out.jornadaTotal.horas*10)/10;
  }}
  return out;
}}
// Grafico de puntos (SVG puro, sin librerias externas — el iframe del Planificador no
// tiene salida a CDN) con la evolucion de horas acumuladas en el mes y la proyeccion de
// cierre. Puntos azules = horas reales acumuladas dia a dia; linea punteada gris = jornada
// esperada (referencia lineal); punto naranja = proyeccion de cierre segun productividad.
function _prodSvgProyeccion(proy){{
  const W=280,H=170,padL=34,padR=10,padT=10,padB=22;
  const plotW=W-padL-padR,plotH=H-padT-padB;
  const ultimoDia=proy.ultimoDia||30;
  const maxY=Math.max(proy.jornadaTotal.horas||0,proy.horasAcum||0,proy.proyeccion||0,1)*1.12;
  const xPix=d=>padL+((Math.max(1,Math.min(d,ultimoDia))-1)/(Math.max(1,ultimoDia-1)))*plotW;
  const yPix=h=>padT+plotH-(Math.min(h,maxY)/maxY)*plotH;

  if(!proy.puntos.length){{
    return '<div class="prod-det-empty">Sin datos suficientes para graficar la proyección del mes.</div>';
  }}

  // Linea guia punteada: jornada esperada acumulada de forma pareja de dia 1 a ultimoDia.
  const guiaPts=`${{xPix(1)}},${{yPix(0)}} ${{xPix(ultimoDia)}},${{yPix(proy.jornadaTotal.horas||0)}}`;

  const puntosSvg=proy.puntos.map(p=>
    `<circle cx="${{xPix(p.dia).toFixed(1)}}" cy="${{yPix(p.acumulado).toFixed(1)}}" r="3" fill="#1b6ec2"></circle>`
  ).join('');

  let proyeccionSvg='';
  if(proy.proyeccion!=null){{
    const px=xPix(ultimoDia).toFixed(1),py=yPix(proy.proyeccion).toFixed(1);
    proyeccionSvg=`<circle cx="${{px}}" cy="${{py}}" r="4.5" fill="#dd6b20" stroke="#fff" stroke-width="1"></circle>`
      +`<text x="${{px}}" y="${{(yPix(proy.proyeccion)-8).toFixed(1)}}" text-anchor="end" font-size="9" fill="#dd6b20" font-weight="700">${{proy.proyeccion.toFixed(1)}} h</text>`;
  }}
  // Punto "hoy" (ultimo dato real) destacado, si el mes seleccionado sigue en curso.
  let hoySvg='';
  if(proy.esMesActual&&proy.puntos.length){{
    const ultimo=proy.puntos[proy.puntos.length-1];
    hoySvg=`<circle cx="${{xPix(ultimo.dia).toFixed(1)}}" cy="${{yPix(ultimo.acumulado).toFixed(1)}}" r="4.5" fill="#0b2e63" stroke="#fff" stroke-width="1"></circle>`;
  }}

  const ejeY0=yPix(0).toFixed(1);
  return `<svg viewBox="0 0 ${{W}} ${{H}}" width="100%" style="max-width:280px;display:block;">`
    +`<line x1="${{padL}}" y1="${{padT}}" x2="${{padL}}" y2="${{ejeY0}}" stroke="#cfd6dd"></line>`
    +`<line x1="${{padL}}" y1="${{ejeY0}}" x2="${{W-padR}}" y2="${{ejeY0}}" stroke="#cfd6dd"></line>`
    +`<polyline points="${{guiaPts}}" fill="none" stroke="#94a3b8" stroke-width="1.2" stroke-dasharray="3,3"></polyline>`
    +puntosSvg+hoySvg+proyeccionSvg
    +`<text x="${{padL}}" y="${{H-6}}" font-size="9" fill="#667">Día 1</text>`
    +`<text x="${{W-padR}}" y="${{H-6}}" text-anchor="end" font-size="9" fill="#667">Día ${{ultimoDia}}</text>`
    +`<text x="2" y="${{padT+8}}" font-size="9" fill="#667">${{Math.round(maxY)}} h</text>`
    +`</svg>`
    +`<div class="prod-det-empty" style="margin-top:2px;">🔵 Horas acumuladas · 🟠 Proyección al cierre`
    +(proy.productividad!=null?` (productividad ${{(proy.productividad*100).toFixed(0)}}%)`:'')+'</div>';
}}
function renderProdDetalle(tec,mes){{
  const el=document.getElementById('prodDetalle');
  if(!el)return;
  const filasResumen=_prodRegistrosSucursal().filter(r=>r.mecanico===tec&&r.mes===mes);
  const horas=filasResumen.reduce((a,r)=>a+(Number(r.total_horas)||0),0);
  const nOt=filasResumen.reduce((a,r)=>a+(Number(r.n_ot)||0),0);
  const disp=_prodJornadaDisponible(mes,tec);
  const pct=disp.horas?(horas/disp.horas*100):null;

  const productos=(prodData&&Array.isArray(prodData.detalle_producto))
    ?prodData.detalle_producto.filter(r=>r.mecanico===tec&&r.mes===mes).sort((a,b)=>b.horas-a.horas)
    :[];

  /* 29/07/2026 (bug reportado por Cristobal con captura) — este era el ULTIMO
     punto que seguia cruzando el roster de Admin -> Tecnicos contra el nombre
     de Produccion Tecnicos por IGUALDAD EXACTA. Desde que el consolidador
     homologa los nombres al canonico de la nomina ("GARCIA OVALLE CRISTIAN
     CLAUDIO") y el roster sigue en formato corto ("CRISTIAN GARCIA"), el
     findIndex devolvia -1 y las 2 secciones que dependen de idxTec
     ("Vehiculos en Taller asignados" y "Tempario vs Tiempo Asignado vs Reloj
     de Taller") aparecian VACIAS aunque el tecnico si tuviera vehiculos
     asignados en el Planificador/JPCB. Se intenta primero el match exacto
     (rapido, y cubre los nombres crudos aun sin homologar) y si no calza se
     cae al match por tokens, el mismo criterio de renderProdDiagNombres y de
     _match_nomina_tecnico() en consolidar_OTs.py. */
  let idxTec=tecnicos.findIndex(t=>normSuc(t)===normSuc(tec));
  if(idxTec<0)idxTec=tecnicos.findIndex(t=>_prodMismoTecnico(t,tec));
  const asignadosTaller=(idxTec>=0)
    ?ordenes.filter(o=>o.tecnico===idxTec&&!o.cerrada)
    :[];

  const tecEsc=esc(tec).replace(/'/g,"&#39;");
  let html=`<div class="prod-det-head">👤 ${{esc(tec)}} — ${{esc(_prodMesLabel(mes))}}`
    +`<button class="prod-det-close" onclick="prodSeleccionarTec('${{tecEsc}}')">✕ cerrar</button></div>`;
  html+=`<div class="prod-kpis-det">`
    +`<div class="prod-kpi"><b>${{horas.toFixed(1)}} h</b><span>Horas facturadas</span></div>`
    +`<div class="prod-kpi"><b>${{nOt}}</b><span># OT</span></div>`
    +`<div class="prod-kpi"><b>${{disp.horas.toFixed(1)}} h</b><span>Horas disponibles acumuladas (${{disp.dias}} días hábiles)</span></div>`
    +`<div class="prod-kpi"><b>${{pct==null?'--':pct.toFixed(0)+'%'}}</b>`
    +`<span>Productividad (${{horas.toFixed(1)}} / ${{disp.horas.toFixed(1)}} h)</span></div>`
    +(disp.diasNoDisp?`<div class="prod-kpi"><b>${{disp.diasNoDisp}}</b>`
      +`<span>Días no disponibles (−${{disp.horasNoDisp.toFixed(1)}} h)</span></div>`:'')
    +`</div>`;

  // Vehiculos en Taller asignados a este tecnico (todas las ordenes activas, no solo las
  // detenidas) — integrado a Produccion Tecnicos a pedido de Cristobal (22/07/2026), para
  // ver de un vistazo la carga real de taller de cada tecnico junto a su produccion
  // facturada. Reutiliza ordenes/STOPS/tipoInfo ya cargados por el Planificador (misma
  // sucursal) — no requiere ningun dato nuevo. Ubicada primero (antes que el detalle de
  // produccion/horas de BDFlexline), a pedido de Cristobal.
  html+='<div class="prod-det-sub">🚗 Vehículos en Taller asignados</div>';
  if(!asignadosTaller.length){{
    html+='<div class="prod-det-empty">Sin vehículos en Taller asignados a este técnico.</div>';
  }} else {{
    html+='<table class="ctgrid prod-mini"><thead><tr>'
      +'<th>Patente</th><th>N° OT</th><th>Modelo</th><th>Servicio</th><th>Detención</th><th>Comentario</th>'
      +'<th>Fecha ingreso Taller</th><th>ETA</th><th>N° caso</th><th>N° pedido</th></tr></thead><tbody>'
      +asignadosTaller.map(o=>{{
        const st=o.stop?STOPS.find(s=>s.id===o.stop):null;
        const servicio=o.servicio||o.comentarios||o.mantencion||tipoInfo(o).label||'';
        return `<tr><td>${{esc(o.patente)}}</td><td>${{esc(o.ot)||'--'}}</td><td>${{esc(o.modelo)}}</td><td>${{esc(servicio)}}</td>`
          +`<td>${{esc(st?st.t:'--')}}</td><td>${{esc(o.comentario2||'')}}</td>`
          +`<td>${{esc(o.ingreso?isoToDdmmyyyy(o.ingreso)+(o.ingreso_taller?' '+o.ingreso_taller:''):'--')}}</td>`
          +`<td>${{esc(o.eta)||'--'}}</td><td>${{esc(o.numero_caso)||'--'}}</td><td>${{esc(o.n_pedido)||'--'}}</td></tr>`;
      }}).join('')
      +'</tbody></table>';
  }}

  // Tempario vs Tiempo Asignado vs Reloj real de Taller (24/07/2026, a pedido de
  // Cristobal): compara, por cada vehiculo de este tecnico en el mes seleccionado, las 3
  // fuentes de duracion disponibles — el tempario oficial (horas_tempario, del Cotizador
  // de Mantenciones), el tiempo asignado a mano en Control de Taller (Ingreso Taller /
  // Salida Taller, via duracionTallerMin) y el reloj REAL de trabajo en taller (arranca
  // solo al entrar a "En Proceso" y se detiene al pasar a "Lavado"/"Entrega" — ver
  // _actualizarRelojTaller). Se filtra por mes de Ingreso (misma fuente local que
  // "Vehiculos en Taller asignados" de arriba, no depende de BDFlexline).
  const otsRelojMes=(idxTec>=0)
    ?ordenes.filter(o=>o.tecnico===idxTec&&(o.ingreso||'').slice(0,7)===mes)
    :[];
  html+='<div class="prod-det-sub">⏱ Tempario vs Tiempo Asignado vs Reloj de Taller</div>';
  if(!otsRelojMes.length){{
    html+='<div class="prod-det-empty">Sin vehículos de este técnico en el mes seleccionado (según fecha de Ingreso).</div>';
  }} else {{
    html+='<table class="ctgrid prod-mini"><thead><tr>'
      +'<th>Patente</th><th>Servicio</th><th style="text-align:right">Tempario</th>'
      +'<th style="text-align:right">Asignado (Ingreso/Salida Taller)</th><th style="text-align:right">Reloj real</th></tr></thead><tbody>'
      +otsRelojMes.map(o=>{{
        const servicio=o.servicio||o.comentarios||o.mantencion||tipoInfo(o).label||'';
        const durAsig=duracionTallerMin(o);
        const txtTemp=(typeof o.horas_tempario==='number'&&o.horas_tempario>0)?o.horas_tempario.toFixed(1)+' h':'--';
        const txtAsig=(durAsig!==null)?(durAsig/60).toFixed(1)+' h':'--';
        return `<tr><td>${{esc(o.patente)}}</td><td>${{esc(servicio)}}</td>`
          +`<td style="text-align:right">${{txtTemp}}</td>`
          +`<td style="text-align:right">${{txtAsig}}</td>`
          +`<td style="text-align:right">${{relojTallerTexto(o)}}</td></tr>`;
      }}).join('')
      +'</tbody></table>';
  }}

  html+='<div class="prod-det-sub">🔧 Detalle por producto (Mano de Obra)</div>';
  if(!productos.length){{
    html+='<div class="prod-det-empty">Sin detalle por producto disponible para este mes (el detalle solo cubre los últimos 3 meses).</div>';
  }} else {{
    html+='<table class="ctgrid prod-mini"><thead><tr><th>Producto</th><th style="text-align:right">Horas</th><th style="text-align:right">Cantidad</th></tr></thead><tbody>'
      +productos.map(p=>`<tr><td>${{esc(p.producto)}}</td><td style="text-align:right">${{(Number(p.horas)||0).toFixed(1)}}</td><td style="text-align:right">${{p.cantidad}}</td></tr>`).join('')
      +'</tbody></table>';
  }}

  // Detalle por OT — linea por linea (sin agrupar), tal cual viene de BDFlexline: Nº OT ·
  // Producto · Precio Lista · Total Horas · Comi_Vta. Agregado 21/07/2026 a pedido de
  // Cristobal para replicar la sabana de ejemplo (mismas OT+producto pueden repetirse con
  // horas/comision distintas — cada linea es una transaccion real, no se suman entre si).
  const otsDetalle=(prodData&&Array.isArray(prodData.detalle_ot))
    ?prodData.detalle_ot.filter(r=>r.mecanico===tec&&r.mes===mes)
    :[];
  html+='<div class="prod-det-sub">🧾 Detalle por OT</div>';
  if(!otsDetalle.length){{
    html+='<div class="prod-det-empty">Sin detalle por OT disponible para este mes (el detalle solo cubre los últimos 3 meses).</div>';
  }} else {{
    const tablaOt='<table class="ctgrid prod-mini"><thead><tr>'
      +'<th>Nº OT</th><th>Producto</th><th style="text-align:right">Total Horas</th></tr></thead><tbody>'
      +otsDetalle.map(r=>`<tr><td>${{esc(r.nro_ot||'--')}}</td><td>${{esc(r.producto)}}</td>`
        +`<td style="text-align:right">${{(Number(r.horas)||0).toFixed(2)}}</td></tr>`).join('')
      +'</tbody></table>';
    const proy=_prodDatosProyeccion(otsDetalle,mes);
    html+='<div class="prod-ot-flex">'
      +`<div class="prod-ot-tabla">${{tablaOt}}</div>`
      +`<div class="prod-ot-chart">${{_prodSvgProyeccion(proy)}}</div>`
      +'</div>';
  }}
  el.innerHTML=html;
}}

/* ─── JPCB ─── */
// Timestamp corto (fecha+hora local del navegador) para dejar registro de quien hizo el
// ultimo cambio de etapa — solo visual/informativo, no es un log de auditoria formal.
function nowStrCorto(){{
  const n=new Date();
  return n.toLocaleDateString('es-CL')+' '+n.toLocaleTimeString('es-CL',{{hour:'2-digit',minute:'2-digit'}});
}}
// Se llama cada vez que se cambia la Etapa de una orden (drag&drop en el tablero, el
// selector de Control de Taller, o el selector del modal de detalle) — deja registrado
// quien fue y cuando, para mostrarlo directo en la tarjeta del JPCB.
function marcarCambioEtapa(o){{
  o.etapa_usuario=USUARIO;
  o.etapa_fecha=nowStrCorto();
  _actualizarRelojTaller(o);
}}
/* ─── Reloj de Taller (24/07/2026, a pedido de Cristobal) ───
   Mide el tiempo REAL de trabajo en el taller: arranca la primera vez que la orden
   entra a "En Proceso" y se detiene la primera vez que pasa a "Lavado" o "Entrega" —
   pero SOLO si alcanzo a pasar por En Proceso primero (los botones rapidos "📥
   Recepcion"/"🧼 Lavado directo", que saltan etapas, nunca arrancan el reloj por si
   solos: si una orden llega a Lavado sin haber pasado por En Proceso, el reloj queda
   sin datos — `relojTallerTexto` devuelve "--"). Se engancha directo dentro de
   `marcarCambioEtapa(o)`, el punto unico que YA se llama en los 5 lugares donde cambia
   `o.etapa` (drag&drop del JPCB, el selector de Etapa de Control de Taller/Vehiculos en
   Taller, el selector del modal de detalle, y los botones rapidos de Recepcion/Lavado)
   — no hace falta tocar cada uno por separado. Una vez arrancado o detenido, no se
   vuelve a tocar (si la orden retrocede a En Proceso de nuevo, el reloj YA iniciado no
   se reinicia — sigue siendo el tiempo desde la primera vez que entro a trabajarse). */
function _actualizarRelojTaller(o){{
  if(o.etapa==='en_proceso'&&!o.reloj_inicio_ts){{
    o.reloj_inicio_ts=Date.now();
    o.reloj_inicio_txt=nowStrCorto();
  }} else if((o.etapa==='lavado'||o.etapa==='entrega')&&o.reloj_inicio_ts&&!o.reloj_fin_ts){{
    o.reloj_fin_ts=Date.now();
    o.reloj_fin_txt=nowStrCorto();
  }}
}}
// Minutos transcurridos del reloj de taller — null si nunca arranco (nunca paso por "En
// Proceso"). Si arranco pero aun no se detiene (sigue "En Proceso"/"En Prueba"), calcula
// contra el momento actual (reloj "en vivo").
function relojTallerMin(o){{
  if(!o.reloj_inicio_ts)return null;
  const fin=o.reloj_fin_ts||Date.now();
  return Math.max(0,Math.round((fin-o.reloj_inicio_ts)/60000));
}}
function relojTallerTexto(o){{
  const min=relojTallerMin(o);
  if(min===null)return'--';
  const txt=(min/60).toFixed(1)+' h';
  return o.reloj_fin_ts?txt:(txt+' · ⏱ en curso');
}}
// Agregar/editar un comentario directo desde la tarjeta del tablero (sin tener que ir a
// Control de Taller) — reutiliza el mismo campo `comentario2` ("Comentario adicional")
// que ya existe en la tabla, asi ambos lugares muestran siempre lo mismo.
function agregarComentarioTablero(id){{
  const o=byId(id);if(!o)return;
  const txt=prompt(`Comentario para ${{o.patente}}:`,o.comentario2||'');
  if(txt===null)return;
  o.comentario2=txt.trim();
  o.comentario2_usuario=USUARIO;
  o.comentario2_fecha=nowStrCorto();
  renderJPCB();saveCtrl();
  toast(o.comentario2?'💬 Comentario guardado':'💬 Comentario eliminado');
}}
// Botones rapidos "Recepcion" y "Lavado" (23/07/2026, a pedido de Cristobal) — mueven la
// orden DIRECTO a la Etapa correspondiente sin pasar por el resto del flujo (una cita
// puede recepcionarse o lavarse fuera de orden, ej. lavado antes de entrar a taller).
// Ademas de mover la Etapa, dejan una marca PERSISTENTE (recepcionado/lavado_hecho, con
// usuario y fecha) que se mantiene aunque la orden avance despues a otra Etapa — asi
// siempre queda visible que el vehiculo YA fue recepcionado / YA paso por lavado, sin
// importar en que Etapa este ahora.
function marcarRecepcionado(id){{
  const o=byId(id);if(!o)return;
  o.etapa='ingreso_taller';
  o.stop=null;
  o.recepcionado=true;
  o.recepcion_usuario=USUARIO;
  o.recepcion_fecha=nowStrCorto();
  marcarCambioEtapa(o);
  renderJPCB();saveCtrl();
  toast(`📥 ${{o.patente}} recepcionado — pasa a Ingreso a Taller`);
}}
function marcarLavado(id){{
  const o=byId(id);if(!o)return;
  o.etapa='lavado';
  o.stop=null;
  o.lavado_hecho=true;
  o.lavado_usuario=USUARIO;
  o.lavado_fecha=nowStrCorto();
  marcarCambioEtapa(o);
  renderJPCB();saveCtrl();
  toast(`🧼 ${{o.patente}} enviado a Lavado`);
}}
function cardHTML(o){{
  const ti=tipoInfo(o);
  const sStop=o.stop?STOPS.find(s=>s.id===o.stop)?.t||'':'';
  const tecNom=o.tecnico!==null&&o.tecnico!==undefined&&tecnicos[o.tecnico]?tecnicos[o.tecnico]:'';
  const _noShow=esNoAsiste(o.ot,o.patente);
  const _idxEt=ETAPAS.findIndex(e=>e.id===o.etapa);
  const _mostrarVCU=esFord(o)&&_idxEt>=IDX_EN_PROCESO;
  const _vcuOk=_mostrarVCU?vcuCompleto(o):false;
  const _esReagenda=_noShow&&o.estadoCita==='reagenda';
  // Boton "Finalizado" — mismo permiso que Asiste/No Asiste/Reagenda (PUEDE_CONFIRMAR_CITAS,
  // ya incluye a los editores completos). Reutiliza _cerrarOrdenInterno via
  // marcarFinalizadoCita(): la orden pasa al Historial de Taller igual que "🔒 Cerrar".
  const _btnFinalizarCita=PUEDE_CONFIRMAR_CITAS?`<button title="Marcar como Finalizado — pasa al Historial de Taller" onclick="event.stopPropagation();marcarFinalizadoCita('${{o.id}}')">✅ Finalizado</button>`:'';
  return `<div class="card${{_noShow?' no-asiste':''}}" draggable="${{PUEDE_EDITAR}}" data-id="${{o.id}}" style="border-left-color:${{ti.border}};background:${{_noShow?'#f2f2f2':ti.color}}">
    ${{_esReagenda?`<div class="cita-reagenda">🔁 Reagendado para ${{esc(o.fecha_reagenda||'--')}}</div>`:(_noShow?`<div class="cita-noasiste">🚫 Cliente no asiste</div>`:'')}}
    ${{sStop?`<span class="cbadge">⛔ ${{sStop}}</span><br>`:''}}<b>${{o.patente}}</b> <span class="cmeta">${{o.cliente||''}}</span>
    <div><span class="cot">🧾 OT ${{esc(o.ot||'--')}}</span></div>
    <div class="cinfo">${{o.modelo||''}}${{o.modelo&&(o.servicio||o.mantencion)?' · ':''}}${{o.servicio||o.mantencion||ti.label}}</div>
    ${{o.hora_rec?`<div class="cmeta">🕐 Ingreso: ${{o.hora_rec}}</div>`:''}}
    ${{o.hora_compromiso?`<div class="centrega">⏰ Entrega: ${{o.hora_compromiso}}</div>`:''}}
    ${{esCruceNoche(o)?`<div class="centrega">🌙 Sale del taller: ${{fechaSalidaTallerTexto(o)}}</div>`:''}}
    ${{tecNom?`<div class="cmeta">🔧 ${{tecNom}}</div>`:''}}
    ${{o.etapa_usuario?`<div class="cmeta" title="Ultimo cambio de etapa">👤 ${{esc(o.etapa_usuario)}} · ${{o.etapa_fecha||''}}</div>`:''}}
    ${{o.comentario2?`<div class="ccoment">💬 ${{esc(o.comentario2)}}</div>`:''}}
    ${{_mostrarVCU?`<div class="vcu-badge ${{_vcuOk?'ok':'pend'}}" onclick="event.stopPropagation();abrirVCU('${{o.id}}')">📋 VCU ${{_vcuOk?'✅':'⚠️ Falta'}}</div>`:''}}
    ${{o.recepcionado?`<div class="qbadge recepcion">📥 Recepcionado</div>`:''}}
    ${{o.lavado_hecho?`<div class="qbadge lavado">🧼 Lavado</div>`:''}}
    ${{PUEDE_EDITAR?`<div class="cacts">
      ${{STOPS.map(s=>`<button title="${{s.t}}" onclick="event.stopPropagation();setStop('${{o.id}}','${{s.id}}')">⛔</button>`).join('')}}
      ${{o.stop?`<button onclick="event.stopPropagation();clearStop('${{o.id}}')">✅</button>`:''}}
      <button title="Agregar/editar comentario" onclick="event.stopPropagation();agregarComentarioTablero('${{o.id}}')">💬</button>
      <button title="${{_noShow?'Reactivar':'Marcar cliente no asiste'}}" onclick="event.stopPropagation();toggleNoAsiste('${{(o.ot||'').replace(/'/g,"\\'")}}','${{(o.patente||'').replace(/'/g,"\\'")}}')">${{_noShow?'↩️':'🚫'}}</button>
      ${{o.etapa!=='ingreso_taller'?`<button title="Marcar Recepcion → pasa directo a Ingreso Taller" onclick="event.stopPropagation();marcarRecepcionado('${{o.id}}')">📥</button>`:''}}
      ${{o.etapa!=='lavado'?`<button title="Enviar directo a Lavado" onclick="event.stopPropagation();marcarLavado('${{o.id}}')">🧼</button>`:''}}
      ${{_btnFinalizarCita}}
      <button onclick="event.stopPropagation();abrirDetalle('${{o.id}}')">✏️</button>
    </div>`:(PUEDE_CONFIRMAR_CITAS?`<div class="cacts">${{_btnFinalizarCita}}<button onclick="event.stopPropagation();abrirDetalle('${{o.id}}')">👁</button></div>`:`<div class="cacts"><button onclick="event.stopPropagation();abrirDetalle('${{o.id}}')">👁</button></div>`)}}
  </div>`;
}}
/* Fila de titulos "responsable" sobre el tablero JPCB (Asesor/Torre Control/Tecnico/
   Asesor) — cada bloque mide exactamente el ancho de las columnas que agrupa (.col es
   flex:0 0 160px con gap 8px en .kanban), asi queda alineado sin depender de un layout
   de tabla. Puramente informativo — no toca ninguna logica de datos ni de columnas. */
function renderJPCBGrupos(){{
  const el=document.getElementById('jpcbGroups');
  if(!el)return;
  el.innerHTML=GRUPOS_RESPONSABLE.map(g=>{{
    const n=g.ids.length;
    const w=n*160+(n-1)*8;
    return `<div class="kg-block" style="flex-basis:${{w}}px;width:${{w}}px">${{esc(g.label)}}</div>`;
  }}).join('');
}}
function renderJPCB(){{
  renderJPCBGrupos();
  // "Citas <fecha>" (23/07/2026) — primera columna: citas de la Agenda aun SIN
  // confirmar (estadoCita==='pendiente'). No es una Etapa real, no tiene data-etapa
  // (no se puede soltar una tarjeta arrastrada ahi), y sus tarjetas usan
  // cardCitaPendienteHTML() (3 botones) en vez de cardHTML().
  const _citasPend=ordenes.filter(o=>o.estadoCita==='pendiente'&&!o.cerrada&&ordenArea(o)===currentArea);
  const _fechaCitasTxt=(typeof planDates!=='undefined'&&planDates[0])?formatDate(planDates[0]):'';
  const _colCitas=`<div class="col"><div class="col-head" style="background:#0b7d43">📅 Citas ${{_fechaCitasTxt}}<br><span class="cnt">(${{_citasPend.length}})</span></div>
      <div class="drop">${{_citasPend.map(cardCitaPendienteHTML).join('')}}</div></div>`;
  // "No asiste" es una columna propia (no una Etapa mas) — las ordenes marcadas se
  // sacan de su columna habitual y se agrupan aca, para no mezclarlas con el resto
  // del flujo (14/07/2026, a pedido de Cristobal). El campo o.etapa no se toca —
  // al reactivar el cliente, la orden vuelve a aparecer en la etapa que ya tenia.
  const _noAsisteCards=ordenes.filter(o=>esNoAsiste(o.ot,o.patente)&&!o.cerrada&&ordenArea(o)===currentArea);
  const _colNoAsiste=`<div class="col"><div class="col-head" style="background:#7a1f1f">🚫 No asiste<br><span class="cnt">(${{_noAsisteCards.length}})</span></div>
      <div class="drop">${{_noAsisteCards.map(cardHTML).join('')}}</div></div>`;
  const _colsEtapas=ETAPAS.map(et=>{{
    // Tambien se ocultan del JPCB las ordenes con Estado Campaña de "salida"
    // (Quiebre Stock/Cliente desiste/Falla servidor) — siguen visibles en Control de
    // Taller/Vehiculos en Taller, solo dejan de aparecer en el tablero. Y las que aun
    // no fueron confirmadas (citaConfirmada, ver comentario de arriba) — esas viven
    // en la columna "Citas" hasta que alguien presione Asiste.
    const cards=ordenes.filter(o=>o.etapa===et.id&&!o.stop&&!o.cerrada&&citaConfirmada(o)&&
      !_ESTADOS_CAMPANA_OCULTAN_JPCB.includes(o.estado_campana)&&
      ordenArea(o)===currentArea);
    return`<div class="col"><div class="col-head" style="background:${{et.color}}">${{et.t}}<br><span class="cnt">(${{cards.length}})</span></div>
      <div class="drop" data-etapa="${{et.id}}">${{cards.map(cardHTML).join('')}}</div></div>`;
  }}).join('');
  document.getElementById('jpcbBoard').innerHTML=_colCitas+_colNoAsiste+_colsEtapas;
  document.getElementById('stopBoard').innerHTML=STOPS.map(s=>{{
    const cards=ordenes.filter(o=>o.stop===s.id&&!o.cerrada&&ordenArea(o)===currentArea);
    return`<div class="col"><div class="col-head" style="background:#b33">${{s.t}}<br><span class="cnt">(${{cards.length}})</span></div>
      <div class="drop stop-drop" data-stop="${{s.id}}">${{cards.map(cardHTML).join('')}}</div></div>`;
  }}).join('');
  renderFinalizadosSemana();
  wireDnD();
}}
function wireDnD(){{
  if(!PUEDE_EDITAR)return;
  document.querySelectorAll('.card[draggable="true"]').forEach(el=>{{
    el.addEventListener('dragstart',e=>{{e.dataTransfer.setData('cid',el.dataset.id);el.classList.add('dragging');}});
    el.addEventListener('dragend',()=>el.classList.remove('dragging'));
  }});
  document.querySelectorAll('.drop[data-etapa]').forEach(z=>{{
    z.addEventListener('dragover',e=>{{e.preventDefault();z.classList.add('over');}});
    z.addEventListener('dragleave',()=>z.classList.remove('over'));
    z.addEventListener('drop',e=>{{e.preventDefault();z.classList.remove('over');const o=byId(e.dataTransfer.getData('cid'));if(o){{
      if(_avanceBloqueadoPorVCU(o,z.dataset.etapa)){{
        alert(`🚫 No se puede avanzar de etapa — falta completar el VCU (Hoja Multipuntos Ford) de ${{o.patente}}.`);
        return;
      }}
      o.etapa=z.dataset.etapa;o.stop=null;marcarCambioEtapa(o);renderJPCB();saveCtrl();}}}});
  }});
  document.querySelectorAll('.stop-drop[data-stop]').forEach(z=>{{
    z.addEventListener('dragover',e=>{{e.preventDefault();z.classList.add('over');}});
    z.addEventListener('dragleave',()=>z.classList.remove('over'));
    z.addEventListener('drop',e=>{{e.preventDefault();z.classList.remove('over');const o=byId(e.dataTransfer.getData('cid'));if(o){{o.stop=z.dataset.stop;renderJPCB();saveCtrl();}}}});
  }});
}}
function setStop(id,sid){{const o=byId(id);if(o){{o.stop=sid;renderJPCB();saveCtrl();}}}}
function clearStop(id){{const o=byId(id);if(o){{o.stop=null;renderJPCB();saveCtrl();}}}}

/* ─── Planificador — Date tabs ─── */
function renderDateTabs(){{
  document.getElementById('dateTabs').innerHTML=planDates.map((d,i)=>{{
    const dow=d.getDay(),fecha=formatDate(d);
    const lbl=i===0?'📅 Hoy':i===1?'📅 Manana':i===2?'📅 Pasado manana':'📅 +'+i+' dias';
    const dayStr=DIAS[dow]+' '+d.getDate()+' '+MESES[d.getMonth()];
    const cls='dtab'+(i===0?' active':'');
    return`<button class="${{cls}}" data-date="${{fecha}}" onclick="selectDate(this)">${{lbl}} — ${{dayStr}}</button>`;
  }}).join('');
}}
function selectDate(btn){{
  document.querySelectorAll('.dtab').forEach(b=>b.classList.remove('active'));
  btn.classList.add('active');selectedDate=btn.dataset.date;renderPlanView();
}}

/* ─── Planificador — Render ─── */
function renderPlanView(){{renderProgramacion(selectedDate);renderPlanificador(selectedDate);renderLegend(selectedDate);}}

function normSuc(s){{return String(s||'').normalize('NFD').replace(/[̀-ͯ]/g,'').toUpperCase().trim();}}
function getCitas(dateStr){{
  if(!agendaData)return[];
  let suc=null;
  if(agendaData.sucursales){{
    // Match tolerante a mayusculas/tildes: la Agenda a veces usa un nombre distinto
    // al del PBI para la misma sucursal (ej. "Chillán Viejo" vs "CHILLAN VIEJO").
    if(agendaData.sucursales[SUCURSAL]!==undefined){{
      suc=agendaData.sucursales[SUCURSAL];
    }} else {{
      const target=normSuc(SUCURSAL);
      const key=Object.keys(agendaData.sucursales).find(k=>normSuc(k)===target);
      suc=key?agendaData.sucursales[key]:null;
    }}
  }} else {{
    suc=agendaData[SUCURSAL];
  }}
  if(!suc)return[];
  if(Array.isArray(suc))return suc;
  return suc[dateStr]||[];
}}
function getBloques(dateStr){{return ctrlData?.[SUCURSAL]?.bloques?.[dateStr]||[];}}

/* ─── "Eliminada del tablero" — evita que autoImportarCitas() vuelva a crear una
   orden que el usuario borro a mano (eliminarOrdenCT), sin importar la patente. La
   cita de la Agenda puede seguir "ingresada" indefinidamente (ej. el reporte no se
   actualiza, o el vehiculo sigue fisicamente ahi aunque ya no se quiera gestionar) —
   antes, cualquier orden eliminada volvia a aparecer sola en la siguiente carga de
   la pagina. Se identifica por patente + Folio OT (misma clave que usa el matching
   de autoImportarCitas), asi si el MISMO vehiculo vuelve mas adelante con una OT
   nueva, esa SI se importa normal — solo queda bloqueada la combinacion exacta que
   se elimino. Se guarda en ctrlData[SUCURSAL].eliminadas (objeto clave -> true),
   viaja en el mismo control_taller.json. 22/07/2026, a pedido de Cristobal ("no
   solamente es eso [las de prueba], porque hay otras patentes y sigue pasando lo
   mismo independiente de la patente"). */
function _ordenEliminadaKey(pat,ot){{return normPat(pat)+'|'+String(ot||'').trim();}}
function _eliminadasMap(){{
  if(!ctrlData)ctrlData={{}};
  if(!ctrlData[SUCURSAL])ctrlData[SUCURSAL]={{}};
  if(!ctrlData[SUCURSAL].eliminadas)ctrlData[SUCURSAL].eliminadas={{}};
  return ctrlData[SUCURSAL].eliminadas;
}}
function _marcarOrdenEliminada(pat,ot){{
  _eliminadasMap()[_ordenEliminadaKey(pat,ot)]=true;
}}
function _ordenFueEliminada(pat,ot){{
  return !!_eliminadasMap()[_ordenEliminadaKey(pat,ot)];
}}

/* ─── "Cliente no asiste" — marca disponible en toda la app (Programacion, JPCB,
   Control de Taller, Vehiculos en Taller) para el mismo caso/cita, sin importar donde
   se marque. Se identifica por Folio OT (OC) — la misma clave que ya usa el resto del
   codigo para emparejar una cita de la Agenda con su orden en Control de Taller/JPCB.
   Se guarda en ctrlData[SUCURSAL].no_show (objeto oc -> {{usuario,fecha}}), que viaja
   dentro del mismo control_taller.json que ya se guarda con saveCtrl() — no requiere
   ningun archivo ni endpoint nuevo. Si la cita no trae OC (caso raro), se usa la patente
   como respaldo. */
function _noShowMap(){{
  if(!ctrlData)ctrlData={{}};
  if(!ctrlData[SUCURSAL])ctrlData[SUCURSAL]={{}};
  if(!ctrlData[SUCURSAL].no_show)ctrlData[SUCURSAL].no_show={{}};
  return ctrlData[SUCURSAL].no_show;
}}

/* --- Disponibilidad de tecnicos (29/07/2026, pedido de Cristobal) ---------
   Periodos en que un tecnico NO esta disponible (vacaciones, licencia,
   permiso, capacitacion). Esos dias habiles no suman horas disponibles al
   denominador de la productividad — ver _prodJornada().
   Se guarda en el MISMO control_taller.json que ya usa el Planificador:
     ctrlData[SUCURSAL].no_disponible = {{ "NOMBRE NORMALIZADO": [
        {{desde:"AAAA-MM-DD", hasta:"AAAA-MM-DD", motivo:"", usuario:"", fecha:""}} ] }}
   La clave es el nombre normalizado (normSuc) del mecanico tal como aparece en
   Produccion Tecnicos, para que sobreviva a diferencias de tilde/mayuscula. */
function _noDispMap(){{
  if(!ctrlData)ctrlData={{}};
  if(!ctrlData[SUCURSAL])ctrlData[SUCURSAL]={{}};
  if(!ctrlData[SUCURSAL].no_disponible)ctrlData[SUCURSAL].no_disponible={{}};
  return ctrlData[SUCURSAL].no_disponible;
}}
function _noDispPeriodos(tec){{
  const arr=_noDispMap()[normSuc(tec)];
  return Array.isArray(arr)?arr.filter(p=>p&&p.desde):[];
}}
// Las fechas son ISO (AAAA-MM-DD): la comparacion de strings ya ordena bien.
function _prodDiaEnPeriodos(iso,periodos){{
  return periodos.some(p=>iso>=p.desde&&iso<=(p.hasta||p.desde));
}}
function _noDispActivoHoy(tec){{
  return _prodDiaEnPeriodos(isoToday(),_noDispPeriodos(tec));
}}
function agregarNoDisponible(tec,desde,hasta,motivo){{
  if(!PUEDE_DISPONIBILIDAD)return;
  if(!desde)return;
  if(hasta&&hasta<desde){{const t=desde;desde=hasta;hasta=t;}}
  const m=_noDispMap(),k=normSuc(tec);
  if(!Array.isArray(m[k]))m[k]=[];
  m[k].push({{desde:desde,hasta:hasta||desde,motivo:motivo||'',
              usuario:USUARIO,fecha:nowStrCorto()}});
  saveCtrl();renderProdTabla();
}}
function quitarNoDisponible(tec,idx){{
  if(!PUEDE_DISPONIBILIDAD)return;
  const m=_noDispMap(),k=normSuc(tec);
  if(!Array.isArray(m[k]))return;
  m[k].splice(idx,1);
  if(!m[k].length)delete m[k];
  saveCtrl();renderProdTabla();
}}
const NODISP_MOTIVOS=['Vacaciones','Licencia médica','Permiso','Capacitación','Otro'];
let _noDispTec=null;
function abrirNoDisponible(tecEsc){{
  if(!PUEDE_DISPONIBILIDAD)return;
  const tmp=document.createElement('textarea');tmp.innerHTML=tecEsc;_noDispTec=tmp.value;
  renderNoDisponible();
  const ov=document.getElementById('nodisp-modal-overlay');
  if(ov)ov.style.display='flex';
}}
function cerrarNoDisponible(){{
  const ov=document.getElementById('nodisp-modal-overlay');
  if(ov)ov.style.display='none';
  _noDispTec=null;
}}
function renderNoDisponible(){{
  const el=document.getElementById('nodispBody');
  if(!el||!_noDispTec)return;
  const tec=_noDispTec,tecEsc=esc(tec).replace(/'/g,"&#39;");
  const periodos=_noDispPeriodos(tec);
  const activo=_noDispActivoHoy(tec);
  let html=`<div class="prod-det-head">🕐 Disponibilidad — ${{esc(tec)}}`
    +`<button class="prod-det-close" onclick="cerrarNoDisponible()">✕ cerrar</button></div>`;
  html+=`<div class="prod-det-empty" style="margin-bottom:8px">`
    +(activo?`🚫 <b>Hoy está marcado como NO disponible.</b>`
            :`✅ <b>Hoy está disponible.</b>`)
    +` Los días hábiles dentro de los períodos de abajo no suman horas disponibles`
    +` al calcular su productividad.</div>`;
  html+=`<div class="prod-det-sub">Períodos no disponibles</div>`;
  if(!periodos.length){{
    html+=`<div class="prod-det-empty">Sin períodos registrados — el técnico cuenta con jornada completa todos los días hábiles.</div>`;
  }} else {{
    html+='<table class="ctgrid prod-mini"><thead><tr><th>Desde</th><th>Hasta</th><th>Motivo</th><th>Registrado por</th><th></th></tr></thead><tbody>';
    periodos.forEach((p,i)=>{{
      html+=`<tr><td>${{esc(p.desde)}}</td><td>${{esc(p.hasta||p.desde)}}</td>`
        +`<td>${{esc(p.motivo||'--')}}</td>`
        +`<td style="font-size:11px;color:#778">${{esc(p.usuario||'')}}<br>${{esc(p.fecha||'')}}</td>`
        +`<td style="text-align:center"><button class="ct-del" title="Quitar período"`
        +` onclick="quitarNoDisponible('${{tecEsc}}',${{i}});renderNoDisponible();">🗑</button></td></tr>`;
    }});
    html+='</tbody></table>';
  }}
  html+=`<div class="prod-det-sub" style="margin-top:10px">Agregar período</div>`
    +`<div class="nodisp-form">`
    +`<label>Desde <input type="date" id="nodispDesde" class="prod-input"></label>`
    +`<label>Hasta <input type="date" id="nodispHasta" class="prod-input"></label>`
    +`<label>Motivo <select id="nodispMotivo" class="prod-select">`
    +NODISP_MOTIVOS.map(mv=>`<option value="${{esc(mv)}}">${{esc(mv)}}</option>`).join('')
    +`</select></label>`
    +`<button class="nodisp-btn add" onclick="guardarNoDisponible('${{tecEsc}}')">➕ Marcar no disponible</button>`
    +`</div>`;
  el.innerHTML=html;
}}
function guardarNoDisponible(tecEsc){{
  const tmp=document.createElement('textarea');tmp.innerHTML=tecEsc;const tec=tmp.value;
  const d=document.getElementById('nodispDesde'),h=document.getElementById('nodispHasta'),
        mo=document.getElementById('nodispMotivo');
  const desde=d?d.value:'',hasta=h?h.value:'';
  if(!desde){{alert('Selecciona al menos la fecha "Desde".');return;}}
  agregarNoDisponible(tec,desde,hasta||desde,mo?mo.value:'');
  renderNoDisponible();
}}
function _noShowKey(oc,pat){{const k=String(oc||'').trim();return k||('PAT:'+normPat(pat||''));}}
function esNoAsiste(oc,pat){{return !!_noShowMap()[_noShowKey(oc,pat)];}}
function toggleNoAsiste(oc,pat){{
  const key=_noShowKey(oc,pat);
  if(!key||key==='PAT:')return;
  const m=_noShowMap();
  if(m[key]){{
    delete m[key];
    toast(`↩️ ${{pat||oc}} vuelve a estar activo`);
  }}else{{
    m[key]={{usuario:USUARIO,fecha:nowStrCorto()}};
    toast(`🚫 ${{pat||oc}} marcado como "Cliente no asiste"`);
  }}
  if(currentView==='plan')renderPlanView();
  renderControlTaller();renderVehiculosTaller();renderJPCB();
  saveCtrl();
}}

/* ─── "Citas <fecha>" — etapa nueva del JPCB (23/07/2026, a pedido de Cristobal): las
   citas de la Agenda aterrizan aca SIN CONFIRMAR (no en Recepcion como antes) y recien
   pasan al flujo normal (Recepcion, Control de Taller, Vehiculos en Taller, grid
   Tecnico x Hora) cuando alguien confirma "Asiste". Se reutiliza el mismo mapa de
   "Cliente no asiste" (_noShowMap) para No Asiste/Reagenda, asi la columna "🚫 No
   asiste" del JPCB (y el resto de la app que ya usa esNoAsiste()) no necesita tocarse.
   citaConfirmada() es el gate unico: una orden solo cuenta como "en el taller" si
   estadoCita==='asiste' (o no tiene el campo — ordenes viejas/alta manual, retrocompat)
   Y ademas no esta marcada "no asiste" por ningun mecanismo (nuevo o el ya existente). */
function citaConfirmada(o){{
  return (o.estadoCita||'asiste')==='asiste' && !esNoAsiste(o.ot,o.patente);
}}
function _marcarNoShowInterno(oc,pat,extra){{
  const key=_noShowKey(oc,pat);
  if(!key||key==='PAT:')return;
  _noShowMap()[key]=Object.assign({{usuario:USUARIO,fecha:nowStrCorto()}},extra||{{}});
}}
function marcarAsisteCita(id){{
  const o=byId(id);if(!o)return;
  o.estadoCita='asiste';
  o.fecha_reagenda='';
  const key=_noShowKey(o.ot,o.patente);
  if(key&&key!=='PAT:')delete _noShowMap()[key];
  if(currentView==='plan')renderPlanView();
  renderControlTaller();renderVehiculosTaller();renderJPCB();
  saveCtrl();
  toast(`✅ ${{o.patente}} confirmado — pasa a Recepcion`);
}}
function marcarNoAsisteCita(id){{
  const o=byId(id);if(!o)return;
  o.estadoCita='no_asiste';
  o.fecha_reagenda='';
  _marcarNoShowInterno(o.ot,o.patente,{{tipo:'no_asiste'}});
  if(currentView==='plan')renderPlanView();
  renderControlTaller();renderVehiculosTaller();renderJPCB();
  saveCtrl();
  toast(`🚫 ${{o.patente}} marcado como "Cliente no asiste"`);
}}
function marcarReagendaCita(id){{
  const o=byId(id);if(!o)return;
  const actual=o.fecha_reagenda||'';
  const fecha=prompt('Nueva fecha de reagendamiento (DD/MM/AAAA):',actual);
  if(fecha===null)return;
  const f=fecha.trim();
  if(!f){{alert('Debes ingresar la fecha de reagendamiento.');return;}}
  o.estadoCita='reagenda';
  o.fecha_reagenda=f;
  _marcarNoShowInterno(o.ot,o.patente,{{tipo:'reagenda',fecha_reagenda:f}});
  if(currentView==='plan')renderPlanView();
  renderControlTaller();renderVehiculosTaller();renderJPCB();
  saveCtrl();
  toast(`🔁 ${{o.patente}} reagendado para ${{f}}`);
}}
/* Tarjeta de la columna "Citas <fecha>" — misma info de siempre (cardHTML) pero con
   3 botones grandes en vez de las acciones normales; no es arrastrable (no forma
   parte del flujo de Etapas todavia). */
function cardCitaPendienteHTML(o){{
  const ti=tipoInfo(o);
  return `<div class="card cita-pend" data-id="${{o.id}}" style="background:${{ti.color}}">
    <b>${{o.patente}}</b> <span class="cmeta">${{o.cliente||''}}</span>
    <div><span class="cot">🧾 OT ${{esc(o.ot||'--')}}</span></div>
    <div class="cinfo">${{o.modelo||''}}${{o.modelo&&(o.servicio||o.mantencion)?' · ':''}}${{o.servicio||o.mantencion||ti.label}}</div>
    ${{o.hora_rec?`<div class="cmeta">🕐 Ingreso: ${{o.hora_rec}}</div>`:''}}
    ${{o.asesor?`<div class="cmeta">🧑 ${{esc(o.asesor)}}</div>`:''}}
    ${{PUEDE_CONFIRMAR_CITAS?`<div class="cacts-cita">
      <button class="btn-asiste" onclick="event.stopPropagation();marcarAsisteCita('${{o.id}}')">✅ Asiste</button>
      <button class="btn-noasiste" onclick="event.stopPropagation();marcarNoAsisteCita('${{o.id}}')">🚫 No Asiste</button>
      <button class="btn-reagenda" onclick="event.stopPropagation();marcarReagendaCita('${{o.id}}')">🔁 Reagenda</button>
    </div>`:''}}
  </div>`;
}}

/* ─── Lista de asesores de esta sucursal, armada desde la Agenda Curifor (no un
   catalogo fijo) — junta los valores unicos del campo asesor de todas las citas
   (todas las fechas que trae agendaData para esta sucursal), asi el desplegable
   de Asesor en Control de Taller siempre refleja quien atiende realmente aqui.
   22/07/2026: si un asesor real todavia no aparece en ninguna cita de la Agenda para
   esta sucursal (ej. recien contratado, o la Agenda aun no trae ninguna cita suya en
   la ventana descargada), se puede agregar a mano con "➕ Agregar asesor" — queda
   guardado en ctrlData[SUCURSAL].asesores_extra (viaja en control_taller.json, mismo
   archivo de siempre) y se suma SIEMPRE a la lista de aca en adelante, sin depender
   de que la Agenda lo traiga. */
function getAsesoresSucursal(){{
  const set=new Set();
  if(agendaData&&agendaData.sucursales){{
    let suc=agendaData.sucursales[SUCURSAL];
    if(suc===undefined){{
      const target=normSuc(SUCURSAL);
      const key=Object.keys(agendaData.sucursales).find(k=>normSuc(k)===target);
      suc=key?agendaData.sucursales[key]:null;
    }}
    if(suc){{
      const addFrom=arr=>(arr||[]).forEach(c=>{{const a=(c.asesor||'').trim();if(a)set.add(a);}});
      if(Array.isArray(suc))addFrom(suc);
      else Object.values(suc).forEach(addFrom);
    }}
  }}
  (ctrlData?.[SUCURSAL]?.asesores_extra||[]).forEach(a=>{{const n=(a||'').trim();if(n)set.add(n);}});
  return[...set].sort((a,b)=>a.localeCompare(b));
}}
function agregarAsesorManual(){{
  const nombre=(prompt('Nombre del asesor a agregar (para esta sucursal):')||'').trim();
  if(!nombre)return;
  if(!ctrlData)ctrlData={{}};
  if(!ctrlData[SUCURSAL])ctrlData[SUCURSAL]={{}};
  if(!Array.isArray(ctrlData[SUCURSAL].asesores_extra))ctrlData[SUCURSAL].asesores_extra=[];
  if(!ctrlData[SUCURSAL].asesores_extra.some(a=>normSuc(a)===normSuc(nombre))){{
    ctrlData[SUCURSAL].asesores_extra.push(nombre);
  }}
  asesoresSucursal=getAsesoresSucursal();
  renderControlTaller();renderVehiculosTaller();
  saveCtrl();
  toast(`✅ "${{nombre}}" agregado a la lista de asesores de ${{SUCURSAL}}`);
}}
function isTomorrow(dateStr){{return dateStr===formatDate(planDates[1]);}}

function renderProgramacion(dateStr){{
  const citas=getCitas(dateStr).filter(c=>citaArea(c)===currentArea);
  const bloques=getBloques(dateStr).filter(b=>detectArea(b.servicio)===currentArea);
  const asignadosOc=new Set(bloques.map(b=>String(b.oc)));
  const el=document.getElementById('progList');
  if(!citas.length){{el.innerHTML='<div style="padding:20px 10px;color:#888;text-align:center;font-size:12px">Sin citas agendadas para este dia</div>';return;}}
  // Conteo de citas por patente en el dia completo (todas las asesores): sirve para
  // avisar cuando una patente tiene mas de una OT agendada (ej. Mantencion + Garantia/
  // Recall del mismo vehiculo) y para agruparlas visualmente una bajo la otra.
  const patCounts={{}};
  citas.forEach(c=>{{const p=(c.patente||'').replace(/\?/g,'').trim().toUpperCase();if(p)patCounts[p]=(patCounts[p]||0)+1;}});
  const groups={{}};
  // Se ordena por patente primero (para que las citas de un mismo vehiculo con distinta
  // OT queden siempre adyacentes, una bajo la otra) y dentro de cada patente por horario.
  [...citas].sort((a,b)=>{{
    const pa=(a.patente||'').replace(/\?/g,'').trim().toUpperCase();
    const pb=(b.patente||'').replace(/\?/g,'').trim().toUpperCase();
    if(pa!==pb)return pa.localeCompare(pb);
    return(a.horario||'').localeCompare(b.horario||'');
  }}).forEach(c=>{{const a=c.asesor||'Sin asignar';if(!groups[a])groups[a]=[];groups[a].push(c);}});
  let html='';
  for(const[asesor,cs]of Object.entries(groups)){{
    html+=`<div class="prog-group"><div class="prog-asesor">👤 ${{asesor}} (${{cs.length}})</div>`;
    html+=cs.map(c=>{{
      const oc=String(c.oc||c.patente||'');
      const asig=asignadosOc.has(oc);
      const ingr=c.ingresado||false;
      const ico=c.estado==='finalizado'?'🧍':(ingr?'🎟️':'🚗');
      const tip=ingr?'Ingreso al taller':'Pendiente de ingreso';
      const _noShow=esNoAsiste(c.oc,c.patente);
      const ac=(asig?' asignado':'')+(_noShow?' no-asiste':'');
      const _patClean=(c.patente||'').replace(/\?/g,'').trim()||'--';
      const _citaJson=JSON.stringify(c).replace(/'/g,"&#39;");
      // Color segun tipo de servicio (Recall/Mantencion/Diagnostico/Reparacion/Otro) —
      // mismo criterio y paleta que ya usan JPCB y Control de Taller (TIPOS/detectTipo),
      // asi el color significa lo mismo en toda la app.
      const _ti=TIPOS[detectTipo(c)]||TIPOS.ot;
      const _multi=patCounts[_patClean.toUpperCase()]>1;
      // El fondo de color por tipo solo se aplica si la cita NO esta ya asignada a un
      // tecnico — el fondo verde de "asignado" (clase .asignado) sigue teniendo prioridad
      // visual en ese caso. El borde izquierdo de color, en cambio, se muestra siempre
      // (inline, gana sobre la clase) para que el tipo de servicio se reconozca de un
      // vistazo incluso en tarjetas ya asignadas.
      const _bgStyle=(asig||_noShow)?'':`background:${{_ti.color}};`;
      const _ocEsc=oc.replace(/'/g,"\\'");
      const _patEsc=_patClean.replace(/'/g,"\\'");
      return`<div class="cita-card${{ac}}" draggable="${{PUEDE_EDITAR&&!asig}}" data-oc="${{oc}}" data-fecha="${{dateStr}}" data-cita='${{_citaJson}}' title="Clic para ver detalle · ${{tip}} · ${{_ti.label}}" onclick="mostrarDetalleCita(this)" style="border-left:4px solid ${{_ti.border}};${{_bgStyle}}">
        ${{_noShow?`<div class="cita-noasiste">🚫 Cliente no asiste</div>`:''}}
        ${{_multi?`<div style="font-size:9px;font-weight:700;color:#5a3d00;background:#ffe9b3;border-radius:2px;padding:1px 4px;display:inline-block;margin-bottom:2px">🔗 ${{patCounts[_patClean.toUpperCase()]}} OT del vehiculo</div>`:''}}
        <div class="cita-top"><span class="cita-hora">${{c.horario||'--'}}</span><span class="cita-status" title="${{tip}}">${{ico}}</span></div>
        <div class="cita-tipo" style="font-size:9px;font-weight:700;color:${{_ti.border}}">${{_ti.label}}</div>
        <div class="cita-plate">${{_patClean}}</div>
        <div class="cita-info">${{c.modelo||''}}${{c.anio?' ('+c.anio+')':''}}</div>
        <div class="cita-svc">${{c.servicio||c.mantencion||'--'}}</div>
        <div class="cita-cliente">${{c.nombre||c.cliente||''}}</div>
        ${{(c.sucursal&&c.sucursal!==SUCURSAL)?`<div style="font-size:9px;color:#fff;background:#555;border-radius:2px;padding:1px 4px;display:inline-block;margin-top:2px">${{c.sucursal}}</div>`:''}}
        ${{asig?'<div class="cita-asig">✅ Asignado</div>':''}}
        ${{PUEDE_EDITAR?`<button class="cita-noasiste-btn" onclick="event.stopPropagation();toggleNoAsiste('${{_ocEsc}}','${{_patEsc}}')">${{_noShow?'↩️ Reactivar':'🚫 No asiste'}}</button>`:''}}
      </div>`;
    }}).join('');
    html+='</div>';
  }}
  el.innerHTML=html;
  if(PUEDE_EDITAR)el.querySelectorAll('.cita-card[draggable="true"]').forEach(c=>{{
    c.addEventListener('dragstart',e=>{{e.dataTransfer.setData('plan_oc',c.dataset.oc);e.dataTransfer.setData('plan_fecha',c.dataset.fecha);c.classList.add('dragging');}});
    c.addEventListener('dragend',()=>c.classList.remove('dragging'));
  }});
}}

function renderPlanificador(dateStr){{
  const END=getEnd(dateStr);const g=document.getElementById('planGrid');
  let html='<thead><tr><th class="corner">Tecnico \ Hora</th>';
  for(let m=START;m<END;m+=STEP)html+=`<th class="time">${{hhmm(m)}}</th>`;
  html+='</tr></thead><tbody>';
  tecnicos.forEach((t,ti)=>{{
    html+=`<tr><th class="tec">${{t}}</th>`;
    for(let m=START;m<END;m+=STEP)html+=`<td class="slot" data-tec="${{ti}}" data-min="${{m}}" data-fecha="${{dateStr}}"></td>`;
    html+='</tr>';
  }});
  if(!tecnicos.length)html+=`<tr><td colspan="999" style="padding:24px;color:#888;text-align:center">Sin tecnicos configurados. El Admin puede agregarlos en Administracion → Tecnicos.</td></tr>`;
  g.innerHTML=html+'</tbody>';
  getBloques(dateStr).filter(b=>detectArea(b.servicio)===currentArea).forEach(b=>{{
    const sm=parseHH(b.ini||'08:30');
    const cell=g.querySelector(`td[data-tec="${{b.tec}}"][data-min="${{sm}}"]`);
    if(!cell)return;
    const totalDur=b.dur||60;
    const span=Math.max(Math.round(totalDur/STEP),1);
    const div=document.createElement('div');
    div.className='gblock';div.dataset.bid=b.id;
    div.style.width=(span*COLW-3)+'px';
    // Si el bloque trae horas_tempario, la barra se pinta en 2 colores: la porcion
    // que corresponde al tempario (azul) y la que la excede — asignada/extendida a
    // mano, por ejemplo por un atraso (ambar). Sin tempario, se ve como antes.
    // El dato se busca en vivo (no solo el guardado en el bloque al crearlo): asi,
    // bloques viejos (creados antes de que existiera este campo, o antes de re-correr
    // la consolidacion) se pintan solos apenas la orden/cita asociada tenga el dato,
    // sin que haya que reasignar el bloque a mano.
    const ordenAsoc=ordenes.find(o=>{{const bb='ct'+o.id;return bb===b.id||String(b.id).startsWith(bb+'_');}});
    let bHoras=(typeof b.horas_tempario==='number'&&b.horas_tempario>0)?b.horas_tempario:null;
    if(bHoras===null){{
      if(ordenAsoc&&typeof ordenAsoc.horas_tempario==='number'&&ordenAsoc.horas_tempario>0){{
        bHoras=ordenAsoc.horas_tempario;
      }} else {{
        const citaAsoc=getCitas(dateStr).find(c=>String(c.oc||c.patente)===String(b.oc));
        if(citaAsoc&&typeof citaAsoc.horas_tempario==='number'&&citaAsoc.horas_tempario>0)bHoras=citaAsoc.horas_tempario;
      }}
    }}
    // Horario Entrega (compromiso con el cliente) — puramente informativo en el grid,
    // no se usa para calcular ni acomodar el bloque, solo se muestra si existe.
    const horaEntregaB=ordenAsoc?.hora_compromiso||'';
    const entregaInfo=horaEntregaB?`<div class="gentrega">⏰ Entrega: ${{horaEntregaB}}</div>`:'';
    let tempInfo='';
    const tempMin=bHoras?Math.round(bHoras*60):null;
    if(tempMin!==null){{
      if(tempMin<totalDur){{
        const pct=Math.max(0,Math.min(100,(tempMin/totalDur)*100));
        div.style.background=`linear-gradient(to right, #bcd4f0 0%, #bcd4f0 ${{pct}}%, #ffdca8 ${{pct}}%, #ffdca8 100%)`;
        tempInfo=`<div class="gtemp" title="Tempario ${{bHoras.toFixed(1)}}h · asignado ${{(totalDur/60).toFixed(1)}}h">⏳ ${{bHoras.toFixed(1)}}h/${{(totalDur/60).toFixed(1)}}h</div>`;
      }} else {{
        div.style.background='#bcd4f0';
        tempInfo=`<div class="gtemp" title="Tempario ${{bHoras.toFixed(1)}}h">⏳ ${{bHoras.toFixed(1)}}h</div>`;
      }}
    }}
    // Turno que cruza a la jornada siguiente (ver upsertBloqueDesdeOrden): b.cont/'sigue'
    // marca el bloque de la jornada de ingreso (banda hasta el cierre del dia), b.cont
    // ==='viene' marca el bloque de la jornada siguiente (banda desde la apertura) — se
    // avisa con un badge para que quede claro que es el mismo vehiculo, no uno nuevo.
    const contInfo=b.contInfo?`<div class="gtemp" style="background:#4a2a6b;color:#fff" title="${{esc(b.contInfo)}}">${{b.contInfo}}</div>`:'';
    div.innerHTML=`<b>${{b.patente||b.oc||'--'}}</b><div style="font-size:10px;color:#445;white-space:nowrap;overflow:hidden;text-overflow:ellipsis">${{b.modelo||''}}</div><div class="gtime">${{b.ini}}—${{hhmm(sm+totalDur)}}</div>${{tempInfo}}${{entregaInfo}}${{contInfo}}
      ${{PUEDE_EDITAR?`<button class="rm-btn" onclick="event.stopPropagation();removeBloque('${{b.id}}','${{dateStr}}')">✕</button>`:''}}`;
    div.addEventListener('click',()=>abrirDetalleBloque(b));
    cell.appendChild(div);
  }});
  if(!PUEDE_EDITAR)return;
  g.querySelectorAll('td.slot').forEach(td=>{{
    td.addEventListener('dragover',e=>{{e.preventDefault();td.classList.add('over');}});
    td.addEventListener('dragleave',()=>td.classList.remove('over'));
    td.addEventListener('drop',e=>{{
      e.preventDefault();td.classList.remove('over');
      const oc=e.dataTransfer.getData('plan_oc');
      const fecha=e.dataTransfer.getData('plan_fecha');
      if(oc)asignarCita(oc,+td.dataset.tec,+td.dataset.min,fecha||dateStr);
    }});
  }});
}}

function renderLegend(dateStr){{
  const _tiposLeg=Object.values(TIPOS).map(t=>
    `<div class="it"><span class="sw" style="background:${{t.color}};border-color:${{t.border}}"></span> ${{t.label}}</div>`
  ).join('');
  document.getElementById('planLegend').innerHTML=`<b>Estado:</b>
    <div class="it"><span style="font-size:15px">🎟️</span> Ingreso al taller</div>
    <div class="it"><span style="font-size:15px">🚗</span> Pendiente de ingreso</div>
    <div class="it"><span style="font-size:15px">🧍</span> Servicio finalizado</div>
    <div class="it"><span class="sw" style="background:#d6dce5;border-color:#1b6ec2"></span> Asignado a tecnico (sin dato de tempario)</div>
    <div class="it"><span class="sw" style="background:#bcd4f0;border-color:#1b6ec2"></span> Dentro del tiempo del tempario</div>
    <div class="it"><span class="sw" style="background:#ffdca8;border-color:#c87900"></span> Excede el tempario (ajuste manual/atraso)</div>
    <b style="margin-left:14px">Tipo de servicio:</b>
    ${{_tiposLeg}}`;
}}

function asignarCita(oc,tec,min,dateStr){{
  const citas=getCitas(dateStr);
  const cita=citas.find(c=>String(c.oc||c.patente)===String(oc));
  // 23/07/2026, a pedido de Cristobal: no se puede agendar en el grid Tecnico x Hora
  // una cita cuya orden todavia no confirmo asistencia (o esta marcada No Asiste/
  // Reagenda) — primero hay que confirmar "Asiste" en la columna "Citas" del JPCB.
  const _ordenCita=ordenes.find(o=>String(o.ot||o.patente)===String(oc)&&!o.cerrada);
  if(_ordenCita&&!citaConfirmada(_ordenCita)){{
    alert(`🚫 ${{cita?.patente||oc}} todavia no confirma asistencia — confirma "Asiste" en la columna "Citas" del JPCB antes de asignarle horario.`);
    return;
  }}
  if(!ctrlData)ctrlData={{}};
  if(!ctrlData[SUCURSAL])ctrlData[SUCURSAL]={{tecnicos,ordenes,bloques:{{}}}};
  if(!ctrlData[SUCURSAL].bloques)ctrlData[SUCURSAL].bloques={{}};
  if(!ctrlData[SUCURSAL].bloques[dateStr])ctrlData[SUCURSAL].bloques[dateStr]=[];
  const durMin=(typeof cita?.horas_tempario==='number'&&cita.horas_tempario>0)?Math.round(cita.horas_tempario*60):60;
  const end=getEnd(dateStr);
  let finMin=Math.min(min+durMin,end);
  if(finMin<=min)finMin=Math.min(min+STEP,end);
  // Regla imperativa (30/07/2026): un tecnico no puede tener 2 trabajos a la vez — se
  // valida ANTES de tocar nada, excluyendo el propio bloque de esta cita (si ya tenia
  // uno asignado antes) para poder reasignarla/moverla sin toparse consigo misma.
  const otrosDelTec=(ctrlData[SUCURSAL].bloques[dateStr]||[]).filter(b=>+b.tec===+tec&&String(b.oc)!==String(oc));
  const conflicto=otrosDelTec.find(b=>{{const bIni=parseHH(b.ini),bFin=bIni+(b.dur||60);return min<bFin&&finMin>bIni;}});
  if(conflicto){{
    alert(`🚫 ${{tecnicos[tec]||'Este tecnico'}} ya tiene asignado ${{conflicto.patente||conflicto.oc||'otro vehiculo'}} el ${{dateStr}} de ${{conflicto.ini}} a ${{hhmm(parseHH(conflicto.ini)+(conflicto.dur||60))}} — un tecnico no puede tener 2 trabajos al mismo tiempo.`);
    return;
  }}
  ctrlData[SUCURSAL].bloques[dateStr]=ctrlData[SUCURSAL].bloques[dateStr].filter(b=>String(b.oc)!==String(oc));
  const id='b'+Date.now()+Math.random().toString(36).slice(2,5);
  ctrlData[SUCURSAL].bloques[dateStr].push({{id,tec,ini:hhmm(min),dur:finMin-min,oc:String(oc),
    patente:cita?.patente||oc,cliente:cita?.nombre||cita?.cliente||'',
    modelo:cita?.modelo||'',servicio:cita?.servicio||cita?.mantencion||'',
    horas_tempario:(typeof cita?.horas_tempario==='number'?cita.horas_tempario:null)}});
  renderPlanView();saveCtrl();
  toast(`✅ ${{cita?.patente||oc}} → ${{tecnicos[tec]||'T'+tec}} a las ${{hhmm(min)}}`);
}}
function removeBloque(bid,dateStr){{
  if(!ctrlData?.[SUCURSAL]?.bloques?.[dateStr])return;
  ctrlData[SUCURSAL].bloques[dateStr]=ctrlData[SUCURSAL].bloques[dateStr].filter(b=>b.id!==bid);
  renderPlanView();saveCtrl();toast('Asignacion quitada');
}}
function deleteBloque(id){{
  if(!ctrlData?.[SUCURSAL]?.bloques)return;
  for(const[date,bls]of Object.entries(ctrlData[SUCURSAL].bloques)){{
    const idx=bls.findIndex(b=>b.id===id);if(idx>-1){{bls.splice(idx,1);break;}}
  }}
  closeModal();renderPlanView();saveCtrl();toast('Asignacion quitada');
}}

function abrirDetalle(id){{
  const o=byId(id);if(!o)return;
  const ti=tipoInfo(o);
  const etNom=ETAPAS.find(e=>e.id===o.etapa)?.t||o.etapa||'--';
  const stNom=o.stop?STOPS.find(s=>s.id===o.stop)?.t||'--':'--';
  const tecNom=o.tecnico!==null&&o.tecnico!==undefined&&tecnicos[o.tecnico]?tecnicos[o.tecnico]:'Sin asignar';
  const _noShow=esNoAsiste(o.ot,o.patente);
  document.getElementById('modal-title').textContent=`Orden — ${{o.patente}} — OT ${{o.ot||'--'}} — ${{o.modelo||''}}`;
  document.getElementById('modal-del-btn').style.display='none';modalBloqueId=null;
  let body=`${{_noShow?`<div class="cita-noasiste" style="margin-bottom:8px">🚫 Cliente no asiste</div>`:''}}<div class="det-g">
    <div><b>N° OT / OC:</b> <span class="cot">🧾 ${{esc(o.ot||'--')}}</span></div><div><b>Patente:</b> ${{o.patente}}</div>
    <div><b>Modelo:</b> ${{o.modelo||'--'}}</div><div><b>Cliente:</b> ${{o.cliente||'--'}}</div>
    <div><b>Asesor:</b> ${{o.asesor||'--'}}</div>
    <div><b>Servicio:</b> ${{o.servicio||ti.label}}</div><div><b>Etapa:</b> ${{etNom}}</div>
    <div><b>Detencion:</b> ${{stNom}}</div><div><b>Tecnico:</b> ${{tecNom}}</div>
    <div><b>Tiempo estimado:</b> ${{tiempoEstimadoTexto(o)}}</div><div><b>Mantencion:</b> ${{o.mantencion||'--'}}</div>
    <div><b>Horario Ingreso (Agenda):</b> ${{o.hora_rec||'--'}}</div><div><b>Ingreso Taller:</b> ${{o.ingreso_taller||'--'}}</div>
    <div><b>Salida Taller:</b> ${{esCruceNoche(o)?'🌙 '+fechaSalidaTallerTexto(o):(o.salida_taller||'--')}}</div><div></div>
    <div><b>Ingreso (fecha):</b> ${{o.ingreso||'--'}}</div><div><b>Salida (fecha):</b> ${{o.salida||'--'}}</div>
    <div><b>ETA:</b> ${{o.eta||'--'}}</div><div><b>Dias faltantes ETA:</b> ${{calcDiasFaltantesEta(o)}}</div>
    <div><b>Ultimo cambio de etapa:</b> ${{o.etapa_usuario?esc(o.etapa_usuario)+' · '+(o.etapa_fecha||''):'--'}}</div><div></div>
  </div>
  ${{o.hora_compromiso?`<div style="margin-top:8px"><span class="centrega">⏰ Horario de Entrega (compromiso): ${{o.hora_compromiso}}</span></div>`:''}}
  ${{esFord(o)?`<div style="margin-top:8px"><span class="vcu-badge ${{vcuCompleto(o)?'ok':'pend'}}" onclick="abrirVCU('${{o.id}}')">📋 VCU (Hoja Multipuntos Ford) ${{vcuCompleto(o)?'✅ Completo':'⚠️ Pendiente de completar'}}</span></div>`:''}}
  ${{o.recepcionado||o.lavado_hecho?`<div style="margin-top:8px">
    ${{o.recepcionado?`<span class="qbadge recepcion">📥 Recepcionado — ${{esc(o.recepcion_usuario||'')}} · ${{o.recepcion_fecha||''}}</span>`:''}}
    ${{o.lavado_hecho?`<span class="qbadge lavado">🧼 Lavado — ${{esc(o.lavado_usuario||'')}} · ${{o.lavado_fecha||''}}</span>`:''}}
  </div>`:''}}
  <div style="margin-top:8px"><b>Comentario del tablero:</b><br>
    <span style="white-space:pre-wrap">${{o.comentario2?esc(o.comentario2):'--'}}</span>
    ${{o.comentario2&&o.comentario2_usuario?`<div class="cmeta" style="margin-top:2px">— ${{esc(o.comentario2_usuario)}} · ${{o.comentario2_fecha||''}}</div>`:''}}
  </div>`;
  if(PUEDE_EDITAR){{
    const opcTec=`<option value="">Sin asignar</option>`+tecnicos.map((t,i)=>`<option value="${{i}}" ${{o.tecnico===i?'selected':''}}>${{t}}</option>`).join('');
    const opcEt=ETAPAS.map(e=>`<option value="${{e.id}}" ${{o.etapa===e.id?'selected':''}}>${{e.t}}</option>`).join('');
    body+=`<hr>
    <div style="margin-bottom:8px"><b>Tecnico:</b><select onchange="editField('${{o.id}}','tecnico',this.value===''?null:+this.value)" style="margin-left:8px">${{opcTec}}</select></div>
    <div style="margin-bottom:8px"><b>Horario Ingreso (Agenda):</b><input type="time" value="${{o.hora_rec||''}}" onchange="editField('${{o.id}}','hora_rec',this.value)" style="margin-left:8px" title="Mismo campo/alcance que 'Horario Ingreso' en Control de Taller — informativo, no alimenta el Planificador de Tecnicos"></div>
    <div style="margin-bottom:8px"><b>Horario Entrega (compromiso):</b><input type="time" value="${{o.hora_compromiso||''}}" onchange="editField('${{o.id}}','hora_compromiso',this.value)" style="margin-left:8px" title="Mismo campo/alcance que 'Horario Entrega' en Control de Taller — informativo, no alimenta el Planificador de Tecnicos"></div>
    <div style="margin-bottom:8px"><b>Ingreso Taller:</b><input type="time" value="${{o.ingreso_taller||''}}" onchange="editField('${{o.id}}','ingreso_taller',this.value)" style="margin-left:8px" title="Alimenta el Planificador de Tecnicos y el Tiempo Estimado"></div>
    <div style="margin-bottom:8px"><b>Salida Taller:</b><input type="time" value="${{o.salida_taller||''}}" onchange="editField('${{o.id}}','salida_taller',this.value)" style="margin-left:8px" title="Alimenta el Planificador de Tecnicos y el Tiempo Estimado. Si la hora de salida es menor o igual a la de ingreso, se interpreta como turno que cruza a la jornada siguiente."></div>
    <div style="margin-bottom:8px"><b>Duracion (min):</b><input type="number" value="${{o.duracion_min||60}}" min="30" max="480" step="30" onchange="editField('${{o.id}}','duracion_min',+this.value)" style="margin-left:8px;width:80px"></div>
    <div style="margin-bottom:8px"><b>Ingreso (fecha):</b><input type="date" value="${{o.ingreso||''}}" onchange="editField('${{o.id}}','ingreso',this.value)" style="margin-left:8px" title="Mismo campo/alcance que 'Ingreso' en Control de Taller — alimenta el Planificador de Tecnicos"></div>
    <div style="margin-bottom:8px"><b>Salida (fecha):</b><input type="date" value="${{o.salida||''}}" onchange="editField('${{o.id}}','salida',this.value)" style="margin-left:8px" title="Mismo campo/alcance que 'Salida' en Control de Taller — solo informativo"></div>
    <div style="margin-bottom:8px"><b>ETA:</b><input type="date" value="${{o.eta||''}}" onchange="editField('${{o.id}}','eta',this.value)" style="margin-left:8px" title="Mismo campo/alcance que 'ETA' en Control de Taller"></div>
    <div style="margin-bottom:8px"><b>Etapa:</b><select onchange="cambiarEtapaDesdeModal('${{o.id}}',this.value)" style="margin-left:8px">${{opcEt}}</select></div>
    <div style="display:flex;gap:6px;flex-wrap:wrap">
      <button onclick="closeModal();agregarComentarioTablero('${{o.id}}')">💬 Agregar/editar comentario</button>
      <button onclick="closeModal();toggleNoAsiste('${{(o.ot||'').replace(/'/g,"\\'")}}','${{(o.patente||'').replace(/'/g,"\\'")}}')">${{_noShow?'↩️ Reactivar':'🚫 Marcar cliente no asiste'}}</button>
      ${{o.etapa!=='ingreso_taller'?`<button onclick="closeModal();marcarRecepcionado('${{o.id}}')">📥 Marcar Recepcion → Ingreso Taller</button>`:''}}
      ${{o.etapa!=='lavado'?`<button onclick="closeModal();marcarLavado('${{o.id}}')">🧼 Enviar a Lavado</button>`:''}}
      <button onclick="closeModal();cerrarCita('${{o.id}}')">✅ Finalizar</button>
      <button style="color:#b33" onclick="closeModal();eliminarOrdenCT('${{o.id}}')">🗑 Eliminar cita del tablero</button>
    </div>`;
  }}
  document.getElementById('modal-body').innerHTML=body;
  document.getElementById('modal-overlay').classList.add('open');
}}
function abrirDetalleBloque(b){{
  document.getElementById('modal-title').textContent=`Cita — ${{b.patente||b.oc||'--'}}`;
  document.getElementById('modal-del-btn').style.display=PUEDE_EDITAR?'block':'none';
  modalBloqueId=b.id;
  document.getElementById('modal-body').innerHTML=`<div class="det-g">
    <div><b>OC / Folio:</b> ${{b.oc||'--'}}</div><div><b>Patente:</b> ${{b.patente||'--'}}</div>
    <div><b>Cliente:</b> ${{b.cliente||'--'}}</div><div><b>Modelo:</b> ${{b.modelo||'--'}}</div>
    <div><b>Servicio:</b> ${{b.servicio||'--'}}</div><div><b>Tecnico:</b> ${{tecnicos[b.tec]||'--'}}</div>
    <div><b>Hora inicio:</b> ${{b.ini||'--'}}</div><div><b>Duracion:</b> ${{b.dur||60}} min</div>
  </div>${{PUEDE_EDITAR?'<hr><div style="font-size:12px;color:#888">Usa 🗑 Quitar para liberar este slot.</div>':''}}`;
  document.getElementById('modal-overlay').classList.add('open');
}}
function closeModal(){{document.getElementById('modal-overlay').classList.remove('open');}}
function editField(id,field,val){{
  const o=byId(id);if(!o)return;
  // Tecnico/Ingreso Taller/Salida Taller/Ingreso (fecha)/Salida (fecha): regla
  // imperativa — un tecnico no puede tener 2 trabajos a la vez. Si hay choque, se
  // revierte todo y se vuelve a pintar el modal para que los campos muestren el valor
  // original (el select/input ya habia mostrado la eleccion nueva del usuario).
  if(['tecnico','ingreso_taller','salida_taller','ingreso','salida'].includes(field)){{
    const res=_aplicarCambioAgenda(o,field,val);
    if(!res.ok){{
      alert(res.mensaje);
      abrirDetalle(id);
      return;
    }}
    saveCtrl();
    if(currentView==='plan')renderPlanView();
    return;
  }}
  o[field]=val;
  if(field==='etapa')marcarCambioEtapa(o);
  saveCtrl();
  if(currentView==='plan')renderPlanView();
}}
// Cambio de Etapa desde el select del modal de detalle — a diferencia de editField()
// directo, este primero verifica el bloqueo por VCU (Ford, etapa > En Proceso). Si esta
// bloqueado, avisa y vuelve a pintar el modal (revierte visualmente el select) sin
// aplicar el cambio; si esta permitido, aplica el cambio normal y cierra el modal.
function cambiarEtapaDesdeModal(id,val){{
  const o=byId(id);if(!o)return;
  if(_avanceBloqueadoPorVCU(o,val)){{
    alert(`🚫 No se puede avanzar de etapa — falta completar el VCU (Hoja Multipuntos Ford) de ${{o.patente}}.`);
    abrirDetalle(id);
    return;
  }}
  editField(id,'etapa',val);
  closeModal();
  renderJPCB();
}}

// =============================================================
// VCU — render / guardado / PDF (ver VCU_SCHEMA mas arriba)
// =============================================================
function _vcuFieldById(id){{
  for(const sec of VCU_SCHEMA){{
    const f=sec.fields.find(x=>x.id===id);
    if(f)return f;
  }}
  return null;
}}
// Mapa de coordenadas (% del ancho/alto de cada imagen real) — estimado visualmente
// campo por campo sobre las 2 paginas reales del PDF de Ford. p:1 = pagina 1 (Asesor +
// Tecnico primera mitad), p:2 = pagina 2 (Neumaticos + Diagnostico + Firmas).
const VCU_POS={{
  fecha:{{p:1,type:'text',x:12.2,y:21.1,w:17.0}},
  or:{{p:1,type:'text',x:34.8,y:21.1,w:14.5}},
  linea:{{p:1,type:'text',x:12.2,y:22.9,w:11.5}},
  modelo:{{p:1,type:'text',x:29.8,y:22.9,w:19.5}},
  vin:{{p:1,type:'text',x:16.2,y:24.7,w:32.5}},
  fl_fugas:{{p:1,type:'sino',six:22.35,siy:38.6,nox:27.25,noy:38.6,cmb:47.4}},
  fl_aceite_motor:{{p:1,type:'sino',six:9.18,siy:42.9,nox:10.59,noy:42.9}},
  fl_fluido_freno:{{p:1,type:'sino',six:9.18,siy:44.9,nox:10.59,noy:44.9}},
  fl_embrague:{{p:1,type:'sino',six:9.18,siy:46.8,nox:10.59,noy:46.8}},
  fl_dir_hid:{{p:1,type:'sino',six:20.95,siy:42.9,nox:22.41,noy:42.9}},
  fl_limpiaparab:{{p:1,type:'sino',six:20.95,siy:44.9,nox:22.41,noy:44.9}},
  fl_lineas_comb:{{p:1,type:'sino',six:20.95,siy:46.8,nox:22.41,noy:46.8}},
  fl_transmision:{{p:1,type:'sino',six:34.14,siy:42.9,nox:35.73,noy:42.9}},
  fl_refrigerante:{{p:1,type:'sino',six:34.14,siy:44.9,nox:35.73,noy:44.9}},
  fl_diferencial:{{p:1,type:'sino',six:34.14,siy:46.8,nox:35.73,noy:46.8}},
  plumillas:{{p:1,type:'sino',six:34.5,siy:52.4,nox:37.75,noy:52.4,cmb:47.4}},
  luces:{{p:1,type:'sino',six:8.9,siy:57.4,nox:10.4,noy:57.4,cmb:47.4}},
  parabrisas:{{p:1,type:'sino',six:8.9,siy:59.3,nox:10.4,noy:59.3,cmb:47.4}},
  cristales:{{p:1,type:'sino',six:8.9,siy:61.2,nox:10.4,noy:61.2,cmb:47.4}},
  bat_estado:{{p:1,type:'semaforo',x:8.9,y:64.5,cmb:47.4}},
  // Grafico triangular "Nivel de carga de Bateria": va de 0% (vertice, punta baja) a
  // 100% (extremo derecho, mas alto) — x0/x100 son los limites horizontales reales del
  // grafico medidos sobre la imagen, yBase es la linea base (0% de altura) y yTop es la
  // altura del extremo derecho (100%). La marca se dibuja como una linea vertical que
  // interpola la altura segun el % elegido, igual que haria un tecnico a mano.
  bat_nivel_carga:{{p:1,type:'trislider',x0:21.5,x100:42.5,yBase:76.8,yTop:72.35}},
  bat_cca_real:{{p:1,type:'text',x:8.75,y:74.9,w:7.5}},
  bat_cca_fabrica:{{p:1,type:'text',x:38.5,y:74.9,w:9.0}},
  bat_recuperacion:{{p:1,type:'sino',six:47.75,siy:76.9,nox:49.3,noy:76.9}},
  cod_verificacion:{{p:1,type:'radio',o1x:30.0,o1y:81.4,o2x:39.5,o2y:81.4}},
  cod_relenti:{{p:1,type:'radio',o1x:30.0,o1y:83.3,o2x:39.5,o2y:83.3}},
  mangueras_motor:{{p:1,type:'semaforo',x:52.0,y:27.0,cmb:90.4}},
  mangueras_refrig:{{p:1,type:'semaforo',x:52.0,y:29.9,cmb:90.4}},
  correa_accesorios:{{p:1,type:'semaforo',x:52.0,y:32.5,cmb:90.4}},
  frenos_sistema:{{p:1,type:'semaforo',x:52.0,y:36.7,cmb:90.4}},
  suspension:{{p:1,type:'semaforo',x:52.0,y:41.0,cmb:90.4}},
  direccion:{{p:1,type:'semaforo',x:52.0,y:43.6,cmb:90.4}},
  escape:{{p:1,type:'semaforo',x:52.0,y:47.8,cmb:90.4}},
  tren_motriz_del:{{p:1,type:'semaforo',x:52.0,y:51.9,cmb:90.4}},
  tren_motriz_tra:{{p:1,type:'semaforo',x:52.0,y:54.6,cmb:90.4}},
  ac_funcionamiento:{{p:1,type:'semaforo',x:52.0,y:58.8,cmb:90.4}},
  ac_filtro_cabina:{{p:1,type:'semaforo',x:52.0,y:61.2,cmb:90.4}},
  filtro_aire:{{p:1,type:'semaforo',x:52.0,y:65.3,cmb:90.4}},
  filtro_combustible:{{p:1,type:'semaforo',x:52.0,y:67.8,cmb:90.4}},
  parte_inferior_obs:{{p:1,type:'textarea',x:52.4,y:82.3,w:38.8,h:7.2}},
  reinicio_aceite:{{p:2,type:'check',x:6.5,y:26.0}},
  comentarios:{{p:2,type:'textarea',x:5.5,y:28.3,w:17.0,h:32.4}},
  ndi_labrado:{{p:2,type:'semaforo_num',x:23.0,y:27.4,vx:47.5,vy:27.4,vw:3.5}},
  ndi_desgaste:{{p:2,type:'text',x:30.0,y:28.7,w:15}},
  ndi_presion:{{p:2,type:'semaforo_num',x:23.0,y:30.1,vx:51.5,vy:30.1,vw:3.0}},
  ndi_pastillas:{{p:2,type:'semaforo_num',x:23.0,y:31.4,vx:47.5,vy:31.4,vw:3.5}},
  ndi_disco:{{p:2,type:'semaforo_num',x:23.0,y:32.8,vx:47.5,vy:32.8,vw:3.5}},
  ndd_labrado:{{p:2,type:'semaforo_num',x:60.0,y:27.4,vx:87.75,vy:27.4,vw:3.5}},
  ndd_desgaste:{{p:2,type:'text',x:67.0,y:28.7,w:15}},
  ndd_presion:{{p:2,type:'semaforo_num',x:60.0,y:30.1,vx:91.5,vy:30.1,vw:3.0}},
  ndd_pastillas:{{p:2,type:'semaforo_num',x:60.0,y:31.4,vx:87.75,vy:31.4,vw:3.5}},
  ndd_disco:{{p:2,type:'semaforo_num',x:60.0,y:32.8,vx:87.75,vy:32.8,vw:3.5}},
  nti_labrado:{{p:2,type:'semaforo_num',x:23.0,y:42.6,vx:47.5,vy:42.6,vw:3.5}},
  nti_desgaste:{{p:2,type:'text',x:30.0,y:45.1,w:15}},
  nti_presion:{{p:2,type:'semaforo_num',x:23.0,y:47.7,vx:51.5,vy:47.7,vw:3.0}},
  nti_pastillas:{{p:2,type:'semaforo_num',x:23.0,y:50.1,vx:47.5,vy:50.1,vw:3.5}},
  nti_tambor:{{p:2,type:'semaforo_num',x:23.0,y:52.4,vx:47.5,vy:52.4,vw:3.5}},
  ntd_labrado:{{p:2,type:'semaforo_num',x:60.0,y:42.6,vx:87.75,vy:42.6,vw:3.5}},
  ntd_desgaste:{{p:2,type:'text',x:67.0,y:45.1,w:15}},
  ntd_presion:{{p:2,type:'semaforo_num',x:60.0,y:47.7,vx:91.5,vy:47.7,vw:3.0}},
  ntd_pastillas:{{p:2,type:'semaforo_num',x:60.0,y:50.1,vx:87.75,vy:50.1,vw:3.5}},
  ntd_tambor:{{p:2,type:'semaforo_num',x:60.0,y:52.4,vx:87.75,vy:52.4,vw:3.5}},
  nrep_presion:{{p:2,type:'semaforo_num',x:60.0,y:59.4,vx:87.5,vy:59.4,vw:3.0}},
  diag_sintoma:{{p:2,type:'textarea',x:5.5,y:66.3,w:28,h:13.5}},
  diag_componente:{{p:2,type:'textarea',x:34.0,y:66.3,w:28,h:13.5}},
  diag_causa_raiz:{{p:2,type:'textarea',x:64.5,y:66.3,w:28,h:13.5}},
  // Firmas: se autocompletan desde el Asesor/Tecnico ya asignados a la orden en
  // Control de Taller (no hay que volver a tipearlos) — ver prefill en vcuFormHTML.
  nombre_asesor:{{p:2,type:'text',x:9.5,y:95.9,w:38}},
  nombre_tecnico:{{p:2,type:'text',x:56.5,y:95.9,w:38}},
}};
const VCU_OF_CW=1.05, VCU_OF_CH=1.35; // tamano estandar de casilla, % ancho/alto imagen
// Mueve en vivo la linea marcadora del grafico triangular "Nivel de carga de Bateria"
// mientras se arrastra el control deslizante — misma interpolacion que usa fieldMark()
// en el PDF, para que lo que se ve en pantalla sea igual a lo impreso.
function _vcuBatMarkerUpdate(el){{
  const pos=VCU_POS[el.dataset.vcu];
  if(!pos)return;
  const v=Math.max(0,Math.min(100,Number(el.value)||0));
  const x=pos.x0+(pos.x100-pos.x0)*(v/100);
  const yTop=pos.yBase-(pos.yBase-pos.yTop)*(v/100);
  const marker=el.nextElementSibling;
  if(marker){{
    marker.style.left=x+'%';
    marker.style.top=yTop+'%';
    marker.style.height=(pos.yBase-yTop)+'%';
  }}
  const lbl=marker?marker.nextElementSibling:null;
  if(lbl){{lbl.style.left=x+'%';lbl.style.top=(yTop-1.2)+'%';lbl.textContent=v+'%';}}
}}
function _vcuOverlayField(f,datos){{
  const pos=VCU_POS[f.id];
  if(!pos)return '';
  const v=datos[f.id]!==undefined&&datos[f.id]!==null?datos[f.id]:'';
  const dis=PUEDE_EDITAR?'':'disabled';
  const CW=VCU_OF_CW, CH=VCU_OF_CH;
  const cmbHTML=(y)=>{{
    if(pos.cmb===undefined)return '';
    const cv=!!datos[f.id+'_cambiado'];
    return '<input type="checkbox" class="vcuf-of-box" title="Cambiado" style="left:'+pos.cmb+'%;top:'+y+'%;width:'+CW+'%;height:'+CH+'%" data-vcu="'+f.id+'_cambiado" '+(cv?'checked':'')+' '+dis+'>';
  }};
  if(pos.type==='sino'){{
    return '<input type="radio" class="vcuf-of-box" title="SI" style="left:'+pos.six+'%;top:'+pos.siy+'%;width:'+CW+'%;height:'+CH+'%" name="vcuf_'+f.id+'" data-vcu="'+f.id+'" value="si" '+(v==='si'?'checked':'')+' '+dis+'>'
      +'<input type="radio" class="vcuf-of-box" title="NO" style="left:'+pos.nox+'%;top:'+pos.noy+'%;width:'+CW+'%;height:'+CH+'%" name="vcuf_'+f.id+'" data-vcu="'+f.id+'" value="no" '+(v==='no'?'checked':'')+' '+dis+'>'
      +cmbHTML(pos.siy);
  }}
  if(pos.type==='semaforo'||pos.type==='semaforo_num'){{
    const dx=1.6;
    let html='<input type="radio" class="vcuf-of-box vcuf-of-verde" title="Verificado y aprobado" style="left:'+pos.x+'%;top:'+pos.y+'%;width:'+CW+'%;height:'+CH+'%" name="vcuf_'+f.id+'" data-vcu="'+f.id+'" value="verde" '+(v==='verde'?'checked':'')+' '+dis+'>'
      +'<input type="radio" class="vcuf-of-box vcuf-of-amarillo" title="Puede requerir atencion en el futuro" style="left:'+(pos.x+dx)+'%;top:'+pos.y+'%;width:'+CW+'%;height:'+CH+'%" name="vcuf_'+f.id+'" data-vcu="'+f.id+'" value="amarillo" '+(v==='amarillo'?'checked':'')+' '+dis+'>'
      +'<input type="radio" class="vcuf-of-box vcuf-of-rojo" title="Requiere atencion inmediata" style="left:'+(pos.x+2*dx)+'%;top:'+pos.y+'%;width:'+CW+'%;height:'+CH+'%" name="vcuf_'+f.id+'" data-vcu="'+f.id+'" value="rojo" '+(v==='rojo'?'checked':'')+' '+dis+'>'
      +cmbHTML(pos.y);
    if(pos.type==='semaforo_num'){{
      const vn=datos[f.id+'_valor']!==undefined&&datos[f.id+'_valor']!==null?datos[f.id+'_valor']:'';
      html+='<input type="text" class="vcuf-of-text" style="left:'+pos.vx+'%;top:'+pos.vy+'%;width:'+pos.vw+'%;text-align:right" data-vcu="'+f.id+'_valor" value="'+esc(vn)+'" '+dis+'>';
    }}
    return html;
  }}
  if(pos.type==='radio'){{
    const opts=f.opts||[];
    let html='';
    if(opts[0])html+='<input type="radio" class="vcuf-of-box" title="'+esc(opts[0][1]||'')+'" style="left:'+pos.o1x+'%;top:'+pos.o1y+'%;width:'+CW+'%;height:'+CH+'%" name="vcuf_'+f.id+'" data-vcu="'+f.id+'" value="'+opts[0][0]+'" '+(v===opts[0][0]?'checked':'')+' '+dis+'>';
    if(opts[1])html+='<input type="radio" class="vcuf-of-box" title="'+esc(opts[1][1]||'')+'" style="left:'+pos.o2x+'%;top:'+pos.o2y+'%;width:'+CW+'%;height:'+CH+'%" name="vcuf_'+f.id+'" data-vcu="'+f.id+'" value="'+opts[1][0]+'" '+(v===opts[1][0]?'checked':'')+' '+dis+'>';
    return html;
  }}
  if(pos.type==='check'){{
    return '<input type="checkbox" class="vcuf-of-box" style="left:'+pos.x+'%;top:'+pos.y+'%;width:'+CW+'%;height:'+CH+'%" data-vcu="'+f.id+'" '+(v?'checked':'')+' '+dis+'>';
  }}
  if(pos.type==='textarea'){{
    return '<textarea class="vcuf-of-textarea" style="left:'+pos.x+'%;top:'+pos.y+'%;width:'+pos.w+'%;height:'+pos.h+'%" data-vcu="'+f.id+'" '+dis+'>'+esc(v)+'</textarea>';
  }}
  if(pos.type==='date'){{
    return '<input type="date" class="vcuf-of-text" style="left:'+pos.x+'%;top:'+pos.y+'%;width:'+pos.w+'%" data-vcu="'+f.id+'" value="'+esc(v)+'" '+dis+'>';
  }}
  if(pos.type==='trislider'){{
    const vv=(v!==''&&v!==undefined&&v!==null)?Math.max(0,Math.min(100,Number(v)||0)):0;
    const mx=pos.x0+(pos.x100-pos.x0)*(vv/100);
    const myTop=pos.yBase-(pos.yBase-pos.yTop)*(vv/100);
    const mh=pos.yBase-myTop;
    return '<input type="range" class="vcuf-of-range" min="0" max="100" style="left:'+pos.x0+'%;top:'+(pos.yBase+0.8)+'%;width:'+(pos.x100-pos.x0)+'%" data-vcu="'+f.id+'" value="'+vv+'" oninput="_vcuBatMarkerUpdate(this)" '+dis+'>'
      +'<div class="vcuf-bat-marker" style="left:'+mx+'%;top:'+myTop+'%;height:'+mh+'%"></div>'
      +'<div class="vcuf-bat-marker-lbl" style="left:'+mx+'%;top:'+(myTop-1.2)+'%">'+vv+'%</div>';
  }}
  return '<input type="text" class="vcuf-of-text'+(v?'':' vcuf-of-empty')+'" style="left:'+pos.x+'%;top:'+pos.y+'%;width:'+pos.w+'%" data-vcu="'+f.id+'" value="'+esc(v)+'" '+dis+'>';
}}
function vcuFormHTML(o){{
  const datos=vcuDatos(o);
  if(!datos.or)datos.or=o.ot||'';
  if(!datos.modelo)datos.modelo=o.modelo||'';
  if(!datos.vin)datos.vin=o.vin||'';
  if(!datos.nombre_asesor)datos.nombre_asesor=o.asesor||'';
  if(!datos.nombre_tecnico){{
    const _tn=(o.tecnico!==null&&o.tecnico!==undefined&&tecnicos[o.tecnico])?tecnicos[o.tecnico]:'';
    if(_tn)datos.nombre_tecnico=_tn;
  }}
  if(!datos.fecha)datos.fecha=isoToday();
  let camposP1='',camposP2='';
  for(const sec of VCU_SCHEMA){{
    for(const f of sec.fields){{
      const pos=VCU_POS[f.id];
      if(!pos)continue;
      const html=_vcuOverlayField(f,datos);
      if(pos.p===1)camposP1+=html; else camposP2+=html;
    }}
  }}
  return ''
    +'<div class="vcuf-overlay-wrap"><img src="'+VCU_IMG_P1+'" alt="VCU Ford pagina 1">'+camposP1+'</div>'
    +'<div class="vcuf-overlay-wrap"><img src="'+VCU_IMG_P2+'" alt="VCU Ford pagina 2">'+camposP2+'</div>';
}}
let vcuOrdenId=null;
function abrirVCU(id){{
  const o=byId(id);if(!o)return;
  vcuOrdenId=id;
  document.getElementById('vcu-modal-title').textContent=`📋 VCU — Hoja Multipuntos Ford — ${{o.patente}} — ${{o.modelo||''}}`;
  document.getElementById('vcu-modal-body').innerHTML=vcuFormHTML(o);
  const est=vcuEstado(o);
  document.getElementById('vcu-modal-estado').textContent=est&&est.completo?`✅ Completo — ${{est.tecnico||''}} · ${{est.fecha||''}}`:'⚠️ Pendiente de completar. Todos los campos con * son obligatorios.';
  document.getElementById('vcu-modal-overlay').classList.add('open');
}}
function cerrarVCU(){{document.getElementById('vcu-modal-overlay').classList.remove('open');vcuOrdenId=null;}}
function _vcuLeerFormulario(){{
  const datos={{}};
  document.querySelectorAll('#vcu-modal-body [data-vcu]').forEach(el=>{{
    const key=el.dataset.vcu;
    if(el.type==='checkbox')datos[key]=el.checked;
    else if(el.type==='radio'){{if(el.checked)datos[key]=el.value;}}
    else datos[key]=el.value;
  }});
  return datos;
}}
function vcuGuardar(marcarCompleto){{
  if(!PUEDE_EDITAR){{toast('Modo solo lectura — no se puede editar el VCU');return;}}
  const o=byId(vcuOrdenId);if(!o)return;
  const datos=_vcuLeerFormulario();
  if(marcarCompleto){{
    const faltan=vcuFaltantes(datos);
    if(faltan.length){{
      alert(`Faltan ${{faltan.length}} campo(s) obligatorio(s) por completar:\\n\\n`+faltan.slice(0,15).map(f=>'• '+f.label).join('\\n')+(faltan.length>15?`\\n... y ${{faltan.length-15}} mas`:''));
      return;
    }}
  }}
  const m=_vcuMap();
  const prev=m[o.id];
  m[o.id]={{datos,completo:marcarCompleto?true:!!(prev&&prev.completo),tecnico:USUARIO,fecha:nowStrCorto()}};
  saveCtrl();
  renderJPCB();renderControlTaller();renderVehiculosTaller();
  const est=m[o.id];
  document.getElementById('vcu-modal-estado').textContent=est.completo?`✅ Completo — ${{est.tecnico||''}} · ${{est.fecha||''}}`:'⚠️ Pendiente de completar. Todos los campos con * son obligatorios.';
  toast(marcarCompleto?'✅ VCU marcado como completo':'💾 Borrador de VCU guardado');
}}
function _vcuLabelValor(f,datos){{
  const v=datos[f.id];
  if(f.type==='sino')return v==='si'?'SI':v==='no'?'NO':'--';
  if(f.type==='semaforo'){{
    const map={{verde:'🟢 Verificado y aprobado',amarillo:'🟡 Atencion en el futuro',rojo:'🔴 Atencion inmediata'}};
    return map[v]||'--';
  }}
  if(f.type==='semaforo_num'){{
    const map={{verde:'🟢',amarillo:'🟡',rojo:'🔴'}};
    const val=datos[f.id+'_valor']||'--';
    return `${{map[v]||'⚪'}} ${{val}}`;
  }}
  if(f.type==='radio'){{
    const opt=(f.opts||[]).find(([val])=>val===v);
    return opt?opt[1]:'--';
  }}
  if(f.type==='check')return v?'Si':'No';
  if(f.type==='trislider')return(v!==undefined&&v!==null&&String(v).trim()!=='')?`${{v}}%`:'--';
  return(v!==undefined&&v!==null&&String(v).trim()!=='')?v:'--';
}}
function vcuDescargarPDF(){{
  const o=byId(vcuOrdenId);if(!o)return;
  // El admin puede descargar el VCU aunque falten campos obligatorios por completar
  // (a pedido de Cristobal, 15/07/2026) — el resto de los usuarios sigue necesitando
  // marcarlo como completo antes de poder descargarlo.
  const esAdmin=(USUARIO||'').toLowerCase()==='cjerez@curifor.com';
  const est=vcuEstado(o);
  if(!esAdmin&&(!est||!est.completo)){{
    alert('🚫 Para descargar el VCU primero hay que completar todos los campos obligatorios y guardarlo como "completo".');
    return;
  }}
  _generarPdfVCU(o, est||{{datos:_vcuLeerFormulario(),tecnico:USUARIO,fecha:nowStrCorto()}});
}}
function _generarPdfVCU(o,est){{
  // Igual que el formulario en pantalla: se dibuja ENCIMA de la imagen real del PDF
  // de Ford (no una recreacion), esta vez con marcas/texto estaticos listos para
  // imprimir en vez de inputs editables. Reutiliza el mismo VCU_POS de coordenadas.
  const datos=est.datos||{{}};
  const CW=VCU_OF_CW, CH=VCU_OF_CH;
  const mark=(x,y)=>'<div class="vcu-pdf-mark" style="left:'+x+'%;top:'+y+'%;width:'+CW+'%;height:'+CH+'%">&#10003;</div>';
  const txt=(x,y,w,val,align)=>'<div class="vcu-pdf-text" style="left:'+x+'%;top:'+y+'%;width:'+(w||10)+'%;'+(align?('text-align:'+align+';'):'')+'">'+esc(val||'')+'</div>';
  const area=(x,y,w,h,val)=>'<div class="vcu-pdf-area" style="left:'+x+'%;top:'+y+'%;width:'+w+'%;height:'+h+'%">'+esc(val||'').replace(/\\n/g,'<br>')+'</div>';

  function cmbMark(f,y){{
    const pos=VCU_POS[f.id];
    if(!pos||pos.cmb===undefined)return '';
    return datos[f.id+'_cambiado']?mark(pos.cmb,y):'';
  }}

  function fieldMark(f){{
    const pos=VCU_POS[f.id];
    if(!pos)return '';
    const v=datos[f.id];
    if(pos.type==='sino'){{
      let html='';
      if(v==='si')html+=mark(pos.six,pos.siy);
      else if(v==='no')html+=mark(pos.nox,pos.noy);
      html+=cmbMark(f,pos.siy);
      return html;
    }}
    if(pos.type==='semaforo'||pos.type==='semaforo_num'){{
      const dx=1.6;
      let html='';
      if(v==='verde')html+=mark(pos.x,pos.y);
      else if(v==='amarillo')html+=mark(pos.x+dx,pos.y);
      else if(v==='rojo')html+=mark(pos.x+2*dx,pos.y);
      if(pos.type==='semaforo_num'){{
        html+=txt(pos.vx,pos.vy,pos.vw,datos[f.id+'_valor'],'right');
      }}
      html+=cmbMark(f,pos.y);
      return html;
    }}
    if(pos.type==='radio'){{
      const opts=f.opts||[];
      if(opts[0]&&v===opts[0][0])return mark(pos.o1x,pos.o1y);
      if(opts[1]&&v===opts[1][0])return mark(pos.o2x,pos.o2y);
      return '';
    }}
    if(pos.type==='check'){{
      return v?mark(pos.x,pos.y):'';
    }}
    if(pos.type==='textarea'){{
      return area(pos.x,pos.y,pos.w,pos.h,v);
    }}
    if(pos.type==='trislider'){{
      if(v===undefined||v===null||v==='')return '';
      const vv=Math.max(0,Math.min(100,Number(v)||0));
      const bx=pos.x0+(pos.x100-pos.x0)*(vv/100);
      const byTop=pos.yBase-(pos.yBase-pos.yTop)*(vv/100);
      const bh=pos.yBase-byTop;
      return '<div class="vcu-pdf-batline" style="left:'+bx+'%;top:'+byTop+'%;height:'+bh+'%"></div>'
        +'<div class="vcu-pdf-batlbl" style="left:'+bx+'%;top:'+(byTop-1.0)+'%">'+vv+'%</div>';
    }}
    return txt(pos.x,pos.y,pos.w,v);
  }}

  let camposP1='',camposP2='';
  for(const sec of VCU_SCHEMA){{
    for(const f of sec.fields){{
      const pos=VCU_POS[f.id];
      if(!pos)continue;
      const html=fieldMark(f);
      if(pos.p===1)camposP1+=html; else camposP2+=html;
    }}
  }}

  const CSS='@page{{size:landscape;margin:0;}}'
    +'*{{box-sizing:border-box;font-family:Arial,Helvetica,sans-serif;}}'
    +'body{{margin:0;padding:0;background:#e9edf1;}}'
    +'.vcu-pdf-meta{{text-align:center;font-size:11px;color:#556;padding:6px 4px;}}'
    +'.vcu-pdf-page{{position:relative;width:100%;max-width:1400px;margin:0 auto 16px;line-height:0;background:#fff;}}'
    +'.vcu-pdf-page img{{display:block;width:100%;height:auto;}}'
    +'.vcu-pdf-mark{{position:absolute;display:flex;align-items:center;justify-content:center;font-weight:900;color:#000;font-size:1.15vw;line-height:1;}}'
    +'.vcu-pdf-text{{position:absolute;font-size:1.05vw;font-weight:700;color:#111;line-height:1.15;white-space:nowrap;overflow:hidden;}}'
    +'.vcu-pdf-area{{position:absolute;font-size:0.92vw;font-weight:600;color:#111;line-height:1.25;overflow:hidden;}}'
    +'.vcu-pdf-batline{{position:absolute;width:.3%;background:#000;}}'
    +'.vcu-pdf-batlbl{{position:absolute;font-size:.9vw;font-weight:700;color:#000;transform:translate(-50%,-100%);white-space:nowrap;}}'
    +'@media print{{body{{background:#fff;}}.vcu-pdf-page{{page-break-after:always;max-width:none;}}.vcu-pdf-page:last-child{{page-break-after:auto;}}.no-print{{display:none;}}}}';

  const meta='<div class="vcu-pdf-meta">Generado '+esc(new Date().toLocaleDateString('es-CL'))+' &middot; Sucursal '+esc(SUCURSAL)+' &middot; Completado por '+esc(est.tecnico||'')+' el '+esc(est.fecha||'')+' &middot; Patente '+esc(o.patente)+'</div>';

  const htmlDoc='<!DOCTYPE html><html lang="es"><head><meta charset="UTF-8"><title>VCU '+esc(o.patente)+'</title><style>'+CSS+'</style></head><body>'
    +meta
    +'<div class="vcu-pdf-page"><img src="'+VCU_IMG_P1+'" alt="VCU Ford pagina 1">'+camposP1+'</div>'
    +'<div class="vcu-pdf-page"><img src="'+VCU_IMG_P2+'" alt="VCU Ford pagina 2">'+camposP2+'</div>'
    +'<div class="no-print" style="text-align:center;margin:14px 0;"><button onclick="window.print()" style="padding:8px 20px;font-size:13px;cursor:pointer">Imprimir / Guardar como PDF</button></div>'
    +'</body></html>';

  const w=window.open('','_blank');
  if(!w){{toast('El navegador bloqueo la ventana emergente — habilitala para descargar el VCU');return;}}
  w.document.write(htmlDoc);
  w.document.close();
}}

loadData();
</script>
</body>
</html>"""



# ============================================================
#   HELPERS DE PERSISTENCIA
#
#   MIGRACIÓN A SUPABASE (10-08-2026)
#   ---------------------------------
#   Estas funciones eran la "base de datos" de la app: leían y escribían los
#   JSON en el repositorio de GitHub. Ahora, si hay credencial de Supabase,
#   leen y escriben en la tabla `documentos`; si no la hay, siguen usando
#   GitHub exactamente como antes.
#
#   Los NOMBRES y lo que DEVUELVEN no cambian, a propósito: los 44 lugares
#   que las llaman quedan intactos y ninguna pantalla se toca. El `sello`
#   que devuelve Supabase cumple el mismo papel que el `sha` de GitHub.
#
#   Que funcione sin credencial no es adorno: permite publicar el código sin
#   cortar nada y encender la migración cuando se quiera, con un solo secreto.
# ============================================================
def _github_headers():
    return {
        "Authorization": f"token {GITHUB_TOKEN}",
        "Accept": "application/vnd.github.v3+json",
    }


def _leer_json_github_raw(nombre_archivo):
    """Lee un JSON por su nombre. Retorna (sello, dict) o (None, {})."""
    if _datos.disponible():
        sello, datos = _datos.leer_con_sello(nombre_archivo)
        return (sello, datos) if datos is not None else (None, {})

    url = f"https://api.github.com/repos/{GITHUB_USUARIO}/{GITHUB_REPO}/contents/{nombre_archivo}"
    try:
        resp = requests.get(url, headers=_github_headers(), timeout=15, verify=False)
        if resp.status_code == 404:
            return None, {}
        resp.raise_for_status()
        info = resp.json()
        sha = info.get("sha")
        raw = info.get("content", "").replace("\n", "").strip()
        if raw:
            datos = json.loads(base64.b64decode(raw).decode("utf-8"))
        else:
            dl = info.get("download_url", "")
            r2 = requests.get(dl, timeout=30, verify=False)
            datos = r2.json()
        return sha, datos
    except Exception:
        return None, {}


def _leer_json_github_blob(nombre_archivo):
    """
    Lee un JSON directo del blob del último commit en 'main' vía Git Data API.
    A diferencia de _leer_json_github_raw / raw.githubusercontent.com, esto NUNCA
    pasa por el CDN público (Fastly) que puede servir una copia cacheada varios
    minutos después de un commit nuevo. Tampoco tiene el límite de 1MB de la
    Contents API (que para archivos grandes redirige a un download_url que sigue
    siendo raw.githubusercontent.com, reintroduciendo el mismo problema).
    Devuelve el dict, o None si algo falla.

    Con Supabase ese problema del CDN no existe —una consulta a la base siempre
    ve lo último—, así que las dos funciones de lectura resuelven igual.
    """
    if _datos.disponible():
        return _datos.leer(nombre_archivo)

    try:
        base_url = f"https://api.github.com/repos/{GITHUB_USUARIO}/{GITHUB_REPO}"
        hdrs = _github_headers()
        # La Contents API devuelve el "sha" del blob en una sola llamada, sin importar
        # el tamaño del archivo (para archivos >1MB solo omite "content", el sha siempre
        # viene) — evita la llamada a git/ref + git/trees?recursive=1 (que además lista
        # TODO el árbol del repo) que se usaba antes solo para llegar a ese mismo sha.
        # Menos llamadas = menos latencia en cada rerun de Streamlit que dependa de esto.
        r = requests.get(f"{base_url}/contents/{nombre_archivo}", headers=hdrs,
                         params={"ref": "main"}, timeout=15, verify=False)
        r.raise_for_status()
        blob_sha = r.json().get("sha")
        if not blob_sha:
            return None
        r = requests.get(f"{base_url}/git/blobs/{blob_sha}", headers=hdrs, timeout=30, verify=False)
        r.raise_for_status()
        raw = r.json().get("content", "").replace("\n", "")
        return json.loads(base64.b64decode(raw).decode("utf-8"))
    except Exception:
        return None


class _RespuestaDoc:
    """Imita la respuesta de `requests` para un documento leído de Supabase.

    Los cargadores de la app (`cargar_datos`, `_cargar_agenda_hoy`, …) piden el
    documento con `requests.get(url)` y después leen `.status_code`, `.json()`
    y a veces el formato propio de la Contents API de GitHub (`sha` + `content`
    en base64). En vez de reescribir cada uno de esos cargadores —que es donde
    se rompen las cosas— se les devuelve algo con la misma forma. Así el código
    que interpreta la respuesta queda intacto y solo cambia de dónde salió.
    """

    def __init__(self, datos, sello=None, formato_contents=False):
        self._datos = datos
        self.status_code = 200 if datos is not None else 404
        if datos is None:
            self._payload = {}
        elif formato_contents:
            crudo = json.dumps(datos, ensure_ascii=False).encode("utf-8")
            self._payload = {
                "sha": sello or "",
                "content": base64.b64encode(crudo).decode("ascii"),
                "encoding": "base64",
            }
        else:
            self._payload = datos

    def json(self):
        return self._payload

    def raise_for_status(self):
        if self.status_code != 200:
            raise ValueError(f"documento no encontrado ({self.status_code})")


def _get_doc(url, **kw):
    """Reemplazo de `requests.get(url)` para los documentos de la app.

    Con Supabase disponible, saca el nombre del documento de la URL y lo lee de
    la base; si no, hace la llamada a GitHub tal cual se hacía. Devuelve el
    formato de la Contents API cuando la URL es de esa API, porque hay
    cargadores que leen `sha` y `content` de ahí.
    """
    if _datos.disponible():
        es_contents = "/contents/" in url
        nombre = (url.split("/contents/")[-1] if es_contents
                  else url.split("/main/")[-1]).split("?")[0].strip("/")
        if nombre:
            sello, datos = _datos.leer_con_sello(nombre)
            if datos is not None:
                return _RespuestaDoc(datos, sello, formato_contents=es_contents)
            return _RespuestaDoc(None)
    return requests.get(url, **kw)


def _guardar_json_github_raw(nombre_archivo, datos_dict, mensaje_commit):
    """Guarda un JSON por su nombre. Retorna True/False."""
    if _datos.disponible():
        return _datos.guardar(nombre_archivo, datos_dict, mensaje_commit)

    url = f"https://api.github.com/repos/{GITHUB_USUARIO}/{GITHUB_REPO}/contents/{nombre_archivo}"
    try:
        sha, _ = _leer_json_github_raw(nombre_archivo)
        contenido_b64 = base64.b64encode(
            json.dumps(datos_dict, ensure_ascii=False, indent=2).encode("utf-8")
        ).decode()
        payload = {"message": mensaje_commit, "content": contenido_b64}
        if sha:
            payload["sha"] = sha
        resp = requests.put(url, headers=_github_headers(), json=payload, timeout=30, verify=False)
        return resp.status_code in [200, 201]
    except Exception:
        return False


# ============================================================
#   GESTIÓN DE USUARIOS (PBKDF2-SHA256)
# ============================================================
def _hash_pwd(password, salt=None):
    """Retorna (hash_hex, salt_hex). Si salt es None, genera uno nuevo."""
    if salt is None:
        salt = secrets.token_hex(16)
    h = hashlib.pbkdf2_hmac(
        "sha256", password.encode("utf-8"), salt.encode("utf-8"), 100_000
    ).hex()
    return h, salt


def _verificar_pwd(password, stored_hash, salt):
    h, _ = _hash_pwd(password, salt)
    return h == stored_hash


def _leer_usuarios():
    _, datos = _leer_json_github_raw(GITHUB_USUARIOS)
    return datos.get("usuarios", [])


def _guardar_usuarios(usuarios_list):
    _guardar_json_github_raw(
        GITHUB_USUARIOS,
        {"usuarios": usuarios_list},
        f"Usuarios actualizados — {ahora_chile()}",
    )


def _buscar_usuario(email, usuarios_list):
    email = email.strip().lower()
    for u in usuarios_list:
        if u.get("email", "").lower() == email:
            return u
    return None


# ============================================================
#   RESTRICCIÓN DE ACCESO POR SUCURSAL (20/07/2026)
#   -------------------------------------------------
#   Cada usuario (excepto el admin) puede quedar limitado a ver
#   solo los datos de su(s) sucursal(es) — tanto en Control y
#   Gestión Post Venta como en el Planificador de Taller y el
#   Asistente App. Se guarda en su registro de
#   usuarios_curifor.json:
#     - sucursal_home: sucursal detectada desde la nómina (informativo)
#     - sucursales_permitidas: lista de sucursales que puede ver.
#       Si la lista está vacía o la clave no existe -> SIN restricción
#       (ve todas). El admin la edita libremente desde 🛡️ Admin.
# ============================================================
def _aplicar_restriccion_nomina(u):
    """Si el usuario aun no tiene `sucursales_permitidas` definido, intenta
    detectar su sucursal desde la nomina (NOMINA_SUCURSAL_POR_EMAIL) y lo deja
    restringido a esa sola sucursal por defecto. Si su correo no aparece en la
    nomina, no se toca (queda sin restriccion hasta que el admin se la asigne
    a mano). Muta el dict `u` in-place; no hace nada si ya existe la clave
    (para no pisar una restriccion/ampliacion que el admin ya haya definido)."""
    if "sucursales_permitidas" in u:
        return
    suc = NOMINA_SUCURSAL_POR_EMAIL.get((u.get("email") or "").strip().lower())
    if suc:
        u["sucursal_home"] = suc
        u["sucursales_permitidas"] = [suc]


def _sucursales_permitidas_usuario(email, usuarios_list=None):
    """Devuelve la lista de sucursales que el usuario puede ver, o None si no
    tiene ninguna restriccion (ve todas). El admin siempre ve todas."""
    if email == ADMIN_EMAIL:
        return None
    if usuarios_list is None:
        usuarios_list = _leer_usuarios()
    u = _buscar_usuario(email, usuarios_list)
    if not u:
        return None
    permitidas = u.get("sucursales_permitidas") or []
    return permitidas if permitidas else None


def _puede_usar_planificador(email, usuarios_list=None):
    """
    Acceso al modulo 'Planificador de Taller': siempre permitido para el
    admin (ADMIN_EMAIL); para el resto, si su registro en
    usuarios_curifor.json tiene el flag 'puede_planificador' en True
    (gestionable desde Admin -> Usuarios) O el flag 'puede_confirmar_citas'
    (23/07/2026 — permiso limitado para asesores: entran al modulo pero solo
    pueden usar los botones Asiste/No Asiste/Reagenda de la columna "Citas",
    ver _puede_confirmar_citas() mas abajo).
    """
    email = (email or "").strip().lower()
    if email == ADMIN_EMAIL:
        return True
    if usuarios_list is None:
        usuarios_list = _leer_usuarios()
    _u = _buscar_usuario(email, usuarios_list)
    return bool(_u and (_u.get("puede_planificador", False) or _u.get("puede_confirmar_citas", False)))


def _puede_editar_planificador(email, usuarios_list=None):
    """
    Permiso de EDICIÓN dentro del Planificador de Taller (mover tarjetas del
    JPCB, asignar técnicos/horarios, editar Control de Taller, cerrar
    citas, etc.) — distinto del acceso de solo VER el módulo
    (_puede_usar_planificador). Siempre permitido para el admin; para el
    resto, solo si su registro en usuarios_curifor.json tiene el flag
    'puede_editar_planificador' en True (gestionable desde Admin -> Usuarios
    -> Permisos de módulos). 10/07/2026: antes esto estaba fijo en código
    (solo el admin podía editar, sin ningún control desde Admin) — a pedido
    de Cristóbal se hizo administrable como el resto de los permisos.
    """
    email = (email or "").strip().lower()
    if email == ADMIN_EMAIL:
        return True
    if usuarios_list is None:
        usuarios_list = _leer_usuarios()
    _u = _buscar_usuario(email, usuarios_list)
    return bool(_u and _u.get("puede_editar_planificador", False))


def _puede_confirmar_citas(email, usuarios_list=None):
    """
    Permiso LIMITADO dentro del Planificador de Taller (23/07/2026, a pedido
    de Cristobal para el listado de asesores de "Asesores.xlsx"): solo puede
    usar los 3 botones "Asiste"/"No Asiste"/"Reagenda" de la columna "Citas
    <fecha>" del JPCB (confirmar si el vehiculo agendado llego o no) — NO
    incluye el resto de _puede_editar_planificador (no puede arrastrar
    tarjetas, asignar tecnicos/horarios ni editar Control de Taller). Siempre
    permitido para el admin; para el resto, solo si su registro en
    usuarios_curifor.json tiene el flag 'puede_confirmar_citas' en True
    (gestionable desde Admin -> Usuarios -> Permisos de modulos, o via el
    boton de carga masiva para ASESORES_CONFIRMAR_CITAS).
    """
    email = (email or "").strip().lower()
    if email == ADMIN_EMAIL:
        return True
    if usuarios_list is None:
        usuarios_list = _leer_usuarios()
    _u = _buscar_usuario(email, usuarios_list)
    return bool(_u and _u.get("puede_confirmar_citas", False))


def _puede_disponibilidad_tecnicos(email, usuarios_list=None):
    """
    Permiso LIMITADO dentro de Produccion Tecnicos (29/07/2026, a pedido de
    Cristobal — pensado para Torre de Control): marcar a un tecnico como NO
    DISPONIBLE en un rango de fechas (vacaciones, licencia medica, permiso,
    capacitacion), para que esos dias habiles no cuenten como horas
    disponibles al calcular su % de productividad. NO habilita nada mas del
    Planificador. Siempre permitido para el admin; para el resto, solo si su
    registro en usuarios_curifor.json tiene el flag
    'puede_disponibilidad_tecnicos' en True (gestionable desde Admin ->
    Usuarios -> Permisos de modulos).
    """
    email = (email or "").strip().lower()
    if email == ADMIN_EMAIL:
        return True
    if usuarios_list is None:
        usuarios_list = _leer_usuarios()
    _u = _buscar_usuario(email, usuarios_list)
    return bool(_u and _u.get("puede_disponibilidad_tecnicos", False))


def _puede_usar_prepicking(email, usuarios_list=None):
    """
    Acceso a la pestaña 'Pre-picking' dentro del Planificador de Taller —
    permiso INDEPENDIENTE de _puede_usar_planificador (13/07/2026, a pedido de
    Cristóbal): alguien puede tener acceso al Planificador completo sin ver
    Pre-picking, y viceversa no aplica (Pre-picking vive dentro del módulo, así
    que igual requiere haber entrado al Planificador). Siempre permitido para
    el admin; para el resto, solo si su registro en usuarios_curifor.json
    tiene el flag 'puede_prepicking' en True (Admin -> Usuarios -> Permisos
    de módulos).
    """
    email = (email or "").strip().lower()
    if email == ADMIN_EMAIL:
        return True
    if usuarios_list is None:
        usuarios_list = _leer_usuarios()
    _u = _buscar_usuario(email, usuarios_list)
    return bool(_u and _u.get("puede_prepicking", False))


def _puede_usar_indicadores(email, usuarios_list=None):
    """
    Acceso al modulo 'Indicadores Post Venta' (Power BI embebido):
    siempre permitido para el admin (ADMIN_EMAIL); para el resto, solo
    si su registro en usuarios_curifor.json tiene el flag
    'puede_indicadores' en True (gestionable desde Admin -> Usuarios).
    """
    email = (email or "").strip().lower()
    if email == ADMIN_EMAIL:
        return True
    if usuarios_list is None:
        usuarios_list = _leer_usuarios()
    _u = _buscar_usuario(email, usuarios_list)
    return bool(_u and _u.get("puede_indicadores", False))


def _puede_usar_control(email, usuarios_list=None):
    """
    Acceso al modulo 'Control y Gestion Post Venta' (dashboard principal):
    siempre permitido para el admin (ADMIN_EMAIL); para el resto, solo si
    su registro en usuarios_curifor.json tiene el flag 'puede_control' en
    True (gestionable desde Admin -> Usuarios -> Permisos de modulos).
    """
    email = (email or "").strip().lower()
    if email == ADMIN_EMAIL:
        return True
    if usuarios_list is None:
        usuarios_list = _leer_usuarios()
    _u = _buscar_usuario(email, usuarios_list)
    return bool(_u and _u.get("puede_control", False))


def _puede_usar_asistente_app(email, usuarios_list=None):
    """
    Acceso al modulo 'Asistente App' (busqueda rapida por patente/folio,
    sin IA — solo cruce de datos propio): siempre permitido para el admin
    (ADMIN_EMAIL); para el resto, solo si su registro en
    usuarios_curifor.json tiene el flag 'puede_asistente_app' en True
    (gestionable desde Admin -> Usuarios -> Permisos de modulos).
    Bloqueado por defecto — 08/07/2026.
    """
    email = (email or "").strip().lower()
    if email == ADMIN_EMAIL:
        return True
    if usuarios_list is None:
        usuarios_list = _leer_usuarios()
    _u = _buscar_usuario(email, usuarios_list)
    return bool(_u and _u.get("puede_asistente_app", False))


def _puede_usar_cotizador(email, usuarios_list=None):
    """
    Acceso al modulo 'Cotizador de Mantenciones' (cotiza mantenciones
    preventivas por marca/modelo/version/anio con precios, repuestos,
    stock, adicionales y packs — plataforma web embebida): siempre
    permitido para el admin (ADMIN_EMAIL); para el resto, solo si su
    registro en usuarios_curifor.json tiene el flag 'puede_cotizador' en
    True (gestionable desde Admin -> Usuarios -> Permisos de modulos).
    Bloqueado por defecto — 21/07/2026.
    """
    email = (email or "").strip().lower()
    if email == ADMIN_EMAIL:
        return True
    if usuarios_list is None:
        usuarios_list = _leer_usuarios()
    _u = _buscar_usuario(email, usuarios_list)
    return bool(_u and _u.get("puede_cotizador", False))


def _puede_usar_campanas(email, usuarios_list=None):
    """
    Acceso al modulo 'Revision de Campañas' (datos de la Agenda Ford — casos
    con campaña/boletín pendiente o vencida): siempre permitido para el admin
    (ADMIN_EMAIL); para el resto, solo si su registro en usuarios_curifor.json
    tiene el flag 'puede_campanas' en True (gestionable desde Admin ->
    Usuarios -> Permisos de modulos). Bloqueado por defecto — 28/07/2026.
    """
    email = (email or "").strip().lower()
    if email == ADMIN_EMAIL:
        return True
    if usuarios_list is None:
        usuarios_list = _leer_usuarios()
    _u = _buscar_usuario(email, usuarios_list)
    return bool(_u and _u.get("puede_campanas", False))


def _puede_usar_cuenta_ficha(email, usuarios_list=None):
    """
    Acceso al modulo 'Cuenta Ficha' (saldos disponibles del cliente en su
    cuenta corriente + historial completo de sus OT con todos los documentos
    posteriores): siempre permitido para el admin (ADMIN_EMAIL); para el
    resto, solo si su registro en usuarios_curifor.json tiene el flag
    'puede_cuenta_ficha' en True (gestionable desde Admin -> Usuarios ->
    Permisos de modulos). Bloqueado por defecto — 31/07/2026.
    """
    email = (email or "").strip().lower()
    if email == ADMIN_EMAIL:
        return True
    if usuarios_list is None:
        usuarios_list = _leer_usuarios()
    _u = _buscar_usuario(email, usuarios_list)
    return bool(_u and _u.get("puede_cuenta_ficha", False))


def _puede_usar_informes_gestion(email, usuarios_list=None):
    """
    Acceso al modulo 'Informes de Gestion' (reportes de gestion por marca que
    Cristobal envia a las automotoras: AG con sus hojas HYU/FOR/D&P por
    sucursal, y el IMOP de Ford mes a mes): siempre permitido para el admin
    (ADMIN_EMAIL); para el resto, solo si su registro en usuarios_curifor.json
    tiene el flag 'puede_informes_gestion' en True (gestionable desde Admin ->
    Usuarios -> Permisos de modulos). Bloqueado por defecto — 04/08/2026.
    """
    email = (email or "").strip().lower()
    if email == ADMIN_EMAIL:
        return True
    if usuarios_list is None:
        usuarios_list = _leer_usuarios()
    _u = _buscar_usuario(email, usuarios_list)
    return bool(_u and _u.get("puede_informes_gestion", False))


def _puede_usar_loaners(email, usuarios_list=None):
    """
    Acceso al modulo 'Loaners' (flota de vehiculos de cortesia que se prestan
    al cliente mientras su unidad esta en taller — reemplaza al Excel
    "LOANERS 2.xlsx"): siempre permitido para el admin (ADMIN_EMAIL); para el
    resto, solo si su registro en usuarios_curifor.json tiene el flag
    'puede_loaners' en True (gestionable desde Admin -> Usuarios -> Permisos
    de modulos). Bloqueado por defecto — 06/08/2026.
    """
    email = (email or "").strip().lower()
    if email == ADMIN_EMAIL:
        return True
    if usuarios_list is None:
        usuarios_list = _leer_usuarios()
    _u = _buscar_usuario(email, usuarios_list)
    return bool(_u and _u.get("puede_loaners", False))


def _puede_usar_agenda_taller(email, usuarios_list=None):
    """
    Acceso al modulo 'Agenda de Taller' (plataforma nueva de agendamiento de
    citas de mantencion por sucursal/dia/hora, embebida por iframe — reemplaza
    a agenda.curifor.cl): siempre permitido para el admin (ADMIN_EMAIL); para
    el resto, solo si su registro en usuarios_curifor.json tiene el flag
    'puede_agenda_taller' en True (gestionable desde Admin -> Usuarios ->
    Permisos de modulos). Bloqueado por defecto — 29/07/2026.
    """
    email = (email or "").strip().lower()
    if email == ADMIN_EMAIL:
        return True
    if usuarios_list is None:
        usuarios_list = _leer_usuarios()
    _u = _buscar_usuario(email, usuarios_list)
    return bool(_u and _u.get("puede_agenda_taller", False))


def _puede_usar_recepcion(email, usuarios_list=None):
    """
    Acceso al modulo 'Recepcion de Vehiculos' (checklist de accesorios,
    inspeccion, firma del cliente e ingreso a taller, embebido por iframe —
    reemplaza a la recepcion anterior): siempre permitido para el admin
    (ADMIN_EMAIL); para el resto, solo si su registro en usuarios_curifor.json
    tiene el flag 'puede_recepcion' en True (gestionable desde Admin ->
    Usuarios -> Permisos de modulos). Bloqueado por defecto — 29/07/2026.
    """
    email = (email or "").strip().lower()
    if email == ADMIN_EMAIL:
        return True
    if usuarios_list is None:
        usuarios_list = _leer_usuarios()
    _u = _buscar_usuario(email, usuarios_list)
    return bool(_u and _u.get("puede_recepcion", False))


# ============================================================
#   NOTIFICACIONES
# ============================================================
def _leer_notificaciones():
    _, datos = _leer_json_github_raw(GITHUB_NOTIFICACIONES)
    return datos.get("notificaciones", [])


def _guardar_notificaciones(notifs):
    _guardar_json_github_raw(
        GITHUB_NOTIFICACIONES,
        {"notificaciones": notifs},
        f"Notificaciones — {ahora_chile()}",
    )


def _crear_notificacion(remitente, destinatario, folio_ot, extracto):
    """Agrega una notificación no leída para el destinatario."""
    notifs = _leer_notificaciones()
    notifs.append({
        "id":           str(uuid.uuid4()),
        "remitente":    remitente,
        "destinatario": destinatario,
        "folio_ot":     folio_ot,
        "extracto":     extracto[:200],
        "fecha":        ahora_chile(),
        "leida":        False,
    })
    _guardar_notificaciones(notifs)


def _marcar_leidas(email):
    """Marca como leídas todas las notificaciones del usuario."""
    notifs = _leer_notificaciones()
    changed = False
    for n in notifs:
        if n.get("destinatario", "").lower() == email.lower() and not n.get("leida"):
            n["leida"] = True
            changed = True
    if changed:
        _guardar_notificaciones(notifs)


# ============================================================
#   AUDIT LOG
# ============================================================
def _registrar_audit(usuario, accion, detalle, folio_ot=""):
    """Append-only audit log con ventana deslizante de 2000 entradas."""
    try:
        _, datos = _leer_json_github_raw(GITHUB_AUDIT)
        registros = datos.get("registros", [])
        registros.append({
            "fecha":    ahora_chile(),
            "usuario":  usuario,
            "accion":   accion,
            "detalle":  detalle,
            "folio_ot": folio_ot,
        })
        # Ventana deslizante
        if len(registros) > 2000:
            registros = registros[-2000:]
        _guardar_json_github_raw(
            GITHUB_AUDIT,
            {"registros": registros},
            f"Audit: {accion} — {usuario}",
        )
    except Exception:
        pass  # El audit nunca debe bloquear la operación principal


def _leer_audit():
    """Lee el historial de auditoría desde GitHub."""
    try:
        _, datos = _leer_json_github_raw(GITHUB_AUDIT)
        return datos.get("registros", [])
    except Exception:
        return []


# ============================================================
#   PRESENCIA EN LÍNEA (heartbeat)
# ============================================================
def _actualizar_heartbeat(email: str):
    """
    Escribe/actualiza el timestamp del usuario en online_users.json.
    Con debounce de 5 minutos para no saturar la API de GitHub — esto corre a
    nivel de modulo, en CADA rerun de CADA usuario y CADA pestana/modo de la
    app (antes del if de modo), asi que cada llamada bloqueante de red que
    haga cuesta tiempo de carga para todos. Antes: debounce de 90s + la version
    generica de _guardar_json_github_raw() (que internamente vuelve a leer el
    archivo solo para sacar el sha) sumaban 3 llamadas HTTP secuenciales por
    heartbeat, con timeouts de hasta 15/15/30s cada una — un peor caso de mas
    de 40s de app "colgada" para todos, ademas de sumarse a la lentitud
    reportada en cada rerun (15/07/2026). Ahora: debounce mas largo (5 min en
    vez de 90s, la presencia no necesita mas precision que esa) + solo 2
    llamadas (lee sha+datos una vez, reutiliza ese mismo sha al escribir en
    vez de volver a leerlo) + timeouts cortos para no dejar la app entera
    esperando si GitHub responde lento en un momento puntual.
    """
    import time as _time
    ahora_ts = _time.time()
    ultima = st.session_state.get("_last_heartbeat_ts", 0)
    if ahora_ts - ultima < 300:
        return  # Demasiado pronto, saltamos

    # Con Supabase el latido es una lectura y una escritura a la base, no dos
    # llamadas a la API de GitHub que ademas dejaban un commit por usuario cada
    # cinco minutos. Ese ruido de commits desaparece.
    if _datos.disponible():
        try:
            datos = _datos.leer(GITHUB_ONLINE) or {}
            usuarios = datos.get("usuarios", {})
            usuarios[email] = {"last_seen": ahora_chile(), "last_seen_ts": int(ahora_ts)}
            usuarios = {k: v for k, v in usuarios.items()
                        if ahora_ts - v.get("last_seen_ts", 0) < 86400}
            _datos.guardar(GITHUB_ONLINE, {"usuarios": usuarios}, f"heartbeat {email}")
            st.session_state["_last_heartbeat_ts"] = ahora_ts
        except Exception:
            pass          # la presencia nunca puede frenar la app
        return

    try:
        url = f"https://api.github.com/repos/{GITHUB_USUARIO}/{GITHUB_REPO}/contents/{GITHUB_ONLINE}"
        hdrs = _github_headers()
        r = requests.get(url, headers=hdrs, timeout=6, verify=False)
        sha = None
        datos = {}
        if r.status_code == 200:
            info = r.json()
            sha = info.get("sha")
            raw = (info.get("content") or "").replace("\n", "").strip()
            if raw:
                datos = json.loads(base64.b64decode(raw).decode("utf-8"))
        usuarios = datos.get("usuarios", {})
        usuarios[email] = {
            "last_seen":    ahora_chile(),
            "last_seen_ts": int(ahora_ts),
        }
        # Limpiar entradas muy antiguas (>24h) para no crecer indefinidamente
        usuarios = {
            k: v for k, v in usuarios.items()
            if ahora_ts - v.get("last_seen_ts", 0) < 86400
        }
        contenido_b64 = base64.b64encode(
            json.dumps({"usuarios": usuarios}, ensure_ascii=False, indent=2).encode("utf-8")
        ).decode()
        payload = {"message": f"heartbeat {email}", "content": contenido_b64}
        if sha:
            payload["sha"] = sha
        requests.put(url, headers=hdrs, json=payload, timeout=6, verify=False)
        st.session_state["_last_heartbeat_ts"] = ahora_ts
    except Exception:
        pass  # Nunca bloquear la app por esto


def _leer_online_users(ventana_seg: int = 300) -> list[dict]:
    """
    Devuelve la lista de usuarios activos en los últimos `ventana_seg` segundos.
    """
    import time as _time
    try:
        _, datos = _leer_json_github_raw(GITHUB_ONLINE)
        usuarios = datos.get("usuarios", {})
        ahora_ts = _time.time()
        activos = []
        for email, info in usuarios.items():
            delta = ahora_ts - info.get("last_seen_ts", 0)
            if delta <= ventana_seg:
                activos.append({
                    "email":      email,
                    "last_seen":  info.get("last_seen", "—"),
                    "hace":       int(delta),
                })
        activos.sort(key=lambda x: x["hace"])
        return activos
    except Exception:
        return []


# ============================================================
#   AUTENTICACIÓN POR USUARIO
# ============================================================
def check_password():
    if st.session_state.get("authenticated"):
        return

    step = st.session_state.get("login_step", 1)

    st.markdown(f"""
    <div style="max-width:420px; margin:60px auto 0; text-align:center;">
        <div style="background:#ffffff; border-radius:10px; padding:10px 20px;
                    display:inline-block; margin-bottom:1.2rem;
                    box-shadow:0 2px 12px rgba(0,0,0,0.12);">
            <img src="{LOGO_DATA_URI}" style="max-width:180px; height:44px;
                 object-fit:contain; display:block;" />
        </div>
        <h2 style="color:var(--text-color); margin-top:0; font-weight:700;">Control y Gestión Post Venta</h2>
        <h3 style="color:var(--text-color); margin-top:0; font-weight:400; opacity:0.7;">Curifor S.A</h3>
        <p style="color:var(--text-color); font-size:0.78rem; opacity:0.4; margin:0 0 1.5rem 0;
                  font-style:italic;">Desarrollado por: Cristóbal Jerez J.</p>
    </div>
    """, unsafe_allow_html=True)

    col_l, col_c, col_r = st.columns([1, 2, 1])
    with col_c:
        # ── Paso 1: Ingresar correo ──
        if step == 1:
            st.markdown("**Ingresa tu correo institucional:**")
            email_in = st.text_input("Correo", key="li_email",
                                     placeholder="tucorreo@curifor.com")
            if st.button("Continuar →", use_container_width=True, type="primary"):
                email_in = email_in.strip().lower()
                if not email_in.endswith(DOMINIO_PERMITIDO):
                    st.error(f"Solo se permiten correos {DOMINIO_PERMITIDO}")
                else:
                    usuarios = _leer_usuarios()
                    u = _buscar_usuario(email_in, usuarios)
                    st.session_state["_li_email"] = email_in
                    st.session_state["_li_usuarios"] = usuarios
                    if u and u.get("password_hash"):
                        st.session_state["login_step"] = 2   # usuario existente
                    else:
                        st.session_state["login_step"] = 3   # usuario nuevo
                    st.rerun()
            st.caption(f"Solo cuentas terminadas en {DOMINIO_PERMITIDO}")

        # ── Paso 2: Usuario existente — ingresar contraseña ──
        elif step == 2:
            email_in = st.session_state.get("_li_email", "")
            st.info(f"👤 {email_in}")
            pwd = st.text_input("Contraseña", type="password", key="li_pwd")
            c1, c2 = st.columns(2)
            with c1:
                if st.button("← Volver", use_container_width=True):
                    st.session_state["login_step"] = 1
                    st.rerun()
            with c2:
                if st.button("Ingresar", use_container_width=True, type="primary"):
                    usuarios = st.session_state.get("_li_usuarios", _leer_usuarios())
                    u = _buscar_usuario(email_in, usuarios)
                    if u and _verificar_pwd(pwd, u["password_hash"], u["salt"]):
                        _era_temp = bool(u.get("temp_pwd", False))
                        # Actualizar ultimo_login
                        for uu in usuarios:
                            if uu.get("email", "").lower() == email_in:
                                uu["ultimo_login"] = ahora_chile()
                                _aplicar_restriccion_nomina(uu)
                                break
                        _guardar_usuarios(usuarios)
                        _registrar_audit(email_in, "LOGIN", "Login exitoso")
                        if _era_temp:
                            # Contraseña provisoria (asignada por el admin) —
                            # obliga a crear una contraseña nueva y permanente
                            # antes de dejarlo entrar (paso 4).
                            st.session_state["_li_usuarios"] = usuarios
                            st.session_state["login_step"] = 4
                            st.rerun()
                        else:
                            st.session_state.authenticated = True
                            st.session_state.user_email    = email_in
                            st.session_state.login_step    = 1
                            st.session_state.pop("_li_email", None)
                            st.session_state.pop("_li_usuarios", None)
                            st.rerun()
                    else:
                        _registrar_audit(email_in, "LOGIN_FAIL", "Contraseña incorrecta")
                        st.error("Contraseña incorrecta.")

        # ── Paso 3: Usuario nuevo — crear contraseña ──
        elif step == 3:
            email_in = st.session_state.get("_li_email", "")
            st.info(f"👤 {email_in}")
            st.success("✅ Primera vez — crea tu contraseña personal:")
            pwd1 = st.text_input("Nueva contraseña", type="password", key="li_pwd1")
            pwd2 = st.text_input("Confirmar contraseña", type="password", key="li_pwd2")
            c1, c2 = st.columns(2)
            with c1:
                if st.button("← Volver", use_container_width=True):
                    st.session_state["login_step"] = 1
                    st.rerun()
            with c2:
                if st.button("Crear y entrar", use_container_width=True, type="primary"):
                    if len(pwd1) < 6:
                        st.error("La contraseña debe tener al menos 6 caracteres.")
                    elif pwd1 != pwd2:
                        st.error("Las contraseñas no coinciden.")
                    else:
                        h, s = _hash_pwd(pwd1)
                        usuarios = st.session_state.get("_li_usuarios", _leer_usuarios())
                        nombre = email_in.split("@")[0]
                        # Actualizar si existe sin hash, o agregar nuevo
                        u_existente = _buscar_usuario(email_in, usuarios)
                        if u_existente:
                            u_existente["password_hash"] = h
                            u_existente["salt"] = s
                            u_existente["ultimo_login"] = ahora_chile()
                            _aplicar_restriccion_nomina(u_existente)
                        else:
                            nuevo_u = {
                                "email":         email_in,
                                "nombre":        nombre,
                                "password_hash": h,
                                "salt":          s,
                                "activo":        True,
                                "creado":        ahora_chile(),
                                "ultimo_login":  ahora_chile(),
                                "temp_pwd":      False,
                            }
                            _aplicar_restriccion_nomina(nuevo_u)
                            usuarios.append(nuevo_u)
                        _guardar_usuarios(usuarios)
                        _registrar_audit(email_in, "REGISTRO", "Cuenta creada")
                        st.session_state.authenticated = True
                        st.session_state.user_email    = email_in
                        st.session_state.login_step    = 1
                        st.session_state.pop("_li_email", None)
                        st.session_state.pop("_li_usuarios", None)
                        st.rerun()

        # ── Paso 4: Contraseña provisoria — obligatorio crear una nueva ──
        elif step == 4:
            email_in = st.session_state.get("_li_email", "")
            st.info(f"👤 {email_in}")
            st.warning(
                "🔑 Iniciaste sesión con una contraseña **provisoria**. "
                "Antes de continuar, crea tu contraseña nueva y permanente:"
            )
            pwd1 = st.text_input("Nueva contraseña", type="password", key="li_pwd1_temp")
            pwd2 = st.text_input("Confirmar contraseña", type="password", key="li_pwd2_temp")
            if st.button("Crear contraseña y entrar", use_container_width=True, type="primary"):
                if len(pwd1) < 6:
                    st.error("La contraseña debe tener al menos 6 caracteres.")
                elif pwd1 != pwd2:
                    st.error("Las contraseñas no coinciden.")
                else:
                    h, s = _hash_pwd(pwd1)
                    usuarios = st.session_state.get("_li_usuarios", _leer_usuarios())
                    u_existente = _buscar_usuario(email_in, usuarios)
                    if u_existente:
                        u_existente["password_hash"] = h
                        u_existente["salt"] = s
                        u_existente["temp_pwd"] = False
                        _guardar_usuarios(usuarios)
                        _registrar_audit(email_in, "CAMBIO_PWD", "Contraseña provisoria reemplazada por una nueva")
                    st.session_state.authenticated = True
                    st.session_state.user_email    = email_in
                    st.session_state.login_step    = 1
                    st.session_state.pop("_li_email", None)
                    st.session_state.pop("_li_usuarios", None)
                    st.rerun()

    st.stop()


check_password()
usuario_activo = st.session_state.get("user_email", "usuario@curifor.com")
# Heartbeat — registra presencia del usuario (debounce 90s, nunca bloquea)
_actualizar_heartbeat(usuario_activo)


# ============================================================
#   CSS — DISEÑO CORPORATIVO CURIFOR (light + dark adaptive)
# ============================================================
st.markdown("""
<style>
/* ── Variables de tema ───────────────────────────────────── */
/* Streamlit expone --background-color, --secondary-background-color,
   --text-color, --primary-color en :root según el tema activo.    */

/* ── KPI Cards — gradiente + sombra, mismo lenguaje que Cuenta Ficha ── */
.kpi-box {
    background: linear-gradient(135deg, #0d2f5a 0%, #1a4f8a 100%);
    border-radius: 14px;
    padding: 18px 14px 16px;
    text-align: center;
    border: none;
    box-shadow: 0 6px 18px rgba(13,47,90,0.22);
    transition: box-shadow 0.2s, transform 0.15s;
    position: relative;
}
.kpi-box:hover { box-shadow: 0 10px 26px rgba(13,47,90,0.32); transform: translateY(-2px); }
.kpi-box.rojo    { background: linear-gradient(135deg,#7a1a20 0%,#c0392b 100%); box-shadow:0 6px 18px rgba(192,57,43,.24); }
.kpi-box.naranja { background: linear-gradient(135deg,#8a4a06 0%,#dd6b20 100%); box-shadow:0 6px 18px rgba(221,107,32,.24); }
.kpi-box.amarillo{ background: linear-gradient(135deg,#8a6300 0%,#d99100 100%); box-shadow:0 6px 18px rgba(217,145,0,.24); }
.kpi-box.verde   { background: linear-gradient(135deg,#0b6b3a 0%,#18a05a 100%); box-shadow:0 6px 18px rgba(11,107,58,.24); }
.kpi-box.azul    { background: linear-gradient(135deg,#0d3c73 0%,#2f7edb 100%); box-shadow:0 6px 18px rgba(47,126,219,.24); }
.kpi-icono { font-size: 1.25rem; margin-bottom: 4px; opacity: .92; }
.kpi-num {
    font-size: 2.1rem; font-weight: 800;
    color: #ffffff;
    margin: 0; letter-spacing: -1px; line-height: 1.1;
}
.kpi-num.rojo, .kpi-num.naranja, .kpi-num.amarillo,
.kpi-num.verde, .kpi-num.azul { color: #ffffff; }
.kpi-label {
    font-size: 0.66rem;
    color: #ffffff;
    opacity: 0.85;
    margin-top: 6px;
    font-weight: 700;
    letter-spacing: 0.07em;
    text-transform: uppercase;
}

/* ── Chips y estados vacíos reutilizables (mismo lenguaje que Cuenta Ficha) */
.info-chip{display:inline-block;font-size:.72rem;font-weight:600;
    padding:3px 11px;border-radius:20px;margin-right:4px;
    background:rgba(74,122,181,.14);color:#2b5ea7;white-space:nowrap;}
.info-chip.ok  { background:rgba(34,197,94,.15);  color:#15803d; }
.info-chip.warn{ background:rgba(239,68,68,.14);   color:#b91c1c; }
.info-chip.mut { background:rgba(128,128,128,.16); color:#6b7280; }
.empty-state{border:1px dashed rgba(128,128,128,.4);border-radius:12px;
    padding:26px;text-align:center;color:var(--text-color);opacity:.55;}

/* ── Section titles ──────────────────────────────────────── */
.section-title {
    font-size: 0.75rem;
    font-weight: 700;
    color: var(--text-color);
    opacity: 0.85;
    text-transform: uppercase;
    letter-spacing: 0.08em;
    border-bottom: 1px solid rgba(128,128,128,0.18);
    padding: 0 0 7px 10px;
    margin-bottom: 14px;
    border-left: 3px solid #4a7ab5;
    line-height: 1.4;
}

/* ── Header corporativo — siempre oscuro (branding) ─────── */
.curifor-header {
    background: linear-gradient(135deg, #0c243d 0%, #1a3a5c 55%, #1e4a72 100%);
    border-radius: 12px;
    padding: 16px 22px;
    display: flex;
    align-items: center;
    gap: 18px;
    margin-bottom: 8px;
    box-shadow: 0 4px 18px rgba(0,0,0,0.22);
}
.logo-pill {
    background: #ffffff;
    border-radius: 8px;
    padding: 8px 14px;
    display: flex;
    align-items: center;
    justify-content: center;
    flex-shrink: 0;
    box-shadow: 0 2px 8px rgba(0,0,0,0.18);
}
.logo-pill img { height: 38px; display: block; }
.curifor-header-text h2 {
    margin: 0;
    color: #f0f6ff;
    font-size: 1.2rem;
    font-weight: 700;
    letter-spacing: -0.02em;
}
.curifor-header-text p {
    margin: 2px 0 0;
    color: rgba(200,220,255,0.55);
    font-size: 0.73rem;
}
.curifor-header-text .dev-credit {
    margin: 5px 0 0;
    color: rgba(200,220,255,0.38);
    font-size: 0.65rem;
    font-style: italic;
}
.curifor-badge {
    margin-left: auto;
    background: rgba(255,255,255,0.1);
    border: 1px solid rgba(255,255,255,0.18);
    border-radius: 20px;
    padding: 4px 12px;
    color: rgba(200,220,255,0.8);
    font-size: 0.7rem;
    font-weight: 500;
    white-space: nowrap;
}

/* ── Comentario cards ────────────────────────────────────── */
.comentario-card {
    background: var(--secondary-background-color);
    border: 1px solid rgba(128,128,128,0.15);
    border-left: 4px solid #4a7ab5;
    border-radius: 8px;
    padding: 12px 16px;
    margin-bottom: 10px;
    box-shadow: 0 1px 5px rgba(0,0,0,0.07);
}
.comentario-meta {
    font-size: 0.72rem;
    color: var(--text-color);
    opacity: 0.45;
    margin-bottom: 5px;
    font-weight: 500;
}
.comentario-texto {
    font-size: 0.88rem;
    color: var(--text-color);
    line-height: 1.55;
}

/* ── Notificación cards ──────────────────────────────────── */
.notif-card {
    background: var(--secondary-background-color);
    border: 1px solid rgba(249,115,22,0.25);
    border-left: 4px solid #f97316;
    border-radius: 8px;
    padding: 12px 16px;
    margin-bottom: 8px;
    box-shadow: 0 1px 5px rgba(0,0,0,0.07);
}
.notif-card.leida {
    border-color: rgba(128,128,128,0.15);
    border-left-color: rgba(128,128,128,0.35);
    opacity: 0.65;
}

/* ── Radio navegación — ocultar burbuja, estilo chip ─────── */
div[data-testid="stRadio"] > label { display: none !important; }
div[data-testid="stRadio"] [role="radiogroup"] {
    gap: 4px !important;
    flex-wrap: wrap !important;
}
div[data-testid="stRadio"] [role="radiogroup"] > label {
    background: var(--secondary-background-color) !important;
    border: 1px solid rgba(128,128,128,0.25) !important;
    border-radius: 7px !important;
    padding: 5px 11px !important;
    font-size: 0.78rem !important;
    font-weight: 500 !important;
    color: var(--text-color) !important;
    opacity: 0.7;
    cursor: pointer;
    transition: opacity 0.15s, border-color 0.15s;
}
div[data-testid="stRadio"] [role="radiogroup"] > label:hover {
    opacity: 1 !important;
    border-color: #4a7ab5 !important;
}
/* Seleccionado */
div[data-testid="stRadio"] [role="radiogroup"] > label:has(input:checked) {
    background: #1a3a5c !important;
    color: #ffffff !important;
    opacity: 1 !important;
    border-color: #1a3a5c !important;
    font-weight: 600 !important;
}
/* Ocultar burbuja del radio */
div[data-testid="stRadio"] [role="radiogroup"] > label > div:first-child {
    display: none !important;
}

/* ── Botones primarios ───────────────────────────────────── */
button[kind="primary"] {
    background: linear-gradient(135deg, #1a3a5c, #2b5ea7) !important;
    border: none !important;
    border-radius: 8px !important;
    font-weight: 600 !important;
    box-shadow: 0 2px 8px rgba(26,58,92,0.3) !important;
}
button[kind="primary"]:hover {
    box-shadow: 0 4px 14px rgba(26,58,92,0.4) !important;
    transform: translateY(-1px);
}

/* ── Métricas nativas ────────────────────────────────────── */
div[data-testid="stMetricValue"] {
    font-size: 1.25rem !important;
    font-weight: 700 !important;
}
div[data-testid="stMetricLabel"] {
    font-size: 0.72rem !important;
    font-weight: 600 !important;
    text-transform: uppercase;
    letter-spacing: 0.05em;
    opacity: 0.6;
}

/* ── Alerts ──────────────────────────────────────────────── */
div[data-testid="stAlert"] {
    border-radius: 8px !important;
    font-size: 0.84rem !important;
}
</style>
""", unsafe_allow_html=True)


# ============================================================
#   GENERADOR EXCEL FACTURAS X
# ============================================================
def _generar_excel_fx(df_fx, total_fact, total_costo, fecha_str):
    """Genera Excel con formato amigable para gestión de Facturas X."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = "Facturas X"

    AZUL_OSC = "1A3A5C"
    GRIS_FILA = "F2F6FA"
    AZUL_HDR2 = "D9E6F2"

    font_titulo  = Font(name="Arial", bold=True, size=13, color="FFFFFF")
    font_resumen = Font(name="Arial", bold=True, size=10, color=AZUL_OSC)
    font_header  = Font(name="Arial", bold=True, size=10, color="FFFFFF")
    font_data    = Font(name="Arial", size=10)
    font_total   = Font(name="Arial", bold=True, size=10)

    fill_titulo = PatternFill("solid", fgColor=AZUL_OSC)
    fill_header = PatternFill("solid", fgColor=AZUL_OSC)
    fill_resumen = PatternFill("solid", fgColor=AZUL_HDR2)
    fill_gris   = PatternFill("solid", fgColor=GRIS_FILA)
    fill_total  = PatternFill("solid", fgColor=AZUL_HDR2)

    aln_center = Alignment(horizontal="center", vertical="center")
    aln_left   = Alignment(horizontal="left",   vertical="center")
    aln_right  = Alignment(horizontal="right",  vertical="center")

    borde = Border(bottom=Side(style="thin", color="CCCCCC"))

    COLS = list(df_fx.columns)
    N = len(COLS)
    FMT_PESO = '$#,##0'
    NUM_COLS = {"Total OT $", "Costo Vale Consumo"}

    # Fila 1 — Título
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=N)
    c = ws["A1"]
    c.value     = f"Facturas X — Curifor S.A  |  {fecha_str}"
    c.font      = font_titulo
    c.fill      = fill_titulo
    c.alignment = aln_center
    ws.row_dimensions[1].height = 28

    # Fila 2 — Resumen
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=N)
    c = ws["A2"]
    c.value     = f"Total: {total_fact:,} facturas     |     Costo total Vale Consumo: ${total_costo:,.0f}"
    c.font      = font_resumen
    c.fill      = fill_resumen
    c.alignment = aln_center
    ws.row_dimensions[2].height = 18

    # Fila 3 — separación
    ws.row_dimensions[3].height = 6

    # Fila 4 — Encabezados
    HDR = 4
    for ci, col in enumerate(COLS, 1):
        c = ws.cell(row=HDR, column=ci, value=col)
        c.font = font_header; c.fill = fill_header; c.alignment = aln_center
    ws.row_dimensions[HDR].height = 20

    # Datos
    for ri, (_, row) in enumerate(df_fx.iterrows(), start=HDR + 1):
        fill_row = fill_gris if ri % 2 == 0 else None
        for ci, col in enumerate(COLS, 1):
            val = row[col]
            c = ws.cell(row=ri, column=ci)
            if col in NUM_COLS:
                try: c.value = float(val)
                except: c.value = 0
                c.number_format = FMT_PESO
                c.alignment = aln_right
            else:
                c.value = str(val) if val not in (None, "") else ""
                c.alignment = aln_left
            c.font = font_data; c.border = borde
            if fill_row: c.fill = fill_row
        ws.row_dimensions[ri].height = 16

    # Fila de totales — merge dinámico hasta antes de la primera columna numérica
    TOT = HDR + len(df_fx) + 1
    _num_positions = sorted(COLS.index(col) + 1 for col in NUM_COLS if col in COLS)
    _merge_end = (_num_positions[0] - 1) if _num_positions else (N - 1)
    if _merge_end >= 1:
        ws.merge_cells(start_row=TOT, start_column=1, end_row=TOT, end_column=_merge_end)
    c = ws.cell(row=TOT, column=1)
    c.value = f"TOTAL: {len(df_fx):,} facturas"
    c.font = font_total; c.fill = fill_total; c.alignment = aln_right
    for col in NUM_COLS:
        if col in COLS:
            ci = COLS.index(col) + 1
            cl = get_column_letter(ci)
            c = ws.cell(row=TOT, column=ci)
            c.value = f"=SUM({cl}{HDR+1}:{cl}{TOT-1})"
            c.number_format = FMT_PESO
            c.font = font_total; c.fill = fill_total; c.alignment = aln_right
    ws.row_dimensions[TOT].height = 18

    # Anchos
    ANCHOS = {
        "N° Factura X": 18, "Tipo": 22, "Fecha Factura": 14, "Folio OT": 10,
        "Sucursal": 14, "Tipo Venta": 20, "Patente": 10, "RUT Cliente": 22,
        "💰 Abono Cliente": 36, "Fecha Anticipo": 14,
        "Marca": 12, "Asesor": 24, "Total OT $": 14, "Costo Vale Consumo": 20,
    }
    for ci, col in enumerate(COLS, 1):
        ws.column_dimensions[get_column_letter(ci)].width = ANCHOS.get(col, 15)

    # Auto-filtro y freeze
    ws.auto_filter.ref = f"A{HDR}:{get_column_letter(N)}{HDR}"
    ws.freeze_panes = f"A{HDR + 1}"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


# ============================================================
#   GENERADOR EXCEL — INFORME POR ASESOR (OTs pendientes)
# ============================================================
def generar_excel_asesor(df_as, asesores_sel, fecha_act):
    """Informe CONSOLIDADO de OTs pendientes de uno o varios asesores.

    Es un solo archivo con todos los asesores juntos (a propósito: una misma
    persona puede figurar con más de un nombre de usuario en las OT, así que
    separar por nombre partiría su carga en pedazos). La columna "Asesor"
    permite igual distinguir cada alias, y el Resumen trae un desglose por
    asesor para verlos comparados.

    3 hojas:
      · Resumen      — indicadores globales + desgloses (asesor, sucursal,
                       rango, tipo de venta, categoría, marca).
      · Listado OTs  — una fila por OT: folio, neto, costo del Vale de
                       Consumo y los folios de cada tipo de documento.
      · Documentos   — una fila por documento posterior (OT · tipo · folio),
                       para filtrar o armar dinámicas.
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    AZUL_OSC  = "1A3A5C"
    AZUL_MED  = "2C5F8D"
    AZUL_CLR  = "D9E6F2"
    GRIS_FILA = "F2F6FA"
    ROJO      = "C0392B"

    f_titulo  = Font(name="Arial", bold=True, size=14, color="FFFFFF")
    f_sub     = Font(name="Arial", bold=True, size=10, color=AZUL_OSC)
    f_seccion = Font(name="Arial", bold=True, size=11, color="FFFFFF")
    f_header  = Font(name="Arial", bold=True, size=10, color="FFFFFF")
    f_data    = Font(name="Arial", size=10)
    f_bold    = Font(name="Arial", bold=True, size=10)
    f_kpi_lbl = Font(name="Arial", size=10, color="444444")
    f_kpi_val = Font(name="Arial", bold=True, size=12, color=AZUL_OSC)
    f_kpi_rj  = Font(name="Arial", bold=True, size=12, color=ROJO)

    fill_titulo  = PatternFill("solid", fgColor=AZUL_OSC)
    fill_sub     = PatternFill("solid", fgColor=AZUL_CLR)
    fill_seccion = PatternFill("solid", fgColor=AZUL_MED)
    fill_header  = PatternFill("solid", fgColor=AZUL_OSC)
    fill_gris    = PatternFill("solid", fgColor=GRIS_FILA)
    fill_total   = PatternFill("solid", fgColor=AZUL_CLR)

    aln_c = Alignment(horizontal="center", vertical="center")
    aln_l = Alignment(horizontal="left",   vertical="center")
    aln_r = Alignment(horizontal="right",  vertical="center")
    aln_w = Alignment(horizontal="left",   vertical="top", wrap_text=True)
    borde = Border(bottom=Side(style="thin", color="CCCCCC"))

    FMT_PESO = '"$"#,##0'
    FMT_INT  = '#,##0'
    FMT_PCT  = '0.0%'
    FMT_TXT  = '@'

    RANGOS = ["0-30", "31-60", "61-90", "91 o más"]

    def _costo_vale(reps):
        if not isinstance(reps, list):
            try:
                reps = json.loads(reps) if isinstance(reps, str) else []
            except Exception:
                reps = []
        _t = 0.0
        for _r in (reps or []):
            if isinstance(_r, dict):
                try:
                    _t += float(str(_r.get("costo_total", 0) or 0).replace(",", "."))
                except Exception:
                    pass
        return round(_t)

    _d = df_as.copy()
    for _c in ["ASESOR", "SUCURSAL", "RANGO", "TIPO VENTA", "MARCA", "MODELO",
               "PATENTE", "CATEGORIA", "FECHA OT", "OBSERVACION OT", "NOTAS",
               "AVANCE - GESTIÓN", "ULTIMA_EDICION"]:
        if _c not in _d.columns:
            _d[_c] = ""
        _d[_c] = _d[_c].fillna("").astype(str).str.strip()
    _d["_NETO"]  = pd.to_numeric(_d.get("NETO"), errors="coerce").fillna(0)
    _d["_DIAS"]  = pd.to_numeric(_d.get("DIAS APERTURA"), errors="coerce").fillna(0)
    _d["_COSTO"] = (_d["repuestos_actual"].apply(_costo_vale)
                    if "repuestos_actual" in _d.columns else 0)
    _d["_CAT"]   = _d["CATEGORIA"].replace("", "Sin categoría")
    _d["_ASES"]  = _d["ASESOR"].replace("", "Sin asesor")
    # Evita filas con la etiqueta en blanco en los desgloses del Resumen
    _d["_MARCA"] = _d["MARCA"].replace("", "Sin marca")
    _d["_TVTA"]  = _d["TIPO VENTA"].replace("", "Sin tipo de venta")
    _d["_SUC"]   = _d["SUCURSAL"].replace("", "Sin sucursal")
    _d["_RANGO"] = _d["RANGO"].replace("", "Sin rango")

    total     = len(_d)
    n_crit    = int((_d["RANGO"] == "91 o más").sum())
    n_urg     = int((_d["RANGO"] == "61-90").sum())
    n_ate     = int((_d["RANGO"] == "31-60").sum())
    n_rec     = int((_d["RANGO"] == "0-30").sum())
    neto_tot  = float(_d["_NETO"].sum())
    costo_tot = float(_d["_COSTO"].sum())
    dias_prom = float(_d["_DIAS"].mean()) if total else 0.0
    dias_max  = float(_d["_DIAS"].max())  if total else 0.0
    sin_gest  = int((_d["AVANCE - GESTIÓN"] == "").sum())
    sin_cat   = int((_d["CATEGORIA"] == "").sum())
    neto_cero = int((_d["_NETO"] == 0).sum())
    con_vale  = int((_d["_COSTO"] > 0).sum())

    wb = Workbook()

    # =====================================================
    #   HOJA 1 — RESUMEN
    # =====================================================
    ws = wb.active
    ws.title = "Resumen"
    N_RES = 10

    def _fila_titulo(fila, texto, sub=""):
        ws.merge_cells(start_row=fila, start_column=1, end_row=fila, end_column=N_RES)
        c = ws.cell(row=fila, column=1, value=texto)
        c.font, c.fill, c.alignment = f_titulo, fill_titulo, aln_c
        ws.row_dimensions[fila].height = 30
        if sub:
            ws.merge_cells(start_row=fila + 1, start_column=1, end_row=fila + 1, end_column=N_RES)
            c2 = ws.cell(row=fila + 1, column=1, value=sub)
            c2.font, c2.fill, c2.alignment = f_sub, fill_sub, aln_c
            ws.row_dimensions[fila + 1].height = 18
            return fila + 2
        return fila + 1

    def _seccion(fila, texto):
        ws.merge_cells(start_row=fila, start_column=1, end_row=fila, end_column=N_RES)
        c = ws.cell(row=fila, column=1, value=texto)
        c.font, c.fill, c.alignment = f_seccion, fill_seccion, aln_l
        ws.row_dimensions[fila].height = 20
        return fila + 1

    def _tabla(fila, headers, rows, fmts=None, totales=None):
        """Pinta una tabla. fmts: dict {indice_col: number_format}."""
        fmts = fmts or {}
        for ci, h in enumerate(headers, 1):
            c = ws.cell(row=fila, column=ci, value=h)
            c.font, c.fill, c.alignment = f_header, fill_header, aln_c
        ws.row_dimensions[fila].height = 18
        r = fila
        for ri, row in enumerate(rows):
            r = fila + 1 + ri
            for ci, val in enumerate(row, 1):
                c = ws.cell(row=r, column=ci)
                c.value = val
                c.font, c.border = f_data, borde
                if ci in fmts:
                    c.number_format = fmts[ci]
                    c.alignment = aln_r
                else:
                    c.alignment = aln_l if ci == 1 else aln_c
                if ri % 2 == 1:
                    c.fill = fill_gris
        if totales:
            r += 1
            for ci, val in enumerate(totales, 1):
                c = ws.cell(row=r, column=ci, value=val)
                c.font, c.fill = f_bold, fill_total
                c.alignment = aln_r if ci in fmts else (aln_l if ci == 1 else aln_c)
                if ci in fmts:
                    c.number_format = fmts[ci]
        return r + 2

    _lst_as = [a for a in asesores_sel if str(a).strip()]
    _txt_as = " · ".join(_lst_as) if _lst_as else "Todos los asesores del listado filtrado"
    if len(_txt_as) > 190:
        _txt_as = _txt_as[:187] + "…"

    fila = _fila_titulo(1, "INFORME DE OTs PENDIENTES POR ASESOR — CURIFOR S.A",
                        f"Asesor(es): {_txt_as}     |     Datos al {fecha_act}     |     "
                        f"{total:,} OT(s) pendientes")
    fila += 1

    # ---- KPIs ----
    fila = _seccion(fila, "INDICADORES GLOBALES")
    _pct_crit = (n_crit / total) if total else 0.0
    KPIS = [
        ("Total OTs pendientes", total,      FMT_INT,  False),
        ("🔴 Críticas (más de 90 días)", n_crit, FMT_INT, True),
        ("🟠 Urgentes (61-90 días)", n_urg,  FMT_INT,  False),
        ("🟡 Atención (31-60 días)", n_ate,  FMT_INT,  False),
        ("🟢 Recientes (0-30 días)", n_rec,  FMT_INT,  False),
        ("% Críticas sobre el total", _pct_crit, FMT_PCT, True),
        ("Días de apertura promedio", round(dias_prom, 1), '#,##0.0', False),
        ("Días de apertura máximo",   dias_max, FMT_INT, False),
        ("Neto total en juego",       neto_tot, FMT_PESO, False),
        ("Costo total Vale de Consumo", costo_tot, FMT_PESO, False),
        ("OTs con Vale de Consumo cargado", con_vale, FMT_INT, False),
        ("OTs con Neto $0",           neto_cero, FMT_INT, True),
        ("OTs sin Avance / Gestión escrito", sin_gest, FMT_INT, True),
        ("OTs sin Categoría asignada",       sin_cat,  FMT_INT, True),
    ]
    _fila_kpi = fila
    for _i, (_lbl, _val, _fmt, _alerta) in enumerate(KPIS):
        _r = _fila_kpi + (_i // 2)
        _c0 = 1 if _i % 2 == 0 else 6
        ws.merge_cells(start_row=_r, start_column=_c0, end_row=_r, end_column=_c0 + 2)
        cl = ws.cell(row=_r, column=_c0, value=_lbl)
        cl.font, cl.alignment, cl.border = f_kpi_lbl, aln_l, borde
        ws.merge_cells(start_row=_r, start_column=_c0 + 3, end_row=_r, end_column=_c0 + 4)
        cv = ws.cell(row=_r, column=_c0 + 3, value=_val)
        cv.font = f_kpi_rj if (_alerta and _val) else f_kpi_val
        cv.alignment, cv.number_format, cv.border = aln_r, _fmt, borde
        ws.row_dimensions[_r].height = 18
    fila = _fila_kpi + ((len(KPIS) + 1) // 2) + 1

    # ---- Desglose por asesor ----
    def _bloque_grupo(col_grupo, titulo, etiqueta, orden_fijo=None):
        _rows, _tot_n = [], 0
        _grupos = (orden_fijo if orden_fijo is not None
                   else list(_d.groupby(col_grupo).size().sort_values(ascending=False).index))
        for _g in _grupos:
            _sub = _d[_d[col_grupo] == _g]
            _n = len(_sub)
            if _n == 0 and orden_fijo is None:
                continue
            _tot_n += _n
            _rows.append([
                str(_g), _n,
                int((_sub["RANGO"] == "0-30").sum()),
                int((_sub["RANGO"] == "31-60").sum()),
                int((_sub["RANGO"] == "61-90").sum()),
                int((_sub["RANGO"] == "91 o más").sum()),
                (int((_sub["RANGO"] == "91 o más").sum()) / _n) if _n else 0,
                round(float(_sub["_DIAS"].mean()), 1) if _n else 0,
                float(_sub["_NETO"].sum()),
                float(_sub["_COSTO"].sum()),
            ])
        _f = _seccion(fila, titulo)
        return _tabla(
            _f,
            [etiqueta, "OTs", "0-30", "31-60", "61-90", "+90", "% Críticas",
             "Días prom.", "Neto total", "Costo Vale"],
            _rows,
            fmts={2: FMT_INT, 3: FMT_INT, 4: FMT_INT, 5: FMT_INT, 6: FMT_INT,
                  7: FMT_PCT, 8: '#,##0.0', 9: FMT_PESO, 10: FMT_PESO},
            totales=["TOTAL", total, n_rec, n_ate, n_urg, n_crit, _pct_crit,
                     round(dias_prom, 1), neto_tot, costo_tot],
        )

    fila = _bloque_grupo("_ASES", "DESGLOSE POR ASESOR", "Asesor")
    fila = _bloque_grupo("_SUC", "DESGLOSE POR SUCURSAL", "Sucursal")

    # ---- Rango / Tipo de venta / Categoría / Marca ----
    def _bloque_simple(fila_ini, titulo, etiqueta, col, orden_fijo=None):
        _grupos = (orden_fijo if orden_fijo is not None
                   else list(_d.groupby(col).size().sort_values(ascending=False).index))
        _rows = []
        for _g in _grupos:
            _sub = _d[_d[col] == _g]
            _n = len(_sub)
            if _n == 0 and orden_fijo is None:
                continue
            _rows.append([str(_g), _n, (_n / total) if total else 0,
                          float(_sub["_NETO"].sum()), float(_sub["_COSTO"].sum())])
        _f = _seccion(fila_ini, titulo)
        return _tabla(_f, [etiqueta, "OTs", "% del total", "Neto total", "Costo Vale"],
                      _rows,
                      fmts={2: FMT_INT, 3: FMT_PCT, 4: FMT_PESO, 5: FMT_PESO},
                      totales=["TOTAL", total, 1.0, neto_tot, costo_tot])

    _rangos_pres = RANGOS + (["Sin rango"] if (_d["_RANGO"] == "Sin rango").any() else [])
    fila = _bloque_simple(fila, "DESGLOSE POR RANGO DE DÍAS", "Rango", "_RANGO",
                          orden_fijo=_rangos_pres)
    fila = _bloque_simple(fila, "DESGLOSE POR TIPO DE VENTA", "Tipo de venta", "_TVTA")
    fila = _bloque_simple(fila, "DESGLOSE POR CATEGORÍA", "Categoría", "_CAT")
    fila = _bloque_simple(fila, "DESGLOSE POR MARCA", "Marca", "_MARCA")

    # ---- Documentos posteriores (resumen) ----
    _rows_doc = []
    for _nom, _k in DOCS_CONFIG:
        _cn = f"N_{_k}"
        _n_docs = int(pd.to_numeric(_d.get(_cn, 0), errors="coerce").fillna(0).sum()) if _cn in _d.columns else 0
        _n_ots  = int((pd.to_numeric(_d.get(_cn, 0), errors="coerce").fillna(0) > 0).sum()) if _cn in _d.columns else 0
        _rows_doc.append([_nom, _n_ots, (_n_ots / total) if total else 0, _n_docs])
    fila = _seccion(fila, "DOCUMENTOS POSTERIORES ASOCIADOS")
    fila = _tabla(fila, ["Tipo de documento", "OTs con el documento",
                         "% de las OTs", "N° de documentos"],
                  _rows_doc, fmts={2: FMT_INT, 3: FMT_PCT, 4: FMT_INT})

    _ANCHOS_RES = {1: 34, 2: 12, 3: 10, 4: 10, 5: 10, 6: 10, 7: 12, 8: 12, 9: 18, 10: 18}
    for _ci, _w in _ANCHOS_RES.items():
        ws.column_dimensions[get_column_letter(_ci)].width = _w
    ws.freeze_panes = "A4"
    ws.sheet_view.showGridLines = False

    # =====================================================
    #   HOJA 2 — LISTADO DE OTs
    # =====================================================
    ws2 = wb.create_sheet("Listado OTs")
    COLS_L = ["Folio OT", "Asesor", "Sucursal", "Patente", "Marca", "Modelo",
              "Tipo Venta", "Categoría", "Fecha OT", "Días", "Rango",
              "Neto", "Costo Vale Consumo"]
    for _nom, _k in DOCS_CONFIG:
        COLS_L += [f"# {_nom}", f"Folios {_nom}"]
    COLS_L += ["Observación OT", "Notas", "Avance / Gestión", "Última edición"]
    NL = len(COLS_L)
    _i_neto  = COLS_L.index("Neto") + 1
    _i_costo = COLS_L.index("Costo Vale Consumo") + 1
    _i_dias  = COLS_L.index("Días") + 1
    _cols_folio = {COLS_L.index(f"Folios {n}") + 1 for n, _ in DOCS_CONFIG}
    _cols_ndoc  = {COLS_L.index(f"# {n}") + 1 for n, _ in DOCS_CONFIG}

    ws2.merge_cells(start_row=1, start_column=1, end_row=1, end_column=NL)
    c = ws2.cell(row=1, column=1,
                 value=f"LISTADO DE OTs PENDIENTES — {_txt_as}   |   {total:,} OT(s)   |   {fecha_act}")
    c.font, c.fill, c.alignment = f_titulo, fill_titulo, aln_c
    ws2.row_dimensions[1].height = 26
    ws2.row_dimensions[2].height = 6

    HDR2 = 3
    for ci, h in enumerate(COLS_L, 1):
        c = ws2.cell(row=HDR2, column=ci, value=h)
        c.font, c.fill, c.alignment = f_header, fill_header, aln_c
    ws2.row_dimensions[HDR2].height = 22

    _d_ord = _d.sort_values(["_DIAS"], ascending=False)
    _fila_doc = []          # para la hoja 3
    for ri, (_, row) in enumerate(_d_ord.iterrows(), start=HDR2 + 1):
        _vals = [
            str(row.get("FOLIO OT", "")), row["_ASES"], row["SUCURSAL"], row["PATENTE"],
            row["MARCA"], row["MODELO"], row["TIPO VENTA"], row["_CAT"],
            row["FECHA OT"], int(row["_DIAS"]), row["RANGO"],
            float(row["_NETO"]), float(row["_COSTO"]),
        ]
        for _nom, _k in DOCS_CONFIG:
            _n_d = int(pd.to_numeric(pd.Series([row.get(f"N_{_k}", 0)]), errors="coerce").fillna(0).iloc[0])
            _fol = str(row.get(f"FOLIOS_{_k}", "") or "").strip()
            _vals += [_n_d, _fol]
            for _f1 in [x.strip() for x in _fol.split(",") if x.strip()]:
                _fila_doc.append([str(row.get("FOLIO OT", "")), row["_ASES"], row["SUCURSAL"],
                                  row["PATENTE"], row["RANGO"], int(row["_DIAS"]), _nom, _f1])
        _vals += [row["OBSERVACION OT"], row["NOTAS"], row["AVANCE - GESTIÓN"],
                  row["ULTIMA_EDICION"]]

        _es_crit = (row["RANGO"] == "91 o más")
        for ci, val in enumerate(_vals, 1):
            c = ws2.cell(row=ri, column=ci)
            c.value = val
            c.font, c.border = f_data, borde
            if ci in (_i_neto, _i_costo):
                c.number_format, c.alignment = FMT_PESO, aln_r
            elif ci in _cols_ndoc or ci == _i_dias:
                c.number_format, c.alignment = FMT_INT, aln_c
            elif ci in _cols_folio or ci == 1:
                c.number_format, c.alignment = FMT_TXT, aln_l
            else:
                c.alignment = aln_l
            if ri % 2 == 0:
                c.fill = fill_gris
        if _es_crit:
            ws2.cell(row=ri, column=COLS_L.index("Rango") + 1).font = Font(
                name="Arial", size=10, bold=True, color=ROJO)
        ws2.row_dimensions[ri].height = 15

    _fin2 = HDR2 + len(_d_ord)
    _tr = _fin2 + 1
    ws2.merge_cells(start_row=_tr, start_column=1, end_row=_tr, end_column=_i_neto - 1)
    c = ws2.cell(row=_tr, column=1, value=f"TOTAL: {total:,} OT(s) pendientes")
    c.font, c.fill, c.alignment = f_bold, fill_total, aln_r
    for _ic in (_i_neto, _i_costo):
        _cl = get_column_letter(_ic)
        c = ws2.cell(row=_tr, column=_ic)
        c.value = f"=SUM({_cl}{HDR2+1}:{_cl}{_fin2})" if total else 0
        c.number_format, c.font, c.fill, c.alignment = FMT_PESO, f_bold, fill_total, aln_r
    for _ic in _cols_ndoc:
        _cl = get_column_letter(_ic)
        c = ws2.cell(row=_tr, column=_ic)
        c.value = f"=SUM({_cl}{HDR2+1}:{_cl}{_fin2})" if total else 0
        c.number_format, c.font, c.fill, c.alignment = FMT_INT, f_bold, fill_total, aln_c

    _ANCHOS_L = {"Folio OT": 12, "Asesor": 26, "Sucursal": 16, "Patente": 10,
                 "Marca": 14, "Modelo": 20, "Tipo Venta": 20, "Categoría": 18,
                 "Fecha OT": 12, "Días": 8, "Rango": 11, "Neto": 15,
                 "Costo Vale Consumo": 19, "Observación OT": 40, "Notas": 40,
                 "Avance / Gestión": 40, "Última edición": 26}
    for ci, h in enumerate(COLS_L, 1):
        _w = _ANCHOS_L.get(h)
        if _w is None:
            _w = 8 if h.startswith("# ") else 26
        ws2.column_dimensions[get_column_letter(ci)].width = _w
    ws2.auto_filter.ref = f"A{HDR2}:{get_column_letter(NL)}{max(_fin2, HDR2)}"
    ws2.freeze_panes = f"C{HDR2 + 1}"

    # =====================================================
    #   HOJA 3 — DOCUMENTOS (una fila por documento)
    # =====================================================
    ws3 = wb.create_sheet("Documentos")
    COLS_D = ["Folio OT", "Asesor", "Sucursal", "Patente", "Rango", "Días",
              "Tipo de Documento", "N° / Folio del Documento"]
    ND = len(COLS_D)
    ws3.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ND)
    c = ws3.cell(row=1, column=1,
                 value=f"DOCUMENTOS POSTERIORES POR OT — {len(_fila_doc):,} documento(s)   |   {fecha_act}")
    c.font, c.fill, c.alignment = f_titulo, fill_titulo, aln_c
    ws3.row_dimensions[1].height = 26
    ws3.row_dimensions[2].height = 6
    HDR3 = 3
    for ci, h in enumerate(COLS_D, 1):
        c = ws3.cell(row=HDR3, column=ci, value=h)
        c.font, c.fill, c.alignment = f_header, fill_header, aln_c
    ws3.row_dimensions[HDR3].height = 22
    for ri, fila_d in enumerate(_fila_doc, start=HDR3 + 1):
        for ci, val in enumerate(fila_d, 1):
            c = ws3.cell(row=ri, column=ci)
            c.value = val
            c.font, c.border = f_data, borde
            if ci == 6:
                c.number_format, c.alignment = FMT_INT, aln_c
            elif ci in (1, 8):
                c.number_format, c.alignment = FMT_TXT, aln_l
            else:
                c.alignment = aln_l if ci in (2, 3, 7) else aln_c
            if ri % 2 == 0:
                c.fill = fill_gris
    if not _fila_doc:
        ws3.merge_cells(start_row=HDR3 + 1, start_column=1, end_row=HDR3 + 1, end_column=ND)
        c = ws3.cell(row=HDR3 + 1, column=1,
                     value="Las OTs de este listado aún no tienen documentos posteriores asociados.")
        c.font, c.alignment = f_data, aln_c
    _ANCHOS_D = {1: 12, 2: 26, 3: 16, 4: 10, 5: 11, 6: 8, 7: 22, 8: 26}
    for _ci, _w in _ANCHOS_D.items():
        ws3.column_dimensions[get_column_letter(_ci)].width = _w
    ws3.auto_filter.ref = f"A{HDR3}:{get_column_letter(ND)}{max(HDR3 + len(_fila_doc), HDR3)}"
    ws3.freeze_panes = f"A{HDR3 + 1}"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


# ============================================================
#   GENERADOR DE TEXTO DE CORREO — OTs >90 días por Asesor
# ============================================================
_DIAS_SEMANA_ES = ["lunes", "martes", "miércoles", "jueves", "viernes", "sábado", "domingo"]


def generar_texto_correo_asesor(df_as, nombre_saludo, fecha_act, fecha_limite,
                                incluir_folios=True, max_folios=15,
                                firma="Cristóbal Jerez\nPost Venta — Curifor S.A."):
    """Texto BREVE e IMPERATIVO para exigirle a un asesor que gestione sus OTs >90 días.

    Se arma con los datos reales de las OTs filtradas (las del asesor seleccionado),
    para pegarlo directo en el correo. Devuelve (asunto, cuerpo).
    """
    def _costo(reps):
        if not isinstance(reps, list):
            try:
                reps = json.loads(reps) if isinstance(reps, str) else []
            except Exception:
                reps = []
        _t = 0.0
        for _r in (reps or []):
            if isinstance(_r, dict):
                try:
                    _t += float(str(_r.get("costo_total", 0) or 0).replace(",", "."))
                except Exception:
                    pass
        return round(_t)

    _d = df_as.copy()
    for _c in ["RANGO", "SUCURSAL", "AVANCE - GESTIÓN", "FOLIO OT",
               "TIPO VENTA", "CATEGORIA"]:
        if _c not in _d.columns:
            _d[_c] = ""
        _d[_c] = _d[_c].fillna("").astype(str).str.strip()
    _d["_DIAS"] = pd.to_numeric(_d.get("DIAS APERTURA"), errors="coerce").fillna(0)
    _d["_NETO"] = pd.to_numeric(_d.get("NETO"), errors="coerce").fillna(0)

    total = len(_d)
    d90   = _d[_d["RANGO"] == "91 o más"].copy()
    n90   = len(d90)
    _nom  = str(nombre_saludo).strip() or "colega"
    # En el correo basta la fecha del dato, sin la hora de la consolidación
    fecha_act = str(fecha_act).split(" ")[0]

    _sucs = [s for s in _d["SUCURSAL"].unique() if s]
    _suc_txt = " / ".join(sorted(_sucs)[:3]) if _sucs else "—"

    if isinstance(fecha_limite, str):
        _lim_txt = fecha_limite
    else:
        _lim_txt = (f"{_DIAS_SEMANA_ES[fecha_limite.weekday()]} "
                    f"{fecha_limite.strftime('%d/%m/%Y')}")

    # --- Sin OTs críticas: mensaje corto, no se inventa una urgencia ---
    if n90 == 0:
        asunto = f"OT abiertas — {_nom} ({_suc_txt})"
        cuerpo = (
            f"Estimado/a {_nom}:\n\n"
            f"Al {fecha_act} no tienes ninguna OT con más de 90 días. "
            f"Tienes {total} OT abiertas en total.\n\n"
            f"Mantén el ritmo: antes del {_lim_txt}, deja escrito en la aplicación "
            f"en qué va cada una y cierra las que ya estén listas.\n\n"
            f"Recuerda que el control y la gestión de las OT se lleva en nuestra "
            f"aplicación: lo que no esté escrito ahí no cuenta como gestionado.\n\n"
            f"Saludos,\n{firma}"
        )
        return asunto, cuerpo

    d90 = d90.sort_values("_DIAS", ascending=False)
    pct     = n90 / total * 100 if total else 0
    neto90  = float(d90["_NETO"].sum())
    costo90 = float(d90["repuestos_actual"].apply(_costo).sum()) if "repuestos_actual" in d90.columns else 0.0
    sin_g   = int((d90["AVANCE - GESTIÓN"] == "").sum())
    _f_ant  = str(d90.iloc[0]["FOLIO OT"])
    _d_ant  = int(d90.iloc[0]["_DIAS"])

    asunto = f"URGENTE · {n90} OT con más de 90 días — {_nom} ({_suc_txt})"

    _l = []
    _l.append(f"Estimado/a {_nom}:")
    _l.append("")
    # Garantías dentro de las críticas: por Tipo de Venta (VTA GARANTIA) o por
    # Categoría (GARANTIA / GARANTIA EXTENDIDA / SUBIR — mismo criterio que usa
    # el Informe por Área para el área Garantía).
    _tv_g  = d90["TIPO VENTA"].str.upper().str.contains("GARANT", na=False)
    _cat_u = d90["CATEGORIA"].str.upper()
    _cat_g = _cat_u.str.contains("GARANTIA", na=False) | (_cat_u == "SUBIR")
    _es_gar = _tv_g | _cat_g
    n_gar    = int(_es_gar.sum())
    neto_gar = float(d90.loc[_es_gar, "_NETO"].sum())

    def _m(v):
        # Separador de miles chileno (punto), solo para los montos
        return f"{v:,.0f}".replace(",", ".")

    _l.append(
        f"Al {fecha_act} tienes {total} OT abiertas y {n90} de ellas llevan más de "
        f"90 días ({pct:.0f}% del total): ${_m(neto90)} en venta y ${_m(costo90)} en "
        f"repuestos ya cargados. La más antigua es la OT {_f_ant}, con {_d_ant} días."
    )
    _l.append("")
    if n_gar:
        _l.append(
            f"De esas {n90}, {n_gar} son garantías (${_m(neto_gar)}): "
            "presenta su respaldo a la marca antes de que se venzan los plazos."
        )
    else:
        _l.append(f"Ninguna de esas {n90} OT es garantía.")
    _l.append("")
    _l.append(f"Antes del {_lim_txt}:")
    _l.append(f"1. Escribe en la aplicación en qué va cada una de esas {n90} OT"
              + (f" ({sin_g} todavía no tienen ninguna nota)." if sin_g else "."))
    _l.append("2. Cierra las que ya estén listas.")
    _l.append("3. Avísame cuáles están detenidas en taller y desde cuándo.")
    if incluir_folios:
        _fols = [str(f) for f in d90["FOLIO OT"].tolist() if str(f).strip()][:max_folios]
        if _fols:
            _resto = n90 - len(_fols)
            _l.append("")
            _l.append("OT a revisar: " + ", ".join(_fols)
                      + (f" y {_resto} más (van en el archivo adjunto)." if _resto > 0 else "."))
    _l.append("")
    _l.append(
        "Recuerda que el control y la gestión de las OT se lleva en nuestra aplicación: "
        "todo el respaldo de lo que hagas tiene que quedar escrito ahí. Es lo que reviso "
        "para hacer el seguimiento, así que lo que no esté en la aplicación no cuenta "
        "como gestionado."
    )
    _l.append("")
    _l.append("Te adjunto el detalle de cada OT.")
    _l.append("")
    _l.append(f"Saludos,\n{firma}")
    return asunto, "\n".join(_l)


# ============================================================
#   GENERADOR DE INFORME PDF (una página A4 horizontal)
# ============================================================
def generar_pdf_informe(df_inf, filtros_desc, fecha_act, logo_b64_str):
    """PDF informe ejecutivo — UNA página A4 landscape, sin listados individuales."""
    (A4, landscape, colors, cm, mm, rl_canvas, Table, TableStyle,
     plt, mpatches, ImageReader) = _importar_pdf_libs()

    # ── Paleta corporativa ─────────────────────────────────
    AZUL     = colors.HexColor("#0c243d")
    AZUL_MED = colors.HexColor("#1a3a5c")
    AZUL_CLR = colors.HexColor("#4a7ab5")
    ROJO     = colors.HexColor("#ef4444")
    NARANJA  = colors.HexColor("#f97316")
    AMARILLO = colors.HexColor("#f59e0b")
    VERDE    = colors.HexColor("#22c55e")
    GRIS_BG  = colors.HexColor("#f8fafc")
    GRIS_ALT = colors.HexColor("#f1f5f9")
    GRIS_LIN = colors.HexColor("#cbd5e1")
    BLANCO   = colors.white

    PAGE_W, PAGE_H = landscape(A4)
    M = 1.1 * cm

    buf = io.BytesIO()
    c = rl_canvas.Canvas(buf, pagesize=landscape(A4))
    c.setTitle("Informe Control y Gestión Post Venta — Curifor S.A")

    # ── Métricas base ──────────────────────────────────────
    total     = len(df_inf)
    if total == 0:
        c.showPage(); c.save(); buf.seek(0); return buf.getvalue()

    # Normalizar ASESOR (trim + colapsar espacios + mayúsculas) para que un mismo
    # asesor no aparezca duplicado en ningún ranking/tabla por variaciones de
    # mayúscula/minúscula o espacios extra (ej. "Eduardo Ortiz" vs "EDUARDO ORTIZ",
    # o "Cristobal  Saavedra" con doble espacio vs "Cristobal Saavedra"). Se aplica
    # una sola vez acá porque el resto de la función (page 1, page 3, páginas por
    # sucursal) parte siempre de df_inf o de una copia de él.
    df_inf = df_inf.copy()
    if "ASESOR" in df_inf.columns:
        df_inf["ASESOR"] = (
            df_inf["ASESOR"].fillna("").astype(str)
            .str.strip()
            .str.replace(r"\s+", " ", regex=True)
            .str.upper()
        )

    criticas  = int((df_inf["RANGO"] == "91 o más").sum())
    urgentes  = int((df_inf["RANGO"] == "61-90").sum())
    atencion  = int((df_inf["RANGO"] == "31-60").sum())
    recientes = int((df_inf["RANGO"] == "0-30").sum())
    idx_crit  = criticas / total * 100

    dias_prom = 0.0
    if "DIAS APERTURA" in df_inf.columns:
        _dn = pd.to_numeric(df_inf["DIAS APERTURA"], errors="coerce").dropna()
        dias_prom = float(_dn.mean()) if len(_dn) else 0.0

    neto_total = 0
    if "NETO" in df_inf.columns:
        neto_total = pd.to_numeric(df_inf["NETO"], errors="coerce").fillna(0).sum()

    sin_gestion = 0
    if "AVANCE - GESTIÓN" in df_inf.columns:
        sin_gestion = int((df_inf["AVANCE - GESTIÓN"].fillna("").str.strip() == "").sum())

    # Sucursal stats
    rangos = ["0-30", "31-60", "61-90", "91 o más"]
    suc_stats = []
    for suc, grp in df_inf.groupby("SUCURSAL"):
        t_s = len(grp)
        crit_s = int((grp["RANGO"] == "91 o más").sum())
        _dp = pd.to_numeric(grp.get("DIAS APERTURA", pd.Series(dtype=float)), errors="coerce").mean()
        dp_str = f"{_dp:.0f}" if not pd.isna(_dp) else "—"
        pct = crit_s / t_s * 100 if t_s else 0
        suc_stats.append({
            "suc": str(suc)[:18],
            **{r: int((grp["RANGO"] == r).sum()) for r in rangos},
            "total": t_s,
            "pct_crit": f"{pct:.1f}%",
            "dias_prom": dp_str,
        })
    suc_stats.sort(key=lambda x: x["total"], reverse=True)

    # Total de páginas del informe: 1 (resumen) + 1 (sucursal/costo vale) +
    # 1 (detalle crítico tipo venta/categoría) + 1 página por cada sucursal con datos
    N_SUCURSALES_CON_DATOS = len(suc_stats)
    TOTAL_PAGINAS = 4 + N_SUCURSALES_CON_DATOS

    suc_data = [["Sucursal", "0-30", "31-60", "61-90", ">90", "Total", "% Críticas", "Días Prom."]]
    for s in suc_stats[:10]:
        suc_data.append([s["suc"], s["0-30"], s["31-60"], s["61-90"],
                         s["91 o más"], s["total"], s["pct_crit"], s["dias_prom"]])
    suc_data.append(["TOTAL", recientes, atencion, urgentes, criticas,
                     total, f"{idx_crit:.1f}%", f"{dias_prom:.0f}"])

    # Tipo de venta
    tv_grp = df_inf[df_inf["TIPO VENTA"] != ""].groupby("TIPO VENTA").size().sort_values(ascending=False)
    tv_data = [["Tipo de Venta", "N°", "%"]]
    for tipo, n in tv_grp.items():
        tv_data.append([str(tipo)[:24], n, f"{n/total*100:.1f}%"])
    tv_data.append(["TOTAL", total, "100%"])

    # Asesor stats
    asesor_data = [["Asesor", "OTs", "% Total", "Críticas", "Días Prom."]]
    if "ASESOR" in df_inf.columns:
        as_stats = []
        for asesor, grp in df_inf[df_inf["ASESOR"] != ""].groupby("ASESOR"):
            t_a = len(grp)
            crit_a = int((grp["RANGO"] == "91 o más").sum())
            _dp = pd.to_numeric(grp.get("DIAS APERTURA", pd.Series(dtype=float)), errors="coerce").mean()
            dp_str = f"{_dp:.0f}" if not pd.isna(_dp) else "—"
            as_stats.append({"asesor": str(asesor)[:22], "total": t_a,
                             "criticas": crit_a, "dp": dp_str,
                             "pct": f"{t_a/total*100:.1f}%"})
        as_stats.sort(key=lambda x: x["total"], reverse=True)
        for a in as_stats[:10]:
            asesor_data.append([a["asesor"], a["total"], a["pct"], a["criticas"], a["dp"]])

    # Marca
    marca_grp = df_inf[df_inf["MARCA"] != ""].groupby("MARCA").size().sort_values(ascending=False)
    marca_data = [["Marca", "N°", "%"]]
    for m, n in marca_grp.head(9).items():
        marca_data.append([str(m)[:18], n, f"{n/total*100:.1f}%"])

    # Categoría
    cat_col = df_inf["CATEGORIA"].fillna("Sin categoría").replace("", "Sin categoría")
    cat_grp = cat_col.groupby(cat_col).size().sort_values(ascending=False)
    cat_data = [["Categoría", "N°", "%"]]
    for cat, n in cat_grp.items():
        cat_data.append([str(cat)[:20], n, f"{n/total*100:.1f}%"])

    # Neto en juego × Tipo de Venta × Rango de días
    neto_x_data = None
    if "NETO" in df_inf.columns and "TIPO VENTA" in df_inf.columns:
        _df_n = df_inf.copy()
        _df_n["_NETO_V"] = pd.to_numeric(_df_n["NETO"], errors="coerce").fillna(0)
        _df_n["_TV_V"]   = _df_n["TIPO VENTA"].replace("", "Sin tipo")

        def _fmt_neto(v):
            if v == 0: return "—"
            if abs(v) >= 1_000_000_000: return f"${v/1_000_000_000:.1f}B"
            if abs(v) >= 1_000_000:     return f"${v/1_000_000:.1f}M"
            return f"${v/1_000:.0f}K"

        _rangos_p = ["0-30", "31-60", "61-90", "91 o más"]
        neto_x_data = [["Tipo de Venta", "0-30", "31-60", "61-90", ">90 d", "OTs", "Neto Total"]]
        for _tv in sorted(_df_n["_TV_V"].unique()):
            _g = _df_n[_df_n["_TV_V"] == _tv]
            _row = [str(_tv)[:16]]
            for _r in _rangos_p:
                _row.append(_fmt_neto(_g[_g["RANGO"] == _r]["_NETO_V"].sum()))
            _row.extend([len(_g), _fmt_neto(_g["_NETO_V"].sum())])
            neto_x_data.append(_row)
        neto_x_data.append([
            "TOTAL",
            _fmt_neto(_df_n[_df_n["RANGO"] == "0-30"]["_NETO_V"].sum()),
            _fmt_neto(_df_n[_df_n["RANGO"] == "31-60"]["_NETO_V"].sum()),
            _fmt_neto(_df_n[_df_n["RANGO"] == "61-90"]["_NETO_V"].sum()),
            _fmt_neto(_df_n[_df_n["RANGO"] == "91 o más"]["_NETO_V"].sum()),
            total,
            _fmt_neto(_df_n["_NETO_V"].sum()),
        ])

    # ── Gráfico de barras ──────────────────────────────────
    fig, ax = plt.subplots(figsize=(3.2, 1.85))
    bar_labels = ["0-30", "31-60", "61-90", ">90 d"]
    bar_vals   = [recientes, atencion, urgentes, criticas]
    bar_clrs   = ["#22c55e", "#f59e0b", "#f97316", "#ef4444"]
    bars = ax.bar(bar_labels, bar_vals, color=bar_clrs, width=0.58, zorder=3,
                  edgecolor="white", linewidth=0.5)
    ax.set_facecolor("#f8fafc"); fig.patch.set_facecolor("#f8fafc")
    ax.grid(axis="y", color="#e2e8f0", linewidth=0.55, zorder=0)
    for sp in ["top","right"]: ax.spines[sp].set_visible(False)
    for sp in ["left","bottom"]: ax.spines[sp].set_color("#e2e8f0")
    ax.tick_params(labelsize=7.5, colors="#1a3a5c")
    _mx = max(bar_vals) if bar_vals else 1
    for bar, val in zip(bars, bar_vals):
        if val > 0:
            ax.text(bar.get_x() + bar.get_width()/2, bar.get_height() + _mx*0.02,
                    str(val), ha="center", va="bottom", fontsize=8.5,
                    fontweight="bold", color="#0c243d")
    fig.tight_layout(pad=0.3)
    chart_buf = io.BytesIO()
    fig.savefig(chart_buf, format="png", dpi=150, bbox_inches="tight", facecolor="#f8fafc")
    plt.close(fig)
    chart_buf.seek(0)

    # ── Logo ──────────────────────────────────────────────
    logo_img_buf = None
    if logo_b64_str:
        try:
            logo_img_buf = io.BytesIO(base64.b64decode(logo_b64_str))
        except Exception:
            pass

    # ── Helper: título de sección ──────────────────────────
    def sec_title(txt, x, y, w, clr=AZUL_MED):
        c.saveState()
        c.setFillColor(clr)
        c.setFont("Helvetica-Bold", 6.2)
        c.drawString(x, y, txt.upper())
        c.setStrokeColor(clr)
        c.setLineWidth(0.65)
        c.line(x, y - 2, x + w, y - 2)
        c.restoreState()
        return y - 9

    # ── Helper: tabla Platypus ─────────────────────────────
    def make_tbl(data, cw, hdr_clr=AZUL_MED, totals=False, red_col=None):
        style = [
            ("BACKGROUND",   (0, 0), (-1, 0), hdr_clr),
            ("TEXTCOLOR",    (0, 0), (-1, 0), BLANCO),
            ("FONTNAME",     (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE",     (0, 0), (-1, 0), 6.2),
            ("FONTNAME",     (0, 1), (-1, -1), "Helvetica"),
            ("FONTSIZE",     (0, 1), (-1, -1), 6.2),
            ("ROWBACKGROUNDS",(0, 1), (-1, -2 if totals else -1), [BLANCO, GRIS_ALT]),
            ("GRID",         (0, 0), (-1, -1), 0.25, GRIS_LIN),
            ("LINEBELOW",    (0, 0), (-1, 0), 0.7, AZUL_CLR),
            ("TOPPADDING",   (0, 0), (-1, -1), 2.2),
            ("BOTTOMPADDING",(0, 0), (-1, -1), 2.2),
            ("LEFTPADDING",  (0, 0), (-1, -1), 3.5),
            ("RIGHTPADDING", (0, 0), (-1, -1), 3),
            ("ALIGN",        (0, 0), (0, -1), "LEFT"),
            ("ALIGN",        (1, 0), (-1, -1), "CENTER"),
            ("VALIGN",       (0, 0), (-1, -1), "MIDDLE"),
        ]
        if totals:
            style += [
                ("BACKGROUND", (0, -1), (-1, -1), AZUL),
                ("TEXTCOLOR",  (0, -1), (-1, -1), BLANCO),
                ("FONTNAME",   (0, -1), (-1, -1), "Helvetica-Bold"),
                ("LINEABOVE",  (0, -1), (-1, -1), 0.7, AZUL_CLR),
            ]
        if red_col is not None:
            for ri in range(1, len(data)):
                try:
                    val = data[ri][red_col]
                    if isinstance(val, int) and val > 0:
                        style.append(("TEXTCOLOR", (red_col, ri), (red_col, ri), ROJO))
                        style.append(("FONTNAME",  (red_col, ri), (red_col, ri), "Helvetica-Bold"))
                except Exception:
                    pass
        t = Table([[str(cell) for cell in row] for row in data], colWidths=cw)
        t.setStyle(TableStyle(style))
        return t

    # ═══════════════ DIBUJAR PÁGINA ═══════════════
    W, H = PAGE_W, PAGE_H

    # ── HEADER ────────────────────────────────────────────
    HDR_H = 1.58 * cm
    hdr_y = H - M - HDR_H
    c.saveState()
    c.setFillColor(AZUL)
    c.roundRect(M, hdr_y, W - 2*M, HDR_H, 5, fill=1, stroke=0)
    c.restoreState()

    if logo_img_buf:
        try:
            pw, ph = 2.85*cm, HDR_H - 6
            px0 = M + 5
            py0 = hdr_y + 3
            c.saveState()
            c.setFillColor(BLANCO)
            c.roundRect(px0, py0, pw, ph, 4, fill=1, stroke=0)
            c.restoreState()
            c.drawImage(ImageReader(logo_img_buf), px0 + 4, py0 + 3,
                        width=pw - 8, height=ph - 6,
                        preserveAspectRatio=True, mask="auto")
        except Exception:
            pass

    tx = M + 3.2*cm
    c.saveState()
    c.setFont("Helvetica-Bold", 10.5)
    c.setFillColor(BLANCO)
    c.drawString(tx, hdr_y + HDR_H*0.58, "INFORME OTs PENDIENTES — CURIFOR S.A")
    c.setFont("Helvetica", 6.3)
    c.setFillColor(colors.HexColor("#b0c8e8"))
    c.drawString(tx, hdr_y + HDR_H*0.22, f"Generado: {fecha_act}   |   Datos al: {fecha_act}")
    c.setFont("Helvetica", 5.7)
    c.setFillColor(colors.HexColor("#8ab0d4"))
    c.drawRightString(W - M - 5, hdr_y + HDR_H*0.44, filtros_desc[:120])
    c.restoreState()

    # ── KPI STRIP ─────────────────────────────────────────
    KPI_Y = hdr_y - 0.25*cm
    KPI_H = 1.42*cm
    KPI_GAP = 0.16*cm
    CW = W - 2*M
    KW = (CW - 4*KPI_GAP) / 5

    kpi_defs = [
        (f"{total:,}",     "TOTAL OTs PENDIENTES",  AZUL_CLR, "#eef4fc"),
        (f"{criticas:,}",  "CRÍTICAS  >90 DÍAS",    ROJO,     "#fef2f2"),
        (f"{urgentes:,}",  "URGENTES  61-90 DÍAS",  NARANJA,  "#fff7ed"),
        (f"{atencion:,}",  "ATENCIÓN  31-60 DÍAS",  AMARILLO, "#fffbeb"),
        (f"{recientes:,}", "RECIENTES  0-30 DÍAS",  VERDE,    "#f0fdf4"),
    ]
    for i, (val, lbl, top, bg) in enumerate(kpi_defs):
        kx = M + i*(KW + KPI_GAP)
        ky = KPI_Y - KPI_H
        c.saveState()
        c.setFillColor(colors.HexColor(bg))
        c.setStrokeColor(GRIS_LIN)
        c.setLineWidth(0.4)
        c.roundRect(kx, ky, KW, KPI_H, 4, fill=1, stroke=1)
        c.setFillColor(top)
        c.rect(kx, ky + KPI_H - 3, KW, 3, fill=1, stroke=0)
        c.setFont("Helvetica-Bold", 15.5)
        c.setFillColor(top)
        c.drawCentredString(kx + KW/2, ky + KPI_H*0.46, val)
        c.setFont("Helvetica-Bold", 4.8)
        c.setFillColor(colors.HexColor("#64748b"))
        c.drawCentredString(kx + KW/2, ky + KPI_H*0.16, lbl)
        c.restoreState()

    # ── CONTENT AREA ──────────────────────────────────────
    CY = KPI_Y - KPI_H - 0.28*cm          # top of content
    CH = CY - M - 0.38*cm                  # available height

    LEFT_W  = CW * 0.385
    RIGHT_W = CW * 0.615 - 0.28*cm
    RIGHT_X = M + LEFT_W + 0.28*cm

    # ─── COLUMNA IZQUIERDA ───────────────────────────────
    cy_l = CY

    # Gráfico
    next_l = sec_title("Distribución por rango de días", M, cy_l, LEFT_W)
    CH_IMG = 4.4*cm
    c.drawImage(ImageReader(chart_buf), M, next_l - CH_IMG,
                width=LEFT_W, height=CH_IMG, preserveAspectRatio=True)
    cy_l = next_l - CH_IMG - 0.36*cm

    # Tipo de venta
    next_l = sec_title("OTs por tipo de venta", M, cy_l, LEFT_W)
    tv_cw  = [LEFT_W*0.56, LEFT_W*0.22, LEFT_W*0.22]
    tv_tbl = make_tbl(tv_data, tv_cw, totals=True)
    _, tv_h = tv_tbl.wrapOn(c, LEFT_W, 999)
    tv_tbl.drawOn(c, M, next_l - tv_h)
    cy_l = next_l - tv_h - 0.32*cm

    # Tabla: Neto en juego × tipo de venta × rango
    _FOOTER_Y = M + 0.52*cm
    if neto_x_data and cy_l > _FOOTER_Y + 1.5*cm:
        next_l2 = sec_title("Neto en juego — tipo de venta × rango de días", M, cy_l, LEFT_W)
        _nt_cw = [LEFT_W*0.25, LEFT_W*0.10, LEFT_W*0.10,
                  LEFT_W*0.10, LEFT_W*0.10, LEFT_W*0.08, LEFT_W*0.27]
        _diff_nt = LEFT_W - sum(_nt_cw)
        _nt_cw[-1] += _diff_nt
        _hdr_nt  = neto_x_data[0]
        _tot_nt  = neto_x_data[-1]
        _rows_nt = neto_x_data[1:-1]
        avail_nt = next_l2 - _FOOTER_Y
        while len(_rows_nt) >= 1:
            _nt_tbl = make_tbl([_hdr_nt] + _rows_nt + [_tot_nt], _nt_cw, totals=True)
            _, _nt_h = _nt_tbl.wrapOn(c, LEFT_W, 999)
            if _nt_h <= avail_nt:
                break
            _rows_nt = _rows_nt[:-1]
        if _rows_nt:
            _nt_tbl.drawOn(c, M, next_l2 - _nt_h)

    # ─── COLUMNA DERECHA ─────────────────────────────────
    cy_r = CY

    # Tabla sucursales
    next_r = sec_title("OTs por sucursal y rango de días", RIGHT_X, cy_r, RIGHT_W)
    suc_cw = [RIGHT_W*0.20, RIGHT_W*0.09, RIGHT_W*0.09, RIGHT_W*0.09,
              RIGHT_W*0.09, RIGHT_W*0.10, RIGHT_W*0.17, RIGHT_W*0.13]
    diff_s = RIGHT_W - sum(suc_cw)
    suc_cw[-1] += diff_s
    suc_tbl = make_tbl(suc_data, suc_cw, totals=True, red_col=4)
    _, suc_h = suc_tbl.wrapOn(c, RIGHT_W, 999)
    suc_tbl.drawOn(c, RIGHT_X, next_r - suc_h)
    cy_r = next_r - suc_h - 0.32*cm

    # 3 sub-columnas: asesor | marca | categoría
    SUB_GAP = 0.18*cm
    SUB_W   = (RIGHT_W - 2*SUB_GAP) / 3
    SX2 = RIGHT_X + SUB_W + SUB_GAP
    SX3 = SX2 + SUB_W + SUB_GAP

    # Asesor
    next_r2 = sec_title("Por asesor (top 10)", RIGHT_X, cy_r, SUB_W)
    as_cw = [SUB_W*0.44, SUB_W*0.14, SUB_W*0.14, SUB_W*0.14, SUB_W*0.14]
    diff_a = SUB_W - sum(as_cw)
    as_cw[0] += diff_a
    as_tbl = make_tbl(asesor_data, as_cw)
    _, as_h = as_tbl.wrapOn(c, SUB_W, 999)
    as_tbl.drawOn(c, RIGHT_X, next_r2 - as_h)

    # Marca
    next_r3 = sec_title("Por marca", SX2, cy_r, SUB_W)
    ma_cw = [SUB_W*0.60, SUB_W*0.20, SUB_W*0.20]
    ma_tbl = make_tbl(marca_data, ma_cw)
    _, ma_h = ma_tbl.wrapOn(c, SUB_W, 999)
    ma_tbl.drawOn(c, SX2, next_r3 - ma_h)

    # Categoría
    next_r4 = sec_title("Por categoría", SX3, cy_r, SUB_W)
    ca_cw = [SUB_W*0.60, SUB_W*0.20, SUB_W*0.20]
    ca_tbl = make_tbl(cat_data, ca_cw)
    _, ca_h = ca_tbl.wrapOn(c, SUB_W, 999)
    ca_tbl.drawOn(c, SX3, next_r4 - ca_h)

    # ── ESPACIO INFERIOR DERECHA: OTs >90d x tipo venta ──
    cy_r_bot = min(next_r2 - as_h, next_r3 - ma_h, next_r4 - ca_h) - 0.32*cm

    # Tabla OTs >90d por tipo de venta (ancho completo de la columna derecha)
    crit_tv_data = None
    if criticas > 0:
        _crit_tv = (df_inf[df_inf["RANGO"] == "91 o más"]
                    .assign(_TV=lambda d: d["TIPO VENTA"].replace("", "Sin tipo"))
                    .groupby("_TV").size().sort_values(ascending=False))
        crit_tv_data = [["Tipo de Venta", "OTs >90d", "% rango"]]
        for tipo, n in _crit_tv.items():
            crit_tv_data.append([str(tipo)[:22], n, f"{n/criticas*100:.1f}%"])
        crit_tv_data.append(["TOTAL", criticas, "100%"])

    FOOTER_Y = M + 0.52*cm   # límite inferior seguro (justo sobre la línea del pie)
    _SEC_H   = 12            # altura aproximada del título de sección

    if crit_tv_data and cy_r_bot > FOOTER_Y + _SEC_H + 0.8*cm:
        nxt = sec_title("OTs >90 DÍAS POR TIPO DE VENTA", RIGHT_X, cy_r_bot, RIGHT_W, clr=ROJO)
        avail_ctv = nxt - FOOTER_Y
        ctv_cw = [RIGHT_W*0.56, RIGHT_W*0.22, RIGHT_W*0.22]
        _header  = crit_tv_data[0]
        _total   = crit_tv_data[-1]
        _rows    = crit_tv_data[1:-1]
        while len(_rows) >= 1:
            _t = make_tbl([_header] + _rows + [_total], ctv_cw, hdr_clr=ROJO, totals=True)
            _, _h = _t.wrapOn(c, RIGHT_W, 999)
            if _h <= avail_ctv:
                break
            _rows = _rows[:-1]
        if _rows:
            ctv_tbl = make_tbl([_header] + _rows + [_total], ctv_cw, hdr_clr=ROJO, totals=True)
            _, ctv_h = ctv_tbl.wrapOn(c, RIGHT_W, 999)
            ctv_tbl.drawOn(c, RIGHT_X, nxt - ctv_h)

    # ── PIE ───────────────────────────────────────────────
    fy = M + 0.08*cm
    c.saveState()
    c.setStrokeColor(AZUL_CLR)
    c.setLineWidth(0.35)
    c.line(M, fy + 0.38*cm, W - M, fy + 0.38*cm)
    c.setFont("Helvetica", 5.3)
    c.setFillColor(colors.HexColor("#94a3b8"))
    c.drawCentredString(W/2, fy + 0.08*cm,
        "Curifor S.A  ·  Sistema de Seguimiento OTs  ·  Documento Confidencial — Solo para uso interno")
    c.drawRightString(W - M, fy + 0.08*cm, f"Pág. 1/{TOTAL_PAGINAS}  ·  {fecha_act}")
    c.restoreState()

    c.showPage()

    # ═══════════════ PÁGINA 2: DESGLOSE POR SUCURSAL + COSTO VALE ═══════════════

    # ── Costo Vale de Consumo por OT ──────────────────────────────────────────
    def _safe_costo_vale(reps):
        if not isinstance(reps, list):
            try:
                import json as _json
                reps = _json.loads(reps) if isinstance(reps, str) else []
            except Exception:
                reps = []
        return sum(
            float(str(r.get("costo_total", 0) or 0).replace(",", "."))
            for r in (reps or [])
        )

    _df_p2 = df_inf.copy()
    if "repuestos_actual" in _df_p2.columns:
        _df_p2["_costo_vale"] = _df_p2["repuestos_actual"].apply(_safe_costo_vale)
    else:
        _df_p2["_costo_vale"] = 0.0

    neto_total_p2 = pd.to_numeric(_df_p2["NETO"], errors="coerce").fillna(0).sum() if "NETO" in _df_p2.columns else 0.0
    vale_total = float(_df_p2["_costo_vale"].sum())

    def _fmt_M(v):
        if v == 0: return "—"
        if abs(v) >= 1_000_000_000: return f"${v/1_000_000_000:.2f}B"
        if abs(v) >= 1_000_000:     return f"${v/1_000_000:.1f}M"
        if abs(v) >= 1_000:         return f"${v/1_000:.0f}K"
        return f"${v:.0f}"

    # ── Sucursal stats página 2 ────────────────────────────────────────────────
    rangos_p2 = ["0-30", "31-60", "61-90", "91 o más"]
    suc2_rows = []
    for _suc2, _grp2 in _df_p2.groupby("SUCURSAL"):
        _t2    = len(_grp2)
        _crit2 = int((_grp2["RANGO"] == "91 o más").sum())
        _dp2   = pd.to_numeric(_grp2.get("DIAS APERTURA", pd.Series(dtype=float)), errors="coerce").mean()
        _dp2s  = f"{_dp2:.0f}" if not pd.isna(_dp2) else "—"
        _pct2  = f"{_crit2/_t2*100:.1f}%" if _t2 else "0.0%"
        _neto2 = pd.to_numeric(_grp2["NETO"], errors="coerce").fillna(0).sum() if "NETO" in _grp2.columns else 0.0
        _vale2 = float(_grp2["_costo_vale"].sum())
        suc2_rows.append({
            "suc": str(_suc2)[:18],
            **{r: int((_grp2["RANGO"] == r).sum()) for r in rangos_p2},
            "total": _t2, "pct_crit": _pct2, "dias": _dp2s,
            "neto": _neto2, "vale": _vale2,
        })
    suc2_rows.sort(key=lambda x: x["total"], reverse=True)

    suc2_data = [["Sucursal", "0-30", "31-60", "61-90", ">90", "Total", "% Crít.", "Días Prom.", "Neto OTs", "Costo Vale Consumo"]]
    for _s2 in suc2_rows:
        suc2_data.append([
            _s2["suc"], _s2["0-30"], _s2["31-60"], _s2["61-90"], _s2["91 o más"],
            _s2["total"], _s2["pct_crit"], _s2["dias"],
            _fmt_M(_s2["neto"]), _fmt_M(_s2["vale"]),
        ])
    suc2_data.append([
        "TOTAL GLOBAL", recientes, atencion, urgentes, criticas,
        total, f"{idx_crit:.1f}%", f"{dias_prom:.0f}",
        _fmt_M(neto_total_p2), _fmt_M(vale_total),
    ])

    # ── Tipo venta × Costo Vale ────────────────────────────────────────────────
    tv_vale_data = [["Tipo de Venta", "N° OTs", "Neto OTs", "Costo Vale Consumo", "Vale / Neto"]]
    if "TIPO VENTA" in _df_p2.columns:
        _tv_rows_unsorted = []
        for _tv2, _g2 in _df_p2[_df_p2["TIPO VENTA"] != ""].groupby("TIPO VENTA"):
            _n2   = len(_g2)
            _nt2  = pd.to_numeric(_g2["NETO"], errors="coerce").fillna(0).sum() if "NETO" in _g2.columns else 0.0
            _v2   = float(_g2["_costo_vale"].sum())
            _ratio = f"{_v2/_nt2*100:.1f}%" if _nt2 > 0 else "—"
            _tv_rows_unsorted.append([str(_tv2)[:22], _n2, _fmt_M(_nt2), _fmt_M(_v2), _ratio])
        _tv_rows_unsorted.sort(key=lambda r: r[1], reverse=True)
        tv_vale_data += _tv_rows_unsorted
        _ratio_tot = f"{vale_total/neto_total_p2*100:.1f}%" if neto_total_p2 > 0 else "—"
        tv_vale_data.append(["TOTAL", total, _fmt_M(neto_total_p2), _fmt_M(vale_total), _ratio_tot])

    # ── HEADER página 2 ───────────────────────────────────────────────────────
    _hdr_y2 = H - M - HDR_H
    c.saveState()
    c.setFillColor(AZUL)
    c.roundRect(M, _hdr_y2, W - 2*M, HDR_H, 5, fill=1, stroke=0)
    c.restoreState()
    if logo_img_buf:
        try:
            logo_img_buf.seek(0)
            _pw2, _ph2 = 2.85*cm, HDR_H - 6
            c.saveState()
            c.setFillColor(BLANCO)
            c.roundRect(M + 5, _hdr_y2 + 3, _pw2, _ph2, 4, fill=1, stroke=0)
            c.restoreState()
            c.drawImage(ImageReader(logo_img_buf), M + 9, _hdr_y2 + 6,
                        width=_pw2 - 8, height=_ph2 - 6,
                        preserveAspectRatio=True, mask="auto")
        except Exception:
            pass
    _tx2 = M + 3.2*cm
    c.saveState()
    c.setFont("Helvetica-Bold", 10.5)
    c.setFillColor(BLANCO)
    c.drawString(_tx2, _hdr_y2 + HDR_H*0.58, "DESGLOSE POR SUCURSAL Y COSTO VALE DE CONSUMO — CURIFOR S.A")
    c.setFont("Helvetica", 6.3)
    c.setFillColor(colors.HexColor("#b0c8e8"))
    c.drawString(_tx2, _hdr_y2 + HDR_H*0.22, f"Generado: {fecha_act}   |   {filtros_desc[:100]}")
    c.setFont("Helvetica", 5.7)
    c.setFillColor(colors.HexColor("#8ab0d4"))
    c.drawRightString(W - M - 5, _hdr_y2 + HDR_H*0.44, f"Total Vale Consumo Global: {_fmt_M(vale_total)}")
    c.restoreState()

    # ── KPI STRIP página 2 ────────────────────────────────────────────────────
    _KPI_Y2 = _hdr_y2 - 0.25*cm
    _KPI_H2 = 1.42*cm
    _KPI_GAP2 = 0.16*cm
    _KW2 = (CW - 3*_KPI_GAP2) / 4

    _kpi2_defs = [
        (_fmt_M(vale_total),       "COSTO TOTAL VALE CONSUMO",  AZUL_CLR, "#eef4fc"),
        (_fmt_M(neto_total_p2),    "NETO TOTAL EN JUEGO",       AZUL_MED, "#e8f0fa"),
        (f"{total:,}",             "TOTAL OTs PENDIENTES",      AZUL,     "#dde9f7"),
        (f"{len(suc2_rows)}",      "SUCURSALES CON DATOS",      AZUL_CLR, "#eef4fc"),
    ]
    for _ki2, (_val2k, _lbl2k, _clr2k, _bg2k) in enumerate(_kpi2_defs):
        _kx2 = M + _ki2 * (_KW2 + _KPI_GAP2)
        _ky2 = _KPI_Y2 - _KPI_H2
        c.saveState()
        c.setFillColor(colors.HexColor(_bg2k))
        c.setStrokeColor(_clr2k)
        c.setLineWidth(0.5)
        c.roundRect(_kx2, _ky2, _KW2, _KPI_H2, 4, fill=1, stroke=1)
        c.setFillColor(_clr2k)
        c.setFont("Helvetica-Bold", 12)
        c.drawCentredString(_kx2 + _KW2/2, _ky2 + _KPI_H2*0.52, str(_val2k))
        c.setFont("Helvetica", 5.5)
        c.setFillColor(colors.HexColor("#64748b"))
        c.drawCentredString(_kx2 + _KW2/2, _ky2 + _KPI_H2*0.2, _lbl2k)
        c.restoreState()

    _cy_p2 = _KPI_Y2 - _KPI_H2 - 0.45*cm

    # ── Tabla sucursal detallada ───────────────────────────────────────────────
    _cy_p2 = sec_title("OTs PENDIENTES POR SUCURSAL — DETALLE CON COSTO VALE DE CONSUMO",
                        M, _cy_p2, CW, clr=AZUL_MED)
    _suc2_cw = [
        CW*0.135, CW*0.05, CW*0.055, CW*0.055, CW*0.055,
        CW*0.055, CW*0.065, CW*0.07, CW*0.13, CW*0.13,
    ]
    _suc2_tbl = make_tbl(suc2_data, _suc2_cw, totals=True, red_col=4)
    _, _suc2_h = _suc2_tbl.wrapOn(c, CW, 999)
    _suc2_tbl.drawOn(c, M, _cy_p2 - _suc2_h)
    _cy_p2 -= _suc2_h + 0.5*cm

    # ── Tabla tipo de venta × Costo Vale (izq) + Resumen global (der) ─────────
    if len(tv_vale_data) > 1 and _cy_p2 > M + 2*cm:
        _tv2_left_w  = CW * 0.50
        _tv2_right_w = CW * 0.46
        _tv2_gap     = CW * 0.04

        _cy_left  = _cy_p2
        _cy_right = _cy_p2

        _cy_left = sec_title("NETO Y COSTO VALE DE CONSUMO POR TIPO DE VENTA",
                              M, _cy_left, _tv2_left_w, clr=AZUL_MED)
        _tv_cw = [_tv2_left_w*0.37, _tv2_left_w*0.14, _tv2_left_w*0.20, _tv2_left_w*0.22, _tv2_left_w*0.07]
        _FOOTER_Y2 = M + 0.52*cm
        _avail_tv2 = _cy_left - _FOOTER_Y2
        _tv_hdr  = tv_vale_data[0]
        _tv_tot  = tv_vale_data[-1]
        _tv_rows = tv_vale_data[1:-1]
        while len(_tv_rows) >= 1:
            _tv_tbl2 = make_tbl([_tv_hdr] + _tv_rows + [_tv_tot], _tv_cw, totals=True)
            _, _tv2_h = _tv_tbl2.wrapOn(c, _tv2_left_w, 999)
            if _tv2_h <= _avail_tv2:
                break
            _tv_rows = _tv_rows[:-1]
        _tv_tbl2 = make_tbl([_tv_hdr] + _tv_rows + [_tv_tot], _tv_cw, totals=True)
        _, _tv2_h = _tv_tbl2.wrapOn(c, _tv2_left_w, 999)
        _tv_tbl2.drawOn(c, M, _cy_left - _tv2_h)

        _ley_x = M + _tv2_left_w + _tv2_gap
        _cy_right = sec_title("RESUMEN GLOBAL DEL INFORME",
                               _ley_x, _cy_right, _tv2_right_w, clr=AZUL_MED)

        _resumen_items = [
            ("Total OTs pendientes",          f"{total:,}"),
            ("OTs críticas (>90 días)",        f"{criticas:,}  ({idx_crit:.1f}%)"),
            ("Días promedio apertura",          f"{dias_prom:.0f} días"),
            ("Neto total en juego",             _fmt_M(neto_total_p2)),
            ("Costo total vale de consumo",     _fmt_M(vale_total)),
            ("Sucursales con OTs pendientes",   f"{len(suc2_rows)}"),
            ("Sin gestión / avance",            f"{sin_gestion:,} OTs"),
        ]
        c.saveState()
        c.setFillColor(colors.HexColor("#f0f5ff"))
        c.setStrokeColor(colors.HexColor("#c7d9f8"))
        c.setLineWidth(0.5)
        _box_h2 = len(_resumen_items) * 9.5 + 10
        c.roundRect(_ley_x, _cy_right - _box_h2, _tv2_right_w, _box_h2, 4, fill=1, stroke=1)
        c.restoreState()
        _ry2 = _cy_right - 6
        for _lbl2r, _val2r in _resumen_items:
            c.saveState()
            c.setFont("Helvetica-Bold", 5.8)
            c.setFillColor(AZUL_MED)
            c.drawString(_ley_x + 6, _ry2, _lbl2r + ":")
            c.setFont("Helvetica", 5.8)
            c.setFillColor(AZUL)
            c.drawRightString(_ley_x + _tv2_right_w - 6, _ry2, str(_val2r))
            c.restoreState()
            c.saveState()
            c.setStrokeColor(colors.HexColor("#dbeafe"))
            c.setLineWidth(0.3)
            c.line(_ley_x + 4, _ry2 - 1.5, _ley_x + _tv2_right_w - 4, _ry2 - 1.5)
            c.restoreState()
            _ry2 -= 9.5

    # ── PIE página 2 ──────────────────────────────────────────────────────────
    _fy2 = M + 0.08*cm
    c.saveState()
    c.setStrokeColor(AZUL_CLR)
    c.setLineWidth(0.35)
    c.line(M, _fy2 + 0.38*cm, W - M, _fy2 + 0.38*cm)
    c.setFont("Helvetica", 5.3)
    c.setFillColor(colors.HexColor("#94a3b8"))
    c.drawCentredString(W/2, _fy2 + 0.08*cm,
        "Curifor S.A  ·  Sistema de Seguimiento OTs  ·  Documento Confidencial — Solo para uso interno")
    c.drawRightString(W - M, _fy2 + 0.08*cm, f"Pág. 2/{TOTAL_PAGINAS}  ·  {fecha_act}")
    c.restoreState()

    c.showPage()

    # ═══ PÁGINA 3: DETALLE CRÍTICO (61-90 y 91 o más) POR TIPO DE VENTA Y CATEGORÍA ═══
    _RANGOS_CRIT = ["61-90", "91 o más"]
    _df_crit_pag = _df_p2[_df_p2["RANGO"].isin(_RANGOS_CRIT)].copy()
    _n_crit_pag  = len(_df_crit_pag)

    def _tabla_critica_por(col_nombre, titulo_col, ancho_col1):
        """Arma [Valor | 61-90 | % | 91 o más | % | Total crítico | % del crítico]."""
        _base = df_inf.copy()
        if col_nombre == "CATEGORIA":
            _serie = _base["CATEGORIA"].fillna("Sin categoría").replace("", "Sin categoría")
        else:
            _serie = _base[col_nombre].replace("", "Sin tipo")
        _base["_GRP"] = _serie
        _rows_out = []
        for _val, _g in _base.groupby("_GRP"):
            _n6190 = int((_g["RANGO"] == "61-90").sum())
            _n90m  = int((_g["RANGO"] == "91 o más").sum())
            _ntot  = _n6190 + _n90m
            if _ntot == 0:
                continue
            _rows_out.append({
                "val": str(_val)[:24], "n6190": _n6190, "n90m": _n90m, "ntot": _ntot,
            })
        _rows_out.sort(key=lambda r: r["ntot"], reverse=True)
        _tdata = [[titulo_col, "61-90", "%", "91 o más", "%", "Total", "% del crítico"]]
        for _r in _rows_out:
            _p6190 = f"{_r['n6190']/_n_crit_pag*100:.1f}%" if _n_crit_pag else "0.0%"
            _p90m  = f"{_r['n90m']/_n_crit_pag*100:.1f}%"  if _n_crit_pag else "0.0%"
            _ptot  = f"{_r['ntot']/_n_crit_pag*100:.1f}%"  if _n_crit_pag else "0.0%"
            _tdata.append([_r["val"], _r["n6190"], _p6190, _r["n90m"], _p90m, _r["ntot"], _ptot])
        _tdata.append([
            "TOTAL",
            sum(r["n6190"] for r in _rows_out), "100%",
            sum(r["n90m"]  for r in _rows_out), "100%",
            _n_crit_pag, "100%",
        ])
        return _tdata

    _tv_crit_data  = _tabla_critica_por("TIPO VENTA", "Tipo de Venta", 0)
    _cat_crit_data = _tabla_critica_por("CATEGORIA",  "Categoría", 0)

    _hdr_y3 = H - M - HDR_H
    c.saveState()
    c.setFillColor(AZUL)
    c.roundRect(M, _hdr_y3, W - 2*M, HDR_H, 5, fill=1, stroke=0)
    c.restoreState()
    if logo_img_buf:
        try:
            logo_img_buf.seek(0)
            _pw3, _ph3 = 2.85*cm, HDR_H - 6
            c.saveState()
            c.setFillColor(BLANCO)
            c.roundRect(M + 5, _hdr_y3 + 3, _pw3, _ph3, 4, fill=1, stroke=0)
            c.restoreState()
            c.drawImage(ImageReader(logo_img_buf), M + 9, _hdr_y3 + 6,
                        width=_pw3 - 8, height=_ph3 - 6,
                        preserveAspectRatio=True, mask="auto")
        except Exception:
            pass
    _tx3 = M + 3.2*cm
    c.saveState()
    c.setFont("Helvetica-Bold", 10.5)
    c.setFillColor(BLANCO)
    c.drawString(_tx3, _hdr_y3 + HDR_H*0.58,
                 "DETALLE CRÍTICO — OTs 61-90 Y 91 O MÁS DÍAS — CURIFOR S.A")
    c.setFont("Helvetica", 6.3)
    c.setFillColor(colors.HexColor("#b0c8e8"))
    c.drawString(_tx3, _hdr_y3 + HDR_H*0.22, f"Generado: {fecha_act}   |   {filtros_desc[:100]}")
    c.setFont("Helvetica", 5.7)
    c.setFillColor(colors.HexColor("#8ab0d4"))
    c.drawRightString(W - M - 5, _hdr_y3 + HDR_H*0.44,
                       f"Total OTs en rango crítico (61-90 + >90): {_n_crit_pag:,}")
    c.restoreState()

    # KPI strip: 61-90, >90, total crítico, % del total pendiente
    _KPI_Y3 = _hdr_y3 - 0.25*cm
    _KPI_H3 = 1.42*cm
    _KPI_GAP3 = 0.16*cm
    _KW3 = (CW - 3*_KPI_GAP3) / 4
    _pct_crit_pag = _n_crit_pag / total * 100 if total else 0
    _kpi3_defs = [
        (f"{urgentes:,}",       "OTs 61-90 DÍAS",               NARANJA, "#fff7ed"),
        (f"{criticas:,}",       "OTs 91 O MÁS DÍAS",            ROJO,    "#fef2f2"),
        (f"{_n_crit_pag:,}",    "TOTAL EN RANGO CRÍTICO",       AZUL_MED,"#e8f0fa"),
        (f"{_pct_crit_pag:.1f}%","% DEL TOTAL DE OTs PENDIENTES",AZUL_CLR,"#eef4fc"),
    ]
    for _ki3, (_val3k, _lbl3k, _clr3k, _bg3k) in enumerate(_kpi3_defs):
        _kx3 = M + _ki3 * (_KW3 + _KPI_GAP3)
        _ky3 = _KPI_Y3 - _KPI_H3
        c.saveState()
        c.setFillColor(colors.HexColor(_bg3k))
        c.setStrokeColor(_clr3k)
        c.setLineWidth(0.5)
        c.roundRect(_kx3, _ky3, _KW3, _KPI_H3, 4, fill=1, stroke=1)
        c.setFillColor(_clr3k)
        c.setFont("Helvetica-Bold", 15.5)
        c.drawCentredString(_kx3 + _KW3/2, _ky3 + _KPI_H3*0.46, str(_val3k))
        c.setFont("Helvetica-Bold", 4.8)
        c.setFillColor(colors.HexColor("#64748b"))
        c.drawCentredString(_kx3 + _KW3/2, _ky3 + _KPI_H3*0.16, _lbl3k)
        c.restoreState()

    _cy3 = _KPI_Y3 - _KPI_H3 - 0.4*cm
    _LEFT_W3  = CW * 0.49
    _RIGHT_W3 = CW * 0.49
    _RIGHT_X3 = M + _LEFT_W3 + CW*0.02

    _cy3_l = sec_title("OTs 61-90 y >90 días — por tipo de venta", M, _cy3, _LEFT_W3, clr=ROJO)
    _tvc_cw = [_LEFT_W3*0.28, _LEFT_W3*0.11, _LEFT_W3*0.11, _LEFT_W3*0.11,
               _LEFT_W3*0.11, _LEFT_W3*0.10, _LEFT_W3*0.18]
    _diff_tvc = _LEFT_W3 - sum(_tvc_cw)
    _tvc_cw[0] += _diff_tvc
    _tvc_tbl = make_tbl(_tv_crit_data, _tvc_cw, hdr_clr=ROJO, totals=True)
    _, _tvc_h = _tvc_tbl.wrapOn(c, _LEFT_W3, 999)
    _tvc_tbl.drawOn(c, M, _cy3_l - _tvc_h)

    _cy3_r = sec_title("OTs 61-90 y >90 días — por categoría", _RIGHT_X3, _cy3, _RIGHT_W3, clr=ROJO)
    _catc_cw = [_RIGHT_W3*0.28, _RIGHT_W3*0.11, _RIGHT_W3*0.11, _RIGHT_W3*0.11,
                _RIGHT_W3*0.11, _RIGHT_W3*0.10, _RIGHT_W3*0.18]
    _diff_catc = _RIGHT_W3 - sum(_catc_cw)
    _catc_cw[0] += _diff_catc
    _catc_tbl = make_tbl(_cat_crit_data, _catc_cw, hdr_clr=ROJO, totals=True)
    _, _catc_h = _catc_tbl.wrapOn(c, _RIGHT_W3, 999)
    _catc_tbl.drawOn(c, _RIGHT_X3, _cy3_r - _catc_h)

    _fy3 = M + 0.08*cm
    c.saveState()
    c.setStrokeColor(AZUL_CLR)
    c.setLineWidth(0.35)
    c.line(M, _fy3 + 0.38*cm, W - M, _fy3 + 0.38*cm)
    c.setFont("Helvetica", 5.3)
    c.setFillColor(colors.HexColor("#94a3b8"))
    c.drawCentredString(W/2, _fy3 + 0.08*cm,
        "Curifor S.A  ·  Sistema de Seguimiento OTs  ·  Documento Confidencial — Solo para uso interno")
    c.drawRightString(W - M, _fy3 + 0.08*cm, f"Pág. 3/{TOTAL_PAGINAS}  ·  {fecha_act}")
    c.restoreState()

    c.showPage()

    # ═══ PÁGINA 4: FACTURAS X PENDIENTES — CRUCE CON ABONOS/ANTICIPOS DE CLIENTE ═══
    def _parse_anticipo_pdf(_ot_row):
        _a = _ot_row.get("anticipo", {}) if hasattr(_ot_row, "get") else {}
        if not isinstance(_a, dict):
            if isinstance(_a, str):
                try:
                    _a = json.loads(_a)
                except Exception:
                    _a = {}
            else:
                _a = {}
        return _a

    _rows_fx_pdf = []
    for _, _ot4 in df_inf.iterrows():
        _ant4 = _parse_anticipo_pdf(_ot4)
        _monto_ant4 = float(_ant4.get("total", 0) or 0)
        _tiene_ant4 = bool(_ant4.get("tiene_saldo", False)) and _monto_ant4 > 0
        _suc4    = str(_ot4.get("SUCURSAL", "") or "").strip() or "Sin sucursal"
        _fol_ot4 = str(_ot4.get("FOLIO OT", "") or "").strip()
        _asesor4 = str(_ot4.get("ASESOR", "") or "").strip()
        _folios_fc4  = str(_ot4.get("FOLIOS_FACT_CLIENTE",  "") or "").strip()
        _folios_fca4 = str(_ot4.get("FOLIOS_FACT_COMPANIA", "") or "").strip()
        for _fnum4 in [f.strip() for f in _folios_fc4.split(",") if f.strip().upper().startswith("X")]:
            _rows_fx_pdf.append({"num": _fnum4, "tipo": "Factura Cliente", "sucursal": _suc4,
                                  "folio_ot": _fol_ot4, "asesor": _asesor4,
                                  "con_abono": _tiene_ant4, "monto_ant": _monto_ant4})
        for _fnum4 in [f.strip() for f in _folios_fca4.split(",") if f.strip().upper().startswith("X")]:
            _rows_fx_pdf.append({"num": _fnum4, "tipo": "Factura Compañía", "sucursal": _suc4,
                                  "folio_ot": _fol_ot4, "asesor": _asesor4,
                                  "con_abono": _tiene_ant4, "monto_ant": _monto_ant4})

    _tot_fx      = len(_rows_fx_pdf)
    _con_abono   = sum(1 for r in _rows_fx_pdf if r["con_abono"])
    _sin_abono   = _tot_fx - _con_abono
    _pct_abono4  = _con_abono / _tot_fx * 100 if _tot_fx else 0

    _hdr_y4 = H - M - HDR_H
    c.saveState()
    c.setFillColor(AZUL)
    c.roundRect(M, _hdr_y4, W - 2*M, HDR_H, 5, fill=1, stroke=0)
    c.restoreState()
    if logo_img_buf:
        try:
            logo_img_buf.seek(0)
            _pw4, _ph4 = 2.85*cm, HDR_H - 6
            c.saveState()
            c.setFillColor(BLANCO)
            c.roundRect(M + 5, _hdr_y4 + 3, _pw4, _ph4, 4, fill=1, stroke=0)
            c.restoreState()
            c.drawImage(ImageReader(logo_img_buf), M + 9, _hdr_y4 + 6,
                        width=_pw4 - 8, height=_ph4 - 6,
                        preserveAspectRatio=True, mask="auto")
        except Exception:
            pass
    _tx4 = M + 3.2*cm
    c.saveState()
    c.setFont("Helvetica-Bold", 10.5)
    c.setFillColor(BLANCO)
    c.drawString(_tx4, _hdr_y4 + HDR_H*0.58,
                 "FACTURAS X PENDIENTES — CRUCE CON ABONOS DE CLIENTE — CURIFOR S.A")
    c.setFont("Helvetica", 6.3)
    c.setFillColor(colors.HexColor("#b0c8e8"))
    c.drawString(_tx4, _hdr_y4 + HDR_H*0.22, f"Generado: {fecha_act}   |   {filtros_desc[:100]}")
    c.setFont("Helvetica", 5.7)
    c.setFillColor(colors.HexColor("#8ab0d4"))
    c.drawRightString(W - M - 5, _hdr_y4 + HDR_H*0.44, f"Total Facturas X: {_tot_fx:,}")
    c.restoreState()

    _KPI_Y4 = _hdr_y4 - 0.25*cm
    _KPI_H4 = 1.42*cm
    _KPI_GAP4 = 0.16*cm
    _KW4 = (CW - 3*_KPI_GAP4) / 4
    _kpi4_defs = [
        (f"{_tot_fx:,}",      "TOTAL FACTURAS X PENDIENTES", AZUL_CLR, "#eef4fc"),
        (f"{_con_abono:,}",   "CON ABONO / ANTICIPO",        VERDE,    "#f0fdf4"),
        (f"{_sin_abono:,}",   "SIN ABONO / ANTICIPO",        ROJO,     "#fef2f2"),
        (f"{_pct_abono4:.1f}%","% CON ABONO ASOCIADO",       AZUL_MED, "#e8f0fa"),
    ]
    for _ki4, (_val4k, _lbl4k, _clr4k, _bg4k) in enumerate(_kpi4_defs):
        _kx4 = M + _ki4*(_KW4 + _KPI_GAP4)
        _ky4 = _KPI_Y4 - _KPI_H4
        c.saveState()
        c.setFillColor(colors.HexColor(_bg4k))
        c.setStrokeColor(_clr4k)
        c.setLineWidth(0.5)
        c.roundRect(_kx4, _ky4, _KW4, _KPI_H4, 4, fill=1, stroke=1)
        c.setFillColor(_clr4k)
        c.setFont("Helvetica-Bold", 15.5)
        c.drawCentredString(_kx4 + _KW4/2, _ky4 + _KPI_H4*0.46, str(_val4k))
        c.setFont("Helvetica-Bold", 4.8)
        c.setFillColor(colors.HexColor("#64748b"))
        c.drawCentredString(_kx4 + _KW4/2, _ky4 + _KPI_H4*0.16, _lbl4k)
        c.restoreState()

    _cy4 = _KPI_Y4 - _KPI_H4 - 0.5*cm

    # --- Facturas X por sucursal: cantidad y cuántas tienen abono asociado ---
    _cy4 = sec_title("Facturas X por sucursal — cantidad y cuántas tienen abono asociado", M, _cy4, CW)
    _suc_fx_rows = {}
    for r in _rows_fx_pdf:
        d = _suc_fx_rows.setdefault(r["sucursal"], {"tot": 0, "con": 0})
        d["tot"] += 1
        if r["con_abono"]:
            d["con"] += 1
    _suc_fx_data = [["Sucursal", "Total Facturas X", "Con Abono", "Sin Abono", "% Con Abono"]]
    for _suc_k, _d in sorted(_suc_fx_rows.items(), key=lambda kv: kv[1]["tot"], reverse=True):
        _sin_k = _d["tot"] - _d["con"]
        _pct_k = f"{_d['con']/_d['tot']*100:.1f}%" if _d["tot"] else "0.0%"
        _suc_fx_data.append([str(_suc_k)[:22], _d["tot"], _d["con"], _sin_k, _pct_k])
    _suc_fx_data.append(["TOTAL", _tot_fx, _con_abono, _sin_abono,
                          f"{_pct_abono4:.1f}%"])
    _sfx_cw = [CW*0.30, CW*0.18, CW*0.17, CW*0.17, CW*0.18]
    _sfx_tbl = make_tbl(_suc_fx_data, _sfx_cw, totals=True, red_col=3)
    _, _sfx_h = _sfx_tbl.wrapOn(c, CW, 999)
    _sfx_tbl.drawOn(c, M, _cy4 - _sfx_h)
    _cy4 -= _sfx_h + 0.5*cm

    # --- Facturas X por tipo de documento: mismo cruce con abono ---
    _tipo_fx_rows = {}
    for r in _rows_fx_pdf:
        d = _tipo_fx_rows.setdefault(r["tipo"], {"tot": 0, "con": 0})
        d["tot"] += 1
        if r["con_abono"]:
            d["con"] += 1
    if _tipo_fx_rows:
        _cy4 = sec_title("Facturas X por tipo de documento — cantidad y cuántas tienen abono asociado",
                          M, _cy4, CW)
        _tipo_fx_data = [["Tipo de Documento", "Total Facturas X", "Con Abono", "Sin Abono", "% Con Abono"]]
        for _tipo_k, _d in sorted(_tipo_fx_rows.items(), key=lambda kv: kv[1]["tot"], reverse=True):
            _sin_k2 = _d["tot"] - _d["con"]
            _pct_k2 = f"{_d['con']/_d['tot']*100:.1f}%" if _d["tot"] else "0.0%"
            _tipo_fx_data.append([_tipo_k, _d["tot"], _d["con"], _sin_k2, _pct_k2])
        _tipo_fx_data.append(["TOTAL", _tot_fx, _con_abono, _sin_abono, f"{_pct_abono4:.1f}%"])
        _tfx_cw = [CW*0.30, CW*0.18, CW*0.17, CW*0.17, CW*0.18]
        _tfx_tbl = make_tbl(_tipo_fx_data, _tfx_cw, totals=True, red_col=3)
        _, _tfx_h = _tfx_tbl.wrapOn(c, CW, 999)
        _tfx_tbl.drawOn(c, M, _cy4 - _tfx_h)

    _fy4 = M + 0.08*cm
    c.saveState()
    c.setStrokeColor(AZUL_CLR)
    c.setLineWidth(0.35)
    c.line(M, _fy4 + 0.38*cm, W - M, _fy4 + 0.38*cm)
    c.setFont("Helvetica", 5.3)
    c.setFillColor(colors.HexColor("#94a3b8"))
    c.drawCentredString(W/2, _fy4 + 0.08*cm,
        "Curifor S.A  ·  Sistema de Seguimiento OTs  ·  Documento Confidencial — Solo para uso interno")
    c.drawRightString(W - M, _fy4 + 0.08*cm, f"Pág. 4/{TOTAL_PAGINAS}  ·  {fecha_act}")
    c.restoreState()

    c.showPage()

    # ═══════════════ PÁGINAS 5+: UNA PÁGINA DETALLADA POR SUCURSAL ═══════════════
    _pag_num = 4
    for _s2 in suc2_rows:   # ya viene ordenado por total desc (ver arriba)
        _pag_num += 1
        _suc_nombre = _s2["suc"]
        _grp_suc = _df_p2[_df_p2["SUCURSAL"].astype(str).str.strip().str[:18] == _suc_nombre]
        if _grp_suc.empty:
            # fallback por si el truncado a 18 caracteres genera colisión/():
            _grp_suc = _df_p2[_df_p2["SUCURSAL"] == _suc_nombre]
        _tot_suc  = len(_grp_suc)
        _crit_suc = int((_grp_suc["RANGO"] == "91 o más").sum())
        _urg_suc  = int((_grp_suc["RANGO"] == "61-90").sum())
        _neto_suc = pd.to_numeric(_grp_suc["NETO"], errors="coerce").fillna(0).sum() if "NETO" in _grp_suc.columns else 0.0
        _vale_suc = float(_grp_suc["_costo_vale"].sum()) if "_costo_vale" in _grp_suc.columns else 0.0

        # Header
        _hdr_ys = H - M - HDR_H
        c.saveState()
        c.setFillColor(AZUL)
        c.roundRect(M, _hdr_ys, W - 2*M, HDR_H, 5, fill=1, stroke=0)
        c.restoreState()
        if logo_img_buf:
            try:
                logo_img_buf.seek(0)
                _pws, _phs = 2.85*cm, HDR_H - 6
                c.saveState()
                c.setFillColor(BLANCO)
                c.roundRect(M + 5, _hdr_ys + 3, _pws, _phs, 4, fill=1, stroke=0)
                c.restoreState()
                c.drawImage(ImageReader(logo_img_buf), M + 9, _hdr_ys + 6,
                            width=_pws - 8, height=_phs - 6,
                            preserveAspectRatio=True, mask="auto")
            except Exception:
                pass
        _txs = M + 3.2*cm
        c.saveState()
        c.setFont("Helvetica-Bold", 10.5)
        c.setFillColor(BLANCO)
        c.drawString(_txs, _hdr_ys + HDR_H*0.58, f"DETALLE POR SUCURSAL — {_suc_nombre.upper()}")
        c.setFont("Helvetica", 6.3)
        c.setFillColor(colors.HexColor("#b0c8e8"))
        c.drawString(_txs, _hdr_ys + HDR_H*0.22, f"Generado: {fecha_act}   |   {filtros_desc[:90]}")
        c.setFont("Helvetica", 5.7)
        c.setFillColor(colors.HexColor("#8ab0d4"))
        c.drawRightString(W - M - 5, _hdr_ys + HDR_H*0.44, f"{_tot_suc:,} OT(s) pendientes en esta sucursal")
        c.restoreState()

        # KPI strip
        _KPI_Ys = _hdr_ys - 0.25*cm
        _KPI_Hs = 1.42*cm
        _KPI_GAPs = 0.16*cm
        _KWs = (CW - 4*_KPI_GAPs) / 5
        _pct_crit_s = _crit_suc / _tot_suc * 100 if _tot_suc else 0
        _kpis_defs = [
            (f"{_tot_suc:,}",     "TOTAL OTs SUCURSAL",     AZUL_CLR, "#eef4fc"),
            (f"{_crit_suc:,}",    "CRÍTICAS  >90 DÍAS",     ROJO,     "#fef2f2"),
            (f"{_urg_suc:,}",     "URGENTES  61-90 DÍAS",   NARANJA,  "#fff7ed"),
            (_fmt_M(_neto_suc),   "NETO TOTAL EN JUEGO",    AZUL_MED, "#e8f0fa"),
            (_fmt_M(_vale_suc),   "COSTO VALE DE CONSUMO",  AZUL,     "#dde9f7"),
        ]
        for _i_s, (_val_s, _lbl_s, _clr_s, _bg_s) in enumerate(_kpis_defs):
            _kx_s = M + _i_s*(_KWs + _KPI_GAPs)
            _ky_s = _KPI_Ys - _KPI_Hs
            c.saveState()
            c.setFillColor(colors.HexColor(_bg_s))
            c.setStrokeColor(GRIS_LIN)
            c.setLineWidth(0.4)
            c.roundRect(_kx_s, _ky_s, _KWs, _KPI_Hs, 4, fill=1, stroke=1)
            c.setFillColor(_clr_s)
            c.rect(_kx_s, _ky_s + _KPI_Hs - 3, _KWs, 3, fill=1, stroke=0)
            c.setFont("Helvetica-Bold", 13.5)
            c.drawCentredString(_kx_s + _KWs/2, _ky_s + _KPI_Hs*0.46, str(_val_s))
            c.setFont("Helvetica-Bold", 4.6)
            c.setFillColor(colors.HexColor("#64748b"))
            c.drawCentredString(_kx_s + _KWs/2, _ky_s + _KPI_Hs*0.16, _lbl_s)
            c.restoreState()

        _cys = _KPI_Ys - _KPI_Hs - 0.32*cm
        _LEFT_Ws  = CW * 0.42
        _RIGHT_Ws = CW * 0.55
        _RIGHT_Xs = M + _LEFT_Ws + CW*0.03

        # --- Columna izquierda: Costos y Netos (por rango y por asesor) + Categoría ---
        def _fmt_M_s(v):
            if v == 0: return "—"
            if abs(v) >= 1_000_000_000: return f"${v/1_000_000_000:.2f}B"
            if abs(v) >= 1_000_000:     return f"${v/1_000_000:.1f}M"
            if abs(v) >= 1_000:         return f"${v/1_000:.0f}K"
            return f"${v:.0f}"

        _FOOTER_Ys_L = M + 0.52*cm

        # 1) Costos y Netos por rango de días
        _cys_l = sec_title("Costos y Netos por rango de días", M, _cys, _LEFT_Ws)
        _cn_rango_rows = []
        for _r in rangos:
            _g_r = _grp_suc[_grp_suc["RANGO"] == _r]
            _n_r = len(_g_r)
            _neto_r = pd.to_numeric(_g_r["NETO"], errors="coerce").fillna(0).sum() if "NETO" in _g_r.columns else 0.0
            _vale_r = float(_g_r["_costo_vale"].sum()) if "_costo_vale" in _g_r.columns else 0.0
            _cn_rango_rows.append([_r, _n_r, _fmt_M_s(_neto_r), _fmt_M_s(_vale_r)])
        _cn_rango_data = ([["Rango", "N°", "Neto", "Vale Consumo"]] + _cn_rango_rows +
                           [["TOTAL", _tot_suc, _fmt_M_s(_neto_suc), _fmt_M_s(_vale_suc)]])
        _cnr_cw = [_LEFT_Ws*0.28, _LEFT_Ws*0.14, _LEFT_Ws*0.29, _LEFT_Ws*0.29]
        _cnr_tbl = make_tbl(_cn_rango_data, _cnr_cw, totals=True)
        _, _cnr_h = _cnr_tbl.wrapOn(c, _LEFT_Ws, 999)
        _cnr_tbl.drawOn(c, M, _cys_l - _cnr_h)
        _cys_l -= _cnr_h + 0.32*cm

        # 2) y 3) Costos y Netos por asesor + Por categoría — el espacio restante se
        #    reparte de forma PROPORCIONAL a lo que necesita cada tabla en su versión
        #    COMPLETA (sin recortar), en vez de que la de Asesor se quede con todo el
        #    espacio disponible y deje a Categoría resumida a 1-2 filas aunque sobre
        #    espacio en blanco en la página. Bug real reportado por Cristóbal: con
        #    sucursales de muchos asesores (ej. 10), categorías reales como GARANTIA
        #    quedaban escondidas dentro de "Otros" sin necesidad, con espacio de sobra
        #    sin usar debajo. Si ambas tablas completas caben en lo que queda, se
        #    muestran las 2 enteras sin recortar nada.
        _MIN_ESPACIO_L = _SEC_H + 26
        _SEC_TITLE_H = 9 + 0.32 * cm   # título de sección + separación entre tablas

        _as_cn_full = []
        if "ASESOR" in _grp_suc.columns:
            for _asesor_cn, _g_cn in _grp_suc[_grp_suc["ASESOR"] != ""].groupby("ASESOR"):
                _neto_cn = pd.to_numeric(_g_cn["NETO"], errors="coerce").fillna(0).sum() if "NETO" in _g_cn.columns else 0.0
                _vale_cn = float(_g_cn["_costo_vale"].sum()) if "_costo_vale" in _g_cn.columns else 0.0
                _as_cn_full.append({"asesor": str(_asesor_cn)[:18], "n": len(_g_cn), "neto": _neto_cn, "vale": _vale_cn})
            _as_cn_full.sort(key=lambda r: r["neto"], reverse=True)

        _cat_col_s = _grp_suc["CATEGORIA"].fillna("Sin categoría").replace("", "Sin categoría") if "CATEGORIA" in _grp_suc.columns else pd.Series(dtype=str)
        _cat_grp_s = _cat_col_s.groupby(_cat_col_s).size().sort_values(ascending=False) if len(_cat_col_s) else pd.Series(dtype=int)
        _cat_full = [{"cat": str(_cat)[:22], "n": int(_n)} for _cat, _n in _cat_grp_s.items()]

        _cn_as_cw = [_LEFT_Ws*0.34, _LEFT_Ws*0.12, _LEFT_Ws*0.27, _LEFT_Ws*0.27]
        _cd_cw    = [_LEFT_Ws*0.6, _LEFT_Ws*0.2, _LEFT_Ws*0.2]
        _hdr_as   = ["Asesor", "N°", "Neto", "Vale Consumo"]
        _hdr_cat  = ["Categoría", "N°", "%"]
        _tot_as   = ["TOTAL", sum(r["n"] for r in _as_cn_full),
                     _fmt_M_s(sum(r["neto"] for r in _as_cn_full)), _fmt_M_s(sum(r["vale"] for r in _as_cn_full))]
        _tot_cat  = ["TOTAL", sum(r["n"] for r in _cat_full), "100.0%"]

        def _fila_as(r):
            return [r["asesor"], r["n"], _fmt_M_s(r["neto"]), _fmt_M_s(r["vale"])]

        def _otros_as(extra):
            return [f"Otros ({len(extra)})", sum(r["n"] for r in extra),
                    _fmt_M_s(sum(r["neto"] for r in extra)), _fmt_M_s(sum(r["vale"] for r in extra))]

        def _fila_cat(r):
            return [r["cat"], r["n"], f"{r['n']/_tot_suc*100:.1f}%" if _tot_suc else "0.0%"]

        def _otros_cat(extra):
            _n_ex = sum(r["n"] for r in extra)
            return [f"Otros ({len(extra)})", _n_ex, f"{_n_ex/_tot_suc*100:.1f}%" if _tot_suc else "0.0%"]

        def _armar_tabla_izq(rows_full, hdr, fila_fn, otros_fn, cw, tot_row, avail_h):
            """Arma una tabla con recorte + fila 'Otros (N)' hasta que quepa en
            avail_h (o quede 1 sola fila) — nunca se pierde ningún dato del total."""
            if not rows_full:
                return None, 0
            _shown = list(rows_full)
            while True:
                _rows_d = [fila_fn(r) for r in _shown]
                _extra = rows_full[len(_shown):]
                _data = [hdr] + _rows_d + ([otros_fn(_extra)] if _extra else []) + [tot_row]
                _tbl = make_tbl(_data, cw, totals=True)
                _, _h = _tbl.wrapOn(c, sum(cw), 999)
                if _h <= avail_h or len(_shown) <= 1:
                    return _tbl, _h
                _shown = _shown[:-1]

        # Alturas naturales (sin recorte) de cada tabla, para decidir el reparto
        _, _h_as_full  = _armar_tabla_izq(_as_cn_full, _hdr_as, _fila_as, _otros_as, _cn_as_cw, _tot_as, 9999)
        _, _h_cat_full = _armar_tabla_izq(_cat_full, _hdr_cat, _fila_cat, _otros_cat, _cd_cw, _tot_cat, 9999)

        _avail_izq_resto = _cys_l - _FOOTER_Ys_L
        _necesita_as  = (_h_as_full + _SEC_TITLE_H) if _as_cn_full else 0
        _necesita_cat = (_h_cat_full + _SEC_TITLE_H) if _cat_full else 0
        _necesita_tot = _necesita_as + _necesita_cat

        if _necesita_tot == 0 or _necesita_tot <= _avail_izq_resto:
            # Alcanza para mostrar AMBAS tablas completas, sin recortar nada
            _cap_as, _cap_cat = 9999, 9999
        else:
            # No alcanza para las 2 completas: reparto proporcional a lo que cada
            # una necesita en su versión completa, con un piso de 30% para que
            # ninguna se quede sin espacio por culpa de la otra
            _prop_as = _necesita_as / _necesita_tot
            _prop_as = min(max(_prop_as, 0.30), 0.70)
            _cap_as  = _avail_izq_resto * _prop_as - _SEC_TITLE_H
            _cap_cat = _avail_izq_resto * (1 - _prop_as) - _SEC_TITLE_H

        # 2) Costos y Netos por asesor
        if _as_cn_full and _cap_as > 0:
            _cys_l = sec_title("Costos y Netos por asesor", M, _cys_l, _LEFT_Ws)
            _avail_as_real = min(_cap_as, _cys_l - _FOOTER_Ys_L)
            _tbl_as_cn, _h_as_cn = _armar_tabla_izq(_as_cn_full, _hdr_as, _fila_as, _otros_as,
                                                      _cn_as_cw, _tot_as, _avail_as_real)
            if _tbl_as_cn is not None and _h_as_cn <= _cys_l - _FOOTER_Ys_L:
                _tbl_as_cn.drawOn(c, M, _cys_l - _h_as_cn)
                _cys_l -= _h_as_cn + 0.32*cm

        # 3) Por categoría
        if _cat_full and _cap_cat > 0 and _cys_l - _FOOTER_Ys_L > _MIN_ESPACIO_L:
            _cys_l = sec_title("Por categoría", M, _cys_l, _LEFT_Ws)
            _avail_cat_real = min(_cap_cat, _cys_l - _FOOTER_Ys_L)
            _cd_tbl, _cd_h = _armar_tabla_izq(_cat_full, _hdr_cat, _fila_cat, _otros_cat,
                                                _cd_cw, _tot_cat, _avail_cat_real)
            if _cd_tbl is not None and _cd_h <= _cys_l - _FOOTER_Ys_L:
                _cd_tbl.drawOn(c, M, _cys_l - _cd_h)
                _cys_l -= _cd_h + 0.32*cm

        # 4) Si queda espacio abajo: OTs con Neto $0 — por asesor y rango de días
        if _cys_l - _FOOTER_Ys_L > _MIN_ESPACIO_L and "NETO" in _grp_suc.columns and "ASESOR" in _grp_suc.columns:
            _neto_num_s = pd.to_numeric(_grp_suc["NETO"], errors="coerce").fillna(0)
            _grp_neto0 = _grp_suc[_neto_num_s == 0]
            if not _grp_neto0.empty:
                _tmp0 = _grp_neto0.copy()
                _tmp0["_AS0"] = _tmp0["ASESOR"].replace("", "Sin asesor")
                _rows0_full = []
                for _as0, _g0 in _tmp0.groupby("_AS0"):
                    _n0_0 = int((_g0["RANGO"] == "0-30").sum())
                    _n0_1 = int((_g0["RANGO"] == "31-60").sum())
                    _n0_2 = int((_g0["RANGO"] == "61-90").sum())
                    _n0_3 = int((_g0["RANGO"] == "91 o más").sum())
                    _n0_tot = _n0_0 + _n0_1 + _n0_2 + _n0_3
                    if _n0_tot == 0:
                        continue
                    _rows0_full.append({"val": str(_as0)[:16], "n0": _n0_0, "n1": _n0_1,
                                         "n2": _n0_2, "n3": _n0_3, "ntot": _n0_tot})
                _rows0_full.sort(key=lambda r: r["ntot"], reverse=True)
                if _rows0_full:
                    _hdr0 = ["Asesor", "0-30", "31-60", "61-90", ">90", "Total"]
                    _tot_row0 = ["TOTAL",
                                 sum(r["n0"] for r in _rows0_full), sum(r["n1"] for r in _rows0_full),
                                 sum(r["n2"] for r in _rows0_full), sum(r["n3"] for r in _rows0_full),
                                 sum(r["ntot"] for r in _rows0_full)]
                    _cw0 = [_LEFT_Ws*0.30, _LEFT_Ws*0.14, _LEFT_Ws*0.14,
                            _LEFT_Ws*0.14, _LEFT_Ws*0.14, _LEFT_Ws*0.14]
                    _y0 = sec_title(f"OTs con Neto $0 — por asesor y rango de días ({len(_grp_neto0)})",
                                     M, _cys_l, _LEFT_Ws, clr=NARANJA)
                    _avail0 = _y0 - _FOOTER_Ys_L
                    _shown0 = list(_rows0_full)
                    while True:
                        _data0_rows = [[r["val"], r["n0"], r["n1"], r["n2"], r["n3"], r["ntot"]] for r in _shown0]
                        _extra0 = _rows0_full[len(_shown0):]
                        if _extra0:
                            _agg0 = [f"Otros ({len(_extra0)})",
                                     sum(r["n0"] for r in _extra0), sum(r["n1"] for r in _extra0),
                                     sum(r["n2"] for r in _extra0), sum(r["n3"] for r in _extra0),
                                     sum(r["ntot"] for r in _extra0)]
                            _full0 = [_hdr0] + _data0_rows + [_agg0, _tot_row0]
                        else:
                            _full0 = [_hdr0] + _data0_rows + [_tot_row0]
                        _tbl0 = make_tbl(_full0, _cw0, hdr_clr=NARANJA, totals=True)
                        _, _h0 = _tbl0.wrapOn(c, _LEFT_Ws, 999)
                        if _h0 <= _avail0 or len(_shown0) <= 1:
                            break
                        _shown0 = _shown0[:-1]
                    if _h0 <= _avail0:
                        _tbl0.drawOn(c, M, _y0 - _h0)
                        _cys_l = _y0 - _h0 - 0.32*cm

        # --- Columna derecha: OTs por Asesor y por Tipo de Venta, con TODOS los rangos ---
        _cys_r = _cys
        _FOOTER_Ys = M + 0.52*cm

        def _rango_rows_por(col_nombre, top_n=None):
            """Agrupa _grp_suc por col_nombre, contando OTs en cada uno de los 4 rangos."""
            if col_nombre not in _grp_suc.columns:
                return []
            _serie = _grp_suc[col_nombre].replace("", "Sin dato")
            _tmp = _grp_suc.copy()
            _tmp["_GRP"] = _serie
            _out = []
            for _val, _g in _tmp.groupby("_GRP"):
                _n0 = int((_g["RANGO"] == "0-30").sum())
                _n1 = int((_g["RANGO"] == "31-60").sum())
                _n2 = int((_g["RANGO"] == "61-90").sum())
                _n3 = int((_g["RANGO"] == "91 o más").sum())
                _ntot = _n0 + _n1 + _n2 + _n3
                if _ntot == 0:
                    continue
                _out.append({"val": str(_val)[:20], "n0": _n0, "n1": _n1, "n2": _n2, "n3": _n3, "ntot": _ntot})
            _out.sort(key=lambda r: r["ntot"], reverse=True)
            return _out

        def _dibujar_tabla_rango(rows_full, titulo, y_top, avail_h, hdr_clr=AZUL_MED):
            """Dibuja una tabla [Valor|0-30|31-60|61-90|>90|Total], recortando filas y
            agregando un renglón 'Otros (N)' si no caben todas — nunca se pierde OT alguna,
            nunca queda una tabla cortada a media fila."""
            if not rows_full:
                return y_top
            _hdr = ["", "0-30", "31-60", "61-90", ">90", "Total"]
            _tot_row = ["TOTAL",
                        sum(r["n0"] for r in rows_full), sum(r["n1"] for r in rows_full),
                        sum(r["n2"] for r in rows_full), sum(r["n3"] for r in rows_full),
                        sum(r["ntot"] for r in rows_full)]
            _cw = [_RIGHT_Ws*0.34, _RIGHT_Ws*0.13, _RIGHT_Ws*0.13,
                   _RIGHT_Ws*0.13, _RIGHT_Ws*0.13, _RIGHT_Ws*0.14]
            _y = sec_title(titulo, _RIGHT_Xs, y_top, _RIGHT_Ws, clr=hdr_clr)
            _avail = _y - _FOOTER_Ys if avail_h is None else min(avail_h, _y - _FOOTER_Ys)
            _shown = list(rows_full)
            while True:
                _data_rows = [[r["val"], r["n0"], r["n1"], r["n2"], r["n3"], r["ntot"]] for r in _shown]
                _extra = rows_full[len(_shown):]
                if _extra:
                    _agg = ["Otros (%d)" % len(_extra),
                            sum(r["n0"] for r in _extra), sum(r["n1"] for r in _extra),
                            sum(r["n2"] for r in _extra), sum(r["n3"] for r in _extra),
                            sum(r["ntot"] for r in _extra)]
                    _full_data = [_hdr] + _data_rows + [_agg, _tot_row]
                else:
                    _full_data = [_hdr] + _data_rows + [_tot_row]
                _tbl = make_tbl(_full_data, _cw, hdr_clr=hdr_clr, totals=True)
                _, _h = _tbl.wrapOn(c, _RIGHT_Ws, 999)
                if _h <= _avail or len(_shown) <= 1:
                    break
                _shown = _shown[:-1]
            _tbl.drawOn(c, _RIGHT_Xs, _y - _h)
            return _y - _h - 0.32*cm

        def _tabla_asesor_tipoventa(y_top, avail_h, max_cols_tv=5):
            """Cruce Asesor × Tipo de Venta: cuántas OTs tiene cada asesor de la sucursal
            por cada tipo de venta. Los tipos de venta menos frecuentes se agrupan en
            'Otros' para no desbordar el ancho de la tabla; los asesores que no caben
            se agregan en una fila 'Otros asesores (N)' — igual que el resto de tablas
            de este informe, nunca se pierde ninguna OT del conteo."""
            if "ASESOR" not in _grp_suc.columns or "TIPO VENTA" not in _grp_suc.columns:
                return y_top
            _tmp = _grp_suc.copy()
            _tmp["_AS"] = _tmp["ASESOR"].replace("", "Sin asesor")
            _tmp["_TV"] = _tmp["TIPO VENTA"].replace("", "Sin tipo")
            _tv_counts = _tmp["_TV"].value_counts()
            if _tv_counts.empty:
                return y_top
            _top_tv  = list(_tv_counts.head(max_cols_tv).index)
            _hay_otr = len(_tv_counts) > len(_top_tv)
            _tmp["_TVB"] = _tmp["_TV"].apply(lambda t: t if t in _top_tv else "Otros")
            _cols_tv = _top_tv + (["Otros"] if _hay_otr else [])
            _pivot = _tmp.groupby(["_AS", "_TVB"]).size().unstack(fill_value=0)
            for _col in _cols_tv:
                if _col not in _pivot.columns:
                    _pivot[_col] = 0
            _pivot = _pivot[_cols_tv]
            _pivot["Total"] = _pivot.sum(axis=1)
            _pivot = _pivot.sort_values("Total", ascending=False)

            _hdr = ["Asesor"] + [str(t)[:11] for t in _cols_tv] + ["Total"]
            _rows_full = [[str(_as)[:16]] + [int(row[c]) for c in _cols_tv] + [int(row["Total"])]
                          for _as, row in _pivot.iterrows()]
            _tot_row = ["TOTAL"] + [int(_pivot[c].sum()) for c in _cols_tv] + [int(_pivot["Total"].sum())]
            _n_cols  = len(_hdr)
            _cw = [_RIGHT_Ws*0.24] + [_RIGHT_Ws*0.76/(_n_cols-1)]*(_n_cols-1)

            _y = sec_title("OTs por asesor y tipo de venta", _RIGHT_Xs, y_top, _RIGHT_Ws)
            _avail = _y - _FOOTER_Ys if avail_h is None else min(avail_h, _y - _FOOTER_Ys)
            _shown = list(_rows_full)
            while True:
                _extra = _rows_full[len(_shown):]
                if _extra:
                    _agg = [f"Otros asesores ({len(_extra)})"] + [
                        sum(r[i] for r in _extra) for i in range(1, _n_cols)
                    ]
                    _data = [_hdr] + _shown + [_agg, _tot_row]
                else:
                    _data = [_hdr] + _shown + [_tot_row]
                _tbl2 = make_tbl(_data, _cw, totals=True)
                _, _h2 = _tbl2.wrapOn(c, _RIGHT_Ws, 999)
                if _h2 <= _avail or len(_shown) <= 1:
                    break
                _shown = _shown[:-1]
            _tbl2.drawOn(c, _RIGHT_Xs, _y - _h2)
            return _y - _h2 - 0.32*cm

        _rows_asesor = _rango_rows_por("ASESOR")

        # Se reparte el espacio disponible en 2 mitades (asesor/rango arriba,
        # asesor/tipo de venta abajo) para que ninguna de las 2 tablas quede cortada.
        _avail_total_der = _cys_r - _FOOTER_Ys
        _avail_asesor = _avail_total_der * 0.55 if _rows_asesor else _avail_total_der
        _cys_r = _dibujar_tabla_rango(_rows_asesor, "OTs por asesor y rango de días",
                                       _cys_r, _avail_asesor)
        _cys_r = _tabla_asesor_tipoventa(_cys_r, None)

        # Footer
        _fys = M + 0.08*cm
        c.saveState()
        c.setStrokeColor(AZUL_CLR)
        c.setLineWidth(0.35)
        c.line(M, _fys + 0.38*cm, W - M, _fys + 0.38*cm)
        c.setFont("Helvetica", 5.3)
        c.setFillColor(colors.HexColor("#94a3b8"))
        c.drawCentredString(W/2, _fys + 0.08*cm,
            "Curifor S.A  ·  Sistema de Seguimiento OTs  ·  Documento Confidencial — Solo para uso interno")
        c.drawRightString(W - M, _fys + 0.08*cm, f"Pág. {_pag_num}/{TOTAL_PAGINAS}  ·  {fecha_act}")
        c.restoreState()

        c.showPage()

    c.save()
    buf.seek(0)
    return buf.getvalue()


def generar_pdf_sucursal_anio(df_inf, filtros_desc, fecha_act, logo_b64_str):
    """PDF — UNA página A4 landscape, mismo estilo visual del informe ejecutivo:
    OTs por Sucursal × Año de Apertura, Costo del Vale de Consumo por Sucursal ×
    Año de Apertura (repuestos_actual, campo costo_total — ya viene multiplicado
    por cantidad, no se debe volver a multiplicar por costo_unitario) y Cantidad
    de OTs por Marca.

    El año de apertura sale de la columna "AÑO" del PBI (Seguimiento Servicio
    Técnico, columna BF) — NO se deriva de FECHA OT. Las OT sin ese dato (error
    de sistema / no vino en el PBI) se agrupan aparte, con etiqueta
    "Desconocida / Error sistema", en vez de perderse o mezclarse mal."""
    (A4, landscape, colors, cm, mm, rl_canvas, Table, TableStyle,
     plt, mpatches, ImageReader) = _importar_pdf_libs()

    # ── Paleta corporativa (misma que el informe ejecutivo) ────
    AZUL     = colors.HexColor("#0c243d")
    AZUL_MED = colors.HexColor("#1a3a5c")
    AZUL_CLR = colors.HexColor("#4a7ab5")
    ROJO     = colors.HexColor("#ef4444")
    GRIS_ALT = colors.HexColor("#f1f5f9")
    GRIS_LIN = colors.HexColor("#cbd5e1")
    BLANCO   = colors.white

    PAGE_W, PAGE_H = landscape(A4)
    M = 1.1 * cm

    buf = io.BytesIO()
    c = rl_canvas.Canvas(buf, pagesize=landscape(A4))
    c.setTitle("Informe OTs por Sucursal y Año — Curifor S.A")

    total = len(df_inf)
    if total == 0:
        c.showPage(); c.save(); buf.seek(0); return buf.getvalue()

    # ── Año de apertura: columna "AÑO" del PBI (Seguimiento Servicio Técnico,
    # columna BF) — es el dato correcto, no se debe derivar de FECHA OT.
    _ANIO_DESCONOCIDO = "Desconocida /\nError sistema"

    def _anio_de(v):
        s = str(v).strip()
        if not s or s.upper() in ("NAN", "NONE", "NAT", "0", "0.0"):
            return _ANIO_DESCONOCIDO
        try:
            s_num = str(int(float(s)))
        except Exception:
            return _ANIO_DESCONOCIDO
        return s_num if len(s_num) == 4 else _ANIO_DESCONOCIDO

    # Costo del Vale de Consumo — suma de costo_total por repuesto (el campo
    # costo_total ya viene como cantidad × costo_unitario; no se recalcula).
    def _safe_costo_vale(reps):
        if not isinstance(reps, list):
            try:
                import json as _json
                reps = _json.loads(reps) if isinstance(reps, str) else []
            except Exception:
                reps = []
        return sum(
            float(str(r.get("costo_total", 0) or 0).replace(",", "."))
            for r in (reps or [])
        )

    def _cat_de(v):
        s = str(v).strip()
        return s if s and s.upper() not in ("NAN", "NONE") else "Sin categoría"

    _df = df_inf.copy()
    _df["_ANIO"] = _df["AÑO"].apply(_anio_de) if "AÑO" in _df.columns else _ANIO_DESCONOCIDO
    _df["_COSTO_VALE"] = (_df["repuestos_actual"].apply(_safe_costo_vale)
                           if "repuestos_actual" in _df.columns else 0.0)
    _df["_CATEGORIA"] = (_df["CATEGORIA"].apply(_cat_de)
                          if "CATEGORIA" in _df.columns else "Sin categoría")

    anios = sorted([a for a in _df["_ANIO"].unique() if a != _ANIO_DESCONOCIDO], key=int)
    anios_cols = anios + [_ANIO_DESCONOCIDO] if (_df["_ANIO"] == _ANIO_DESCONOCIDO).any() else anios
    if not anios_cols:
        anios_cols = [_ANIO_DESCONOCIDO]

    def _fmt_M(v):
        if v == 0: return "—"
        if abs(v) >= 1_000_000_000: return f"${v/1_000_000_000:.2f}B"
        if abs(v) >= 1_000_000:     return f"${v/1_000_000:.1f}M"
        if abs(v) >= 1_000:         return f"${v/1_000:.0f}K"
        return f"${v:.0f}"

    # ── Pivot: cantidad de OTs por Sucursal × Año ──────────────
    pivot_cant = _df.groupby(["SUCURSAL", "_ANIO"]).size().unstack(fill_value=0)
    for a in anios_cols:
        if a not in pivot_cant.columns:
            pivot_cant[a] = 0
    pivot_cant = pivot_cant[anios_cols]
    pivot_cant["Total"] = pivot_cant.sum(axis=1)
    pivot_cant = pivot_cant.sort_values("Total", ascending=False)

    cant_data = [["Sucursal"] + anios_cols + ["Total"]]
    for suc, row in pivot_cant.iterrows():
        cant_data.append([str(suc)[:20]] + [int(row[a]) for a in anios_cols] + [int(row["Total"])])
    cant_data.append(["TOTAL"] + [int(pivot_cant[a].sum()) for a in anios_cols] + [int(pivot_cant["Total"].sum())])

    # ── Pivot: costo Vale de Consumo por Sucursal × Año ────────
    pivot_costo = _df.groupby(["SUCURSAL", "_ANIO"])["_COSTO_VALE"].sum().unstack(fill_value=0)
    for a in anios_cols:
        if a not in pivot_costo.columns:
            pivot_costo[a] = 0
    pivot_costo = pivot_costo[anios_cols]
    pivot_costo["Total"] = pivot_costo.sum(axis=1)
    pivot_costo = pivot_costo.reindex(pivot_cant.index)  # mismo orden de sucursales

    costo_data = [["Sucursal"] + anios_cols + ["Total"]]
    for suc, row in pivot_costo.iterrows():
        costo_data.append([str(suc)[:20]] + [_fmt_M(row[a]) for a in anios_cols] + [_fmt_M(row["Total"])])
    costo_data.append(["TOTAL"] + [_fmt_M(pivot_costo[a].sum()) for a in anios_cols] + [_fmt_M(pivot_costo["Total"].sum())])

    vale_total = float(pivot_costo["Total"].sum())

    # ── Pivot Sucursal × Categoría × Año (hoja/página nueva) ───
    # 3 dimensiones no caben en una sola tabla plana normal, así que se
    # muestra 1 fila por combinación (Sucursal, Categoría) que tenga al
    # menos 1 OT, con una columna por Año (mismas anios_cols de arriba) +
    # Total — ordenadas igual que la tabla de cantidad de arriba (misma
    # sucursal primero, más "ocupada" primero; dentro de cada sucursal,
    # la categoría con más OTs primero).
    _cat_pivot = _df.groupby(["SUCURSAL", "_CATEGORIA", "_ANIO"]).size().unstack(fill_value=0)
    for a in anios_cols:
        if a not in _cat_pivot.columns:
            _cat_pivot[a] = 0
    _cat_pivot = _cat_pivot[anios_cols]
    _cat_pivot["Total"] = _cat_pivot.sum(axis=1)
    _cat_pivot = _cat_pivot[_cat_pivot["Total"] > 0].reset_index()

    _orden_suc = list(pivot_cant.index)
    _cat_pivot["_ord_suc"] = _cat_pivot["SUCURSAL"].apply(
        lambda s: _orden_suc.index(s) if s in _orden_suc else len(_orden_suc))
    _cat_pivot = _cat_pivot.sort_values(["_ord_suc", "Total"], ascending=[True, False])

    cat_rows_all = [
        [str(r["SUCURSAL"])[:18], str(r["_CATEGORIA"])[:22]]
        + [int(r[a]) for a in anios_cols] + [int(r["Total"])]
        for _, r in _cat_pivot.iterrows()
    ]
    _cat_hdr_row = ["Sucursal", "Categoría"] + anios_cols + ["Total"]

    _CW_early = PAGE_W - 2 * M   # mismo valor que CW mas abajo (aun no definido aca)
    _n_year_cols_cat = len(anios_cols)
    _cat_lbl1_w = _CW_early * 0.15
    _cat_lbl2_w = _CW_early * 0.19
    _cat_rest_w = _CW_early - _cat_lbl1_w - _cat_lbl2_w
    cat_cw = ([_cat_lbl1_w, _cat_lbl2_w]
              + [_cat_rest_w / (_n_year_cols_cat + 1)] * _n_year_cols_cat
              + [_cat_rest_w / (_n_year_cols_cat + 1)])
    cat_cw[-1] += _CW_early - sum(cat_cw)

    def make_cat_tbl(data):
        style = [
            ("BACKGROUND",   (0, 0), (-1, 0), AZUL_MED),
            ("TEXTCOLOR",    (0, 0), (-1, 0), BLANCO),
            ("FONTNAME",     (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE",     (0, 0), (-1, 0), 6.2),
            ("FONTNAME",     (0, 1), (-1, -1), "Helvetica"),
            ("FONTSIZE",     (0, 1), (-1, -1), 6.2),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [BLANCO, GRIS_ALT]),
            ("GRID",         (0, 0), (-1, -1), 0.25, GRIS_LIN),
            ("LINEBELOW",    (0, 0), (-1, 0), 0.7, AZUL_CLR),
            ("TOPPADDING",   (0, 0), (-1, -1), 2.2),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 2.2),
            ("LEFTPADDING",  (0, 0), (-1, -1), 3.5),
            ("RIGHTPADDING", (0, 0), (-1, -1), 3),
            ("ALIGN",        (0, 0), (1, -1), "LEFT"),
            ("ALIGN",        (2, 0), (-1, -1), "CENTER"),
            ("VALIGN",       (0, 0), (-1, -1), "MIDDLE"),
        ]
        t = Table([[str(cell) for cell in row] for row in data], colWidths=cat_cw)
        t.setStyle(TableStyle(style))
        return t

    # ── Cantidad de OTs por Marca (todas, no solo top N) ───────
    marca_grp = (_df[_df["MARCA"] != ""].groupby("MARCA").size().sort_values(ascending=False)
                 if "MARCA" in _df.columns else pd.Series(dtype=int))
    marca_rows = [[str(m)[:22], int(n), f"{n/total*100:.1f}%"] for m, n in marca_grp.items()]
    n_marcas = len(marca_rows)

    # ── Logo ────────────────────────────────────────────────────
    logo_img_buf = None
    if logo_b64_str:
        try:
            logo_img_buf = io.BytesIO(base64.b64decode(logo_b64_str))
        except Exception:
            pass

    # ── Helper: título de sección ──────────────────────────────
    def sec_title(txt, x, y, w, clr=AZUL_MED):
        c.saveState()
        c.setFillColor(clr)
        c.setFont("Helvetica-Bold", 6.2)
        c.drawString(x, y, txt.upper())
        c.setStrokeColor(clr)
        c.setLineWidth(0.65)
        c.line(x, y - 2, x + w, y - 2)
        c.restoreState()
        return y - 9

    # ── Helper: tabla Platypus ──────────────────────────────────
    def make_tbl(data, cw, hdr_clr=AZUL_MED, totals=False):
        style = [
            ("BACKGROUND",   (0, 0), (-1, 0), hdr_clr),
            ("TEXTCOLOR",    (0, 0), (-1, 0), BLANCO),
            ("FONTNAME",     (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE",     (0, 0), (-1, 0), 6.2),
            ("FONTNAME",     (0, 1), (-1, -1), "Helvetica"),
            ("FONTSIZE",     (0, 1), (-1, -1), 6.2),
            ("ROWBACKGROUNDS", (0, 1), (-1, -2 if totals else -1), [BLANCO, GRIS_ALT]),
            ("GRID",         (0, 0), (-1, -1), 0.25, GRIS_LIN),
            ("LINEBELOW",    (0, 0), (-1, 0), 0.7, AZUL_CLR),
            ("TOPPADDING",   (0, 0), (-1, -1), 2.2),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 2.2),
            ("LEFTPADDING",  (0, 0), (-1, -1), 3.5),
            ("RIGHTPADDING", (0, 0), (-1, -1), 3),
            ("ALIGN",        (0, 0), (0, -1), "LEFT"),
            ("ALIGN",        (1, 0), (-1, -1), "CENTER"),
            ("VALIGN",       (0, 0), (-1, -1), "MIDDLE"),
        ]
        if totals:
            style += [
                ("BACKGROUND", (0, -1), (-1, -1), AZUL),
                ("TEXTCOLOR",  (0, -1), (-1, -1), BLANCO),
                ("FONTNAME",   (0, -1), (-1, -1), "Helvetica-Bold"),
                ("LINEABOVE",  (0, -1), (-1, -1), 0.7, AZUL_CLR),
            ]
        t = Table([[str(cell) for cell in row] for row in data], colWidths=cw)
        t.setStyle(TableStyle(style))
        return t

    # ═══════════════ DIBUJAR PÁGINA ═══════════════
    W, H = PAGE_W, PAGE_H

    # ── HEADER ──────────────────────────────────────────────────
    HDR_H = 1.58 * cm
    hdr_y = H - M - HDR_H
    c.saveState()
    c.setFillColor(AZUL)
    c.roundRect(M, hdr_y, W - 2*M, HDR_H, 5, fill=1, stroke=0)
    c.restoreState()

    if logo_img_buf:
        try:
            pw, ph = 2.85*cm, HDR_H - 6
            px0 = M + 5
            py0 = hdr_y + 3
            c.saveState()
            c.setFillColor(BLANCO)
            c.roundRect(px0, py0, pw, ph, 4, fill=1, stroke=0)
            c.restoreState()
            c.drawImage(ImageReader(logo_img_buf), px0 + 4, py0 + 3,
                        width=pw - 8, height=ph - 6,
                        preserveAspectRatio=True, mask="auto")
        except Exception:
            pass

    tx = M + 3.2*cm
    c.saveState()
    c.setFont("Helvetica-Bold", 10.5)
    c.setFillColor(BLANCO)
    c.drawString(tx, hdr_y + HDR_H*0.58, "OTs PENDIENTES POR SUCURSAL Y AÑO DE APERTURA — CURIFOR S.A")
    c.setFont("Helvetica", 6.3)
    c.setFillColor(colors.HexColor("#b0c8e8"))
    c.drawString(tx, hdr_y + HDR_H*0.22, f"Generado: {fecha_act}   |   Datos al: {fecha_act}")
    c.setFont("Helvetica", 5.7)
    c.setFillColor(colors.HexColor("#8ab0d4"))
    c.drawRightString(W - M - 5, hdr_y + HDR_H*0.44, filtros_desc[:120])
    c.restoreState()

    # ── KPI STRIP ───────────────────────────────────────────────
    KPI_Y = hdr_y - 0.25*cm
    KPI_H = 1.42*cm
    KPI_GAP = 0.16*cm
    CW = W - 2*M
    KW = (CW - 3*KPI_GAP) / 4

    kpi_defs = [
        (f"{total:,}",              "TOTAL OTs PENDIENTES",       AZUL_CLR, "#eef4fc"),
        (_fmt_M(vale_total),        "COSTO TOTAL VALE CONSUMO",   AZUL_MED, "#e8f0fa"),
        (f"{len(pivot_cant)}",      "SUCURSALES CON DATOS",       AZUL,     "#dde9f7"),
        (f"{n_marcas}",             "MARCAS DISTINTAS",           AZUL_CLR, "#eef4fc"),
    ]
    for i, (val, lbl, top, bg) in enumerate(kpi_defs):
        kx = M + i*(KW + KPI_GAP)
        ky = KPI_Y - KPI_H
        c.saveState()
        c.setFillColor(colors.HexColor(bg))
        c.setStrokeColor(GRIS_LIN)
        c.setLineWidth(0.4)
        c.roundRect(kx, ky, KW, KPI_H, 4, fill=1, stroke=1)
        c.setFillColor(top)
        c.rect(kx, ky + KPI_H - 3, KW, 3, fill=1, stroke=0)
        c.setFont("Helvetica-Bold", 15.5)
        c.setFillColor(top)
        c.drawCentredString(kx + KW/2, ky + KPI_H*0.46, val)
        c.setFont("Helvetica-Bold", 4.8)
        c.setFillColor(colors.HexColor("#64748b"))
        c.drawCentredString(kx + KW/2, ky + KPI_H*0.16, lbl)
        c.restoreState()

    # ── CONTENT AREA ────────────────────────────────────────────
    CY = KPI_Y - KPI_H - 0.28*cm
    FOOTER_Y = M + 0.52*cm

    LEFT_W  = CW * 0.485
    RIGHT_W = CW * 0.485
    RIGHT_X = M + LEFT_W + CW*0.03

    n_year_cols = len(anios_cols)

    # Tabla cantidad (izquierda)
    next_l = sec_title("OTs por sucursal y año de apertura", M, CY, LEFT_W)
    cant_cw = [LEFT_W*0.30] + [LEFT_W*0.70/(n_year_cols+1)]*n_year_cols + [LEFT_W*0.70/(n_year_cols+1)]
    diff_c = LEFT_W - sum(cant_cw)
    cant_cw[-1] += diff_c
    cant_tbl = make_tbl(cant_data, cant_cw, totals=True)
    _, cant_h = cant_tbl.wrapOn(c, LEFT_W, 999)
    cant_tbl.drawOn(c, M, next_l - cant_h)

    # Tabla costo (derecha)
    next_r = sec_title("Costo Vale de Consumo por sucursal y año de apertura", RIGHT_X, CY, RIGHT_W)
    costo_cw = [RIGHT_W*0.28] + [RIGHT_W*0.72/(n_year_cols+1)]*n_year_cols + [RIGHT_W*0.72/(n_year_cols+1)]
    diff_co = RIGHT_W - sum(costo_cw)
    costo_cw[-1] += diff_co
    costo_tbl = make_tbl(costo_data, costo_cw, totals=True)
    _, costo_h = costo_tbl.wrapOn(c, RIGHT_W, 999)
    costo_tbl.drawOn(c, RIGHT_X, next_r - costo_h)

    cy2 = min(next_l - cant_h, next_r - costo_h) - 0.34*cm

    # ── Tabla marca (ancho completo, columnas dinámicas según espacio) ────
    # En vez de repartir la lista en 2 mitades fijas y recortar cada una por
    # separado (lo que dejaba un salto invisible en el medio del ranking si
    # no cabían todas las filas), se calcula primero cuántas marcas caben en
    # total, y si no caben todas se muestran las top N + una fila "Otras
    # marcas" agregada — así el ranking mostrado siempre es continuo.
    if marca_rows:
        nxt_m = sec_title("Cantidad de OTs por marca", M, cy2, CW)
        avail_h = nxt_m - FOOTER_Y

        MARCA_FONT = 5.7
        MARCA_PAD  = 1.4
        hdr_ma = ["Marca", "N°", "%"]

        def make_marca_tbl(data, cw):
            style = [
                ("BACKGROUND",   (0, 0), (-1, 0), AZUL_MED),
                ("TEXTCOLOR",    (0, 0), (-1, 0), BLANCO),
                ("FONTNAME",     (0, 0), (-1, 0), "Helvetica-Bold"),
                ("FONTSIZE",     (0, 0), (-1, 0), MARCA_FONT),
                ("FONTNAME",     (0, 1), (-1, -1), "Helvetica"),
                ("FONTSIZE",     (0, 1), (-1, -1), MARCA_FONT),
                ("ROWBACKGROUNDS", (0, 1), (-1, -1), [BLANCO, GRIS_ALT]),
                ("GRID",         (0, 0), (-1, -1), 0.25, GRIS_LIN),
                ("LINEBELOW",    (0, 0), (-1, 0), 0.7, AZUL_CLR),
                ("TOPPADDING",   (0, 0), (-1, -1), MARCA_PAD),
                ("BOTTOMPADDING", (0, 0), (-1, -1), MARCA_PAD),
                ("LEFTPADDING",  (0, 0), (-1, -1), 3.5),
                ("RIGHTPADDING", (0, 0), (-1, -1), 3),
                ("ALIGN",        (0, 0), (0, -1), "LEFT"),
                ("ALIGN",        (1, 0), (-1, -1), "CENTER"),
                ("VALIGN",       (0, 0), (-1, -1), "MIDDLE"),
            ]
            t = Table([[str(cell) for cell in row] for row in data], colWidths=cw)
            t.setStyle(TableStyle(style))
            return t

        # Altura real de encabezado y de una fila de datos (misma para
        # cualquier cantidad de columnas, porque todas parten del mismo Y)
        _probe_cw = [CW*0.6, CW*0.2, CW*0.2]
        _h_hdr  = make_marca_tbl([hdr_ma], _probe_cw).wrapOn(c, CW, 999)[1]
        _h_2row = make_marca_tbl([hdr_ma, marca_rows[0]], _probe_cw).wrapOn(c, CW, 999)[1]
        _row_h  = max(_h_2row - _h_hdr, 1)
        _cap_col = max(int((avail_h - _h_hdr) // _row_h), 0)

        # Menor número de columnas (hasta 3) que alcance a mostrar todas las marcas
        NCOLS, cap_total = 3, _cap_col * 3
        for _nc in (1, 2, 3):
            _cap = _cap_col * _nc
            if _cap >= n_marcas:
                NCOLS, cap_total = _nc, _cap
                break

        if cap_total >= n_marcas:
            filas_mostrar = marca_rows
        else:
            _n_top = max(cap_total - 1, 0)
            _top_rows = marca_rows[:_n_top]
            _otras = marca_rows[_n_top:]
            if _otras:
                _n_otras = sum(r[1] for r in _otras)
                _pct_otras = f"{_n_otras/total*100:.1f}%"
                filas_mostrar = _top_rows + [[f"Otras marcas ({len(_otras)})", _n_otras, _pct_otras]]
            else:
                filas_mostrar = _top_rows

        if filas_mostrar:
            _por_col = (len(filas_mostrar) + NCOLS - 1) // NCOLS
            _cols_rows = [filas_mostrar[i:i + _por_col] for i in range(0, len(filas_mostrar), _por_col)]

            MA_GAP = 0.28*cm
            _n_cols_real = len(_cols_rows)
            MA_W = (CW - MA_GAP * (_n_cols_real - 1)) / _n_cols_real
            ma_cw = [MA_W*0.60, MA_W*0.20, MA_W*0.20]

            for _i, _col_rows in enumerate(_cols_rows):
                _cx = M + _i*(MA_W + MA_GAP)
                _t_ma = make_marca_tbl([hdr_ma] + _col_rows, ma_cw)
                _, _h_ma = _t_ma.wrapOn(c, MA_W, 999)
                _t_ma.drawOn(c, _cx, nxt_m - _h_ma)

    # ── Paginación de la hoja nueva (Sucursal x Categoría x Año) ──
    # Se calcula ANTES del pie de la página 1 para poder mostrar el total
    # de páginas correcto ("Pág. 1/N") sin adivinar.
    if cat_rows_all:
        _h_hdr_cat  = make_cat_tbl([_cat_hdr_row]).wrapOn(c, CW, 999)[1]
        _h_2row_cat = make_cat_tbl([_cat_hdr_row, cat_rows_all[0]]).wrapOn(c, CW, 999)[1]
        _row_h_cat  = max(_h_2row_cat - _h_hdr_cat, 1)
        _cont_avail_h = (hdr_y - 0.3*cm) - FOOTER_Y - 10  # 10pt para el titulo de seccion
        _rows_per_cont_page = max(int((_cont_avail_h - _h_hdr_cat) // _row_h_cat), 1)
        _cat_paginas = [cat_rows_all[i:i + _rows_per_cont_page]
                        for i in range(0, len(cat_rows_all), _rows_per_cont_page)]
    else:
        _cat_paginas = []
    TOTAL_PAGINAS = 1 + len(_cat_paginas)

    # ── PIE (pagina 1) ────────────────────────────────────────────
    fy = M + 0.08*cm
    c.saveState()
    c.setStrokeColor(AZUL_CLR)
    c.setLineWidth(0.35)
    c.line(M, fy + 0.38*cm, W - M, fy + 0.38*cm)
    c.setFont("Helvetica", 5.3)
    c.setFillColor(colors.HexColor("#94a3b8"))
    c.drawCentredString(W/2, fy + 0.08*cm,
        "Curifor S.A  ·  Sistema de Seguimiento OTs  ·  Documento Confidencial — Solo para uso interno")
    c.drawRightString(W - M, fy + 0.08*cm, f"Pág. 1/{TOTAL_PAGINAS}  ·  {fecha_act}")
    c.restoreState()

    c.showPage()

    # ── Página(s) nueva(s): OTs por Sucursal, Categoría y Año ─────
    for _pi, _rows_pg in enumerate(_cat_paginas, start=1):
        c.saveState()
        c.setFillColor(AZUL)
        c.roundRect(M, hdr_y, W - 2*M, HDR_H, 5, fill=1, stroke=0)
        c.restoreState()

        if logo_img_buf:
            try:
                logo_img_buf.seek(0)
                pw, ph = 2.85*cm, HDR_H - 6
                px0 = M + 5
                py0 = hdr_y + 3
                c.saveState()
                c.setFillColor(BLANCO)
                c.roundRect(px0, py0, pw, ph, 4, fill=1, stroke=0)
                c.restoreState()
                c.drawImage(ImageReader(logo_img_buf), px0 + 4, py0 + 3,
                            width=pw - 8, height=ph - 6,
                            preserveAspectRatio=True, mask="auto")
            except Exception:
                pass

        tx2 = M + 3.2*cm
        c.saveState()
        c.setFont("Helvetica-Bold", 10.5)
        c.setFillColor(BLANCO)
        c.drawString(tx2, hdr_y + HDR_H*0.58,
                      "OTs POR SUCURSAL, CATEGORÍA Y AÑO DE APERTURA — CURIFOR S.A")
        c.setFont("Helvetica", 6.3)
        c.setFillColor(colors.HexColor("#b0c8e8"))
        c.drawString(tx2, hdr_y + HDR_H*0.22, f"Generado: {fecha_act}   |   Datos al: {fecha_act}")
        c.setFont("Helvetica", 5.7)
        c.setFillColor(colors.HexColor("#8ab0d4"))
        c.drawRightString(W - M - 5, hdr_y + HDR_H*0.44, filtros_desc[:120])
        c.restoreState()

        _cy_cat = hdr_y - 0.3*cm
        _titulo_cat = "OTs por sucursal, categoría y año de apertura"
        if len(_cat_paginas) > 1:
            _titulo_cat += f" (continúa {_pi}/{len(_cat_paginas)})"
        _nxt_cat = sec_title(_titulo_cat, M, _cy_cat, CW)
        _tbl_cat = make_cat_tbl([_cat_hdr_row] + _rows_pg)
        _, _h_cat = _tbl_cat.wrapOn(c, CW, 999)
        _tbl_cat.drawOn(c, M, _nxt_cat - _h_cat)

        fy2 = M + 0.08*cm
        c.saveState()
        c.setStrokeColor(AZUL_CLR)
        c.setLineWidth(0.35)
        c.line(M, fy2 + 0.38*cm, W - M, fy2 + 0.38*cm)
        c.setFont("Helvetica", 5.3)
        c.setFillColor(colors.HexColor("#94a3b8"))
        c.drawCentredString(W/2, fy2 + 0.08*cm,
            "Curifor S.A  ·  Sistema de Seguimiento OTs  ·  Documento Confidencial — Solo para uso interno")
        c.drawRightString(W - M, fy2 + 0.08*cm, f"Pág. {_pi + 1}/{TOTAL_PAGINAS}  ·  {fecha_act}")
        c.restoreState()

        c.showPage()

    c.save()
    buf.seek(0)
    return buf.getvalue()


# ============================================================
#   INFORME POR ÁREA — mapeo CATEGORIA -> Área (fijado por Cristóbal 13/07/2026)
# ============================================================
AREA_ORDEN = ["Servicio Técnico", "Garantía", "Interno", "DyP", "Sin categoría"]
_AREA_CATS = {
    "Servicio Técnico": {"CG", "CLIENTE", "COBRAR", "ERROR", "FACTURADA", "REALIZAR", "REVISAR", "ST"},
    "Garantía":          {"GARANTIA", "SUBIR"},
    "Interno":           {"INTERNO", "CI"},
    "DyP":               {"DYP"},
}


def _area_de_categoria(cat):
    """Traduce el codigo real de CATEGORIA (CG, CI, DYP, etc.) al Area de negocio.
    Cualquier codigo no mapeado explicitamente (incluye vacio, ST ya esta en
    Servicio Tecnico, SUBIR ya esta en Garantia) cae en 'Sin categoria'."""
    s = str(cat or "").strip().upper()
    for area, cats in _AREA_CATS.items():
        if s in cats:
            return area
    return "Sin categoría"


def generar_pdf_por_area(df_inf, filtros_desc, fecha_act, logo_b64_str):
    """PDF — 2 páginas A4 horizontal, mismo estilo visual del resto de informes.
    Agrupa las OTs por ÁREA (Servicio Técnico, Garantía, Interno, DyP, Sin
    categoría) según la columna CATEGORIA — ver _AREA_CATS arriba.

    Página 1: OTs por Área x Rango de días de apertura, Costo del Vale de
    Consumo por Área x Rango (repuestos_actual, campo costo_total — ya viene
    multiplicado por cantidad), y OTs por Área x Sucursal.

    Página 2: análisis mensual por Área — de las OTs que HOY superan 90 días
    de apertura (snapshot al momento de generar el informe, no historico),
    agrupadas por el mes de FECHA OT (mes de apertura), con su cantidad y el
    costo del Vale de Consumo de ESAS OT (no de todas las del mes/area — a
    pedido de Cristobal, responde "cuanto llevamos atrasado y cuanto cuesta").
    """
    (A4, landscape, colors, cm, mm, rl_canvas, Table, TableStyle,
     plt, mpatches, ImageReader) = _importar_pdf_libs()

    AZUL     = colors.HexColor("#0c243d")
    AZUL_MED = colors.HexColor("#1a3a5c")
    AZUL_CLR = colors.HexColor("#4a7ab5")
    GRIS_ALT = colors.HexColor("#f1f5f9")
    GRIS_LIN = colors.HexColor("#cbd5e1")
    BLANCO   = colors.white

    PAGE_W, PAGE_H = landscape(A4)
    M = 1.1 * cm

    buf = io.BytesIO()
    c = rl_canvas.Canvas(buf, pagesize=landscape(A4))
    c.setTitle("Informe OTs por Área — Curifor S.A")

    total = len(df_inf)
    if total == 0:
        c.showPage(); c.save(); buf.seek(0); return buf.getvalue()

    def _safe_costo_vale(reps):
        if not isinstance(reps, list):
            try:
                import json as _json
                reps = _json.loads(reps) if isinstance(reps, str) else []
            except Exception:
                reps = []
        return sum(
            float(str(r.get("costo_total", 0) or 0).replace(",", "."))
            for r in (reps or [])
        )

    def _fmt_M(v):
        if v == 0: return "—"
        if abs(v) >= 1_000_000_000: return f"${v/1_000_000_000:.2f}B"
        if abs(v) >= 1_000_000:     return f"${v/1_000_000:.1f}M"
        if abs(v) >= 1_000:         return f"${v/1_000:.0f}K"
        return f"${v:.0f}"

    def _mes_de_fecha(v):
        s = str(v or "").strip()
        if not s:
            return "Sin fecha"
        d = pd.to_datetime(s, dayfirst=True, errors="coerce")
        if pd.isna(d):
            return "Sin fecha"
        return d.strftime("%Y-%m")

    _df = df_inf.copy()
    _df["_AREA"] = _df["CATEGORIA"].apply(_area_de_categoria) if "CATEGORIA" in _df.columns else "Sin categoría"
    _df["_COSTO_VALE"] = (_df["repuestos_actual"].apply(_safe_costo_vale)
                           if "repuestos_actual" in _df.columns else 0.0)
    _df["_DIAS"] = pd.to_numeric(_df["DIAS APERTURA"], errors="coerce").fillna(0) if "DIAS APERTURA" in _df.columns else 0

    areas_presentes = [a for a in AREA_ORDEN if (_df["_AREA"] == a).any()]
    if not areas_presentes:
        areas_presentes = AREA_ORDEN

    rangos_orden = ["0-30", "31-60", "61-90", "91 o más"]

    # ── Pivot 1: cantidad de OTs por Área x Rango ──────────────
    pivot_cant = _df.groupby(["_AREA", "RANGO"]).size().unstack(fill_value=0)
    for r in rangos_orden:
        if r not in pivot_cant.columns:
            pivot_cant[r] = 0
    pivot_cant = pivot_cant.reindex(areas_presentes).fillna(0)[rangos_orden]
    pivot_cant["Total"] = pivot_cant.sum(axis=1)

    cant_data = [["Área"] + rangos_orden + ["Total"]]
    for area in areas_presentes:
        row = pivot_cant.loc[area]
        cant_data.append([area] + [int(row[r]) for r in rangos_orden] + [int(row["Total"])])
    cant_data.append(["TOTAL"] + [int(pivot_cant[r].sum()) for r in rangos_orden] + [int(pivot_cant["Total"].sum())])

    # ── Pivot 2: costo Vale de Consumo por Área x Rango ────────
    pivot_costo = _df.groupby(["_AREA", "RANGO"])["_COSTO_VALE"].sum().unstack(fill_value=0)
    for r in rangos_orden:
        if r not in pivot_costo.columns:
            pivot_costo[r] = 0
    pivot_costo = pivot_costo.reindex(areas_presentes).fillna(0)[rangos_orden]
    pivot_costo["Total"] = pivot_costo.sum(axis=1)

    costo_data = [["Área"] + rangos_orden + ["Total"]]
    for area in areas_presentes:
        row = pivot_costo.loc[area]
        costo_data.append([area] + [_fmt_M(row[r]) for r in rangos_orden] + [_fmt_M(row["Total"])])
    costo_data.append(["TOTAL"] + [_fmt_M(pivot_costo[r].sum()) for r in rangos_orden] + [_fmt_M(pivot_costo["Total"].sum())])

    vale_total = float(pivot_costo["Total"].sum())

    # ── Pivot 3: OTs por Área x Sucursal (ancho completo) ──────
    pivot_suc = _df.groupby(["SUCURSAL", "_AREA"]).size().unstack(fill_value=0) if "SUCURSAL" in _df.columns else pd.DataFrame()
    for a in areas_presentes:
        if a not in pivot_suc.columns:
            pivot_suc[a] = 0
    if not pivot_suc.empty:
        pivot_suc = pivot_suc[areas_presentes]
        pivot_suc["Total"] = pivot_suc.sum(axis=1)
        pivot_suc = pivot_suc.sort_values("Total", ascending=False)

    suc_data = [["Sucursal"] + areas_presentes + ["Total"]]
    for suc, row in pivot_suc.iterrows():
        suc_data.append([str(suc)[:20]] + [int(row[a]) for a in areas_presentes] + [int(row["Total"])])
    if not pivot_suc.empty:
        suc_data.append(["TOTAL"] + [int(pivot_suc[a].sum()) for a in areas_presentes] + [int(pivot_suc["Total"].sum())])
    else:
        suc_data.append(["TOTAL"] + [0 for _ in areas_presentes] + [0])

    # ── OTs con más de 90 días de apertura (KPI + página 2) ────
    df90 = _df[_df["_DIAS"] > 90].copy()
    n_90 = len(df90)
    if not df90.empty:
        df90["_MES"] = df90["FECHA OT"].apply(_mes_de_fecha) if "FECHA OT" in df90.columns else "Sin fecha"

    # ── Logo, helpers de dibujo (mismo estilo que los otros informes) ──
    logo_img_buf = None
    if logo_b64_str:
        try:
            logo_img_buf = io.BytesIO(base64.b64decode(logo_b64_str))
        except Exception:
            pass

    def sec_title(txt, x, y, w, clr=AZUL_MED):
        c.saveState()
        c.setFillColor(clr)
        c.setFont("Helvetica-Bold", 6.2)
        c.drawString(x, y, txt.upper())
        c.setStrokeColor(clr)
        c.setLineWidth(0.65)
        c.line(x, y - 2, x + w, y - 2)
        c.restoreState()
        return y - 9

    def make_tbl(data, cw, hdr_clr=AZUL_MED, totals=False, font=6.2):
        style = [
            ("BACKGROUND",   (0, 0), (-1, 0), hdr_clr),
            ("TEXTCOLOR",    (0, 0), (-1, 0), BLANCO),
            ("FONTNAME",     (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE",     (0, 0), (-1, 0), font),
            ("FONTNAME",     (0, 1), (-1, -1), "Helvetica"),
            ("FONTSIZE",     (0, 1), (-1, -1), font),
            ("ROWBACKGROUNDS", (0, 1), (-1, -2 if totals else -1), [BLANCO, GRIS_ALT]),
            ("GRID",         (0, 0), (-1, -1), 0.25, GRIS_LIN),
            ("LINEBELOW",    (0, 0), (-1, 0), 0.7, AZUL_CLR),
            ("TOPPADDING",   (0, 0), (-1, -1), 2.2),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 2.2),
            ("LEFTPADDING",  (0, 0), (-1, -1), 3.5),
            ("RIGHTPADDING", (0, 0), (-1, -1), 3),
            ("ALIGN",        (0, 0), (0, -1), "LEFT"),
            ("ALIGN",        (1, 0), (-1, -1), "CENTER"),
            ("VALIGN",       (0, 0), (-1, -1), "MIDDLE"),
        ]
        if totals:
            style += [
                ("BACKGROUND", (0, -1), (-1, -1), AZUL),
                ("TEXTCOLOR",  (0, -1), (-1, -1), BLANCO),
                ("FONTNAME",   (0, -1), (-1, -1), "Helvetica-Bold"),
                ("LINEABOVE",  (0, -1), (-1, -1), 0.7, AZUL_CLR),
            ]
        t = Table([[str(cell) for cell in row] for row in data], colWidths=cw)
        t.setStyle(TableStyle(style))
        return t

    def draw_header(titulo, pag_lbl):
        HDR_H = 1.58 * cm
        hdr_y = PAGE_H - M - HDR_H
        c.saveState()
        c.setFillColor(AZUL)
        c.roundRect(M, hdr_y, PAGE_W - 2*M, HDR_H, 5, fill=1, stroke=0)
        c.restoreState()
        if logo_img_buf:
            try:
                pw, ph = 2.85*cm, HDR_H - 6
                px0 = M + 5
                py0 = hdr_y + 3
                c.saveState()
                c.setFillColor(BLANCO)
                c.roundRect(px0, py0, pw, ph, 4, fill=1, stroke=0)
                c.restoreState()
                logo_img_buf.seek(0)
                c.drawImage(ImageReader(logo_img_buf), px0 + 4, py0 + 3,
                            width=pw - 8, height=ph - 6,
                            preserveAspectRatio=True, mask="auto")
            except Exception:
                pass
        tx = M + 3.2*cm
        c.saveState()
        c.setFont("Helvetica-Bold", 10.5)
        c.setFillColor(BLANCO)
        c.drawString(tx, hdr_y + HDR_H*0.58, titulo)
        c.setFont("Helvetica", 6.3)
        c.setFillColor(colors.HexColor("#b0c8e8"))
        c.drawString(tx, hdr_y + HDR_H*0.22, f"Generado: {fecha_act}   |   Datos al: {fecha_act}")
        c.setFont("Helvetica", 5.7)
        c.setFillColor(colors.HexColor("#8ab0d4"))
        c.drawRightString(PAGE_W - M - 5, hdr_y + HDR_H*0.44, filtros_desc[:120])
        c.restoreState()
        return hdr_y

    def draw_footer(pag_lbl):
        fy = M + 0.08*cm
        c.saveState()
        c.setStrokeColor(AZUL_CLR)
        c.setLineWidth(0.35)
        c.line(M, fy + 0.38*cm, PAGE_W - M, fy + 0.38*cm)
        c.setFont("Helvetica", 5.3)
        c.setFillColor(colors.HexColor("#94a3b8"))
        c.drawCentredString(PAGE_W/2, fy + 0.08*cm,
            "Curifor S.A  ·  Sistema de Seguimiento OTs  ·  Documento Confidencial — Solo para uso interno")
        c.drawRightString(PAGE_W - M, fy + 0.08*cm, f"{pag_lbl}  ·  {fecha_act}")
        c.restoreState()

    CW = PAGE_W - 2*M
    FOOTER_Y = M + 0.52*cm

    # ═══════════════ PÁGINA 1 — Rango, Costo y Sucursal por Área ═══════════════
    hdr_y = draw_header("OTs PENDIENTES POR ÁREA — CURIFOR S.A", "Pág. 1/2")

    KPI_Y = hdr_y - 0.25*cm
    KPI_H = 1.42*cm
    KPI_GAP = 0.16*cm
    KW = (CW - 3*KPI_GAP) / 4
    kpi_defs = [
        (f"{total:,}",              "TOTAL OTs PENDIENTES",      AZUL_CLR, "#eef4fc"),
        (_fmt_M(vale_total),        "COSTO TOTAL VALE CONSUMO",  AZUL_MED, "#e8f0fa"),
        (f"{n_90:,}",               "OTs CON MÁS DE 90 DÍAS",    AZUL,     "#dde9f7"),
        (f"{len(areas_presentes)}", "ÁREAS CON DATOS",           AZUL_CLR, "#eef4fc"),
    ]
    for i, (val, lbl, top, bg) in enumerate(kpi_defs):
        kx = M + i*(KW + KPI_GAP)
        ky = KPI_Y - KPI_H
        c.saveState()
        c.setFillColor(colors.HexColor(bg))
        c.setStrokeColor(GRIS_LIN)
        c.setLineWidth(0.4)
        c.roundRect(kx, ky, KW, KPI_H, 4, fill=1, stroke=1)
        c.setFillColor(top)
        c.rect(kx, ky + KPI_H - 3, KW, 3, fill=1, stroke=0)
        c.setFont("Helvetica-Bold", 15.5)
        c.setFillColor(top)
        c.drawCentredString(kx + KW/2, ky + KPI_H*0.46, val)
        c.setFont("Helvetica-Bold", 4.8)
        c.setFillColor(colors.HexColor("#64748b"))
        c.drawCentredString(kx + KW/2, ky + KPI_H*0.16, lbl)
        c.restoreState()

    CY = KPI_Y - KPI_H - 0.28*cm
    LEFT_W  = CW * 0.485
    RIGHT_W = CW * 0.485
    RIGHT_X = M + LEFT_W + CW*0.03
    n_r = len(rangos_orden)

    next_l = sec_title("OTs por Área y Rango de días de apertura", M, CY, LEFT_W)
    cant_cw = [LEFT_W*0.30] + [LEFT_W*0.70/(n_r+1)]*n_r + [LEFT_W*0.70/(n_r+1)]
    cant_cw[-1] += LEFT_W - sum(cant_cw)
    cant_tbl = make_tbl(cant_data, cant_cw, totals=True)
    _, cant_h = cant_tbl.wrapOn(c, LEFT_W, 999)
    cant_tbl.drawOn(c, M, next_l - cant_h)

    next_r = sec_title("Costo Vale de Consumo por Área y Rango de días", RIGHT_X, CY, RIGHT_W)
    costo_cw = [RIGHT_W*0.28] + [RIGHT_W*0.72/(n_r+1)]*n_r + [RIGHT_W*0.72/(n_r+1)]
    costo_cw[-1] += RIGHT_W - sum(costo_cw)
    costo_tbl = make_tbl(costo_data, costo_cw, totals=True)
    _, costo_h = costo_tbl.wrapOn(c, RIGHT_W, 999)
    costo_tbl.drawOn(c, RIGHT_X, next_r - costo_h)

    cy2 = min(next_l - cant_h, next_r - costo_h) - 0.34*cm

    nxt_s = sec_title("OTs por Área y Sucursal", M, cy2, CW)
    n_a = len(areas_presentes)
    suc_cw = [CW*0.22] + [CW*0.78/(n_a+1)]*n_a + [CW*0.78/(n_a+1)]
    suc_cw[-1] += CW - sum(suc_cw)
    suc_tbl = make_tbl(suc_data, suc_cw, totals=True, font=5.9)
    _, suc_h = suc_tbl.wrapOn(c, CW, 999)
    suc_tbl.drawOn(c, M, nxt_s - suc_h)

    draw_footer("Pág. 1/2")
    c.showPage()

    # ═══════════════ PÁGINA 2 — Análisis mensual: OTs >90 días por Área ═══════════════
    hdr_y2 = draw_header("ANÁLISIS MENSUAL POR ÁREA — OTs >90 DÍAS Y COSTO VALE DE CONSUMO", "Pág. 2/2")
    CY2 = hdr_y2 - 0.35*cm

    nota_y = sec_title(
        f"OTs con más de 90 días de apertura al día de hoy, agrupadas por mes en que se abrieron "
        f"(FECHA OT) — celda: cantidad de OT (costo Vale de Consumo de esas OT)  ·  {n_90:,} OT en total",
        M, CY2, CW
    )

    if df90.empty:
        c.saveState()
        c.setFont("Helvetica", 8.5)
        c.setFillColor(colors.HexColor("#64748b"))
        c.drawString(M, nota_y - 18, "No hay OTs con más de 90 días de apertura en los datos incluidos en este informe.")
        c.restoreState()
    else:
        cant_mes  = df90.groupby(["_MES", "_AREA"]).size().unstack(fill_value=0)
        costo_mes = df90.groupby(["_MES", "_AREA"])["_COSTO_VALE"].sum().unstack(fill_value=0)
        for a in areas_presentes:
            if a not in cant_mes.columns:  cant_mes[a] = 0
            if a not in costo_mes.columns: costo_mes[a] = 0

        meses_reales = sorted([m for m in df90["_MES"].unique() if m != "Sin fecha"], reverse=True)
        tiene_sin_fecha = (df90["_MES"] == "Sin fecha").any()

        MAX_MESES = 18
        meses_mostrar = meses_reales[:MAX_MESES]
        meses_agrupar = meses_reales[MAX_MESES:]

        areas_tot_n = {a: 0 for a in areas_presentes}
        areas_tot_c = {a: 0.0 for a in areas_presentes}

        def _fila_mes(mes_label, meses_incluidos):
            fila = [mes_label]
            tot_n, tot_c = 0, 0.0
            for a in areas_presentes:
                n = int(sum(cant_mes.loc[m, a] for m in meses_incluidos if m in cant_mes.index))
                v = float(sum(costo_mes.loc[m, a] for m in meses_incluidos if m in costo_mes.index))
                areas_tot_n[a] += n
                areas_tot_c[a] += v
                tot_n += n; tot_c += v
                fila.append(f"{n} ({_fmt_M(v)})" if n else "—")
            fila.append(f"{tot_n} ({_fmt_M(tot_c)})" if tot_n else "—")
            return fila

        mes_data = [["Mes de apertura"] + areas_presentes + ["Total"]]
        for m in meses_mostrar:
            mes_data.append(_fila_mes(m, [m]))
        if meses_agrupar:
            mes_data.append(_fila_mes(f"Anteriores ({len(meses_agrupar)} meses)", meses_agrupar))
        if tiene_sin_fecha:
            mes_data.append(_fila_mes("Sin fecha", ["Sin fecha"]))

        gran_tot_n = sum(areas_tot_n.values())
        gran_tot_c = sum(areas_tot_c.values())
        fila_total = ["TOTAL"]
        for a in areas_presentes:
            fila_total.append(f"{areas_tot_n[a]} ({_fmt_M(areas_tot_c[a])})" if areas_tot_n[a] else "—")
        fila_total.append(f"{gran_tot_n} ({_fmt_M(gran_tot_c)})" if gran_tot_n else "—")
        mes_data.append(fila_total)

        n_a2 = len(areas_presentes)
        mes_cw = [CW*0.16] + [CW*0.84/(n_a2+1)]*n_a2 + [CW*0.84/(n_a2+1)]
        mes_cw[-1] += CW - sum(mes_cw)
        mes_tbl = make_tbl(mes_data, mes_cw, totals=True, font=6.0)
        _, mes_h = mes_tbl.wrapOn(c, CW, 999)
        # Si no caben todas las filas en la pagina (muchos meses distintos), se
        # recorta a lo que quepa entre el titulo y el pie — no debería pasar con
        # el tope de 18 meses + Anteriores + Sin fecha + Total, pero por seguridad.
        max_h_disp = nota_y - 18 - FOOTER_Y
        if mes_h > max_h_disp and len(mes_data) > 3:
            mes_tbl = make_tbl([mes_data[0]] + mes_data[1:2] + [mes_data[-1]], mes_cw, totals=True, font=6.0)
            _, mes_h = mes_tbl.wrapOn(c, CW, 999)
        mes_tbl.drawOn(c, M, nota_y - 18 - mes_h)

    draw_footer("Pág. 2/2")
    c.showPage()
    c.save()
    buf.seek(0)
    return buf.getvalue()


# ============================================================
#   CARGA DE DATOS
# ============================================================
# TTL de 5 minutos (antes 45 s) — 06/08/2026.
# datos_dashboard.json pesa 5,76 MB y esta funcion corre en CADA rerun de CADA
# usuario. Con TTL de 45 s, la cache (compartida por todas las sesiones) vencia
# ~80 veces por hora y el usuario que la disparaba se quedaba esperando la
# descarga completa mientras el resto quedaba bloqueado detras — esa congelacion
# periodica es lo que se reportaba como "la App se pega" y, cuando se alargaba,
# tumbaba el websocket y devolvia a la gente a la pantalla de inicio con la
# sesion vacia.
# NO retrasa ver las ediciones hechas desde la App: al guardar en Detalle y
# Edicion se llama a cargar_datos.clear() (linea ~13757), que invalida esta
# cache para TODAS las sesiones al instante. El TTL solo gobierna cuanto tarda
# en notarse una consolidacion hecha por fuera (el BAT, 1 vez al dia), y el
# boton "🔄 Actualizar datos" sigue forzando la relectura cuando haga falta.
@st.cache_data(ttl=300)
def cargar_datos():
    try:
        # Lectura vía Git Data API (blob del commit HEAD) — nunca pasa por el CDN público
        # de raw.githubusercontent.com, que puede servir una copia cacheada varios minutos
        # después de un commit nuevo y hacer que ediciones recién guardadas "desaparezcan"
        # al recargar. Funciona también con datos_dashboard.json > 1MB.
        data = _leer_json_github_blob(GITHUB_ARCHIVO)
        if not data:
            # Fallback: raw.githubusercontent.com con cache-busting, por si la API falla/rate-limit
            r = _get_doc(URL_DATOS, params={"_": datetime.now().timestamp()}, timeout=15, verify=False)
            r.raise_for_status()
            data = r.json()
        df = pd.DataFrame(data["ots"])
        return df, data.get("fecha_actualizacion", "Sin fecha")
    except Exception as e:
        return pd.DataFrame(), f"Error al cargar: {e}"


@st.cache_data(ttl=45)
def cargar_comentarios():
    try:
        r = _get_doc(URL_COMENTARIOS, params={"_": datetime.now().timestamp()}, timeout=15, verify=False)
        if r.status_code == 404:
            return pd.DataFrame(columns=["folio_ot", "autor", "fecha", "comentario", "mencionado"])
        r.raise_for_status()
        data = r.json()
        df = pd.DataFrame(data.get("comentarios", []))
        if df.empty:
            return pd.DataFrame(columns=["folio_ot", "autor", "fecha", "comentario", "mencionado"])
        if "mencionado" not in df.columns:
            df["mencionado"] = ""
        return df
    except Exception:
        return pd.DataFrame(columns=["folio_ot", "autor", "fecha", "comentario", "mencionado"])


@st.cache_data(ttl=60)
def _cargar_agenda_hoy():
    """Carga agenda_hoy.json desde GitHub (Python, sin fetch de browser)."""
    try:
        url = (f"https://raw.githubusercontent.com/"
               f"{GITHUB_USUARIO}/{GITHUB_REPO}/main/agenda_hoy.json")
        r = _get_doc(url, params={"_": datetime.now().timestamp()}, timeout=10, verify=False)
        if r.status_code == 200:
            return r.json()
    except Exception:
        pass
    return {}


@st.cache_data(ttl=300)
def _cargar_stock_repuestos():
    """
    Carga stock_repuestos.json (catálogo completo de Stock Repestos Costo.xlsx,
    ~33.000 productos, subido por consolidar_OTs.py — PASO 10). Es un archivo
    grande (~10 MB), así que se lee vía blob de Git (misma vía que
    datos_dashboard.json) para evitar el límite de 1MB / caché del CDN de
    raw.githubusercontent.com. Devuelve (lista_productos, fecha_actualizacion).
    08/07/2026.
    """
    try:
        data = _leer_json_github_blob("stock_repuestos.json")
        if data:
            return data.get("productos", []), data.get("fecha_actualizacion", "")
    except Exception:
        pass
    return [], ""


@st.cache_data(ttl=60)
def _cargar_ctrl_taller():
    """
    Carga control_taller.json desde GitHub API (Python, sin fetch de browser).

    ⚠️ INCIDENTE 05/08/2026 — LEER ANTES DE TOCAR ESTA FUNCIÓN.
    La Contents API devuelve `content: ""` (vacío) para archivos de MÁS DE 1 MB;
    no da error, simplemente no manda el contenido. La versión anterior hacía
    `data = ... if raw else {}` — o sea, apenas control_taller.json superó el
    megabyte (llegó a 1.058.920 bytes, apenas 10 KB por encima del límite), el
    Planificador empezó a cargar con `ctrlData = {}`. A partir de ahí, la primera
    persona que guardaba algo subía un archivo que contenía SOLO su sucursal:
    se perdieron las 11 sucursales, los 46 técnicos configurados y 800+ órdenes,
    y volvía a pasar con cada guardado.

    Ahora: si la Contents API viene vacía, se lee el blob por Git Data API
    (que NO tiene límite de tamaño), igual que ya hace _leer_json_github_blob()
    para datos_dashboard.json. El `sha` que devuelve la Contents API para un
    archivo es exactamente el sha del blob de git, así que el que se obtiene
    del árbol sirve igual para el PUT posterior.

    NUNCA devolver {} en silencio cuando el archivo sí existe: quien llama no
    puede distinguir "no hay datos" de "no los pude leer", y esa ambigüedad es
    justamente lo que borró los datos.
    """
    _base = f"https://api.github.com/repos/{GITHUB_USUARIO}/{GITHUB_REPO}"

    # 1) Vía normal (Contents API) — sirve mientras el archivo pese menos de 1 MB
    try:
        r = _get_doc(f"{_base}/contents/control_taller.json",
                     headers=_github_headers(), timeout=10, verify=False)
        if r.status_code == 200:
            j   = r.json()
            sha = j.get("sha", "")
            raw = (j.get("content") or "").replace("\n", "")
            if raw:
                return json.loads(base64.b64decode(raw).decode("utf-8")), sha
    except Exception:
        pass

    # 2) Respaldo por Git Data API (sin límite de 1 MB, y sin pasar por el CDN).
    #    Con Supabase NO se usa: si el documento no está en la base, buscarlo en
    #    GitHub traería una copia vieja y el tablero mostraría órdenes que ya no
    #    son las vigentes. Mejor devolver vacío y que se note.
    if _datos.disponible():
        return {}, ""

    try:
        h = _github_headers()
        r = requests.get(f"{_base}/git/ref/heads/main", headers=h, timeout=10, verify=False)
        r.raise_for_status()
        commit_sha = r.json()["object"]["sha"]
        r = requests.get(f"{_base}/git/trees/{commit_sha}", headers=h,
                         params={"recursive": "1"}, timeout=15, verify=False)
        r.raise_for_status()
        blob_sha = next((it.get("sha") for it in r.json().get("tree", [])
                         if it.get("path") == "control_taller.json"), None)
        if blob_sha:
            r = requests.get(f"{_base}/git/blobs/{blob_sha}", headers=h, timeout=30, verify=False)
            r.raise_for_status()
            raw = (r.json().get("content") or "").replace("\n", "")
            if raw:
                return json.loads(base64.b64decode(raw).decode("utf-8")), blob_sha
    except Exception:
        pass

    return {}, ""


@st.cache_data(ttl=60)
def _cargar_prepicking():
    """
    Carga prepicking_estados.json desde GitHub API (Python, sin fetch de
    browser) — mismo patrón que _cargar_ctrl_taller(). Guarda el estado
    Realizado/Pendiente de cada cita de Pre-picking, por sucursal + fecha +
    OC (13/07/2026, independiente de agenda_hoy.json para que no se pierda
    cada vez que se refresca la agenda). Devuelve (data, sha).
    Estructura: {"SUCURSAL": {"DD/MM/AAAA": {"oc": "realizado"|"pendiente"}}}
    """
    try:
        url = (f"https://api.github.com/repos/"
               f"{GITHUB_USUARIO}/{GITHUB_REPO}/contents/prepicking_estados.json")
        r = _get_doc(url, headers=_github_headers(), timeout=10, verify=False)
        if r.status_code == 200:
            j    = r.json()
            sha  = j.get("sha", "")
            raw  = (j.get("content") or "").replace("\n", "")
            data = json.loads(base64.b64decode(raw).decode("utf-8")) if raw else {}
            return data, sha
    except Exception:
        pass
    return {}, ""


@st.cache_data(ttl=600)
def _cargar_tempario():
    """Carga tempario.json (horas de mano de obra por marca/modelo/km, subido por
    consolidar_OTs.py). Archivo chico (~20 KB) y casi estático — TTL largo.
    Usado por el Asistente App (sección 5.3 de la Constitución). 10/07/2026."""
    try:
        url = (f"https://raw.githubusercontent.com/"
               f"{GITHUB_USUARIO}/{GITHUB_REPO}/main/tempario.json")
        r = _get_doc(url, params={"_": datetime.now().timestamp()}, timeout=15, verify=False)
        if r.status_code == 200:
            return r.json().get("modelos", [])
    except Exception:
        pass
    return []


@st.cache_data(ttl=300)
def _cargar_produccion_tecnicos():
    """
    Carga produccion_tecnicos.json (horas facturadas por tecnico/mecanico, subido
    por consolidar_OTs.py desde BDFlexline — PASO 11). Alimenta la pestaña
    "📊 Producción Técnicos" del Planificador de Taller. Vía blob de Git (misma
    que datos_dashboard.json/stock_repuestos.json) por si el archivo crece más
    de 1 MB con el historial acumulado. Devuelve (resumen, detalle_producto,
    detalle_ot, fecha_act) — resumen es mecanico+sucursal+mes (historial largo,
    tabla principal); detalle_producto es mecanico+mes+producto agregado (solo
    meses recientes, para el drill-down por técnico); detalle_ot es la misma
    ventana de meses pero SIN agregar (linea por linea: Nº OT/Producto/Precio
    Lista/Total Horas/Comi_Vta — agregado 21/07/2026 a pedido de Cristóbal).
    20/07/2026.
    """
    try:
        data = _leer_json_github_blob(GITHUB_PRODUCCION_TECNICOS)
        if data:
            return (data.get("resumen", []), data.get("detalle_producto", []),
                     data.get("detalle_ot", []), data.get("fecha_actualizacion", ""))
    except Exception:
        pass
    return [], [], [], ""


@st.cache_data(ttl=120)
def _cargar_campanas():
    """
    Carga campanas_curifor.json — "Revisión de Campañas Ford", subido
    por consolidar_OTs.py (PASO 12) desde un archivo de OTRO proyecto local
    ('*_Consolidado_Curifor_2Tandas.xlsx'). Vía blob de Git (misma vía
    autoritativa que datos_dashboard.json) para no depender del CDN cacheado.
    Devuelve (lista_de_casos, fecha_actualizacion, archivo_origen) — vacío/""
    si el archivo aún no existe (primera vez, antes de correr el consolidador
    con este paso nuevo). 28/07/2026.
    """
    try:
        data = _leer_json_github_blob(GITHUB_CAMPANAS)
        if data:
            return (data.get("campanas", []), data.get("fecha_actualizacion", ""),
                     data.get("archivo_origen", ""))
    except Exception:
        pass
    return [], "", ""


@st.cache_data(ttl=300)
def _cargar_cuenta_ficha():
    """
    Carga cuenta_ficha.json — modulo "Cuenta Ficha", subido por
    consolidar_OTs.py (PASO 13). El payload viaja comprimido (gzip+base64,
    campo "gz") porque en crudo son ~9 MB: saldos de cuenta corriente por
    cliente + su historial de OT con todos los documentos posteriores.

    Se lee por blob de Git (via autoritativa, sin el CDN cacheado) igual que
    datos_dashboard.json. Devuelve (clientes, resumen, fecha_actualizacion).
    31/07/2026.
    """
    try:
        data = _leer_json_github_blob(GITHUB_CUENTA_FICHA)
        if data:
            _gz = data.get("gz", "")
            if _gz:
                payload = json.loads(
                    gzip.decompress(base64.b64decode(_gz)).decode("utf-8")
                )
            else:                       # compatibilidad: payload sin comprimir
                payload = data
            return (payload.get("clientes", []),
                    payload.get("resumen", data.get("resumen", {})),
                    data.get("fecha_actualizacion", ""))
    except Exception:
        pass
    return [], {}, ""


@st.cache_data(ttl=180)
def _cargar_informes_gestion():
    """
    Carga informes_gestion.json — modulo "Informes de Gestion", subido por
    informes_gestion.py (Actualizar_Informes_Gestion.bat). El payload viaja
    comprimido (gzip+base64, campo "gz").

    Contiene:
      * ag    : historico del "NUEVO REPORTE DE GESTION" por sucursal
                (hojas HYU / FOR / D&P, 12 meses cada una)
      * ford  : historico del "Informe IMOP" (una hoja por mes)
      * actual: indicadores del mes en curso calculados desde la carpeta
                Alimentacion (pasos vehiculares y facturacion)
      * snapshots: una foto por cada carga, para ver el avance semanal

    Se lee por blob de Git (via autoritativa, sin el CDN cacheado). 04/08/2026.
    """
    try:
        data = _leer_json_github_blob(GITHUB_INFORMES_GESTION)
        if data:
            _gz = data.get("gz", "")
            payload = (json.loads(gzip.decompress(base64.b64decode(_gz)).decode("utf-8"))
                       if _gz else data)
            return payload
    except Exception:
        pass
    return {}


@st.cache_data(ttl=45)
def _cf_comentarios_por_ot():
    """
    Comentarios escritos desde la App (comentarios_log.json), agrupados por
    Folio OT, para mostrarlos dentro de Cuenta Ficha. Se leen EN VIVO (no van
    dentro de cuenta_ficha.json) para que un comentario recién escrito en
    "Documentos y Comentarios" aparezca aquí sin esperar la consolidación.
    Devuelve { FOLIO: [ {autor, fecha, texto}, ... ] }.  31/07/2026.
    """
    try:
        _dfc = cargar_comentarios()
        if _dfc is None or _dfc.empty or "folio_ot" not in _dfc.columns:
            return {}
        out = {}
        for _r in _dfc.itertuples(index=False):
            _f = str(getattr(_r, "folio_ot", "")).strip()
            _f = _f.lstrip("0") or _f
            if not _f:
                continue
            out.setdefault(_f, []).append({
                "autor": str(getattr(_r, "autor", "") or ""),
                "fecha": str(getattr(_r, "fecha", "") or ""),
                "texto": str(getattr(_r, "comentario", "") or ""),
            })
        return out
    except Exception:
        return {}


@st.cache_data(ttl=120)
def _cf_gestion_por_ot():
    """
    Gestión escrita desde la App sobre cada OT (Categoría, Observación OT,
    Notas y Avance/Gestión de datos_dashboard.json), para mostrarla dentro de
    Cuenta Ficha. Solo existe para las OT que hoy están pendientes/anuladas
    (son las únicas que viajan en el dashboard) — que son justamente las
    destacadas en la ficha. Devuelve { FOLIO: {campo: valor} }.  31/07/2026.
    """
    try:
        _dfd, _ = cargar_datos()
        if _dfd is None or _dfd.empty or "FOLIO OT" not in _dfd.columns:
            return {}
        _campos = [c for c in ["CATEGORIA", "OBSERVACION OT", "NOTAS",
                               "AVANCE - GESTIÓN", "ULTIMA_EDICION"]
                   if c in _dfd.columns]
        if not _campos:
            return {}
        out = {}
        for _fila in _dfd[["FOLIO OT"] + _campos].itertuples(index=False, name=None):
            _f = str(_fila[0]).strip()
            _f = _f.lstrip("0") or _f
            if not _f:
                continue
            _vals = {}
            for _i, _c in enumerate(_campos):
                _v = str(_fila[_i + 1] or "").strip()
                if _v and _v.lower() not in ("nan", "none"):
                    _vals[_c] = _v
            if any(_vals.get(_k) for _k in
                   ("CATEGORIA", "OBSERVACION OT", "NOTAS", "AVANCE - GESTIÓN")):
                out[_f] = _vals
        return out
    except Exception:
        return {}


@st.cache_data(ttl=20)
def _cargar_cf_revisados():
    """
    Marcas "Revisado" de Cuenta Ficha — COMPARTIDAS por todo el equipo
    (decision de Cristobal 31/07/2026): si alguien revisa un cliente, queda
    marcado para todos, con quien lo reviso y cuando.
    Devuelve { RUT: {"usuario", "fecha", "nota"} }.
    """
    try:
        _, datos = _leer_json_github_raw(GITHUB_CUENTA_FICHA_REV)
        if isinstance(datos, dict):
            _r = datos.get("revisados", {})
            if isinstance(_r, dict):
                return _r
    except Exception:
        pass
    return {}


def _guardar_cf_revisado(rut, usuario, marcar=True, nota=""):
    """
    Marca/desmarca un cliente como revisado. Relee el archivo fresco antes de
    escribir para no pisar lo que otro usuario haya marcado mientras tanto.
    """
    try:
        _, datos = _leer_json_github_raw(GITHUB_CUENTA_FICHA_REV)
        if not isinstance(datos, dict):
            datos = {}
        revs = datos.get("revisados", {})
        if not isinstance(revs, dict):
            revs = {}
        if marcar:
            revs[str(rut)] = {
                "usuario": usuario,
                "fecha":   ahora_chile(),
                "nota":    (nota or "").strip()[:300],
            }
        else:
            revs.pop(str(rut), None)
        datos["revisados"] = revs
        ok = _guardar_json_github_raw(
            GITHUB_CUENTA_FICHA_REV, datos,
            f"Cuenta Ficha: {'revisado' if marcar else 'sin revisar'} {rut} ({usuario})",
        )
        return ok
    except Exception:
        return False


@st.cache_data(ttl=600)
def _cargar_cotizador_gz():
    """
    Carga cotizador_data.json (bundle del módulo 'Cotizador de Mantenciones':
    índice de marcas/modelos/versiones + stock + 273 pautas + CSS/JS/HTML +
    librería XLSX + logo, todo comprimido gzip+base64 en el campo 'gz', ~1.1 MB).
    Se lee vía blob de Git (misma vía autoritativa que datos_dashboard.json) para
    no depender del CDN cacheado. El string gz se inyecta tal cual en el iframe y
    se descomprime en el navegador con DecompressionStream. Casi estático (solo
    cambia si se regenera el bundle desde el Cotizador), por eso TTL largo.
    21/07/2026. Devuelve (gz_b64, actualizado) o ("", "") si no está disponible.
    """
    try:
        data = _leer_json_github_blob("cotizador_data.json")
        if data and data.get("gz"):
            return data.get("gz", ""), data.get("actualizado", "")
    except Exception:
        pass
    return "", ""


@st.cache_data(ttl=600)
def _cargar_stock_completo_gz():
    """
    Catálogo COMPLETO de Stock de Repuestos (stock_repuestos.json, ~33.000
    filas producto+bodega, generado por consolidar_OTs.py — PASO 10) comprimido
    gzip+base64 para inyectar en el iframe del Planificador.

    Por qué existe esta función además de _cargar_stock_repuestos() (que ya
    usa el Asistente App): esa devuelve la lista completa sin comprimir —
    perfecta para procesarla en Python, pero inyectar ~10 MB de JSON plano
    dentro del HTML del Planificador sería muy pesado. Acá se recorta cada
    registro a solo los 4 campos que necesita el JS de Pre-picking para buscar
    "código relacionado" y su stock (producto/descripción/bodega/stock — se
    descartan familia/subfamilia/costo/etc.) y se comprime con gzip antes de
    convertir a base64, igual que ya se hace con cotizador_data.json — el
    navegador lo descomprime con DecompressionStream.

    Esto permite que las alternativas "también sirve" de Pre-picking busquen
    contra el catálogo REAL completo (33.000 códigos) en vez de solo los ~400
    códigos acotados que trae el bundle del Cotizador de Mantenciones —
    restaurando la misma amplitud que tenía el pipeline viejo (sesión 14/07,
    "_buscar_repuestos_compatibles"). 22/07/2026.
    Devuelve (gz_b64, fecha_actualizacion) o ("", "") si no está disponible.
    """
    try:
        data = _leer_json_github_blob("stock_repuestos.json")
        if not data:
            return "", ""
        productos = data.get("productos", [])
        if not productos:
            return "", ""
        compacto = [
            {"p": p.get("producto", ""), "d": p.get("descripcion", ""),
             "b": p.get("bodega", ""), "s": p.get("stock", 0) or 0}
            for p in productos
        ]
        raw = json.dumps(compacto, ensure_ascii=False).encode("utf-8")
        gz_b64 = base64.b64encode(gzip.compress(raw)).decode("ascii")
        return gz_b64, data.get("fecha_actualizacion", "")
    except Exception:
        return "", ""


@st.cache_data(ttl=120)
def cargar_historial_cierres():
    url = (
        f"https://raw.githubusercontent.com/"
        f"{GITHUB_USUARIO}/{GITHUB_REPO}/main/historial_cierres.json"
    )
    try:
        r = _get_doc(url, params={"_": datetime.now().timestamp()}, timeout=15, verify=False)
        if r.status_code == 404:
            return pd.DataFrame(), []
        r.raise_for_status()
        data = r.json()
        registros = data.get("registros", [])
        if not registros:
            return pd.DataFrame(), []
        resumen = []
        for reg in registros:
            resumen.append({
                "Fecha":        reg.get("fecha", ""),
                "OTs Cerradas": reg.get("total_cerradas", 0),
                "OTs Nuevas":   reg.get("total_nuevas", 0),
                "OTs Activas":  reg.get("total_activas", 0),
            })
        return pd.DataFrame(resumen), registros
    except Exception:
        return pd.DataFrame(), []


@st.cache_data(ttl=120)
def cargar_ranking_cierres():
    try:
        r = _get_doc(URL_RANKING, params={"_": datetime.now().timestamp()}, timeout=15, verify=False)
        if r.status_code == 404:
            return None
        r.raise_for_status()
        return r.json()
    except Exception:
        return None


@st.cache_data(ttl=30)
def cargar_notificaciones_cache():
    return _leer_notificaciones()


@st.cache_data(ttl=120)
def cargar_usuarios_cache():
    return _leer_usuarios()


_KPI_ICONS = {
    "rojo":    "🔴",
    "naranja": "🟠",
    "amarillo":"🟡",
    "verde":   "🟢",
    "":        "📋",
}

def kpi_card(label, valor, color=""):
    icono = _KPI_ICONS.get(color, "📋")
    st.markdown(f"""
    <div class="kpi-box {color}">
        <div class="kpi-icono">{icono}</div>
        <p class="kpi-num {color}">{valor:,}</p>
        <p class="kpi-label">{label}</p>
    </div>
    """, unsafe_allow_html=True)


# ============================================================
#   FUNCIONES GITHUB PARA DATOS PRINCIPALES
# ============================================================
def _folio_limpio_de_display(display_val):
    _emojis_conocidos = {"🔴", "🟡", "🟢", "🔵"}
    s = str(display_val).strip()
    partes = s.split(" ", 1)
    if len(partes) == 2 and partes[0] in _emojis_conocidos:
        return partes[1]
    return s


def _nombre_de_url(url_api):
    """El nombre del documento a partir de la URL de la Contents API.

    Existe porque esta función recibe la URL ya armada en vez del nombre, y con
    Supabase lo que se necesita es el nombre. Se saca de la propia URL para no
    tener que tocar los lugares que la llaman.
    """
    return url_api.split("/contents/")[-1].split("?")[0].strip("/")


def _leer_json_github(url_api):
    if _datos.disponible():
        nombre = _nombre_de_url(url_api)
        sello, datos = _datos.leer_con_sello(nombre)
        if datos is None:
            # Igual que antes: si no se pudo leer, se levanta el error en vez de
            # devolver un documento vacío que luego se guardaría encima del bueno.
            raise ValueError(f"No se pudo leer '{nombre}' desde Supabase.")
        return sello, datos

    resp = requests.get(url_api, headers=_github_headers(), timeout=15, verify=False)
    resp.raise_for_status()
    info = resp.json()
    if "sha" not in info:
        raise ValueError(f"Respuesta inesperada de GitHub: {str(info)[:120]}")
    sha = info["sha"]
    raw_content = info.get("content", "").replace("\n", "").strip()
    if raw_content:
        datos = json.loads(base64.b64decode(raw_content).decode("utf-8"))
    else:
        dl_url = info.get("download_url", "")
        if not dl_url:
            raise ValueError("Archivo demasiado grande y sin download_url.")
        r2 = requests.get(dl_url, timeout=30, verify=False)
        r2.raise_for_status()
        datos = r2.json()
    return sha, datos


def _subir_json_github(nombre_archivo, datos_dict, mensaje_commit):
    if _datos.disponible():
        if not _datos.guardar(nombre_archivo, datos_dict, mensaje_commit):
            raise RuntimeError(f"No se pudo guardar '{nombre_archivo}' en Supabase.")
        return True

    hdrs     = _github_headers()
    base_url = f"https://api.github.com/repos/{GITHUB_USUARIO}/{GITHUB_REPO}"
    contenido_b64 = base64.b64encode(
        json.dumps(datos_dict, ensure_ascii=False, indent=2).encode("utf-8")
    ).decode()
    r = requests.post(f"{base_url}/git/blobs",
                      headers=hdrs,
                      json={"content": contenido_b64, "encoding": "base64"},
                      timeout=60, verify=False)
    r.raise_for_status()
    blob_sha = r.json()["sha"]
    r = requests.get(f"{base_url}/git/ref/heads/main", headers=hdrs, timeout=15, verify=False)
    r.raise_for_status()
    commit_sha = r.json()["object"]["sha"]
    r = requests.get(f"{base_url}/git/commits/{commit_sha}", headers=hdrs, timeout=15, verify=False)
    r.raise_for_status()
    tree_sha = r.json()["tree"]["sha"]
    r = requests.post(f"{base_url}/git/trees",
                      headers=hdrs,
                      json={"base_tree": tree_sha,
                            "tree": [{"path": nombre_archivo, "mode": "100644",
                                      "type": "blob", "sha": blob_sha}]},
                      timeout=30, verify=False)
    r.raise_for_status()
    new_tree_sha = r.json()["sha"]
    r = requests.post(f"{base_url}/git/commits",
                      headers=hdrs,
                      json={"message": mensaje_commit, "tree": new_tree_sha, "parents": [commit_sha]},
                      timeout=30, verify=False)
    r.raise_for_status()
    new_commit_sha = r.json()["sha"]
    r = requests.patch(f"{base_url}/git/refs/heads/main",
                       headers=hdrs,
                       json={"sha": new_commit_sha},
                       timeout=15, verify=False)
    r.raise_for_status()
    return True


def guardar_en_github(df_editado, df_original, usuario_activo):
    cols_editables = ["CATEGORIA", "OBSERVACION OT", "NOTAS", "AVANCE - GESTIÓN"]
    cambios = {}
    for i, row_nuevo in df_editado.iterrows():
        folio = _folio_limpio_de_display(row_nuevo.get("FOLIO OT", ""))
        if i >= len(df_original):
            continue
        row_orig = df_original.iloc[i]
        if any(str(row_nuevo.get(c, "")).strip() != str(row_orig.get(c, "")).strip()
               for c in cols_editables if c in df_editado.columns):
            cambios[folio] = {c: str(row_nuevo.get(c, "")).strip()
                              for c in cols_editables if c in df_editado.columns}
    if not cambios:
        return False, "No se detectaron cambios para guardar."
    url_api = (f"https://api.github.com/repos/{GITHUB_USUARIO}/"
               f"{GITHUB_REPO}/contents/{GITHUB_ARCHIVO}")
    try:
        _, datos = _leer_json_github(url_api)
        ahora = ahora_chile()
        for ot in datos["ots"]:
            folio = str(ot.get("FOLIO OT", "")).strip()
            if folio in cambios:
                for col, val in cambios[folio].items():
                    ot[col] = val
                ot["ULTIMA_EDICION"] = f"{usuario_activo} — {ahora}"
        _subir_json_github(GITHUB_ARCHIVO, datos, f"Edición de {usuario_activo} — {ahora}")
        _registrar_audit(usuario_activo, "EDICION", f"{len(cambios)} OT(s) editadas", ", ".join(list(cambios.keys())[:5]))
        return True, f"✅ {len(cambios)} OT(s) guardadas correctamente."
    except Exception as e:
        return False, f"Error inesperado: {e}"


def guardar_colores_github(cambios_color: dict):
    url_api = (f"https://api.github.com/repos/{GITHUB_USUARIO}/"
               f"{GITHUB_REPO}/contents/{GITHUB_ARCHIVO}")
    try:
        _, datos = _leer_json_github(url_api)
        for ot in datos["ots"]:
            folio = str(ot.get("FOLIO OT", "")).strip()
            if folio in cambios_color:
                ot["_MARCA_COLOR_"] = str(cambios_color[folio])
        ahora = ahora_chile()
        _subir_json_github(GITHUB_ARCHIVO, datos, f"Marcas de color — {ahora}")
        return True, f"✅ {len(cambios_color)} marca(s) guardada(s)."
    except Exception as e:
        return False, f"Error inesperado: {e}"


def guardar_etapa_jpcb_github(cambios_etapa: dict):
    """Actualiza ETAPA_JPCB en datos_dashboard.json. cambios_etapa = {folio: etapa}"""
    url_api = (f"https://api.github.com/repos/{GITHUB_USUARIO}/"
               f"{GITHUB_REPO}/contents/{GITHUB_ARCHIVO}")
    try:
        _, datos = _leer_json_github(url_api)
        for ot in datos["ots"]:
            folio = str(ot.get("FOLIO OT", "")).strip()
            if folio in cambios_etapa:
                ot["ETAPA_JPCB"] = str(cambios_etapa[folio])
        ahora = ahora_chile()
        _subir_json_github(GITHUB_ARCHIVO, datos, f"JPCB etapas — {ahora}")
        return True, f"✅ {len(cambios_etapa)} OT(s) actualizadas en el tablero JPCB."
    except Exception as e:
        return False, f"Error inesperado: {e}"


def agregar_comentario_github(folio_ot, autor, texto, mencionados=None):
    """Agrega un comentario y crea notificaciones para cada @mencionado.
    mencionados puede ser una lista de emails o un string (retrocompatibilidad).
    """
    # Normalizar a lista
    if mencionados is None:
        mencionados = []
    elif isinstance(mencionados, str):
        mencionados = [mencionados] if mencionados.strip() else []
    # Filtrar al propio autor
    mencionados = [m.strip() for m in mencionados if m.strip().lower() != autor.strip().lower()]

    url_api = (f"https://api.github.com/repos/{GITHUB_USUARIO}/"
               f"{GITHUB_REPO}/contents/{GITHUB_COMENTARIOS}")
    try:
        sha = None
        comentarios = []
        resp = _get_doc(url_api, headers=_github_headers(), timeout=15, verify=False)
        if resp.status_code == 200:
            info = resp.json()
            sha  = info["sha"]
            datos = json.loads(base64.b64decode(info["content"]).decode("utf-8"))
            comentarios = datos.get("comentarios", [])
        elif resp.status_code != 404:
            resp.raise_for_status()

        comentarios.append({
            "folio_ot":   str(folio_ot).strip(),
            "autor":      str(autor).strip(),
            "fecha":      ahora_chile(),
            "comentario": str(texto).strip(),
            "mencionado": ", ".join(mencionados),
        })

        # La lectura de arriba ya viene de Supabase cuando está disponible; la
        # escritura tiene que ir al MISMO lado o los comentarios se guardarían
        # en un archivo que nadie vuelve a leer.
        if _datos.disponible():
            _ok_guardado = _datos.guardar(GITHUB_COMENTARIOS, {"comentarios": comentarios},
                                          f"Comentario OT {folio_ot} — {autor}")
        else:
            nuevo_b64 = base64.b64encode(
                json.dumps({"comentarios": comentarios}, ensure_ascii=False, indent=2).encode("utf-8")
            ).decode()
            payload = {"message": f"Comentario OT {folio_ot} — {autor}", "content": nuevo_b64}
            if sha:
                payload["sha"] = sha
            resp_put = requests.put(url_api, headers=_github_headers(),
                                    json=payload, timeout=30, verify=False)
            _ok_guardado = resp_put.status_code in [200, 201]

        if _ok_guardado:
            # Audit
            _registrar_audit(autor, "COMENTARIO", f"Comentario en OT {folio_ot}", folio_ot)
            # Notificación para cada mencionado
            for dest in mencionados:
                try:
                    _crear_notificacion(
                        remitente=autor,
                        destinatario=dest,
                        folio_ot=str(folio_ot).strip(),
                        extracto=texto[:200],
                    )
                    _registrar_audit(autor, "MENCION", f"Mencionó a {dest} en OT {folio_ot}", folio_ot)
                except Exception:
                    pass
            return True, "✅ Comentario guardado correctamente."
        return False, f"Error al guardar ({resp_put.status_code})"
    except Exception as e:
        return False, f"Error inesperado: {e}"


# ============================================================
#   CAMBIO DE CONTRASEÑA
# ============================================================
def cambiar_password(email, pwd_actual, pwd_nueva):
    usuarios = _leer_usuarios()
    u = _buscar_usuario(email, usuarios)
    if not u:
        return False, "Usuario no encontrado."
    if not _verificar_pwd(pwd_actual, u["password_hash"], u["salt"]):
        return False, "Contraseña actual incorrecta."
    if len(pwd_nueva) < 6:
        return False, "La nueva contraseña debe tener al menos 6 caracteres."
    h, s = _hash_pwd(pwd_nueva)
    for uu in usuarios:
        if uu.get("email", "").lower() == email.lower():
            uu["password_hash"] = h
            uu["salt"] = s
            uu["temp_pwd"] = False
            break
    _guardar_usuarios(usuarios)
    _registrar_audit(email, "CAMBIO_PWD", "Contraseña cambiada")
    return True, "✅ Contraseña actualizada correctamente."


def _generar_password_provisoria():
    """Genera una contraseña provisoria legible (8 caracteres, sin ambigüedades
    tipo 0/O o 1/l) para que el admin se la pueda dictar/escribir a la persona
    sin errores de transcripción."""
    _alfabeto = "ABCDEFGHJKMNPQRSTUVWXYZabcdefghijkmnpqrstuvwxyz23456789"
    return "".join(secrets.choice(_alfabeto) for _ in range(8))


def asignar_password_provisoria(email, admin_email):
    """Genera y guarda una contraseña provisoria para `email` (admin la
    reparte al usuario que olvidó la suya). Marca `temp_pwd: True` para que,
    en el próximo login exitoso con esa contraseña, la app OBLIGUE a crear
    una contraseña nueva y permanente antes de dejarlo entrar (ver
    check_password(), paso 4). Devuelve (ok, mensaje, password_plano) — el
    password en texto plano solo se muestra una vez, en pantalla, nunca se
    guarda en ningún archivo."""
    email = (email or "").strip().lower()
    usuarios = _leer_usuarios()
    u = _buscar_usuario(email, usuarios)
    if not u:
        return False, "Usuario no encontrado (debe haberse registrado antes al menos una vez).", ""
    pwd_temp = _generar_password_provisoria()
    h, s = _hash_pwd(pwd_temp)
    u["password_hash"] = h
    u["salt"] = s
    u["temp_pwd"] = True
    _guardar_usuarios(usuarios)
    _registrar_audit(admin_email, "RESET_PWD", f"Contraseña provisoria asignada a {email}")
    return True, f"✅ Contraseña provisoria generada para {email}.", pwd_temp


# ============================================================
#   SELECTOR DE MODO  — pantalla de bienvenida post-login
# ============================================================
if not st.session_state.get("app_mode"):

    def _hex_rgba(hexcolor, alpha):
        """'#1a3a5c' -> 'rgba(26,58,92,alpha)' (fondos translúcidos por módulo)."""
        h = hexcolor.lstrip("#")
        r, g, b = int(h[0:2], 16), int(h[2:4], 16), int(h[4:6], 16)
        return f"rgba({r},{g},{b},{alpha})"

    st.markdown("""
    <style>
      .mod-card{position:relative;border-radius:16px;overflow:hidden;
                background:var(--secondary-background-color);
                border:1px solid rgba(128,128,128,.16);
                box-shadow:0 3px 14px rgba(0,0,0,.08);
                transition:transform .16s ease, box-shadow .16s ease;
                min-height:232px;display:flex;flex-direction:column;}
      .mod-card:hover{transform:translateY(-3px);box-shadow:0 12px 30px rgba(0,0,0,.16);}
      .mod-card-top{height:5px;background:var(--accent,#1a3a5c);}
      .mod-card-body{padding:24px 22px 20px;text-align:center;flex:1;
                     display:flex;flex-direction:column;align-items:center;}
      .mod-icon{width:60px;height:60px;border-radius:16px;display:flex;
                align-items:center;justify-content:center;font-size:1.9rem;
                margin-bottom:14px;}
      .mod-card-body h3{margin:0 0 8px;font-size:1.02rem;font-weight:700;
                        color:var(--accent,#1a3a5c);letter-spacing:-.2px;}
      .mod-card-body p{margin:0;color:var(--text-color);opacity:.62;
                       font-size:.82rem;line-height:1.5;}
      .mod-card.locked{opacity:.55;}
      .mod-card.locked .mod-card-top{background:rgba(128,128,128,.4);}
      .mod-card.locked h3{color:var(--text-color);opacity:.75;}
      .mod-icon.locked-icon{background:rgba(128,128,128,.16);color:#888;}
    </style>
    """, unsafe_allow_html=True)

    st.markdown(
        f'''<div class="curifor-header">
            <div class="logo-pill"><img src="{LOGO_DATA_URI}" /></div>
            <div class="curifor-header-text">
                <h2>Bienvenido, {usuario_activo.split('@')[0]}</h2>
                <p>¿Qué módulo deseas abrir hoy?</p>
                <span class="dev-credit">Curifor S.A</span>
            </div>
        </div>''',
        unsafe_allow_html=True,
    )
    st.markdown("<div style='height:22px'></div>", unsafe_allow_html=True)

    _usuarios_bienvenida = _leer_usuarios()

    def _tarjeta_modulo(icono, color, titulo, desc, puede, key_suffix, mode_value, primary=False):
        """Tarjeta de acceso a un módulo — misma lógica de siempre (permiso +
        session_state['app_mode'] + rerun), solo se modernizó el HTML/CSS."""
        if puede:
            st.markdown(f"""
            <div class="mod-card" style="--accent:{color};">
              <div class="mod-card-top"></div>
              <div class="mod-card-body">
                <div class="mod-icon" style="background:{_hex_rgba(color, 0.14)};color:{color};">{icono}</div>
                <h3>{titulo}</h3>
                <p>{desc}</p>
              </div>
            </div>""", unsafe_allow_html=True)
            st.markdown("<div style='height:10px'></div>", unsafe_allow_html=True)
            if st.button("Entrar al módulo →", key=f"btn_modo_{key_suffix}",
                         use_container_width=True,
                         type=("primary" if primary else "secondary")):
                st.session_state["app_mode"] = mode_value
                st.rerun()
        else:
            st.markdown(f"""
            <div class="mod-card locked">
              <div class="mod-card-top"></div>
              <div class="mod-card-body">
                <div class="mod-icon locked-icon">🔒</div>
                <h3>{titulo}</h3>
                <p>Acceso restringido. Solicita autorización al administrador
                   (cjerez@curifor.com) para usar este módulo.</p>
              </div>
            </div>""", unsafe_allow_html=True)
            st.markdown("<div style='height:10px'></div>", unsafe_allow_html=True)
            st.button("🔒 Sin acceso", key=f"btn_modo_{key_suffix}_bloqueado",
                      use_container_width=True, disabled=True)

    # Icono · color de identidad · título · descripción · función de permiso ·
    # sufijo de key · valor de app_mode · botón primario. El orden es el mismo
    # que antes (Control primero, Cuenta Ficha al final); solo cambió que ahora
    # viven en una lista + helper en vez de 18 bloques HTML repetidos a mano.
    _MODULOS_BIENVENIDA = [
        ("📋", "#1a3a5c", "Control y Gestión Post Venta",
         "Dashboard completo: detalle, edición, documentos, comentarios, "
         "análisis, rankings y administración.",
         _puede_usar_control, "ots", "ots", True),
        ("🔄", "#2e7d32", "Planificador de Taller",
         "JPCB por patente, Agenda Curifor, Control de Taller y Vehículos "
         "en Taller — todo en tiempo real.",
         _puede_usar_planificador, "plan", "planificador", False),
        ("📈", "#b7791f", "Indicadores Post Venta",
         "Avances de facturación y KPIs del área, directo desde Power BI "
         "en tiempo real.",
         _puede_usar_indicadores, "ind", "indicadores", False),
        ("🤖", "#6b46c1", "Asistente App",
         "Consulta rápida por patente o folio: OT abierta, sucursal, "
         "costos y Vale de Consumo.",
         _puede_usar_asistente_app, "asis", "asistente", False),
        ("🧮", "#0b7d43", "Cotizador de Mantenciones",
         "Cotiza mantenciones por marca, modelo, versión y año: operaciones, "
         "repuestos con stock, adicionales y packs.",
         _puede_usar_cotizador, "cot", "cotizador", False),
        ("🚩", "#c0392b", "Revisión de Campañas Ford",
         "Casos con campaña/boletín pendiente (Agenda Ford): vencidas, "
         "por revisar y ya revisadas.",
         _puede_usar_campanas, "camp", "campanas", False),
        ("🗓️", "#0b5cad", "Agenda de Taller",
         "Agenda de citas de mantención por sucursal, día y hora. "
         "Reemplaza a la agenda web anterior.",
         _puede_usar_agenda_taller, "agenda", "agenda", False),
        ("📋", "#6a1b9a", "Recepción de Vehículos",
         "Recepción: checklist de accesorios, inspección, firma e ingreso "
         "a taller. Reemplaza a la recepción anterior.",
         _puede_usar_recepcion, "recep", "recepcion", False),
        ("💳", "#0b6b3a", "Cuenta Ficha",
         "Saldo disponible del cliente por sucursal e historial completo "
         "de sus OT con todos sus documentos.",
         _puede_usar_cuenta_ficha, "cuenta_ficha", "cuenta_ficha", False),
        ("📑", "#7b341e", "Informes de Gestión",
         "Reportes por marca (Ford, AG) con el histórico mes a mes y el "
         "avance del mes en curso desde la Alimentación semanal.",
         _puede_usar_informes_gestion, "infges", "informes_gestion", False),
        ("🚗", "#12694d", "Loaners",
         "Flota de vehículos de cortesía: qué unidad está disponible, cuál "
         "está prestada, a qué cliente y desde cuándo.",
         _puede_usar_loaners, "loaners", "loaners", False),
        ("🪟", "#455a64", "Vista Dividida",
         "Abre 2 módulos a la vez, lado a lado: Control y Gestión, Planificador, "
         "Indicadores, Campañas Ford o Cuenta Ficha.",
         (lambda email, usuarios_list=None: True), "split", "split", False),
    ]

    for _fila_ini in range(0, len(_MODULOS_BIENVENIDA), 3):
        _cols_fila = st.columns(3, gap="medium")
        for _col, _mod_info in zip(_cols_fila, _MODULOS_BIENVENIDA[_fila_ini:_fila_ini + 3]):
            _icono, _color, _titulo, _desc, _fn_permiso, _key_suf, _mode_val, _primary = _mod_info
            with _col:
                _tarjeta_modulo(_icono, _color, _titulo, _desc,
                                _fn_permiso(usuario_activo, _usuarios_bienvenida),
                                _key_suf, _mode_val, _primary)
        st.markdown("<div style='height:18px'></div>", unsafe_allow_html=True)

    st.stop()


# ============================================================
#   INDICADORES POST VENTA  — modo independiente
#   Informe Power BI embebido via link publico "Publish to web"
#   (iframe, sin login de Microsoft). La URL NO se hardcodea aqui
#   porque app.py vive en un repo publico de GitHub — se lee desde
#   Streamlit Secrets (PBI_INDICADORES_URL, solo visible para el
#   admin). Acceso al modulo restringido (puede_indicadores).
#   Se ubica ANTES de cargar_datos() porque no depende del JSON
#   de OTs — abre al instante aunque los datos fallen. 07/07/2026
# ============================================================
def _render_indicadores(pane=None):

    if not _puede_usar_indicadores(usuario_activo):
        if pane is None:
            st.session_state.pop("app_mode", None)
            st.error("🔒 No tienes autorización para ver los Indicadores Post Venta. "
                     "Solicítala al administrador (cjerez@curifor.com).")
            if st.button("← Volver al inicio", key="ind_sin_acceso_volver"):
                st.rerun()
            st.stop()
        else:
            st.error("🔒 No tienes autorización para ver los Indicadores Post Venta.")
            return

    # Configurar en share.streamlit.io -> Settings -> Secrets:
    #   PBI_INDICADORES_URL = "https://app.powerbi.com/view?r=..."
    _PBI_IND_URL = st.secrets.get("PBI_INDICADORES_URL", "").strip()

    with st.sidebar:
        if pane is None:
            st.markdown(
                f'<img src="{LOGO_DATA_URI}" style="max-width:180px;margin-bottom:0.4rem;"/>',
                unsafe_allow_html=True,
            )
            st.markdown("")
            if st.button("← Volver al inicio", use_container_width=True, key="ind_volver"):
                st.session_state.pop("app_mode", None)
                st.rerun()
            st.divider()

        st.markdown("### 🧭 Páginas del informe")
        st.caption("Navega entre páginas con las flechas o el menú inferior del informe.")
        st.markdown(
            """
            <div style="display:flex;flex-wrap:wrap;gap:6px;margin-top:6px;">
                <span class="info-chip">1 · Post Venta General</span>
                <span class="info-chip">2 · Servicio Técnico</span>
                <span class="info-chip">3 · DyP</span>
                <span class="info-chip">4 · Avance Facturación</span>
                <span class="info-chip">5 · Venta Repuestos</span>
                <span class="info-chip">6 · Pronóstico Ventas</span>
            </div>
            """,
            unsafe_allow_html=True,
        )

        if pane is None:
            st.divider()
            st.markdown(f"**Usuario:** `{usuario_activo}`")
            if st.button("🚪 Cerrar sesión", use_container_width=True, key="ind_logout"):
                for _k in ["authenticated", "user_email", "app_mode"]:
                    st.session_state.pop(_k, None)
                st.rerun()

    st.markdown(
        f'''<div class="curifor-header">
            <div class="logo-pill"><img src="{LOGO_DATA_URI}" /></div>
            <div class="curifor-header-text">
                <h2>Indicadores Post Venta</h2>
                <p>Avances de facturación · Power BI en tiempo real</p>
                <span class="dev-credit">Curifor S.A</span>
            </div>
            <span class="curifor-badge">📈 Power BI</span>
        </div>''',
        unsafe_allow_html=True,
    )

    if not _PBI_IND_URL:
        st.warning(
            "⚙️ El informe aún no está configurado. El administrador debe "
            "agregar `PBI_INDICADORES_URL` en los Secrets de Streamlit Cloud "
            "(Settings → Secrets)."
        )
    else:
        components.iframe(_PBI_IND_URL, height=(820 if pane is None else 560), scrolling=True)

    if pane is None:
        st.stop()   # ← Detener aquí: no ejecutar el resto del app (modo OTs)


if st.session_state.get("app_mode") == "indicadores":
    _render_indicadores()


# ============================================================
#   COTIZADOR DE MANTENCIONES  — modo independiente
#   -------------------------------------------------
#   Plataforma web del "Cotizador de Mantenciones" (marca/modelo/versión/
#   año → operaciones, repuestos con stock, adicionales y packs, Excel)
#   embebida como componente HTML. Todos los datos (índice + stock + 273
#   pautas), el CSS, el JS del cotizador, la librería XLSX y el logo viajan
#   comprimidos gzip+base64 en cotizador_data.json (campo 'gz', ~1.1 MB) y
#   se descomprimen en el navegador con DecompressionStream — así el iframe
#   no necesita salida a CDN (bloqueada por la CSP de Streamlit Cloud, mismo
#   patrón ya usado en el Planificador). Se ubica ANTES de cargar_datos()
#   porque no depende del JSON de OTs. Acceso restringido (puede_cotizador).
#   21/07/2026.
# ============================================================
if st.session_state.get("app_mode") == "cotizador":

    if not _puede_usar_cotizador(usuario_activo):
        st.session_state.pop("app_mode", None)
        st.error("🔒 No tienes autorización para usar el Cotizador de Mantenciones. "
                 "Solicítala al administrador (cjerez@curifor.com).")
        if st.button("← Volver al inicio", key="cot_sin_acceso_volver"):
            st.rerun()
        st.stop()

    with st.sidebar:
        st.markdown(
            f'<img src="{LOGO_DATA_URI}" style="max-width:180px;margin-bottom:0.4rem;"/>',
            unsafe_allow_html=True,
        )
        st.markdown("")
        if st.button("← Volver al inicio", use_container_width=True, key="cot_volver"):
            st.session_state.pop("app_mode", None)
            st.rerun()

        st.divider()
        st.markdown("### 🧮 Cotizador de Mantenciones")
        st.caption(
            "Elige marca → modelo → versión (y año en Ford) y presiona "
            "**Ver plan de mantención**. Verás el carrusel de revisiones con "
            "operaciones, repuestos con disponibilidad en bodega, servicios "
            "adicionales y packs. Puedes descargar la cotización a Excel."
        )
        st.caption(
            "Valores referenciales con IVA incluido, en CLP. Herramienta de "
            "apoyo para Servicio; confirmar precios y stock con el sistema."
        )

        st.divider()
        st.markdown(f"**Usuario:** `{usuario_activo}`")
        if st.button("🚪 Cerrar sesión", use_container_width=True, key="cot_logout"):
            for _k in ["authenticated", "user_email", "app_mode"]:
                st.session_state.pop(_k, None)
            st.rerun()

    _cot_gz, _cot_fecha = _cargar_cotizador_gz()

    st.markdown(
        f'''<div class="curifor-header">
            <div class="logo-pill"><img src="{LOGO_DATA_URI}" /></div>
            <div class="curifor-header-text">
                <h2>Cotizador de Mantenciones</h2>
                <p>Mantenciones preventivas por marca, modelo y versión</p>
                <span class="dev-credit">Curifor S.A</span>
            </div>
            {f'<span class="curifor-badge">🧮 {_cot_fecha}</span>' if _cot_fecha else ''}
        </div>''',
        unsafe_allow_html=True,
    )

    if not _cot_gz:
        st.warning(
            "⚙️ El cotizador aún no está disponible. Falta subir "
            "`cotizador_data.json` al repositorio con `Subir_App_GitHub.bat`. "
            "Una vez subido, este módulo cargará automáticamente."
        )
    else:
        _COT_HTML_TPL = """<!doctype html>
<html lang="es-CL"><head><meta charset="utf-8">
<style id="cotizStyle"></style>
</head><body style="margin:0;background:transparent;">
<div id="cotizRoot" style="font-family:system-ui,'Segoe UI',Roboto,Arial,sans-serif;padding:16px;color:#334;">
  <p id="cotizMsg">Cargando cotizador…</p>
</div>
<script id="cotizData" type="application/octet-stream">__GZ_B64__</script>
<script>
(async function(){
  var msg = document.getElementById('cotizMsg');
  try {
    if (typeof DecompressionStream === 'undefined') {
      msg.textContent = 'Tu navegador no soporta la descompresion requerida. Usa Chrome, Edge o Firefox actualizados.';
      return;
    }
    var b64 = document.getElementById('cotizData').textContent.trim();
    var bin = Uint8Array.from(atob(b64), function(c){ return c.charCodeAt(0); });
    var ds = new DecompressionStream('gzip');
    var stream = new Blob([bin]).stream().pipeThrough(ds);
    var buf = await new Response(stream).arrayBuffer();
    var pkg = JSON.parse(new TextDecoder().decode(buf));
    document.getElementById('cotizStyle').textContent = pkg.css;
    document.getElementById('cotizRoot').innerHTML = pkg.body;
    window._COTIZ = { indice: pkg.indice, stock: pkg.stock, pautas: pkg.pautas, logo: pkg.logo };
    var sx = document.createElement('script'); sx.textContent = pkg.xlsx; document.body.appendChild(sx);
    var sj = document.createElement('script'); sj.textContent = pkg.js; document.body.appendChild(sj);
    if (typeof window.__cotizInit === 'function') { window.__cotizInit(); }
  } catch (e) {
    msg.textContent = 'No se pudo cargar el cotizador: ' + e;
  }
})();
</script>
</body></html>"""
        _cot_html = _COT_HTML_TPL.replace("__GZ_B64__", _cot_gz)
        components.html(_cot_html, height=1500, scrolling=True)

    st.stop()   # ← Detener aquí: no ejecutar el resto del app (modo OTs)


# ============================================================
#   AGENDA Y RECEPCIÓN DE TALLER — módulos embebidos por iframe
#   Las nuevas plataformas de agenda y recepción viven en
#   platoniaaa/cotizador-mantenciones (GitHub Pages) y se embeben
#   por URL con ?vista= para abrir cada una en su pestaña. Reemplazan
#   a agenda.curifor.cl y recepcion.curifor.cl. 29/07/2026.
# ============================================================
if st.session_state.get("app_mode") in ("agenda", "recepcion"):
    _tal_modo = st.session_state["app_mode"]

    # 29/07/2026 — guardia de acceso (defensa en profundidad): aunque la
    # tarjeta de la pantalla de bienvenida ya se oculta sin permiso, si
    # alguien llega aca con el app_mode seteado por otra via (session_state
    # viejo, etc.) se corta antes de cargar el iframe.
    _tal_ok = (_puede_usar_agenda_taller(usuario_activo)
               if _tal_modo == "agenda"
               else _puede_usar_recepcion(usuario_activo))
    if not _tal_ok:
        st.session_state.pop("app_mode", None)
        st.error("🔒 No tienes autorización para ver "
                 + ("la Agenda de Taller. " if _tal_modo == "agenda"
                    else "la Recepción de Vehículos. ")
                 + "Solicítala al administrador (cjerez@curifor.com).")
        if st.button("← Volver al inicio", key="tal_sin_acceso_volver"):
            st.rerun()
        st.stop()

    _tal_titulo = "Agenda de Taller" if _tal_modo == "agenda" else "Recepción de Vehículos"
    _tal_desc = ("Agenda de citas por sucursal, día y hora"
                 if _tal_modo == "agenda"
                 else "Recepción de vehículos: checklist, inspección y firma")
    _tal_url = "https://platoniaaa.github.io/cotizador-mantenciones/taller.html?vista=" + _tal_modo

    with st.sidebar:
        st.markdown(
            f'<img src="{LOGO_DATA_URI}" style="max-width:180px;margin-bottom:0.4rem;"/>',
            unsafe_allow_html=True,
        )
        if st.button("← Volver al inicio", use_container_width=True, key="tal_volver"):
            st.session_state.pop("app_mode", None)
            st.rerun()
        st.divider()
        st.markdown(f"**Usuario:** `{usuario_activo}`")
        if st.button("🚪 Cerrar sesión", use_container_width=True, key="tal_logout"):
            for _k in ["authenticated", "user_email", "app_mode"]:
                st.session_state.pop(_k, None)
            st.rerun()

    st.markdown(
        f'''<div class="curifor-header">
            <div class="logo-pill"><img src="{LOGO_DATA_URI}" /></div>
            <div class="curifor-header-text">
                <h2>{_tal_titulo}</h2>
                <p>{_tal_desc}</p>
                <span class="dev-credit">Curifor S.A</span>
            </div>
        </div>''',
        unsafe_allow_html=True,
    )

    components.iframe(_tal_url, height=1600, scrolling=True)
    st.stop()


# ============================================================
#   REVISIÓN DE CAMPAÑAS  — modo independiente
#   -------------------------------------------
#   Datos de la Agenda Ford (archivo de OTRO proyecto local,
#   '*_Consolidado_Curifor_2Tandas.xlsx'), subidos por
#   consolidar_OTs.py (PASO 12) a campanas_curifor.json. Solo viajan los
#   casos que tienen datos en la columna W (Campañas/Boletín). Estado:
#     🔴 Campaña No Realizada — fecha de programación anterior a hoy.
#     🔵 Cita de Hoy          — fecha de programación = hoy, Status En Curso o Agendado.
#     🟡 No revisada          — fecha de hoy en adelante (resto de casos), AA = No revisado.
#     🟢 Cita Revisada        — fecha de hoy en adelante (resto de casos), AA = Revisado.
#   Se ubica ANTES de cargar_datos() porque no depende del JSON de OTs — abre
#   al instante aunque los datos del dashboard principal fallen (mismo patrón
#   que Indicadores/Cotizador). Acceso restringido (puede_campanas) y
#   respeta la misma restricción de sucursal por usuario que el resto de la
#   app (sucursales_permitidas). 28/07/2026.
# ============================================================
def _render_campanas(pane=None):

    if not _puede_usar_campanas(usuario_activo):
        if pane is None:
            st.session_state.pop("app_mode", None)
            st.error("🔒 No tienes autorización para ver Revisión de Campañas Ford. "
                     "Solicítala al administrador (cjerez@curifor.com).")
            if st.button("← Volver al inicio", key="camp_sin_acceso_volver"):
                st.rerun()
            st.stop()
        else:
            st.error("🔒 No tienes autorización para ver Revisión de Campañas Ford.")
            return

    _camp_casos, _camp_fecha_act, _camp_archivo = _cargar_campanas()

    df_camp = pd.DataFrame(_camp_casos)
    for _c in ["sucursal", "estado_color", "estado_texto", "fecha_programacion",
               "modelo", "patente", "chasis", "propietario", "campanas",
               "orden_servicio", "fecha_cierre", "recall_codigos"]:
        if _c not in df_camp.columns:
            df_camp[_c] = ""
    if "recall_obligatorio" not in df_camp.columns:
        df_camp["recall_obligatorio"] = False
    df_camp["recall_obligatorio"] = df_camp["recall_obligatorio"].fillna(False).astype(bool)

    # Restricción de sucursal por usuario — misma lógica que el resto de la app
    # (un usuario con sucursales_permitidas SOLO ve/filtra sus propias sucursales).
    _usuarios_camp = _leer_usuarios()
    _mis_suc_camp = _sucursales_permitidas_usuario(usuario_activo, _usuarios_camp)
    if _mis_suc_camp:
        _mis_suc_camp_norm = {s.strip().upper() for s in _mis_suc_camp}
        df_camp = df_camp[df_camp["sucursal"].str.upper().isin(_mis_suc_camp_norm)].reset_index(drop=True)

    with st.sidebar:
        if pane is None:
            st.markdown(
                f'<img src="{LOGO_DATA_URI}" style="max-width:180px;margin-bottom:0.4rem;"/>',
                unsafe_allow_html=True,
            )
            st.markdown("")
            if st.button("← Volver al inicio", use_container_width=True, key="camp_volver"):
                st.session_state.pop("app_mode", None)
                st.rerun()
            st.divider()

        st.markdown("### 🚩 Revisión de Campañas Ford")
        st.markdown(
            """
            <div style="display:flex;flex-direction:column;gap:6px;margin:6px 0 10px;">
                <span class="info-chip" style="background:rgba(192,57,43,.14);color:#c0392b;">
                    🔴 Vencida sin revisar</span>
                <span class="info-chip" style="background:rgba(47,126,219,.14);color:#2f7edb;">
                    🔵 Cita de hoy · En Curso/Agendado</span>
                <span class="info-chip" style="background:rgba(217,145,0,.14);color:#b97900;">
                    🟡 Próxima · aún sin revisar</span>
                <span class="info-chip ok">🟢 Próxima · ya revisada</span>
            </div>
            """,
            unsafe_allow_html=True,
        )
        if _mis_suc_camp:
            st.caption(f"🔒 Acceso limitado a: {', '.join(_mis_suc_camp)}")

        st.markdown("### Filtros")
        _suc_opts_camp = sorted(df_camp["sucursal"].unique().tolist())
        _sel_suc_camp = st.multiselect("Sucursal", _suc_opts_camp, placeholder="Todas", key="camp_f_suc")
        _sel_estado_camp = st.multiselect(
            "Estado",
            ["🔴 Campaña No Realizada", "🔵 Cita de Hoy", "🟡 No revisada", "🟢 Cita Revisada"],
            placeholder="Todos", key="camp_f_estado",
        )
        _busq_camp = st.text_input(
            "Buscar (patente, propietario, campaña...)", "", key="camp_f_busq"
        )
        _solo_recall_camp = st.checkbox(
            "🚨 Solo Recall Obligatorio (FSA)", value=False, key="camp_f_recall",
            help="Muestra solo las citas cuya columna Campañas/Boletín trae al "
                 "menos uno de los códigos FSA de recall obligatorio.",
        )

        st.divider()
        if st.button("🔄 Actualizar datos", use_container_width=True, key="camp_actualizar"):
            st.cache_data.clear()
            st.rerun()

        if pane is None:
            st.divider()
            st.markdown(f"**Usuario:** `{usuario_activo}`")
            if st.button("🚪 Cerrar sesión", use_container_width=True, key="camp_logout"):
                for _k in ["authenticated", "user_email", "app_mode"]:
                    st.session_state.pop(_k, None)
                st.rerun()

    st.markdown(
        f'''<div class="curifor-header">
            <div class="logo-pill"><img src="{LOGO_DATA_URI}" /></div>
            <div class="curifor-header-text">
                <h2>Revisión de Campañas Ford</h2>
                <p>Agenda Ford{(" · " + _camp_archivo) if _camp_archivo else ""}{(" · Actualizado " + _camp_fecha_act) if _camp_fecha_act else ""}</p>
                <span class="dev-credit">Curifor S.A</span>
            </div>
            <span class="curifor-badge" style="background:rgba(192,57,43,.22);border-color:rgba(192,57,43,.4);">🚩 Ford</span>
        </div>''',
        unsafe_allow_html=True,
    )

    if not _camp_casos:
        st.warning(
            "⚙️ Aún no hay datos de Revisión de Campañas Ford. El administrador debe "
            "correr `Ejecutar_Consolidacion.bat` con el archivo de la Agenda "
            "Ford disponible en su carpeta correspondiente."
        )
        if pane is None:
            st.stop()
        return

    if df_camp.empty:
        st.info("No hay casos de campañas para las sucursales que puedes ver.")
        if pane is None:
            st.stop()
        return

    df_camp_f = df_camp.copy()
    if _sel_suc_camp:
        df_camp_f = df_camp_f[df_camp_f["sucursal"].isin(_sel_suc_camp)]
    if _sel_estado_camp:
        _map_estado_sel = {
            "🔴 Campaña No Realizada": "rojo",
            "🔵 Cita de Hoy": "azul",
            "🟡 No revisada": "amarillo",
            "🟢 Cita Revisada": "verde",
        }
        _colores_sel = [_map_estado_sel[e] for e in _sel_estado_camp]
        df_camp_f = df_camp_f[df_camp_f["estado_color"].isin(_colores_sel)]
    if _busq_camp:
        _mask_camp = pd.Series([False] * len(df_camp_f), index=df_camp_f.index)
        for _col in ["patente", "propietario", "campanas", "modelo", "chasis"]:
            _mask_camp |= df_camp_f[_col].astype(str).str.contains(_busq_camp, case=False, na=False)
        df_camp_f = df_camp_f[_mask_camp]
    if _solo_recall_camp:
        df_camp_f = df_camp_f[df_camp_f["recall_obligatorio"] == True]

    st.caption(f"Mostrando {len(df_camp_f):,} de {len(df_camp):,} caso(s) con campaña/boletín")

    _n_recall_camp = int(df_camp_f["recall_obligatorio"].sum())
    if _n_recall_camp:
        st.warning(
            f"🚨 **{_n_recall_camp} cita(s) con Recall Obligatorio (FSA)** — "
            "campañas que Ford exige realizar sin excepción. Revísalas primero."
        )

    # ---- KPIs ----
    _n_total_camp = len(df_camp_f)
    _n_rojo_camp  = int((df_camp_f["estado_color"] == "rojo").sum())
    _n_azul_camp  = int((df_camp_f["estado_color"] == "azul").sum())
    _n_amar_camp  = int((df_camp_f["estado_color"] == "amarillo").sum())
    _n_verde_camp = int((df_camp_f["estado_color"] == "verde").sum())

    _kc1, _kc2, _kc3, _kc4, _kc5, _kc6 = st.columns(6)
    with _kc1: kpi_card("Total con campaña", _n_total_camp)
    with _kc2: kpi_card("🚨 Recall Obligatorio", _n_recall_camp, "naranja")
    with _kc3: kpi_card("🔴 Campaña No Realizada", _n_rojo_camp, "rojo")
    with _kc4: kpi_card("🔵 Cita de Hoy", _n_azul_camp, "azul")
    with _kc5: kpi_card("🟡 No revisada", _n_amar_camp, "amarillo")
    with _kc6: kpi_card("🟢 Cita Revisada", _n_verde_camp, "verde")

    st.markdown("<div style='height:14px'></div>", unsafe_allow_html=True)

    # ---- Tabla: solo Sucursal + las columnas que se pidieron (M/L/N/P/W) ----
    _map_emoji_camp = {"rojo": "🔴", "azul": "🔵", "amarillo": "🟡", "verde": "🟢"}
    df_camp_tabla = df_camp_f.copy()
    df_camp_tabla["Estado"] = df_camp_tabla.apply(
        lambda r: f"{_map_emoji_camp.get(r['estado_color'], '')} {r['estado_texto']}", axis=1
    )
    df_camp_tabla["Recall FSA"] = df_camp_tabla.apply(
        lambda r: f"🚨 {r['recall_codigos']}" if r["recall_obligatorio"] else "", axis=1
    )
    df_camp_tabla = df_camp_tabla.rename(columns={
        "sucursal": "Sucursal", "orden_servicio": "N° OT",
        "fecha_programacion": "Fecha Programación", "patente": "Placa / Patente",
        "modelo": "Modelo", "chasis": "Chasis", "propietario": "Nombre Propietario",
        "campanas": "Campañas/Boletín", "fecha_cierre": "Fecha Cierre",
    })
    _cols_tabla_camp = ["Estado", "Recall FSA", "Sucursal", "N° OT", "Fecha Programación",
                        "Placa / Patente", "Modelo", "Chasis", "Nombre Propietario",
                        "Campañas/Boletín", "Fecha Cierre"]
    _orden_estado = {"🔴 Campaña No Realizada": 0, "🔵 Cita de Hoy": 1,
                      "🟡 No revisada": 2, "🟢 Cita Revisada": 3}
    # Las citas con Recall Obligatorio (🚨) siempre suben al principio de la
    # tabla, sin importar su Estado -- son campañas que Ford exige sí o sí.
    df_camp_tabla["_orden_recall"] = (~df_camp_tabla["recall_obligatorio"]).astype(int)
    df_camp_tabla["_orden_estado"] = df_camp_tabla["Estado"].map(lambda v: _orden_estado.get(v, 9))
    df_camp_tabla = df_camp_tabla.sort_values(by=["_orden_recall", "_orden_estado"])
    st.dataframe(
        df_camp_tabla[_cols_tabla_camp], hide_index=True, use_container_width=True, height=560,
        column_config={
            "Estado": st.column_config.TextColumn(width="medium"),
            "Recall FSA": st.column_config.TextColumn(
                width="medium",
                help="🚨 = la campaña incluye un código FSA de recall obligatorio de Ford",
            ),
            "Campañas/Boletín": st.column_config.TextColumn(width="medium"),
        },
    )

    _buf_camp = io.BytesIO()
    with pd.ExcelWriter(_buf_camp, engine="openpyxl") as _writer_camp:
        df_camp_tabla[_cols_tabla_camp].to_excel(_writer_camp, index=False, sheet_name="Revisión de Campañas Ford")
        _ws_camp = _writer_camp.sheets["Revisión de Campañas Ford"]
        for _col_cells in _ws_camp.columns:
            _max_len = max((len(str(_cell.value)) if _cell.value is not None else 0) for _cell in _col_cells)
            _ws_camp.column_dimensions[_col_cells[0].column_letter].width = min(_max_len + 4, 60)
    _buf_camp.seek(0)
    st.download_button(
        "⬇️ Descargar como Excel", _buf_camp,
        f"Revision_Campanas_Ford_{datetime.now().strftime('%Y%m%d')}.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        use_container_width=True, key="camp_download",
    )

    if pane is None:
        st.stop()


if st.session_state.get("app_mode") == "campanas":
    _render_campanas()


# ============================================================
#   INFORMES DE GESTION — modo independiente
#   Reportes de gestion que Cristobal envia a cada automotora, organizados
#   por marca en un submenu: Ford, AG, GM y Omoda-Jaecoo (estas dos ultimas
#   quedan pendientes). Los datos los genera informes_gestion.py
#   (Actualizar_Informes_Gestion.bat) y viajan en informes_gestion.json:
#     * historico mes a mes (AG por sucursal, Ford por hoja mensual)
#     * mes en curso calculado desde la carpeta Alimentacion
#   Acceso restringido por usuario (puede_informes_gestion). 04/08/2026.
# ============================================================
def _render_informes_gestion(pane=None):

    if not _puede_usar_informes_gestion(usuario_activo):
        if pane is None:
            st.session_state.pop("app_mode", None)
            st.error("🔒 No tienes autorización para ver Informes de Gestión. "
                     "Solicítala al administrador (cjerez@curifor.com).")
            if st.button("← Volver al inicio", key="ig_sin_acceso_volver"):
                st.rerun()
            st.stop()
        else:
            st.error("🔒 No tienes autorización para ver Informes de Gestión.")
            return

    _k = f"ig{pane or ''}_"
    _ig = _cargar_informes_gestion()

    # ---------- estilos del modulo ----------
    st.markdown("""
    <style>
      .ig-hero{border-radius:18px;padding:20px 26px;margin-bottom:16px;color:#fff;
               background:linear-gradient(135deg,#5c2410 0%,#9c4221 55%,#c05621 100%);
               box-shadow:0 10px 26px rgba(92,36,16,.28);}
      .ig-hero h2{margin:0;font-size:1.5rem;font-weight:800;letter-spacing:-.4px;color:#fff;}
      .ig-hero p{margin:5px 0 0;opacity:.9;font-size:.85rem;}
      .ig-kpis{display:flex;gap:14px;flex-wrap:wrap;margin:4px 0 18px;}
      .ig-kpi{flex:1;min-width:150px;border-radius:14px;padding:15px 18px;
              background:linear-gradient(135deg,#0d2f5a 0%,#1a4f8a 100%);color:#fff;
              box-shadow:0 6px 18px rgba(13,47,90,.22);}
      .ig-kpi.verde{background:linear-gradient(135deg,#0b6b3a 0%,#18a05a 100%);
                    box-shadow:0 6px 18px rgba(11,107,58,.22);}
      .ig-kpi.ambar{background:linear-gradient(135deg,#9a6100 0%,#d99100 100%);
                    box-shadow:0 6px 18px rgba(154,97,0,.22);}
      .ig-kpi.rojo{background:linear-gradient(135deg,#8c2a1f 0%,#c0392b 100%);
                   box-shadow:0 6px 18px rgba(140,42,31,.22);}
      .ig-kpi.gris{background:linear-gradient(135deg,#3d4b5c 0%,#5b6d80 100%);
                   box-shadow:0 6px 18px rgba(61,75,92,.22);}
      .ig-kpi .v{font-size:1.7rem;font-weight:800;line-height:1.15;letter-spacing:-.5px;}
      .ig-kpi .l{font-size:.72rem;opacity:.88;text-transform:uppercase;
                 letter-spacing:.6px;margin-top:4px;}
      .ig-kpi .d{font-size:.72rem;opacity:.95;margin-top:6px;font-weight:600;}
      .ig-chips{display:flex;gap:6px;flex-wrap:wrap;margin:2px 0 12px;}
      .ig-chip{font-size:.72rem;padding:4px 11px;border-radius:20px;
               background:rgba(156,66,33,.13);color:#9c4221;font-weight:600;white-space:nowrap;}
      .ig-chip.ok{background:rgba(24,160,90,.14);color:#0b6b3a;}
      .ig-chip.curso{background:rgba(217,145,0,.16);color:#8a5c00;}
      .ig-chip.mut{background:rgba(128,128,128,.14);color:#6b7885;}
      .ig-sec{font-size:1rem;font-weight:700;color:var(--text-color);
              margin:20px 0 8px;display:flex;align-items:center;gap:9px;}
      .ig-sec:before{content:"";width:5px;height:19px;border-radius:3px;background:#9c4221;}
      .ig-tblwrap{border:1px solid rgba(128,128,128,.25);border-radius:14px;
                  overflow:hidden;margin-bottom:6px;}
      table.ig-tbl{width:100%;border-collapse:collapse;font-size:.86rem;}
      table.ig-tbl th{background:linear-gradient(135deg,#0d2f5a 0%,#1a4f8a 100%);
                      color:#fff;font-weight:700;padding:10px 12px;text-align:right;
                      font-size:.76rem;text-transform:uppercase;letter-spacing:.4px;}
      table.ig-tbl th:first-child{text-align:left;}
      table.ig-tbl td{padding:8px 12px;text-align:right;
                      border-top:1px solid rgba(128,128,128,.16);
                      color:var(--text-color);font-variant-numeric:tabular-nums;}
      table.ig-tbl td:first-child{text-align:left;font-weight:500;}
      table.ig-tbl tr.sec td{background:rgba(128,128,128,.10);font-weight:700;
                             font-size:.74rem;text-transform:uppercase;
                             letter-spacing:.6px;color:#7a8794;padding:7px 12px;}
      table.ig-tbl tr.tot td{background:rgba(26,79,138,.09);font-weight:800;}
      table.ig-tbl tr.mes-curso td{background:rgba(217,145,0,.10);}
      table.ig-tbl td.nulo{color:rgba(128,128,128,.45);}
      table.ig-tbl td.mon{font-weight:600;}
      table.ig-tbl tr:hover td{background:rgba(26,79,138,.05);}
      table.ig-tbl tr.sec:hover td{background:rgba(128,128,128,.10);}
      .ig-empty{border:1px dashed rgba(128,128,128,.4);border-radius:14px;padding:30px;
                text-align:center;color:#7a8794;}
      .ig-delta{font-weight:700;}
      .ig-delta.up{color:#18a05a;} .ig-delta.dn{color:#c0392b;}
    </style>
    """, unsafe_allow_html=True)

    # ---------- helpers de formato ----------
    def _n(v):
        """Numero entero con separador de miles chileno."""
        if v is None or (isinstance(v, float) and v != v):
            return None
        try:
            return f"{int(round(float(v))):,}".replace(",", ".")
        except Exception:
            return None

    def _m(v):
        """Monto en pesos."""
        t = _n(v)
        return f"$ {t}" if t is not None else None

    def _celda(v, monto=False):
        t = _m(v) if monto else _n(v)
        if t is None or t in ("0", "$ 0"):
            return f'<td class="nulo">{t or "–"}</td>' if t else '<td class="nulo">–</td>'
        return f'<td class="{"mon" if monto else ""}">{t}</td>'

    def _kpi(valor, label, clase="", detalle=""):
        _d = f'<div class="d">{detalle}</div>' if detalle else ""
        return (f'<div class="ig-kpi {clase}"><div class="v">{valor}</div>'
                f'<div class="l">{label}</div>{_d}</div>')

    def _delta(actual, previo):
        """Variacion vs el mes anterior, en texto."""
        if not previo or actual is None:
            return ""
        d = actual - previo
        if d == 0:
            return "= igual que el mes anterior"
        pct = (d / previo * 100) if previo else 0
        return f'{"▲" if d > 0 else "▼"} {_n(abs(d))} ({pct:+.0f}%) vs mes anterior'

    # Que filas van en formato moneda. No basta con mirar el nombre de la fila
    # (la fila "TOTAL" del bloque de ingresos no dice "Venta"), asi que tambien
    # se arrastra la seccion en la que va cada fila.
    _SEC_MONTO = ("VENTA", "PAGO", "$", "INGRESO", "MATERIAL", "REPUESTO",
                  "MONTO", "INVENTARIO", "FACTURA")
    _NO_MONTO = ("HORA", "DIA", "PERSONAL", "PASO", "UNIDAD", "VEHICULO",
                 "ROTACION", "%")

    def _es_monto(label, seccion=""):
        t, s = _norm_txt_ig(label), _norm_txt_ig(seccion)
        if any(p in t for p in _NO_MONTO):
            return False
        if any(p in t for p in _SEC_MONTO) or "MESON" in t:
            return True
        return any(p in s for p in _SEC_MONTO)

    # ---------- hero ----------
    _fecha_act = _ig.get("fecha_actualizacion", "")
    st.markdown(f"""
    <div class="ig-hero">
      <h2>📑 Informes de Gestión</h2>
      <p>Reportes por marca · histórico mes a mes y avance del mes en curso
         {('&nbsp;·&nbsp; actualizado ' + _fecha_act) if _fecha_act else ''}</p>
    </div>""", unsafe_allow_html=True)

    if pane is None:
        with st.sidebar:
            st.markdown(f'<img src="{LOGO_DATA_URI}" style="max-width:170px;'
                        'margin-bottom:.5rem;" />', unsafe_allow_html=True)
            if st.button("← Volver al inicio", use_container_width=True,
                         key=f"{_k}volver"):
                st.session_state.pop("app_mode", None)
                st.rerun()
            st.divider()

    if not _ig:
        st.markdown(
            '<div class="ig-empty"><b>Todavía no hay datos cargados.</b><br>'
            'Corre <code>Actualizar_Informes_Gestion.bat</code> en la carpeta del '
            'proyecto para leer la carpeta <i>Informe de gestión</i> y publicar '
            'los datos en la App.</div>', unsafe_allow_html=True)
        if pane is None:
            st.stop()
        return

    _actual = _ig.get("actual", {}) or {}
    _periodo_curso = _actual.get("periodo", "")
    try:
        _im = int(str(_periodo_curso)[4:6]) - 1
        _mes_corto, _mes_largo = MESES_CORTOS_IG[_im], MESES_LARGOS_IG[_im]
    except Exception:
        _im, _mes_corto, _mes_largo = None, "", ""

    # ---------- submenu por marca ----------
    _MARCAS = [("🔵", "Ford"), ("🟠", "AG"), ("⚪", "GM"), ("⚪", "Omoda - Jaecoo")]
    _marca = st.radio(
        "Marca", [f"{i}  {n}" for i, n in _MARCAS], horizontal=True,
        label_visibility="collapsed", key=f"{_k}marca")
    _marca_n = _marca.split("  ", 1)[-1]

    # =========================================================
    #   FORD  — Informe IMOP, una sucursal a la vez
    # =========================================================
    if _marca_n == "Ford":
        _ford = _ig.get("ford", {}) or {}
        _meses_ford = _ford.get("meses", {}) or {}
        if not _meses_ford:
            st.markdown('<div class="ig-empty">No se encontró el informe IMOP de '
                        'Ford en la carpeta <i>Informe de gestión</i>.</div>',
                        unsafe_allow_html=True)
            if pane is None:
                st.stop()
            return

        _orden_meses = [m for m in MESES_LARGOS_IG if m in _meses_ford]

        def _mes_tiene_datos(m):
            # Los meses que aun no ocurren vienen con las celdas en 0, no vacias.
            return any(any(v for v in f["valores"].values())
                       for f in _meses_ford[m]["filas"])

        _con_datos = [m for m in _orden_meses if _mes_tiene_datos(m)]
        _idx_def = (_orden_meses.index(_con_datos[-1]) if _con_datos
                    else len(_orden_meses) - 1)

        _cols_todas = []
        for _m2 in _orden_meses:
            for _c in _meses_ford[_m2].get("columnas", []):
                if _c not in _cols_todas:
                    _cols_todas.append(_c)

        # "Rancagua Livianos" -> sucursal "Rancagua", area "Livianos"
        _AREAS_IG = ["Meson", "Livianos", "Pesados", "Camiones", "DYP", "DyP"]

        def _split_col(c):
            for a in sorted(_AREAS_IG, key=len, reverse=True):
                if _norm_txt_ig(c).endswith(_norm_txt_ig(a)):
                    return c[: len(c) - len(a)].strip(), a
            return c.strip(), ""

        _sucs_ford, _por_suc = [], {}
        for _c in _cols_todas:
            _s, _a = _split_col(_c)
            if _s not in _por_suc:
                _por_suc[_s] = []
                _sucs_ford.append(_s)
            _por_suc[_s].append((_c, _a))

        _c1, _c2 = st.columns([3, 2])
        with _c1:
            _suc_sel = st.selectbox("🏢 Sucursal", _sucs_ford, key=f"{_k}ford_suc")
        with _c2:
            _mes_sel = st.selectbox("📅 Mes del informe", _orden_meses,
                                    index=_idx_def, key=f"{_k}ford_mes")

        _hoja = _meses_ford[_mes_sel]
        _pares = _por_suc[_suc_sel]
        _areas_lbl = [a or c for c, a in _pares]

        _i_mes = _orden_meses.index(_mes_sel)
        _mes_prev = _orden_meses[_i_mes - 1] if _i_mes > 0 else None

        def _val(mes, label, col):
            for f in _meses_ford[mes]["filas"]:
                if f["label"] == label:
                    return f["valores"].get(col)
            return None

        def _suma_fila(mes, label):
            return sum((_val(mes, label, c) or 0) for c, _a in _pares)

        _tot_atn = _suma_fila(_mes_sel, "TOTAL")
        _tot_prev = _suma_fila(_mes_prev, "TOTAL") if _mes_prev else 0
        _mant = _suma_fila(_mes_sel, "Mantenciones")
        _gar = _suma_fila(_mes_sel, "Garantía")
        _int = _suma_fila(_mes_sel, "Interno")
        # El informe trae 2 filas "TOTAL": la 1a es del volumen de atenciones y
        # la 2a la de ingresos por ventas.
        _venta_tot = _venta_comp = 0
        _vistos = 0
        for _f in _hoja["filas"]:
            if _f["label"] == "TOTAL":
                _vistos += 1
                if _vistos == 2:
                    _venta_tot = sum((_f["valores"].get(c) or 0) for c, _a in _pares)
                    # base comparable con la Alimentacion: sin Mesón (son ventas
                    # de mostrador sin OT, que no viajan en la sábana)
                    _venta_comp = sum((_f["valores"].get(c) or 0) for c, _a in _pares
                                      if _norm_txt_ig(_a) != "MESON")

        st.markdown(
            f'<div class="ig-chips">'
            f'<span class="ig-chip">🏢 {_suc_sel}</span>'
            f'<span class="ig-chip">📅 {_mes_sel}</span>'
            f'<span class="ig-chip mut">{len(_pares)} área(s)</span>'
            + (f'<span class="ig-chip curso">Mes en curso: {_mes_largo}</span>'
               if _mes_largo else '') +
            f'</div>', unsafe_allow_html=True)

        st.markdown(
            '<div class="ig-kpis">'
            + _kpi(_n(_tot_atn) or "0", "Atenciones del mes", "",
                   _delta(_tot_atn, _tot_prev))
            + _kpi(_n(_mant) or "0", "Mantenciones", "verde")
            + _kpi(_n(_gar) or "0", "Garantía", "ambar")
            + _kpi(_n(_int) or "0", "Interno", "gris")
            + _kpi(_m(_venta_tot) or "$ 0", "Ingresos por ventas", "rojo")
            + '</div>', unsafe_allow_html=True)

        # ---- tablas: volumen e ingresos ----
        def _tabla_ford(desde, hasta, titulo, monto=False):
            _filas_html = []
            for _f in _hoja["filas"][desde:hasta]:
                if _f["tipo"] == "seccion":
                    _filas_html.append(
                        f'<tr class="sec"><td colspan="{len(_pares) + 1}">'
                        f'{_f["label"]}</td></tr>')
                    continue
                _hay = any(_f["valores"].get(c) is not None for c, _a in _pares)
                if not _hay and _f["tipo"] != "total":
                    continue
                # todo el bloque de ingresos va en moneda, incluida su fila TOTAL
                _mon = monto or _es_monto(_f["label"])
                _tds = "".join(_celda(_f["valores"].get(c), _mon) for c, _a in _pares)
                _cls = ' class="tot"' if _f["tipo"] == "total" else ""
                _filas_html.append(f'<tr{_cls}><td>{_f["label"]}</td>{_tds}</tr>')
            if not _filas_html:
                return ""
            _ths = "".join(f"<th>{a}</th>" for a in _areas_lbl)
            return (f'<div class="ig-sec">{titulo}</div>'
                    f'<div class="ig-tblwrap"><table class="ig-tbl">'
                    f'<thead><tr><th>Ítem</th>{_ths}</tr></thead>'
                    f'<tbody>{"".join(_filas_html)}</tbody></table></div>')

        # el informe trae 2 bloques: volumen (unidades) e ingresos (montos)
        _corte = len(_hoja["filas"])
        for _i, _f in enumerate(_hoja["filas"]):
            if _f["tipo"] == "seccion" and "INGRESO" in _norm_txt_ig(_f["label"]):
                _corte = _i
                break
        st.markdown(_tabla_ford(0, _corte, "📊 Volumen de atenciones (unidades)"),
                    unsafe_allow_html=True)
        st.markdown(_tabla_ford(_corte, len(_hoja["filas"]),
                                "💰 Ingresos por ventas", monto=True),
                    unsafe_allow_html=True)

        # ---- avance del mes en curso ----
        # El taller movil va APARTE de las sucursales (decision de Cristóbal):
        # las columnas "Camión Taller" del IMOP muestran el movil consolidado de
        # todas las sucursales, y una sucursal normal no incluye sus móviles.
        _es_movil_col = any(p in _norm_txt_ig(_suc_sel)
                            for p in ("CAMION TALLER", "MOVIL"))
        _f_actual = {}
        if _es_movil_col:
            _mov = {}
            for _s, _d in (_actual.get("ford", {}) or {}).items():
                for _kk, _vv in ((_d.get("movil") or {}).get("pasos") or {}).items():
                    _mov[_kk] = _mov.get(_kk, 0) + _vv
            if _mov.get("total"):
                _f_actual = {"movil": {"pasos": _mov}}
        else:
            for _s, _d in (_actual.get("ford", {}) or {}).items():
                if _norm_txt_ig(_s) == _norm_txt_ig(_suc_sel):
                    _f_actual = {_a: _v for _a, _v in _d.items() if _a != "movil"}
                    break
        if _f_actual:
            st.markdown(
                f'<div class="ig-sec">🟡 Avance del mes en curso — {_mes_largo or _periodo_curso}'
                f'</div>', unsafe_allow_html=True)
            _rows = []
            for _area, _lbl in [("livianos", "Livianos"), ("pesados", "Pesados / Camiones"),
                                ("dyp", "DYP"), ("movil", "Taller Móvil")]:
                _d = _f_actual.get(_area) or {}
                _p = _d.get("pasos") or {}
                _fa = _d.get("facturacion") or {}
                if not (_p.get("total") or _fa.get("total")):
                    continue
                _rows.append((_lbl, _p, _fa))
            if _rows:
                _tp = sum(p.get("total", 0) for _, p, _ in _rows)
                _tv = sum(f.get("total", 0) for _, _, f in _rows)
                st.markdown(
                    '<div class="ig-kpis">'
                    + _kpi(_n(_tp) or "0", f"Pasos en {_mes_corto}", "ambar",
                           _delta(_tp, _tot_atn))
                    + _kpi(_m(_tv) or "$ 0", f"Venta en {_mes_corto}", "rojo",
                           _delta(_tv, _venta_comp))
                    + '</div>', unsafe_allow_html=True)
                _hf = []
                for _lbl, _p, _fa in _rows:
                    _hf.append(
                        f'<tr><td>{_lbl}</td>'
                        + _celda(_p.get("cliente")) + _celda(_p.get("garantia"))
                        + _celda(_p.get("interno")) + _celda(_p.get("seguro"))
                        + f'<td><b>{_n(_p.get("total")) or "–"}</b></td>'
                        + _celda(_fa.get("mano_obra"), True)
                        + _celda(_fa.get("terceros_total"), True)
                        + _celda(_fa.get("insumos"), True)
                        + _celda(_fa.get("repuestos"), True)
                        + f'<td class="mon"><b>{_m(_fa.get("total")) or "–"}</b></td></tr>')
                st.markdown(
                    '<div class="ig-tblwrap"><table class="ig-tbl"><thead><tr>'
                    '<th>Área</th><th>Cliente</th><th>Garantía</th><th>Interno</th>'
                    '<th>Seguro</th><th>Total pasos</th><th>Mano de obra</th>'
                    '<th>Trabajos 3os</th><th>Insumos</th><th>Repuestos</th>'
                    '<th>Total venta</th></tr></thead><tbody>'
                    + "".join(_hf) + '</tbody></table></div>', unsafe_allow_html=True)
                st.caption(
                    "Un VIN = un paso, aunque tenga varios documentos en el mes "
                    "(se cuenta en la categoría de su documento de mayor prioridad). "
                    "Montos = suma de TOTAL NETA de *Producción Mensual Post-Venta*.")

        # ---- descarga ----
        _rows_x = []
        for _f in _hoja["filas"]:
            if _f["tipo"] == "seccion":
                continue
            _rows_x.append({"Ítem": _f["label"],
                            **{a or c: _f["valores"].get(c) for c, a in _pares}})
        _buf = io.BytesIO()
        with pd.ExcelWriter(_buf, engine="openpyxl") as _w:
            pd.DataFrame(_rows_x).to_excel(_w, index=False, sheet_name=_mes_sel[:28])
        st.download_button(
            "⬇️ Descargar esta sucursal a Excel", _buf.getvalue(),
            f"IMOP_Ford_{_suc_sel.replace(' ', '_')}_{_mes_sel}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True, key=f"{_k}ford_dl")

    # =========================================================
    #   AG  — NUEVO REPORTE DE GESTION, sucursal por sucursal
    # =========================================================
    elif _marca_n == "AG":
        _ag = _ig.get("ag", {}) or {}
        if not _ag:
            st.markdown('<div class="ig-empty">No se encontraron los archivos '
                        '<i>NUEVO REPORTE DE GESTIÓN</i> en la carpeta '
                        '<i>Informe de gestión</i>.</div>', unsafe_allow_html=True)
            if pane is None:
                st.stop()
            return

        _sucs = sorted(_ag, key=lambda s: _ag[s]["sucursal"])
        _labels = {s: _ag[s]["sucursal"] for s in _sucs}
        _NOMBRES = {"HYU": "🚙 Hyundai", "FOR": "🚐 Fortaleza",
                    "D&P": "🎨 Desabolladura y Pintura"}

        _c1, _c2 = st.columns([3, 4])
        with _c1:
            _suc_key = st.selectbox("🏢 Sucursal", _sucs,
                                    format_func=lambda s: _labels[s],
                                    key=f"{_k}ag_suc")
        _hojas = _ag[_suc_key]["hojas"]
        _disp = [h for h in ["HYU", "FOR", "D&P"] if h in _hojas]
        with _c2:
            _hoja_sel = st.radio("Hoja del reporte", _disp,
                                 format_func=lambda h: _NOMBRES.get(h, h),
                                 horizontal=True, key=f"{_k}ag_hoja")
        if _hoja_sel not in _hojas:      # cambio de sucursal con otra hoja activa
            _hoja_sel = _disp[0]
        _h = _hojas[_hoja_sel]

        # ultimo mes con datos del historico
        _idx_ult = None
        for _f in _h["filas"]:
            for _i in range(11, -1, -1):
                if _f["valores"][_i]:
                    _idx_ult = max(_idx_ult if _idx_ult is not None else 0, _i)
                    break
        _ult = MESES_CORTOS_IG[_idx_ult] if _idx_ult is not None else ""

        st.markdown(
            f'<div class="ig-chips">'
            f'<span class="ig-chip">🏢 {_labels[_suc_key]}</span>'
            f'<span class="ig-chip">{_NOMBRES.get(_hoja_sel, _hoja_sel)}</span>'
            + (f'<span class="ig-chip mut">{_h.get("codigo", "")}</span>'
               if _h.get("codigo") else '')
            + (f'<span class="ig-chip ok">Histórico hasta {_ult}</span>' if _ult else '')
            + '</div>', unsafe_allow_html=True)

        def _fila_ag(label):
            for _f in _h["filas"]:
                if _norm_txt_ig(_f["label"]) == _norm_txt_ig(label):
                    return _f["valores"]
            return [None] * 12

        # ---- KPIs del ultimo mes cerrado ----
        if _idx_ult is not None:
            _prev = _idx_ult - 1 if _idx_ult > 0 else None

            def _kv(label):
                _v = _fila_ag(label)
                return (_v[_idx_ult] or 0, (_v[_prev] or 0) if _prev is not None else 0)

            if _hoja_sel == "D&P":
                _a, _ap = _kv("Cliente")
                _b, _bp = _kv("Compañía de seguros")
                _t, _tp = _kv("Total pasos pagados por")
                _mo, _mop = _kv("Total Venta")
                st.markdown(
                    '<div class="ig-kpis">'
                    + _kpi(_n(_t) or "0", f"Pasos en {_ult}", "", _delta(_t, _tp))
                    + _kpi(_n(_a) or "0", "Pagados por cliente", "verde")
                    + _kpi(_n(_b) or "0", "Pagados por compañía", "ambar")
                    + _kpi(_m(_mo) or "$ 0", "Venta mano de obra", "rojo",
                           _delta(_mo, _mop))
                    + '</div>', unsafe_allow_html=True)
            else:
                _s, _sp = _kv("Paso Cliente - Servicio")
                _g, _gp = _kv("Paso Cliente - Garantía")
                _i2, _ip = _kv("Paso Cliente - Interno")
                _t, _tp = _kv("Total paso")
                st.markdown(
                    '<div class="ig-kpis">'
                    + _kpi(_n(_t) or "0", f"Pasos de la marca en {_ult}", "",
                           _delta(_t, _tp))
                    + _kpi(_n(_s) or "0", "Servicio (cliente)", "verde", _delta(_s, _sp))
                    + _kpi(_n(_g) or "0", "Garantía", "ambar", _delta(_g, _gp))
                    + _kpi(_n(_i2) or "0", "Interno", "gris", _delta(_i2, _ip))
                    + '</div>', unsafe_allow_html=True)

        # ---- tabla historica 12 meses ----
        st.markdown('<div class="ig-sec">📊 Histórico del año</div>',
                    unsafe_allow_html=True)
        # Titulos de seccion reales del reporte AG. Una fila sin valores que NO
        # este aqui (ej. "Flotista", "Otros") es un dato en blanco, no un titulo:
        # se muestra como fila normal con guiones para no confundir.
        _SEC_AG = {"PASO VEHICULAR", "VENTA MANO DE OBRA $", "HORAS DEL MES",
                   "MATERIALES Y LUBRICANTES $", "VENTAS NETAS", "REPUESTOS",
                   "PASOS PAGADOS POR", "PASOS POR MARCAS", "VENTA DE REPUESTOS $",
                   "HORAS DE ASISTENCIA PROMEDIO TECNICOS"}

        _filas_html = []
        _sec_actual = ""
        for _f in _h["filas"]:
            if _f["tipo"] == "cabecera":
                _sec_actual = _f["label"]
                continue
            _es_titulo = (_f["tipo"] == "seccion"
                          and (_norm_txt_ig(_f["label"]) in _SEC_AG
                               or _f["label"].strip().endswith("$")))
            if _es_titulo:
                _sec_actual = _f["label"]
                _filas_html.append(
                    f'<tr class="sec"><td colspan="13">{_f["label"]}</td></tr>')
                continue
            _mon = _es_monto(_f["label"], _sec_actual)
            _tds = "".join(_celda(_v, _mon) for _v in _f["valores"])
            _cls = ' class="tot"' if _f["tipo"] == "total" else ""
            _filas_html.append(f'<tr{_cls}><td>{_f["label"]}</td>{_tds}</tr>')
        _ths = "".join(f"<th>{m}</th>" for m in MESES_CORTOS_IG)
        st.markdown(
            f'<div class="ig-tblwrap"><table class="ig-tbl">'
            f'<thead><tr><th>Concepto</th>{_ths}</tr></thead>'
            f'<tbody>{"".join(_filas_html)}</tbody></table></div>',
            unsafe_allow_html=True)

        # ---- avance del mes en curso ----
        _a_actual = (_actual.get("ag", {}) or {}).get(_suc_key, {})
        if _a_actual:
            st.markdown(
                f'<div class="ig-sec">🟡 Avance del mes en curso — '
                f'{_mes_largo or _periodo_curso}</div>', unsafe_allow_html=True)
            if _hoja_sel in ("HYU", "FOR"):
                _p = _a_actual.get(_hoja_sel, {}) or {}
                _ref = _fila_ag("Total paso")[_idx_ult] if _idx_ult is not None else 0
                st.markdown(
                    '<div class="ig-kpis">'
                    + _kpi(_n(_p.get("total")) or "0", f"Pasos en {_mes_corto}", "ambar",
                           _delta(_p.get("total"), _ref))
                    + _kpi(_n(_p.get("cliente")) or "0", "Servicio (cliente)", "verde")
                    + _kpi(_n(_p.get("garantia")) or "0", "Garantía", "gris")
                    + _kpi(_n(_p.get("interno")) or "0", "Interno", "gris")
                    + '</div>', unsafe_allow_html=True)
                _tt = _a_actual.get("total_taller")
                if _tt:
                    st.markdown(
                        f'<div class="ig-chips">'
                        f'<span class="ig-chip">Taller completo: {_n(_tt)} pasos</span>'
                        f'<span class="ig-chip mut">{_n(_a_actual.get("total_taller_ag", 0))} marcas AG '
                        f'+ {_n(_a_actual.get("total_taller_ford", 0))} Ford</span>'
                        f'<span class="ig-chip mut">Otras marcas: '
                        f'{_n(_tt - (_p.get("total") or 0))}</span></div>',
                        unsafe_allow_html=True)
                if _hoja_sel == "FOR":
                    _pm = _a_actual.get("FOR_por_marca", {}) or {}
                    if any(_pm.values()):
                        _tds = "".join(
                            f'<tr><td>{k.title()}</td>{_celda(v)}</tr>'
                            for k, v in _pm.items())
                        st.markdown(
                            '<div class="ig-tblwrap" style="max-width:340px">'
                            '<table class="ig-tbl"><thead><tr><th>Marca</th>'
                            f'<th>Pasos {_mes_corto}</th></tr></thead>'
                            f'<tbody>{_tds}</tbody></table></div>',
                            unsafe_allow_html=True)
            else:
                _p = _a_actual.get("D&P", {}) or {}
                st.markdown(
                    '<div class="ig-kpis">'
                    + _kpi(_n(_p.get("total")) or "0", f"Pasos en {_mes_corto}", "ambar")
                    + _kpi(_n(_p.get("cliente")) or "0", "Pagados por cliente", "verde")
                    + _kpi(_n(_p.get("seguro")) or "0", "Pagados por compañía", "gris")
                    + _kpi(_n(_p.get("interno")) or "0", "Interno", "gris")
                    + '</div>', unsafe_allow_html=True)
                _pm = _a_actual.get("D&P_por_marca", {}) or {}
                if any(_pm.values()):
                    _tds = "".join(f'<tr><td>{k.title()}</td>{_celda(v)}</tr>'
                                   for k, v in _pm.items())
                    st.markdown(
                        '<div class="ig-tblwrap" style="max-width:340px">'
                        '<table class="ig-tbl"><thead><tr><th>Marca</th>'
                        f'<th>Pasos {_mes_corto}</th></tr></thead>'
                        f'<tbody>{_tds}</tbody></table></div>',
                        unsafe_allow_html=True)
            st.caption(
                "Un VIN = un paso, aunque tenga varios documentos en el mes. "
                "La venta de mano de obra y repuestos de AG todavía no se puede "
                "calcular: falta el archivo de facturación de AG en *Alimentación*.")
        else:
            st.info("Sin datos de Alimentación para esta sucursal en el mes en curso.")

        # ---- descarga ----
        _rows_x = [{"Concepto": _f["label"],
                    **{MESES_CORTOS_IG[_i]: _f["valores"][_i] for _i in range(12)}}
                   for _f in _h["filas"] if _f["tipo"] not in ("cabecera", "seccion")]
        _buf = io.BytesIO()
        with pd.ExcelWriter(_buf, engine="openpyxl") as _w:
            pd.DataFrame(_rows_x).to_excel(
                _w, index=False, sheet_name=_hoja_sel.replace("&", "y")[:28])
        st.download_button(
            "⬇️ Descargar esta hoja a Excel", _buf.getvalue(),
            f"Reporte_AG_{_labels[_suc_key].replace(' ', '_')}_{_hoja_sel.replace('&', 'y')}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True, key=f"{_k}ag_dl")

    # =========================================================
    #   GM / OMODA - JAECOO  (pendientes)
    # =========================================================
    else:
        st.markdown(
            f'<div class="ig-empty"><b>{_marca_n}</b> todavía no está configurado.'
            '<br><br>Para habilitarlo necesito el formato del informe de gestión de '
            'esta marca (el archivo que envías a la automotora) y su archivo de '
            'alimentación en la carpeta <i>Alimentación</i>.</div>',
            unsafe_allow_html=True)

    # ---------- avance entre cargas ----------
    _snaps = _ig.get("snapshots", []) or []
    if len(_snaps) > 1:
        with st.expander(f"📈 Avance entre cargas ({len(_snaps)} corridas)"):
            _rows = []
            for _s in _snaps:
                _res = _s.get("resumen", {}) or {}
                _rows.append({
                    "Carga": _s.get("fecha", ""),
                    "Período": _s.get("periodo", ""),
                    "Pasos AG": sum(v.get("HYU_total", 0) + v.get("FOR_total", 0)
                                    for v in (_res.get("ag", {}) or {}).values()),
                    "Pasos Ford": sum(a.get("pasos", 0)
                                      for suc in (_res.get("ford", {}) or {}).values()
                                      for a in suc.values()),
                    "Venta Ford $": sum(a.get("venta", 0)
                                        for suc in (_res.get("ford", {}) or {}).values()
                                        for a in suc.values()),
                })
            st.dataframe(pd.DataFrame(_rows), hide_index=True,
                         use_container_width=True)
            st.caption("Cada fila es una corrida del BAT. Sirve para ver cuánto "
                       "avanzó el mes entre una semana y la siguiente.")

    if pane is None:
        with st.sidebar:
            st.markdown(f"**Usuario:** `{usuario_activo}`")
            if st.button("🔄 Actualizar datos", use_container_width=True,
                         key=f"{_k}refresh"):
                st.cache_data.clear()
                st.rerun()
            st.divider()
            if st.button("🚪 Cerrar sesión", use_container_width=True,
                         key=f"{_k}logout"):
                st.session_state.authenticated = False
                st.session_state.user_email = ""
                st.rerun()
        st.stop()


if st.session_state.get("app_mode") == "informes_gestion":
    _render_informes_gestion()


# ============================================================
#   CUENTA FICHA — modo independiente
#   Saldo disponible del cliente en su cuenta corriente (Informe Ficha
#   Cuenta del ERP) + su historial completo de OT con todos los documentos
#   posteriores. El puente entre ambos mundos es Patente-Cliente.xlsx
#   (patente <-> RUT), igual que lo pidio Cristobal. Los datos los genera
#   consolidar_OTs.py (PASO 13) y viajan comprimidos en cuenta_ficha.json.
#   Acceso restringido por usuario (puede_cuenta_ficha). 31/07/2026.
# ============================================================
def _render_cuenta_ficha(pane=None):

    if not _puede_usar_cuenta_ficha(usuario_activo):
        if pane is None:
            st.session_state.pop("app_mode", None)
            st.error("🔒 No tienes autorización para ver Cuenta Ficha. "
                     "Solicítala al administrador (cjerez@curifor.com).")
            if st.button("← Volver al inicio", key="cf_sin_acceso_volver"):
                st.rerun()
            st.stop()
        else:
            st.error("🔒 No tienes autorización para ver Cuenta Ficha.")
            return

    _cf_clientes, _cf_resumen, _cf_fecha_act = _cargar_cuenta_ficha()
    _cf_revisados = _cargar_cf_revisados()
    # Comentarios y notas escritos desde la App (en vivo, no van en el JSON)
    _cf_coments = _cf_comentarios_por_ot()
    _cf_gestion = _cf_gestion_por_ot()

    def _cf_notas_de(o):
        """Comentarios + gestión de una OT. Devuelve (comentarios, gestion, total)."""
        _c = _cf_coments.get(o["ot"], [])
        _g = _cf_gestion.get(o["ot"], {})
        _n_g = sum(1 for _k in ("CATEGORIA", "OBSERVACION OT", "NOTAS",
                                "AVANCE - GESTIÓN") if _g.get(_k))
        return _c, _g, len(_c) + _n_g

    def _cf_fecha_key(f):
        """'31/07/2026 10:05' -> '2026-07-31 10:05' para poder ordenar."""
        _s = str(f or "").strip()
        if len(_s) >= 10 and _s[2] == "/" and _s[5] == "/":
            return f"{_s[6:10]}-{_s[3:5]}-{_s[0:2]}{_s[10:]}"
        return _s

    # ---------- helpers de formato ----------
    def _cf_money(v):
        try:
            n = int(round(float(v)))
        except Exception:
            return "$0"
        signo = "-" if n < 0 else ""
        return signo + "$" + f"{abs(n):,}".replace(",", ".")

    def _cf_money_corto(v):
        try:
            n = float(v)
        except Exception:
            return "$0"
        if abs(n) >= 1_000_000:
            return f"${n/1_000_000:,.1f}M".replace(",", ".")
        if abs(n) >= 1_000:
            return f"${n/1_000:,.0f}K".replace(",", ".")
        return _cf_money(n)

    def _cf_esc(t):
        return (str(t or "").replace("&", "&amp;").replace("<", "&lt;")
                .replace(">", "&gt;").replace('"', "&quot;"))

    _CF_EST_COLOR = {
        "Pendiente":  ("#c0392b", "#fdecea", "🔴"),
        "Anulado":    ("#7f8c8d", "#f0f2f4", "⚫"),
        "Cerrado":    ("#1e8449", "#eafaf1", "🟢"),
        "Finalizado": ("#1565c0", "#e8f1fb", "🔵"),
    }

    # ---------- estilos del modulo ----------
    st.markdown("""
    <style>
      .cf-kpis{display:flex;gap:14px;flex-wrap:wrap;margin:6px 0 18px;}
      .cf-kpi{flex:1;min-width:165px;border-radius:14px;padding:16px 18px;
              background:linear-gradient(135deg,#0d2f5a 0%,#1a4f8a 100%);color:#fff;
              box-shadow:0 6px 18px rgba(13,47,90,.22);}
      .cf-kpi.verde{background:linear-gradient(135deg,#0b6b3a 0%,#18a05a 100%);
                    box-shadow:0 6px 18px rgba(11,107,58,.22);}
      .cf-kpi.ambar{background:linear-gradient(135deg,#9a6100 0%,#d99100 100%);
                    box-shadow:0 6px 18px rgba(154,97,0,.22);}
      .cf-kpi.gris{background:linear-gradient(135deg,#3d4b5c 0%,#5b6d80 100%);
                   box-shadow:0 6px 18px rgba(61,75,92,.22);}
      .cf-kpi .v{font-size:1.75rem;font-weight:800;line-height:1.15;letter-spacing:-.5px;}
      .cf-kpi .l{font-size:.74rem;opacity:.88;text-transform:uppercase;
                 letter-spacing:.6px;margin-top:4px;}
      .cf-card{border:1px solid rgba(128,128,128,.28);border-left:6px solid #1a4f8a;
               border-radius:14px;padding:14px 16px;background:var(--secondary-background-color);
               min-height:158px;}
      .cf-card.rev{border-left-color:#18a05a;}
      .cf-card.sinsaldo{border-left-color:#95a5a6;opacity:.9;}
      .cf-card.deudor{border-left-color:#c0392b;}
      .cf-card .rut{font-size:.78rem;color:#7a8794;font-weight:700;letter-spacing:.4px;}
      .cf-card .nom{font-size:1rem;font-weight:700;color:var(--text-color);
                    line-height:1.25;margin:2px 0 8px;min-height:2.4em;}
      .cf-card .saldo{font-size:1.5rem;font-weight:800;color:#0b6b3a;letter-spacing:-.5px;}
      .cf-card .saldo.neg{color:#c0392b;}
      .cf-card .saldo.cero{color:#7a8794;}
      .cf-chips{display:flex;gap:6px;flex-wrap:wrap;margin-top:8px;}
      .cf-chip{font-size:.7rem;padding:3px 9px;border-radius:20px;background:rgba(26,79,138,.12);
               color:#1a4f8a;font-weight:600;white-space:nowrap;}
      .cf-chip.ok{background:rgba(24,160,90,.14);color:#0b6b3a;}
      .cf-chip.warn{background:rgba(192,57,43,.12);color:#c0392b;}
      .cf-chip.mut{background:rgba(128,128,128,.14);color:#6b7885;}
      .cf-hero{border-radius:18px;padding:22px 26px;margin-bottom:16px;color:#fff;
               background:linear-gradient(135deg,#0d2f5a 0%,#1a4f8a 55%,#2274c4 100%);
               box-shadow:0 10px 26px rgba(13,47,90,.25);}
      .cf-hero h2{margin:0;font-size:1.55rem;font-weight:800;letter-spacing:-.4px;color:#fff;}
      .cf-hero .rut{font-size:.9rem;opacity:.85;letter-spacing:1px;font-weight:600;}
      .cf-hero .saldo{font-size:2.5rem;font-weight:800;letter-spacing:-1px;line-height:1.1;}
      .cf-hero .saldo small{font-size:.8rem;font-weight:600;opacity:.8;display:block;
                            text-transform:uppercase;letter-spacing:.8px;}
      .cf-hero .cf-chip{background:rgba(255,255,255,.18);color:#fff;}
      .cf-sucbox{border:1px solid rgba(128,128,128,.25);border-radius:12px;padding:12px 14px;
                 background:var(--secondary-background-color);margin-bottom:10px;}
      .cf-sucbox .s{font-weight:700;color:var(--text-color);font-size:.92rem;}
      .cf-sucbox .m{font-size:1.25rem;font-weight:800;color:#0b6b3a;}
      .cf-otline{display:flex;gap:10px;align-items:center;flex-wrap:wrap;}
      .cf-badge{font-size:.72rem;font-weight:700;padding:3px 10px;border-radius:8px;}
      .cf-empty{border:1px dashed rgba(128,128,128,.4);border-radius:12px;padding:26px;
                text-align:center;color:#7a8794;}
    </style>
    """, unsafe_allow_html=True)

    # ---------- sidebar ----------
    with st.sidebar:
        if pane is None:
            st.markdown(
                f'<img src="{LOGO_DATA_URI}" style="max-width:180px;margin-bottom:0.4rem;"/>',
                unsafe_allow_html=True,
            )
            st.markdown("")
            if st.button("← Volver al inicio", use_container_width=True, key="cf_volver"):
                for _k in ["app_mode", "cf_sel", "cf_pag", "cf_pag_val"]:
                    st.session_state.pop(_k, None)
                st.rerun()
            st.divider()

        st.markdown("### 💳 Cuenta Ficha")
        st.caption(
            "Saldo disponible del cliente en su cuenta corriente y todo su "
            "historial de OT con los documentos de cada una. Las OT pendientes "
            "van siempre destacadas."
        )

        st.markdown("### Filtros")
        _cf_busq = st.text_input("Buscar (RUT, nombre o patente)", "", key="cf_f_busq")

        _cf_sucs_all = sorted({s["suc"] for c in _cf_clientes for s in c.get("sucursales", [])})
        _cf_f_suc = st.multiselect("Sucursal del saldo", _cf_sucs_all,
                                   placeholder="Todas", key="cf_f_suc")
        _cf_f_estado = st.radio(
            "Estado de la cuenta",
            ["💰 Con saldo a favor", "Todos", "⚖️ Sin saldo (netea 0)", "🔻 Deudores"],
            index=0, key="cf_f_estado",
        )
        _cf_f_rev = st.radio(
            "Revisión",
            ["Todos", "⏳ Pendientes de revisar", "✅ Ya revisados"],
            index=0, key="cf_f_rev",
        )
        _cf_f_ot = st.checkbox("Solo clientes con OT pendientes", value=False, key="cf_f_ot")
        _cf_orden = st.selectbox(
            "Ordenar por",
            ["Saldo (mayor primero)", "Saldo (menor primero)", "Nombre (A-Z)",
             "OT pendientes", "N° de movimientos"],
            key="cf_f_orden",
        )

        st.divider()
        if st.button("🔄 Actualizar datos", use_container_width=True, key="cf_actualizar"):
            st.cache_data.clear()
            st.rerun()

        if pane is None:
            st.divider()
            st.markdown(f"**Usuario:** `{usuario_activo}`")
            if st.button("🚪 Cerrar sesión", use_container_width=True, key="cf_logout"):
                for _k in ["authenticated", "user_email", "app_mode"]:
                    st.session_state.pop(_k, None)
                st.rerun()

    # ---------- header ----------
    _cf_meses = (_cf_resumen or {}).get("meses_historial", 24)
    st.markdown(
        f'''<div class="curifor-header">
            <div class="logo-pill"><img src="{LOGO_DATA_URI}" /></div>
            <div class="curifor-header-text">
                <h2>Cuenta Ficha</h2>
                <p>Saldos de cliente + historial de OT (últimos {_cf_meses} meses)
                   {(" · Actualizado " + _cf_fecha_act) if _cf_fecha_act else ""}</p>
                <span class="dev-credit">Curifor S.A</span>
            </div>
        </div>''',
        unsafe_allow_html=True,
    )

    if not _cf_clientes:
        st.warning(
            "⚙️ Aún no hay datos de Cuenta Ficha. El administrador debe correr "
            "`Ejecutar_Consolidacion.bat` con el archivo **INFORME FICHA CUENTA "
            "(.xls)** y **Patente-Cliente.xlsx** en la carpeta PBI — el PASO 13 "
            "genera y sube `cuenta_ficha.json`."
        )
        if pane is None:
            st.stop()
        return

    # ============================================================
    #   FICHA DE UN CLIENTE (detalle)
    # ============================================================
    _cf_sel = st.session_state.get("cf_sel")
    _cf_map = {c["rut"]: c for c in _cf_clientes}

    if _cf_sel and _cf_sel in _cf_map:
        cli = _cf_map[_cf_sel]
        _rev = _cf_revisados.get(cli["rut"])

        _cn1, _cn2 = st.columns([1, 4])
        with _cn1:
            if st.button("← Volver al listado", use_container_width=True, key="cf_back"):
                st.session_state.pop("cf_sel", None)
                st.rerun()

        # Un cliente sin "tiene_saldo" pero con total > 0 es DEUDOR (debe plata):
        # se muestra en rojo y en negativo para que no se confunda con saldo a favor.
        _es_deudor   = (not cli["tiene_saldo"]) and cli["saldo"] > 0
        _saldo_hero  = -cli["saldo"] if _es_deudor else (cli["saldo"] if cli["tiene_saldo"] else 0)
        _lbl_hero    = ("Saldo disponible" if cli["tiene_saldo"] and cli["saldo"] > 0
                        else ("Saldo deudor" if _es_deudor else "Sin saldo a favor"))
        _chips_hero = "".join(
            f'<span class="cf-chip">{_cf_esc(s["suc"])} · {_cf_money(s["monto"])}</span>'
            for s in cli.get("sucursales", [])[:6]
        )
        _rev_hero = (
            f'<span class="cf-chip">✅ Revisado por {_cf_esc(_rev.get("usuario",""))} · {_cf_esc(_rev.get("fecha",""))}</span>'
            if _rev else '<span class="cf-chip">⏳ Sin revisar</span>'
        )
        st.markdown(f"""
        <div class="cf-hero">
          <div style="display:flex;justify-content:space-between;gap:24px;flex-wrap:wrap;align-items:flex-start;">
            <div style="flex:2;min-width:260px;">
              <div class="rut">{_cf_esc(cli['rut'])}</div>
              <h2>{_cf_esc(cli['nombre'])}</h2>
              <div class="cf-chips" style="margin-top:10px;">
                {_rev_hero}
                <span class="cf-chip">🔧 {cli['n_ot']} OT</span>
                <span class="cf-chip">🔴 {cli['n_ot_pend']} pendiente(s)</span>
                <span class="cf-chip">🚗 {len(cli.get('patentes', []))} patente(s)</span>
                <span class="cf-chip">📄 {cli['n_mov']} movimiento(s)</span>
              </div>
            </div>
            <div style="flex:1;min-width:220px;text-align:right;">
              <div class="saldo"><small>{_lbl_hero}</small>{_cf_money(_saldo_hero)}</div>
              <div class="cf-chips" style="justify-content:flex-end;margin-top:10px;">{_chips_hero}</div>
            </div>
          </div>
        </div>
        """, unsafe_allow_html=True)

        _rv1, _rv2, _rv3 = st.columns([1.4, 1.4, 3])
        with _rv1:
            if _rev:
                if st.button("↩️ Quitar revisado", use_container_width=True, key="cf_unrev_det"):
                    if _guardar_cf_revisado(cli["rut"], usuario_activo, marcar=False):
                        _cargar_cf_revisados.clear()
                        st.rerun()
                    else:
                        st.error("No se pudo guardar. Reintenta.")
            else:
                if st.button("✅ Marcar como revisado", type="primary",
                             use_container_width=True, key="cf_rev_det"):
                    if _guardar_cf_revisado(cli["rut"], usuario_activo, marcar=True,
                                            nota=st.session_state.get("cf_nota_det", "")):
                        _cargar_cf_revisados.clear()
                        st.rerun()
                    else:
                        st.error("No se pudo guardar. Reintenta.")
        with _rv2:
            st.text_input("Nota (opcional)", "", key="cf_nota_det",
                          label_visibility="collapsed", placeholder="Nota de la revisión…")
        with _rv3:
            if _rev and _rev.get("nota"):
                st.caption(f"📝 {_rev['nota']}")

        _ots_cli = cli.get("ots", [])

        # Comentarios y notas de gestión de TODAS las OT de este cliente
        _notas_cli = []
        for _o in _ots_cli:
            _cms, _ges, _tot_n = _cf_notas_de(_o)
            for _cm in _cms:
                _notas_cli.append({"tipo": "comentario", "ot": _o, **_cm})
            if _ges:
                _notas_cli.append({"tipo": "gestion", "ot": _o, "g": _ges,
                                   "fecha": _ges.get("ULTIMA_EDICION", "")})
        _n_notas = len(_notas_cli)

        _tab_mov, _tab_ot, _tab_com, _tab_veh = st.tabs(
            [f"💰 Saldo y movimientos ({cli['n_mov']})",
             f"🔧 Historial de OT ({cli['n_ot']})",
             f"💬 Comentarios y notas ({_n_notas})",
             f"🚗 Vehículos ({len(cli.get('patentes', []))})"]
        )

        # ---- Movimientos ----
        with _tab_mov:
            if not cli.get("movimientos"):
                st.markdown('<div class="cf-empty">Este cliente no tiene movimientos en el informe.</div>',
                            unsafe_allow_html=True)
            else:
                _sucs = cli.get("sucursales", [])
                _cols_s = st.columns(min(4, max(1, len(_sucs))))
                for _i, _s in enumerate(_sucs):
                    with _cols_s[_i % len(_cols_s)]:
                        st.markdown(
                            f'<div class="cf-sucbox"><div class="s">🏢 {_cf_esc(_s["suc"])}</div>'
                            f'<div class="m">{_cf_money(_s["monto"])}</div>'
                            f'<div style="font-size:.75rem;color:#7a8794;">{_s["n"]} movimiento(s)</div></div>',
                            unsafe_allow_html=True)

                _df_mov = pd.DataFrame([{
                    "Sucursal":   m.get("local", ""),
                    "Documento":  m.get("documento", ""),
                    "N° Ref.":    m.get("nro", ""),
                    "Fecha":      m.get("fecha", ""),
                    "Monto":      int(m.get("saldo") or 0),
                    "Glosa / Detalle": m.get("glosa", ""),
                } for m in cli["movimientos"]])
                st.dataframe(
                    _df_mov, hide_index=True, use_container_width=True,
                    column_config={
                        "Monto": st.column_config.NumberColumn(format="$%d", width="small"),
                        "Glosa / Detalle": st.column_config.TextColumn(width="large"),
                    },
                )

        # ---- Historial de OT ----
        with _tab_ot:
            if not _ots_cli:
                st.markdown(
                    '<div class="cf-empty">Este cliente no tiene OT en los últimos '
                    f'{_cf_meses} meses.<br><span style="font-size:.85rem;">Puede que sus '
                    'vehículos no estén asociados a su RUT en <b>Patente-Cliente.xlsx</b>.</span></div>',
                    unsafe_allow_html=True)
            else:
                _pend = [o for o in _ots_cli if o["est"] == "Pendiente"]
                if _pend:
                    st.markdown(
                        f'<div style="background:#fdecea;border-left:5px solid #c0392b;'
                        f'border-radius:10px;padding:10px 14px;margin-bottom:10px;">'
                        f'<b style="color:#c0392b;">🔴 {len(_pend)} OT pendiente(s)</b> '
                        f'<span style="color:#7a4a45;font-size:.85rem;">— revisar antes de '
                        f'disponer del saldo</span></div>', unsafe_allow_html=True)

                _fo1, _fo2, _fo3, _fo4 = st.columns([1.3, 1.3, 1, 1.6])
                with _fo1:
                    _f_est = st.multiselect(
                        "Estado", sorted({o["est"] for o in _ots_cli}),
                        placeholder="Todos", key="cf_ot_est")
                with _fo2:
                    _f_suc_ot = st.multiselect(
                        "Sucursal", sorted({o["suc"] for o in _ots_cli if o["suc"]}),
                        placeholder="Todas", key="cf_ot_suc")
                with _fo3:
                    _f_anio = st.multiselect(
                        "Año", sorted({(o.get("fec_ord") or "")[:4] for o in _ots_cli if o.get("fec_ord")},
                                      reverse=True),
                        placeholder="Todos", key="cf_ot_anio")
                with _fo4:
                    _f_busq_ot = st.text_input(
                        "Buscar en el historial", "", key="cf_ot_busq",
                        placeholder="N° OT, patente, documento, glosa…")

                def _cf_ot_match(o):
                    if _f_est and o["est"] not in _f_est:
                        return False
                    if _f_suc_ot and o["suc"] not in _f_suc_ot:
                        return False
                    if _f_anio and (o.get("fec_ord") or "")[:4] not in _f_anio:
                        return False
                    if _f_busq_ot:
                        q = _f_busq_ot.strip().lower()
                        blob = " ".join([
                            o.get("ot", ""), o.get("num", ""), o.get("pat", ""),
                            o.get("mod", ""), o.get("mar", ""), o.get("ase", ""),
                            o.get("glosa", ""), o.get("tv", ""), o.get("estd", ""),
                            " ".join(d["n"] for lst in o.get("docs", {}).values() for d in lst),
                        ]).lower()
                        if q not in blob:
                            return False
                    return True

                _ots_f = [o for o in _ots_cli if _cf_ot_match(o)]
                # Pendientes primero, luego por fecha descendente
                _ots_f.sort(key=lambda o: (0 if o["est"] == "Pendiente" else 1,
                                           "" if o["est"] == "Pendiente" else "",
                                           o.get("fec_ord") or ""),
                            reverse=False)
                _ots_f = ([o for o in _ots_f if o["est"] == "Pendiente"][::-1] +
                          sorted([o for o in _ots_f if o["est"] != "Pendiente"],
                                 key=lambda o: o.get("fec_ord") or "", reverse=True))

                st.caption(f"Mostrando {len(_ots_f):,} de {len(_ots_cli):,} OT")

                _POR_PAG_OT = 15
                _tot_pag_ot = max(1, (len(_ots_f) + _POR_PAG_OT - 1) // _POR_PAG_OT)
                _pag_ot = 1
                if _tot_pag_ot > 1:
                    _pag_ot = st.number_input(
                        f"Página (de {_tot_pag_ot})", min_value=1, max_value=_tot_pag_ot,
                        value=1, step=1, key="cf_ot_pag")
                _ini_ot = (int(_pag_ot) - 1) * _POR_PAG_OT

                for o in _ots_f[_ini_ot:_ini_ot + _POR_PAG_OT]:
                    _col, _bg, _emo = _CF_EST_COLOR.get(o["est"], ("#555", "#eee", "⚪"))
                    _ndocs = sum(len(v) for v in o.get("docs", {}).values())
                    _cms_ot, _ges_ot, _n_notas_ot = _cf_notas_de(o)
                    _titulo = (f"{_emo} OT {o['ot']} · {o.get('fec','')} · "
                               f"{o.get('pat','')} {o.get('mar','')} {o.get('mod','')} · "
                               f"{o['est']} · {_ndocs} doc."
                               + (f" · 💬 {_n_notas_ot}" if _n_notas_ot else ""))
                    with st.expander(_titulo, expanded=(o["est"] == "Pendiente" and len(_pend) <= 3)):
                        _d1, _d2, _d3, _d4 = st.columns(4)
                        _d1.markdown(f"**Sucursal**  \n{_cf_esc(o.get('suc','—'))}")
                        _d2.markdown(f"**Asesor**  \n{_cf_esc(o.get('ase','—')) or '—'}")
                        _d3.markdown(f"**Tipo venta**  \n{_cf_esc(o.get('tv','—')) or '—'}")
                        _d4.markdown(f"**Neto**  \n{_cf_money(o.get('neto') or 0)}")
                        _e1, _e2, _e3 = st.columns([1, 1, 2])
                        _e1.markdown(f"**Días apertura**  \n{o.get('dias','—') or '—'}")
                        _e2.markdown(f"**Estado detallado**  \n{_cf_esc(o.get('estd','')) or '—'}")
                        _e3.markdown(f"**Glosa**  \n{_cf_esc(o.get('glosa','')) or '—'}")
                        if o.get("cierre"):
                            st.caption(f"🏷️ Tipo de cierre: {o['cierre']}")

                        st.markdown("**📑 Documentos posteriores**")
                        if not o.get("docs"):
                            st.caption("Esta OT no tiene documentos asociados en el PBI.")
                        else:
                            _rows_doc = [{
                                "Documento": _tipo,
                                "N° / Folio": _d["n"],
                                "Fecha": _d.get("f", ""),
                            } for _tipo, _lst in o["docs"].items() for _d in _lst]
                            st.dataframe(pd.DataFrame(_rows_doc), hide_index=True,
                                         use_container_width=True)

                        # --- Gestión escrita desde la App (Detalle y Edición) ---
                        if _ges_ot:
                            st.markdown("**📝 Gestión de la OT** _(escrita en la App)_")
                            _gc1, _gc2 = st.columns(2)
                            _gc1.markdown(
                                f"**Categoría:** {_cf_esc(_ges_ot.get('CATEGORIA','')) or '—'}  \n"
                                f"**Observación OT:** {_cf_esc(_ges_ot.get('OBSERVACION OT','')) or '—'}")
                            _gc2.markdown(
                                f"**Notas:** {_cf_esc(_ges_ot.get('NOTAS','')) or '—'}  \n"
                                f"**Avance / Gestión:** {_cf_esc(_ges_ot.get('AVANCE - GESTIÓN','')) or '—'}")
                            if _ges_ot.get("ULTIMA_EDICION"):
                                st.caption(f"✍️ Última edición: {_ges_ot['ULTIMA_EDICION']}")

                        # --- Comentarios escritos desde la App ---
                        st.markdown(f"**💬 Comentarios ({len(_cms_ot)})**")
                        if not _cms_ot:
                            st.caption("Esta OT aún no tiene comentarios en la App.")
                        else:
                            for _cm in sorted(_cms_ot,
                                              key=lambda c: _cf_fecha_key(c.get("fecha")),
                                              reverse=True):
                                st.markdown(f"""
                                <div class="comentario-card">
                                  <div class="comentario-meta">🗓 {_cf_esc(_cm.get('fecha',''))}
                                       &nbsp;·&nbsp; 👤 {_cf_esc(_cm.get('autor',''))}</div>
                                  <div class="comentario-texto">{_cf_esc(_cm.get('texto',''))}</div>
                                </div>""", unsafe_allow_html=True)

                # Exportar la ficha completa a Excel
                _buf_cf = io.BytesIO()
                with pd.ExcelWriter(_buf_cf, engine="openpyxl") as _xl:
                    pd.DataFrame([{
                        "RUT": cli["rut"], "Cliente": cli["nombre"],
                        "Saldo disponible": cli["saldo"] if cli["tiene_saldo"] else 0,
                        "Sucursal(es)": " / ".join(s["suc"] for s in cli.get("sucursales", [])),
                        "N° movimientos": cli["n_mov"], "N° OT": cli["n_ot"],
                        "OT pendientes": cli["n_ot_pend"],
                    }]).to_excel(_xl, index=False, sheet_name="Resumen")
                    pd.DataFrame([{
                        "Sucursal": m.get("local", ""), "Documento": m.get("documento", ""),
                        "N° Ref.": m.get("nro", ""), "Fecha": m.get("fecha", ""),
                        "Monto": int(m.get("saldo") or 0), "Glosa": m.get("glosa", ""),
                    } for m in cli.get("movimientos", [])]).to_excel(
                        _xl, index=False, sheet_name="Movimientos")
                    pd.DataFrame([{
                        "N° OT": o["ot"], "Fecha": o.get("fec", ""), "Estado": o["est"],
                        "Sucursal": o.get("suc", ""), "Patente": o.get("pat", ""),
                        "Marca": o.get("mar", ""), "Modelo": o.get("mod", ""),
                        "Tipo venta": o.get("tv", ""), "Asesor": o.get("ase", ""),
                        "Neto": o.get("neto", ""), "Días": o.get("dias", ""),
                        "Documentos": " | ".join(
                            f"{t}: {', '.join(d['n'] for d in l)}"
                            for t, l in o.get("docs", {}).items()),
                        "Glosa": o.get("glosa", ""),
                    } for o in _ots_cli]).to_excel(_xl, index=False, sheet_name="Historial OT")
                st.download_button(
                    "⬇️ Descargar ficha completa (Excel)", _buf_cf.getvalue(),
                    f"CuentaFicha_{cli['rut']}_{datetime.now().strftime('%Y%m%d')}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True, key="cf_dl_ficha",
                )

        # ---- Comentarios y notas escritos desde la App ----
        with _tab_com:
            if not _notas_cli:
                st.markdown(
                    '<div class="cf-empty">Ninguna OT de este cliente tiene todavía '
                    'comentarios ni notas escritas desde la App.<br>'
                    '<span style="font-size:.85rem;">Se escriben en '
                    '<b>Control y Gestión Post Venta → 📄 Documentos y Comentarios</b> '
                    '(comentarios) y en <b>✏️ Detalle y Edición</b> (Categoría, '
                    'Observación OT, Notas y Avance/Gestión).</span></div>',
                    unsafe_allow_html=True)
            else:
                _n_cm = sum(1 for _n in _notas_cli if _n["tipo"] == "comentario")
                st.caption(
                    f"💬 {_n_cm} comentario(s) y 📝 {_n_notas - _n_cm} OT con notas de "
                    f"gestión. Se leen en vivo desde la App — lo que escribas en "
                    f"Documentos y Comentarios aparece acá de inmediato."
                )
                _solo_pend_nc = st.checkbox(
                    "Mostrar solo las de OT pendientes", value=False, key="cf_nc_pend")
                _lista_nc = [_n for _n in _notas_cli
                             if (not _solo_pend_nc) or _n["ot"]["est"] == "Pendiente"]
                _lista_nc.sort(key=lambda n: _cf_fecha_key(n.get("fecha")), reverse=True)

                for _n in _lista_nc[:120]:
                    _o = _n["ot"]
                    _emo_nc = _CF_EST_COLOR.get(_o["est"], ("", "", "⚪"))[2]
                    _cab = (f'<span style="font-size:.78rem;color:#7a8794;">'
                            f'{_emo_nc} OT <b>{_cf_esc(_o["ot"])}</b> · {_cf_esc(_o.get("fec",""))} · '
                            f'{_cf_esc(_o.get("pat",""))} · {_cf_esc(_o.get("suc",""))} · '
                            f'{_cf_esc(_o["est"])}</span>')
                    if _n["tipo"] == "comentario":
                        st.markdown(f"""
                        <div class="comentario-card">
                          {_cab}
                          <div class="comentario-meta" style="margin-top:4px;">🗓 {_cf_esc(_n.get('fecha',''))}
                               &nbsp;·&nbsp; 👤 {_cf_esc(_n.get('autor',''))}</div>
                          <div class="comentario-texto">{_cf_esc(_n.get('texto',''))}</div>
                        </div>""", unsafe_allow_html=True)
                    else:
                        _g = _n["g"]
                        _partes = []
                        for _et, _k in (("Categoría", "CATEGORIA"),
                                        ("Observación OT", "OBSERVACION OT"),
                                        ("Notas", "NOTAS"),
                                        ("Avance / Gestión", "AVANCE - GESTIÓN")):
                            if _g.get(_k):
                                _partes.append(f"<b>{_et}:</b> {_cf_esc(_g[_k])}")
                        st.markdown(f"""
                        <div class="comentario-card" style="border-left-color:#9a6100;">
                          {_cab}
                          <div class="comentario-meta" style="margin-top:4px;">📝 Notas de gestión
                               {(' &nbsp;·&nbsp; ✍️ ' + _cf_esc(_g.get('ULTIMA_EDICION',''))) if _g.get('ULTIMA_EDICION') else ''}</div>
                          <div class="comentario-texto">{'<br>'.join(_partes)}</div>
                        </div>""", unsafe_allow_html=True)

                if len(_lista_nc) > 120:
                    st.caption(f"…y {len(_lista_nc) - 120} más. Usa el historial de OT "
                               f"para llegar a una OT puntual.")

        # ---- Vehículos ----
        with _tab_veh:
            _pats = cli.get("patentes", [])
            if not _pats:
                st.markdown(
                    '<div class="cf-empty">Este RUT no tiene patentes asociadas en '
                    '<b>Patente-Cliente.xlsx</b>, así que no se le puede cruzar historial de OT.</div>',
                    unsafe_allow_html=True)
            else:
                _por_pat = {}
                for o in _ots_cli:
                    d = _por_pat.setdefault(o["pat"], {"n": 0, "pend": 0, "mar": "", "mod": "", "ult": ""})
                    d["n"] += 1
                    d["pend"] += 1 if o["est"] == "Pendiente" else 0
                    if not d["mar"]:
                        d["mar"], d["mod"] = o.get("mar", ""), o.get("mod", "")
                    if (o.get("fec_ord") or "") > d["ult"]:
                        d["ult"] = o.get("fec", "")
                st.dataframe(
                    pd.DataFrame([{
                        "Patente": p,
                        "Marca":   _por_pat.get(p, {}).get("mar", ""),
                        "Modelo":  _por_pat.get(p, {}).get("mod", ""),
                        "OT":      _por_pat.get(p, {}).get("n", 0),
                        "OT pendientes": _por_pat.get(p, {}).get("pend", 0),
                        "Última OT": _por_pat.get(p, {}).get("ult", ""),
                    } for p in _pats]).sort_values(["OT pendientes", "OT"], ascending=False),
                    hide_index=True, use_container_width=True,
                )

        if pane is None:
            st.stop()
        return

    # ============================================================
    #   LISTADO DE CLIENTES
    # ============================================================
    _cf_lista = list(_cf_clientes)

    if _cf_f_estado == "💰 Con saldo a favor":
        _cf_lista = [c for c in _cf_lista if c["tiene_saldo"] and c["saldo"] > 0]
    elif _cf_f_estado == "⚖️ Sin saldo (netea 0)":
        _cf_lista = [c for c in _cf_lista if not c["tiene_saldo"] and c["saldo"] == 0]
    elif _cf_f_estado == "🔻 Deudores":
        _cf_lista = [c for c in _cf_lista if not c["tiene_saldo"] and c["saldo"] > 0]

    if _cf_f_suc:
        _cf_lista = [c for c in _cf_lista
                     if any(s["suc"] in _cf_f_suc for s in c.get("sucursales", []))]
    if _cf_f_rev == "⏳ Pendientes de revisar":
        _cf_lista = [c for c in _cf_lista if c["rut"] not in _cf_revisados]
    elif _cf_f_rev == "✅ Ya revisados":
        _cf_lista = [c for c in _cf_lista if c["rut"] in _cf_revisados]
    if _cf_f_ot:
        _cf_lista = [c for c in _cf_lista if c["n_ot_pend"] > 0]
    if _cf_busq:
        _q = _cf_busq.strip().lower()
        _cf_lista = [c for c in _cf_lista
                     if _q in c["rut"].lower() or _q in c["nombre"].lower()
                     or any(_q in p.lower() for p in c.get("patentes", []))]

    if _cf_orden == "Saldo (mayor primero)":
        _cf_lista.sort(key=lambda c: -c["saldo"])
    elif _cf_orden == "Saldo (menor primero)":
        _cf_lista.sort(key=lambda c: c["saldo"])
    elif _cf_orden == "Nombre (A-Z)":
        _cf_lista.sort(key=lambda c: c["nombre"])
    elif _cf_orden == "OT pendientes":
        _cf_lista.sort(key=lambda c: (-c["n_ot_pend"], -c["saldo"]))
    else:
        _cf_lista.sort(key=lambda c: -c["n_mov"])

    _monto_f = sum(c["saldo"] for c in _cf_lista if c["tiene_saldo"])
    _n_rev_f = sum(1 for c in _cf_lista if c["rut"] in _cf_revisados)
    _n_pend_ot = sum(1 for c in _cf_lista if c["n_ot_pend"] > 0)
    st.markdown(f"""
    <div class="cf-kpis">
      <div class="cf-kpi verde"><div class="v">{_cf_money(_monto_f)}</div>
           <div class="l">Saldo disponible (filtrado)</div></div>
      <div class="cf-kpi"><div class="v">{len(_cf_lista):,}</div>
           <div class="l">Clientes en pantalla</div></div>
      <div class="cf-kpi ambar"><div class="v">{len(_cf_lista) - _n_rev_f:,}</div>
           <div class="l">Pendientes de revisar</div></div>
      <div class="cf-kpi gris"><div class="v">{_n_rev_f:,}</div>
           <div class="l">Ya revisados</div></div>
      <div class="cf-kpi"><div class="v">{_n_pend_ot:,}</div>
           <div class="l">Con OT pendientes</div></div>
    </div>
    """.replace(",", "."), unsafe_allow_html=True)

    if not _cf_lista:
        st.markdown('<div class="cf-empty">Ningún cliente calza con los filtros actuales.</div>',
                    unsafe_allow_html=True)
        if pane is None:
            st.stop()
        return

    _POR_PAG = 12
    _tot_pag = max(1, (len(_cf_lista) + _POR_PAG - 1) // _POR_PAG)

    # La página se guarda aparte del widget ("cf_pag_val"): Streamlit descarta el
    # estado de un widget que deja de dibujarse, y al entrar a la ficha de un
    # cliente el paginador no existe — por eso al volver caía siempre en la 1.
    # Se conserva el valor y se ajusta si los filtros dejaron menos páginas.
    _pag_guardada = st.session_state.get("cf_pag_val", 1)
    try:
        _pag_guardada = min(max(1, int(_pag_guardada)), _tot_pag)
    except (TypeError, ValueError):
        _pag_guardada = 1
    if "cf_pag" in st.session_state:
        try:
            st.session_state["cf_pag"] = min(max(1, int(st.session_state["cf_pag"])), _tot_pag)
        except (TypeError, ValueError):
            st.session_state["cf_pag"] = _pag_guardada

    _p1, _p2 = st.columns([1, 4])
    with _p1:
        _pag = st.number_input(f"Página (de {_tot_pag})", min_value=1, max_value=_tot_pag,
                               value=_pag_guardada, step=1, key="cf_pag")
    st.session_state["cf_pag_val"] = int(_pag)
    _ini = (int(_pag) - 1) * _POR_PAG
    _pagina = _cf_lista[_ini:_ini + _POR_PAG]
    with _p2:
        st.caption(f"Mostrando {_ini + 1}–{_ini + len(_pagina)} de {len(_cf_lista):,} cliente(s)")

    for _fila in range(0, len(_pagina), 3):
        _cols = st.columns(3)
        for _j, cli in enumerate(_pagina[_fila:_fila + 3]):
            with _cols[_j]:
                _rev = _cf_revisados.get(cli["rut"])
                _deudor = (not cli["tiene_saldo"]) and cli["saldo"] > 0
                _cls = "cf-card"
                if _rev:
                    _cls += " rev"
                elif _deudor:
                    _cls += " deudor"
                elif not cli["tiene_saldo"]:
                    _cls += " sinsaldo"
                _saldo_card = (cli["saldo"] if cli["tiene_saldo"]
                               else (-cli["saldo"] if _deudor else 0))
                _saldo_cls = ("" if cli["tiene_saldo"] and cli["saldo"] > 0
                              else ("neg" if _deudor else "cero"))
                _suc_txt = cli.get("suc_principal", "") or "Sin sucursal"
                _n_suc = len(cli.get("sucursales", []))
                _chips = f'<span class="cf-chip">🏢 {_cf_esc(_suc_txt)}</span>'
                if _n_suc > 1:
                    _chips += f'<span class="cf-chip mut">+{_n_suc-1} sucursal(es)</span>'
                if cli["n_ot_pend"]:
                    _chips += f'<span class="cf-chip warn">🔴 {cli["n_ot_pend"]} OT pendiente(s)</span>'
                _chips += f'<span class="cf-chip mut">🔧 {cli["n_ot"]} OT</span>'
                # Comentarios / notas escritos desde la App en las OT de este cliente
                _n_nc = sum(_cf_notas_de(_o)[2] for _o in cli.get("ots", []))
                if _n_nc:
                    _chips += f'<span class="cf-chip">💬 {_n_nc} nota(s)</span>'
                if _rev:
                    _chips += f'<span class="cf-chip ok">✅ {_cf_esc(_rev.get("usuario","").split("@")[0])}</span>'
                st.markdown(f"""
                <div class="{_cls}">
                  <div class="rut">{_cf_esc(cli['rut'])}</div>
                  <div class="nom">{_cf_esc(cli['nombre'][:58])}</div>
                  <div class="saldo {_saldo_cls}">{_cf_money(_saldo_card)}</div>
                  <div style="font-size:.72rem;color:#7a8794;">{cli['n_mov']} movimiento(s)</div>
                  <div class="cf-chips">{_chips}</div>
                </div>""", unsafe_allow_html=True)
                _b1, _b2 = st.columns([2, 1])
                with _b1:
                    if st.button("📂 Abrir ficha", key=f"cf_open_{cli['rut']}",
                                 use_container_width=True):
                        st.session_state["cf_sel"] = cli["rut"]
                        st.rerun()
                with _b2:
                    if _rev:
                        if st.button("↩️", key=f"cf_unrev_{cli['rut']}",
                                     use_container_width=True, help="Quitar 'revisado'"):
                            if _guardar_cf_revisado(cli["rut"], usuario_activo, marcar=False):
                                _cargar_cf_revisados.clear()
                                st.rerun()
                    else:
                        if st.button("✅", key=f"cf_rev_{cli['rut']}",
                                     use_container_width=True, help="Marcar como revisado"):
                            if _guardar_cf_revisado(cli["rut"], usuario_activo, marcar=True):
                                _cargar_cf_revisados.clear()
                                st.rerun()
                st.markdown("<div style='height:6px'></div>", unsafe_allow_html=True)

    st.divider()
    _df_exp = pd.DataFrame([{
        "RUT": c["rut"], "Cliente": c["nombre"],
        "Saldo disponible": c["saldo"] if c["tiene_saldo"] else 0,
        "Sucursal principal": c.get("suc_principal", ""),
        "Sucursales": " / ".join(s["suc"] for s in c.get("sucursales", [])),
        "N° movimientos": c["n_mov"], "N° OT": c["n_ot"], "OT pendientes": c["n_ot_pend"],
        "Patentes": " / ".join(c.get("patentes", [])[:20]),
        "Revisado por": _cf_revisados.get(c["rut"], {}).get("usuario", ""),
        "Fecha revisión": _cf_revisados.get(c["rut"], {}).get("fecha", ""),
    } for c in _cf_lista])
    _buf_lst = io.BytesIO()
    with pd.ExcelWriter(_buf_lst, engine="openpyxl") as _xl:
        _df_exp.to_excel(_xl, index=False, sheet_name="Cuenta Ficha")
    st.download_button(
        "⬇️ Descargar listado filtrado (Excel)", _buf_lst.getvalue(),
        f"CuentaFicha_Listado_{datetime.now().strftime('%Y%m%d')}.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        use_container_width=True, key="cf_dl_listado",
    )

    if pane is None:
        st.stop()


if st.session_state.get("app_mode") == "cuenta_ficha":
    _render_cuenta_ficha()


# ============================================================
#   LOANERS  — modo independiente
#   Flota de vehiculos de cortesia (loaners) que se prestan al
#   cliente mientras su unidad esta en el taller. Reemplaza al
#   Excel "LOANERS 2.xlsx" que se llevaba a mano: misma
#   distribucion de columnas y misma logica — la columna
#   "Fecha Solicitud" dice literalmente "Disponible" cuando la
#   unidad esta libre, y trae la fecha del prestamo cuando esta
#   asignada a un cliente.
#
#   La identidad de cada unidad (VIN / Modelo / Patente Loaner,
#   columnas B, C y D) vive fija en LOANERS_FLOTA y NO es
#   editable — decision de Cristobal (06/08/2026): los usuarios
#   solo editan lo que cambia dia a dia, que es la Sucursal
#   (columna A) y las columnas E a K. Eso editable se guarda en
#   loaners.json en GitHub, compartido por todo el equipo (mismo
#   mecanismo que las marcas "Revisado" de Cuenta Ficha).
#
#   Se ubica ANTES de cargar_datos() porque no depende del JSON
#   de OTs: abre al instante aunque los datos fallen.
#   Acceso restringido (puede_loaners) — 06/08/2026.
# ============================================================

# Flota fija. Cada unidad se identifica por su VIN (es la clave con la que se
# guarda lo editado en loaners.json, por eso no puede cambiar). Los valores de
# sucursal / fecha / cliente que van aca son los del Excel original y actuan
# solo como estado INICIAL: apenas alguien guarda un cambio en la App, manda
# lo que este en loaners.json. Para sumar una unidad nueva a la flota basta
# con agregarle una entrada a esta lista.
LOANERS_FLOTA = [
    {"vin": "3FTTW8JA9SRA52067", "modelo": "MAVERICK XLT", "patente": "VJGR22",
     "sucursal": "Linderos (RANCAGUA)", "fecha_solicitud": "2026-07-20", "kms_salida": 8669,
     "vin_cliente": "3FMCR9E92RRE77861", "nombre_cliente": "ESTEBAN VALENTIN",
     "modelo_cliente": "BRONCO", "fecha_ot": "2026-07-06", "caso_sf": "1620439"},
    {"vin": "8AFBR01C9SJ472257", "modelo": "RANGER", "patente": "VJDW16",
     "sucursal": "Linderos (RANCAGUA)", "fecha_solicitud": "2026-07-30", "kms_salida": 21014,
     "vin_cliente": "3FTTW8M39RRB81966", "nombre_cliente": "INGENIERIA Y CONSTRUCCION TORSAN SPA",
     "modelo_cliente": "MAVERICK", "fecha_ot": "2026-07-30", "caso_sf": "1638862"},
    {"vin": "8AFBR00C2RJ409139", "modelo": "RANGER", "patente": "VJDW14",
     "sucursal": "Linderos (RANCAGUA)", "fecha_solicitud": "2026-06-23", "kms_salida": 19832,
     "vin_cliente": "1FTFW5L52SFB46008", "nombre_cliente": "JORGE TAGLE",
     "modelo_cliente": "F150", "fecha_ot": "2026-06-10", "caso_sf": "1621501"},
    {"vin": "8AFBR00E0SJ476717", "modelo": "RANGER", "patente": "VJDW89",
     "sucursal": "Linderos (RANCAGUA)", "fecha_solicitud": "2026-07-15", "kms_salida": 29399,
     "vin_cliente": "8AFBR01K3SJ412241",
     "nombre_cliente": "SERVICIOS MÉDICOS JUAN PABLO FIGUEROA VERCELLINO E.I.R.L.",
     "modelo_cliente": "RANGER", "fecha_ot": "2026-05-12", "caso_sf": "1630450"},
    {"vin": "LJXCU2BB3SHF56680", "modelo": "TERRITORY TITANIUM", "patente": "VJFC18",
     "sucursal": "Talca 1 (Rancagua)", "fecha_solicitud": "", "kms_salida": 10127,
     "vin_cliente": "", "nombre_cliente": "", "modelo_cliente": "",
     "fecha_ot": "", "caso_sf": ""},
    {"vin": "3FTTW8S94RRB81297", "modelo": "MAVERICK LARIAT", "patente": "VJGV10",
     "sucursal": "Talca 2", "fecha_solicitud": "2026-07-27", "kms_salida": 12057,
     "vin_cliente": "8AFAR23Y4PJ334729",
     "nombre_cliente": "JOSE FRANCISCO JAVIER MARTINEZ VEILLON",
     "modelo_cliente": "RANGER", "fecha_ot": "2026-07-27", "caso_sf": "Sin N°"},
    {"vin": "8AFBR01C4SJ469086", "modelo": "RANGER", "patente": "VJGS25",
     "sucursal": "Talca 2", "fecha_solicitud": "", "kms_salida": 14598,
     "vin_cliente": "", "nombre_cliente": "", "modelo_cliente": "",
     "fecha_ot": "", "caso_sf": ""},
    {"vin": "3FTTW8M3XRRB81541", "modelo": "MAVERICK LARIAT", "patente": "VJGT90",
     "sucursal": "Curicó (LINDEROS)", "fecha_solicitud": "2026-08-03", "kms_salida": 10806,
     "vin_cliente": "1FTFW3L54SFC54179",
     "nombre_cliente": "ASESORIAS E INVERSIONES EMF LIMITADA",
     "modelo_cliente": "F150", "fecha_ot": "2026-08-03", "caso_sf": "Sin N°"},
    {"vin": "8AFBR00E5SJ420109", "modelo": "RANGER", "patente": "VJDX53",
     "sucursal": "Curicó", "fecha_solicitud": "", "kms_salida": 5243,
     "vin_cliente": "", "nombre_cliente": "", "modelo_cliente": "",
     "fecha_ot": "", "caso_sf": ""},
    {"vin": "3FTTW8H97RRB81530", "modelo": "MAVERICK XLT", "patente": "VJDX81",
     "sucursal": "Chillán", "fecha_solicitud": "", "kms_salida": 11239,
     "vin_cliente": "", "nombre_cliente": "", "modelo_cliente": "",
     "fecha_ot": "", "caso_sf": ""},
]

# Sucursales base del desplegable. Se les suman las que ya esten en uso en los
# datos y las que el usuario agregue a mano (se guardan en loaners.json), para
# poder escribir variantes como "Linderos (RANCAGUA)" cuando la unidad esta
# prestada a otra sucursal — mismo patron que "asesores_extra" del Planificador.
LOANERS_SUCURSALES_BASE = [
    "Linderos", "Rancagua", "Talca 1", "Talca 2", "Curicó", "Chillán",
    "Chillán Viejo", "Placilla", "Lo Blanco", "CD Repuestos",
]

# Campos que el usuario puede editar (todo lo que no es la identidad del vehiculo)
LOANERS_CAMPOS_EDIT = [
    "sucursal", "fecha_solicitud", "kms_salida", "vin_cliente",
    "nombre_cliente", "modelo_cliente", "fecha_ot", "caso_sf",
]
# Campos del cliente: se limpian solos al marcar la unidad como Disponible
# (misma logica del Excel, donde las filas disponibles van en blanco de G a K).
LOANERS_CAMPOS_CLIENTE = ["vin_cliente", "nombre_cliente", "modelo_cliente",
                          "fecha_ot", "caso_sf"]

LOANERS_DISPONIBLE = "Disponible"


def _loaners_a_date(v):
    """'2026-07-20' / date / Timestamp -> datetime.date. '' o 'Disponible' -> None."""
    if v is None:
        return None
    try:
        if pd.isna(v):          # cubre NaT / NaN que devuelve el data_editor
            return None
    except Exception:
        pass
    if isinstance(v, datetime):  # datetime y pd.Timestamp
        return v.date()
    if not isinstance(v, str) and hasattr(v, "isoformat"):   # datetime.date
        try:
            return datetime.strptime(str(v)[:10], "%Y-%m-%d").date()
        except Exception:
            return None
    s = str(v).strip()
    if not s or s.lower() == LOANERS_DISPONIBLE.lower():
        return None
    for _f in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y"):
        try:
            return datetime.strptime(s[:10], _f).date()
        except Exception:
            continue
    return None


def _loaners_a_iso(v):
    """Valor de fecha -> 'AAAA-MM-DD' o '' si no hay fecha valida."""
    d = _loaners_a_date(v)
    return d.strftime("%Y-%m-%d") if d else ""


def _loaners_fecha_txt(v):
    """Valor de fecha -> 'DD/MM/AAAA' para mostrar en pantalla."""
    d = _loaners_a_date(v)
    return d.strftime("%d/%m/%Y") if d else ""


def _loaners_disponible(fila):
    """Una unidad esta disponible si su Fecha Solicitud no es una fecha real
    (vacia o el literal 'Disponible') — misma regla que usa el Excel."""
    return _loaners_a_date(fila.get("fecha_solicitud")) is None


def _loaners_dias(fila):
    """Dias que lleva prestada la unidad (None si esta disponible)."""
    d = _loaners_a_date(fila.get("fecha_solicitud"))
    if not d:
        return None
    return (datetime.now(_TZ_CHILE).date() - d).days


@st.cache_data(ttl=20)
def _cargar_loaners_estado():
    """
    Lee loaners.json de GitHub. Devuelve ({vin: {campos editados}},
    [sucursales agregadas a mano]). Si el archivo aun no existe (primera vez),
    devuelve vacio y el modulo muestra los valores iniciales de LOANERS_FLOTA.
    """
    try:
        _, datos = _leer_json_github_raw(GITHUB_LOANERS)
        if isinstance(datos, dict):
            _l = datos.get("loaners", {})
            _s = datos.get("sucursales_extra", [])
            return (_l if isinstance(_l, dict) else {},
                    [str(x) for x in _s] if isinstance(_s, list) else [])
    except Exception:
        pass
    return {}, []


def _loaners_filas():
    """Flota fija + lo que este guardado en GitHub (lo guardado manda)."""
    _estado, _ = _cargar_loaners_estado()
    filas = []
    for base in LOANERS_FLOTA:
        fila = dict(base)
        fila["_editado_por"] = ""
        fila["_editado"] = ""
        guardado = _estado.get(base["vin"], {})
        if isinstance(guardado, dict):
            for _k in LOANERS_CAMPOS_EDIT:
                if _k in guardado:
                    fila[_k] = guardado[_k]
            fila["_editado_por"] = str(guardado.get("_editado_por", ""))
            fila["_editado"] = str(guardado.get("_editado", ""))
        filas.append(fila)
    return filas


def _guardar_loaners(cambios, usuario):
    """
    cambios = {vin: {campo: valor}}. Relee el archivo fresco antes de escribir
    para no pisar lo que otro usuario haya guardado mientras tanto (mismo
    criterio que _guardar_cf_revisado de Cuenta Ficha).
    """
    if not cambios:
        return True
    try:
        _, datos = _leer_json_github_raw(GITHUB_LOANERS)
        if not isinstance(datos, dict):
            datos = {}
        loaners = datos.get("loaners", {})
        if not isinstance(loaners, dict):
            loaners = {}
        for vin, campos in cambios.items():
            reg = loaners.get(vin, {})
            if not isinstance(reg, dict):
                reg = {}
            for _k, _v in campos.items():
                if _k in LOANERS_CAMPOS_EDIT:
                    reg[_k] = _v
            reg["_editado_por"] = usuario
            reg["_editado"] = ahora_chile()
            loaners[vin] = reg
        datos["loaners"] = loaners
        return _guardar_json_github_raw(
            GITHUB_LOANERS, datos,
            f"Loaners: {len(cambios)} unidad(es) actualizada(s) ({usuario})",
        )
    except Exception:
        return False


def _guardar_loaners_sucursal_extra(nombre, usuario):
    """Agrega una sucursal escrita a mano al desplegable (compartida)."""
    nombre = str(nombre or "").strip()
    if not nombre:
        return False
    try:
        _, datos = _leer_json_github_raw(GITHUB_LOANERS)
        if not isinstance(datos, dict):
            datos = {}
        extras = datos.get("sucursales_extra", [])
        if not isinstance(extras, list):
            extras = []
        if any(str(x).strip().lower() == nombre.lower() for x in extras):
            return True
        extras.append(nombre)
        datos["sucursales_extra"] = extras
        return _guardar_json_github_raw(
            GITHUB_LOANERS, datos, f"Loaners: sucursal '{nombre}' agregada ({usuario})")
    except Exception:
        return False


def generar_excel_loaners(filas):
    """
    Exporta el listado completo con el MISMO formato del Excel original
    (LOANERS 2.xlsx): las 11 columnas en el mismo orden, el mismo color por
    columna, el literal "Disponible" en Fecha Solicitud cuando la unidad esta
    libre y los mismos anchos de columna. Devuelve los bytes del .xlsx.
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    # Colores tomados del archivo original (temas de Office ya resueltos a RGB)
    GRIS_HDR = "D9D9D9"   # encabezado
    AZUL_A   = "4E95D9"   # Sucursal
    VERDE_B  = "C2F1C8"   # VIN Loaner
    CELES_C  = "C1E5F5"   # Modelo Loaner
    NARAN_D  = "F6C6AD"   # Patente Loaner
    CELES_I  = "83CBEB"   # Modelo Unidad Cliente
    GRIS_DIS = "D9D9D9"   # Fecha Solicitud cuando dice "Disponible"

    ENCABEZADOS = [
        "Sucursal", "VIN\xa0Loaner", "Modelo\xa0Loaner", "Patente\xa0Loaner",
        "Fecha Solicitud", "KMS Salida", "Vin Cliente", "Nombre Cliente",
        "Modelo Unidad Cliente", "Fecha OT", "N° caso Salesforce",
    ]
    ANCHOS = [23.453125, 24.0, 20.81640625, 11.81640625, 13.0, 11.453125,
              22.7265625, 38.36328125, 18.7265625, 11.7265625, 15.0]
    FILL_COL = {1: AZUL_A, 2: VERDE_B, 3: CELES_C, 4: NARAN_D, 9: CELES_I}

    wb = Workbook()
    ws = wb.active
    ws.title = "Hoja1"

    f_hdr = Font(name="Calibri", bold=True, size=11, color="000000")
    f_dat = Font(name="Calibri", size=11)
    borde = Border(*[Side(style="thin", color="BFBFBF")] * 4)

    for i, h in enumerate(ENCABEZADOS, start=1):
        c = ws.cell(row=1, column=i, value=h)
        c.font = f_hdr
        c.fill = PatternFill("solid", fgColor=GRIS_HDR)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = borde
        ws.column_dimensions[get_column_letter(i)].width = ANCHOS[i - 1]
    ws.row_dimensions[1].height = 30

    for r, fila in enumerate(filas, start=2):
        disponible = _loaners_disponible(fila)
        _f_sol = LOANERS_DISPONIBLE if disponible else _loaners_a_date(fila.get("fecha_solicitud"))
        _f_ot = _loaners_a_date(fila.get("fecha_ot"))
        try:
            _kms = int(float(fila.get("kms_salida") or 0))
        except Exception:
            _kms = None

        valores = [
            str(fila.get("sucursal", "") or ""),
            str(fila.get("vin", "") or ""),
            str(fila.get("modelo", "") or ""),
            str(fila.get("patente", "") or ""),
            _f_sol,
            _kms,
            "" if disponible else str(fila.get("vin_cliente", "") or ""),
            "" if disponible else str(fila.get("nombre_cliente", "") or ""),
            "" if disponible else str(fila.get("modelo_cliente", "") or ""),
            None if disponible else _f_ot,
            "" if disponible else str(fila.get("caso_sf", "") or ""),
        ]
        for i, v in enumerate(valores, start=1):
            c = ws.cell(row=r, column=i, value=(v if v != "" else None))
            c.font = f_dat
            c.border = borde
            if i in FILL_COL:
                c.fill = PatternFill("solid", fgColor=FILL_COL[i])
            if i == 5 and disponible:
                c.fill = PatternFill("solid", fgColor=GRIS_DIS)
                c.alignment = Alignment(horizontal="center")
            if i in (5, 10) and not isinstance(v, str) and v is not None:
                c.number_format = "DD/MM/YYYY"
                c.alignment = Alignment(horizontal="center")
            if i == 6:
                c.number_format = "#,##0"

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:K{max(1, len(filas) + 1)}"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


def _render_loaners(pane=None):

    if not _puede_usar_loaners(usuario_activo):
        if pane is None:
            st.session_state.pop("app_mode", None)
            st.error("🔒 No tienes autorización para ver el módulo Loaners. "
                     "Solicítala al administrador (cjerez@curifor.com).")
            if st.button("← Volver al inicio", key="ln_sin_acceso_volver"):
                st.rerun()
            st.stop()
        else:
            st.error("🔒 No tienes autorización para ver el módulo Loaners.")
            return

    _pfx = f"ln_{pane or 'std'}_"          # prefijo de keys (unico por panel)
    _filas = _loaners_filas()
    _, _sucs_extra = _cargar_loaners_estado()

    def _ln_esc(t):
        return (str(t or "").replace("&", "&amp;").replace("<", "&lt;")
                .replace(">", "&gt;").replace('"', "&quot;"))

    # Opciones del desplegable de Sucursal: base + las que ya estan en uso +
    # las agregadas a mano. Se ordena dejando primero las base.
    _sucs_uso = [str(f.get("sucursal", "")).strip() for f in _filas if str(f.get("sucursal", "")).strip()]
    _sucs_opts = []
    for _s in LOANERS_SUCURSALES_BASE + sorted(set(_sucs_uso + _sucs_extra)):
        if _s and _s not in _sucs_opts:
            _sucs_opts.append(_s)

    # ---------- estilos del modulo ----------
    st.markdown("""
    <style>
      .ln-hero{border-radius:18px;padding:22px 26px;margin-bottom:16px;color:#fff;
               background:linear-gradient(135deg,#0b3d2e 0%,#12694d 55%,#1aa06f 100%);
               box-shadow:0 10px 26px rgba(11,61,46,.25);}
      .ln-hero h2{margin:0;font-size:1.55rem;font-weight:800;letter-spacing:-.4px;color:#fff;}
      .ln-hero p{margin:6px 0 0;font-size:.9rem;opacity:.9;}
      .ln-kpis{display:flex;gap:14px;flex-wrap:wrap;margin:6px 0 18px;}
      .ln-kpi{flex:1;min-width:150px;border-radius:14px;padding:16px 18px;color:#fff;
              background:linear-gradient(135deg,#0d2f5a 0%,#1a4f8a 100%);
              box-shadow:0 6px 18px rgba(13,47,90,.22);}
      .ln-kpi.verde{background:linear-gradient(135deg,#0b6b3a 0%,#18a05a 100%);
                    box-shadow:0 6px 18px rgba(11,107,58,.22);}
      .ln-kpi.rojo{background:linear-gradient(135deg,#8c1d13 0%,#c0392b 100%);
                   box-shadow:0 6px 18px rgba(140,29,19,.22);}
      .ln-kpi.ambar{background:linear-gradient(135deg,#9a6100 0%,#d99100 100%);
                    box-shadow:0 6px 18px rgba(154,97,0,.22);}
      .ln-kpi.gris{background:linear-gradient(135deg,#3d4b5c 0%,#5b6d80 100%);
                   box-shadow:0 6px 18px rgba(61,75,92,.22);}
      .ln-kpi .v{font-size:1.75rem;font-weight:800;line-height:1.15;letter-spacing:-.5px;}
      .ln-kpi .l{font-size:.74rem;opacity:.88;text-transform:uppercase;
                 letter-spacing:.6px;margin-top:4px;}
      .ln-card{border:1px solid rgba(128,128,128,.28);border-left:6px solid #18a05a;
               border-radius:14px;padding:14px 16px;margin-bottom:12px;
               background:var(--secondary-background-color);}
      .ln-card.asig{border-left-color:#c0392b;}
      .ln-card .pat{font-size:1.15rem;font-weight:800;color:var(--text-color);
                    letter-spacing:.5px;}
      .ln-card .mod{font-size:.85rem;color:#7a8794;font-weight:600;margin-bottom:6px;}
      .ln-card .vin{font-size:.72rem;color:#8b97a4;letter-spacing:.4px;}
      .ln-card .cli{font-size:.9rem;color:var(--text-color);margin-top:8px;line-height:1.45;}
      .ln-card .cli b{color:#1a4f8a;}
      .ln-chips{display:flex;gap:6px;flex-wrap:wrap;margin-top:8px;}
      .ln-chip{font-size:.7rem;padding:3px 9px;border-radius:20px;font-weight:600;
               background:rgba(26,79,138,.12);color:#1a4f8a;white-space:nowrap;}
      .ln-chip.ok{background:rgba(24,160,90,.16);color:#0b6b3a;}
      .ln-chip.warn{background:rgba(192,57,43,.14);color:#c0392b;}
      .ln-chip.mut{background:rgba(128,128,128,.16);color:#6b7885;}
      .ln-empty{border:1px dashed rgba(128,128,128,.4);border-radius:12px;padding:26px;
                text-align:center;color:#7a8794;}
    </style>
    """, unsafe_allow_html=True)

    # ---------- sidebar ----------
    with st.sidebar:
        if pane is None:
            st.markdown(
                f'<img src="{LOGO_DATA_URI}" style="max-width:180px;margin-bottom:0.4rem;"/>',
                unsafe_allow_html=True,
            )
            st.markdown("")
            if st.button("← Volver al inicio", use_container_width=True, key=_pfx + "volver"):
                st.session_state.pop("app_mode", None)
                st.rerun()
            st.divider()

        st.markdown("### 🚗 Loaners")
        st.caption(
            "Flota de vehículos de cortesía: qué unidad está disponible, cuál "
            "está prestada, a qué cliente y desde cuándo."
        )

        st.markdown("### Filtros")
        _f_busq = st.text_input("Buscar (patente, VIN, cliente, N° caso)", "",
                                key=_pfx + "f_busq")
        _f_suc = st.multiselect("Sucursal", sorted({str(f.get("sucursal", "")).strip()
                                                    for f in _filas if str(f.get("sucursal", "")).strip()}),
                                placeholder="Todas", key=_pfx + "f_suc")
        _f_estado = st.radio("Estado", ["Todos", "🟢 Disponibles", "🔴 Asignados"],
                             index=0, key=_pfx + "f_estado")

        st.divider()
        if st.button("🔄 Actualizar datos", use_container_width=True, key=_pfx + "refresh"):
            _cargar_loaners_estado.clear()
            st.rerun()

        if pane is None:
            st.divider()
            st.markdown(f"**Usuario:** `{usuario_activo}`")
            if st.button("🚪 Cerrar sesión", use_container_width=True, key=_pfx + "logout"):
                for _k in ["authenticated", "user_email", "app_mode"]:
                    st.session_state.pop(_k, None)
                st.rerun()

    # ---------- hero ----------
    _n_total = len(_filas)
    _n_disp = sum(1 for f in _filas if _loaners_disponible(f))
    _n_asig = _n_total - _n_disp
    _dias_todos = [d for d in (_loaners_dias(f) for f in _filas) if d is not None]
    _max_dias = max(_dias_todos) if _dias_todos else 0
    _n_sucs = len({str(f.get("sucursal", "")).strip() for f in _filas
                   if str(f.get("sucursal", "")).strip()})

    st.markdown(
        f'''<div class="ln-hero">
              <h2>🚗 Loaners — Vehículos de cortesía</h2>
              <p>Control de la flota de reemplazo: disponibilidad, cliente asignado,
                 kilometraje de salida y N° de caso Salesforce.</p>
            </div>''',
        unsafe_allow_html=True,
    )

    st.markdown(
        '<div class="ln-kpis">'
        f'<div class="ln-kpi"><div class="v">{_n_total}</div><div class="l">Flota total</div></div>'
        f'<div class="ln-kpi verde"><div class="v">{_n_disp}</div><div class="l">Disponibles</div></div>'
        f'<div class="ln-kpi rojo"><div class="v">{_n_asig}</div><div class="l">Asignados</div></div>'
        f'<div class="ln-kpi ambar"><div class="v">{_max_dias}</div><div class="l">Días del préstamo más largo</div></div>'
        f'<div class="ln-kpi gris"><div class="v">{_n_sucs}</div><div class="l">Sucursales</div></div>'
        '</div>',
        unsafe_allow_html=True,
    )

    # ---------- filtros aplicados ----------
    _vis = list(_filas)
    if _f_suc:
        _vis = [f for f in _vis if str(f.get("sucursal", "")).strip() in _f_suc]
    if _f_estado == "🟢 Disponibles":
        _vis = [f for f in _vis if _loaners_disponible(f)]
    elif _f_estado == "🔴 Asignados":
        _vis = [f for f in _vis if not _loaners_disponible(f)]
    if _f_busq.strip():
        _q = _f_busq.strip().lower()
        _vis = [f for f in _vis if any(
            _q in str(f.get(_c, "")).lower()
            for _c in ("patente", "vin", "modelo", "vin_cliente", "nombre_cliente",
                       "modelo_cliente", "caso_sf", "sucursal"))]

    if len(_vis) != len(_filas):
        st.caption(f"Mostrando **{len(_vis)}** de **{len(_filas)}** unidades "
                   "(filtros activos en el panel izquierdo).")

    _tab_tabla, _tab_cards = st.tabs(["📋 Listado y edición", "🗂️ Vista tarjetas"])

    # =========================================================
    #   LISTADO EDITABLE
    # =========================================================
    with _tab_tabla:
        st.caption(
            "Editables: **Estado**, **Sucursal** y de **Fecha Solicitud** en adelante. "
            "VIN, Modelo y Patente del loaner son fijos (identifican la unidad). "
            "Al marcar una unidad como 🟢 **Disponible** se limpian los datos del "
            "cliente, igual que en el Excel; si la marcas 🔴 **Asignado** sin fecha, "
            "se usa la fecha de hoy."
        )

        if not _vis:
            st.markdown('<div class="ln-empty">Sin unidades con los filtros actuales.</div>',
                        unsafe_allow_html=True)
        else:
            _rows = []
            for f in _vis:
                _disp = _loaners_disponible(f)
                _d = _loaners_dias(f)
                try:
                    _kms_v = int(float(f.get("kms_salida") or 0))
                except Exception:
                    _kms_v = 0
                _rows.append({
                    "Estado": "🟢 Disponible" if _disp else "🔴 Asignado",
                    "Sucursal": str(f.get("sucursal", "") or ""),
                    "VIN Loaner": str(f.get("vin", "") or ""),
                    "Modelo Loaner": str(f.get("modelo", "") or ""),
                    "Patente Loaner": str(f.get("patente", "") or ""),
                    "Fecha Solicitud": _loaners_a_date(f.get("fecha_solicitud")),
                    "KMS Salida": _kms_v,
                    "Vin Cliente": str(f.get("vin_cliente", "") or ""),
                    "Nombre Cliente": str(f.get("nombre_cliente", "") or ""),
                    "Modelo Unidad Cliente": str(f.get("modelo_cliente", "") or ""),
                    "Fecha OT": _loaners_a_date(f.get("fecha_ot")),
                    "N° caso Salesforce": str(f.get("caso_sf", "") or ""),
                    "Días prestado": (_d if _d is not None else None),
                    "Última edición": (f"{f.get('_editado_por','')} — {f.get('_editado','')}"
                                       if f.get("_editado") else ""),
                })
            _df_ln = pd.DataFrame(_rows)
            _df_ln["Fecha Solicitud"] = pd.to_datetime(_df_ln["Fecha Solicitud"], errors="coerce")
            _df_ln["Fecha OT"] = pd.to_datetime(_df_ln["Fecha OT"], errors="coerce")

            _opts_estado = ["🟢 Disponible", "🔴 Asignado"]
            _df_ed = st.data_editor(
                _df_ln, hide_index=True, use_container_width=True,
                key=_pfx + "editor", num_rows="fixed",
                column_config={
                    "Estado": st.column_config.SelectboxColumn(
                        "Estado", options=_opts_estado, required=True, width="small",
                        help="Disponible = la unidad está libre. Al guardarla como "
                             "Disponible se borran los datos del cliente."),
                    "Sucursal": st.column_config.SelectboxColumn(
                        "Sucursal", options=_sucs_opts, width="medium",
                        help="Si falta una, agrégala más abajo en '➕ Agregar sucursal a la lista'."),
                    "VIN Loaner": st.column_config.TextColumn("VIN Loaner", disabled=True, width="medium"),
                    "Modelo Loaner": st.column_config.TextColumn("Modelo Loaner", disabled=True),
                    "Patente Loaner": st.column_config.TextColumn("Patente Loaner", disabled=True, width="small"),
                    "Fecha Solicitud": st.column_config.DateColumn(
                        "Fecha Solicitud", format="DD/MM/YYYY", width="small"),
                    "KMS Salida": st.column_config.NumberColumn(
                        "KMS Salida", format="%d", min_value=0, step=1, width="small"),
                    "Vin Cliente": st.column_config.TextColumn("Vin Cliente", width="medium"),
                    "Nombre Cliente": st.column_config.TextColumn("Nombre Cliente", width="large"),
                    "Modelo Unidad Cliente": st.column_config.TextColumn("Modelo Unidad Cliente"),
                    "Fecha OT": st.column_config.DateColumn("Fecha OT", format="DD/MM/YYYY", width="small"),
                    "N° caso Salesforce": st.column_config.TextColumn("N° caso Salesforce", width="small"),
                    "Días prestado": st.column_config.NumberColumn(
                        "Días", disabled=True, format="%d", width="small",
                        help="Días transcurridos desde la Fecha de Solicitud"),
                    "Última edición": st.column_config.TextColumn(
                        "Última edición", disabled=True, width="medium"),
                },
            )

            _c_guardar, _c_info = st.columns([1, 3])
            with _c_guardar:
                _click_guardar = st.button("💾 Guardar cambios", type="primary",
                                           use_container_width=True, key=_pfx + "guardar")
            with _c_info:
                st.caption("Los cambios se guardan en GitHub y quedan visibles para todo "
                           "el equipo, con tu correo y la fecha de la edición.")

            if _click_guardar:
                _por_vin = {str(f.get("vin")): f for f in _filas}
                _cambios = {}
                for _, _r in _df_ed.iterrows():
                    _vin = str(_r["VIN Loaner"]).strip()
                    _orig = _por_vin.get(_vin)
                    if not _orig:
                        continue
                    _disp_nuevo = str(_r["Estado"]).startswith("🟢")
                    _nuevo = {
                        "sucursal": str(_r["Sucursal"] or "").strip(),
                        "kms_salida": (int(_r["KMS Salida"]) if pd.notna(_r["KMS Salida"]) else 0),
                    }
                    if _disp_nuevo:
                        # Unidad libre: replica exactamente el Excel (E dice
                        # "Disponible" y las columnas del cliente van vacias).
                        _nuevo["fecha_solicitud"] = LOANERS_DISPONIBLE
                        for _c in LOANERS_CAMPOS_CLIENTE:
                            _nuevo[_c] = ""
                    else:
                        # Asignado sin fecha -> se asume que salio hoy (si no, la
                        # unidad volveria a verse como Disponible al recargar).
                        _nuevo["fecha_solicitud"] = (
                            _loaners_a_iso(_r["Fecha Solicitud"])
                            or datetime.now(_TZ_CHILE).strftime("%Y-%m-%d"))
                        _nuevo["vin_cliente"] = str(_r["Vin Cliente"] or "").strip()
                        _nuevo["nombre_cliente"] = str(_r["Nombre Cliente"] or "").strip()
                        _nuevo["modelo_cliente"] = str(_r["Modelo Unidad Cliente"] or "").strip()
                        _nuevo["fecha_ot"] = _loaners_a_iso(_r["Fecha OT"])
                        _nuevo["caso_sf"] = str(_r["N° caso Salesforce"] or "").strip()

                    _dif = False
                    for _k, _v in _nuevo.items():
                        _ant = _orig.get(_k, "")
                        if _k == "kms_salida":
                            try:
                                _ant = int(float(_ant or 0))
                            except Exception:
                                _ant = 0
                        elif _k in ("fecha_solicitud", "fecha_ot"):
                            _ant = _loaners_a_iso(_ant) or (
                                LOANERS_DISPONIBLE if (_k == "fecha_solicitud"
                                                       and _loaners_disponible(_orig)) else "")
                        else:
                            _ant = str(_ant or "").strip()
                        if str(_ant) != str(_v):
                            _dif = True
                            break
                    if _dif:
                        _cambios[_vin] = _nuevo

                if not _cambios:
                    st.info("No hay cambios que guardar.")
                else:
                    with st.spinner("Guardando en GitHub..."):
                        _ok = _guardar_loaners(_cambios, usuario_activo)
                    if _ok:
                        _cargar_loaners_estado.clear()
                        st.success(f"✅ {len(_cambios)} unidad(es) actualizada(s).")
                        st.rerun()
                    else:
                        st.error("No se pudo guardar en GitHub. Reintenta en unos segundos.")

        with st.expander("➕ Agregar sucursal a la lista", expanded=False):
            st.caption(
                "Para casos como *Linderos (RANCAGUA)*, cuando la unidad de una "
                "sucursal está prestada a otra. Queda disponible para todo el equipo."
            )
            _c_new1, _c_new2 = st.columns([3, 1])
            with _c_new1:
                _nueva_suc = st.text_input("Nombre de la sucursal", "",
                                           key=_pfx + "nueva_suc",
                                           label_visibility="collapsed",
                                           placeholder="Ej: Linderos (RANCAGUA)")
            with _c_new2:
                if st.button("➕ Agregar", use_container_width=True, key=_pfx + "btn_nueva_suc"):
                    if not _nueva_suc.strip():
                        st.warning("Escribe un nombre primero.")
                    elif _guardar_loaners_sucursal_extra(_nueva_suc, usuario_activo):
                        _cargar_loaners_estado.clear()
                        st.success(f"✅ '{_nueva_suc.strip()}' agregada.")
                        st.rerun()
                    else:
                        st.error("No se pudo guardar la sucursal.")

    # =========================================================
    #   VISTA TARJETAS
    # =========================================================
    with _tab_cards:
        if not _vis:
            st.markdown('<div class="ln-empty">Sin unidades con los filtros actuales.</div>',
                        unsafe_allow_html=True)
        else:
            _ordenadas = sorted(
                _vis, key=lambda f: (_loaners_disponible(f), -(_loaners_dias(f) or 0)))
            _cols_card = st.columns(2, gap="medium")
            for _i, f in enumerate(_ordenadas):
                _disp = _loaners_disponible(f)
                _d = _loaners_dias(f)
                try:
                    _kms_txt = f"{int(float(f.get('kms_salida') or 0)):,}".replace(",", ".")
                except Exception:
                    _kms_txt = "--"
                _chips = [f'<span class="ln-chip">{_ln_esc(f.get("sucursal") or "Sin sucursal")}</span>']
                if _disp:
                    _chips.append('<span class="ln-chip ok">🟢 Disponible</span>')
                else:
                    _chips.append('<span class="ln-chip warn">🔴 Asignado</span>')
                    if _d is not None:
                        _cl = "warn" if _d >= 30 else "mut"
                        _chips.append(f'<span class="ln-chip {_cl}">⏱ {_d} día(s)</span>')
                _chips.append(f'<span class="ln-chip mut">🛣️ {_kms_txt} km</span>')

                if _disp:
                    _cuerpo = ('<div class="cli" style="color:#7a8794;">'
                               'Unidad libre — sin cliente asignado.</div>')
                else:
                    _cuerpo = (
                        '<div class="cli">'
                        f'<b>Cliente:</b> {_ln_esc(f.get("nombre_cliente") or "--")}<br>'
                        f'<b>Su unidad:</b> {_ln_esc(f.get("modelo_cliente") or "--")} '
                        f'· <span class="vin">{_ln_esc(f.get("vin_cliente") or "--")}</span><br>'
                        f'<b>Solicitud:</b> {_loaners_fecha_txt(f.get("fecha_solicitud")) or "--"} '
                        f'· <b>Fecha OT:</b> {_loaners_fecha_txt(f.get("fecha_ot")) or "--"}<br>'
                        f'<b>N° caso:</b> {_ln_esc(f.get("caso_sf") or "--")}'
                        '</div>')

                with _cols_card[_i % 2]:
                    st.markdown(
                        f'''<div class="ln-card {"" if _disp else "asig"}">
                              <div class="pat">{_ln_esc(f.get("patente"))}</div>
                              <div class="mod">{_ln_esc(f.get("modelo"))}</div>
                              <div class="vin">VIN {_ln_esc(f.get("vin"))}</div>
                              <div class="ln-chips">{"".join(_chips)}</div>
                              {_cuerpo}
                            </div>''',
                        unsafe_allow_html=True,
                    )

    # =========================================================
    #   EXPORTAR
    # =========================================================
    st.divider()
    _c_exp1, _c_exp2 = st.columns([1, 3])
    with _c_exp1:
        if st.button("📥 Exportar listado (Excel)", type="primary",
                     use_container_width=True, key=_pfx + "btn_export"):
            try:
                st.session_state[_pfx + "xlsx"] = generar_excel_loaners(_filas)
            except Exception as _e:
                st.session_state.pop(_pfx + "xlsx", None)
                st.error(f"No se pudo generar el Excel: {_e}")
    with _c_exp2:
        st.caption(
            "Genera el listado **completo** de la flota (no depende de los filtros), "
            "con las mismas 11 columnas y el mismo formato del archivo original."
        )

    if st.session_state.get(_pfx + "xlsx"):
        st.download_button(
            "⬇️ Descargar LOANERS.xlsx",
            data=st.session_state[_pfx + "xlsx"],
            file_name=f"LOANERS_{datetime.now(_TZ_CHILE).strftime('%Y%m%d_%H%M')}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True, key=_pfx + "dl",
        )

    if pane is None:
        st.stop()


if st.session_state.get("app_mode") == "loaners":
    _render_loaners()



# ============================================================
#   CARGA Y NORMALIZACIÓN
# ============================================================
df_raw, fecha_actualizacion = cargar_datos()

if df_raw.empty:
    st.error(f"⚠️ No se pudieron cargar los datos. {fecha_actualizacion}")
    st.info("Verifica que el script de consolidación haya corrido.")
    st.stop()

for col in ["RANGO","SUCURSAL","TIPO VENTA","TIPO CLIENTE","MARCA","ASESOR","ESTADO",
            "CATEGORIA","OBSERVACION OT","NOTAS","AVANCE - GESTIÓN","ULTIMA_EDICION",
            "_MARCA_COLOR_","ETAPA_JPCB","AÑO"]:
    if col not in df_raw.columns:
        df_raw[col] = ""
    df_raw[col] = df_raw[col].fillna("").astype(str).str.strip()

for _, key in DOCS_CONFIG:
    for prefix in (f"N_{key}", f"FOLIOS_{key}"):
        if prefix not in df_raw.columns:
            df_raw[prefix] = "" if prefix.startswith("FOLIOS") else 0
    df_raw[f"N_{key}"] = pd.to_numeric(df_raw[f"N_{key}"], errors="coerce").fillna(0).astype(int)
    df_raw[f"FOLIOS_{key}"] = df_raw[f"FOLIOS_{key}"].fillna("").astype(str)

if "NETO" in df_raw.columns:
    df_raw["NETO"] = (
        df_raw["NETO"].astype(str).str.strip()
        .str.replace(".", "", regex=False)
        .str.replace(",", ".", regex=False)
    )
    df_raw["NETO"] = pd.to_numeric(df_raw["NETO"], errors="coerce").fillna(0).astype(int)

# ============================================================
#   RESTRICCIÓN DE ACCESO POR SUCURSAL
#   -------------------------------------------------
#   `df_raw_full` conserva TODAS las sucursales (se usa para armar el
#   selector del Planificador de Taller, para que un usuario restringido
#   igual pueda entrar a su sucursal aunque esta no tenga OTs pendientes
#   en este momento). `df_raw` (la que consumen el resto de las pestañas
#   de Control y Gestión Post Venta y el Asistente App) queda filtrada a
#   solo las sucursales que el usuario tiene permitidas.
# ============================================================
df_raw_full = df_raw.copy()
_usuarios_restr_cache = _leer_usuarios()
_mis_sucursales = _sucursales_permitidas_usuario(usuario_activo, _usuarios_restr_cache)
if _mis_sucursales:
    _mis_sucursales_norm = {s.strip().upper() for s in _mis_sucursales}
    df_raw = df_raw[df_raw["SUCURSAL"].str.upper().isin(_mis_sucursales_norm)].reset_index(drop=True)


# ============================================================
#   PLANIFICADOR DE TALLER  — modo independiente
#   (JPCB por patente + Agenda Curifor + Control de Taller +
#   Vehiculos en Taller. Este es el desarrollo que antes vivia
#   como pestaña "Planificador de Taller" dentro del modulo
#   "Control y Gestion Post Venta" — se movio aqui para que sea
#   su propio modulo standalone. Reemplaza al JPCB antiguo por
#   Folio OT / ETAPA_JPCB, que fue eliminado — 01/07/2026.)
# ============================================================
def _render_planificador(pane=None):

    if not _puede_usar_planificador(usuario_activo):
        if pane is None:
            st.session_state.pop("app_mode", None)
            st.error("🔒 No tienes autorización para usar el Planificador de Taller. "
                     "Solicítala al administrador (cjerez@curifor.com).")
            if st.button("← Volver al inicio", key="plan_sin_acceso_volver"):
                st.rerun()
            st.stop()
        else:
            st.error("🔒 No tienes autorización para usar el Planificador de Taller.")
            return

    # Sin "Todas" — el planificador es por taller especifico. Se arma desde
    # df_raw_full (TODAS las sucursales reales) para que un usuario restringido
    # por sucursal pueda igual entrar a la suya aunque hoy no tenga OTs
    # pendientes en Control y Gestion Post Venta.
    _sucs_planif = sorted(df_raw_full["SUCURSAL"].unique().tolist())
    if _mis_sucursales:
        _mis_sucursales_norm_pl = {s.strip().upper() for s in _mis_sucursales}
        _sucs_planif = [s for s in _sucs_planif if s.strip().upper() in _mis_sucursales_norm_pl]
    # Recuperar ultima sucursal usada en esta sesion
    _suc_default = st.session_state.get("planif_suc_last", _sucs_planif[0] if _sucs_planif else "")
    _suc_idx     = _sucs_planif.index(_suc_default) if _suc_default in _sucs_planif else 0

    with st.sidebar:
        if pane is None:
            st.markdown(
                f'<img src="{LOGO_DATA_URI}" style="max-width:180px;margin-bottom:0.4rem;"/>',
                unsafe_allow_html=True,
            )
            if st.button("← Volver al inicio", use_container_width=True, key="plan_volver"):
                st.session_state.pop("app_mode", None)
                st.rerun()
            st.divider()

        st.markdown("### Sucursal / Taller")
        _suc_planif = st.selectbox("Sucursal / Taller", _sucs_planif, index=_suc_idx,
                                    key="planif_suc_standalone", label_visibility="collapsed")
        st.session_state["planif_suc_last"] = _suc_planif

        st.divider()
        if st.button("🔄 Actualizar datos", use_container_width=True, key="plan_refresh"):
            st.cache_data.clear()
            st.rerun()
        st.caption("Presiona 'Actualizar datos' para traer los cambios mas recientes de otros usuarios "
                   "(Torre Control, tecnicos, etc.) — tambien se mezclan solos cada vez que guardas algo.")

        if pane is None:
            st.divider()
            st.markdown(f"**Usuario:** `{usuario_activo}`")
            if st.button("🚪 Cerrar sesión", use_container_width=True, key="plan_logout"):
                for _k in ["authenticated", "user_email", "app_mode"]:
                    st.session_state.pop(_k, None)
                st.rerun()

    st.markdown(
        f'''<div class="curifor-header">
            <div class="logo-pill"><img src="{LOGO_DATA_URI}" /></div>
            <div class="curifor-header-text">
                <h2>Planificador de Taller</h2>
                <p>JPCB por patente · Agenda Curifor · Control de Taller · Vehículos en Taller</p>
                <span class="dev-credit">Curifor S.A</span>
            </div>
        </div>''',
        unsafe_allow_html=True,
    )

    _puede_ed = _puede_editar_planificador(usuario_activo)
    _puede_conf_citas = _puede_confirmar_citas(usuario_activo)
    _txt_ed = ("✅ Modo edición (Admin)" if usuario_activo == ADMIN_EMAIL
               else "✅ Modo edición" if _puede_ed
               else "📅 Solo puede confirmar Asiste/No Asiste/Reagenda" if _puede_conf_citas
               else "👁️ Solo lectura")
    st.caption(f"Sucursal: **{_suc_planif}** · {_txt_ed}")

    if not _suc_planif:
        st.warning("Selecciona una sucursal para cargar el planificador.")
        if pane is None:
            st.stop()
        return

    # Defensa en profundidad: si por algun motivo quedo seleccionada una
    # sucursal fuera de las permitidas (ej. session_state de antes de que se
    # activara la restriccion), se bloquea aca tambien, no solo en el selector.
    if _mis_sucursales and _suc_planif.strip().upper() not in {s.strip().upper() for s in _mis_sucursales}:
        st.error(f"🔒 No tienes autorización para ver la sucursal '{_suc_planif}'. "
                 f"Tu acceso está limitado a: {', '.join(_mis_sucursales)}.")
        if pane is None:
            st.stop()
        return

    _agenda_data_py             = _cargar_agenda_hoy()
    _ctrl_data_py, _ctrl_sha_py = _cargar_ctrl_taller()
    _pp_data_py, _pp_sha_py     = _cargar_prepicking()
    _puede_pp                   = _puede_usar_prepicking(usuario_activo)
    _prod_resumen_py, _prod_detalle_py, _prod_detalle_ot_py, _prod_fecha_py = _cargar_produccion_tecnicos()
    _cot_gz_py, _cot_fecha_py = _cargar_cotizador_gz()
    # El catalogo completo de Stock (~1-1.5 MB comprimido) solo se carga/inyecta
    # si el usuario tiene acceso a Pre-picking — para el resto del Planificador
    # no aporta nada y solo pesaria la pagina de mas.
    _stock_gz_py = _cargar_stock_completo_gz()[0] if _puede_pp else ""

    _html_planif = _generar_html_planificador(
        sucursal    = _suc_planif,
        usuario     = usuario_activo,
        puede_editar= _puede_ed,
        token       = GITHUB_TOKEN,
        github_user = GITHUB_USUARIO,
        github_repo = GITHUB_REPO,
        agenda_data = _agenda_data_py,
        ctrl_data   = _ctrl_data_py,
        ctrl_sha    = _ctrl_sha_py,
        puede_confirmar_citas = _puede_conf_citas,
        puede_disponibilidad  = _puede_disponibilidad_tecnicos(usuario_activo),
        puede_prepicking = _puede_pp,
        prepicking_data  = _pp_data_py,
        prepicking_sha   = _pp_sha_py,
        logo_data_uri    = LOGO_DATA_URI,
        produccion_data  = ({"resumen": _prod_resumen_py, "detalle_producto": _prod_detalle_py,
                             "detalle_ot": _prod_detalle_ot_py, "fecha_actualizacion": _prod_fecha_py}
                            if _prod_resumen_py else None),
        cotizador_gz     = _cot_gz_py,
        stock_completo_gz = _stock_gz_py,
    )
    components.html(_html_planif, height=(940 if pane is None else 620), scrolling=True)

    if pane is None:
        st.stop()   # ← Detener aquí: no ejecutar el resto del app (modo OTs)


if st.session_state.get("app_mode") == "planificador":
    _render_planificador()


def _render_control(pane=None):

    # ============================================================
    #   CONTROL Y GESTIÓN POST VENTA — guardia de acceso
    #   (defensa en profundidad: si alguien llega aquí sin autorización,
    #   por ejemplo manipulando el estado de sesión, se corta el flujo)
    # ============================================================
    if not _puede_usar_control(usuario_activo):
        if pane is None:
            st.session_state.pop("app_mode", None)
            st.error("🔒 No tienes autorización para el módulo Control y Gestión Post Venta. "
                     "Solicítala al administrador (cjerez@curifor.com).")
            if st.button("← Volver al inicio", key="ots_sin_acceso_volver"):
                st.rerun()
            st.stop()
        else:
            st.error("🔒 No tienes autorización para el módulo Control y Gestión Post Venta.")
            return


    # ============================================================
    #   NOTIFICACIONES — conteo para badge
    # ============================================================
    _notifs_raw     = cargar_notificaciones_cache()
    _notifs_propias = [n for n in _notifs_raw
                       if n.get("destinatario", "").lower() == usuario_activo.lower()
                       and not n.get("leida")]
    _badge_count    = len(_notifs_propias)


    # ============================================================
    #   NAVEGACIÓN ENTRANTE → Detalle y Edición (desde Repuestos Pendientes)
    #   Debe ejecutarse ANTES de crear el buscador del sidebar para poder
    #   precargar su valor con el folio de la OT pendiente destino.
    # ============================================================
    _nav_det = st.session_state.pop("_nav_buscar_folio", None)
    if _nav_det:
        st.session_state["busqueda_global"] = str(_nav_det)
        st.session_state["_nav_radio"] = "✏️ Detalle y Edición"


    # ============================================================
    #   SIDEBAR
    # ============================================================
    with st.sidebar:
        if pane is None:
            st.markdown(f'''<img src="{LOGO_DATA_URI}" style="max-width:180px; margin-bottom:0.5rem;" />''', unsafe_allow_html=True)
            st.markdown(f"**Última actualización:**  \n{fecha_actualizacion}")
            if st.button("🏠 Inicio", use_container_width=True, key="ots_volver_inicio",
                         help="Volver al selector de módulos"):
                st.session_state.pop("app_mode", None)
                st.rerun()
            st.divider()

        # Badge de notificaciones (clicable → navega a tab Notificaciones)
        if _badge_count > 0:
            if st.button(
                f"🔔 {_badge_count} notificación(es) sin leer",
                use_container_width=True,
                key="btn_sidebar_notifs",
                help="Ir a la pestaña de Notificaciones",
            ):
                st.session_state["_ir_a_notifs"] = True
                st.rerun()

        if _mis_sucursales:
            st.caption(f"🔒 Acceso limitado a: {', '.join(_mis_sucursales)}")

        st.markdown("### Filtros")
        sel_sucursal = st.multiselect("Sucursal",
                                      sorted(df_raw["SUCURSAL"].unique().tolist()),
                                      placeholder="Todas")
        sel_rango    = st.multiselect("Rango de días",
                                      ["0-30","31-60","61-90","91 o más"],
                                      placeholder="Todos")
        sel_tipo     = st.multiselect("Tipo de venta",
                                      sorted(df_raw["TIPO VENTA"].unique().tolist()),
                                      placeholder="Todos")
        sel_marca    = st.multiselect("Marca",
                                      sorted([m for m in df_raw["MARCA"].unique() if m]),
                                      placeholder="Todas")
        sel_asesor   = st.multiselect("Asesor",
                                      sorted([a for a in df_raw["ASESOR"].unique() if a]),
                                      placeholder="Todos")
        sel_categoria = st.multiselect("Categoría",
                                       sorted([c for c in df_raw["CATEGORIA"].unique()
                                               if c and c.strip().lower() not in
                                               ("nan", "none", "sin categoría", "sin categoria", "")]),
                                       placeholder="Todas")
        sel_documento = st.multiselect("Tiene documento",
                                       [nombre for nombre, _ in DOCS_CONFIG],
                                       placeholder="Todos")

        st.markdown("**Período (Fecha OT)**")
        _fechas_ot_todas = (pd.to_datetime(df_raw["FECHA OT"], dayfirst=True, errors="coerce")
                            if "FECHA OT" in df_raw.columns else pd.Series([], dtype="datetime64[ns]"))
        _fecha_min_ot = _fechas_ot_todas.min()
        _fecha_max_ot = _fechas_ot_todas.max()
        if pd.isna(_fecha_min_ot): _fecha_min_ot = pd.Timestamp.now()
        if pd.isna(_fecha_max_ot): _fecha_max_ot = pd.Timestamp.now()
        _fmin_d, _fmax_d = _fecha_min_ot.date(), _fecha_max_ot.date()

        _cf1, _cf2 = st.columns(2)
        with _cf1:
            fecha_ot_desde = st.date_input("Desde", value=_fmin_d, min_value=_fmin_d,
                                           max_value=_fmax_d, key="filtro_fecha_ot_desde",
                                           format="DD/MM/YYYY")
        with _cf2:
            fecha_ot_hasta = st.date_input("Hasta", value=_fmax_d, min_value=_fmin_d,
                                           max_value=_fmax_d, key="filtro_fecha_ot_hasta",
                                           format="DD/MM/YYYY")
        if fecha_ot_desde > fecha_ot_hasta:
            st.caption("⚠️ 'Desde' es posterior a 'Hasta' — se invierten automáticamente.")
            fecha_ot_desde, fecha_ot_hasta = fecha_ot_hasta, fecha_ot_desde
        _periodo_ot_activo = (fecha_ot_desde != _fmin_d) or (fecha_ot_hasta != _fmax_d)

        busqueda = st.text_input("Buscar (folio, patente, asesor...)", key="busqueda_global")

        if pane is None:
            st.divider()
            st.markdown(f"**Usuario:** `{usuario_activo}`")

            # Cambiar contraseña
            with st.expander("🔑 Cambiar contraseña"):
                cp1 = st.text_input("Contraseña actual", type="password", key="cp_actual")
                cp2 = st.text_input("Nueva contraseña",  type="password", key="cp_nueva")
                cp3 = st.text_input("Confirmar nueva",   type="password", key="cp_confirma")
                if st.button("Actualizar contraseña", use_container_width=True):
                    if cp2 != cp3:
                        st.error("Las contraseñas nuevas no coinciden.")
                    else:
                        ok, msg = cambiar_password(usuario_activo, cp1, cp2)
                        if ok:
                            st.success(msg)
                        else:
                            st.error(msg)

        if st.button("🔄 Actualizar datos", use_container_width=True):
            st.cache_data.clear()
            st.session_state.pop("_color_overrides", None)
            st.rerun()

        if pane is None:
            st.divider()
            if st.button("🚪 Cerrar sesión", use_container_width=True):
                st.session_state.authenticated = False
                st.session_state.user_email    = ""
                st.rerun()


    # ============================================================
    #   FILTROS
    # ============================================================
    df = df_raw.copy()
    if sel_sucursal:  df = df[df["SUCURSAL"].isin(sel_sucursal)]
    if sel_rango:     df = df[df["RANGO"].isin(sel_rango)]
    if sel_tipo:      df = df[df["TIPO VENTA"].isin(sel_tipo)]
    if sel_marca:     df = df[df["MARCA"].isin(sel_marca)]
    if sel_asesor:    df = df[df["ASESOR"].isin(sel_asesor)]
    if sel_categoria: df = df[df["CATEGORIA"].isin(sel_categoria)]
    if sel_documento:
        keys_sel = [key for nombre, key in DOCS_CONFIG if nombre in sel_documento]
        mask_doc = pd.Series([False] * len(df), index=df.index)
        for key in keys_sel:
            col = f"N_{key}"
            if col in df.columns:
                mask_doc |= df[col] > 0
        df = df[mask_doc]
    if _periodo_ot_activo and "FECHA OT" in df.columns:
        _fecha_ot_dt_f = pd.to_datetime(df["FECHA OT"], dayfirst=True, errors="coerce")
        df = df[(_fecha_ot_dt_f.dt.date >= fecha_ot_desde) & (_fecha_ot_dt_f.dt.date <= fecha_ot_hasta)]
    if busqueda:
        mask = pd.Series([False]*len(df), index=df.index)
        for col in ["FOLIO OT","PATENTE","ASESOR","MODELO","GLOSA TRABAJO"]:
            if col in df.columns:
                mask |= df[col].astype(str).str.contains(busqueda, case=False, na=False)
        df = df[mask]


    # ============================================================
    #   HEADER + KPIs
    # ============================================================
    st.markdown(
        f'''<div class="curifor-header">
            <div class="logo-pill">
                <img src="{LOGO_DATA_URI}" />
            </div>
            <div class="curifor-header-text">
                <h2>Control y Gestión Post Venta — Curifor S.A</h2>
                <p>Sistema de seguimiento de órdenes de trabajo</p>
                <span class="dev-credit">Desarrollado por: Cristóbal Jerez J.</span>
            </div>
            <div class="curifor-badge">⏱ Act: {fecha_actualizacion}</div>
        </div>''',
        unsafe_allow_html=True
    )

    filtros_activos = (len(sel_sucursal)+len(sel_rango)+len(sel_tipo)
                       +len(sel_marca)+len(sel_asesor)+len(sel_categoria)
                       +len(sel_documento)+(1 if busqueda else 0)
                       +(1 if _periodo_ot_activo else 0))
    if filtros_activos:
        st.markdown(
            f'<div style="margin:6px 0 14px;">'
            f'<span class="info-chip ok">🔍 {len(df):,} de {len(df_raw):,} OTs</span>'
            f'<span class="info-chip">{filtros_activos} filtro(s) activo(s)</span>'
            f'</div>',
            unsafe_allow_html=True,
        )
    else:
        st.markdown(
            f'<div style="margin:6px 0 14px;">'
            f'<span class="info-chip mut">Total: {len(df_raw):,} OTs pendientes</span>'
            f'</div>',
            unsafe_allow_html=True,
        )

    total     = len(df)
    criticas  = int((df["RANGO"]=="91 o más").sum())
    urgentes  = int((df["RANGO"]=="61-90").sum())
    atencion  = int((df["RANGO"]=="31-60").sum())
    recientes = int((df["RANGO"]=="0-30").sum())

    c1,c2,c3,c4,c5 = st.columns(5)
    with c1: kpi_card("OTs Pendientes",      total)
    with c2: kpi_card("Críticas  · >90 d",  criticas,  "rojo")
    with c3: kpi_card("Urgentes · 61-90 d", urgentes,  "naranja")
    with c4: kpi_card("Atención · 31-60 d", atencion,  "amarillo")
    with c5: kpi_card("Recientes · 0-30 d", recientes, "verde")

    st.divider()


    # ============================================================
    #   NAVEGACIÓN — persistida en session_state (sin JavaScript)
    #   Clave: siempre escribir session_state["_nav_radio"] con el
    #   LABEL actual del tab deseado ANTES de renderizar el radio.
    #   Así Streamlit nunca reseteará la selección a Resumen.
    # ============================================================

    _TAB_RESUMEN   = 0
    _TAB_DETALLE   = 1
    _TAB_DOCS      = 2
    _TAB_HISTORIAL = 3
    _TAB_SUCURSAL  = 4
    _TAB_ASESOR    = 5
    _TAB_ANALISIS  = 6
    _TAB_RANKING   = 7
    _TAB_NOTIFS    = 8
    _TAB_REPUESTOS = 9
    _TAB_CLIENTE   = 10
    _TAB_FACT_X    = 11
    _TAB_ADMIN     = 12

    _badge_str  = "  🔴" if _badge_count > 0 else ""
    _tab_labels = [
        "📊 Resumen",
        "✏️ Detalle y Edición",
        "📄 Documentos y Comentarios",
        "📋 Historial de Comentarios",
        "🏢 Por Sucursal",
        "👤 Por Asesor",
        "📈 Análisis y Tendencias",
        "🏆 Ranking Cierres >90d",
        f"🔔 Notificaciones{_badge_str}",
        "🧰 Repuestos Pendientes",
        "🔍 Búsqueda de Cliente",
        "🧾 Facturas X",
    ]
    if usuario_activo == ADMIN_EMAIL:
        _tab_labels.append("🛡️ Admin")

    # ── Navegación programática: SOLO escribir _nav_radio cuando sea necesario ─
    # Si lo escribimos en cada rerun, cancelamos el clic del usuario.
    # El radio con key="_nav_radio" persiste solo entre reruns normales.
    _folio_nav = st.session_state.get("_nav_folio", None)
    if _folio_nav:
        del st.session_state["_nav_folio"]
        st.session_state["sel_folio_docs"] = _folio_nav
        st.session_state["_nav_radio"] = _tab_labels[_TAB_DOCS]

    if st.session_state.pop("_ir_a_notifs", False):
        st.session_state["_nav_radio"] = _tab_labels[_TAB_NOTIFS]

    # Si el label de notifs cambió (badge apareció/desapareció) y el usuario
    # está en esa pestaña, actualizar el label almacenado para que no se rompa
    _stored_nav = st.session_state.get("_nav_radio", "")
    if _stored_nav and _stored_nav.startswith("🔔"):
        st.session_state["_nav_radio"] = _tab_labels[_TAB_NOTIFS]

    _selected_label = st.radio(
        "Navegación",
        _tab_labels,
        horizontal=True,
        label_visibility="collapsed",
        key="_nav_radio",
    )
    _tab = _tab_labels.index(_selected_label) if _selected_label in _tab_labels else 0

    st.divider()



    # ---- TAB 1: RESUMEN ----------------------------------------
    if _tab == 0:
        col_izq, col_der = st.columns(2)
        with col_izq:
            st.markdown('<p class="section-title">OT por Tipo de Venta</p>', unsafe_allow_html=True)
            tv = df[df["TIPO VENTA"]!=""].groupby("TIPO VENTA").size().reset_index(name="Cantidad")
            tv = tv.sort_values("Cantidad", ascending=False)
            if total > 0:
                tv["% Total"] = (tv["Cantidad"]/total*100).round(1).astype(str)+"%"
            st.dataframe(tv, hide_index=True, use_container_width=True,
                         column_config={"Cantidad": st.column_config.NumberColumn(format="%d")})
        with col_der:
            st.markdown('<p class="section-title">OT por Marca</p>', unsafe_allow_html=True)
            marca = df[df["MARCA"]!=""].groupby("MARCA").size().reset_index(name="Cantidad")
            marca = marca.sort_values("Cantidad", ascending=False)
            st.dataframe(marca, hide_index=True, use_container_width=True,
                         column_config={"Cantidad": st.column_config.NumberColumn(format="%d")})
        st.markdown('<p class="section-title">OT por Sucursal y Rango</p>', unsafe_allow_html=True)
        rangos_orden = ["0-30","31-60","61-90","91 o más"]
        pivot = df.groupby(["SUCURSAL","RANGO"]).size().unstack(fill_value=0)
        for r in rangos_orden:
            if r not in pivot.columns: pivot[r] = 0
        pivot = pivot[rangos_orden]
        pivot["Total"] = pivot.sum(axis=1)
        pivot = pivot.sort_values("Total", ascending=False).reset_index()
        st.dataframe(pivot, hide_index=True, use_container_width=True)

        # ---- Botón Generar Informe PDF --------------------------------
        st.divider()
        _btn_col, _info_col = st.columns([1, 3])
        with _btn_col:
            _gen_pdf = st.button("📊 Generar Informe PDF", type="primary",
                                 use_container_width=True,
                                 help="Genera un informe ejecutivo en PDF: resumen general, desglose por sucursal y "
                                      "costo del Vale de Consumo, detalle crítico (61-90 y >90 días) por tipo de "
                                      "venta/categoría, y una página completa por cada sucursal (rango, categoría, "
                                      "asesor, costos y las OTs más críticas)")
        with _info_col:
            _filtros_activos_resumen = (len(sel_sucursal)+len(sel_rango)+len(sel_tipo)+len(sel_marca)+(1 if busqueda else 0)+(1 if _periodo_ot_activo else 0))
            if _filtros_activos_resumen:
                st.info(f"El informe considerará los **{len(df):,} datos filtrados** actualmente ({_filtros_activos_resumen} filtro(s) activo(s)).")
            else:
                st.info(f"El informe incluirá el **total de {len(df):,} OTs pendientes** (sin filtros activos).")

        if _gen_pdf:
            _filtros_desc = "Sin filtros (todos los datos)"
            _parts = []
            if sel_sucursal: _parts.append(f"Sucursales: {', '.join(sel_sucursal)}")
            if sel_rango:    _parts.append(f"Rango: {', '.join(sel_rango)}")
            if sel_tipo:     _parts.append(f"Tipo: {', '.join(sel_tipo)}")
            if sel_marca:    _parts.append(f"Marca: {', '.join(sel_marca)}")
            if _periodo_ot_activo: _parts.append(f"Período: {fecha_ot_desde.strftime('%d/%m/%Y')} al {fecha_ot_hasta.strftime('%d/%m/%Y')}")
            if busqueda:     _parts.append(f"Búsqueda: {busqueda}")
            if _parts:       _filtros_desc = "  |  ".join(_parts)

            with st.spinner("Generando informe PDF..."):
                try:
                    _pdf_bytes = generar_pdf_informe(df, _filtros_desc, ahora_chile(), LOGO_B64)
                    _nombre_pdf = f"Informe_OTs_{datetime.now(_TZ_CHILE).strftime('%Y%m%d_%H%M')}.pdf"
                    st.download_button(
                        label="⬇️ Descargar Informe PDF",
                        data=_pdf_bytes,
                        file_name=_nombre_pdf,
                        mime="application/pdf",
                        use_container_width=True,
                    )
                    st.success("✅ Informe generado. Presiona el botón de descarga.")
                except Exception as _e:
                    st.error(f"Error al generar el PDF: {_e}")

        # ---- Botón Generar Informe PDF por Sucursal y Año -------------
        st.divider()
        _btn_col2, _info_col2 = st.columns([1, 3])
        with _btn_col2:
            _gen_pdf2 = st.button("🏢 Informe por Sucursal y Año", type="secondary",
                                  use_container_width=True,
                                  help="Genera un PDF con OTs por sucursal y año de apertura, su costo "
                                       "(Vale de Consumo) bajo el mismo cruce, cantidad de OTs por marca, "
                                       "y una hoja adicional con el cruce Sucursal × Categoría × Año")
        with _info_col2:
            _filtros_activos_resumen2 = (len(sel_sucursal)+len(sel_rango)+len(sel_tipo)+len(sel_marca)+(1 if busqueda else 0)+(1 if _periodo_ot_activo else 0))
            if _filtros_activos_resumen2:
                st.info(f"El informe considerará los **{len(df):,} datos filtrados** actualmente ({_filtros_activos_resumen2} filtro(s) activo(s)).")
            else:
                st.info(f"El informe incluirá el **total de {len(df):,} OTs pendientes** (sin filtros activos).")

        if _gen_pdf2:
            _filtros_desc2 = "Sin filtros (todos los datos)"
            _parts2 = []
            if sel_sucursal: _parts2.append(f"Sucursales: {', '.join(sel_sucursal)}")
            if sel_rango:    _parts2.append(f"Rango: {', '.join(sel_rango)}")
            if sel_tipo:     _parts2.append(f"Tipo: {', '.join(sel_tipo)}")
            if sel_marca:    _parts2.append(f"Marca: {', '.join(sel_marca)}")
            if _periodo_ot_activo: _parts2.append(f"Período: {fecha_ot_desde.strftime('%d/%m/%Y')} al {fecha_ot_hasta.strftime('%d/%m/%Y')}")
            if busqueda:     _parts2.append(f"Búsqueda: {busqueda}")
            if _parts2:      _filtros_desc2 = "  |  ".join(_parts2)

            with st.spinner("Generando informe PDF..."):
                try:
                    _pdf_bytes2 = generar_pdf_sucursal_anio(df, _filtros_desc2, ahora_chile(), LOGO_B64)
                    _nombre_pdf2 = f"Informe_Sucursal_Anio_{datetime.now(_TZ_CHILE).strftime('%Y%m%d_%H%M')}.pdf"
                    st.download_button(
                        label="⬇️ Descargar Informe por Sucursal y Año",
                        data=_pdf_bytes2,
                        file_name=_nombre_pdf2,
                        mime="application/pdf",
                        use_container_width=True,
                        key="btn_download_pdf_suc_anio",
                    )
                    st.success("✅ Informe generado. Presiona el botón de descarga.")
                except Exception as _e2:
                    st.error(f"Error al generar el PDF: {_e2}")

        # ---- Botón Generar Informe PDF por Área -----------------------
        st.divider()
        _btn_col3, _info_col3 = st.columns([1, 3])
        with _btn_col3:
            _gen_pdf3 = st.button("🧩 Informe por Área", type="secondary",
                                  use_container_width=True,
                                  help="Genera un PDF de 2 páginas: OTs y costo del Vale de Consumo por Área "
                                       "(Servicio Técnico, Garantía, Interno, DyP) cruzadas con rango de días y "
                                       "sucursal, más un análisis mensual de las OT con más de 90 días de apertura")
        with _info_col3:
            _filtros_activos_resumen3 = (len(sel_sucursal)+len(sel_rango)+len(sel_tipo)+len(sel_marca)+(1 if busqueda else 0)+(1 if _periodo_ot_activo else 0))
            if _filtros_activos_resumen3:
                st.info(f"El informe considerará los **{len(df):,} datos filtrados** actualmente ({_filtros_activos_resumen3} filtro(s) activo(s)).")
            else:
                st.info(f"El informe incluirá el **total de {len(df):,} OTs pendientes** (sin filtros activos).")

        if _gen_pdf3:
            _filtros_desc3 = "Sin filtros (todos los datos)"
            _parts3 = []
            if sel_sucursal: _parts3.append(f"Sucursales: {', '.join(sel_sucursal)}")
            if sel_rango:    _parts3.append(f"Rango: {', '.join(sel_rango)}")
            if sel_tipo:     _parts3.append(f"Tipo: {', '.join(sel_tipo)}")
            if sel_marca:    _parts3.append(f"Marca: {', '.join(sel_marca)}")
            if _periodo_ot_activo: _parts3.append(f"Período: {fecha_ot_desde.strftime('%d/%m/%Y')} al {fecha_ot_hasta.strftime('%d/%m/%Y')}")
            if busqueda:     _parts3.append(f"Búsqueda: {busqueda}")
            if _parts3:      _filtros_desc3 = "  |  ".join(_parts3)

            with st.spinner("Generando informe PDF..."):
                try:
                    _pdf_bytes3 = generar_pdf_por_area(df, _filtros_desc3, ahora_chile(), LOGO_B64)
                    _nombre_pdf3 = f"Informe_Por_Area_{datetime.now(_TZ_CHILE).strftime('%Y%m%d_%H%M')}.pdf"
                    st.download_button(
                        label="⬇️ Descargar Informe por Área",
                        data=_pdf_bytes3,
                        file_name=_nombre_pdf3,
                        mime="application/pdf",
                        use_container_width=True,
                        key="btn_download_pdf_area",
                    )
                    st.success("✅ Informe generado. Presiona el botón de descarga.")
                except Exception as _e3:
                    st.error(f"Error al generar el PDF: {_e3}")


    # ---- TAB 2: DETALLE Y EDICIÓN --------------------------------
    elif _tab == 1:
        _RANGO_ICONS  = {"91 o más": "🔴", "61-90": "🟠", "31-60": "🟡", "0-30": "🟢"}
        _EMOJI_COLOR  = {"🔴": "🔴 Rojo", "🟡": "🟡 Amarillo", "🟢": "🟢 Verde", "🔵": "🔵 Azul"}
        _COLOR_EMOJI  = {v: k for k, v in _EMOJI_COLOR.items()}
        _EMOJIS_FOLIO = set(_EMOJI_COLOR.keys())

        def _folio_a_display(folio_clean, color):
            emoji = _COLOR_EMOJI.get(color, "")
            return f"{emoji} {folio_clean}".strip() if emoji else folio_clean

        def _display_a_folio(display_val):
            partes = str(display_val).strip().split(" ", 1)
            if len(partes) == 2 and partes[0] in _EMOJIS_FOLIO:
                return partes[1]
            return str(display_val).strip()

        def _display_a_color(display_val):
            partes = str(display_val).strip().split(" ", 1)
            if len(partes) == 2 and partes[0] in _EMOJIS_FOLIO:
                return _EMOJI_COLOR.get(partes[0], "")
            return ""

        # Costo del Vale de Consumo por OT — suma de costo_total de cada repuesto.
        # costo_total YA viene como cantidad x costo_unitario en el JSON, no se
        # vuelve a multiplicar (mismo criterio que los informes PDF y Facturas X).
        def _costo_vale_ot(reps):
            if not isinstance(reps, list):
                try:
                    reps = json.loads(reps) if isinstance(reps, str) else []
                except Exception:
                    reps = []
            _t = 0.0
            for _r in (reps or []):
                if not isinstance(_r, dict):
                    continue
                try:
                    _t += float(str(_r.get("costo_total", 0) or 0).replace(",", "."))
                except Exception:
                    pass
            # El peso chileno no tiene decimales: se redondea para que la tabla,
            # el total del encabezado y el Excel muestren exactamente lo mismo.
            return round(_t)

        COLS_FOLIO_RANGO = [c for c in ["FOLIO OT", "PATENTE", "RANGO"] if c in df.columns]
        COLS_ED  = [c for c in ["CATEGORIA", "OBSERVACION OT", "NOTAS", "AVANCE - GESTIÓN"] if c in df.columns]
        COLS_RO  = [c for c in ["SUCURSAL", "DIAS APERTURA", "FECHA OT",
                                 "TIPO VENTA", "TIPO CLIENTE", "MARCA", "MODELO",
                                 "ASESOR", "ESTADO", "NETO"]
                    if c in df.columns]
        # Por cada tipo de documento: primero el contador (N_) y a continuacion
        # sus folios reales (FOLIOS_), para poder leerlos uno al lado del otro.
        COLS_DOC = []
        for _, _k_doc in DOCS_CONFIG:
            if f"N_{_k_doc}" in df.columns:
                COLS_DOC.append(f"N_{_k_doc}")
            if f"FOLIOS_{_k_doc}" in df.columns:
                COLS_DOC.append(f"FOLIOS_{_k_doc}")
        ULTIMA   = ["ULTIMA_EDICION"] if "ULTIMA_EDICION" in df.columns else []
        COLS_ALL = COLS_FOLIO_RANGO + COLS_ED + COLS_RO + ULTIMA + COLS_DOC

        df_ed = df[COLS_ALL].copy().reset_index(drop=True)

        # Columna de costo, inmediatamente despues de NETO (o al final si no existe)
        _costos_vale = (df["repuestos_actual"].apply(_costo_vale_ot).tolist()
                        if "repuestos_actual" in df.columns else [0.0] * len(df_ed))
        _pos_costo = (df_ed.columns.get_loc("NETO") + 1) if "NETO" in df_ed.columns else len(df_ed.columns)
        df_ed.insert(_pos_costo, "COSTO_VALE", _costos_vale)

        if "_color_overrides" not in st.session_state:
            st.session_state["_color_overrides"] = {}

        _folios_limpios = df_ed["FOLIO OT"].astype(str).tolist()
        _colores_gh = (df["_MARCA_COLOR_"].fillna("").astype(str).tolist()
                       if "_MARCA_COLOR_" in df.columns else [""] * len(df_ed))
        _colores_cargados = []
        for _f, _c_gh in zip(_folios_limpios, _colores_gh):
            _override = st.session_state["_color_overrides"].get(_f)
            if _override is not None:
                _colores_cargados.append(_override if _override else "—")
            else:
                _colores_cargados.append(_c_gh if _c_gh else "—")

        if "RANGO" in df_ed.columns:
            df_ed["RANGO"] = df_ed["RANGO"].map(
                lambda r: f"{_RANGO_ICONS.get(r, '⚪')} {r}" if r else r
            )

        _folio_col_idx = df_ed.columns.get_loc("PATENTE" if "PATENTE" in df_ed.columns else "FOLIO OT") + 1
        df_ed.insert(_folio_col_idx, "COLOR", _colores_cargados)
        df_ed.insert(0, "_ir_", False)

        _b1, _b2 = st.columns([3, 2])
        with _b1:
            st.info(
                "✏️ **Columnas editables:** Categoría · Observación OT · Notas · Avance / Gestión  \n"
                "Son las únicas modificables. Guárdalas con **💾 Guardar cambios**."
            )
        with _b2:
            st.info(
                "🎨 **Colorear Folio OT:**  \n"
                "Haz clic en la celda del Folio → elige el color en el menú desplegable.  \n"
                "Se guarda automáticamente al seleccionar."
            )

        _SORT_OPTS = {
            "Folio OT (defecto)":                           ("FOLIO OT",        True),
            "Días apertura ↓  (más antiguos primero)":      ("DIAS APERTURA",   False),
            "Neto ↓  (mayor valor primero)":                ("NETO",            False),
            "Costo Vale de Consumo ↓  (mayor primero)":     ("COSTO_VALE",      False),
            "Sucursal  A→Z":                                ("SUCURSAL",        True),
            "Asesor  A→Z":                                  ("ASESOR",          True),
            "Categoría  A→Z":                               ("CATEGORIA",       True),
            "Última edición ↓  (más reciente primero)":     ("ULTIMA_EDICION",  False),
            "# Liquidaciones ↓":                            ("N_LIQ_ST",        False),
            "# Facturas Cliente ↓":                         ("N_FACT_CLIENTE",  False),
        }
        _srt_col1, _srt_col2 = st.columns([3, 2])
        with _srt_col1:
            _prev_sort = st.session_state.get("_prev_sort_select", "")
            _sel_sort = st.selectbox("🔀 Ordenar tabla por", list(_SORT_OPTS.keys()),
                                     key="_sort_select")
            if _sel_sort != _prev_sort and _prev_sort:
                st.session_state["det_page"] = 0
            st.session_state["_prev_sort_select"] = _sel_sort
        with _srt_col2:
            st.caption(" ")
            _tot_vale_lst = float(pd.to_numeric(df_ed.get("COSTO_VALE"), errors="coerce").fillna(0).sum()) \
                            if "COSTO_VALE" in df_ed.columns else 0.0
            st.caption(f"📋 Mostrando **{len(df_ed):,}** OTs en este listado.  \n"
                       f"📦 Costo Vale de Consumo: **${_tot_vale_lst:,.0f}**")
        _sort_col_name, _sort_ascending = _SORT_OPTS[_sel_sort]

        _cats_base = ["Sin categoría", "GARANTIA", "GARANTIA EXTENDIDA", "CLIENTE",
                      "SEGUROS", "INTERNO", "ENACOM", "PERDIDA TOTAL"]
        _cats_en_datos = sorted(set(
            v for v in df_raw["CATEGORIA"].fillna("").astype(str).unique()
            if v.strip() and v.strip().lower() not in ("nan", "none", "sin categoría", "sin categoria", "")
        ))
        _opciones_categoria = _cats_base + [c for c in _cats_en_datos if c not in _cats_base]

        cc = {
            "_ir_": st.column_config.CheckboxColumn("📄", width="small",
                                                     help="Marca para ir a Documentos y Comentarios"),
            "FOLIO OT":         st.column_config.TextColumn("Folio OT",       disabled=True, width="small"),
            "COLOR":            st.column_config.SelectboxColumn("🎨 Color",
                                    options=["—", "🔴 Rojo", "🟡 Amarillo", "🟢 Verde", "🔵 Azul"],
                                    width="small"),
            "RANGO":            st.column_config.TextColumn("Rango",           disabled=True, width="small"),
            "CATEGORIA":        st.column_config.SelectboxColumn("Categoría",  options=_opciones_categoria),
            "OBSERVACION OT":   st.column_config.TextColumn("Observación OT",  width="large"),
            "NOTAS":            st.column_config.TextColumn("Notas",            width="large"),
            "AVANCE - GESTIÓN": st.column_config.TextColumn("Avance / Gestión", width="large"),
            "SUCURSAL":         st.column_config.TextColumn("Sucursal",        disabled=True),
            "DIAS APERTURA":    st.column_config.NumberColumn("Días",          disabled=True, format="%d", width="small"),
            "FECHA OT":         st.column_config.TextColumn("Fecha OT",        disabled=True),
            "TIPO VENTA":       st.column_config.TextColumn("Tipo venta",      disabled=True),
            "TIPO CLIENTE":     st.column_config.TextColumn(
                "Tipo Cliente", disabled=True, width="medium",
                help="Tipo de cliente que trae el Seguimiento de Servicio Técnico "
                     "(columna BM del PBI): *PARTICULAR, *CURIFOR, *GARANTIA o la "
                     "compañía de seguro correspondiente (CIA-SEG …)."),
            "MARCA":            st.column_config.TextColumn("Marca",           disabled=True),
            "MODELO":           st.column_config.TextColumn("Modelo",          disabled=True),
            "PATENTE":          st.column_config.TextColumn("Patente",         disabled=True, width="small"),
            "ASESOR":           st.column_config.TextColumn("Asesor",          disabled=True),
            "ESTADO":           st.column_config.TextColumn("Estado",          disabled=True),
            "NETO":             st.column_config.NumberColumn("Neto",          disabled=True, format="$%d"),
            "COSTO_VALE":       st.column_config.NumberColumn("Costo Vale Consumo", disabled=True,
                                    format="$%d",
                                    help="Suma del costo de los repuestos del Vale de Consumo de la OT"),
            "ULTIMA_EDICION":   st.column_config.TextColumn("Última edición",  disabled=True),
        }
        for _nom_doc, key in DOCS_CONFIG:
            cc[f"N_{key}"] = st.column_config.NumberColumn(
                f"# {_nom_doc}", disabled=True, format="%d", width="small")
            cc[f"FOLIOS_{key}"] = st.column_config.TextColumn(
                f"Folios {_nom_doc}", disabled=True, width="medium",
                help=f"Número(s) de documento de {_nom_doc} asociados a la OT")

        _NUMERIC_SORT_COLS = {"DIAS APERTURA", "NETO", "COSTO_VALE"} | {f"N_{k}" for _, k in DOCS_CONFIG}
        if _sort_col_name in df_ed.columns:
            if _sort_col_name in _NUMERIC_SORT_COLS:
                _sort_key = pd.to_numeric(df_ed[_sort_col_name], errors="coerce").fillna(0)
                df_ed = df_ed.iloc[_sort_key.sort_values(ascending=_sort_ascending).index].reset_index(drop=True)
            else:
                df_ed = df_ed.sort_values(_sort_col_name, ascending=_sort_ascending).reset_index(drop=True)

        df_ed_pag = df_ed.copy().reset_index(drop=True)
        _det_cur_page = 0  # sin paginación, siempre página 0

        df_editado = st.data_editor(
            df_ed_pag, hide_index=True, use_container_width=True,
            height=500, column_config=cc, key="editor_principal",
        )

        def _norm_color(v):
            s = str(v).strip() if v is not None else ""
            return "" if s.lower() in ("none", "nan", "", "—", "-") else s

        _cambios_color = {}
        for _i in range(len(df_editado)):
            _color_nuevo = _norm_color(df_editado.iloc[_i].get("COLOR"))
            _color_orig  = _norm_color(df_ed_pag.iloc[_i].get("COLOR"))
            if _color_nuevo != _color_orig:
                _fc = _display_a_folio(str(df_editado.iloc[_i]["FOLIO OT"]))
                _cambios_color[_fc] = _color_nuevo
        if _cambios_color:
            _ok_c, _msg_c = guardar_colores_github(_cambios_color)
            if _ok_c:
                st.session_state["_color_overrides"].update(_cambios_color)
                st.toast("🎨 Color guardado", icon="✅")
            else:
                st.toast(f"⚠️ No se pudo guardar el color: {_msg_c}", icon="❌")

        _sel_rows = df_editado[df_editado["_ir_"] == True]
        if not _sel_rows.empty:
            _folio_nav_display = str(_sel_rows.iloc[0]["FOLIO OT"])
            _folio_nav_limpio  = _display_a_folio(_folio_nav_display)
            _nc1, _nc2, _nc3 = st.columns([2, 2, 1])
            with _nc1:
                st.info(f"📄 OT seleccionada: **{_folio_nav_limpio}**")
            with _nc2:
                st.caption("Presiona el botón para ver documentos y comentarios →")
            with _nc3:
                if st.button("📄 Ver Documentos", type="primary",
                             use_container_width=True, key="btn_nav_tab3"):
                    st.session_state["_nav_folio"] = _folio_nav_limpio
                    st.rerun()

        col_btn, col_info = st.columns([1, 3])
        with col_btn:
            if st.button("💾 Guardar cambios", type="primary", use_container_width=True):
                _df_save = df_editado.drop(columns=["_ir_", "COLOR"], errors="ignore")
                _df_orig = df_ed_pag.drop(columns=["_ir_", "COLOR"], errors="ignore")
                with st.spinner("Guardando en GitHub..."):
                    ok, msg = guardar_en_github(_df_save, _df_orig, usuario_activo)
                if ok:
                    st.success(msg)
                    # Solo se toco datos_dashboard.json — limpiar unicamente esa cache
                    # (no toda la app) para que el resto de usuarios no sufra un
                    # re-fetch innecesario de agenda/comentarios/stock/etc. en su
                    # proximo rerun. Antes esto usaba st.cache_data.clear() (borra
                    # TODO, cache compartida entre sesiones) y era la causa principal
                    # de la lentitud reportada al guardar (15/07/2026).
                    cargar_datos.clear()
                    st.rerun()
                else:
                    st.error(f"⚠️ {msg}  \n**Tus cambios NO fueron guardados.**")
        with col_info:
            st.caption("Categoría, Observación, Notas y Avance/Gestión se guardan en GitHub.")

        st.divider()
        _fecha_hoy = datetime.now(_TZ_CHILE).strftime('%Y%m%d')
        # Para el Excel descargamos TODAS las filas (no solo la página actual)
        _df_xlsx = df_ed.drop(columns=["_ir_", "COLOR"], errors="ignore").copy()
        _renombres_xlsx = {
            "FOLIO OT": "Folio OT", "RANGO": "Rango", "CATEGORIA": "Categoría",
            "OBSERVACION OT": "Observación OT", "NOTAS": "Notas",
            "AVANCE - GESTIÓN": "Avance / Gestión", "SUCURSAL": "Sucursal",
            "DIAS APERTURA": "Días apertura", "FECHA OT": "Fecha OT",
            "TIPO VENTA": "Tipo venta", "TIPO CLIENTE": "Tipo Cliente",
            "MARCA": "Marca", "MODELO": "Modelo",
            "PATENTE": "Patente", "ASESOR": "Asesor", "ESTADO": "Estado",
            "NETO": "Neto", "COSTO_VALE": "Costo Vale Consumo",
            "ULTIMA_EDICION": "Última edición",
        }
        for _nom_doc_x, _key in DOCS_CONFIG:
            _renombres_xlsx[f"N_{_key}"]      = f"# {_nom_doc_x}"
            _renombres_xlsx[f"FOLIOS_{_key}"] = f"Folios {_nom_doc_x}"
        # Los folios son numeros de documento: deben quedar como TEXTO para que
        # Excel no se coma los ceros a la izquierda (ej. vale "0000018346") ni
        # convierta un listado "123, 456" en algo raro.
        _cols_folios_xlsx = [f"FOLIOS_{k}" for _, k in DOCS_CONFIG if f"FOLIOS_{k}" in _df_xlsx.columns]
        for _c_f in _cols_folios_xlsx:
            _df_xlsx[_c_f] = _df_xlsx[_c_f].fillna("").astype(str)
        if "COSTO_VALE" in _df_xlsx.columns:
            _df_xlsx["COSTO_VALE"] = pd.to_numeric(_df_xlsx["COSTO_VALE"], errors="coerce").fillna(0)
        _df_xlsx = _df_xlsx.rename(columns={k: v for k, v in _renombres_xlsx.items() if k in _df_xlsx.columns})
        _buf = io.BytesIO()
        with pd.ExcelWriter(_buf, engine="openpyxl") as _writer:
            _df_xlsx.to_excel(_writer, index=False, sheet_name="OTs Pendientes")
            _ws = _writer.sheets["OTs Pendientes"]
            _hdrs_xlsx = [str(_c.value) if _c.value is not None else "" for _c in _ws[1]]
            for _idx_h, _nombre_h in enumerate(_hdrs_xlsx, start=1):
                _letra_h = _ws.cell(row=1, column=_idx_h).column_letter
                if _nombre_h in ("Neto", "Costo Vale Consumo"):
                    for _fila_h in range(2, _ws.max_row + 1):
                        _ws.cell(row=_fila_h, column=_idx_h).number_format = '"$"#,##0'
                elif _nombre_h.startswith("Folios "):
                    for _fila_h in range(2, _ws.max_row + 1):
                        _ws.cell(row=_fila_h, column=_idx_h).number_format = "@"
            for _col_cells in _ws.columns:
                _max_len = max((len(str(_cell.value)) if _cell.value is not None else 0) for _cell in _col_cells)
                _ws.column_dimensions[_col_cells[0].column_letter].width = min(_max_len + 4, 60)
            _ws.freeze_panes = "A2"
        _buf.seek(0)
        st.download_button("⬇️ Descargar como Excel", _buf,
                           f"OTs_Pendientes_{_fecha_hoy}.xlsx",
                           mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


    # ---- TAB 3: DOCUMENTOS Y COMENTARIOS ------------------------
    elif _tab == 2:
        st.markdown('<p class="section-title">Selecciona una OT para ver sus documentos y comentarios</p>',
                    unsafe_allow_html=True)

        # Usamos df_raw (todas las OT pendientes, sin filtros del sidebar) para que la
        # navegación desde Repuestos Pendientes / Patentes a Contactar siempre encuentre la OT.
        folios = sorted(df_raw["FOLIO OT"].astype(str).unique().tolist())
        folio_sel = st.selectbox("Folio OT", [""]+folios, key="sel_folio_docs")

        if folio_sel:
            filt = df_raw[df_raw["FOLIO OT"].astype(str)==folio_sel]
            if filt.empty:
                st.warning("OT no encontrada en los listados.")
            else:
                ot = filt.iloc[0]
                cc1,cc2,cc3,cc4,cc5 = st.columns(5)
                cc1.metric("Sucursal",  ot.get("SUCURSAL","—"))
                cc2.metric("Asesor",    ot.get("ASESOR","—"))
                cc3.metric("Días",      ot.get("DIAS APERTURA","—"))
                cc4.metric("Rango",     ot.get("RANGO","—"))
                _neto_raw = ot.get("NETO", "")
                try:
                    _neto_fmt = f"${int(float(_neto_raw)):,}" if _neto_raw and _neto_raw not in ("", "—") else "—"
                except Exception:
                    _neto_fmt = _neto_raw or "—"
                cc5.metric("Neto OT", _neto_fmt)
                _rut_ot = str(ot.get("rut_cliente", "") or "").strip()
                st.markdown(f"**Modelo:** {ot.get('MODELO','—')}  &nbsp;|&nbsp;  "
                            f"**Patente:** {ot.get('PATENTE','—')}  &nbsp;|&nbsp;  "
                            f"**Tipo:** {ot.get('TIPO VENTA','—')}"
                            + (f"  &nbsp;|&nbsp;  **RUT Cliente:** {_rut_ot}" if _rut_ot else ""))

                # ── Indicador Cuenta Ficha (Anticipo Taller) ─────────────────────
                _anticipo = ot.get("anticipo", {})
                if not isinstance(_anticipo, dict):
                    try:
                        _anticipo = json.loads(str(_anticipo)) if _anticipo else {}
                    except Exception:
                        _anticipo = {}
                _ant_tiene = _anticipo.get("tiene_saldo", False)
                _ant_total = _anticipo.get("total", 0) or 0
                _ant_nombre = _anticipo.get("nombre", "") or ""
                _ant_movs   = _anticipo.get("movimientos", []) or []
                if _ant_tiene:
                    try:
                        _ant_total_fmt = f"${int(_ant_total):,}"
                    except Exception:
                        _ant_total_fmt = f"${_ant_total}"
                    st.markdown(f"""
    <div style="background:linear-gradient(135deg,#1a4a2e,#22863a);color:white;
                border-radius:12px;padding:14px 20px;margin:10px 0 4px 0;
                display:flex;align-items:center;gap:18px;box-shadow:0 2px 8px rgba(0,0,0,.25);">
        <div style="font-size:2.2rem;line-height:1;">💰</div>
        <div>
            <div style="font-size:0.78rem;font-weight:600;letter-spacing:.06em;
                        text-transform:uppercase;opacity:.85;margin-bottom:2px;">
                Cuenta Ficha — Saldo disponible en anticipo taller
            </div>
            <div style="font-size:0.88rem;opacity:.9;">{_ant_nombre or '—'}
                {(' · RUT: ' + _rut_ot) if _rut_ot else ''}</div>
            <div style="font-size:1.8rem;font-weight:800;letter-spacing:-.01em;margin-top:4px;">
                {_ant_total_fmt}
            </div>
        </div>
    </div>""", unsafe_allow_html=True)
                    if _ant_movs:
                        with st.expander(f"📋 Ver detalle de movimientos ({len(_ant_movs)} línea(s))", expanded=False):
                            _df_movs = pd.DataFrame(_ant_movs)
                            if "saldo" in _df_movs.columns:
                                _df_movs["saldo"] = pd.to_numeric(_df_movs["saldo"], errors="coerce")
                            _df_movs = _df_movs.rename(columns={
                                "documento": "Documento", "nro": "N°", "saldo": "Monto ($)",
                                "fecha": "Fecha", "local": "Sucursal / Local", "glosa": "Glosa",
                            })
                            _cols_mv = [c for c in ["Documento","N°","Monto ($)","Fecha","Sucursal / Local","Glosa"]
                                        if c in _df_movs.columns]
                            st.dataframe(_df_movs[_cols_mv], hide_index=True, use_container_width=True,
                                         column_config={
                                             "Monto ($)": st.column_config.NumberColumn(format="$%,.0f", width="medium"),
                                             "Glosa": st.column_config.TextColumn(width="large"),
                                             "Sucursal / Local": st.column_config.TextColumn(width="large"),
                                             "Documento": st.column_config.TextColumn(width="medium"),
                                         })
                            _total_movs = _df_movs["Monto ($)"].sum() if "Monto ($)" in _df_movs.columns else 0
                            st.markdown(f"<div style='text-align:right;font-weight:600;'>Total: ${_total_movs:,.0f}</div>",
                                        unsafe_allow_html=True)
                elif _anticipo:
                    # Hay datos pero saldo = 0
                    st.markdown(
                        "<div style='background:#f8f0e3;border-left:4px solid #d69e2e;"
                        "border-radius:6px;padding:10px 14px;margin:8px 0 4px 0;font-size:0.88rem;color:#7b4f00;'>"
                        "⚠️ <strong>Sin saldo en Cuenta Ficha</strong> — el cliente no tiene anticipos disponibles."
                        "</div>", unsafe_allow_html=True)
                # Si no hay datos de anticipo, no se muestra nada (cliente sin cruce en el archivo)

                st.divider()

                col_docs, col_gestion = st.columns([3, 2])

                with col_docs:
                    st.markdown('<p class="section-title">Documentos registrados</p>', unsafe_allow_html=True)
                    docs_rows = []
                    for tipo, key in DOCS_CONFIG:
                        n        = int(ot.get(f"N_{key}", 0))
                        fols_doc = str(ot.get(f"FOLIOS_{key}","")).strip()
                        docs_rows.append({"Tipo de Documento": tipo, "Cantidad": n,
                                          "Número(s) / Folio(s)": fols_doc if fols_doc else "—"})
                    df_docs = pd.DataFrame(docs_rows)
                    st.dataframe(df_docs, hide_index=True, use_container_width=True,
                                 column_config={
                                     "Cantidad": st.column_config.NumberColumn(format="%d", width="small"),
                                     "Tipo de Documento": st.column_config.TextColumn(width="medium"),
                                     "Número(s) / Folio(s)": st.column_config.TextColumn(width="large"),
                                 })
                    st.caption(f"Total documentos: **{df_docs['Cantidad'].sum()}**")

                    _reps_actual    = ot.get("repuestos_actual", [])
                    _reps_historico = ot.get("repuestos_historico", [])
                    if _reps_actual:
                        _vales_unicos = list(dict.fromkeys(r["vale"] for r in _reps_actual if r.get("vale")))
                        st.markdown('<p style="margin-top:14px;font-size:0.83rem;color:#888;">🔧 Repuestos — haz clic para ver el detalle:</p>', unsafe_allow_html=True)
                        for _vale in _vales_unicos:
                            _items = [r for r in _reps_actual if r.get("vale") == _vale]
                            _costo_vale = sum(float(r.get("costo_total", 0) or 0) for r in _items)
                            with st.expander(f"📦 Vale {_vale}  —  {len(_items)} ítem(s)  ·  ${_costo_vale:,.0f}", expanded=False):
                                _df_v = pd.DataFrame(_items).rename(columns={
                                    "producto": "Código", "descripcion": "Descripción",
                                    "cantidad": "Cantidad", "costo_unitario": "Costo Unitario", "costo_total": "Costo Total"})
                                for _c in ["Cantidad", "Costo Unitario", "Costo Total"]:
                                    if _c in _df_v.columns:
                                        _df_v[_c] = pd.to_numeric(_df_v[_c], errors="coerce")
                                _cols_v = [c for c in ["Código","Descripción","Cantidad","Costo Unitario","Costo Total"] if c in _df_v.columns]
                                st.dataframe(_df_v[_cols_v], hide_index=True, use_container_width=True,
                                             column_config={
                                                 "Código": st.column_config.TextColumn(width="medium"),
                                                 "Descripción": st.column_config.TextColumn(width="large"),
                                                 "Cantidad": st.column_config.NumberColumn(format="%g", width="small"),
                                                 "Costo Unitario": st.column_config.NumberColumn(format="$%,.0f", width="medium"),
                                                 "Costo Total": st.column_config.NumberColumn(format="$%,.0f", width="medium"),
                                             })
                                st.markdown(f"<div style='text-align:right;font-weight:600;'>Total vale: ${_costo_vale:,.0f}</div>", unsafe_allow_html=True)

                with col_gestion:
                    st.markdown('<p class="section-title">Estado de gestión</p>', unsafe_allow_html=True)
                    st.markdown(f"**Categoría:** {ot.get('CATEGORIA','—')}")
                    st.markdown(f"**Observación OT:** {ot.get('OBSERVACION OT','—')}")
                    st.markdown(f"**Notas:** {ot.get('NOTAS','—')}")
                    st.markdown(f"**Avance / Gestión:** {ot.get('AVANCE - GESTIÓN','—')}")
                    if ot.get("ULTIMA_EDICION","").strip():
                        st.caption(f"Última edición: {ot['ULTIMA_EDICION']}")

                st.divider()

                # ============================================================
                #   REPUESTOS — SEGUIMIENTO DE COMPRAS (en espera / en bodega)
                # ============================================================
                st.markdown('<p class="section-title">🔧 Repuestos (Seguimiento de Compras)</p>',
                            unsafe_allow_html=True)
                st.caption("Arriba (📦 Vale de Consumo) ves lo que ya está **cargado** en la OT. "
                           "Aquí ves lo que falta por consumir: 🟢 lo que **ya está en bodega** "
                           "(pendiente de instalar) y 🔴 lo que **aún no llega a bodega** "
                           "(pendiente de llegar).")

                _reps_compras = ot.get("repuestos_compras", [])
                if not isinstance(_reps_compras, list):
                    _reps_compras = []
                _reps_espera  = [r for r in _reps_compras if not r.get("en_bodega")]
                _reps_bodega  = [r for r in _reps_compras if r.get("en_bodega")]

                def _tabla_repuestos_compras(_lista, _con_fecha):
                    """Construye y muestra la tabla detalle de repuestos de compras."""
                    _rows = []
                    for _r in _lista:
                        _fila = {
                            "OT Origen":            _r.get("ot_origen", ""),
                            "Vía":                  "Patente" if _r.get("via_patente") else "Directo",
                            "Origen":               _r.get("origen", ""),
                            "Producto":             _r.get("producto", ""),
                            "Descripción Producto": _r.get("descripcion", ""),
                            "Bodega":               _r.get("bodega", "") or "—",
                            "Stock":                pd.to_numeric(_r.get("stock", None), errors="coerce"),
                            "Cantidad":             pd.to_numeric(_r.get("cantidad", ""), errors="coerce"),
                            "Costo":                pd.to_numeric(_r.get("costo", None), errors="coerce"),
                            "Estado":               _r.get("estado", ""),
                        }
                        if _con_fecha:
                            _fila["Fecha en bodega"] = _r.get("fecha_bodega", "") or "—"
                        _rows.append(_fila)
                    _df_r = pd.DataFrame(_rows)
                    _cfg = {
                        "OT Origen":            st.column_config.TextColumn(
                            width="small", help="OT donde se pidió el repuesto (puede estar cerrada)"),
                        "Vía":                  st.column_config.TextColumn(
                            width="small", help="Directo o enlazado por patente desde una OT cerrada"),
                        "Origen":               st.column_config.TextColumn(width="small"),
                        "Producto":             st.column_config.TextColumn(width="medium"),
                        "Descripción Producto": st.column_config.TextColumn(width="large"),
                        "Bodega":               st.column_config.TextColumn(
                            width="medium", help="Bodega donde se encuentra el repuesto (Stock repuestos)"),
                        "Stock":                st.column_config.NumberColumn(
                            format="%g", width="small",
                            help="Cantidad disponible en bodega según archivo de Stock"),
                        "Cantidad":             st.column_config.NumberColumn(format="%g", width="small"),
                        "Costo":                st.column_config.NumberColumn(
                            format="$%,.0f", width="medium",
                            help="Costo unitario real del repuesto (Stock Repuestos Costos)"),
                        "Estado":               st.column_config.TextColumn(width="medium"),
                        "Fecha en bodega":      st.column_config.TextColumn(width="small"),
                    }
                    st.dataframe(_df_r, hide_index=True, use_container_width=True, column_config=_cfg)
                    _tot = _df_r["Costo"].fillna(0).sum()
                    st.markdown(
                        f"<div style='text-align:right;font-weight:600;'>Costo total: ${_tot:,.0f}</div>",
                        unsafe_allow_html=True,
                    )

                if not _reps_compras:
                    # ⚪ SEMÁFORO 3 — sin repuestos en espera
                    st.markdown(
                        "<div style='background:#f1f3f5;border-left:5px solid #adb5bd;"
                        "border-radius:8px;padding:12px 16px;'>"
                        "⚪ <b>Sin espera de repuestos</b> — esta OT no tiene repuestos "
                        "asociados en el Seguimiento de Compras.</div>",
                        unsafe_allow_html=True,
                    )
                else:
                    # 🟢 SEMÁFORO — repuestos en bodega pendientes de instalar
                    if _reps_bodega:
                        st.markdown(
                            "<div style='background:#e6f4ea;border-left:5px solid #38a169;"
                            "border-radius:8px;padding:12px 16px;margin-bottom:8px;'>"
                            f"🟢 <b>En bodega — pendientes de instalar</b> — {len(_reps_bodega)} "
                            "repuesto(s) ya llegaron a bodega y faltan por consumir en la OT.</div>",
                            unsafe_allow_html=True,
                        )
                        with st.expander(f"🟢 Ver {len(_reps_bodega)} repuesto(s) en bodega", expanded=True):
                            _tabla_repuestos_compras(_reps_bodega, _con_fecha=True)

                    # 🔴 SEMÁFORO — repuestos pendientes de llegar a bodega
                    if _reps_espera:
                        st.markdown(
                            "<div style='background:#fdeaea;border-left:5px solid #e53e3e;"
                            "border-radius:8px;padding:12px 16px;margin-bottom:8px;'>"
                            f"🔴 <b>Pendientes de llegar a bodega</b> — {len(_reps_espera)} "
                            "repuesto(s) aún no llegan (Pendiente / Abierto con referencia).</div>",
                            unsafe_allow_html=True,
                        )
                        with st.expander(f"🔴 Ver {len(_reps_espera)} repuesto(s) pendiente(s) de llegar", expanded=True):
                            _tabla_repuestos_compras(_reps_espera, _con_fecha=False)

                st.divider()

                df_coms = cargar_comentarios()
                coms_ot = pd.DataFrame()
                if not df_coms.empty and "folio_ot" in df_coms.columns:
                    coms_ot = df_coms[df_coms["folio_ot"].astype(str)==str(folio_sel)].copy()
                    if "fecha" in coms_ot.columns:
                        coms_ot = coms_ot.sort_values("fecha", ascending=False)

                col_form, col_hist = st.columns([2, 3])

                with col_form:
                    st.markdown('<p class="section-title">➕ Agregar comentario</p>', unsafe_allow_html=True)

                    # Lista de usuarios para @mencionar
                    _todos_usuarios = cargar_usuarios_cache()
                    _emails_otros = ["(Ninguno)"] + sorted([
                        u["email"] for u in _todos_usuarios
                        if u.get("email", "").lower() != usuario_activo.lower()
                        and u.get("activo", True)
                    ])

                    # st.form evita que el selectbox dispare un rerun al cambiar,
                    # el rerun solo ocurre al presionar el botón de submit.
                    with st.form(key=f"form_com_{folio_sel}", clear_on_submit=True):
                        _mencionados_sel = st.multiselect(
                            "📣 Mencionar a (opcional):",
                            _emails_otros,
                            help="Todas las personas seleccionadas recibirán una notificación en la app."
                        )
                        nuevo_comentario = st.text_area(
                            "Escribe el avance o comentario",
                            placeholder="Ej: Se contactó al cliente el día de hoy...",
                            height=130,
                        )
                        _submitted = st.form_submit_button(
                            "💬 Guardar comentario",
                            type="primary",
                            use_container_width=True,
                        )

                    if _submitted:
                        if not nuevo_comentario.strip():
                            st.error("El comentario no puede estar vacío.")
                        else:
                            with st.spinner("Guardando..."):
                                ok, msg = agregar_comentario_github(
                                    folio_sel, usuario_activo, nuevo_comentario, _mencionados_sel
                                )
                            if ok:
                                st.success(msg)
                                if _mencionados_sel:
                                    st.info(f"🔔 Notificación enviada a: {', '.join(_mencionados_sel)}")
                                # Solo se tocaron comentarios (y notificaciones si hubo
                                # menciones) — no limpiar toda la cache compartida (ver
                                # nota junto a "Guardar cambios" en Detalle y Edicion).
                                cargar_comentarios.clear()
                                if _mencionados_sel:
                                    cargar_notificaciones_cache.clear()
                            else:
                                st.error(msg)

                with col_hist:
                    st.markdown('<p class="section-title">🕐 Historial de esta OT</p>', unsafe_allow_html=True)
                    if coms_ot.empty:
                        st.info("Esta OT aún no tiene comentarios registrados.")
                    else:
                        for _, row in coms_ot.iterrows():
                            _men = row.get("mencionado", "")
                            if isinstance(_men, list):
                                _men = ", ".join(str(x) for x in _men if x)
                            elif _men is None or (isinstance(_men, float) and pd.isna(_men)):
                                _men = ""
                            else:
                                _men = str(_men)
                            _men_html = f' &nbsp;·&nbsp; 📣 <b>{_men}</b>' if _men.strip() else ""
                            st.markdown(f"""
                            <div class="comentario-card">
                                <div class="comentario-meta">
                                    🗓 {row.get('fecha','—')} &nbsp;·&nbsp; 👤 {row.get('autor','—')}{_men_html}
                                </div>
                                <div class="comentario-texto">{row.get('comentario','')}</div>
                            </div>
                            """, unsafe_allow_html=True)
        else:
            st.info("Selecciona un Folio OT arriba para ver sus documentos y comentarios.")


    # ---- TAB 4: HISTORIAL DE COMENTARIOS -------------------------
    elif _tab == 3:
        st.markdown('<p class="section-title">Sábana completa de comentarios y avances por OT</p>',
                    unsafe_allow_html=True)
        df_hist = cargar_comentarios()
        if df_hist.empty:
            st.info("Aún no hay comentarios registrados.")
        else:
            hc1, hc2, hc3 = st.columns(3)
            with hc1:
                folios_hist  = ["Todos"] + sorted(df_hist["folio_ot"].astype(str).unique().tolist())
                filtro_folio = st.selectbox("Filtrar por Folio OT", folios_hist, key="hist_folio")
            with hc2:
                autores_hist = ["Todos"] + sorted(df_hist["autor"].astype(str).unique().tolist()) if "autor" in df_hist.columns else ["Todos"]
                filtro_autor = st.selectbox("Filtrar por autor", autores_hist, key="hist_autor")
            with hc3:
                busq_hist = st.text_input("Buscar en comentarios", "", key="hist_busq")

            df_mostrar = df_hist.copy()
            if filtro_folio != "Todos":
                df_mostrar = df_mostrar[df_mostrar["folio_ot"].astype(str)==filtro_folio]
            if filtro_autor != "Todos":
                df_mostrar = df_mostrar[df_mostrar["autor"].astype(str)==filtro_autor]
            if busq_hist:
                df_mostrar = df_mostrar[
                    df_mostrar["comentario"].astype(str).str.contains(busq_hist, case=False, na=False)
                ]
            if "fecha" in df_mostrar.columns:
                df_mostrar = df_mostrar.sort_values("fecha", ascending=False)
            st.caption(f"Mostrando {len(df_mostrar):,} de {len(df_hist):,} comentarios")
            st.dataframe(
                df_mostrar.rename(columns={"folio_ot": "Folio OT", "autor": "Autor",
                                           "fecha": "Fecha", "comentario": "Comentario",
                                           "mencionado": "Mencionado"}),
                hide_index=True, use_container_width=True, height=500,
                column_config={
                    "Folio OT":   st.column_config.TextColumn(width="small"),
                    "Autor":      st.column_config.TextColumn(width="medium"),
                    "Fecha":      st.column_config.TextColumn(width="medium"),
                    "Mencionado": st.column_config.TextColumn(width="medium"),
                    "Comentario": st.column_config.TextColumn(width="large"),
                },
            )
            csv_hist = df_mostrar.to_csv(index=False).encode("utf-8-sig")
            st.download_button("⬇️ Descargar historial como CSV", csv_hist,
                               f"Historial_Comentarios_{datetime.now(_TZ_CHILE).strftime('%Y%m%d')}.csv",
                               mime="text/csv", use_container_width=True)


    # ---- TAB 5: POR SUCURSAL ------------------------------------
    elif _tab == 4:
        st.markdown('<p class="section-title">Resumen por Sucursal</p>', unsafe_allow_html=True)
        if df.empty:
            st.info("Sin datos con los filtros actuales.")
        else:
            por_suc = (df.groupby("SUCURSAL").agg(Total=("FOLIO OT","count"))
                       .reset_index().sort_values("Total", ascending=False))
            for col_name in ["0-30","31-60","61-90","91 o más"]:
                sub = df[df["RANGO"]==col_name].groupby("SUCURSAL").size().rename(col_name)
                por_suc = por_suc.merge(sub, on="SUCURSAL", how="left").fillna(0)
                por_suc[col_name] = por_suc[col_name].astype(int)
            st.dataframe(por_suc, hide_index=True, use_container_width=True)
            st.bar_chart(por_suc.set_index("SUCURSAL")["Total"])


    # ---- TAB 6: POR ASESOR --------------------------------------
    elif _tab == 5:
        st.markdown('<p class="section-title">Resumen por Asesor</p>', unsafe_allow_html=True)
        if df.empty:
            st.info("Sin datos con los filtros actuales.")
        else:
            por_asesor = (df.groupby("ASESOR").agg(Total=("FOLIO OT","count"))
                          .reset_index().sort_values("Total", ascending=False))
            st.dataframe(por_asesor, hide_index=True, use_container_width=True)
            st.bar_chart(por_asesor.set_index("ASESOR")["Total"])

            st.divider()
            st.markdown('<p class="section-title">📑 Informe de OTs pendientes por Asesor</p>',
                        unsafe_allow_html=True)
            _as_en_listado = sorted([a for a in df["ASESOR"].unique() if str(a).strip()])
            _as_informe = list(sel_asesor) if sel_asesor else _as_en_listado
            # Si cambian los filtros, el archivo generado antes deja de ofrecerse
            # (evita descargar un informe que ya no corresponde a lo que se ve).
            _sig_informe = (tuple(sorted(_as_informe)), len(df))
            _ic1, _ic2 = st.columns([3, 2])
            with _ic1:
                if sel_asesor:
                    st.caption(
                        f"Se generará con el filtro de Asesor del panel izquierdo: "
                        f"**{' · '.join(sel_asesor)}** — {len(df):,} OT(s) pendientes.  \n"
                        "Todos los asesores van **consolidados en un solo archivo** "
                        "(una misma persona puede figurar con más de un nombre de usuario)."
                    )
                else:
                    st.caption(
                        f"⚠️ No hay ningún asesor filtrado: el informe saldrá con los "
                        f"**{len(_as_en_listado)}** asesores del listado actual "
                        f"({len(df):,} OT(s)).  \nPara acotarlo, elige uno o varios en "
                        "el filtro **Asesor** del panel izquierdo."
                    )
                st.caption(
                    "Contiene: **Resumen** (indicadores globales y desgloses por asesor, "
                    "sucursal, rango, tipo de venta, categoría y marca), **Listado OTs** "
                    "(folio, neto, costo del Vale de Consumo y los folios de cada documento) "
                    "y **Documentos** (una fila por documento posterior)."
                )
            with _ic2:
                st.caption(" ")
                if st.button("📑 Generar informe por Asesor (Excel)", type="primary",
                             use_container_width=True, key="btn_informe_asesor"):
                    with st.spinner("Generando informe..."):
                        try:
                            _xlsx_as = generar_excel_asesor(df, _as_informe, fecha_actualizacion)
                            st.session_state["_xlsx_asesor"] = _xlsx_as
                            st.session_state["_xlsx_asesor_n"] = len(df)
                            st.session_state["_xlsx_asesor_sig"] = _sig_informe
                        except Exception as _e_as:
                            st.session_state.pop("_xlsx_asesor", None)
                            st.error(f"No se pudo generar el informe: {_e_as}")
                if (st.session_state.get("_xlsx_asesor")
                        and st.session_state.get("_xlsx_asesor_sig") == _sig_informe):
                    if len(_as_informe) == 1:
                        _slug = "".join(ch if ch.isalnum() else "_"
                                        for ch in str(_as_informe[0]))[:24].strip("_")
                    else:
                        _slug = f"{len(_as_informe)}_asesores"
                    _nom_arch = (f"Informe_Asesor_{_slug}_"
                                 f"{datetime.now(_TZ_CHILE).strftime('%Y%m%d_%H%M')}.xlsx")
                    st.download_button(
                        f"⬇️ Descargar informe ({st.session_state.get('_xlsx_asesor_n', 0):,} OTs)",
                        st.session_state["_xlsx_asesor"], _nom_arch,
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        use_container_width=True, key="dl_informe_asesor")

            # ---- Texto de correo (solo admin) ----------------------------
            if usuario_activo == ADMIN_EMAIL:
                st.divider()
                st.markdown('<p class="section-title">✉️ Correo de urgencia — OTs sobre 90 días</p>',
                            unsafe_allow_html=True)
                _n90_vista = int((df["RANGO"] == "91 o más").sum())
                if not sel_asesor:
                    st.warning(
                        "El correo es para **una persona**: filtra primero al asesor "
                        "(o sus alias) en el panel izquierdo. Sin filtro, el texto "
                        "mezclaría las OTs de los "
                        f"{len(_as_en_listado)} asesores del listado."
                    )
                else:
                    # El texto se arma con `df`, que YA viene filtrado por el asesor
                    # seleccionado en el panel izquierdo — nombre, sucursal, conteos,
                    # montos y folios salen todos de SUS OTs, no del total.
                    _n90_gar = int((
                        (df["RANGO"] == "91 o más")
                        & (df["TIPO VENTA"].astype(str).str.upper().str.contains("GARANT", na=False)
                           | df["CATEGORIA"].astype(str).str.upper().str.contains("GARANTIA", na=False)
                           | (df["CATEGORIA"].astype(str).str.upper() == "SUBIR"))
                    ).sum())
                    st.caption(
                        f"Datos de: **{' · '.join(sel_asesor)}** — "
                        f"{len(df):,} OT(s) pendientes, de las cuales **{_n90_vista}** "
                        f"superan los 90 días y **{_n90_gar}** son garantías. "
                        "El texto se arma con esas mismas OTs."
                    )
                    st.caption(
                        "ℹ️ Se cuenta como garantía si el **Tipo de Venta** dice GARANTÍA o "
                        "si la **Categoría** es GARANTIA / GARANTIA EXTENDIDA / SUBIR "
                        "(mismo criterio del Informe por Área)."
                    )
                    # Nombre de pila del asesor filtrado (los alias son de la misma persona)
                    _nom_def = str(_as_informe[0]).split()[0].title() if _as_informe else ""
                    _mc1, _mc2, _mc3 = st.columns([2, 2, 2])
                    with _mc1:
                        _nom_mail = st.text_input("Nombre para el saludo", value=_nom_def,
                                                  placeholder="Ej: Rodrigo",
                                                  key="mail_nombre_asesor")
                    with _mc2:
                        # Plazo por defecto: 3 días, corrido al lunes si cae fin de semana
                        _plazo_def = (datetime.now(_TZ_CHILE) + _timedelta(days=3)).date()
                        if _plazo_def.weekday() >= 5:
                            _plazo_def = _plazo_def + _timedelta(days=7 - _plazo_def.weekday())
                        _lim_mail = st.date_input(
                            "Plazo exigido", value=_plazo_def,
                            key="mail_plazo_asesor", format="DD/MM/YYYY")
                    with _mc3:
                        st.caption(" ")
                        _inc_fol = st.checkbox("Incluir folios en el texto", value=True,
                                               key="mail_folios_asesor")
                        _gen_mail = st.button("✉️ Generar texto del correo",
                                              use_container_width=True, key="btn_texto_correo")
                    if _gen_mail:
                        try:
                            _asunto_m, _cuerpo_m = generar_texto_correo_asesor(
                                df, _nom_mail, fecha_actualizacion, _lim_mail,
                                incluir_folios=_inc_fol)
                            st.session_state["_mail_asesor"] = (_asunto_m, _cuerpo_m)
                            st.session_state["_mail_asesor_sig"] = _sig_informe
                        except Exception as _e_m:
                            st.session_state.pop("_mail_asesor", None)
                            st.error(f"No se pudo generar el texto: {_e_m}")
                    if (st.session_state.get("_mail_asesor")
                            and st.session_state.get("_mail_asesor_sig") == _sig_informe):
                        _asunto_m, _cuerpo_m = st.session_state["_mail_asesor"]
                        st.caption("**Asunto** — usa el ícono de copiar de la esquina 📋")
                        st.code(_asunto_m, language=None)
                        st.caption("**Cuerpo del correo** — cópialo y pégalo en Outlook")
                        st.code(_cuerpo_m, language=None)
                        st.caption("💡 Adjunta el informe Excel de más arriba para que el "
                                   "asesor tenga el detalle OT por OT.")


    # ---- TAB 7: ANÁLISIS Y TENDENCIAS ---------------------------
    elif _tab == 6:
        st.markdown('<p class="section-title">📊 Carga actual por Asesor — Distribución por Rango</p>',
                    unsafe_allow_html=True)
        if df.empty:
            st.info("Sin datos con los filtros actuales.")
        else:
            _rangos_orden = ["0-30", "31-60", "61-90", "91 o más"]
            _df_ar = (df[df["ASESOR"] != ""].groupby(["ASESOR", "RANGO"]).size().unstack(fill_value=0))
            for _r in _rangos_orden:
                if _r not in _df_ar.columns: _df_ar[_r] = 0
            _df_ar = _df_ar[_rangos_orden].copy()
            _df_ar["Total"] = _df_ar[_rangos_orden].sum(axis=1)
            _df_ar = _df_ar.sort_values("Total", ascending=False).reset_index()
            _col_tabla, _col_chart = st.columns([2, 3])
            with _col_tabla:
                st.dataframe(_df_ar, hide_index=True, use_container_width=True, height=420,
                             column_config={
                                 "ASESOR": st.column_config.TextColumn("Asesor"),
                                 "0-30": st.column_config.NumberColumn("🟢 0-30",  format="%d"),
                                 "31-60": st.column_config.NumberColumn("🟡 31-60", format="%d"),
                                 "61-90": st.column_config.NumberColumn("🟠 61-90", format="%d"),
                                 "91 o más": st.column_config.NumberColumn("🔴 91+", format="%d"),
                                 "Total": st.column_config.NumberColumn("Total",    format="%d"),
                             })
            with _col_chart:
                st.caption("OTs por asesor según rango de días")
                st.bar_chart(_df_ar.set_index("ASESOR")[_rangos_orden], use_container_width=True, height=380)
        st.divider()
        st.markdown('<p class="section-title">⚠️ Proyección — OTs próximas a escalar de rango</p>',
                    unsafe_allow_html=True)
        st.caption("Las OTs en 61-90 días pasarán a Críticas (>90) si no se resuelven.")
        if not df.empty:
            _cols_proyeccion = [c for c in ["FOLIO OT","SUCURSAL","ASESOR","DIAS APERTURA","CATEGORIA","OBSERVACION OT"] if c in df.columns]
            _prox_criticas = df[df["RANGO"] == "61-90"].copy()
            _prox_urgentes = df[df["RANGO"] == "31-60"].copy()
            _pc1, _pc2 = st.columns(2)
            with _pc1:
                st.markdown(
                    f"<div style='background:#fff3f3;border-left:4px solid #e53e3e;border-radius:6px;padding:10px 14px;margin-bottom:10px;'>"
                    f"<b style='color:#e53e3e'>🔴 Próximamente Críticas</b><br>"
                    f"<span style='font-size:1.8rem;font-weight:700;color:#e53e3e'>{len(_prox_criticas)}</span> "
                    f"<span style='color:#666;font-size:0.85rem'> OTs en rango 61-90 días</span></div>",
                    unsafe_allow_html=True)
                if not _prox_criticas.empty:
                    st.dataframe(_prox_criticas[_cols_proyeccion].sort_values("DIAS APERTURA", ascending=False),
                                 hide_index=True, use_container_width=True, height=280)
                else:
                    st.success("✅ No hay OTs en rango 61-90 días.")
            with _pc2:
                st.markdown(
                    f"<div style='background:#fff8f0;border-left:4px solid #dd6b20;border-radius:6px;padding:10px 14px;margin-bottom:10px;'>"
                    f"<b style='color:#dd6b20'>🟠 Próximamente Urgentes</b><br>"
                    f"<span style='font-size:1.8rem;font-weight:700;color:#dd6b20'>{len(_prox_urgentes)}</span> "
                    f"<span style='color:#666;font-size:0.85rem'> OTs en rango 31-60 días</span></div>",
                    unsafe_allow_html=True)
                if not _prox_urgentes.empty:
                    st.dataframe(_prox_urgentes[_cols_proyeccion].sort_values("DIAS APERTURA", ascending=False),
                                 hide_index=True, use_container_width=True, height=280)
                else:
                    st.success("✅ No hay OTs en rango 31-60 días.")
        st.divider()
        st.markdown('<p class="section-title">📉 Historial de OTs Cerradas por Actualización</p>',
                    unsafe_allow_html=True)
        _df_cierres_resumen, _registros_cierres = cargar_historial_cierres()
        if _df_cierres_resumen.empty:
            st.info("⏳ Aún no hay registros de cierres.")
        else:
            _total_cerradas_hist = _df_cierres_resumen["OTs Cerradas"].sum()
            _total_nuevas_hist   = _df_cierres_resumen["OTs Nuevas"].sum()
            _ultima_act          = _df_cierres_resumen["Fecha"].iloc[-1] if not _df_cierres_resumen.empty else "—"
            _hk1, _hk2, _hk3, _hk4 = st.columns(4)
            with _hk1: kpi_card("Total registros", len(_df_cierres_resumen))
            with _hk2: kpi_card("OTs cerradas (acum.)", int(_total_cerradas_hist), "verde")
            with _hk3: kpi_card("OTs nuevas (acum.)", int(_total_nuevas_hist))
            with _hk4:
                st.markdown(
                    f"<div class='kpi-box'><p class='kpi-label'>Última actualización</p>"
                    f"<p style='font-size:0.95rem;font-weight:600;color:#1a3a5c;margin:4px 0'>{_ultima_act}</p></div>",
                    unsafe_allow_html=True)
            st.markdown("#### Evolución de cierres por actualización")
            st.line_chart(_df_cierres_resumen.set_index("Fecha")[["OTs Cerradas","OTs Nuevas","OTs Activas"]],
                          use_container_width=True, height=280)
            st.markdown("#### Detalle por actualización")
            st.dataframe(_df_cierres_resumen.sort_values("Fecha", ascending=False),
                         hide_index=True, use_container_width=True)
            st.markdown("#### OTs cerradas por registro")
            for _reg in reversed(_registros_cierres[-10:]):
                _fecha_reg  = _reg.get("fecha", "—")
                _n_cerradas = _reg.get("total_cerradas", 0)
                _ots_list   = _reg.get("ots_cerradas", [])
                _label      = f"🗓 {_fecha_reg} — {_n_cerradas} OT(s) cerrada(s)"
                if _ots_list:
                    with st.expander(_label, expanded=False):
                        st.dataframe(pd.DataFrame(_ots_list), hide_index=True, use_container_width=True)
                else:
                    st.caption(f"{_label} (sin detalle)")
            _csv_cierres = _df_cierres_resumen.to_csv(index=False).encode("utf-8-sig")
            st.download_button("⬇️ Descargar historial de cierres como CSV", _csv_cierres,
                               f"Historial_Cierres_{datetime.now().strftime('%Y%m%d')}.csv", mime="text/csv")


    # ---- TAB 7: RANKING SEMANAL OTs QUE CRUZAN 90 DÍAS --------
    elif _tab == 7:
        from datetime import timedelta as _td

        st.markdown('<p class="section-title">🏆 Ranking semanal — OTs que cruzan los 90 días (lunes a lunes)</p>',
                    unsafe_allow_html=True)
        st.caption(
            "Muestra cuántas OTs cruzaron la barrera de los 90 días de apertura en cada semana, "
            "calculado desde **Días de Apertura** de cada OT."
        )

        # ── Preparar base: todas las OTs >90d con fecha_cruce calculada ──
        _df90 = df_raw[df_raw["RANGO"] == "91 o más"].copy()
        _hoy  = datetime.now(_TZ_CHILE).replace(tzinfo=None)

        _df90["_dias_num"] = pd.to_numeric(_df90["DIAS APERTURA"], errors="coerce")
        _df90 = _df90.dropna(subset=["_dias_num"])
        _df90["_apertura"]    = _df90["_dias_num"].apply(lambda d: _hoy - _td(days=int(d)))
        _df90["_cruce_90"]    = _df90["_apertura"] + _td(days=90)
        # Lunes de la semana en que cruzó los 90 días
        _df90["_lunes"]       = _df90["_cruce_90"] - pd.to_timedelta(_df90["_cruce_90"].dt.weekday, unit="d")
        _df90["_lunes"]       = _df90["_lunes"].dt.normalize()
        _df90["Semana"]       = _df90["_lunes"].dt.strftime("Lun %d/%m/%Y")

        _total90 = len(_df90)
        _suc90   = _df90["SUCURSAL"].nunique()
        _as90    = _df90["ASESOR"].nunique()

        # ── KPIs ──────────────────────────────────────────────────────
        _k1, _k2, _k3, _k4 = st.columns(4)
        with _k1: kpi_card("OTs >90 días pendientes", _total90, "rojo")
        with _k2: kpi_card("Sucursales afectadas", _suc90)
        with _k3: kpi_card("Asesores involucrados", _as90)
        with _k4:
            _sem_top = _df90.groupby("Semana").size().idxmax() if _total90 > 0 else "—"
            _sem_top_n = int(_df90.groupby("Semana").size().max()) if _total90 > 0 else 0
            st.markdown(
                f"<div class='kpi-box rojo'>"
                f"<p class='kpi-label'>Semana con más cruces</p>"
                f"<p style='font-size:0.85rem;font-weight:700;color:#e53e3e;margin:4px 0'>{_sem_top}</p>"
                f"<p class='kpi-label'>{_sem_top_n} OTs esa semana</p></div>",
                unsafe_allow_html=True,
            )

        st.divider()

        if _total90 == 0:
            st.success("✅ No hay OTs pendientes con más de 90 días.")
        else:
            # ── Selector de ventana de semanas ────────────────────────
            _ventana = st.radio(
                "Período a mostrar",
                ["Últimas 4 semanas", "Últimas 8 semanas", "Últimas 12 semanas", "Todo el historial"],
                horizontal=True, key="rk_ventana",
            )
            _n_sem = {"Últimas 4 semanas": 4, "Últimas 8 semanas": 8,
                      "Últimas 12 semanas": 12, "Todo el historial": 9999}[_ventana]

            # Ordenar semanas y filtrar
            _semanas_ord = sorted(_df90["_lunes"].unique())
            _semanas_sel = _semanas_ord[-_n_sem:]
            _df90_v = _df90[_df90["_lunes"].isin(_semanas_sel)].copy()

            # ── Vista: Sucursal o Asesor ──────────────────────────────
            _rk_vista = st.radio(
                "Desglose por",
                ["🏢 Sucursal", "👤 Asesor"],
                horizontal=True, key="rk_vista",
            )

            _grupo = "SUCURSAL" if _rk_vista == "🏢 Sucursal" else "ASESOR"

            # Si vista asesor, permitir filtrar por sucursal
            if _grupo == "ASESOR":
                _suc_fil = st.selectbox(
                    "Filtrar por Sucursal",
                    ["Todas"] + sorted(_df90_v["SUCURSAL"].unique().tolist()),
                    key="rk_suc_as_fil",
                )
                if _suc_fil != "Todas":
                    _df90_v = _df90_v[_df90_v["SUCURSAL"] == _suc_fil]

            # ── Tabla pivot: semana × grupo ───────────────────────────
            _pv_sem = (
                _df90_v.groupby(["Semana", "_lunes", _grupo])
                .size()
                .reset_index(name="OTs")
            )
            _pv_wide = (
                _pv_sem.pivot_table(index=["Semana", "_lunes"], columns=_grupo, values="OTs", fill_value=0)
                .reset_index()
                .sort_values("_lunes", ascending=False)
            )
            _pv_wide.insert(2, "Total semana", _pv_wide.iloc[:, 2:].sum(axis=1))
            _pv_wide = _pv_wide.drop(columns=["_lunes"])
            _pv_wide.columns.name = None

            st.markdown(f'<p class="section-title">OTs que cruzaron los 90 días — por semana y {_grupo.lower()}</p>',
                        unsafe_allow_html=True)
            st.dataframe(
                _pv_wide,
                hide_index=True,
                use_container_width=True,
                height=min(500, 36 + 36 * len(_pv_wide)),
                column_config={"Semana": st.column_config.TextColumn(width="medium"),
                               "Total semana": st.column_config.NumberColumn(width="small", format="%d")},
            )

            st.divider()

            # ── Gráfico stacked: semana × grupo ──────────────────────
            st.markdown(f'<p class="section-title">Evolución semanal por {_grupo.lower()}</p>',
                        unsafe_allow_html=True)
            _chart_data = (
                _df90_v.groupby(["_lunes", _grupo]).size().unstack(fill_value=0).sort_index()
            )
            _chart_data.index = _chart_data.index.strftime("Lun %d/%m")
            # Limitar a top 10 en el eje para legibilidad
            _top10 = _df90_v.groupby(_grupo).size().sort_values(ascending=False).head(10).index
            _chart_data = _chart_data[[c for c in _top10 if c in _chart_data.columns]]
            st.bar_chart(_chart_data, use_container_width=True, height=320)

            st.divider()

            # ── Ranking acumulado del período ─────────────────────────
            st.markdown(f'<p class="section-title">Ranking acumulado del período — por {_grupo.lower()} y Tipo de Venta</p>',
                        unsafe_allow_html=True)
            _pv_acum = (
                _df90_v.groupby([_grupo, "TIPO VENTA"]).size().unstack(fill_value=0)
            )
            _pv_acum["Total"] = _pv_acum.sum(axis=1)
            _pv_acum = _pv_acum.sort_values("Total", ascending=False).reset_index()
            _pv_acum.insert(0, "#", range(1, len(_pv_acum) + 1))
            _pv_acum.columns.name = None
            st.dataframe(_pv_acum, hide_index=True, use_container_width=True,
                         height=min(450, 36 + 36 * len(_pv_acum)))

            # ── Descarga ──────────────────────────────────────────────
            _csv_rk = _df90_v[
                [c for c in ["FOLIO OT", "SUCURSAL", "ASESOR", "DIAS APERTURA",
                              "TIPO VENTA", "MARCA", "PATENTE", "CATEGORIA", "Semana"]
                 if c in _df90_v.columns]
            ].sort_values(["Semana", _grupo]).to_csv(index=False).encode("utf-8-sig")
            st.download_button(
                "⬇️ Descargar detalle CSV",
                _csv_rk,
                f"Ranking_90d_{datetime.now(_TZ_CHILE).strftime('%Y%m%d')}.csv",
                mime="text/csv",
            )


    # ---- TAB 9: NOTIFICACIONES ----------------------------------
    elif _tab == 8:
        st.markdown('<p class="section-title">🔔 Mis Notificaciones</p>', unsafe_allow_html=True)

        _mis_notifs = [n for n in _notifs_raw if n.get("destinatario", "").lower() == usuario_activo.lower()]
        _mis_notifs_sorted = sorted(_mis_notifs, key=lambda x: x.get("fecha", ""), reverse=True)

        if not _mis_notifs_sorted:
            st.info("No tienes notificaciones aún. Cuando alguien te mencione en un comentario, aparecerá aquí.")
        else:
            _no_leidas = [n for n in _mis_notifs_sorted if not n.get("leida")]
            _leidas    = [n for n in _mis_notifs_sorted if n.get("leida")]

            if _no_leidas:
                st.markdown(f"**{len(_no_leidas)} notificación(es) sin leer:**")
                for _n in _no_leidas:
                    st.markdown(f"""
                    <div class="notif-card">
                        <div class="comentario-meta">
                            🗓 {_n.get('fecha','—')} &nbsp;·&nbsp; 👤 {_n.get('remitente','—')}
                            &nbsp;·&nbsp; 📋 OT <b>{_n.get('folio_ot','—')}</b>
                        </div>
                        <div class="comentario-texto">{_n.get('extracto','')[:200]}</div>
                    </div>
                    """, unsafe_allow_html=True)
                    if st.button(f"📄 Ir a OT {_n.get('folio_ot','')}",
                                 key=f"notif_ir_{_n.get('id','')}"):
                        st.session_state["_nav_folio"] = _n.get("folio_ot", "")
                        st.rerun()

                if st.button("✅ Marcar todas como leídas", type="secondary"):
                    with st.spinner("Actualizando..."):
                        _marcar_leidas(usuario_activo)
                    # Solo se toco notificaciones.json — no limpiar toda la cache
                    # compartida (ver nota junto a "Guardar cambios" en Detalle y
                    # Edicion).
                    cargar_notificaciones_cache.clear()
                    st.success("Notificaciones marcadas como leídas.")
            else:
                st.success("✅ Sin notificaciones pendientes.")

            if _leidas:
                with st.expander(f"📂 Ver {len(_leidas)} notificación(es) ya leída(s)", expanded=False):
                    for _n in _leidas[:30]:
                        st.markdown(f"""
                        <div class="notif-card leida">
                            <div class="comentario-meta">
                                🗓 {_n.get('fecha','—')} &nbsp;·&nbsp; 👤 {_n.get('remitente','—')}
                                &nbsp;·&nbsp; 📋 OT {_n.get('folio_ot','—')}
                            </div>
                            <div class="comentario-texto" style="color:#999">{_n.get('extracto','')[:150]}</div>
                        </div>
                        """, unsafe_allow_html=True)


    # ---- TAB: REPUESTOS PENDIENTES ------------------------------
    elif _tab == 9:
        import io as _io
        from openpyxl import Workbook as _Workbook
        from openpyxl.styles import (PatternFill as _PFill, Font as _Font,
                                      Alignment as _Align, Border as _Border, Side as _Side)
        from openpyxl.utils import get_column_letter as _gcl

        st.markdown('<p class="section-title">🧰 Repuestos Pendientes — en bodega</p>',
                    unsafe_allow_html=True)

        # ── Filtro de fecha de corte (01/01/2025) ─────────────────────────────
        # FECHA OT en el PBI no es fiable (suele mostrar la fecha de extracción).
        # Usamos DIAS APERTURA: fecha_apertura = fecha_extracción - DIAS_APERTURA.
        # Calculamos el máximo de días permitidos desde la fecha de actualización del JSON.
        _FECHA_CORTE_REP = pd.Timestamp("2025-01-01")
        _fecha_extraccion = pd.to_datetime(
            fecha_actualizacion.split()[0] if fecha_actualizacion else "",
            dayfirst=True, errors="coerce"
        )
        if pd.isna(_fecha_extraccion):
            _fecha_extraccion = pd.Timestamp.now()
        # Días máximos permitidos: diferencia entre fecha extracción y corte
        _MAX_DIAS_APERTURA = (_fecha_extraccion - _FECHA_CORTE_REP).days

        def _dias_ok(_ot_row):
            """True si la OT abrió desde 01/01/2025 según DIAS APERTURA."""
            _d = pd.to_numeric(_ot_row.get("DIAS APERTURA", ""), errors="coerce")
            if pd.isna(_d):
                return False  # sin dato → excluir
            return int(_d) <= _MAX_DIAS_APERTURA

        # Pre-construir lookup de dias de apertura para OTs (para filtrar ot_origen via patente)
        _mapa_dias_ot = {}
        for _, _r_raw in df_raw.iterrows():
            _f_raw = str(_r_raw.get("FOLIO OT", ""))
            _d_raw = _r_raw.get("DIAS APERTURA", "")
            if _f_raw:
                _mapa_dias_ot[_f_raw] = _d_raw

        st.info(
            f"📅 Mostrando OTs abiertas desde **01/01/2025** (≤ {_MAX_DIAS_APERTURA} días de apertura) "
            f"y repuestos en bodega desde esa fecha en adelante.",
            icon="🔍",
        )

        # ── Construir datos completos (siempre incluye espera para stats) ──────
        _rows_rp_all = []
        _ot_bodega = _ot_espera = 0
        _costo_bodega = _costo_espera = 0.0
        _stock_total = 0.0
        for _, _ot in df.iterrows():
            # ── Filtro por fecha de apertura de la OT Pendiente (via DIAS APERTURA) ──
            if not _dias_ok(_ot):
                continue
            _reps = _ot.get("repuestos_compras", [])
            if not isinstance(_reps, list) or not _reps:
                continue
            _hb = [r for r in _reps if r.get("en_bodega")]
            _he = [r for r in _reps if not r.get("en_bodega")]
            if _hb:
                _ot_bodega += 1
                _costo_bodega += sum((pd.to_numeric(r.get("costo", None) if r.get("costo") is not None else r.get("total", ""), errors="coerce") or 0) for r in _hb)
                _stock_total  += sum((r.get("stock") or 0) for r in _hb if r.get("stock") is not None)
            if _he:
                _ot_espera += 1
                _costo_espera += sum((pd.to_numeric(r.get("costo", None) if r.get("costo") is not None else r.get("total", ""), errors="coerce") or 0) for r in _he)
            for _r in _reps:
                # ── Filtro por fecha de ot_origen (via DIAS APERTURA) ──
                _ot_orig = str(_r.get("ot_origen", ""))
                if _ot_orig and _ot_orig in _mapa_dias_ot:
                    _d_orig = _mapa_dias_ot[_ot_orig]
                    _d_orig_n = pd.to_numeric(_d_orig, errors="coerce")
                    if pd.notna(_d_orig_n) and int(_d_orig_n) > _MAX_DIAS_APERTURA:
                        continue
                # ── Filtro por fecha_bodega (repuestos en bodega desde antes de 2025) ──
                _fbod = str(_r.get("fecha_bodega", "") or "")
                if _fbod and _fbod not in ("—", "nan", "None", ""):
                    _fbod_dt = pd.to_datetime(_fbod, dayfirst=True, errors="coerce")
                    if pd.notna(_fbod_dt) and _fbod_dt < _FECHA_CORTE_REP:
                        continue
                _rows_rp_all.append({
                    "OT Pendiente":         str(_ot.get("FOLIO OT", "")),
                    "Sucursal":             str(_ot.get("SUCURSAL", "")),
                    "Patente":              _r.get("patente", "") or str(_ot.get("PATENTE", "")),
                    "OT Origen":            _r.get("ot_origen", ""),
                    "Via":                  "Patente" if _r.get("via_patente") else "Directo",
                    "Producto":             _r.get("producto", ""),
                    "Descripcion Producto": _r.get("descripcion", ""),
                    "Bodega":               _r.get("bodega", "") or "—",
                    "Stock":                pd.to_numeric(_r.get("stock", None), errors="coerce"),
                    "Cantidad":             pd.to_numeric(_r.get("cantidad", ""), errors="coerce"),
                    "Costo":                pd.to_numeric(_r.get("costo", None) if _r.get("costo") is not None else _r.get("total", ""), errors="coerce"),
                    "Estado":               _r.get("estado", ""),
                    "Fecha en bodega":      _r.get("fecha_bodega", "") or "—",
                    "Origen":               _r.get("origen", ""),
                    "en_bodega":            bool(_r.get("en_bodega")),
                })

        _df_rp_all = pd.DataFrame(_rows_rp_all) if _rows_rp_all else pd.DataFrame()
        _df_bod_all = _df_rp_all[_df_rp_all["en_bodega"]] if not _df_rp_all.empty else pd.DataFrame()
        _df_esp_all = _df_rp_all[~_df_rp_all["en_bodega"]] if not _df_rp_all.empty else pd.DataFrame()

        # ── Indicadores ───────────────────────────────────────────────────────
        _ki1, _ki2, _ki3 = st.columns(3)
        _ki1.metric("OTs con repuestos en bodega",   f"{_ot_bodega:,}")
        _ki2.metric("OTs con repuestos en espera",   f"{_ot_espera:,}")
        _ki3.metric("Costo en bodega sin consumo",   f"${_costo_bodega:,.0f}")

        _ki4, _ki5, _ki6 = st.columns(3)
        _ki4.metric("Líneas en bodega",              f"{len(_df_bod_all):,}")
        _ki5.metric("Stock total disponible",        f"{_stock_total:,.0f} un.")
        _ki6.metric("Costo repuestos en espera",     f"${_costo_espera:,.0f}")

        # ── Desgloses por sucursal y origen ───────────────────────────────────
        if not _df_bod_all.empty:
            _col_suc, _col_ori = st.columns(2)

            with _col_suc:
                st.markdown('<p class="section-title">Por sucursal (en bodega)</p>',
                            unsafe_allow_html=True)
                _by_suc = (
                    _df_bod_all.groupby("Sucursal")
                    .agg(OTs=("OT Pendiente", "nunique"),
                         Lineas=("Producto", "count"),
                         Valor=("Costo", "sum"),
                         Stock=("Stock", "sum"))
                    .reset_index()
                    .sort_values("Valor", ascending=False)
                )
                _by_suc["Valor"] = _by_suc["Valor"].apply(lambda x: f"${x:,.0f}")
                _by_suc["Stock"] = _by_suc["Stock"].apply(lambda x: f"{x:,.0f}")
                st.dataframe(_by_suc, hide_index=True, use_container_width=True)

            with _col_ori:
                st.markdown('<p class="section-title">Por origen (en bodega)</p>',
                            unsafe_allow_html=True)
                _by_ori = (
                    _df_bod_all.groupby("Origen")
                    .agg(OTs=("OT Pendiente", "nunique"),
                         Lineas=("Producto", "count"),
                         Valor=("Costo", "sum"),
                         Stock=("Stock", "sum"))
                    .reset_index()
                    .sort_values("Valor", ascending=False)
                )
                _by_ori["Valor"] = _by_ori["Valor"].apply(lambda x: f"${x:,.0f}")
                _by_ori["Stock"] = _by_ori["Stock"].apply(lambda x: f"{x:,.0f}")
                st.dataframe(_by_ori, hide_index=True, use_container_width=True)

        st.divider()

        # ── Tabla detalle ─────────────────────────────────────────────────────
        _incluir_espera = st.checkbox(
            "Incluir también repuestos en espera (aún no llegan a bodega)", value=False
        )

        _rows_rp = []
        for _, _ot in df.iterrows():
            # ── Filtro por fecha de apertura de la OT Pendiente (via DIAS APERTURA) ──
            if not _dias_ok(_ot):
                continue
            _reps = _ot.get("repuestos_compras", [])
            if not isinstance(_reps, list) or not _reps:
                continue
            for _r in _reps:
                _enb = bool(_r.get("en_bodega"))
                if not _enb and not _incluir_espera:
                    continue
                # ── Filtro por fecha de ot_origen (via DIAS APERTURA) ──
                _ot_orig = str(_r.get("ot_origen", ""))
                if _ot_orig and _ot_orig in _mapa_dias_ot:
                    _d_orig = _mapa_dias_ot[_ot_orig]
                    _d_orig_n = pd.to_numeric(_d_orig, errors="coerce")
                    if pd.notna(_d_orig_n) and int(_d_orig_n) > _MAX_DIAS_APERTURA:
                        continue
                # ── Filtro por fecha_bodega (repuestos en bodega desde antes de 2025) ──
                _fbod = str(_r.get("fecha_bodega", "") or "")
                if _fbod and _fbod not in ("—", "nan", "None", ""):
                    _fbod_dt = pd.to_datetime(_fbod, dayfirst=True, errors="coerce")
                    if pd.notna(_fbod_dt) and _fbod_dt < _FECHA_CORTE_REP:
                        continue
                _rows_rp.append({
                    "OT Pendiente":         str(_ot.get("FOLIO OT", "")),
                    "Sucursal":             str(_ot.get("SUCURSAL", "")),
                    "Patente":              _r.get("patente", "") or str(_ot.get("PATENTE", "")),
                    "OT Origen":            _r.get("ot_origen", ""),
                    "Vía":                  "Patente" if _r.get("via_patente") else "Directo",
                    "Producto":             _r.get("producto", ""),
                    "Descripción Producto": _r.get("descripcion", ""),
                    "Bodega":               _r.get("bodega", "") or "—",
                    "Stock":                pd.to_numeric(_r.get("stock", None), errors="coerce"),
                    "Cantidad":             pd.to_numeric(_r.get("cantidad", ""), errors="coerce"),
                    "Costo":                pd.to_numeric(_r.get("costo", None) if _r.get("costo") is not None else _r.get("total", ""), errors="coerce"),
                    "Estado":               _r.get("estado", ""),
                    "Fecha en bodega":      _r.get("fecha_bodega", "") or "—",
                    "Origen":               _r.get("origen", ""),
                })

        if not _rows_rp:
            st.info("No hay repuestos en bodega para el filtro actual del sidebar "
                    "(o aún no se ha corrido la consolidación).")
        else:
            _df_rp = pd.DataFrame(_rows_rp)

            # ── Controles de búsqueda y ordenamiento ────────────────────────────
            _ctrl1, _ctrl2 = st.columns([3, 2])
            with _ctrl1:
                _q = st.text_input("🔍 Buscar (folio, patente, producto, descripción...)",
                                   "", key="rp_busq")
            with _ctrl2:
                _sort_rp_opts = {
                    "Costo ↓":           ("Costo",            False),
                    "OT Pendiente ↑":    ("OT Pendiente",     True),
                    "Fecha bodega ↓":    ("Fecha en bodega",  False),
                    "Sucursal A→Z":      ("Sucursal",         True),
                    "Patente A→Z":       ("Patente",          True),
                }
                _sel_sort_rp = st.selectbox("🔀 Ordenar por", list(_sort_rp_opts.keys()),
                                            key="rp_sort_sel")

            if _q:
                _mask = pd.Series(False, index=_df_rp.index)
                for _c in ["OT Pendiente", "Patente", "OT Origen", "Producto", "Descripción Producto"]:
                    _mask |= _df_rp[_c].astype(str).str.contains(_q, case=False, na=False)
                _df_rp = _df_rp[_mask]

            # Aplicar orden
            _sort_col_rp, _sort_asc_rp = _sort_rp_opts[_sel_sort_rp]
            if _sort_col_rp in _df_rp.columns:
                if _sort_col_rp == "Costo":
                    _df_rp = _df_rp.iloc[
                        pd.to_numeric(_df_rp[_sort_col_rp], errors="coerce").fillna(0)
                        .sort_values(ascending=_sort_asc_rp).index
                    ].reset_index(drop=True)
                else:
                    _df_rp = _df_rp.sort_values(_sort_col_rp, ascending=_sort_asc_rp).reset_index(drop=True)

            # ── Resumen de filas ────────────────────────────────────────────────
            _rp_total_lines = len(_df_rp)
            st.caption(
                f"**{_rp_total_lines:,}** líneas "
                f"&nbsp;·&nbsp; costo total **${_df_rp['Costo'].fillna(0).sum():,.0f}**"
            )

            _df_rp_pag = _df_rp.copy().reset_index(drop=True)
            _df_rp_pag.insert(0, "_ir_", False)

            _cfg_rp = {
                "_ir_": st.column_config.CheckboxColumn(
                    "📄", width="small", help="Marca para ver Documentos y Comentarios de la OT pendiente"),
                "OT Pendiente":         st.column_config.TextColumn("OT Pend.", disabled=True, width="small"),
                "Sucursal":             st.column_config.TextColumn("Sucursal", disabled=True, width="small"),
                "Patente":              st.column_config.TextColumn("Patente",  disabled=True, width="small"),
                "OT Origen":            st.column_config.TextColumn(
                    "OT Orig.", disabled=True, width="small",
                    help="OT donde se solicitó el repuesto (puede estar cerrada)"),
                "Vía":                  st.column_config.TextColumn(
                    "Vía", disabled=True, width="small",
                    help="Directo: la OT pendiente pidió el repuesto. "
                         "Patente: enlazado desde OT cerrada del mismo vehículo."),
                "Producto":             st.column_config.TextColumn("Código",   disabled=True, width="medium"),
                "Descripción Producto": st.column_config.TextColumn("Descripción", disabled=True, width="large"),
                "Bodega":               st.column_config.TextColumn(
                    "Bodega", disabled=True, width="medium",
                    help="Bodega donde se encuentra el repuesto"),
                "Stock":                st.column_config.NumberColumn(
                    "Stock", disabled=True, format="%g", width="small"),
                "Cantidad":             st.column_config.NumberColumn("Cant.", disabled=True, format="%g", width="small"),
                "Costo":                st.column_config.NumberColumn(
                    "Costo $", disabled=True, format="$%,.0f", width="medium",
                    help="Costo unitario real del repuesto (Stock Repuestos Costos)"),
                "Estado":               st.column_config.TextColumn("Estado",  disabled=True, width="medium"),
                "Fecha en bodega":      st.column_config.TextColumn("F. Bodega", disabled=True, width="small"),
                "Origen":               st.column_config.TextColumn("Origen",  disabled=True, width="small"),
            }

            _edited_rp = st.data_editor(
                _df_rp_pag, hide_index=True, use_container_width=True, height=460,
                column_config=_cfg_rp, key="editor_rp",
            )

            _sel_rp = _edited_rp[_edited_rp["_ir_"] == True]
            if not _sel_rp.empty:
                _folio_dest = str(_sel_rp.iloc[0]["OT Pendiente"])
                _cA, _cB = st.columns([3, 1])
                with _cA:
                    st.info(f"OT pendiente seleccionada: **{_folio_dest}**")
                with _cB:
                    if st.button("📄 Ver Documentos y Comentarios", type="primary",
                                 use_container_width=True, key="btn_rp_nav"):
                        st.session_state["_nav_folio"] = _folio_dest
                        st.rerun()

            st.divider()

            # ── Descargas ─────────────────────────────────────────────────────
            _dcol1, _dcol2 = st.columns(2)

            with _dcol1:
                _csv_rp = _df_rp.drop(columns=["_ir_"], errors="ignore").to_csv(index=False).encode("utf-8-sig")
                st.download_button("⬇️ Descargar como CSV", _csv_rp,
                                   f"Repuestos_Pendientes_{datetime.now(_TZ_CHILE).strftime('%Y%m%d')}.csv",
                                   mime="text/csv", use_container_width=True)

            with _dcol2:
                # ── Generar Excel con Resumen + Detalle ───────────────────────
                def _generar_excel_repuestos(_df_det, _df_bod, _by_suc_raw, _by_ori_raw,
                                             _kpis, _fecha_str):
                    """Genera un .xlsx con hoja Resumen y hoja Detalle."""
                    _wb = _Workbook()

                    # ── Estilos comunes ──
                    _hdr_fill   = _PFill(patternType="solid", fgColor="1A3A5C")
                    _hdr_font   = _Font(bold=True, color="FFFFFF", size=11)
                    _sub_fill   = _PFill(patternType="solid", fgColor="D6E4F0")
                    _sub_font   = _Font(bold=True, color="1A3A5C", size=10)
                    _kpi_fill   = _PFill(patternType="solid", fgColor="F0F4F8")
                    _kpi_font   = _Font(bold=True, color="1A3A5C", size=14)
                    _lbl_font   = _Font(color="555555", size=9)
                    _thin        = _Side(style="thin", color="CCCCCC")
                    _brd         = _Border(left=_thin, right=_thin, top=_thin, bottom=_thin)
                    _ctr         = _Align(horizontal="center", vertical="center", wrap_text=True)
                    _lft         = _Align(horizontal="left",  vertical="center", wrap_text=True)

                    def _header(ws, row, col, txt, fill=None, font=None, align=None):
                        c = ws.cell(row=row, column=col, value=txt)
                        if fill:  c.fill  = fill
                        if font:  c.font  = font
                        if align: c.alignment = align
                        c.border = _brd
                        return c

                    # ════════════════════════════════════════════════
                    #   HOJA 1 — RESUMEN
                    # ════════════════════════════════════════════════
                    _ws1 = _wb.active
                    _ws1.title = "Resumen"
                    _ws1.column_dimensions["A"].width = 32
                    _ws1.column_dimensions["B"].width = 18
                    _ws1.column_dimensions["C"].width = 18
                    _ws1.column_dimensions["D"].width = 18
                    _ws1.column_dimensions["E"].width = 18

                    # Título
                    _ws1.merge_cells("A1:E1")
                    _tc = _ws1["A1"]
                    _tc.value = f"Informe Repuestos Pendientes — Curifor S.A  |  {_fecha_str}"
                    _tc.fill  = _PFill(patternType="solid", fgColor="1A3A5C")
                    _tc.font  = _Font(bold=True, color="FFFFFF", size=13)
                    _tc.alignment = _ctr
                    _ws1.row_dimensions[1].height = 28

                    # Fecha generación
                    _ws1.merge_cells("A2:E2")
                    _ws1["A2"].value = f"Generado: {datetime.now(_TZ_CHILE).strftime('%d/%m/%Y %H:%M')}"
                    _ws1["A2"].font  = _Font(italic=True, color="888888", size=9)
                    _ws1["A2"].alignment = _ctr

                    _filt_txt = []
                    if sel_sucursal: _filt_txt.append(f"Sucursal: {', '.join(sel_sucursal)}")
                    if sel_rango:    _filt_txt.append(f"Rango: {', '.join(sel_rango)}")
                    if sel_tipo:     _filt_txt.append(f"Tipo: {', '.join(sel_tipo)}")
                    if sel_marca:    _filt_txt.append(f"Marca: {', '.join(sel_marca)}")
                    if _periodo_ot_activo: _filt_txt.append(f"Período: {fecha_ot_desde.strftime('%d/%m/%Y')} al {fecha_ot_hasta.strftime('%d/%m/%Y')}")
                    if busqueda:     _filt_txt.append(f"Búsqueda: {busqueda}")
                    if _filt_txt:
                        _ws1.merge_cells("A3:E3")
                        _ws1["A3"].value     = "Filtros activos: " + " · ".join(_filt_txt)
                        _ws1["A3"].font      = _Font(italic=True, color="CC5500", size=9)
                        _ws1["A3"].alignment = _ctr
                        _row = 5
                    else:
                        _row = 4

                    # ── KPIs ──
                    _ws1.merge_cells(f"A{_row}:E{_row}")
                    _ws1[f"A{_row}"].value = "Indicadores de Resumen"
                    _ws1[f"A{_row}"].fill  = _sub_fill
                    _ws1[f"A{_row}"].font  = _sub_font
                    _ws1[f"A{_row}"].alignment = _ctr
                    _row += 1

                    _kpi_pairs = [
                        ("OTs con repuestos en bodega",  _kpis["ot_bodega"]),
                        ("OTs con repuestos en espera",  _kpis["ot_espera"]),
                        ("Líneas en bodega",             _kpis["lineas_bod"]),
                        ("Stock total disponible",       _kpis["stock"]),
                        ("Costo en bodega sin consumo",  _kpis["costo_bod"]),
                        ("Costo repuestos en espera",    _kpis["costo_esp"]),
                    ]
                    for _ki, (_lbl, _val) in enumerate(_kpi_pairs):
                        _r_k = _row + _ki
                        _ws1.row_dimensions[_r_k].height = 22
                        _ws1.cell(_r_k, 1, _lbl).font = _lbl_font
                        _ws1.cell(_r_k, 1).alignment  = _lft
                        _vc = _ws1.cell(_r_k, 2, _val)
                        _vc.font      = _Font(bold=True, color="1A3A5C", size=11)
                        _vc.fill      = _kpi_fill
                        _vc.alignment = _ctr
                        _vc.border    = _brd
                    _row += len(_kpi_pairs) + 1

                    # ── Por sucursal ──
                    _ws1.merge_cells(f"A{_row}:E{_row}")
                    _ws1[f"A{_row}"].value = "Resumen por Sucursal (en bodega)"
                    _ws1[f"A{_row}"].fill  = _sub_fill
                    _ws1[f"A{_row}"].font  = _sub_font
                    _ws1[f"A{_row}"].alignment = _ctr
                    _row += 1

                    for _col_i, _col_n in enumerate(["Sucursal","OTs","Líneas","Valor","Stock"], 1):
                        _header(_ws1, _row, _col_i, _col_n, fill=_hdr_fill, font=_hdr_font, align=_ctr)
                    _row += 1

                    for _, _rw in _by_suc_raw.iterrows():
                        _ws1.cell(_row, 1, _rw["Sucursal"]).alignment = _lft
                        _ws1.cell(_row, 2, int(_rw["OTs"])).alignment  = _ctr
                        _ws1.cell(_row, 3, int(_rw["Lineas"])).alignment = _ctr
                        _vc = _ws1.cell(_row, 4, _rw["Valor"])
                        _vc.number_format = '$#,##0'
                        _vc.alignment = _ctr
                        _vc2 = _ws1.cell(_row, 5, _rw["Stock"])
                        _vc2.number_format = '#,##0'
                        _vc2.alignment = _ctr
                        for _ci in range(1, 6):
                            _ws1.cell(_row, _ci).border = _brd
                        _row += 1
                    _row += 1

                    # ── Por origen ──
                    _ws1.merge_cells(f"A{_row}:E{_row}")
                    _ws1[f"A{_row}"].value = "Resumen por Origen (en bodega)"
                    _ws1[f"A{_row}"].fill  = _sub_fill
                    _ws1[f"A{_row}"].font  = _sub_font
                    _ws1[f"A{_row}"].alignment = _ctr
                    _row += 1

                    for _col_i, _col_n in enumerate(["Origen","OTs","Líneas","Valor","Stock"], 1):
                        _header(_ws1, _row, _col_i, _col_n, fill=_hdr_fill, font=_hdr_font, align=_ctr)
                    _row += 1

                    for _, _rw in _by_ori_raw.iterrows():
                        _ws1.cell(_row, 1, _rw["Origen"]).alignment = _lft
                        _ws1.cell(_row, 2, int(_rw["OTs"])).alignment  = _ctr
                        _ws1.cell(_row, 3, int(_rw["Lineas"])).alignment = _ctr
                        _vc = _ws1.cell(_row, 4, _rw["Valor"])
                        _vc.number_format = '$#,##0'
                        _vc.alignment = _ctr
                        _vc2 = _ws1.cell(_row, 5, _rw["Stock"])
                        _vc2.number_format = '#,##0'
                        _vc2.alignment = _ctr
                        for _ci in range(1, 6):
                            _ws1.cell(_row, _ci).border = _brd
                        _row += 1

                    # ════════════════════════════════════════════════
                    #   HOJA 2 — DETALLE
                    # ════════════════════════════════════════════════
                    _ws2 = _wb.create_sheet("Detalle")

                    _det_cols = [c for c in _df_det.columns if c != "_ir_"]
                    _col_widths = {
                        "OT Pendiente": 14, "Sucursal": 16, "Patente": 12,
                        "OT Origen": 14, "Vía": 10, "Producto": 20,
                        "Descripción Producto": 36, "Bodega": 18,
                        "Stock": 10, "Cantidad": 10, "Costo": 14,
                        "Estado": 24, "Fecha en bodega": 16, "Origen": 14,
                    }
                    for _ci, _cn in enumerate(_det_cols, 1):
                        _ws2.column_dimensions[_gcl(_ci)].width = _col_widths.get(_cn, 14)
                        _header(_ws2, 1, _ci, _cn, fill=_hdr_fill, font=_hdr_font, align=_ctr)
                    _ws2.row_dimensions[1].height = 20
                    _ws2.freeze_panes = "A2"

                    _num_cols = {"Stock", "Cantidad", "Costo"}
                    for _ri, (_, _rw) in enumerate(_df_det[_det_cols].iterrows(), 2):
                        _fill_row = _PFill(patternType="solid", fgColor="F8FBFF") if _ri % 2 == 0 else None
                        for _ci, _cn in enumerate(_det_cols, 1):
                            _val = _rw[_cn]
                            if _cn in _num_cols:
                                try: _val = float(_val)
                                except (ValueError, TypeError): _val = None
                            _c = _ws2.cell(_ri, _ci, _val)
                            _c.border    = _brd
                            _c.alignment = _ctr if _cn in _num_cols else _lft
                            if _fill_row: _c.fill = _fill_row
                            if _cn == "Costo" and _val is not None:
                                _c.number_format = '$#,##0'
                            elif _cn in ("Stock", "Cantidad") and _val is not None:
                                _c.number_format = '#,##0'

                    # Auto-filtro en detalle
                    _ws2.auto_filter.ref = (
                        f"A1:{_gcl(len(_det_cols))}1"
                    )

                    _buf = _io.BytesIO()
                    _wb.save(_buf)
                    _buf.seek(0)
                    return _buf.getvalue()

                # Preparar datos raw para Excel (numéricos sin formato string)
                _by_suc_raw = (
                    _df_bod_all.groupby("Sucursal")
                    .agg(OTs=("OT Pendiente", "nunique"),
                         Lineas=("Producto", "count"),
                         Valor=("Costo", "sum"),
                         Stock=("Stock", "sum"))
                    .reset_index()
                    .sort_values("Valor", ascending=False)
                ) if not _df_bod_all.empty else pd.DataFrame(
                    columns=["Sucursal","OTs","Lineas","Valor","Stock"])

                _by_ori_raw = (
                    _df_bod_all.groupby("Origen")
                    .agg(OTs=("OT Pendiente", "nunique"),
                         Lineas=("Producto", "count"),
                         Valor=("Costo", "sum"),
                         Stock=("Stock", "sum"))
                    .reset_index()
                    .sort_values("Valor", ascending=False)
                ) if not _df_bod_all.empty else pd.DataFrame(
                    columns=["Origen","OTs","Lineas","Valor","Stock"])

                _kpis_dict = {
                    "ot_bodega":  _ot_bodega,
                    "ot_espera":  _ot_espera,
                    "lineas_bod": len(_df_bod_all),
                    "stock":      f"{_stock_total:,.0f} un.",
                    "costo_bod":  f"${_costo_bodega:,.0f}",
                    "costo_esp":  f"${_costo_espera:,.0f}",
                }

                _xlsx_bytes = _generar_excel_repuestos(
                    _df_rp, _df_bod_all, _by_suc_raw, _by_ori_raw,
                    _kpis_dict, datetime.now(_TZ_CHILE).strftime("%d/%m/%Y")
                )
                st.download_button(
                    "📊 Descargar informe Excel",
                    _xlsx_bytes,
                    f"Informe_Repuestos_{datetime.now(_TZ_CHILE).strftime('%Y%m%d_%H%M')}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",                use_container_width=True,
                    type="primary",
                )


    # ---- TAB: BÚSQUEDA DE CLIENTE ----------------------------------------
    elif _tab == _TAB_CLIENTE:
        st.markdown('<p class="section-title">🔍 Búsqueda de Cliente</p>', unsafe_allow_html=True)
        st.caption(
            "Busca por nombre, RUT o patente. Verás todas las OTs pendientes agrupadas por cliente, "
            "el monto total pendiente de facturar y el saldo en Cuenta Ficha."
        )

        _q_cli = st.text_input(
            "Nombre, RUT o Patente",
            placeholder="Ej: JUAN PEREZ · 12345678-9 · ABCD12",
            key="cli_busq",
        )

        # ── Construir índice de clientes desde df_raw ────────────────────────
        # Clave primaria: primer RUT del cliente (o patente si no hay RUT)
        _clientes = {}   # { clave: {nombre, ruts, patentes, ots, neto_total, ant_total, tiene_saldo} }

        for _, _ot in df_raw.iterrows():
            _patente = str(_ot.get("PATENTE", "") or "").strip().upper()
            _rut_raw = str(_ot.get("rut_cliente", "") or "").strip()
            _ruts = [r.strip() for r in _rut_raw.split("/") if r.strip()] if _rut_raw else []

            # Anticipo / Cuenta Ficha
            _ant = _ot.get("anticipo", {}) or {}
            if isinstance(_ant, str):
                try:
                    import json as _json
                    _ant = _json.loads(_ant)
                except Exception:
                    _ant = {}
            _ant_total  = float(_ant.get("total", 0) or 0)
            _tiene_saldo = bool(_ant.get("tiene_saldo", False))
            _nombre_ant  = str(_ant.get("nombre", "") or "").strip()

            # NETO de la OT
            try:
                _neto_ot = float(str(_ot.get("NETO", 0) or "0").replace(".", "").replace(",", "."))
            except Exception:
                _neto_ot = 0.0

            # Clave del cliente
            _clave = _ruts[0] if _ruts else (_patente or "SIN_ID")
            _nombre_cli = _nombre_ant or (_ruts[0] if _ruts else _patente)

            if _clave not in _clientes:
                _clientes[_clave] = {
                    "nombre":      _nombre_cli,
                    "ruts":        set(_ruts),
                    "patentes":    set(),
                    "ots":         [],
                    "neto_total":  0.0,
                    "ant_total":   _ant_total,
                    "tiene_saldo": _tiene_saldo,
                    "movimientos": _ant.get("movimientos", []),
                }
            else:
                _clientes[_clave]["ruts"].update(_ruts)
                # Actualizar anticipo si tiene datos
                if _ant_total > 0:
                    _clientes[_clave]["ant_total"]   = _ant_total
                    _clientes[_clave]["tiene_saldo"]  = _tiene_saldo
                    _clientes[_clave]["movimientos"]  = _ant.get("movimientos", [])
                if _nombre_ant:
                    _clientes[_clave]["nombre"] = _nombre_ant

            if _patente:
                _clientes[_clave]["patentes"].add(_patente)

            _clientes[_clave]["neto_total"] += _neto_ot
            _clientes[_clave]["ots"].append({
                "folio":   str(_ot.get("FOLIO OT", "")),
                "sucursal": str(_ot.get("SUCURSAL", "")),
                "asesor":  str(_ot.get("ASESOR", "")),
                "rango":   str(_ot.get("RANGO", "")),
                "dias":    str(_ot.get("DIAS APERTURA", "")),
                "tipo":    str(_ot.get("TIPO VENTA", "")),
                "neto":    str(_ot.get("NETO", "")),
                "patente": _patente,
            })

        # ── Filtro de búsqueda ───────────────────────────────────────────────
        _q = (_q_cli or "").strip().upper()
        if not _q:
            st.divider()
            st.info("👆 Ingresa un nombre, RUT o patente para buscar.")
        else:
            _clientes_filtrados = {}
            for _clave, _cli in _clientes.items():
                _haystack = " ".join([
                    _cli["nombre"],
                    " ".join(_cli["ruts"]),
                    " ".join(_cli["patentes"]),
                ]).upper()
                if _q in _haystack:
                    _clientes_filtrados[_clave] = _cli

            # ── Métricas ─────────────────────────────────────────────────────
            _total_cli  = len(_clientes_filtrados)
            _total_ots  = sum(len(c["ots"]) for c in _clientes_filtrados.values())
            _total_neto = sum(c["neto_total"] for c in _clientes_filtrados.values())
            _con_saldo  = sum(1 for c in _clientes_filtrados.values() if c["tiene_saldo"])

            _km1, _km2, _km3, _km4 = st.columns(4)
            _km1.metric("Clientes encontrados", f"{_total_cli:,}")
            _km2.metric("OTs pendientes", f"{_total_ots:,}")
            _km3.metric("Total a facturar", f"${_total_neto:,.0f}")
            _km4.metric("Con Cuenta Ficha", f"{_con_saldo:,}")

            st.divider()

            if not _clientes_filtrados:
                st.warning(f"No se encontró ningún cliente con '{_q_cli}'.")
            else:
                # Ordenar por neto_total desc
                _items_sorted = sorted(_clientes_filtrados.items(),
                                       key=lambda x: x[1]["neto_total"], reverse=True)

                for _clave, _cli in _items_sorted:
                    _neto_cli = _cli["neto_total"]
                    _nombre_show = _cli["nombre"] or _clave
                    _patentes_str = " · ".join(sorted(_cli["patentes"])) if _cli["patentes"] else "—"
                    _ruts_str = " / ".join(sorted(_cli["ruts"])) if _cli["ruts"] else "—"

                    # Header del expander
                    _exp_title = (
                        f"**{_nombre_show}**  —  "
                        f"RUT: {_ruts_str}  —  "
                        f"Patente(s): {_patentes_str}  —  "
                        f"{len(_cli['ots'])} OT(s)"
                    )
                    with st.expander(_exp_title, expanded=True):
                        # Encabezados de columnas
                        _fcol1, _fcol2, _fcol3, _fcol4, _fcol5, _fcol6, _fcol7 = st.columns(
                            [2, 3, 3, 3, 3, 2, 1]
                        )
                        _fcol1.markdown("**Folio OT**")
                        _fcol2.markdown("**Sucursal**")
                        _fcol3.markdown("**Asesor**")
                        _fcol4.markdown("**Rango**")
                        _fcol5.markdown("**Tipo**")
                        _fcol6.markdown("**Neto OT**")
                        _fcol7.markdown("**Ver**")
                        st.markdown("---")

                        for _idx_ot, _ot_item in enumerate(_cli["ots"]):
                            _fcol1, _fcol2, _fcol3, _fcol4, _fcol5, _fcol6, _fcol7 = st.columns(
                                [2, 3, 3, 3, 3, 2, 1]
                            )
                            _fcol1.markdown(f"`{_ot_item['folio']}`")
                            _fcol2.markdown(_ot_item["sucursal"])
                            _fcol3.markdown(_ot_item["asesor"])
                            _r = _ot_item["rango"]
                            _r_color = {"0-30": "🟢", "31-60": "🟡", "61-90": "🟠", "91 o más": "🔴"}.get(_r, "⚪")
                            _fcol4.markdown(f"{_r_color} {_r} ({_ot_item['dias']}d)")
                            _fcol5.markdown(_ot_item["tipo"])
                            try:
                                _neto_fmt = f"${int(float(_ot_item['neto'])):,}" if _ot_item["neto"] else "—"
                            except Exception:
                                _neto_fmt = "—"
                            _fcol6.markdown(f"**{_neto_fmt}**")
                            if _fcol7.button("📄 Ver", key=f"cli_nav_{_clave}_{_idx_ot}",
                                             help=f"Ir a Documentos y Comentarios de OT {_ot_item['folio']}"):
                                st.session_state["_nav_folio"] = _ot_item["folio"]
                                st.rerun()

                        # Totalizador al pie de cada cliente
                        _cf_sufijo = (f"  ·  💰 Cuenta Ficha: ${_cli['ant_total']:,.0f}"
                                      if _cli["tiene_saldo"] else "")
                        st.markdown(
                            f"<div style='text-align:right;padding-top:6px;font-weight:600;color:#1a3a5c;'>"
                            f"Total pendiente de facturar: ${_neto_cli:,.0f}{_cf_sufijo}"
                            f"</div>",
                            unsafe_allow_html=True,
                        )

                        # Detalle Cuenta Ficha si tiene movimientos
                        if _cli["tiene_saldo"] and _cli.get("movimientos"):
                            with st.expander(f"💰 Movimientos Cuenta Ficha ({len(_cli['movimientos'])} mov.)",
                                             expanded=False):
                                _mov_rows = []
                                for _mv in _cli["movimientos"]:
                                    _mov_rows.append({
                                        "Documento": _mv.get("documento", ""),
                                        "N°":        _mv.get("nro", ""),
                                        "Saldo":     _mv.get("saldo", 0),
                                        "Fecha":     _mv.get("fecha", ""),
                                        "Local":     _mv.get("local", ""),
                                        "Glosa":     _mv.get("glosa", ""),
                                    })
                                _df_mov = pd.DataFrame(_mov_rows)
                                st.dataframe(
                                    _df_mov, hide_index=True, use_container_width=True,
                                    column_config={
                                        "Saldo": st.column_config.NumberColumn(format="$%d", width="small"),
                                    }
                                )
                                st.caption(f"Saldo total en cuenta: **${_cli['ant_total']:,.0f}**")


    # ---- TAB: FACTURAS X ----------------------------------------
    elif _tab == _TAB_FACT_X:
        st.markdown('<p class="section-title">🧾 Facturas X — Listado de facturas con prefijo X</p>',
                    unsafe_allow_html=True)
        st.caption("Se muestran todas las Facturas Cliente y Facturas Compañía cuyo número comienza con X, "
                   "cruzadas con el Costo Total del Vale de Consumo de cada OT. "
                   "Los filtros del sidebar aplican.")

        # Construir tabla: una fila por cada factura X encontrada
        _rows_fx = []
        for _, _ot in df.iterrows():
            # Costo total del Vale de Consumo (repuestos_actual)
            _reps = _ot.get("repuestos_actual", [])
            if not isinstance(_reps, list):
                _reps = []
            _costo_vale = sum(
                float(str(r.get("costo_total", 0) or 0).replace(",", ".")) for r in _reps
            ) if _reps else 0


            # Neto de la OT
            try:
                _neto_ot = float(str(_ot.get("NETO", 0) or "0").replace(".", "").replace(",", "."))
            except Exception:
                _neto_ot = 0.0

            # Anticipo / abono disponible
            _ant = _ot.get("anticipo", {}) or {}
            if isinstance(_ant, str):
                try:
                    _ant = json.loads(_ant)
                except Exception:
                    _ant = {}
            _rut_saldo  = str(_ant.get("rut_saldo", "") or "").strip()
            _monto_ant  = float(_ant.get("total", 0) or 0)
            _tiene_ant  = bool(_ant.get("tiene_saldo", False)) and _monto_ant > 0
            _abono_str  = f"💰 {_rut_saldo} · ${_monto_ant:,.0f}" if _tiene_ant else ""

            # Fecha Anticipo: última fecha de movimientos
            _movs_ant   = _ant.get("movimientos", [])
            if not isinstance(_movs_ant, list):
                _movs_ant = []
            _fechas_ant = [m.get("fecha", "") for m in _movs_ant if m.get("fecha")]
            _fecha_anticipo = max(_fechas_ant) if _fechas_ant else ""

            # Campos de la OT
            _folios_fc  = str(_ot.get("FOLIOS_FACT_CLIENTE",  "") or "").strip()
            _folios_fca = str(_ot.get("FOLIOS_FACT_COMPANIA", "") or "").strip()
            _fecha_fc   = str(_ot.get("FECHA_FACT_CLIENTE",   "") or "").strip()
            _fecha_fca  = str(_ot.get("FECHA_FACT_COMPANIA",  "") or "").strip()
            _rut_cli    = str(_ot.get("rut_cliente", "") or "").strip()
            _patente_fx = str(_ot.get("PATENTE",    "") or "").strip()
            _fol_ot     = str(_ot.get("FOLIO OT",   "") or "").strip()
            _suc_fx     = str(_ot.get("SUCURSAL",   "") or "").strip()
            _tv_fx      = str(_ot.get("TIPO VENTA", "") or "").strip()
            _marca_fx   = str(_ot.get("MARCA",      "") or "").strip()
            _asesor_fx  = str(_ot.get("ASESOR",     "") or "").strip()

            for _fnum in [f.strip() for f in _folios_fc.split(",") if f.strip().upper().startswith("X")]:
                _rows_fx.append({
                    "N° Factura X":       _fnum,
                    "Tipo":               "Factura Cliente",
                    "Fecha Factura":      _fecha_fc,
                    "Folio OT":           _fol_ot,
                    "Sucursal":           _suc_fx,
                    "Tipo Venta":         _tv_fx,
                    "Patente":            _patente_fx,
                    "RUT Cliente":        _rut_cli,
                    "💰 Abono Cliente":   _abono_str,
                    "Fecha Anticipo":     _fecha_anticipo,
                    "Marca":              _marca_fx,
                    "Asesor":             _asesor_fx,
                    "Total OT $":         _neto_ot,
                    "Costo Vale Consumo": _costo_vale,
                    "_sel_":              False,
                })
            for _fnum in [f.strip() for f in _folios_fca.split(",") if f.strip().upper().startswith("X")]:
                _rows_fx.append({
                    "N° Factura X":       _fnum,
                    "Tipo":               "Factura Compañía",
                    "Fecha Factura":      _fecha_fca,
                    "Folio OT":           _fol_ot,
                    "Sucursal":           _suc_fx,
                    "Tipo Venta":         _tv_fx,
                    "Patente":            _patente_fx,
                    "RUT Cliente":        _rut_cli,
                    "💰 Abono Cliente":   _abono_str,
                    "Fecha Anticipo":     _fecha_anticipo,
                    "Marca":              _marca_fx,
                    "Asesor":             _asesor_fx,
                    "Total OT $":         _neto_ot,
                    "Costo Vale Consumo": _costo_vale,
                    "_sel_":              False,
                })

        if not _rows_fx:
            st.info("No hay facturas X registradas con los filtros actuales.")
        else:
            _df_fx = pd.DataFrame(_rows_fx)

            # KPIs
            _kfx1, _kfx2, _kfx3, _kfx4, _kfx5 = st.columns(5)
            _kfx1.metric("Total Facturas X",       f"{len(_df_fx):,}")
            _kfx2.metric("OTs únicas",             f"{_df_fx['Folio OT'].nunique():,}")
            _kfx3.metric("Total OT $ (neto)",      f"${_df_fx['Total OT $'].sum():,.0f}")
            _kfx4.metric("Costo Vale Consumo",     f"${_df_fx['Costo Vale Consumo'].sum():,.0f}")
            _kfx5.metric("💰 Con abono disponible",
                         int((_df_fx["💰 Abono Cliente"] != "").sum()))

            st.divider()

            _col_cfg_fx = {
                "_sel_":              st.column_config.CheckboxColumn("📌", width="small"),
                "N° Factura X":       st.column_config.TextColumn("N° Factura X",    width="medium"),
                "Tipo":               st.column_config.TextColumn("Tipo",            width="medium"),
                "Fecha Factura":      st.column_config.TextColumn("Fecha Factura",   width="small"),
                "Folio OT":           st.column_config.TextColumn("Folio OT",        width="small"),
                "Sucursal":           st.column_config.TextColumn("Sucursal",        width="medium"),
                "Tipo Venta":         st.column_config.TextColumn("Tipo Venta",      width="medium"),
                "Patente":            st.column_config.TextColumn("Patente",         width="small"),
                "RUT Cliente":        st.column_config.TextColumn("RUT Cliente",     width="medium"),
                "💰 Abono Cliente":   st.column_config.TextColumn("💰 Abono Cliente", width="large"),
                "Fecha Anticipo":     st.column_config.TextColumn("Fecha Anticipo",  width="small"),
                "Marca":              st.column_config.TextColumn("Marca",           width="small"),
                "Asesor":             st.column_config.TextColumn("Asesor",          width="medium"),
                "Total OT $":         st.column_config.NumberColumn("Total OT $",          format="$%d"),
                "Costo Vale Consumo": st.column_config.NumberColumn("Costo Vale Consumo",  format="$%d"),
            }

            _df_fx_ed = st.data_editor(
                _df_fx, hide_index=True, use_container_width=True,
                height=560, column_config=_col_cfg_fx, key="fx_editor",
            )

            _sel_fx = _df_fx_ed[_df_fx_ed["_sel_"] == True]
            if not _sel_fx.empty:
                _folio_dest_fx = str(_sel_fx.iloc[0]["Folio OT"])
                _fxc1, _fxc2 = st.columns([3, 1])
                with _fxc1:
                    st.info(f"Factura X seleccionada — OT asociada: **{_folio_dest_fx}**")
                with _fxc2:
                    if st.button("📄 Ver Documentos y Comentarios", type="primary",
                                 use_container_width=True, key="btn_fx_nav"):
                        st.session_state["_nav_folio"] = _folio_dest_fx
                        st.rerun()

            # Descargas
            _fx_dl1, _fx_dl2 = st.columns(2)
            _csv_fx = _df_fx_ed.drop(columns=["_sel_"], errors="ignore").to_csv(
                index=False).encode("utf-8-sig")
            with _fx_dl1:
                st.download_button(
                    "⬇️ Descargar CSV", _csv_fx,
                    f"FacturasX_{ahora_chile().replace('/','_').replace(' ','_').replace(':','-')}.csv",
                    "text/csv",
                )
            with _fx_dl2:
                _fecha_str_fx  = ahora_chile().split()[0]
                _total_cost_fx = _df_fx["Costo Vale Consumo"].sum()
                _xlsx_fx = _generar_excel_fx(
                    _df_fx.drop(columns=["_sel_"], errors="ignore"),
                    len(_df_fx), _total_cost_fx, _fecha_str_fx,
                )
                st.download_button(
                    "⬇️ Descargar Excel", _xlsx_fx,
                    f"FacturasX_{_fecha_str_fx.replace('/','-')}.xlsx",
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )


    # ---- TAB: ADMIN ----------------------------------------------------------
    elif _tab == _TAB_ADMIN and usuario_activo == ADMIN_EMAIL:
        st.markdown('<p class="section-title">🛡️ Panel de Administración — Curifor S.A</p>',
                    unsafe_allow_html=True)

        _adm_tab = st.radio(
            "Sección admin",
            ["👥 Usuarios", "🔧 Técnicos", "📋 Auditoría", "🟢 En línea", "📊 Sistema"],
            horizontal=True, label_visibility="collapsed", key="_adm_subtab",
        )

        # ── 👥 Usuarios ───────────────────────────────────────────────────────
        if _adm_tab == "👥 Usuarios":
            st.markdown("#### Usuarios registrados")
            _usuarios = _leer_usuarios()
            if not _usuarios:
                st.info("No hay usuarios registrados aún.")
            else:
                _df_usr = pd.DataFrame([{
                    "Email":         u.get("email", ""),
                    "Nombre":        u.get("nombre", ""),
                    "Activo":        "✅" if u.get("activo", True) else "❌",
                    "Planificador":  "🔄 Sí" if (u.get("email","").lower()==ADMIN_EMAIL or u.get("puede_planificador", False)) else "🔒 No",
                    "Editar Planif.": "✏️ Sí" if (u.get("email","").lower()==ADMIN_EMAIL or u.get("puede_editar_planificador", False)) else "👁️ Solo ver",
                    "Pre-picking":   "🧩 Sí" if (u.get("email","").lower()==ADMIN_EMAIL or u.get("puede_prepicking", False)) else "🔒 No",
                    "Indicadores":   "📈 Sí" if (u.get("email","").lower()==ADMIN_EMAIL or u.get("puede_indicadores", False)) else "🔒 No",
                    "Asistente App": "🤖 Sí" if (u.get("email","").lower()==ADMIN_EMAIL or u.get("puede_asistente_app", False)) else "🔒 No",
                    "Cotizador":     "🧮 Sí" if (u.get("email","").lower()==ADMIN_EMAIL or u.get("puede_cotizador", False)) else "🔒 No",
                    "Agenda Taller": "🗓️ Sí" if (u.get("email","").lower()==ADMIN_EMAIL or u.get("puede_agenda_taller", False)) else "🔒 No",
                    "Recepción":     "📋 Sí" if (u.get("email","").lower()==ADMIN_EMAIL or u.get("puede_recepcion", False)) else "🔒 No",
                    "Cuenta Ficha":  "💳 Sí" if (u.get("email","").lower()==ADMIN_EMAIL or u.get("puede_cuenta_ficha", False)) else "🔒 No",
                    "Creado":        u.get("creado", ""),
                    "Último login":  u.get("ultimo_login", ""),
                } for u in _usuarios])
                st.dataframe(_df_usr, hide_index=True, use_container_width=True)

            st.divider()
            st.markdown("#### Activar / Desactivar usuario")
            _em_toggle = st.text_input("Email del usuario", key="adm_toggle_email",
                                       placeholder="correo@curifor.com")
            _ac1, _ac2 = st.columns(2)
            with _ac1:
                if st.button("✅ Activar", use_container_width=True, key="btn_activar"):
                    _us = _leer_usuarios()
                    _found = False
                    for _u in _us:
                        if _u.get("email", "").lower() == _em_toggle.strip().lower():
                            _u["activo"] = True
                            _found = True
                    if _found:
                        _guardar_usuarios(_us)
                        st.success(f"Usuario {_em_toggle} activado.")
                        _registrar_audit(usuario_activo, "activar_usuario", _em_toggle)
                    else:
                        st.error("Usuario no encontrado.")
            with _ac2:
                if st.button("❌ Desactivar", use_container_width=True, key="btn_desactivar"):
                    _us = _leer_usuarios()
                    _found = False
                    for _u in _us:
                        if _u.get("email", "").lower() == _em_toggle.strip().lower():
                            _u["activo"] = False
                            _found = True
                    if _found:
                        _guardar_usuarios(_us)
                        st.success(f"Usuario {_em_toggle} desactivado.")
                        _registrar_audit(usuario_activo, "desactivar_usuario", _em_toggle)
                    else:
                        st.error("Usuario no encontrado.")

            st.divider()
            st.markdown("#### 🔐 Permisos de módulos")
            st.caption(
                "Marca los módulos que cada usuario puede abrir desde la pantalla de "
                "bienvenida y presiona **Guardar permisos**. El administrador siempre "
                "tiene acceso a todo (no aparece en la tabla). Un usuario debe haber "
                "iniciado sesión al menos una vez para aparecer aquí."
            )
            _us_perm = _leer_usuarios()
            _rows_perm = [{
                "Email":  _u.get("email", ""),
                "Activo": "✅" if _u.get("activo", True) else "❌",
                "📋 Control y Gestión": bool(_u.get("puede_control", False)),
                "🔄 Planificador":      bool(_u.get("puede_planificador", False)),
                "✏️ Editar Planificador": bool(_u.get("puede_editar_planificador", False)),
                "📅 Confirmar Citas":   bool(_u.get("puede_confirmar_citas", False)),
                "🧩 Pre-picking":       bool(_u.get("puede_prepicking", False)),
                "📈 Indicadores":       bool(_u.get("puede_indicadores", False)),
                "🤖 Asistente App":     bool(_u.get("puede_asistente_app", False)),
                "🧮 Cotizador":         bool(_u.get("puede_cotizador", False)),
                "🚩 Campañas":          bool(_u.get("puede_campanas", False)),
                "🕐 Disponib. Técnicos": bool(_u.get("puede_disponibilidad_tecnicos", False)),
                "🗓️ Agenda Taller":      bool(_u.get("puede_agenda_taller", False)),
                "📋 Recepción":          bool(_u.get("puede_recepcion", False)),
                "💳 Cuenta Ficha":       bool(_u.get("puede_cuenta_ficha", False)),
                "📑 Informes Gestión":   bool(_u.get("puede_informes_gestion", False)),
                "🚗 Loaners":            bool(_u.get("puede_loaners", False)),
            } for _u in _us_perm if _u.get("email", "").lower() != ADMIN_EMAIL]

            if not _rows_perm:
                st.info("Aún no hay otros usuarios registrados.")
            else:
                _df_perm = pd.DataFrame(_rows_perm)
                _df_perm_ed = st.data_editor(
                    _df_perm, hide_index=True, use_container_width=True,
                    key="perm_editor",
                    column_config={
                        "Email":  st.column_config.TextColumn("Email", disabled=True),
                        "Activo": st.column_config.TextColumn("Activo", disabled=True, width="small"),
                        "📋 Control y Gestión": st.column_config.CheckboxColumn(
                            "📋 Control y Gestión",
                            help="Dashboard principal de OTs pendientes"),
                        "🔄 Planificador": st.column_config.CheckboxColumn(
                            "🔄 Planificador",
                            help="Puede ABRIR el Planificador de Taller (JPCB, Agenda, Control de Taller) — solo lectura si no tiene además 'Editar Planificador'"),
                        "✏️ Editar Planificador": st.column_config.CheckboxColumn(
                            "✏️ Editar Planificador",
                            help="Puede mover tarjetas del JPCB, asignar técnicos/horarios, editar Control de Taller y cerrar citas. Requiere también tener marcado '🔄 Planificador' para poder entrar al módulo."),
                        "📅 Confirmar Citas": st.column_config.CheckboxColumn(
                            "📅 Confirmar Citas",
                            help="Permiso LIMITADO: solo puede usar los botones Asiste/No Asiste/Reagenda de la columna 'Citas' del JPCB (confirmar si el vehículo agendado llegó o no). No puede arrastrar tarjetas, asignar técnicos ni editar Control de Taller. Por sí solo ya permite entrar al módulo (no requiere '🔄 Planificador')."),
                        "🧩 Pre-picking": st.column_config.CheckboxColumn(
                            "🧩 Pre-picking",
                            help="Puede ver y usar la pestaña Pre-picking dentro del Planificador de Taller (detalle de mantenciones, repuestos sugeridos, exportar presupuesto y marcar Realizado/Pendiente). Requiere también tener marcado '🔄 Planificador' para poder entrar al módulo."),
                        "📈 Indicadores": st.column_config.CheckboxColumn(
                            "📈 Indicadores",
                            help="Indicadores Post Venta (informe Power BI)"),
                        "🤖 Asistente App": st.column_config.CheckboxColumn(
                            "🤖 Asistente App",
                            help="Consulta rápida por patente/folio (sin IA)"),
                        "🧮 Cotizador": st.column_config.CheckboxColumn(
                            "🧮 Cotizador",
                            help="Cotizador de Mantenciones (marca/modelo/versión → operaciones, repuestos con stock, adicionales, packs y Excel)"),
                        "🚩 Campañas": st.column_config.CheckboxColumn(
                            "🚩 Campañas",
                            help="Revisión de Campañas Ford: casos con campaña/boletín vencidos, por revisar y ya revisados"),
                        "🕐 Disponib. Técnicos": st.column_config.CheckboxColumn(
                            "🕐 Disponib. Técnicos",
                            help="Permiso LIMITADO (pensado para Torre de Control): en Producción Técnicos puede marcar a un técnico como NO DISPONIBLE en un rango de fechas (vacaciones, licencia, permiso, capacitación) para que esos días no cuenten como horas disponibles en su % de productividad. No habilita nada más del Planificador. Requiere también '🔄 Planificador' para poder entrar al módulo."),
                        "🗓️ Agenda Taller": st.column_config.CheckboxColumn(
                            "🗓️ Agenda Taller",
                            help="Agenda de Taller (plataforma nueva): agendar citas de mantención por sucursal, día y hora. Reemplaza a la agenda web anterior."),
                        "📋 Recepción": st.column_config.CheckboxColumn(
                            "📋 Recepción",
                            help="Recepción de Vehículos (plataforma nueva): checklist de accesorios, inspección, firma del cliente e ingreso a taller. Reemplaza a la recepción anterior."),
                        "💳 Cuenta Ficha": st.column_config.CheckboxColumn(
                            "💳 Cuenta Ficha",
                            help="Cuenta Ficha: saldo disponible del cliente en su cuenta corriente (por sucursal y movimiento) e historial completo de sus OT con todos los documentos posteriores. Incluye el botón 'Revisado' compartido por todo el equipo."),
                        "📑 Informes Gestión": st.column_config.CheckboxColumn(
                            "📑 Informes Gestión",
                            help="Informes de Gestión: reportes por marca que se envían a las automotoras (AG con sus hojas Hyundai/Fortaleza/D&P por sucursal, e IMOP de Ford mes a mes), con el histórico y el avance del mes en curso calculado desde la carpeta Alimentación."),
                        "🚗 Loaners": st.column_config.CheckboxColumn(
                            "🚗 Loaners",
                            help="Loaners: flota de vehículos de cortesía que se prestan al cliente mientras su unidad está en taller (disponibilidad, cliente asignado, KMS de salida, fecha de OT y N° de caso Salesforce). Quien tenga el permiso puede ver Y editar el listado."),
                    },
                )
                if st.button("💾 Guardar permisos", type="primary", key="btn_save_perms"):
                    _us_fresh   = _leer_usuarios()
                    _map_flags  = {"📋 Control y Gestión": "puede_control",
                                   "🔄 Planificador":      "puede_planificador",
                                   "✏️ Editar Planificador": "puede_editar_planificador",
                                   "📅 Confirmar Citas":   "puede_confirmar_citas",
                                   "🧩 Pre-picking":       "puede_prepicking",
                                   "📈 Indicadores":       "puede_indicadores",
                                   "🤖 Asistente App":     "puede_asistente_app",
                                   "🧮 Cotizador":         "puede_cotizador",
                                   "🚩 Campañas":          "puede_campanas",
                                   "🕐 Disponib. Técnicos": "puede_disponibilidad_tecnicos",
                                   "🗓️ Agenda Taller":      "puede_agenda_taller",
                                   "📋 Recepción":          "puede_recepcion",
                                   "💳 Cuenta Ficha":       "puede_cuenta_ficha",
                                   "📑 Informes Gestión":   "puede_informes_gestion",
                                   "🚗 Loaners":            "puede_loaners"}
                    _cambios_perm = []
                    for _, _rp in _df_perm_ed.iterrows():
                        _u = _buscar_usuario(str(_rp["Email"]).strip().lower(), _us_fresh)
                        if not _u:
                            continue
                        for _col, _flag in _map_flags.items():
                            _nuevo = bool(_rp[_col])
                            if bool(_u.get(_flag, False)) != _nuevo:
                                _u[_flag] = _nuevo
                                _accion = ("autorizar_" if _nuevo else "revocar_") + _flag.replace("puede_", "")
                                _cambios_perm.append((_accion, _u.get("email", "")))
                    if _cambios_perm:
                        _guardar_usuarios(_us_fresh)
                        for _accion, _em_perm in _cambios_perm:
                            _registrar_audit(usuario_activo, _accion, _em_perm)
                        st.success(f"✅ {len(_cambios_perm)} permiso(s) actualizado(s).")
                        st.rerun()
                    else:
                        st.info("No hay cambios de permisos que guardar.")

            st.divider()
            st.markdown("#### 📅 Permiso masivo — Confirmar Citas (listado de asesores)")
            st.caption(
                "Aplica el permiso limitado **'📅 Confirmar Citas'** (solo botones Asiste/No "
                "Asiste/Reagenda del JPCB) a los 29 correos del listado de asesores subido el "
                "23/07/2026 (`Asesores.xlsx`) — de una sola vez, sin que cada uno tenga que "
                "haber iniciado sesión antes. Si el correo ya existe en el sistema, solo se le "
                "agrega el flag (nada más se toca); si todavía no existe, se pre-registra sin "
                "contraseña — la primera vez que esa persona entre con su correo, la app le va "
                "a pedir crear su contraseña normalmente, igual que cualquier usuario nuevo."
            )
            with st.expander(f"Ver los {len(ASESORES_CONFIRMAR_CITAS)} correos del listado"):
                st.dataframe(
                    pd.DataFrame([{"Email": em, "Nombre": nom} for em, nom in ASESORES_CONFIRMAR_CITAS.items()]),
                    hide_index=True, use_container_width=True,
                )
            if st.button("🚀 Aplicar 'Confirmar Citas' a la lista de asesores", key="btn_bulk_confirmar_citas"):
                _us_bulk = _leer_usuarios()
                _nuevos_bulk = 0
                _actualizados_bulk = 0
                for _em_bulk, _nom_bulk in ASESORES_CONFIRMAR_CITAS.items():
                    _em_bulk = _em_bulk.strip().lower()
                    _u_bulk = _buscar_usuario(_em_bulk, _us_bulk)
                    if _u_bulk:
                        if not _u_bulk.get("puede_confirmar_citas", False):
                            _u_bulk["puede_confirmar_citas"] = True
                            _actualizados_bulk += 1
                    else:
                        _nuevo_u_bulk = {
                            "email":                 _em_bulk,
                            "nombre":                _nom_bulk,
                            "activo":                True,
                            "creado":                ahora_chile(),
                            "puede_confirmar_citas": True,
                        }
                        _aplicar_restriccion_nomina(_nuevo_u_bulk)
                        _us_bulk.append(_nuevo_u_bulk)
                        _nuevos_bulk += 1
                if _nuevos_bulk or _actualizados_bulk:
                    _guardar_usuarios(_us_bulk)
                    _registrar_audit(
                        usuario_activo, "bulk_confirmar_citas",
                        f"{_nuevos_bulk} nuevo(s) + {_actualizados_bulk} actualizado(s) de {len(ASESORES_CONFIRMAR_CITAS)}",
                    )
                    st.success(
                        f"✅ Listo — {_nuevos_bulk} usuario(s) pre-registrado(s) y "
                        f"{_actualizados_bulk} ya existente(s) actualizado(s) con el permiso "
                        f"'Confirmar Citas'."
                    )
                    st.rerun()
                else:
                    st.info("Todos los correos del listado ya tenían el permiso 'Confirmar Citas'.")

            st.divider()
            st.markdown("#### 🏢 Acceso por Sucursal")
            st.caption(
                "Cada usuario detectado en la nómina (Nomina Area PV) queda limitado a ver "
                "solo su sucursal — en Control y Gestión Post Venta, en cualquiera de las "
                "áreas del Planificador de Taller y en el Asistente App. Si 'Sucursales "
                "permitidas' queda vacío, el usuario NO tiene restricción (ve todas). Usa "
                "el selector de abajo para ampliar o cambiar el acceso de alguien puntual."
            )
            _us_suc = _leer_usuarios()
            _rows_suc = [{
                "Email": _u.get("email", ""),
                "Sucursal detectada (nómina)": _u.get("sucursal_home", "") or "—",
                "Sucursales permitidas": ", ".join(_u.get("sucursales_permitidas") or []) or "Todas (sin restricción)",
            } for _u in _us_suc if _u.get("email", "").lower() != ADMIN_EMAIL]

            if not _rows_suc:
                st.info("Aún no hay otros usuarios registrados.")
            else:
                st.dataframe(pd.DataFrame(_rows_suc), hide_index=True, use_container_width=True)

                st.markdown("##### Editar sucursales permitidas de un usuario")
                _emails_suc = sorted([r["Email"] for r in _rows_suc])
                _em_suc_sel = st.selectbox("Usuario", _emails_suc, key="adm_suc_email_sel")
                _u_suc_sel  = _buscar_usuario(_em_suc_sel, _us_suc) or {}
                _sucs_todas = sorted(df_raw_full["SUCURSAL"].unique().tolist())
                _actuales_suc = _u_suc_sel.get("sucursales_permitidas") or []
                _nuevas_suc = st.multiselect(
                    "Sucursales permitidas (vacío = sin restricción, ve todas)",
                    _sucs_todas,
                    default=[s for s in _actuales_suc if s in _sucs_todas],
                    key="adm_suc_multiselect",
                )
                if st.button("💾 Guardar sucursales permitidas", key="btn_save_sucursales"):
                    _us_fresh_suc = _leer_usuarios()
                    _u_fresh_suc  = _buscar_usuario(_em_suc_sel, _us_fresh_suc)
                    if _u_fresh_suc:
                        _u_fresh_suc["sucursales_permitidas"] = _nuevas_suc
                        _guardar_usuarios(_us_fresh_suc)
                        _registrar_audit(
                            usuario_activo, "editar_sucursales_permitidas",
                            f"{_em_suc_sel} -> {', '.join(_nuevas_suc) if _nuevas_suc else 'Todas (sin restricción)'}",
                        )
                        st.success(f"✅ Sucursales permitidas de {_em_suc_sel} actualizadas.")
                        st.rerun()
                    else:
                        st.error("Usuario no encontrado.")

            st.divider()
            st.markdown("#### 🔑 Restablecer contraseña (usuario olvidó la suya)")
            st.caption(
                "Genera una contraseña provisoria para el usuario y compártesela por un "
                "medio seguro (WhatsApp, en persona, etc. — no queda guardada en ninguna "
                "parte, solo se muestra una vez en esta pantalla). Al iniciar sesión con "
                "ella, la app le va a exigir crear su propia contraseña nueva antes de "
                "dejarlo entrar."
            )
            _us_reset = _leer_usuarios()
            _emails_reset = sorted([u.get("email", "") for u in _us_reset
                                    if u.get("email", "").lower() != ADMIN_EMAIL])
            if not _emails_reset:
                st.info("Aún no hay otros usuarios registrados.")
            else:
                _em_reset_sel = st.selectbox("Usuario", _emails_reset, key="adm_reset_email_sel")
                if st.button("🔑 Generar contraseña provisoria", key="btn_reset_pwd"):
                    _ok_reset, _msg_reset, _pwd_temp = asignar_password_provisoria(_em_reset_sel, usuario_activo)
                    if _ok_reset:
                        st.success(_msg_reset)
                        st.code(_pwd_temp, language=None)
                        st.warning(
                            "⚠️ Copia esta contraseña ahora — no se va a volver a mostrar. "
                            "Entrégasela al usuario para que inicie sesión con su correo y "
                            "esta contraseña; la app le pedirá crear una nueva de inmediato."
                        )
                    else:
                        st.error(_msg_reset)

        # ── 🔧 Técnicos ───────────────────────────────────────────────────────
        elif _adm_tab == "🔧 Técnicos":
            st.markdown("#### Técnicos por sucursal")
            st.caption(
                "Define los técnicos de cada sucursal. "
                "El Planificador de Taller usa esta lista para asignar OTs."
            )

            _sha_ctrl, _ctrl_dat = _leer_json_github_raw(GITHUB_CTRL_TALLER)
            if not _ctrl_dat:
                _ctrl_dat = {}

            _sucs_tec    = sorted(df_raw["SUCURSAL"].unique().tolist())
            _suc_sel_tec = st.selectbox("Sucursal", _sucs_tec, key="adm_tec_suc")

            _suc_dat       = _ctrl_dat.get(_suc_sel_tec, {})
            _tecs_actuales = list(_suc_dat.get("tecnicos", []))

            st.markdown(f"**Técnicos en {_suc_sel_tec}:**")
            if _tecs_actuales:
                for _ti, _tn in enumerate(_tecs_actuales):
                    _tc1, _tc2 = st.columns([5, 1])
                    _tc1.markdown(f"• {_tn}")
                    if _tc2.button("✕", key=f"rm_tec_{_ti}", help=f"Eliminar {_tn}"):
                        _tecs_actuales.pop(_ti)
                        if _suc_sel_tec not in _ctrl_dat:
                            _ctrl_dat[_suc_sel_tec] = {}
                        _ctrl_dat[_suc_sel_tec]["tecnicos"] = _tecs_actuales
                        _guardar_json_github_raw(
                            GITHUB_CTRL_TALLER, _ctrl_dat,
                            f"Técnicos {_suc_sel_tec} — {ahora_chile()}"
                        )
                        _registrar_audit(usuario_activo, "eliminar_tecnico",
                                         f"{_tn} de {_suc_sel_tec}")
                        st.success(f"Técnico {_tn} eliminado.")
                        st.rerun()
            else:
                st.info(f"No hay técnicos configurados para {_suc_sel_tec}.")

            st.divider()
            _nuevo_tec = st.text_input("Nombre del técnico a agregar", key="adm_nuevo_tec",
                                       placeholder="Ej: JUAN PÉREZ")
            if st.button("➕ Agregar técnico", key="btn_add_tec", type="primary"):
                _nuevo_tec = (_nuevo_tec or "").strip().upper()
                if not _nuevo_tec:
                    st.error("Ingresa el nombre del técnico.")
                elif _nuevo_tec in _tecs_actuales:
                    st.warning("Ese técnico ya existe en esta sucursal.")
                else:
                    _tecs_actuales.append(_nuevo_tec)
                    if _suc_sel_tec not in _ctrl_dat:
                        _ctrl_dat[_suc_sel_tec] = {}
                    _ctrl_dat[_suc_sel_tec]["tecnicos"] = _tecs_actuales
                    _guardar_json_github_raw(
                        GITHUB_CTRL_TALLER, _ctrl_dat,
                        f"Técnicos {_suc_sel_tec} — {ahora_chile()}"
                    )
                    _registrar_audit(usuario_activo, "agregar_tecnico",
                                     f"{_nuevo_tec} en {_suc_sel_tec}")
                    st.success(f"Técnico {_nuevo_tec} agregado a {_suc_sel_tec}.")
                    st.rerun()

            # ── Mapeo manual tecnico -> sucursal (fallback de Produccion Tecnicos) ──
            # 23/07/2026, a pedido de Cristobal: investigando por que "Produccion
            # Tecnicos" mostraba solo 4 de 10 tecnicos configurados en una sucursal,
            # se confirmo con datos reales que la consulta SQL de BDFlexline trae
            # una porcion real de tecnicos (~1 de cada 3, en la muestra revisada) sin
            # sucursal identificable — vienen SOLO de Tmp_HorasPorTecnico, tabla que
            # no registra sucursal ni fecha, y nunca aparecen en la otra tabla
            # (Tmp_ProduccionMensualMecanicos) para poder cruzarlos por nombre. Esto
            # es una limitacion real del origen de datos, no un bug de la app. Como
            # Cristobal SI sabe en que sucursal trabaja cada tecnico, se agrega un
            # mapeo manual (persistido en tecnicos_sucursal_manual.json) que
            # consolidar_OTs.py usa como ULTIMO respaldo (despues del cruce por
            # nombre contra la otra tabla) antes de darse por vencido.
            st.divider()
            st.markdown("#### 🗺️ Sucursal manual para Producción Técnicos (respaldo)")
            st.caption(
                "Algunos técnicos llegan desde BDFlexline sin sucursal identificable "
                "(limitación real del origen de datos — no un error de la app). Acá "
                "puedes asignarles la sucursal a mano; se usa como último respaldo "
                "la próxima vez que corras `Ejecutar_Consolidacion.bat`."
            )
            _sha_map_tec, _map_tec = _leer_json_github_raw(GITHUB_TECNICOS_SUCURSAL_MANUAL)
            if not isinstance(_map_tec, dict):
                _map_tec = {}

            _prod_resumen_adm, _, _, _ = _cargar_produccion_tecnicos()
            _tecs_sin_suc = sorted({
                str(r.get("mecanico", "")).strip()
                for r in (_prod_resumen_adm or [])
                if str(r.get("mecanico", "")).strip()
                and not str(r.get("sucursal_mecanico", "")).strip()
                and str(r.get("mecanico", "")).strip().upper() not in _map_tec
            })

            if _tecs_sin_suc:
                st.warning(
                    f"⚠️ {len(_tecs_sin_suc)} técnico(s) llegan sin sucursal identificable "
                    f"y aún no tienen mapeo manual: {', '.join(_tecs_sin_suc[:20])}"
                    + (f" (+{len(_tecs_sin_suc)-20} más)" if len(_tecs_sin_suc) > 20 else "")
                )
                _tc_map1, _tc_map2, _tc_map3 = st.columns([3, 3, 2])
                _tec_elegido = _tc_map1.selectbox("Técnico sin sucursal", _tecs_sin_suc, key="adm_map_tec_sel")
                _suc_elegida = _tc_map2.selectbox("Asignar a sucursal", _sucs_tec, key="adm_map_tec_suc")
                if _tc_map3.button("💾 Guardar mapeo", key="btn_map_tec", type="primary", use_container_width=True):
                    _map_tec[_tec_elegido.strip().upper()] = _suc_elegida
                    _guardar_json_github_raw(
                        GITHUB_TECNICOS_SUCURSAL_MANUAL, _map_tec,
                        f"Mapeo manual tecnico->sucursal: {_tec_elegido} -> {_suc_elegida} — {ahora_chile()}"
                    )
                    _registrar_audit(usuario_activo, "mapear_tecnico_sucursal",
                                     f"{_tec_elegido} -> {_suc_elegida}")
                    st.success(f"'{_tec_elegido}' quedó asignado a {_suc_elegida}. Se aplicará en la próxima consolidación.")
                    st.rerun()
            else:
                st.success("✅ No hay técnicos pendientes de mapeo manual (o ya están todos asignados).")

            if _map_tec:
                st.markdown("**Mapeos manuales ya guardados:**")
                for _mt_nombre in sorted(_map_tec.keys()):
                    _mtc1, _mtc2 = st.columns([5, 1])
                    _mtc1.markdown(f"• {_mt_nombre} → **{_map_tec[_mt_nombre]}**")
                    if _mtc2.button("✕", key=f"rm_map_tec_{_mt_nombre}", help=f"Quitar mapeo de {_mt_nombre}"):
                        del _map_tec[_mt_nombre]
                        _guardar_json_github_raw(
                            GITHUB_TECNICOS_SUCURSAL_MANUAL, _map_tec,
                            f"Quitar mapeo manual de {_mt_nombre} — {ahora_chile()}"
                        )
                        _registrar_audit(usuario_activo, "quitar_mapeo_tecnico_sucursal", _mt_nombre)
                        st.rerun()

        # ── 📋 Auditoría ──────────────────────────────────────────────────────
        elif _adm_tab == "📋 Auditoría":
            st.markdown("#### Log de auditoría")
            _audit_data = _leer_audit()
            if not _audit_data:
                st.info("No hay registros de auditoría.")
            else:
                _df_aud = pd.DataFrame(_audit_data[::-1])
                _aud_busq = st.text_input("Buscar en auditoría", "", key="adm_aud_busq")
                if _aud_busq:
                    _mask_a = pd.Series(False, index=_df_aud.index)
                    for _ca in ["usuario", "accion", "detalle", "folio_ot"]:
                        if _ca in _df_aud.columns:
                            _mask_a |= _df_aud[_ca].astype(str).str.contains(
                                _aud_busq, case=False, na=False)
                    _df_aud = _df_aud[_mask_a]
                st.caption(f"Mostrando {len(_df_aud):,} registros")
                st.dataframe(
                    _df_aud.rename(columns={
                        "fecha": "Fecha", "usuario": "Usuario",
                        "accion": "Acción", "detalle": "Detalle", "folio_ot": "Folio OT",
                    }),
                    hide_index=True, use_container_width=True, height=500,
                )
                _csv_aud = _df_aud.to_csv(index=False).encode("utf-8-sig")
                st.download_button("⬇️ Descargar auditoría CSV", _csv_aud,
                                   "audit_log.csv", "text/csv")

        # ── 🟢 En línea ───────────────────────────────────────────────────────
        elif _adm_tab == "🟢 En línea":
            st.markdown("#### Usuarios activos (últimos 5 minutos)")
            _online = _leer_online_users(300)
            if not _online:
                st.info("No hay usuarios activos en este momento.")
            else:
                for _ou in _online:
                    _hace_min = _ou["hace"] // 60
                    _hace_sec = _ou["hace"] % 60
                    st.markdown(
                        f"🟢 **{_ou['email']}** — visto hace "
                        f"{_hace_min}m {_hace_sec}s  ·  último: {_ou['last_seen']}"
                    )
            if st.button("🔄 Actualizar", key="adm_online_ref"):
                st.rerun()

        # ── 📊 Sistema ────────────────────────────────────────────────────────
        elif _adm_tab == "📊 Sistema":
            st.markdown("#### Estadísticas del sistema")
            _sa1, _sa2, _sa3, _sa4 = st.columns(4)
            _sa1.metric("Total OTs (raw)",       f"{len(df_raw):,}")
            _sa2.metric("Usuarios registrados",  len(_leer_usuarios()))
            _sa3.metric("Notificaciones no leídas",
                        len([n for n in _leer_notificaciones() if not n.get("leida")]))
            _sa4.metric("Comentarios totales",   len(cargar_comentarios()))
            st.divider()
            st.caption(
                f"App: curifor-ots.streamlit.app  ·  "
                f"Datos actualizados: {fecha_actualizacion}  ·  "
                f"OTs en sistema: {len(df_raw):,}"
            )



# ============================================================
#   VISTA DIVIDIDA (split screen) — dos módulos lado a lado
#   -------------------------------------------------------
#   Permite abrir simultáneamente 2 de los módulos "pane-aware"
#   (Control y Gestión, Planificador, Indicadores, Campañas Ford,
#   Cuenta Ficha) uno junto al otro, con un selector de proporción
#   (no drag-resize). Exclusión mutua: cada selector excluye lo que
#   el otro panel ya tiene elegido, así nunca se ejecuta el mismo
#   módulo 2 veces en la misma corrida del script (evita IDs de
#   widget duplicados).
# ============================================================
if st.session_state.get("app_mode") == "split":

    _SPLIT_MODULOS = [
        ("Control y Gestión Post Venta", "ots", _render_control),
        ("Planificador de Taller", "planificador", _render_planificador),
        ("Indicadores Post Venta", "indicadores", _render_indicadores),
        ("Revisión de Campañas Ford", "campanas", _render_campanas),
        ("Cuenta Ficha", "cuenta_ficha", _render_cuenta_ficha),
        ("Informes de Gestión", "informes_gestion", _render_informes_gestion),
        ("Loaners", "loaners", _render_loaners),
    ]
    _SPLIT_FN_POR_MODO = {m: fn for _, m, fn in _SPLIT_MODULOS}
    _SPLIT_LABEL_POR_MODO = {m: lbl for lbl, m, _ in _SPLIT_MODULOS}

    _SPLIT_RATIOS = {
        "50 / 50": (1, 1), "60 / 40": (6, 4), "70 / 30": (7, 3),
        "40 / 60": (4, 6), "30 / 70": (3, 7),
    }

    with st.sidebar:
        st.markdown(
            f'<img src="{LOGO_DATA_URI}" style="max-width:180px;margin-bottom:0.4rem;"/>',
            unsafe_allow_html=True,
        )
        if st.button("← Volver al inicio", use_container_width=True, key="split_volver"):
            st.session_state.pop("app_mode", None)
            st.rerun()

        st.divider()
        st.markdown("### 🪟 Vista Dividida")
        _split_ratio_lbl = st.selectbox(
            "Proporción", list(_SPLIT_RATIOS.keys()), index=0, key="split_ratio",
        )
        st.caption(
            "Elige un módulo para cada panel. Un módulo ya abierto en un panel "
            "no se puede repetir en el otro."
        )

        st.divider()
        st.markdown(f"**Usuario:** `{usuario_activo}`")
        if st.button("🚪 Cerrar sesión", use_container_width=True, key="split_logout"):
            for _k in ["authenticated", "user_email", "app_mode"]:
                st.session_state.pop(_k, None)
            st.rerun()

    st.markdown(
        f'''<div class="curifor-header">
            <div class="logo-pill"><img src="{LOGO_DATA_URI}" /></div>
            <div class="curifor-header-text">
                <h2>Vista Dividida</h2>
                <p>Dos módulos a la vez, lado a lado.</p>
                <span class="dev-credit">Curifor S.A</span>
            </div>
        </div>''',
        unsafe_allow_html=True,
    )

    _rat_l, _rat_r = _SPLIT_RATIOS[_split_ratio_lbl]
    _col_l, _col_r = st.columns([_rat_l, _rat_r], gap="medium")

    _SIN_MODULO = "— Selecciona un módulo —"

    # Valores elegidos en la corrida anterior (para calcular exclusion mutua
    # y la posicion por defecto de cada selectbox).
    _modo_prev_l = st.session_state.get("split_sel_L", "")
    _modo_prev_r = st.session_state.get("split_sel_R", "")

    with _col_l:
        _opts_l = [_SIN_MODULO] + [lbl for lbl, m, _ in _SPLIT_MODULOS if m != _modo_prev_r]
        _lbl_prev_l = _SPLIT_LABEL_POR_MODO.get(_modo_prev_l, _SIN_MODULO)
        _idx_l = _opts_l.index(_lbl_prev_l) if _lbl_prev_l in _opts_l else 0
        _sel_lbl_l = st.selectbox("Panel izquierdo", _opts_l, index=_idx_l, key="split_pick_L")
        _sel_modo_l = next((m for lbl, m, _ in _SPLIT_MODULOS if lbl == _sel_lbl_l), "")
        st.session_state["split_sel_L"] = _sel_modo_l
        st.divider()
        if _sel_modo_l:
            _SPLIT_FN_POR_MODO[_sel_modo_l](pane="L")
        else:
            st.info("👈 Elige un módulo para este panel.")

    with _col_r:
        # Usa _sel_modo_l (ya actualizado arriba) para excluir en el acto lo
        # que el panel izquierdo acaba de elegir en esta misma corrida.
        _opts_r = [_SIN_MODULO] + [lbl for lbl, m, _ in _SPLIT_MODULOS if m != _sel_modo_l]
        _lbl_prev_r = _SPLIT_LABEL_POR_MODO.get(_modo_prev_r, _SIN_MODULO)
        _idx_r = _opts_r.index(_lbl_prev_r) if _lbl_prev_r in _opts_r else 0
        _sel_lbl_r = st.selectbox("Panel derecho", _opts_r, index=_idx_r, key="split_pick_R")
        _sel_modo_r = next((m for lbl, m, _ in _SPLIT_MODULOS if lbl == _sel_lbl_r), "")
        st.session_state["split_sel_R"] = _sel_modo_r
        st.divider()
        if _sel_modo_r:
            _SPLIT_FN_POR_MODO[_sel_modo_r](pane="R")
        else:
            st.info("👉 Elige un módulo para este panel.")

    st.stop()


# ============================================================
#   ASISTENTE APP  — modo independiente
#   Busqueda rapida por patente y/o folio OT, sin IA — puro cruce de
#   datos contra df_raw (mismo mecanismo que el consolidador usa para
#   costos de Vale de Consumo). No genera archivos, no edita nada, no
#   sale ningun dato fuera de GitHub/Streamlit. Acceso restringido
#   (puede_asistente_app) — por defecto solo el admin lo ve. 08/07/2026.
# ============================================================
if st.session_state.get("app_mode") == "asistente":

    if not _puede_usar_asistente_app(usuario_activo):
        st.session_state.pop("app_mode", None)
        st.error("🔒 No tienes autorización para usar el Asistente App. "
                 "Solicítala al administrador (cjerez@curifor.com).")
        if st.button("← Volver al inicio", key="asis_sin_acceso_volver"):
            st.rerun()
        st.stop()

    import re as _re_asis
    import unicodedata as _ud_asis
    import difflib as _difflib_asis

    def _asis_norm_txt(s):
        # Minusculas + sin tildes/diacriticos, para que "Muñoz"/"MUÑOZ"/"munoz"
        # y variantes de acentos siempre calcen entre si al comparar texto libre
        # contra los valores reales de SUCURSAL/MARCA/ASESOR/etc. 08/07/2026.
        s = str(s or "").lower().strip()
        s = _ud_asis.normalize("NFKD", s)
        return "".join(c for c in s if not _ud_asis.combining(c))

    def _asis_costo_vale(reps):
        if not isinstance(reps, list):
            return 0.0
        total = 0.0
        for _r in reps:
            try:
                total += float(str(_r.get("costo_total", 0) or 0).replace(",", "."))
            except Exception:
                pass
        return total

    def _asis_conteo_repuestos(reps_compras):
        if not isinstance(reps_compras, list):
            return 0, 0
        _bod = sum(1 for _r in reps_compras if _r.get("en_bodega"))
        _esp = sum(1 for _r in reps_compras if not _r.get("en_bodega"))
        return _bod, _esp

    with st.sidebar:
        st.markdown(
            f'<img src="{LOGO_DATA_URI}" style="max-width:180px;margin-bottom:0.4rem;"/>',
            unsafe_allow_html=True,
        )
        st.markdown(f"**Act:** {fecha_actualizacion}")
        st.markdown("")
        if st.button("← Volver al inicio", use_container_width=True, key="asis_volver"):
            st.session_state.pop("app_mode", None)
            st.rerun()
        st.divider()
        st.markdown(f"**Usuario:** `{usuario_activo}`")
        if st.button("🔄 Actualizar datos", use_container_width=True, key="asis_refresh"):
            st.cache_data.clear()
            st.rerun()
        if st.button("🚪 Cerrar sesión", use_container_width=True, key="asis_logout"):
            for _k in ["authenticated", "user_email", "app_mode"]:
                st.session_state.pop(_k, None)
            st.rerun()

    st.markdown("""
    <style>
      /* Asistente App — identidad morada sobre los componentes nativos de chat */
      div[data-testid="stChatMessageAvatarAssistant"]{
          background:linear-gradient(135deg,#6b46c1,#8f6bd6)!important;
          box-shadow:0 2px 6px rgba(107,70,193,.35);
      }
      div[data-testid="stChatMessageAvatarUser"]{
          background:linear-gradient(135deg,#4a5568,#718096)!important;
      }
      div[data-testid="stChatInput"] textarea{ border-radius:10px!important; }
      div[data-testid="stChatInput"]:focus-within{
          box-shadow:0 0 0 2px rgba(107,70,193,.22);
          border-radius:10px;
      }
      div[data-testid="stVerticalBlockBorderWrapper"]{
          border-radius:14px!important;
          border-color:rgba(107,70,193,.28)!important;
          box-shadow:0 3px 14px rgba(107,70,193,.08);
      }
    </style>
    """, unsafe_allow_html=True)

    st.markdown(
        f'''<div class="curifor-header">
            <div class="logo-pill"><img src="{LOGO_DATA_URI}" /></div>
            <div class="curifor-header-text">
                <h2>🤖 Asistente App</h2>
                <p>Consulta rápida por patente, folio y mucho más</p>
                <span class="dev-credit">Curifor S.A</span>
            </div>
            <span class="curifor-badge" style="background:rgba(107,70,193,.22);border-color:rgba(107,70,193,.4);">🤖 IA determinística</span>
        </div>''',
        unsafe_allow_html=True,
    )

    st.info(
        "Pega una o varias **patentes** y/o **folios OT** (una por línea, o separadas por "
        "coma/espacio) y te muestro si tienen una OT pendiente abierta, en qué sucursal, "
        "el costo del Vale de Consumo y el NETO de la OT."
    )

    _asis_txt = st.text_area(
        "Patentes o folios a consultar",
        placeholder="Ej:\nLYVG96\nSTSG41\n1174878",
        height=140, key="asis_input",
    )

    if st.button("🔎 Consultar", type="primary", key="asis_buscar") and _asis_txt.strip():
        _tokens = [t.strip().upper() for t in _re_asis.split(r"[,\s;]+", _asis_txt) if t.strip()]
        _tokens = list(dict.fromkeys(_tokens))  # dedup preservando orden

        _res_rows = []
        for _tok in _tokens:
            _es_folio = _tok.isdigit()
            if _es_folio:
                _matches = df_raw[df_raw["FOLIO OT"].astype(str).str.strip() == _tok]
            else:
                _matches = df_raw[df_raw["PATENTE"].astype(str).str.strip().str.upper()
                                   .str.replace(" ", "") == _tok.replace(" ", "")]

            if _matches.empty:
                _res_rows.append({
                    "Consulta": _tok, "¿OT Abierta?": "NO",
                    "Folio OT": "", "Sucursal": "", "Rango": "", "Días": "",
                    "Asesor": "", "Costo Vale Consumo": None, "NETO OT": None,
                    "Rep. en Bodega": None, "Rep. en Espera": None, "Nota": "",
                })
                continue

            for _, _m in _matches.iterrows():
                _neto = float(_m.get("NETO", 0) or 0)
                _vale = _asis_costo_vale(_m.get("repuestos_actual"))
                _n_bod, _n_esp = _asis_conteo_repuestos(_m.get("repuestos_compras"))
                _nota = ""
                if _neto == 0 and _vale == 0:
                    _nota = "OT sin datos cargados (sin Vale de Consumo ni NETO)"
                _res_rows.append({
                    "Consulta": _tok, "¿OT Abierta?": "SI",
                    "Folio OT": _m.get("FOLIO OT", ""),
                    "Sucursal": _m.get("SUCURSAL", ""),
                    "Rango": _m.get("RANGO", ""),
                    "Días": _m.get("DIAS APERTURA", ""),
                    "Asesor": _m.get("ASESOR", ""),
                    "Costo Vale Consumo": _vale,
                    "NETO OT": _neto,
                    "Rep. en Bodega": _n_bod,
                    "Rep. en Espera": _n_esp,
                    "Nota": _nota,
                })

        _df_res = pd.DataFrame(_res_rows)
        st.session_state["asis_resultado"] = _df_res

    if "asis_resultado" in st.session_state:
        _df_res = st.session_state["asis_resultado"]
        _n_si = int((_df_res["¿OT Abierta?"] == "SI").sum())
        _n_no = int((_df_res["¿OT Abierta?"] == "NO").sum())
        _c1, _c2 = st.columns(2)
        _c1.metric("Con OT abierta", _n_si)
        _c2.metric("Sin OT abierta", _n_no)

        def _asis_highlight(row):
            # Se fija tambien el color de texto (no solo el fondo) para que se
            # lea bien tanto en modo claro como en modo oscuro de Streamlit —
            # con solo background-color, el texto heredaba el color del tema
            # (blanco en modo oscuro) y quedaba casi invisible sobre fondo claro.
            if row["¿OT Abierta?"] == "SI":
                bg, fg = "#e6f4ea", "#14532d"
            else:
                bg, fg = "#fdecea", "#7a1a13"
            return [f"background-color: {bg}; color: {fg}"] * len(row)

        st.dataframe(
            _df_res.style.apply(_asis_highlight, axis=1),
            hide_index=True, use_container_width=True,
            column_config={
                "Costo Vale Consumo": st.column_config.NumberColumn(format="$ %d"),
                "NETO OT": st.column_config.NumberColumn(format="$ %d"),
            },
        )

        _csv_asis = _df_res.to_csv(index=False).encode("utf-8-sig")
        st.download_button(
            "⬇️ Descargar resultado como CSV", _csv_asis,
            f"Asistente_App_{datetime.now().strftime('%Y%m%d_%H%M')}.csv",
            mime="text/csv",
        )

    # ========================================================
    #   OTRAS CONSULTAS — interprete simple por palabras clave, con
    #   interfaz tipo chat (st.chat_message/st.chat_input). Detecta
    #   nombres de sucursal, asesor, marca, tipo de venta, rango,
    #   "mas de N dias", "top N", "repuestos/stock", "abono/cuenta
    #   ficha", y si no reconoce nada estructurado busca texto libre
    #   en Glosa/Observacion/Notas. Siempre muestra que interpreto,
    #   para que quede claro que es un cruce de datos y no una
    #   respuesta "inventada". Historico de OTs cerradas desde 2020
    #   queda fuera de alcance (solo hay datos de OTs pendientes en
    #   datos_dashboard.json). El stock de repuestos se responde
    #   contra el catalogo completo de stock_repuestos.json (PASO 10
    #   del consolidador), no contra repuestos_compras (que solo
    #   trae lo ligado a una OT). 08/07/2026.
    # ========================================================
    st.divider()
    st.markdown("### 💬 Otras consultas")
    st.caption(
        "Este es un asistente de consulta rápida sobre tus OTs pendientes: reconoce sucursal, "
        "asesor, marca, tipo de venta, categoría, año, mes, modelo, rango, días (\"más de/menos "
        "de/entre X e Y/sobre N días\"), exclusiones (\"excepto Ford\"), \"top N\" (también \"con "
        "menos\"), desgloses (\"por sucursal\"), comparaciones (\"Linderos vs Talca\"), estadísticas "
        "(\"promedio de días\", \"la más antigua\", \"qué porcentaje\"), montos (\"neto total de...\", "
        "\"costo del vale de consumo de...\"), detalle completo de una OT o patente (pega el folio o "
        "la patente), documentos (\"con/sin factura de garantía\"), OTs \"sin gestión\"/\"sin "
        "categoría\", \"repuestos\"/\"stock\", \"abono\"/\"cuenta ficha\", \"vehículos en el taller\" "
        "y \"agenda\"/\"citas de hoy\". Mantiene el hilo de la conversación (\"ordénalas por rango\", "
        "\"y en Curicó?\", \"quita el filtro de marca\", \"empecemos de nuevo\") y si no reconoce "
        "nada estructurado, busca las palabras en Glosa / Observación / Notas. Escribe \"ayuda\" en "
        "el chat para ver la guía completa. No incluye OTs cerradas ni historial anterior — solo lo "
        "que ya ve la app."
    )

    def _asis_indice_clientes(df_fuente):
        _clientes = {}
        for _, _ot in df_fuente.iterrows():
            _patente = str(_ot.get("PATENTE", "") or "").strip().upper()
            _rut_raw = str(_ot.get("rut_cliente", "") or "").strip()
            _ruts = [r.strip() for r in _rut_raw.split("/") if r.strip()] if _rut_raw else []
            _ant = _ot.get("anticipo", {}) or {}
            if isinstance(_ant, str):
                try:
                    _ant = json.loads(_ant)
                except Exception:
                    _ant = {}
            _ant_total = float(_ant.get("total", 0) or 0)
            _tiene_saldo = bool(_ant.get("tiene_saldo", False))
            _nombre_ant = str(_ant.get("nombre", "") or "").strip()
            _clave = _ruts[0] if _ruts else (_patente or "SIN_ID")
            _nombre_cli = _nombre_ant or (_ruts[0] if _ruts else _patente)
            if _clave not in _clientes:
                _clientes[_clave] = {
                    "nombre": _nombre_cli, "ruts": set(_ruts), "patentes": set(),
                    "ant_total": _ant_total, "tiene_saldo": _tiene_saldo,
                    "movimientos": _ant.get("movimientos", []),
                }
            else:
                _clientes[_clave]["ruts"].update(_ruts)
                if _ant_total > 0:
                    _clientes[_clave]["ant_total"] = _ant_total
                    _clientes[_clave]["tiene_saldo"] = _tiene_saldo
                    _clientes[_clave]["movimientos"] = _ant.get("movimientos", [])
                if _nombre_ant:
                    _clientes[_clave]["nombre"] = _nombre_ant
            if _patente:
                _clientes[_clave]["patentes"].add(_patente)
        return _clientes

    # Palabras de relleno conversacional a ignorar en la busqueda de stock —
    # sin esto, una frase natural tipo "hola necesito el stock de este
    # repuesto 80 621151" exigia que TODAS esas palabras (incluyendo "hola",
    # "necesito", "este") aparecieran en el producto/descripcion y no
    # encontraba nada, aunque el codigo 621151 si existiera. 08/07/2026.
    _ASIS_STOPWORDS_STOCK = {
        "hola", "necesito", "quiero", "dame", "puedes", "podrias", "podrías",
        "porfavor", "favor", "porfa", "este", "esta", "esa", "ese", "esas",
        "esos", "cual", "cuales", "cuál", "cuáles", "tiene", "tienen",
        "tienes", "hay", "para", "con", "del", "los", "las", "que", "decir",
        "dime", "una", "uno", "unos", "unas", "gracias", "saludos", "buenas",
        "buenos", "dias", "días", "tardes", "noches", "consulta", "consultar",
        "oye", "mira", "sabes", "saber", "cuanto", "cuánto", "cuantos",
        "cuántos", "cuanta", "cuánta", "cuantas", "cuántas", "por", "sera",
        "será", "existe", "queda", "quedan",
    }

    def _asis_buscar_repuestos(q, catalogo_stock):
        # Busca en el catalogo COMPLETO de Stock Repestos Costo.xlsx (subido por
        # el consolidador a stock_repuestos.json) — independiente de las OTs, a
        # diferencia de repuestos_compras (que solo trae lo que alguna vez se
        # pidio para una OT). 08/07/2026.
        _t = q.lower()
        for _kw in ["stock de", "stock del", "stock", "repuestos de", "repuesto de",
                    "repuestos", "repuesto", "en bodega", "bodega"]:
            _t = _t.replace(_kw, " ")
        # Tokeniza con regex (ignora signos de puntuacion tipo "?"/",") y
        # descarta palabras de relleno conversacional.
        _palabras = [w for w in _re_asis.findall(r"[a-záéíóúñ0-9\-]+", _t)
                     if len(w) >= 3 and w not in _ASIS_STOPWORDS_STOCK]
        _rows = []
        for _p in catalogo_stock:
            _texto_r = f"{_p.get('producto','')} {_p.get('descripcion','')}".lower()
            if _palabras and not all(pw in _texto_r for pw in _palabras):
                continue
            _rows.append({
                "Producto": _p.get("producto", ""),
                "Descripción": _p.get("descripcion", ""),
                "Stock": _p.get("stock", 0) or 0,
                "Stock Proyectado": _p.get("stock_proyectado", 0) or 0,
                "Bodega": _p.get("bodega", "") or "—",
                "Costo": _p.get("costo", 0) or 0,
                "Familia": _p.get("familia", ""),
                "Procedencia": _p.get("procedencia", ""),
            })
        return pd.DataFrame(_rows), _palabras

    # Documentos detectables por "con"/"sin" + palabra clave, mapeados a las
    # columnas N_* que ya calcula el consolidador (una por tipo de documento).
    # 08/07/2026.
    _ASIS_DOC_KEYWORDS = [
        ("factura de garantia", "FACT_GTIA"), ("factura garantia", "FACT_GTIA"),
        ("factura de cliente", "FACT_CLIENTE"), ("factura cliente", "FACT_CLIENTE"),
        ("factura de compania", "FACT_COMPANIA"), ("factura compania", "FACT_COMPANIA"),
        ("cargo de garantia", "CARGO_GTIA"), ("cargo garantia", "CARGO_GTIA"),
        ("cargo interno", "CARGO_INT"),
        ("liquidacion", "LIQ_ST"),
        ("vale de consumo", "VALE_CONSUMO"), ("vale consumo", "VALE_CONSUMO"),
    ]

    # Meses para filtrar por mes de FECHA OT ("OTs de marzo", "abiertas en enero").
    # 09/07/2026.
    _ASIS_MESES = {
        "enero": 1, "febrero": 2, "marzo": 3, "abril": 4, "mayo": 5, "junio": 6,
        "julio": 7, "agosto": 8, "septiembre": 9, "setiembre": 9, "octubre": 10,
        "noviembre": 11, "diciembre": 12,
    }

    # -- Tolerancia a typos (sección 5.1 de la Constitución) --------------------
    # Segundo pase, SOLO cuando el match exacto por palabra completa no encontró
    # nada para una dimensión: compara ventanas de N palabras consecutivas del
    # texto (N = cantidad de palabras del valor real) contra cada valor real vía
    # difflib.SequenceMatcher. Nunca se aplica en silencio — si hay match, se
    # avisa explícitamente "(interpretado de ...)"; si hay ambigüedad entre 2+
    # candidatos con score parecido, no se adivina, se avisa. 10/07/2026.
    def _asis_tokens(_t_norm_local):
        return [(m.group(), m.start(), m.end())
                for m in _re_asis.finditer(r"[a-z0-9]+", _t_norm_local)]

    def _asis_fuzzy_valor(_tokens, _val_norm, _cutoff=0.70):
        _val_words = _val_norm.split()
        _n = len(_val_words)
        if _n == 0 or len(_tokens) < _n:
            return None
        _mejor = None
        for _i in range(len(_tokens) - _n + 1):
            _ventana = " ".join(w for w, _, _ in _tokens[_i:_i + _n])
            _ratio = _difflib_asis.SequenceMatcher(None, _ventana, _val_norm).ratio()
            if _ratio >= _cutoff and (not _mejor or _ratio > _mejor[0]):
                _mejor = (_ratio, _tokens[_i][1], _tokens[_i + _n - 1][2], _ventana)
        return _mejor

    def _asis_buscar_fuzzy_dimension(_tokens, _valores_col):
        """Devuelve ('ok', (ratio,ini,fin,valor,texto)) si hay un candidato claro,
        ('ambiguo', [top3]) si hay 2+ candidatos con score muy parecido, o None."""
        _candidatos = []
        for _val in _valores_col:
            _val_norm = _asis_norm_txt(_val)
            if _val_norm.isdigit() or len(_val_norm) < 4:
                continue  # typo tolerance no aplica a valores muy cortos/numericos
            _res = _asis_fuzzy_valor(_tokens, _val_norm)
            if _res:
                _candidatos.append((_res[0], _res[1], _res[2], _val, _res[3]))
        if not _candidatos:
            return None
        _candidatos.sort(key=lambda x: -x[0])
        _top = _candidatos[0]
        if len(_candidatos) > 1 and (_top[0] - _candidatos[1][0]) < 0.03:
            return ("ambiguo", [c[3] for c in _candidatos[:3]])
        return ("ok", _top)

    # Diccionario de jerga/abreviaciones del taller: se reemplaza ANTES de
    # normalizar, así el resto de la detección (exacta o fuzzy) ve la forma
    # completa. Mantenido a mano con ejemplos reales que Cristóbal reporte.
    _ASIS_JERGA = {
        r"\bsucu\b": "sucursal",
        r"\bgtia\b": "garantia",
        r"\btec\b": "tecnico",
        r"\brepu\b": "repuesto",
    }

    def _asis_detectar_filtros(t):
        _filtros, _desc = {}, []
        for _patron_j, _reemplazo_j in _ASIS_JERGA.items():
            t = _re_asis.sub(_patron_j, _reemplazo_j, t, flags=_re_asis.IGNORECASE)
        _t_norm = _asis_norm_txt(t)
        _tokens_fuzzy = _asis_tokens(_t_norm)
        # Registra los tramos de texto ya "reclamados" por una dimension, para
        # que la misma palabra no pueda calzar en dos columnas distintas a la
        # vez (ej. dato sucio real: existe un MODELO literal "HYUNDAI" ademas
        # de la MARCA "HYUNDAI" — sin esto, una consulta con "Hyundai" arma
        # Marca=HYUNDAI Y Modelo=HYUNDAI al mismo tiempo, un filtro contradictorio
        # que nunca calza con ninguna OT). 08/07/2026.
        _tramos_usados = []
        # Zona de exclusión: cualquier valor de dimensión mencionado DESPUÉS de
        # "excepto"/"salvo"/"sin contar"/"excluyendo"/"que no sea(n)"/"menos las"
        # se interpreta como filtro NEGATIVO (excluir ese valor). Ej: "todas las
        # OT de Linderos excepto Hyundai". 09/07/2026.
        _mo_excl = _re_asis.search(
            r"\b(excepto|salvo|sin contar|excluye(?:ndo)?|que no sean?(?: de)?|"
            r"menos (?:las|los|la|el))\b", _t_norm)
        _excl_start = _mo_excl.start() if _mo_excl else len(_t_norm) + 1
        for _col, _key in [("SUCURSAL", "sucursal"), ("MARCA", "marca"),
                            ("TIPO VENTA", "tipo_venta"), ("ASESOR", "asesor"),
                            ("CATEGORIA", "categoria"), ("AÑO", "anio"),
                            ("MODELO", "modelo")]:
            # _matches_col acumula TODOS los valores positivos que calzan para esta
            # columna (no solo el primero) — permite "Hyundai o Ford"/"Linderos o
            # Talca" en la misma consulta (sección 5.2, condiciones OR dentro de una
            # misma dimensión). 10/07/2026.
            _matches_col = []
            for _val in sorted([v for v in df_raw[_col].astype(str).unique() if v.strip()],
                                key=len, reverse=True):
                # Match por palabra completa (\b...\b), no substring libre — con
                # CATEGORIA trayendo códigos de 2 letras reales ("CI", "CG", "ST")
                # un simple "in" hacia falsos positivos (ej. "ci" adentro de
                # "PlaCIlla"). El limite de palabra evita eso. 08/07/2026.
                _val_norm = _asis_norm_txt(_val)
                _pat_val = r"\b" + _re_asis.escape(_val_norm) + r"\b"
                if _val_norm.isdigit():
                    # Hay modelos que son puramente numéricos en los datos reales
                    # (ej. Peugeot "206"/"308"/"5008", Volvo "760") — sin este
                    # guard, "más de 100 días" detectaría Modelo=100 ademas de
                    # Días > 100. Se excluye si el número va seguido de "día(s)".
                    _pat_val += r"(?!\s*d[ií]as?\b)"
                _mo = _re_asis.search(_pat_val, _t_norm)
                if not _mo:
                    continue
                _ini, _fin = _mo.span()
                if any(not (_fin <= _s0 or _ini >= _s1) for (_s0, _s1) in _tramos_usados):
                    continue  # ese tramo de texto ya lo tomo otra dimension
                if _ini >= _excl_start:
                    _filtros.setdefault("excluir", {}).setdefault(_key, []).append(_val)
                    _desc.append(f"Excluye {_col.title()} = {_val}")
                    _tramos_usados.append((_ini, _fin))
                    continue
                _matches_col.append(_val)
                _tramos_usados.append((_ini, _fin))
            if _matches_col:
                if len(_matches_col) == 1:
                    _filtros[_key] = _matches_col[0]
                    _desc.append(f"{_col.title()} = {_matches_col[0]}")
                else:
                    _filtros[_key] = _matches_col
                    _desc.append(f"{_col.title()} = {' o '.join(_matches_col)}")
            # -- Fallback de tolerancia a typos (sección 5.1) --------------------
            # Solo si el match EXACTO no encontró nada para esta dimensión, y solo
            # sobre tramos de texto que ninguna otra dimensión ya reclamó.
            if not _matches_col:
                _valores_col = [v for v in df_raw[_col].astype(str).unique() if v.strip()]
                _res_fz = _asis_buscar_fuzzy_dimension(_tokens_fuzzy, _valores_col)
                if _res_fz and _res_fz[0] == "ok":
                    _ratio_fz, _ini_fz, _fin_fz, _val_fz, _texto_fz = _res_fz[1]
                    if not any(not (_fin_fz <= _s0 or _ini_fz >= _s1) for (_s0, _s1) in _tramos_usados):
                        _filtros[_key] = _val_fz
                        _desc.append(f'{_col.title()} = {_val_fz} (interpretado de "{_texto_fz}")')
                        _tramos_usados.append((_ini_fz, _fin_fz))
                elif _res_fz and _res_fz[0] == "ambiguo":
                    _desc.append(f"⚠️ No identifiqué con certeza el valor de {_col.title()} — "
                                 f"¿quisiste decir {' o '.join(_res_fz[1])}? Probá escribirlo completo.")
        # Manejo de errores más claro (sección 5.4): si detectamos una frase de
        # exclusión ("excepto"/"salvo"/...) pero ningún valor real quedó registrado
        # como excluido, avisamos en vez de ignorarlo en silencio — antes esto se
        # perdía sin ningún rastro para el usuario. 10/07/2026.
        if _mo_excl and not _filtros.get("excluir"):
            _desc.append("⚠️ Detecté una palabra de exclusión (\"excepto\"/\"salvo\"/...) pero no "
                         "reconocí qué valor querías excluir — revisa que esté escrito igual que "
                         "en los datos (sucursal, marca, asesor, categoría, año o modelo).")
        for _r in ["0-30", "31-60", "61-90", "91 o más", "91 o mas"]:
            if _r.replace(" ", "") in t.replace(" ", ""):
                _filtros["rango"] = "91 o más" if _r.startswith("91") else _r
                _desc.append(f"Rango = {_filtros['rango']}")
                break
        # Días de apertura — se opera sobre _t_norm (ya sin tildes) para no tener
        # que duplicar cada patrón con/sin acento. Se acepta un vocabulario amplio
        # de sinónimos ("sobre", "superior a", "mayor a", "arriba de", etc.) — un
        # asistente que solo entiende "más de N días" es demasiado frágil frente a
        # como la gente realmente pregunta. 08/07/2026.
        _me = _re_asis.search(r"entre\s+(\d+)\s*y\s*(\d+)\s*dias", _t_norm)
        if _me:
            _lo, _hi = int(_me.group(1)), int(_me.group(2))
            _filtros["dias_entre"] = (min(_lo, _hi), max(_lo, _hi))
            _desc.append(f"Días apertura entre {_lo} y {_hi}")
        else:
            _m = _re_asis.search(
                r"(?:mas de|sobre|superior a|por encima de|arriba de|mayor a|mayor que|"
                r"desde)\s+(\d+)\s*dias", _t_norm)
            if _m:
                _filtros["dias_mayor"] = int(_m.group(1))
                _desc.append(f"Días apertura > {_filtros['dias_mayor']}")
            _mn = _re_asis.search(
                r"(?:menos de|inferior a|por debajo de|bajo|menor a|menor que|"
                r"hasta)\s+(\d+)\s*dias", _t_norm)
            if _mn:
                _filtros["dias_menor"] = int(_mn.group(1))
                _desc.append(f"Días apertura < {_filtros['dias_menor']}")
            # Si la consulta menciona "dias" con un numero pero no calzo con
            # ninguno de los patrones de arriba, avisamos en vez de ignorarlo en
            # silencio (ej. "dias" sin ningun comparador reconocible).
            if not _m and not _mn and _re_asis.search(r"\d+\s*dias", _t_norm):
                _filtros["dias_no_reconocido"] = True
        # OTs sin gestión / sin categoría / sin notas / sin observación
        if "sin categoria" in _t_norm:
            _filtros["sin_categoria"] = True
            _desc.append("Sin categoría asignada")
        if "sin notas" in _t_norm:
            _filtros["sin_notas"] = True
            _desc.append("Sin notas")
        if "sin observacion" in _t_norm:
            _filtros["sin_observacion"] = True
            _desc.append("Sin observación OT")
        if any(k in _t_norm for k in ["sin gestion", "sin avance", "sin seguimiento"]):
            _filtros["sin_gestion"] = True
            _desc.append("Sin gestión (notas/observación/avance vacíos)")
        # Documentos presentes/ausentes (factura, vale de consumo, liquidación, etc.)
        for _kw, _dockey in sorted(_ASIS_DOC_KEYWORDS, key=lambda x: len(x[0]), reverse=True):
            if f"sin {_kw}" in _t_norm:
                _filtros["doc_sin"] = _dockey
                _desc.append(f"Sin {_kw}")
                break
            if f"con {_kw}" in _t_norm:
                _filtros["doc_con"] = _dockey
                _desc.append(f"Con {_kw}")
                break
        # Mes de apertura (sobre FECHA OT): "OTs de marzo", "abiertas en enero".
        for _mes_nombre, _mes_num in _ASIS_MESES.items():
            if _re_asis.search(r"\b" + _mes_nombre + r"\b", _t_norm):
                _filtros["mes"] = _mes_num
                _desc.append(f"Mes OT = {_mes_nombre.title()}")
                break
        _top = None
        _mt = _re_asis.search(r"top\s*(\d+)", t)
        if _mt:
            _top = int(_mt.group(1))
        elif "top" in t or "ranking" in t:
            _top = 5
        if _top:
            _filtros["top_n"] = _top
            _desc.append(f"Top {_top}")
            # "top 5 con menos OTs" / "los peores" → ranking ascendente. 09/07/2026.
            if any(k in _t_norm for k in ["con menos", "con menor", "mas bajo", "mas bajas",
                                           "mas bajos", "peores", "ultimos"]):
                _filtros["top_asc"] = True
                _desc.append("(de menor a mayor)")
        _filtros["es_conteo"] = any(k in t for k in
                                    ["cuant", "cuánt", "cantidad de", "número de", "numero de"])
        return _filtros, _desc

    def _asis_primer_valor(_v):
        # Varias ramas puntuales (vehículos en taller, agenda, técnicos, ranking/
        # historial de cierres) solo soportan UN valor de sucursal/asesor a la vez
        # — si el usuario pidió una condición OR ("Linderos o Talca"), esas ramas
        # usan el primero y listo (no es su caso de uso principal). El motor de
        # filtros genérico (_asis_match_dim) sí soporta la lista completa. 10/07/2026.
        return _v[0] if isinstance(_v, list) else _v

    def _asis_match_dim(_df, _col, _val):
        # _val puede ser un string (un solo valor) o una lista (condición OR
        # dentro de la misma dimensión, ej. Marca = Hyundai o Ford). 10/07/2026.
        if isinstance(_val, list):
            return _df[_df[_col].isin(_val)]
        return _df[_df[_col] == _val]

    def _asis_aplicar_filtros(_filtros):
        _df = df_raw.copy()
        if _filtros.get("sucursal"):
            _df = _asis_match_dim(_df, "SUCURSAL", _filtros["sucursal"])
        if _filtros.get("marca"):
            _df = _asis_match_dim(_df, "MARCA", _filtros["marca"])
        if _filtros.get("tipo_venta"):
            _df = _asis_match_dim(_df, "TIPO VENTA", _filtros["tipo_venta"])
        if _filtros.get("asesor"):
            _df = _asis_match_dim(_df, "ASESOR", _filtros["asesor"])
        if _filtros.get("categoria"):
            _df = _asis_match_dim(_df, "CATEGORIA", _filtros["categoria"])
        if _filtros.get("anio"):
            _df = _asis_match_dim(_df, "AÑO", _filtros["anio"])
        if _filtros.get("modelo"):
            _vm = _filtros["modelo"]
            if isinstance(_vm, list):
                _df = _df[_df["MODELO"].astype(str).str.upper().isin([str(x).upper() for x in _vm])]
            else:
                _df = _df[_df["MODELO"].astype(str).str.upper() == str(_vm).upper()]
        if _filtros.get("rango"):
            _df = _df[_df["RANGO"] == _filtros["rango"]]
        if _filtros.get("dias_mayor") is not None:
            _dd = pd.to_numeric(_df["DIAS APERTURA"], errors="coerce").fillna(0)
            _df = _df[_dd > _filtros["dias_mayor"]]
        if _filtros.get("dias_menor") is not None:
            _dd = pd.to_numeric(_df["DIAS APERTURA"], errors="coerce").fillna(0)
            _df = _df[_dd < _filtros["dias_menor"]]
        if _filtros.get("dias_entre") is not None:
            _lo, _hi = _filtros["dias_entre"]
            _dd = pd.to_numeric(_df["DIAS APERTURA"], errors="coerce").fillna(0)
            _df = _df[(_dd >= _lo) & (_dd <= _hi)]
        if _filtros.get("sin_categoria"):
            _df = _df[_df["CATEGORIA"].astype(str).str.strip().isin(["", "Sin categoría"])]
        if _filtros.get("sin_notas"):
            _df = _df[_df["NOTAS"].astype(str).str.strip() == ""]
        if _filtros.get("sin_observacion"):
            _df = _df[_df["OBSERVACION OT"].astype(str).str.strip() == ""]
        if _filtros.get("sin_gestion"):
            _df = _df[
                (_df["NOTAS"].astype(str).str.strip() == "")
                & (_df["OBSERVACION OT"].astype(str).str.strip() == "")
                & (_df["AVANCE - GESTIÓN"].astype(str).str.strip() == "")
            ]
        if _filtros.get("doc_con"):
            _col_n = f"N_{_filtros['doc_con']}"
            if _col_n in _df.columns:
                _df = _df[pd.to_numeric(_df[_col_n], errors="coerce").fillna(0) > 0]
        if _filtros.get("doc_sin"):
            _col_n = f"N_{_filtros['doc_sin']}"
            if _col_n in _df.columns:
                _df = _df[pd.to_numeric(_df[_col_n], errors="coerce").fillna(0) == 0]
        if _filtros.get("mes") and "FECHA OT" in _df.columns:
            _fechas_m = pd.to_datetime(_df["FECHA OT"], dayfirst=True, errors="coerce")
            _df = _df[_fechas_m.dt.month == _filtros["mes"]]
        # Exclusiones ("excepto X"): filtro negativo por valor de dimensión. 09/07/2026.
        for _exc_key, _exc_vals in (_filtros.get("excluir") or {}).items():
            _exc_col = {"sucursal": "SUCURSAL", "marca": "MARCA", "tipo_venta": "TIPO VENTA",
                        "asesor": "ASESOR", "categoria": "CATEGORIA", "anio": "AÑO",
                        "modelo": "MODELO"}.get(_exc_key)
            if _exc_col and _exc_col in _df.columns:
                _df = _df[~_df[_exc_col].isin(_exc_vals)]
        return _df

    # Columnas que se pueden pedir como criterio de orden ("ordenalas por rango",
    # "ordename por dias", "de mayor a menor neto", etc.). 08/07/2026.
    _ASIS_ORDEN_COLS = [
        ("dias de apertura", "DIAS APERTURA"), ("dias apertura", "DIAS APERTURA"),
        ("dias", "DIAS APERTURA"), ("rango", "RANGO"), ("neto", "NETO"),
        ("asesor", "ASESOR"), ("marca", "MARCA"), ("sucursal", "SUCURSAL"),
        ("modelo", "MODELO"), ("categoria", "CATEGORIA"), ("folio", "FOLIO OT"),
        ("patente", "PATENTE"), ("tipo de venta", "TIPO VENTA"), ("año", "AÑO"),
        ("ano", "AÑO"),
    ]

    def _asis_detectar_orden(_t_norm):
        _pide = any(k in _t_norm for k in [
            "ordena", "ordename", "ordenalas", "ordenalos", "ordenalo", "ordenar",
            "orden por", "de mayor a menor", "de menor a mayor",
        ])
        if not _pide:
            return None, False, False
        _col = None
        for _kw, _colname in sorted(_ASIS_ORDEN_COLS, key=lambda x: len(x[0]), reverse=True):
            if _kw in _t_norm:
                _col = _colname
                break
        _asc = any(k in _t_norm for k in ["ascendente", "menor a mayor", "de menor"])
        return _col, _asc, True

    def _asis_aplicar_orden(_df, _orden_col, _orden_asc):
        if _df.empty or not _orden_col or _orden_col not in _df.columns:
            return _df
        if _orden_col == "RANGO":
            _orden_map = {"0-30": 0, "31-60": 1, "61-90": 2, "91 o más": 3}
            return (_df.assign(_ord_=_df["RANGO"].map(_orden_map).fillna(99))
                        .sort_values("_ord_", ascending=_orden_asc).drop(columns="_ord_"))
        if _orden_col == "DIAS APERTURA":
            return (_df.assign(_ord_=pd.to_numeric(_df["DIAS APERTURA"], errors="coerce").fillna(0))
                        .sort_values("_ord_", ascending=_orden_asc).drop(columns="_ord_"))
        if _orden_col == "NETO":
            return _df.sort_values("NETO", ascending=_orden_asc)
        return _df.sort_values(_orden_col, ascending=_orden_asc, na_position="last")

    # Dimensiones reconocibles para desgloses ("por sucursal") y para listar/
    # contar valores distintos ("qué marcas hay", "cuántos asesores"). 09/07/2026.
    _ASIS_DESGLOSE_COLS = {
        "sucursal": "SUCURSAL", "asesor": "ASESOR", "marca": "MARCA", "rango": "RANGO",
        "tipo de venta": "TIPO VENTA", "tipo venta": "TIPO VENTA",
        "categoria": "CATEGORIA", "año": "AÑO", "ano": "AÑO", "modelo": "MODELO",
    }
    _ASIS_PLURAL_COLS = {
        "sucursales": "SUCURSAL", "asesores": "ASESOR", "marcas": "MARCA",
        "modelos": "MODELO", "categorias": "CATEGORIA", "patentes": "PATENTE",
        "clientes": "PATENTE", "vehiculos": "PATENTE", "tipos de venta": "TIPO VENTA",
        "años": "AÑO", "anos": "AÑO",
    }

    # Guía de capacidades para "qué puedes hacer" / "ayuda". 09/07/2026.
    _ASIS_AYUDA_MD = (
        "**Esto es lo que puedo responder con los datos de la app:**\n\n"
        "- **Conteos con filtros combinables**: *\"cuántas OT tiene Linderos de Hyundai con más de "
        "60 días\"* — reconozco sucursal, asesor, marca, modelo, tipo de venta, categoría, año, mes, "
        "rango (0-30/31-60/61-90/91 o más), días (*\"más de/menos de/entre X e Y/sobre N días\"*) y "
        "exclusiones (*\"...excepto Ford\"*).\n"
        "- **Listados y orden**: *\"dame los números de OT\"*, *\"ordénalas por días\"*, *\"de menor a mayor neto\"*.\n"
        "- **Desgloses**: *\"cuántas OT por sucursal\"*, *\"desglose por rango\"*, *\"por mes\"*.\n"
        "- **Rankings**: *\"top 5 asesores\"*, *\"top 3 sucursales por neto total\"*, *\"top 5 con menos OTs\"*.\n"
        "- **Estadísticas**: *\"promedio de días de Placilla\"*, *\"la OT más antigua\"*, *\"las más caras\"*, "
        "*\"qué porcentaje son de garantía\"*.\n"
        "- **Montos**: *\"neto total de Linderos\"*, *\"costo del vale de consumo de Talca\"*.\n"
        "- **Detalle de una OT o patente**: pega el folio o la patente (*\"1142225\"* o *\"detalle de VXWZ73\"*) "
        "y te muestro todo: datos, documentos, gestión, repuestos y comentarios.\n"
        "- **Comparaciones**: *\"compara Linderos con Talca\"*, *\"Hyundai vs Ford\"*.\n"
        "- **Dimensiones**: *\"qué sucursales hay\"*, *\"cuántos asesores tiene Curicó\"*.\n"
        "- **Documentos y gestión**: *\"OT de garantía sin gestión\"*, *\"sin factura de cliente\"*, *\"sin vale de consumo\"*.\n"
        "- **Stock de repuestos**: *\"stock filtro de aceite\"*, *\"stock del repuesto 621151\"*.\n"
        "- **Abonos / Cuenta Ficha**: *\"abono Juan Pérez\"*, *\"saldo de la patente VXWZ73\"*.\n"
        "- **Taller y agenda**: *\"vehículos en el taller de Chillán\"*, *\"citas de hoy en Talca\"*.\n"
        "- **Texto libre**: cualquier palabra del trabajo (*\"choque\"*, *\"parabrisas\"*) la busco en "
        "Glosa / Observación / Notas.\n\n"
        "**Y mantengo el hilo de la conversación**: después de una consulta puedes decir "
        "*\"ordénalas por rango\"*, *\"de esas dame las de más de 60 días\"*, *\"y en Curicó?\"*, "
        "*\"quita el filtro de marca\"* o *\"empecemos de nuevo\"*.\n\n"
        "**También te puedo explicar cómo usar la app**: preguntame *\"¿cómo uso el "
        "planificador?\"*, *\"¿qué es Facturas X?\"* o *\"¿cómo cambio mi contraseña?\"* y te "
        "explico paso a paso."
    )

    # -- Base de conocimiento de "meta-ayuda": cómo USAR la app (no los datos). --
    # Ver CONSTITUCION_ASISTENTE_APP.md sección 5.5. IMPORTANTE: cada vez que un
    # módulo cambia de comportamiento en una sesión futura, hay que actualizar la
    # entrada correspondiente acá — un manual desactualizado es peor que ninguno.
    _ASIS_MANUAL_APP = {
        "control_gestion": {
            "keywords": ["control y gestion post venta", "gestion post venta",
                         "modulo principal", "dashboard principal", "dashboard de ots"],
            "titulo": "📋 Control y Gestión Post Venta",
            "texto": (
                "Es el módulo principal — el dashboard completo de OTs pendientes. Tiene estas "
                "pestañas: **Resumen** (KPIs generales), **Detalle y Edición** (tabla editable), "
                "**Documentos y Comentarios**, **Historial de Comentarios**, **Por Sucursal**, "
                "**Por Asesor**, **Análisis y Tendencias**, **Ranking Cierres >90d**, "
                "**Notificaciones**, **Repuestos Pendientes**, **Patentes a Contactar**, "
                "**Búsqueda de Cliente**, **Facturas X** y **Admin** (si sos administrador). "
                "Pregúntame por el nombre de cualquiera de esas pestañas si querés el detalle."
            ),
        },
        "planificador": {
            "keywords": ["planificador de taller", "planificador", "tablero de taller",
                         "jpcb", "tecnico x hora", "técnico x hora", "agenda curifor",
                         "control de taller", "vehiculos en taller", "vehículos en taller",
                         "historial de taller"],
            "titulo": "🔄 Planificador de Taller",
            "texto": (
                "Tiene 5 vistas, en pestañas arriba (elegí primero la Sucursal/Taller en el panel izquierdo):\n\n"
                "- **JPCB por patente**: kanban de 8 etapas (Recepción → Pedido Repuestos → "
                "Arribo Repuestos → En Proceso → Espera → Control de Calidad → Lavado → "
                "Entrega). Arrastra la tarjeta de un vehículo entre columnas para cambiar su "
                "etapa.\n"
                "- **Planificador (Técnico x Hora)**: agenda diaria por técnico. Arrastra una "
                "cita desde el panel izquierdo 'Programación' hacia una celda del técnico para "
                "asignarla — el bloque se ajusta solo si hay tempario (horas de mano de obra "
                "de la mantención) y evita superponerse con otro vehículo del mismo técnico.\n"
                "- **Control de Taller**: tabla editable con todos los vehículos en el taller "
                "(patente, modelo, mantención, técnico, horario, etapa, comentarios, número de "
                "caso, ETA, auto de reemplazo). Los cambios mueven la tarjeta en el JPCB solos. "
                "Las patentes con cita marcada 🎟️ Ingresado en la Agenda de hoy se agregan "
                "solas; también podés agregar una patente manual con el botón correspondiente.\n"
                "- **Vehículos en Taller**: igual que Control de Taller pero solo los últimos "
                "60 días sin salida registrada.\n"
                "- **Historial de Taller**: órdenes ya cerradas (botón 🔒 Cerrar en Control de "
                "Taller), de solo lectura, con botón ↩️ Reabrir si se cerró por error.\n\n"
                "Acceso restringido: solo admin y usuarios autorizados (Admin → Permisos de "
                "módulos → Planificador)."
            ),
        },
        "indicadores": {
            "keywords": ["indicadores post venta", "indicadores", "power bi",
                         "informe power bi", "kpis de facturacion"],
            "titulo": "📈 Indicadores Post Venta",
            "texto": (
                "Informe de Power BI embebido directo en la app, sin necesidad de cuenta "
                "Microsoft (carga solo). Tiene 6 páginas — se navega con las flechas o el menú "
                "inferior del informe: **1** Post Venta General, **2** Servicio Técnico, "
                "**3** DyP, **4** Avance Facturación, **5** Venta Repuestos, "
                "**6** Pronóstico Ventas.\n\n"
                "Acceso restringido: solo admin y usuarios autorizados (Admin → Permisos de "
                "módulos → Indicadores)."
            ),
        },
        "asistente_app": {
            "keywords": ["asistente app", "este chat", "el asistente", "el chat"],
            "titulo": "🤖 Asistente App",
            "texto": (
                "Este mismo chat. Responde preguntas sobre los datos de la app (conteos, "
                "filtros combinados, rankings, montos, stock de repuestos, abonos de clientes, "
                "detalle de una OT/patente, comparaciones) y también, como esta que me hiciste, "
                "sobre **cómo usar la app**. Escribe **\"ayuda\"** para ver todo lo que puedo "
                "responder sobre los datos."
            ),
        },
        "detalle_edicion": {
            "keywords": ["detalle y edicion", "editar categoria", "tabla editable",
                         "editar una ot", "cambiar categoria"],
            "titulo": "✏️ Detalle y Edición",
            "texto": (
                "Tabla editable con todas las OT pendientes. Editables: **Categoría**, "
                "**Observación OT**, **Notas**, **Avance/Gestión** y el color de marca "
                "(🔴🟡🟢🔵). Se guarda solo al editar la celda (auto-guardado, con aviso de "
                "éxito). Marca el checkbox 📄 de una fila y presiona el botón para saltar "
                "directo a Documentos y Comentarios de esa OT."
            ),
        },
        "documentos_comentarios": {
            "keywords": ["documentos y comentarios", "agregar comentario",
                         "mencionar usuario", "ficha de la ot", "ficha de una ot"],
            "titulo": "📄 Documentos y Comentarios",
            "texto": (
                "Ficha completa de una OT puntual: sus documentos asociados (liquidación, "
                "facturas, cargos, vale de consumo), estado de gestión (categoría/observación/"
                "notas/avance), repuestos del Vale de Consumo y de Seguimiento de Compras "
                "(en bodega / en espera), y el historial de comentarios. Podés escribir un "
                "comentario nuevo y **mencionar a otro usuario** — le llega una notificación 🔔."
            ),
        },
        "historial_comentarios": {
            "keywords": ["historial de comentarios", "log de comentarios",
                         "todos los comentarios"],
            "titulo": "📋 Historial de Comentarios",
            "texto": (
                "Sábana completa de todos los comentarios de todas las OT, filtrable por "
                "folio, autor o texto, y descargable como CSV."
            ),
        },
        "por_sucursal": {
            "keywords": ["por sucursal", "resumen por sucursal"],
            "titulo": "🏢 Por Sucursal",
            "texto": "Tabla y gráfico de cantidad de OT pendientes por sucursal.",
        },
        "por_asesor": {
            "keywords": ["por asesor", "resumen por asesor"],
            "titulo": "👤 Por Asesor",
            "texto": "Tabla y gráfico de cantidad de OT pendientes por asesor.",
        },
        "analisis_tendencias": {
            "keywords": ["analisis y tendencias", "historial de cierres", "tendencias"],
            "titulo": "📈 Análisis y Tendencias",
            "texto": "Historial de cierres de OT a lo largo del tiempo, por sucursal/período.",
        },
        "ranking_cierres": {
            "keywords": ["ranking cierres", "ranking de cierres", "cierres mayor a 90"],
            "titulo": "🏆 Ranking Cierres >90d",
            "texto": "Ranking de asesores/sucursales por cantidad de OT críticas (>90 días) cerradas.",
        },
        "notificaciones": {
            "keywords": ["notificaciones", "menciones", "campanita"],
            "titulo": "🔔 Notificaciones",
            "texto": (
                "Notificaciones de menciones que te hizo otro usuario en un comentario. El "
                "número junto a la campanita 🔔 en la barra lateral indica cuántas no leíste — "
                "se marcan como leídas automáticamente al abrir esta pestaña."
            ),
        },
        "repuestos_pendientes": {
            "keywords": ["repuestos pendientes", "repuestos en bodega", "repuestos en espera"],
            "titulo": "🧰 Repuestos Pendientes",
            "texto": (
                "Lista todos los repuestos en bodega (ya llegaron, Cerrado por sistema) de "
                "todas las OT pendientes, cruzados contra el Vale de Consumo para no mostrar lo "
                "que ya está instalado. Muestra si el repuesto se pidió directo o vía la patente "
                "(cuando se pidió con una OT ya cerrada del mismo vehículo)."
            ),
        },
        "patentes_a_contactar": {
            "keywords": ["patentes a contactar", "contactar cliente", "avisar cliente repuesto"],
            "titulo": "📞 Patentes a Contactar",
            "texto": (
                "Agrupa por patente los vehículos con repuestos en bodega pendientes de "
                "instalar, para que el taller contacte al cliente y coordine la instalación."
            ),
        },
        "busqueda_cliente": {
            "keywords": ["busqueda de cliente", "buscar cliente", "cuenta ficha", "anticipos",
                         "abono", "abonos", "saldo de cliente", "saldo del cliente"],
            "titulo": "🔍 Búsqueda de Cliente",
            "texto": (
                "Busca por nombre, RUT o patente y muestra las OT asociadas, el saldo de la "
                "cuenta ficha (anticipos) y sus movimientos. Solo muestra resultados cuando "
                "escribís algo en el buscador.\n\n"
                "**En este mismo chat** también podés preguntar directo *\"abono Juan Pérez\"* "
                "o *\"saldo de la patente VXWZ73\"* y te muestro el resultado sin ir a la pestaña."
            ),
        },
        "facturas_x": {
            "keywords": ["facturas x", "factura x"],
            "titulo": "🧾 Facturas X",
            "texto": (
                "Lista las facturas con prefijo X, con el abono de cliente disponible, fecha "
                "de factura y fecha de anticipo. Podés marcar el checkbox 📌 de una fila y "
                "saltar directo a Documentos y Comentarios de la OT asociada."
            ),
        },
        "admin": {
            "keywords": ["panel de administracion", "permisos de modulos",
                         "autorizar usuario", "gestionar usuarios", "seccion admin"],
            "titulo": "🛡️ Admin",
            "texto": (
                "Solo visible para el administrador. Gestión de usuarios registrados y sus "
                "permisos de módulos (Control y Gestión, Planificador, Indicadores, Asistente "
                "App) con una tabla de checkboxes + botón Guardar, y configuración de técnicos "
                "por sucursal."
            ),
        },
        "login": {
            "keywords": ["contraseña", "contrasena", "iniciar sesion", "inicio sesion"],
            "titulo": "🔑 Acceso y contraseña",
            "texto": (
                "Ingresás con tu correo @curifor.com. La primera vez creás tu propia "
                "contraseña; las siguientes veces solo la escribís. Para cambiarla, andá al "
                "menú lateral (☰) → 🔑 Cambiar contraseña, dentro del módulo Control y Gestión "
                "Post Venta."
            ),
        },
    }

    # Frases que indican intención real de "cómo se usa X" — se exige al menos una
    # de estas ADEMÁS del nombre del módulo, para no disparar la meta-ayuda por un
    # simple mención de pasada (ej. una futura consulta de datos que solo nombre
    # "facturas x" sin pedir explicación de uso).
    _ASIS_INTENCION_COMO = [
        "como uso", "como se usa", "como funciona", "como puedo", "como hago",
        "que es", "que es el", "que es la", "que hace", "para que sirve",
        "explicame", "explica", "ayuda con", "manual de", "guia de",
        "instrucciones de", "donde encuentro", "donde esta", "donde veo",
        "como entro", "como accedo", "como cierro", "como agrego", "como asigno",
        "como cambio", "como edito", "como creo", "como reabro", "como cambia",
        "como veo", "como puedo ver", "de que manera", "de que forma",
        "me podrias explicar", "me puedes explicar", "no se como", "no entiendo como",
        "tengo dudas de como", "tengo dudas sobre como", "cuentame como", "dime como",
        "dime que es",
    ]
    # Nota: se sacaron a propósito frases genéricas tipo "quiero saber"/"necesito
    # saber"/"información sobre" — colisionaban con consultas de datos reales
    # formuladas educadamente (ej. "necesito saber el abono de Juan Pérez" NO es
    # una pregunta de "cómo se usa", es una consulta de datos real). 10/07/2026,
    # bug real reportado por Cristóbal: "¿cómo puedo ver abono de clientes?" caía
    # en la búsqueda real de clientes en vez de la meta-ayuda porque "abono" no
    # estaba entre las keywords de busqueda_cliente (ya corregido arriba).

    def _asis_buscar_manual_app(_t_norm_local):
        """Busca módulos de la app cuyas keywords calcen en el texto — solo si hay
        una frase de intención de "cómo se usa" (ver _ASIS_INTENCION_COMO), para
        no competir con futuras consultas de datos que solo mencionen el nombre
        de un módulo de pasada."""
        if not any(_fr in _t_norm_local for _fr in _ASIS_INTENCION_COMO):
            return []
        _matches, _tramos_m = [], []
        for _info in _ASIS_MANUAL_APP.values():
            for _kw in sorted(_info["keywords"], key=len, reverse=True):
                _kw_norm = _asis_norm_txt(_kw)
                _mo_m = _re_asis.search(r"\b" + _re_asis.escape(_kw_norm) + r"\b", _t_norm_local)
                if not _mo_m:
                    continue
                _ini_m, _fin_m = _mo_m.span()
                if any(not (_fin_m <= _s0 or _ini_m >= _s1) for (_s0, _s1) in _tramos_m):
                    continue
                _matches.append(_info)
                _tramos_m.append((_ini_m, _fin_m))
                break  # una keyword que calce por módulo alcanza
        return _matches

    def _asis_bloques_detalle_ot(_m):
        """Arma la ficha completa de una OT (fila de df_raw) como bloques de chat."""
        _bl = []
        _folio_d = str(_m.get("FOLIO OT", "")).strip()
        try:
            _neto_d = float(_m.get("NETO", 0) or 0)
        except Exception:
            _neto_d = 0.0
        _vale_d = _asis_costo_vale(_m.get("repuestos_actual"))
        _bl.append({"type": "markdown", "text":
            f"**OT {_folio_d}** · {_m.get('SUCURSAL','—')} · {_m.get('MARCA','—')} "
            f"{_m.get('MODELO','')} · Patente **{_m.get('PATENTE','—')}**\n\n"
            f"- **Asesor:** {_m.get('ASESOR','—')}\n"
            f"- **Tipo de venta:** {_m.get('TIPO VENTA','—')} · **Categoría:** "
            f"{str(_m.get('CATEGORIA','') or '').strip() or 'Sin categoría'}\n"
            f"- **Fecha OT:** {_m.get('FECHA OT','—')} · **Año:** {_m.get('AÑO','—')}\n"
            f"- **Glosa:** {str(_m.get('GLOSA TRABAJO','') or '').strip() or '—'}"})
        _bl.append({"type": "metrics", "items": [
            ("Días apertura", _m.get("DIAS APERTURA", "—")),
            ("Rango", _m.get("RANGO", "—")),
            ("NETO", f"${_neto_d:,.0f}"),
            ("Vale de Consumo", f"${_vale_d:,.0f}"),
        ]})
        _docs_txt = []
        for _tipo_doc, _key_doc in DOCS_CONFIG:
            try:
                _n_doc = int(float(_m.get(f"N_{_key_doc}", 0) or 0))
            except Exception:
                _n_doc = 0
            if _n_doc > 0:
                _fols_doc = str(_m.get(f"FOLIOS_{_key_doc}", "") or "").strip()
                _docs_txt.append(f"**{_tipo_doc}** ({_n_doc}): {_fols_doc or '—'}")
        _bl.append({"type": "markdown", "text":
            "📄 **Documentos:** " + (" · ".join(_docs_txt) if _docs_txt
                                     else "sin documentos registrados")})
        _gest_txt = []
        for _lbl_g, _col_g in [("Observación", "OBSERVACION OT"), ("Notas", "NOTAS"),
                                ("Avance/Gestión", "AVANCE - GESTIÓN")]:
            _v_g = str(_m.get(_col_g, "") or "").strip()
            if _v_g:
                _gest_txt.append(f"**{_lbl_g}:** {_v_g}")
        _bl.append({"type": "markdown", "text":
            "🗂️ " + ("  \n".join(_gest_txt) if _gest_txt
                     else "Sin gestión registrada (observación / notas / avance vacíos)")})
        _reps_a = _m.get("repuestos_actual")
        if isinstance(_reps_a, list) and _reps_a:
            _bl.append({"type": "markdown", "text": "🔧 **Repuestos del Vale de Consumo:**"})
            _bl.append({"type": "dataframe", "df": pd.DataFrame(_reps_a)})
        _reps_c = _m.get("repuestos_compras")
        if isinstance(_reps_c, list) and _reps_c:
            _n_bod_d, _n_esp_d = _asis_conteo_repuestos(_reps_c)
            _bl.append({"type": "markdown", "text":
                f"📦 **Seguimiento de Compras:** {_n_bod_d} repuesto(s) en bodega · "
                f"{_n_esp_d} en espera"})
        try:
            _df_coms_d = cargar_comentarios()
            if not _df_coms_d.empty and "folio_ot" in _df_coms_d.columns:
                _coms_ot_d = _df_coms_d[_df_coms_d["folio_ot"].astype(str) == _folio_d]
                if not _coms_ot_d.empty:
                    _bl.append({"type": "markdown",
                                 "text": f"💬 **Comentarios ({len(_coms_ot_d)}):**"})
                    for _, _cm in (_coms_ot_d.sort_values("fecha", ascending=False)
                                   .head(5).iterrows()):
                        _bl.append({"type": "markdown", "text":
                            f"> 🗓 {_cm.get('fecha','—')} · 👤 {_cm.get('autor','—')}: "
                            f"{_cm.get('comentario','')}"})
        except Exception:
            pass
        return _bl

    # -- Construcción de la respuesta como "bloques" (texto/metricas/tablas), --
    # -- para poder pintarla igual sea la primera vez o al re-renderizar el   --
    # -- historial del chat despues de un rerun.                             --
    def _asis_procesar_consulta(_q_raw):
        _t = _q_raw.strip().lower()
        _t_norm = _asis_norm_txt(_t)
        _bloques = []

        # ---- Meta-ayuda: cómo usar un módulo específico de la app -------------
        # (CONSTITUCION_ASISTENTE_APP.md sección 5.5). Se evalúa ANTES que la ayuda
        # genérica de abajo porque es más específica — si preguntan "cómo uso el
        # planificador" no tiene sentido devolver la lista genérica de consultas de
        # datos, aunque "como se usa"/"como funciona" también matcheen ahí.
        _manual_matches = _asis_buscar_manual_app(_t_norm)
        if _manual_matches:
            for _i_mm, _info_mm in enumerate(_manual_matches):
                if _i_mm > 0:
                    _bloques.append({"type": "divider"})
                _bloques.append({"type": "markdown",
                                  "text": f"### {_info_mm['titulo']}\n\n{_info_mm['texto']}"})
            return _bloques

        # ---- Ayuda: "qué puedes hacer" / "ayuda" ----------------------------
        if any(k in _t_norm for k in ["que puedes hacer", "que sabes hacer", "ayuda",
                                       "como te uso", "como se usa", "como funciona",
                                       "que consultas", "que preguntas", "instrucciones",
                                       "capacidades", "help", "que puedo preguntar"]):
            _bloques.append({"type": "markdown", "text": _ASIS_AYUDA_MD})
            return _bloques

        # ---- Reset explícito del contexto de conversación -------------------
        if any(k in _t_norm for k in ["olvida el filtro", "olvida los filtros",
                                       "olvida el contexto", "borra el contexto",
                                       "limpia el contexto", "limpia los filtros",
                                       "empecemos de nuevo", "partamos de cero",
                                       "empezar de nuevo", "borron y cuenta nueva",
                                       "resetea", "desde cero"]):
            for _k_ctx in ["asis_ultimo_filtro", "asis_ultimo_desc", "asis_ultimo_orden"]:
                st.session_state.pop(_k_ctx, None)
            _bloques.append({"type": "markdown", "text":
                "Listo ✅ — olvidé el contexto de la conversación. "
                "Tu próxima consulta parte de cero."})
            return _bloques

        # ---- Flags: detalle de OT/patente y comparación ----------------------
        # Detalle: si el mensaje trae un folio o patente REAL (validado contra
        # los datos) y pide detalle — o si el mensaje es básicamente solo ese
        # token — se muestra la ficha completa de la OT. 09/07/2026.
        _es_comparacion = any(k in _t_norm for k in [" vs ", " versus ", "compara",
                                                      "comparar", "comparacion",
                                                      "diferencia entre"])
        _folios_set = set(df_raw["FOLIO OT"].astype(str).str.strip())
        _patentes_set = set(df_raw["PATENTE"].astype(str).str.strip().str.upper()
                            .str.replace(" ", ""))
        _patentes_set.discard("")
        # Folios reales pueden tener desde 4 dígitos (ej. 2404) — se valida
        # contra el set real, así un año/modelo suelto no confunde. 09/07/2026.
        _tokens_det = _re_asis.findall(r"[a-z0-9]{4,8}", _t_norm)
        _folio_hit = next((tk for tk in _tokens_det
                           if tk.isdigit() and tk in _folios_set), None)
        _pat_hit = next((tk.upper() for tk in _tokens_det
                         if not tk.isdigit() and len(tk) >= 5
                         and tk.upper() in _patentes_set), None)
        _pide_detalle = any(k in _t_norm for k in [
            "detalle", "informacion", "info de", "datos de", "ficha de", "todo sobre",
            "que sabes de", "que tiene la ot", "estado de la ot", "como va la ot",
            "como esta la ot", "cuentame de",
        ])
        _resto_det = _t_norm
        for _tk_d in _tokens_det:
            _resto_det = _resto_det.replace(_tk_d, " ")
        _resto_words = [w for w in _re_asis.findall(r"[a-z]+", _resto_det)
                        if w not in {"hola", "dame", "la", "las", "el", "los", "de", "del",
                                     "ot", "ots", "folio", "folios", "patente", "patentes",
                                     "por", "favor", "me", "muestra", "muestrame", "ver",
                                     "y", "o", "que", "con", "revisa", "busca", "buscame"}]
        _solo_tokens = bool((_folio_hit or _pat_hit) and not _resto_words)
        _es_detalle = bool((_folio_hit or _pat_hit) and (_pide_detalle or _solo_tokens))

        # Montos: "neto total de...", "costo del vale de consumo de...", "promedio
        # de neto de...", etc. Se evalua primero porque tiene prioridad sobre el
        # resto de las ramas (ninguna otra rama entiende de sumas/promedios).
        # 08/07/2026.
        _quiere_neto = "neto" in _t_norm and any(k in _t_norm for k in
                                                  ["total", "promedio", "suma", "cuanto", "monto"])
        _quiere_vale = any(k in _t_norm for k in ["vale de consumo", "vale consumo"]) and any(
            k in _t_norm for k in ["total", "promedio", "suma", "costo", "cuanto"])

        if (_quiere_neto or _quiere_vale) and not _es_comparacion and not _es_detalle:
            _filtros, _desc = _asis_detectar_filtros(_t)
            # Continuidad también en montos: "¿y cuánto suman esas en neto?" tras
            # un listado reutiliza el filtro del turno anterior. 09/07/2026.
            _ref_prev_m = any(k in _t_norm for k in ["esas", "esos", "estas", "estos",
                                                      "de ellas", "de ellos", "anterior",
                                                      "mismas", "mismos", "ese listado",
                                                      "esa lista", "las de antes"])
            if not _desc and _ref_prev_m and st.session_state.get("asis_ultimo_filtro"):
                _filtros = {**st.session_state["asis_ultimo_filtro"],
                            **{k: v for k, v in _filtros.items() if k != "es_conteo"}}
                _desc = list(st.session_state.get("asis_ultimo_desc", []))
                _bloques.append({"type": "caption",
                                  "text": "(usando el filtro de tu consulta anterior)"})
            if _desc and not _filtros.get("top_n"):
                st.session_state["asis_ultimo_filtro"] = {
                    k: v for k, v in _filtros.items()
                    if k not in ("es_conteo", "dias_no_reconocido")}
                st.session_state["asis_ultimo_desc"] = list(_desc)
            if _filtros.get("dias_no_reconocido"):
                _bloques.append({"type": "caption", "text":
                    "⚠️ Detecté un número de días en tu consulta pero no reconocí cómo compararlo — "
                    "prueba con \"más de N días\", \"menos de N días\", \"entre X y Y días\" o "
                    "\"sobre N días\"."})
            _df_f = _asis_aplicar_filtros(_filtros)
            _etiquetas = []
            if _quiere_neto:
                _etiquetas.append("NETO")
            if _quiere_vale:
                _etiquetas.append("Costo Vale de Consumo")
            _bloques.append({"type": "caption", "text":
                f"Interpreté: monto de {' y '.join(_etiquetas)}"
                + (" · " + " · ".join(_desc) if _desc else " (sobre el total de OTs pendientes)")})
            if _df_f.empty:
                _bloques.append({"type": "warning", "text": "No hay OTs que coincidan con esos filtros."})
            elif _filtros.get("top_n"):
                if "asesor" in _t_norm:
                    _dim = "ASESOR"
                elif "marca" in _t_norm:
                    _dim = "MARCA"
                elif "tipo de venta" in _t_norm or "tipo venta" in _t_norm:
                    _dim = "TIPO VENTA"
                elif "modelo" in _t_norm:
                    _dim = "MODELO"
                elif "categoria" in _t_norm:
                    _dim = "CATEGORIA"
                else:
                    _dim = "SUCURSAL"
                _df_calc = _df_f.copy()
                if _quiere_vale:
                    _df_calc["_VALE_"] = [_asis_costo_vale(r) for r in
                                           _df_calc.get("repuestos_actual", [None] * len(_df_calc))]
                    _metrica_col, _metrica_nombre = "_VALE_", "Costo Vale de Consumo"
                else:
                    _metrica_col, _metrica_nombre = "NETO", "NETO"
                _rank = (_df_calc[_df_calc[_dim] != ""].groupby(_dim)[_metrica_col]
                         .sum().reset_index(name=_metrica_nombre))
                _rank = (_rank.sort_values(_metrica_nombre,
                                           ascending=bool(_filtros.get("top_asc")))
                         .head(_filtros["top_n"]))
                if _rank.empty:
                    _bloques.append({"type": "warning", "text": "No hay datos suficientes para ese ranking."})
                else:
                    _bloques.append({"type": "dataframe", "df": _rank,
                                      "column_config": {_metrica_nombre: st.column_config.NumberColumn(format="$ %d")}})
                    _bloques.append({"type": "bar_chart", "df": _rank, "index": _dim, "value": _metrica_nombre})
            else:
                _items = [("OTs consideradas", len(_df_f))]
                if _quiere_neto:
                    _items.append(("NETO total", f"${_df_f['NETO'].sum():,.0f}"))
                    _items.append(("NETO promedio", f"${_df_f['NETO'].mean():,.0f}"))
                if _quiere_vale:
                    _vales = [_asis_costo_vale(r) for r in
                              _df_f.get("repuestos_actual", [None] * len(_df_f))]
                    _suma_vale = sum(_vales)
                    _items.append(("Costo Vale de Consumo total", f"${_suma_vale:,.0f}"))
                    _items.append(("Costo Vale de Consumo promedio",
                                    f"${(_suma_vale / len(_vales) if _vales else 0):,.0f}"))
                _bloques.append({"type": "metrics", "items": _items})

        elif _es_detalle:
            _matches_det = pd.DataFrame()
            if _folio_hit:
                _matches_det = df_raw[df_raw["FOLIO OT"].astype(str).str.strip() == _folio_hit]
            if _matches_det.empty and _pat_hit:
                _matches_det = df_raw[df_raw["PATENTE"].astype(str).str.strip().str.upper()
                                       .str.replace(" ", "") == _pat_hit]
            _bloques.append({"type": "caption", "text":
                f"Interpreté: detalle completo de la OT/patente «{_folio_hit or _pat_hit}»"})
            if _matches_det.empty:
                _bloques.append({"type": "warning", "text":
                    "No encontré esa OT/patente entre las OTs pendientes."})
            else:
                if len(_matches_det) > 1:
                    _bloques.append({"type": "markdown", "text":
                        f"Esa patente tiene **{len(_matches_det)} OTs pendientes**:"})
                for _, _m_det in _matches_det.iterrows():
                    _bloques.extend(_asis_bloques_detalle_ot(_m_det))
                    _bloques.append({"type": "divider"})

        elif _es_comparacion:
            # Comparación lado a lado entre 2+ valores de una misma dimensión
            # ("compara Linderos con Talca", "Hyundai vs Ford"). 09/07/2026.
            _grupos_cmp, _col_cmp = [], None
            for _col_c in ["SUCURSAL", "ASESOR", "MARCA", "TIPO VENTA", "MODELO", "CATEGORIA"]:
                _vals_c = []
                for _val_c in sorted([v for v in df_raw[_col_c].astype(str).unique()
                                      if v.strip()], key=len, reverse=True):
                    _vn_c = _asis_norm_txt(_val_c)
                    if (_re_asis.search(r"\b" + _re_asis.escape(_vn_c) + r"\b", _t_norm)
                            and _val_c not in _vals_c):
                        _vals_c.append(_val_c)
                if len(_vals_c) >= 2:
                    _col_cmp, _grupos_cmp = _col_c, _vals_c
                    break
            if not _col_cmp:
                _bloques.append({"type": "warning", "text":
                    "Para comparar, menciona dos (o más) valores de la misma dimensión — "
                    "ej: «compara Linderos con Talca» o «Hyundai vs Ford»."})
            else:
                _bloques.append({"type": "caption", "text":
                    f"Interpreté: comparación por {_col_cmp.title()}: "
                    + " vs ".join(_grupos_cmp)})
                _filas_cmp = []
                for _g_cmp in _grupos_cmp:
                    _df_g = df_raw[df_raw[_col_cmp] == _g_cmp]
                    _dias_g = pd.to_numeric(_df_g["DIAS APERTURA"], errors="coerce").fillna(0)
                    try:
                        _neto_g = float(pd.to_numeric(_df_g["NETO"], errors="coerce")
                                        .fillna(0).sum())
                    except Exception:
                        _neto_g = 0.0
                    _filas_cmp.append({
                        _col_cmp.title(): _g_cmp,
                        "OTs": len(_df_g),
                        "% del total": round(len(_df_g) / max(len(df_raw), 1) * 100, 1),
                        "Prom. días": round(float(_dias_g.mean()) if len(_df_g) else 0, 1),
                        "Críticas >90": int((_df_g["RANGO"] == "91 o más").sum()),
                        "NETO total": _neto_g,
                        "Vale Consumo total": sum(_asis_costo_vale(r) for r in
                                                   _df_g.get("repuestos_actual", [])),
                    })
                _df_cmp = pd.DataFrame(_filas_cmp)
                _bloques.append({"type": "dataframe", "df": _df_cmp,
                                  "column_config": {
                                      "NETO total": st.column_config.NumberColumn(format="$ %d"),
                                      "Vale Consumo total": st.column_config.NumberColumn(format="$ %d"),
                                  }})
                _bloques.append({"type": "bar_chart", "df": _df_cmp,
                                  "index": _col_cmp.title(), "value": "OTs"})

        elif any(k in _t for k in ["abono", "anticipo", "cuenta ficha", "saldo"]):
            _t_busq = _t
            for _kw in ["abono de", "abono del", "abono", "anticipo de", "anticipo",
                        "cuenta ficha de", "cuenta ficha", "saldo de", "saldo"]:
                _t_busq = _t_busq.replace(_kw, " ")
            _consulta = _t_busq.strip().upper()
            _clientes_idx = _asis_indice_clientes(df_raw)

            if not _consulta:
                _bloques.append({"type": "caption", "text":
                    "Interpreté: resumen general de abonos / Cuenta Ficha "
                    "(no detecté un cliente específico en la consulta)"})
                _con_saldo = [c for c in _clientes_idx.values()
                              if c["tiene_saldo"] and c["ant_total"] > 0]
                _bloques.append({"type": "metrics", "items": [
                    ("Clientes con saldo", len(_con_saldo)),
                    ("Saldo total", f"${sum(c['ant_total'] for c in _con_saldo):,.0f}"),
                ]})
                if _con_saldo:
                    _df_ab = pd.DataFrame([{
                        "Cliente": c["nombre"], "Saldo": c["ant_total"],
                        "Patentes": ", ".join(sorted(c["patentes"])),
                    } for c in _con_saldo]).sort_values("Saldo", ascending=False)
                    _bloques.append({"type": "dataframe", "df": _df_ab.head(50),
                                      "column_config": {"Saldo": st.column_config.NumberColumn(format="$ %d")}})
            else:
                _encontrados = {}
                for _clave, _c in _clientes_idx.items():
                    _haystack = " ".join([_c["nombre"], " ".join(_c["ruts"]),
                                          " ".join(_c["patentes"])]).upper()
                    if _consulta in _haystack:
                        _encontrados[_clave] = _c
                _bloques.append({"type": "caption",
                                  "text": f"Interpreté: búsqueda de abono/cuenta ficha para «{_consulta}»"})
                if not _encontrados:
                    _bloques.append({"type": "warning",
                                      "text": "No encontré ningún cliente que coincida con esa búsqueda."})
                for _c in _encontrados.values():
                    _bloques.append({"type": "markdown", "text":
                        f"**{_c['nombre']}**  ·  RUT: {', '.join(sorted(_c['ruts'])) or '—'}  ·  "
                        f"Patentes: {', '.join(sorted(_c['patentes'])) or '—'}"})
                    if _c["tiene_saldo"] and _c["ant_total"] > 0:
                        _bloques.append({"type": "success",
                                          "text": f"💰 Saldo en Cuenta Ficha: ${_c['ant_total']:,.0f}"})
                        if _c["movimientos"]:
                            _bloques.append({"type": "dataframe", "df": pd.DataFrame(_c["movimientos"])})
                    else:
                        _bloques.append({"type": "caption", "text": "Sin saldo disponible en Cuenta Ficha."})
                    _bloques.append({"type": "divider"})

        elif any(k in _t_norm for k in ["vehiculos en el taller", "vehiculo en el taller",
                                         "vehiculos en taller", "autos en el taller",
                                         "detenidos en el taller", "en el taller"]):
            _filtros_v, _ = _asis_detectar_filtros(_t)
            _ctrl_data, _ = _cargar_ctrl_taller()
            if not _ctrl_data:
                _bloques.append({"type": "warning", "text":
                    "No pude cargar los datos de Control de Taller (control_taller.json)."})
            else:
                _sucursal_v = _asis_primer_valor(_filtros_v.get("sucursal"))
                _hoy_dt = datetime.now(_TZ_CHILE).replace(tzinfo=None)
                _filas_v = []
                _sucursales_iter = [_sucursal_v] if _sucursal_v else list(_ctrl_data.keys())
                # Restriccion por sucursal (20/07/2026): un usuario limitado nunca puede
                # ver Control de Taller de otra sucursal via el chat, aunque la mencione.
                if _mis_sucursales:
                    _permitidas_norm_v = {s.strip().upper() for s in _mis_sucursales}
                    _sucursales_iter = [s for s in _sucursales_iter if str(s).strip().upper() in _permitidas_norm_v]
                for _suc in _sucursales_iter:
                    _info = _ctrl_data.get(_suc, {}) if isinstance(_ctrl_data, dict) else {}
                    for _o in (_info.get("ordenes", []) if isinstance(_info, dict) else []):
                        if str(_o.get("salida", "")).strip():
                            continue  # ya tiene salida registrada
                        _pat = str(_o.get("patente", "")).strip().upper()
                        if _re_asis.match(r"^SP\d{4}$", _pat):
                            continue  # patente de prueba
                        _ing = str(_o.get("ingreso", "")).strip()
                        _dentro_60 = True
                        if _ing:
                            try:
                                _fd = datetime.strptime(_ing, "%Y-%m-%d")
                                _dentro_60 = (_hoy_dt - _fd).days <= 60
                            except Exception:
                                pass
                        if not _dentro_60:
                            continue
                        _filas_v.append({
                            "Sucursal": _suc, "Patente": _pat, "Modelo": _o.get("modelo", ""),
                            "OT": _o.get("ot", ""), "Técnico": _o.get("tecnico", ""),
                            "Ingreso": _ing, "Mantención": _o.get("mantencion", ""),
                            "Asesor": _o.get("asesor", ""),
                        })
                _desc_v = f" · Sucursal = {_sucursal_v}" if _sucursal_v else ""
                _bloques.append({"type": "caption", "text":
                    f"Interpreté: vehículos actualmente en el taller (sin salida registrada, "
                    f"últimos 60 días){_desc_v}"})
                if not _filas_v:
                    _bloques.append({"type": "warning", "text":
                        "No encontré vehículos en el taller con esos criterios."})
                else:
                    _df_v = pd.DataFrame(_filas_v)
                    _bloques.append({"type": "metrics", "items": [("Vehículos en el taller", len(_df_v))]})
                    _bloques.append({"type": "dataframe", "df": _df_v.head(300)})

        elif any(k in _t_norm for k in ["agenda", "citas de hoy", "cita de hoy", "citas hoy",
                                         "ingresos de hoy", "que citas hay", "citas para hoy"]):
            _filtros_a, _ = _asis_detectar_filtros(_t)
            _agenda_data = _cargar_agenda_hoy()
            if not _agenda_data:
                _bloques.append({"type": "warning", "text":
                    "No pude cargar la Agenda de hoy (agenda_hoy.json)."})
            else:
                _sucursal_a = _asis_primer_valor(_filtros_a.get("sucursal"))
                _hoy_str = datetime.now(_TZ_CHILE).strftime("%d/%m/%Y")
                _sucursales_dict = _agenda_data.get("sucursales", _agenda_data)
                _filas_a = []
                _sucursales_iter_a = [_sucursal_a] if _sucursal_a else list(_sucursales_dict.keys())
                # Restriccion por sucursal (20/07/2026): idem Control de Taller — un
                # usuario limitado no puede consultar la Agenda de otra sucursal via chat.
                if _mis_sucursales:
                    _permitidas_norm_a = {s.strip().upper() for s in _mis_sucursales}
                    _sucursales_iter_a = [s for s in _sucursales_iter_a if str(s).strip().upper() in _permitidas_norm_a]
                for _suc in _sucursales_iter_a:
                    _por_fecha = _sucursales_dict.get(_suc, {})
                    _citas_hoy = _por_fecha if isinstance(_por_fecha, list) else _por_fecha.get(_hoy_str, [])
                    for _c in _citas_hoy:
                        _estado = _c.get("estado") or ("ingresado" if _c.get("ingresado") else "pendiente")
                        _filas_a.append({
                            "Sucursal": _suc, "Horario": _c.get("horario", ""),
                            "Patente": _c.get("patente", ""), "Cliente": _c.get("nombre", ""),
                            "Modelo": _c.get("modelo", ""), "Servicio": _c.get("servicio", ""),
                            "Asesor": _c.get("asesor", ""), "Estado": _estado,
                        })
                _desc_a = f" · Sucursal = {_sucursal_a}" if _sucursal_a else ""
                _bloques.append({"type": "caption", "text":
                    f"Interpreté: citas de hoy ({_hoy_str}) en la Agenda{_desc_a}"})
                if not _filas_a:
                    _bloques.append({"type": "warning", "text":
                        "No hay citas registradas para hoy con esos criterios."})
                else:
                    _df_a = pd.DataFrame(_filas_a)
                    _n_ingresadas  = int((_df_a["Estado"] == "ingresado").sum())
                    _n_pendientes  = int((_df_a["Estado"] == "pendiente").sum())
                    _n_finalizadas = int((_df_a["Estado"] == "finalizado").sum())
                    _bloques.append({"type": "metrics", "items": [
                        ("Total citas hoy", len(_df_a)),
                        ("🎟️ Ingresadas", _n_ingresadas),
                        ("🚗 Pendientes", _n_pendientes),
                        ("🧍 Finalizadas", _n_finalizadas),
                    ]})
                    _bloques.append({"type": "dataframe", "df": _df_a.sort_values("Horario").head(300)})

        elif any(k in _t_norm for k in ["ranking de cierres", "ranking cierres",
                                         "cierres mayor a 90", "cierres sobre 90",
                                         "ot cerradas hace mas de 90", "ranking de ot cerradas",
                                         "ranking de cierre"]):
            # Fuente: sección 5.3 de la Constitución — ranking_cierres.json (agregados
            # de OT que superaron 90 días al momento de cerrarse). 10/07/2026.
            _filtros_rk, _ = _asis_detectar_filtros(_t)
            _ranking_data = cargar_ranking_cierres()
            if not _ranking_data:
                _bloques.append({"type": "warning", "text":
                    "No pude cargar el ranking de cierres (ranking_cierres.json)."})
            else:
                _sucursal_rk = _asis_primer_valor(_filtros_rk.get("sucursal"))
                _asesor_rk   = _asis_primer_valor(_filtros_rk.get("asesor"))
                _desc_rk = []
                if _sucursal_rk: _desc_rk.append(f"Sucursal = {_sucursal_rk}")
                if _asesor_rk:   _desc_rk.append(f"Asesor = {_asesor_rk}")
                _bloques.append({"type": "caption", "text":
                    f"Interpreté: ranking de OT cerradas con más de 90 días "
                    f"(desde {_ranking_data.get('periodo_desde','—')}, actualizado "
                    f"{_ranking_data.get('fecha_generacion','—')})"
                    + (" · " + " · ".join(_desc_rk) if _desc_rk else "")})
                _bloques.append({"type": "metrics", "items": [
                    ("Total OT >90 días cerradas", _ranking_data.get("total_ots_90mas", 0))]})
                if _asesor_rk:
                    _filas_asesor = [r for r in _ranking_data.get("por_asesor", [])
                                      if _asis_norm_txt(r.get("ASESOR", "")) == _asis_norm_txt(_asesor_rk)]
                    if _filas_asesor:
                        _bloques.append({"type": "dataframe", "df": pd.DataFrame(_filas_asesor)})
                    else:
                        _bloques.append({"type": "warning", "text":
                            f"{_asesor_rk} no tiene OT cerradas con más de 90 días en este período."})
                elif _sucursal_rk:
                    _filas_suc = [r for r in _ranking_data.get("por_asesor_sucursal", [])
                                   if _asis_norm_txt(r.get("SUCURSAL", "")) == _asis_norm_txt(_sucursal_rk)]
                    if _filas_suc:
                        _bloques.append({"type": "dataframe",
                                          "df": pd.DataFrame(_filas_suc).sort_values("total", ascending=False)})
                    else:
                        _bloques.append({"type": "warning", "text":
                            f"{_sucursal_rk} no tiene OT cerradas con más de 90 días en este período."})
                else:
                    _bloques.append({"type": "markdown", "text": "**Top asesores (por cantidad):**"})
                    _bloques.append({"type": "dataframe",
                                      "df": pd.DataFrame(_ranking_data.get("por_asesor", [])).head(10)})
                    _bloques.append({"type": "markdown", "text": "**Por sucursal:**"})
                    _bloques.append({"type": "dataframe",
                                      "df": pd.DataFrame(_ranking_data.get("por_sucursal", []))})

        elif any(k in _t_norm for k in ["cuantas ot cerro", "cuantas ots cerro", "cuantas cerro",
                                         "ot cerradas", "ots cerradas", "cuantas se cerraron",
                                         "cuantas cerraron", "cuantas ot se cerraron"]):
            # Fuente: historial_cierres.json — cada corrida de consolidación detecta qué OT
            # dejaron de estar pendientes y las guarda con sus datos. Sección 5.3. 10/07/2026.
            _filtros_hc, _ = _asis_detectar_filtros(_t)
            _df_hist_resumen, _registros_hc = cargar_historial_cierres()
            if not _registros_hc:
                _bloques.append({"type": "warning", "text":
                    "No pude cargar el historial de cierres (historial_cierres.json)."})
            else:
                _filas_cerr = []
                for _reg in _registros_hc:
                    for _ot_c in _reg.get("ots_cerradas", []):
                        _fila_c = dict(_ot_c)
                        _fila_c["Fecha Detección"] = _reg.get("fecha", "")
                        _filas_cerr.append(_fila_c)
                _df_cerr = pd.DataFrame(_filas_cerr)
                _sucursal_hc = _asis_primer_valor(_filtros_hc.get("sucursal"))
                _asesor_hc   = _asis_primer_valor(_filtros_hc.get("asesor"))
                if not _df_cerr.empty:
                    if _sucursal_hc and "SUCURSAL" in _df_cerr.columns:
                        _df_cerr = _df_cerr[_df_cerr["SUCURSAL"].astype(str).str.upper() == _sucursal_hc.upper()]
                    if _asesor_hc and "ASESOR" in _df_cerr.columns:
                        _df_cerr = _df_cerr[_df_cerr["ASESOR"].astype(str).apply(_asis_norm_txt)
                                              == _asis_norm_txt(_asesor_hc)]
                    if "FOLIO OT" in _df_cerr.columns:
                        _df_cerr = _df_cerr.drop_duplicates(subset=["FOLIO OT"])
                _desc_hc = []
                if _sucursal_hc: _desc_hc.append(f"Sucursal = {_sucursal_hc}")
                if _asesor_hc:   _desc_hc.append(f"Asesor = {_asesor_hc}")
                _bloques.append({"type": "caption", "text":
                    "Interpreté: OT detectadas como cerradas en las corridas de consolidación"
                    + (" · " + " · ".join(_desc_hc) if _desc_hc else "")})
                _bloques.append({"type": "caption", "text":
                    "Nota: se detectan cuando la OT deja de aparecer como pendiente en una corrida "
                    "de consolidación — la fecha es cuándo se notó, no necesariamente la fecha "
                    "exacta de cierre."})
                if _df_cerr.empty:
                    _bloques.append({"type": "warning", "text":
                        "No encontré OT cerradas con esos criterios."})
                else:
                    _bloques.append({"type": "metrics", "items": [("OT cerradas detectadas", len(_df_cerr))]})
                    _bloques.append({"type": "dataframe", "df": _df_cerr.head(300)})

        elif any(k in _t_norm for k in ["tempario", "horas de mantencion", "horas de mano de obra",
                                         "horas de servicio", "horas de la mantencion",
                                         "horas de mantencion tiene"]):
            # Fuente: tempario.json (pautas oficiales de mano de obra por marca/modelo/km,
            # 8 marcas cubiertas). Sección 5.3. 10/07/2026.
            _tempario_data = _cargar_tempario()
            if not _tempario_data:
                _bloques.append({"type": "warning", "text":
                    "No pude cargar el tempario (tempario.json)."})
            else:
                _t_sin_esp = _t_norm.replace(" ", "")
                _mo_km = _re_asis.search(r"(\d[\d\.]{2,})\s*km", _t_norm)
                _km_pedido = None
                if _mo_km:
                    try:
                        _km_pedido = int(_mo_km.group(1).replace(".", ""))
                    except Exception:
                        _km_pedido = None
                _match_temp = None
                for _mdl in _tempario_data:
                    _marca_n  = _asis_norm_txt(_mdl.get("marca", ""))
                    _modelo_n = _asis_norm_txt(_mdl.get("modelo", "")).replace(" ", "")
                    if (_marca_n and _re_asis.search(r"\b" + _re_asis.escape(_marca_n) + r"\b", _t_norm)
                            and _modelo_n and _modelo_n in _t_sin_esp):
                        _match_temp = _mdl
                        break
                if not _match_temp:
                    _bloques.append({"type": "warning", "text":
                        "No reconocí la marca/modelo dentro del tempario (cubre BAIC, Ford, JAC, "
                        "Jaecoo, JIM, Mahindra, Omoda, Shineray). Probá con el nombre exacto, ej. "
                        "\"tempario Ford Ranger a los 20000 km\"."})
                else:
                    _km_horas_m = _match_temp.get("km_horas", {})
                    if _km_pedido:
                        _horas_exact = _km_horas_m.get(str(_km_pedido))
                        _nota_km = ""
                        if _horas_exact is None and _km_horas_m:
                            _cercano = min(_km_horas_m.keys(), key=lambda k: abs(int(k) - _km_pedido))
                            _horas_exact = _km_horas_m.get(_cercano)
                            _nota_km = f" (km más cercano disponible: {int(_cercano):,})"
                        _bloques.append({"type": "caption", "text":
                            f"Interpreté: tempario de {_match_temp.get('marcaNombre', '')} "
                            f"{_match_temp.get('modelo', '')} a los {_km_pedido:,} km{_nota_km}"})
                        if _horas_exact is not None:
                            _bloques.append({"type": "success",
                                              "text": f"⏳ {_horas_exact} horas de mano de obra"})
                        else:
                            _bloques.append({"type": "warning", "text": "No hay dato para ese kilometraje."})
                    else:
                        _bloques.append({"type": "caption", "text":
                            f"Interpreté: tabla completa de tempario de "
                            f"{_match_temp.get('marcaNombre', '')} {_match_temp.get('modelo', '')}"})
                        _df_temp = pd.DataFrame(
                            [{"KM": int(k), "Horas": v} for k, v in _km_horas_m.items()])
                        _bloques.append({"type": "dataframe", "df": _df_temp.sort_values("KM")})

        elif any(k in _t_norm for k in ["tecnicos tiene", "tecnicos de la sucursal",
                                         "lista de tecnicos", "cuantos tecnicos",
                                         "que tecnicos hay", "tecnicos configurados"]):
            # Fuente: control_taller.json → ctrl_data[SUCURSAL]["tecnicos"] — ya se carga
            # para "vehículos en el taller", se reutiliza sin fetch nuevo. Sección 5.3. 10/07/2026.
            _filtros_tec, _ = _asis_detectar_filtros(_t)
            _sucursal_tec = _asis_primer_valor(_filtros_tec.get("sucursal"))
            _ctrl_data_tec, _ = _cargar_ctrl_taller()
            if not _ctrl_data_tec:
                _bloques.append({"type": "warning", "text":
                    "No pude cargar los datos de Control de Taller (control_taller.json)."})
            elif not _sucursal_tec:
                _bloques.append({"type": "warning", "text":
                    "Decime de qué sucursal — ej. \"¿qué técnicos tiene Linderos?\"."})
            elif _mis_sucursales and _sucursal_tec.strip().upper() not in {s.strip().upper() for s in _mis_sucursales}:
                _bloques.append({"type": "warning", "text":
                    f"🔒 No tienes acceso a la sucursal '{_sucursal_tec}'. Tu acceso está limitado a: "
                    f"{', '.join(_mis_sucursales)}."})
            else:
                _tecs = (_ctrl_data_tec.get(_sucursal_tec, {}) or {}).get("tecnicos", [])
                _bloques.append({"type": "caption", "text":
                    f"Interpreté: técnicos configurados en {_sucursal_tec}"})
                if not _tecs:
                    _bloques.append({"type": "warning", "text":
                        f"{_sucursal_tec} no tiene técnicos configurados en el Planificador."})
                else:
                    _bloques.append({"type": "metrics", "items": [("Técnicos", len(_tecs))]})
                    _bloques.append({"type": "dataframe", "df": pd.DataFrame({"Técnico": _tecs})})

        elif any(k in _t for k in ["repuesto", "stock", "bodega"]):
            _stock_catalogo, _stock_fecha = _cargar_stock_repuestos()
            if not _stock_catalogo:
                _bloques.append({"type": "warning", "text":
                    "El catálogo de Stock de Repuestos todavía no está disponible. Hay que correr "
                    "Ejecutar_Consolidacion.bat una vez para generarlo (PASO 10)."})
            else:
                _df_reps, _palabras_rep = _asis_buscar_repuestos(_q_raw, _stock_catalogo)
                _suf_fecha = f" (catálogo actualizado {_stock_fecha})" if _stock_fecha else ""
                if _palabras_rep:
                    _bloques.append({"type": "caption", "text":
                        f"Interpreté: búsqueda de stock que contenga «{' '.join(_palabras_rep)}»{_suf_fecha}"})
                else:
                    _bloques.append({"type": "caption", "text":
                        f"Interpreté: resumen general de stock de repuestos{_suf_fecha}"})
                if _df_reps.empty:
                    _bloques.append({"type": "warning",
                                      "text": "No encontré repuestos que coincidan con esa búsqueda."})
                else:
                    _con_stock = int((_df_reps["Stock"] > 0).sum())
                    _sin_stock = int((_df_reps["Stock"] <= 0).sum())
                    _bloques.append({"type": "metrics", "items": [
                        ("Productos encontrados", len(_df_reps)),
                        ("Con stock (> 0)", _con_stock),
                        ("Sin stock", _sin_stock),
                    ]})
                    _df_reps = _df_reps.sort_values("Stock", ascending=False)
                    _bloques.append({"type": "dataframe", "df": _df_reps.head(300),
                                      "column_config": {
                                          "Stock": st.column_config.NumberColumn(format="%d"),
                                          "Stock Proyectado": st.column_config.NumberColumn(format="%d"),
                                          "Costo": st.column_config.NumberColumn(format="$ %d"),
                                      }})
                    if len(_df_reps) > 300:
                        _bloques.append({"type": "caption", "text":
                            f"Mostrando 300 de {len(_df_reps):,} resultados. "
                            "Afina la búsqueda (ej. el nombre del repuesto) para acotar."})

        elif _t_norm.strip(" ?!.,¿¡") in {
            "hola", "gracias", "muchas gracias", "buenas", "buenos dias",
            "buenas tardes", "buenas noches", "chao", "adios", "ok", "listo",
            "genial", "perfecto", "excelente", "de nada", "como estas", "que tal",
        }:
            # Saludo/agradecimiento puro — no debe disparar la reutilización de
            # contexto (si no, "hola" después de una consulta real volvería a
            # mostrar el mismo listado en vez de solo saludar). 08/07/2026.
            _bloques.append({"type": "markdown", "text":
                "¡Hola! Preguntame sobre sucursales, asesores, marcas, montos, stock de repuestos, "
                "vehículos en el taller o lo que necesites de tus OTs pendientes."})

        else:
            _filtros, _desc = _asis_detectar_filtros(_t)
            if _filtros.get("dias_no_reconocido"):
                _bloques.append({"type": "caption", "text":
                    "⚠️ Detecté un número de días en tu consulta pero no reconocí cómo compararlo — "
                    "prueba con \"más de N días\", \"menos de N días\", \"entre X y Y días\" o "
                    "\"sobre N días\"."})

            # -- Continuidad de conversación -------------------------------
            # Cualquier mensaje de seguimiento sin filtros propios ("dame los
            # números de OT", "ordénalas por rango", "y las de esas cuáles
            # son", etc.) reutiliza automáticamente el último filtro que se
            # interpretó con éxito en el chat (guardado en session_state) —
            # ya no depende de reconocer una frase exacta para activarse.
            # Si el mensaje SÍ trae filtros propios pero además suena a
            # refinamiento ("de esas OT...", "y que también...", "filtra
            # también..."), se combinan ambos: se parte del contexto anterior
            # y se sobreescribe/agrega lo nuevo detectado en este mensaje.
            # 08/07/2026.
            _pide_listado = any(k in _t_norm for k in [
                "numero de ot", "numeros de ot", "los numeros", "los folios",
                "el listado", "la lista", "el detalle", "cuales son", "cuales fueron",
                "muestrame las ot", "muestrame los folios", "dame el listado",
                "dame la lista", "dame los folios", "ver las ot", "ver el listado",
                "cuales ot", "que ot son", "que ots son", "cuales ots",
            ])
            _orden_col, _orden_asc, _pide_orden = _asis_detectar_orden(_t_norm)
            _es_refinamiento = any(k in _t_norm for k in [
                "de esas", "de esos", "de esa lista", "de ese listado", "entre esas",
                "entre esos", "y que tambien", "ademas que", "filtra tambien",
                "de esos resultados", "de ese resultado", "y ademas", "tambien de",
                "mismo pero", "lo mismo pero", "misma consulta pero", "las mismas",
                "los mismos",
            ]) or bool(_re_asis.match(r"^\s*(?:y|ahora)\b", _t_norm)) or bool(
                # Demostrativos que refieren al resultado anterior ("ordena ESTAS
                # 31 OTs...", "esas OTs...") — sin esto, un mensaje que además
                # menciona una dimensión ("...de Ranger") se interpretaba como
                # consulta NUEVA y perdía el resto del filtro del hilo. 09/07/2026.
                _re_asis.search(r"\b(estas|estos|esas|esos)\b", _t_norm)
                and st.session_state.get("asis_ultimo_filtro"))
            _tiene_contexto = bool(st.session_state.get("asis_ultimo_filtro"))
            _uso_contexto = False
            # es_conteo debe venir SIEMPRE del mensaje actual, no del contexto
            # guardado — "¿y cuántas son?" tras un listado es un conteo aunque
            # el turno anterior haya sido un listado. 09/07/2026.
            _es_conteo_msg = bool(_filtros.get("es_conteo"))

            # "quita/saca el filtro de X" — elimina una dimensión del contexto y
            # re-muestra el resultado con el resto de los filtros. 09/07/2026.
            _mo_quitar = _re_asis.search(
                r"(?:quita|quitale|saca|sacale|elimina|borra|remueve)\w*\s+"
                r"(?:el |la |los |las )?(?:filtro\s+)?(?:de |del |de la )?"
                r"(sucursal|marca|asesor|rango|dias|categoria|ano|año|modelo|"
                r"tipo de venta|mes)", _t_norm)

            if _mo_quitar and _tiene_contexto:
                _dim_q = _mo_quitar.group(1)
                _keys_q = {
                    "sucursal": ["sucursal"], "marca": ["marca"], "asesor": ["asesor"],
                    "rango": ["rango"], "categoria": ["categoria"],
                    "ano": ["anio"], "año": ["anio"], "modelo": ["modelo"],
                    "tipo de venta": ["tipo_venta"], "mes": ["mes"],
                    "dias": ["dias_mayor", "dias_menor", "dias_entre"],
                }.get(_dim_q, [])
                _pref_q = {
                    "sucursal": ["Sucursal ="], "marca": ["Marca ="],
                    "asesor": ["Asesor ="], "rango": ["Rango ="],
                    "categoria": ["Categoria ="], "ano": ["Año ="], "año": ["Año ="],
                    "modelo": ["Modelo ="], "tipo de venta": ["Tipo Venta ="],
                    "mes": ["Mes OT ="], "dias": ["Días apertura"],
                }.get(_dim_q, [])
                _filtros = dict(st.session_state["asis_ultimo_filtro"])
                for _kq in _keys_q:
                    _filtros.pop(_kq, None)
                if isinstance(_filtros.get("excluir"), dict):
                    for _kq in _keys_q:
                        _filtros["excluir"].pop(_kq, None)
                _desc = [d for d in st.session_state.get("asis_ultimo_desc", [])
                         if not any(p in d for p in _pref_q)]
                _uso_contexto = True
            elif _tiene_contexto and _es_refinamiento:
                _base = dict(st.session_state["asis_ultimo_filtro"])
                _base_desc = list(st.session_state.get("asis_ultimo_desc", []))
                # Un filtro de días nuevo reemplaza al de días anterior (no se
                # acumulan dos condiciones de días contradictorias). 09/07/2026.
                if any(k in _filtros for k in ("dias_mayor", "dias_menor", "dias_entre")):
                    for _kd in ("dias_mayor", "dias_menor", "dias_entre"):
                        _base.pop(_kd, None)
                    _base_desc = [d for d in _base_desc
                                  if not d.startswith("Días apertura")]
                for _k, _v in _filtros.items():
                    if _k == "es_conteo":
                        continue
                    _base[_k] = _v
                # Si este mensaje reemplaza una dimensión que ya estaba en el
                # contexto ("y en Curicó?" tras Linderos), la descripción vieja
                # de esa dimensión se descarta para no mostrar las dos. 09/07/2026.
                _prefijos_nuevos = [d.split("=")[0].strip() + " ="
                                    for d in _desc if "=" in d]
                _base_desc = [d for d in _base_desc
                              if not any(d.startswith(p) for p in _prefijos_nuevos)]
                _filtros = _base
                _desc = _base_desc + [d for d in _desc if d not in _base_desc]
                _uso_contexto = True
            elif not _desc and _tiene_contexto:
                _filtros = dict(st.session_state["asis_ultimo_filtro"])
                _desc = list(st.session_state.get("asis_ultimo_desc", []))
                _uso_contexto = True
            _filtros["es_conteo"] = _es_conteo_msg

            _df_f = _asis_aplicar_filtros(_filtros)
            if _pide_orden:
                _df_f = _asis_aplicar_orden(_df_f, _orden_col, _orden_asc)
                if _orden_col:
                    st.session_state["asis_ultimo_orden"] = (_orden_col, _orden_asc)
            elif _uso_contexto and st.session_state.get("asis_ultimo_orden"):
                # Mantiene el orden pedido en un turno anterior mientras se siga
                # trabajando sobre el mismo hilo. 09/07/2026.
                _oc_prev, _oa_prev = st.session_state["asis_ultimo_orden"]
                _df_f = _asis_aplicar_orden(_df_f, _oc_prev, _oa_prev)

            # Guarda el filtro (ya combinado si hubo refinamiento) para que el
            # próximo turno lo pueda retomar. No se guarda un ranking Top N
            # porque ese no es un "conjunto de OT" reutilizable como los demás.
            if _desc and not _filtros.get("top_n"):
                st.session_state["asis_ultimo_filtro"] = {
                    k: v for k, v in _filtros.items()
                    if k not in ("es_conteo", "dias_no_reconocido")}
                st.session_state["asis_ultimo_desc"] = list(_desc)
                if not _uso_contexto and not _pide_orden:
                    # Consulta nueva desde cero → el orden de un hilo anterior
                    # ya no aplica.
                    st.session_state.pop("asis_ultimo_orden", None)

            # ---- Flags de los tipos de respuesta nuevos: desglose, distintos,
            # ---- estadísticas, porcentaje y listar dimensiones. 09/07/2026.
            _mo_desglose = None if _pide_orden else _re_asis.search(
                r"\bpor\s+(sucursal|asesor|marca|rango|tipo de venta|categoria|"
                r"ano|año|modelo|mes)\b", _t_norm)
            _mo_distinct = _re_asis.search(
                r"cuant\w*\s+(sucursales|asesores|marcas|modelos|categorias|"
                r"patentes|clientes|vehiculos)\b", _t_norm)
            _mo_listar = _re_asis.search(
                r"(?:que|cuales|lista(?:do)?(?: de)?|dime|nombra|enumera)\b[^\.]*?\b"
                r"(sucursales|asesores|marcas|modelos|categorias|tipos de venta)\b",
                _t_norm)
            _pide_stats = any(k in _t_norm for k in [
                "promedio de dias", "dias promedio", "promedio dias",
                "antiguedad promedio", "edad promedio", "maximo de dias",
                "minimo de dias", "dias maximo", "dias minimo", "mas antigua",
                "mas antiguo", "mas vieja", "mas viejo", "mas nueva", "mas nuevo",
                "mas reciente", "mayor neto", "neto mas alto", "mas cara",
                "mas caras", "mayor tiempo",
            ])
            _pide_pct = "porcentaje" in _t_norm or "%" in _t

            _cols_show = [c for c in ["FOLIO OT", "SUCURSAL", "ASESOR", "MARCA", "MODELO",
                                       "PATENTE", "TIPO VENTA", "RANGO", "DIAS APERTURA",
                                       "GLOSA TRABAJO"] if c in _df_f.columns]
            _orden_txt = ""
            if _pide_orden:
                _orden_txt = (f" · ordenado por {_orden_col.title()} "
                              f"({'ascendente' if _orden_asc else 'descendente'})"
                              if _orden_col else " · (no reconocí por qué columna ordenar)")

            if _filtros.get("top_n"):
                if "asesor" in _t:
                    _dim = "ASESOR"
                elif "marca" in _t:
                    _dim = "MARCA"
                elif "tipo de venta" in _t or "tipo venta" in _t:
                    _dim = "TIPO VENTA"
                elif "modelo" in _t_norm:
                    _dim = "MODELO"
                elif "categoria" in _t_norm:
                    _dim = "CATEGORIA"
                else:
                    _dim = "SUCURSAL"
                _bloques.append({"type": "caption", "text":
                    f"Interpreté: ranking Top {_filtros['top_n']} por {_dim.title()}"
                    + (" · " + " · ".join(_desc) if _desc else "")})
                _rank = _df_f[_df_f[_dim] != ""].groupby(_dim).size().reset_index(name="Cantidad")
                _rank = (_rank.sort_values("Cantidad", ascending=bool(_filtros.get("top_asc")))
                         .head(_filtros["top_n"]))
                if _rank.empty:
                    _bloques.append({"type": "warning", "text": "No hay datos suficientes para ese ranking."})
                else:
                    _bloques.append({"type": "dataframe", "df": _rank})
                    _bloques.append({"type": "bar_chart", "df": _rank, "index": _dim, "value": "Cantidad"})

            elif _mo_distinct:
                _col_dis = _ASIS_PLURAL_COLS.get(_mo_distinct.group(1))
                _lbl_dis = _mo_distinct.group(1)
                if _col_dis and _col_dis in _df_f.columns:
                    _n_dis = int(_df_f[_df_f[_col_dis].astype(str).str.strip() != ""]
                                 [_col_dis].nunique())
                    _bloques.append({"type": "caption", "text":
                        f"Interpreté: cantidad de {_lbl_dis} distintos"
                        + (" · " + " · ".join(_desc) if _desc
                           else " (en todas las OTs pendientes)")})
                    _bloques.append({"type": "metrics", "items": [
                        (f"{_lbl_dis.title()} distintos", _n_dis),
                        ("OTs consideradas", len(_df_f)),
                    ]})
                else:
                    _bloques.append({"type": "warning", "text":
                        "No pude identificar esa dimensión en los datos."})

            elif _mo_desglose and (_filtros.get("es_conteo") or any(
                    k in _t_norm for k in ["desglose", "desglosa", "distribucion",
                                            "agrupa", "agrupadas", "agrupados",
                                            "reparto", "separa", "separadas",
                                            "abiertas por", "cuadro"])):
                _dim_g = _mo_desglose.group(1)
                if _dim_g == "mes":
                    _fechas_g = pd.to_datetime(_df_f.get("FECHA OT"), dayfirst=True,
                                                errors="coerce")
                    _tab_g = (_df_f.assign(_MES_=_fechas_g.dt.strftime("%Y-%m"))
                              .dropna(subset=["_MES_"]).groupby("_MES_").size()
                              .reset_index(name="OTs").rename(columns={"_MES_": "Mes"}))
                    _dim_col_g = "Mes"
                else:
                    _col_g = _ASIS_DESGLOSE_COLS.get(_dim_g)
                    if _col_g and _col_g in _df_f.columns:
                        _df_g2 = _df_f[_df_f[_col_g].astype(str).str.strip() != ""]
                    else:
                        _df_g2 = _df_f.iloc[0:0]
                    _tab_g = _df_g2.groupby(_col_g).size().reset_index(name="OTs") \
                        if len(_df_g2) else pd.DataFrame()
                    if len(_tab_g) and "NETO" in _df_g2.columns:
                        _neto_grp = (pd.to_numeric(_df_g2["NETO"], errors="coerce")
                                     .fillna(0).groupby(_df_g2[_col_g]).sum()
                                     .reset_index(name="NETO total"))
                        _tab_g = _tab_g.merge(_neto_grp, on=_col_g, how="left")
                    _dim_col_g = _col_g
                if _tab_g is None or not len(_tab_g):
                    _bloques.append({"type": "warning",
                                      "text": "No hay datos para ese desglose."})
                else:
                    _tot_g = int(_tab_g["OTs"].sum())
                    _tab_g["%"] = (_tab_g["OTs"] / max(_tot_g, 1) * 100).round(1)
                    if _dim_g == "rango":
                        _rmap_g = {"0-30": 0, "31-60": 1, "61-90": 2, "91 o más": 3}
                        _tab_g = (_tab_g.assign(_o_=_tab_g[_dim_col_g].map(_rmap_g)
                                                .fillna(9))
                                  .sort_values("_o_").drop(columns="_o_"))
                    elif _dim_g == "mes":
                        _tab_g = _tab_g.sort_values("Mes")
                    else:
                        _tab_g = _tab_g.sort_values("OTs", ascending=False)
                    _bloques.append({"type": "caption", "text":
                        f"Interpreté: desglose por {_dim_g.title()}"
                        + (" · " + " · ".join(_desc) if _desc
                           else " (todas las OTs pendientes)")})
                    _bloques.append({"type": "metrics",
                                      "items": [("OTs consideradas", _tot_g)]})
                    _cc_g = ({"NETO total": st.column_config.NumberColumn(format="$ %d")}
                             if "NETO total" in _tab_g.columns else None)
                    _bloques.append({"type": "dataframe", "df": _tab_g,
                                      "column_config": _cc_g})
                    _bloques.append({"type": "bar_chart", "df": _tab_g,
                                      "index": _dim_col_g, "value": "OTs"})

            elif _pide_stats:
                _dias_s = pd.to_numeric(_df_f["DIAS APERTURA"], errors="coerce").fillna(0)
                if _df_f.empty:
                    _bloques.append({"type": "warning",
                                      "text": "No hay OTs que coincidan con esos filtros."})
                elif any(k in _t_norm for k in ["mayor neto", "neto mas alto",
                                                 "mas cara", "mas caras"]):
                    _bloques.append({"type": "caption", "text":
                        "Interpreté: OTs con mayor NETO"
                        + (" · " + " · ".join(_desc) if _desc else "")})
                    _neto_s = pd.to_numeric(_df_f["NETO"], errors="coerce").fillna(0)
                    _top_neto = (_df_f.assign(_n_=_neto_s)
                                 .sort_values("_n_", ascending=False).drop(columns="_n_"))
                    _cols_neto = _cols_show + (["NETO"] if "NETO" in _df_f.columns
                                                and "NETO" not in _cols_show else [])
                    _bloques.append({"type": "dataframe",
                                      "df": _top_neto[_cols_neto].head(5),
                                      "column_config": {"NETO": st.column_config.NumberColumn(format="$ %d")}})
                elif any(k in _t_norm for k in ["mas antigua", "mas antiguo", "mas vieja",
                                                 "mas viejo", "mayor tiempo"]):
                    _bloques.append({"type": "caption", "text":
                        "Interpreté: OTs más antiguas (mayor días de apertura)"
                        + (" · " + " · ".join(_desc) if _desc else "")})
                    _bloques.append({"type": "dataframe", "df":
                        (_df_f.assign(_d_=_dias_s).sort_values("_d_", ascending=False)
                         .drop(columns="_d_")[_cols_show].head(5))})
                elif any(k in _t_norm for k in ["mas nueva", "mas nuevo", "mas reciente"]):
                    _bloques.append({"type": "caption", "text":
                        "Interpreté: OTs más recientes (menor días de apertura)"
                        + (" · " + " · ".join(_desc) if _desc else "")})
                    _bloques.append({"type": "dataframe", "df":
                        (_df_f.assign(_d_=_dias_s).sort_values("_d_", ascending=True)
                         .drop(columns="_d_")[_cols_show].head(5))})
                else:
                    _bloques.append({"type": "caption", "text":
                        "Interpreté: estadísticas de antigüedad (días de apertura)"
                        + (" · " + " · ".join(_desc) if _desc
                           else " (todas las OTs pendientes)")})
                    _bloques.append({"type": "metrics", "items": [
                        ("OTs", len(_df_f)),
                        ("Promedio días", f"{_dias_s.mean():,.1f}"),
                        ("Mínimo", int(_dias_s.min())),
                        ("Máximo", int(_dias_s.max())),
                    ]})

            elif _pide_pct and _desc:
                _pct_v = len(_df_f) / max(len(df_raw), 1) * 100
                _bloques.append({"type": "caption", "text":
                    "Interpreté: porcentaje sobre el total de OTs pendientes · "
                    + " · ".join(_desc)})
                _bloques.append({"type": "metrics", "items": [
                    ("OTs que cumplen", len(_df_f)),
                    ("Total pendientes", len(df_raw)),
                    ("Porcentaje", f"{_pct_v:,.1f}%"),
                ]})

            elif _mo_listar:
                _col_l = _ASIS_PLURAL_COLS.get(_mo_listar.group(1))
                if not _col_l or _col_l not in _df_f.columns:
                    _bloques.append({"type": "warning", "text":
                        "No pude identificar esa dimensión en los datos."})
                else:
                    _vc_l = (_df_f[_df_f[_col_l].astype(str).str.strip() != ""]
                             [_col_l].value_counts().reset_index())
                    _vc_l.columns = [_col_l.title(), "OTs"]
                    _bloques.append({"type": "caption", "text":
                        f"Interpreté: listado de {_mo_listar.group(1)}"
                        + (" · " + " · ".join(_desc) if _desc
                           else " (en todas las OTs pendientes)")})
                    _bloques.append({"type": "metrics", "items": [
                        (f"{_mo_listar.group(1).title()} distintos", len(_vc_l))]})
                    _bloques.append({"type": "dataframe", "df": _vc_l.head(300)})

            elif (_pide_listado or _pide_orden
                  or (_uso_contexto and not _filtros.get("es_conteo"))) and _desc:
                _bloques.append({"type": "caption", "text":
                    "Interpreté: listado de OTs · " + " · ".join(_desc) + _orden_txt
                    + (" (usando el filtro de tu consulta anterior)" if _uso_contexto else "")})
                if _df_f.empty:
                    _bloques.append({"type": "warning", "text": "No hay OTs que coincidan con esos filtros."})
                else:
                    _bloques.append({"type": "metrics", "items": [("OTs encontradas", len(_df_f))]})
                    _bloques.append({"type": "dataframe", "df": _df_f[_cols_show].head(300)})
                    if len(_df_f) > 300:
                        _bloques.append({"type": "caption", "text": f"Mostrando 300 de {len(_df_f):,} resultados."})

            elif _filtros.get("es_conteo") and not _desc:
                _bloques.append({"type": "caption", "text": "Interpreté: conteo total (no detecté filtros específicos)"})
                _bloques.append({"type": "metrics", "items": [("Cantidad de OTs pendientes", len(_df_f))]})

            elif _filtros.get("es_conteo"):
                _bloques.append({"type": "caption", "text": "Interpreté: conteo · " + " · ".join(_desc)
                                  + " · tip: escribe \"dame los números de OT\" para ver el listado"})
                _bloques.append({"type": "metrics", "items": [("Cantidad de OTs", len(_df_f))]})

            else:
                # Pidió ordenar/listar pero no hay ni filtros propios ni contexto
                # anterior — avisar en vez de caer a texto libre con ruido. 09/07/2026.
                if not _desc and (_pide_orden or _pide_listado):
                    _bloques.append({"type": "warning", "text":
                        "No tengo una consulta anterior en memoria que ordenar o listar. "
                        "Haz primero una consulta (ej: \"OTs de Linderos\") y después pídeme "
                        "ordenarla o listarla."})
                    return _bloques
                if not _desc:
                    _stop = {"que","con","para","desde","hasta","tiene","tienen","cual","cuales",
                             "cuál","cuáles","cuantas","cuántas","cuantos","cuántos","donde","dónde",
                             "como","cómo","esta","están","son","los","las","del","por","una","unos",
                             "unas","mas","más","dias","días","ots","hay","muestrame","muéstrame",
                             "dame","quiero","ver","todas","todos","folio","patente"}
                    _palabras_lib = [w for w in _re_asis.findall(r"[a-záéíóúñ0-9]{4,}", _t)
                                     if w not in _stop]
                    if _palabras_lib:
                        _mask = pd.Series(False, index=_df_f.index)
                        for _col_txt in ["GLOSA TRABAJO", "OBSERVACION OT", "NOTAS", "AVANCE - GESTIÓN"]:
                            if _col_txt in _df_f.columns:
                                for _w in _palabras_lib:
                                    _mask |= (_df_f[_col_txt].astype(str).str.lower()
                                              .str.contains(_w, na=False, regex=False))
                        _df_f = _df_f[_mask]
                        _desc.append(f"Texto libre: {' '.join(_palabras_lib)}")

                if not _desc:
                    _bloques.append({"type": "warning", "text":
                        "No logré interpretar esa consulta. Prueba mencionando una sucursal, un "
                        "asesor, una marca, un rango de días, \"más de N días\", \"top N\", o alguna "
                        "palabra clave del trabajo (ej: parabrisas, choque, mantención). "
                        "Escribe **\"ayuda\"** para ver todo lo que puedo hacer."})
                else:
                    _bloques.append({"type": "caption", "text": "Interpreté: " + " · ".join(_desc)})
                    _bloques.append({"type": "metrics", "items": [("OTs encontradas", len(_df_f))]})
                    _bloques.append({"type": "dataframe", "df": _df_f[_cols_show].head(300)})
                    if len(_df_f) > 300:
                        _bloques.append({"type": "caption", "text": f"Mostrando 300 de {len(_df_f):,} resultados."})

        return _bloques

    def _asis_pintar_bloques(_bloques):
        for _b in _bloques:
            _tipo = _b["type"]
            if _tipo == "caption":
                st.caption(_b["text"])
            elif _tipo == "markdown":
                st.markdown(_b["text"])
            elif _tipo == "warning":
                st.warning(_b["text"])
            elif _tipo == "success":
                st.success(_b["text"])
            elif _tipo == "divider":
                st.divider()
            elif _tipo == "metrics":
                _cols = st.columns(len(_b["items"]))
                for _col, (_label, _val) in zip(_cols, _b["items"]):
                    _col.metric(_label, _val)
            elif _tipo == "dataframe":
                st.dataframe(_b["df"], hide_index=True, use_container_width=True,
                             column_config=_b.get("column_config"))
            elif _tipo == "bar_chart":
                st.bar_chart(_b["df"].set_index(_b["index"])[_b["value"]])

    # -- Interfaz de chat -----------------------------------------------
    if "asis_chat" not in st.session_state:
        st.session_state["asis_chat"] = []

    _chat_head_col, _chat_clear_col = st.columns([6, 1])
    with _chat_clear_col:
        if st.session_state["asis_chat"] and st.button(
            "🗑️ Limpiar", key="asis_clear_chat", use_container_width=True
        ):
            st.session_state["asis_chat"] = []
            st.rerun()

    _asis_chat_box = st.container(height=430, border=True)
    with _asis_chat_box:
        if not st.session_state["asis_chat"]:
            with st.chat_message("assistant", avatar="🤖"):
                st.markdown(
                    "¡Hola! Pregúntame sobre sucursales, asesores, marcas, rangos de días, "
                    "detalles de una OT o patente, comparaciones, desgloses, montos, stock de "
                    "repuestos o abonos de clientes — y puedo seguir el hilo de la conversación "
                    "(*\"ordénalas por días\"*, *\"y en Curicó?\"*, *\"de esas dame las de más de "
                    "60 días\"*). También te explico **cómo usar la app** (*\"¿cómo uso el "
                    "planificador?\"*). Escribe **\"ayuda\"** para ver todo lo que sé hacer."
                )
        for _turno in st.session_state["asis_chat"]:
            if _turno["role"] == "user":
                with st.chat_message("user", avatar="🧑"):
                    st.markdown(_turno["text"])
            else:
                with st.chat_message("assistant", avatar="🤖"):
                    _asis_pintar_bloques(_turno["bloques"])

    _asis_nueva_consulta = st.chat_input(
        'Escribe tu consulta... ej: "cuántas OT tiene Linderos", "abono Juan Pérez"',
        key="asis_chat_input",
    )
    if _asis_nueva_consulta:
        st.session_state["asis_chat"].append({"role": "user", "text": _asis_nueva_consulta})
        _bloques_resp = _asis_procesar_consulta(_asis_nueva_consulta)
        st.session_state["asis_chat"].append({"role": "assistant", "bloques": _bloques_resp})
        st.rerun()

    st.stop()   # ← Detener aquí: no ejecutar el resto del app (modo OTs)



if st.session_state.get("app_mode") == "ots":
    _render_control()
